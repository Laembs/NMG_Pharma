"""SP11: Synchronisiert app/i18n.py mit translations/NMGone_Translations.xlsx.

Aufruf:
    python scripts/sync_translations.py

Liest das Excel, generiert die TRANSLATIONS-Dict neu, schreibt nach
app/i18n.py. Wird vor jedem Build manuell aufgerufen, wenn der User
eine aktualisierte Excel geliefert hat.
"""
from __future__ import annotations
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
EXCEL = ROOT / "translations" / "NMGone_Translations.xlsx"
I18N = ROOT / "app" / "i18n.py"


def main() -> int:
    if not EXCEL.exists():
        print(f"FEHLT: {EXCEL}", file=sys.stderr)
        return 1
    if not I18N.exists():
        print(f"FEHLT: {I18N}", file=sys.stderr)
        return 1

    wb = load_workbook(EXCEL)
    if "Translations" not in wb.sheetnames:
        print("Sheet 'Translations' nicht gefunden", file=sys.stderr)
        return 1
    ws = wb["Translations"]

    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    try:
        key_idx = header.index("Key (DE original)")
        de_idx = header.index("Deutsch (DE)")
        en_idx = header.index("English (EN)")
        sk_idx = next(i for i, h in enumerate(header) if h.startswith("Slovencina") or h.startswith("Slovenčina"))
        cz_idx = next(i for i, h in enumerate(header) if h.startswith("Cestina") or h.startswith("Čeština"))
    except (ValueError, StopIteration) as exc:
        print(f"Header-Spalten nicht erkannt: {exc}", file=sys.stderr)
        return 1

    translations: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[key_idx]:
            continue
        key = str(row[key_idx]).strip()
        de = str(row[de_idx] or "").strip()
        en = str(row[en_idx] or "").strip()
        sk = str(row[sk_idx] or "").strip()
        cz = str(row[cz_idx] or "").strip()
        # Stub-Fallback: leeres SK/CZ -> DE
        if not sk:
            sk = de
        if not cz:
            cz = de
        translations[key] = {"DE": de, "EN": en, "SK": sk, "CZ": cz}

    def py_escape(s: str) -> str:
        return s.replace("\\", "\\\\").replace('"', '\\"')

    lines = ["TRANSLATIONS: dict[str, dict[str, str]] = {"]
    for key, langs in translations.items():
        lines.append(f'    "{py_escape(key)}": {{')
        for lang in ("DE", "EN", "SK", "CZ"):
            lines.append(f'        "{lang}": "{py_escape(langs[lang])}",')
        lines.append("    },")
    lines.append("}")
    new_dict = "\n".join(lines)

    content = I18N.read_text(encoding="utf-8")
    pattern = r"TRANSLATIONS: dict\[str, dict\[str, str\]\] = \{.*?\n\}"
    new_content, n = re.subn(pattern, new_dict, content, count=1, flags=re.DOTALL)
    if n == 0:
        print("Konnte TRANSLATIONS-Dict in i18n.py nicht finden", file=sys.stderr)
        return 1

    I18N.write_text(new_content, encoding="utf-8")
    print(f"OK: {len(translations)} Eintraege aus {EXCEL.name} -> {I18N.name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
