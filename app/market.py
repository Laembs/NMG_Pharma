from pathlib import Path
from datetime import datetime
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .config import DB_PATH, OUTPUT_DIR
from .db import init_db


def _safe_num(value):
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def _norm_source(datenquelle: str | None) -> str:
    dq = (datenquelle or "ALLE").upper().strip()
    # PK ist die neue fachliche Bezeichnung für die bisherigen NMG-/Partnerkonditions-Auswertungen.
    # Intern bleiben bestehende Altdaten aus Kompatibilitätsgründen als NMG gespeichert.
    if dq == "PK":
        return "NMG"
    return dq if dq in {"NMG", "ZF", "ALLE"} else "ALLE"


def _source_label(dq: str) -> str:
    if dq == "NMG":
        return "PK"
    if dq == "ZF":
        return "ZF"
    return "PK + ZF"


def export_marktanalyse_nicht_nmg(limit: int = 200, min_apotheken: int = 1, datenquelle: str = "ALLE", auswertung_id: int | None = None) -> Path:
    """Exportiert eine Marktanalyse über gespeicherte Auswertungen.

    Ziel: Produkte finden, die keine NMG-Zuordnung haben und bei Kunden hohe Abverkäufe haben.
    Gruppierung erfolgt über PZN. Sortierung: Gesamtabsatz absteigend, dann Anzahl Apotheken.
    """
    if not DB_PATH.exists():
        init_db(DB_PATH)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    dq = _norm_source(datenquelle)
    scope = f"AUSWERTUNG_{auswertung_id}_" if auswertung_id else ""
    out = OUTPUT_DIR / f"Marktanalyse_{scope}Nicht_NMG_{dq}_Top_Absatz_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

    with sqlite3.connect(DB_PATH) as con:
        con.row_factory = sqlite3.Row
        rows = con.execute(
            """
            SELECT
                p.pzn,
                COALESCE(MAX(NULLIF(p.artikelname, '')), '') AS artikelname,
                COALESCE(MAX(NULLIF(p.df, '')), '') AS df,
                COALESCE(MAX(NULLIF(p.pck, '')), '') AS pck,
                COALESCE(MAX(NULLIF(p.herstellerkuerzel, '')), '') AS hersteller,
                COUNT(DISTINCT a.apotheke) AS apotheken,
                COUNT(*) AS vorkommen,
                SUM(COALESCE(p.absatz_6m, 0)) AS gesamt_absatz_6m,
                AVG(NULLIF(p.absatz_6m, 0)) AS durchschnitt_absatz_6m,
                SUM(COALESCE(p.umsatz, 0)) AS gesamt_umsatz,
                MAX(COALESCE(p.absatz_6m, 0)) AS hoechster_einzel_absatz,
                GROUP_CONCAT(DISTINCT a.apotheke) AS apotheken_liste
            FROM tbl_auswertungspositionen p
            JOIN tbl_auswertungen a ON a.id = p.auswertung_id
            WHERE COALESCE(p.ist_nmg_treffer, 0) = 0
              AND COALESCE(p.absatz_6m, 0) > 0
              AND p.pzn IS NOT NULL AND p.pzn <> ''
              AND p.pzn NOT IN (SELECT pzn FROM tbl_nmg_stamm)
              AND (? = 'ALLE' OR COALESCE(a.datenquelle, 'NMG') = ?)
              AND (? IS NULL OR a.id = ?)
            GROUP BY p.pzn
            HAVING COUNT(DISTINCT a.apotheke) >= ?
            ORDER BY gesamt_absatz_6m DESC, apotheken DESC, vorkommen DESC
            LIMIT ?
            """,
            (dq, dq, auswertung_id, auswertung_id, int(min_apotheken), int(limit)),
        ).fetchall()

        summary = con.execute(
            """
            SELECT COUNT(*) AS auswertungen,
                   COALESCE(SUM(anzahl_positionen), 0) AS positionen,
                   COALESCE(SUM(nicht_nmg), 0) AS nicht_nmg,
                   COALESCE(SUM(nmg_treffer), 0) AS nmg_treffer
            FROM tbl_auswertungen
            WHERE (? = 'ALLE' OR COALESCE(datenquelle, 'NMG') = ?)
              AND (? IS NULL OR id = ?)
            """, (dq, dq, auswertung_id, auswertung_id)
        ).fetchone()

    wb = Workbook()
    ws = wb.active
    ws.title = "Nicht-PK Top Absatz"

    title = f"Marktanalyse: Nicht-PK-Produkte mit hohem Absatz ({_source_label(dq)})"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:L1")
    ws["A2"] = f"Basis: {summary['auswertungen']} gespeicherte Auswertungen, {summary['positionen']} Positionen, {summary['nicht_nmg']} Nicht-PK-Positionen"
    ws.merge_cells("A2:L2")
    ws["A3"] = "Hinweis: EK wird nur aus Rohdaten/Fallback verwendet; diese Analyse dient als Kandidatenliste für mögliche neue Produkte."
    ws.merge_cells("A3:L3")

    headers = [
        "Rang", "PZN", "Artikelname", "DF", "PCK", "Hersteller", "Apotheken",
        "Vorkommen", "Gesamtabsatz 6M", "Ø Absatz 6M", "Gesamtumsatz", "Apothekenliste"
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(5, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, row in enumerate(rows, start=1):
        r = idx + 5
        values = [
            idx, row["pzn"], row["artikelname"], row["df"], row["pck"], row["hersteller"],
            row["apotheken"], row["vorkommen"], _safe_num(row["gesamt_absatz_6m"]),
            _safe_num(row["durchschnitt_absatz_6m"]), _safe_num(row["gesamt_umsatz"]),
            row["apotheken_liste"] or "",
        ]
        for c, value in enumerate(values, start=1):
            ws.cell(r, c, value)

    last_row = max(5, len(rows) + 5)
    thin = Side(style="thin", color="B7B7B7")
    for row in ws.iter_rows(min_row=5, max_row=last_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in [9, 10, 11]:
        for r in range(6, last_row + 1):
            ws.cell(r, col).number_format = '#,##0.00'
    widths = [8, 12, 42, 8, 10, 14, 10, 10, 16, 14, 14, 50]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:L{last_row}"

    # Zweites Blatt: gespeicherte Auswertungen als Nachweis.
    ws2 = wb.create_sheet("Gespeicherte Auswertungen")
    with sqlite3.connect(DB_PATH) as con:
        con.row_factory = sqlite3.Row
        auswertungen = con.execute(
            """SELECT datum, apotheke, quelldatei, anzahl_positionen, nmg_treffer, nicht_nmg, gesamt_absatz
                   FROM tbl_auswertungen
                   WHERE (? = 'ALLE' OR COALESCE(datenquelle, 'NMG') = ?)
                     AND (? IS NULL OR id = ?)
                   ORDER BY datum DESC""", (dq, dq, auswertung_id, auswertung_id)
        ).fetchall()
    headers2 = ["Datum", "Apotheke", "Quelldatei", "Positionen", "PK Treffer", "Nicht-PK", "Gesamtabsatz"]
    for c, h in enumerate(headers2, start=1):
        ws2.cell(1, c, h).font = Font(bold=True)
    for r_idx, row in enumerate(auswertungen, start=2):
        for c_idx, key in enumerate(["datum", "apotheke", "quelldatei", "anzahl_positionen", "nmg_treffer", "nicht_nmg", "gesamt_absatz"], start=1):
            ws2.cell(r_idx, c_idx, row[key])
    for idx, width in enumerate([20, 28, 42, 12, 12, 12, 14], start=1):
        ws2.column_dimensions[get_column_letter(idx)].width = width
    ws2.auto_filter.ref = f"A1:G{max(1, len(auswertungen)+1)}"
    wb.save(out)
    return out




def export_marktanalyse_produktchancen(limit: int = 500, min_apotheken: int = 1, datenquelle: str = "ALLE", auswertung_id: int | None = None) -> Path:
    """Erweiterte Marktanalyse für Produktentwicklung.

    Unterscheidet gespeicherte Positionen in:
    - Bereits im NMG-Sortiment
    - NMG-Ersatz vorhanden (Austauschartikel)
    - Neue Produktchance (nicht NMG, kein bekannter Ersatz)
    """
    if not DB_PATH.exists():
        init_db(DB_PATH)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    dq = _norm_source(datenquelle)
    scope = f"AUSWERTUNG_{auswertung_id}_" if auswertung_id else ""
    out = OUTPUT_DIR / f"Marktanalyse_Produktchancen_{scope}{dq}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

    base_sql = """
        WITH gruppiert AS (
            SELECT
                p.pzn,
                COALESCE(MAX(NULLIF(p.artikelname, '')), '') AS artikelname,
                COALESCE(MAX(NULLIF(p.df, '')), '') AS df,
                COALESCE(MAX(NULLIF(p.pck, '')), '') AS pck,
                COALESCE(MAX(NULLIF(p.herstellerkuerzel, '')), '') AS hersteller,
                COUNT(DISTINCT a.apotheke) AS apotheken,
                COUNT(*) AS vorkommen,
                SUM(COALESCE(p.absatz_6m, 0)) AS gesamt_absatz_6m,
                AVG(NULLIF(p.absatz_6m, 0)) AS durchschnitt_absatz_6m,
                SUM(COALESCE(p.umsatz, 0)) AS gesamt_umsatz,
                MAX(COALESCE(p.absatz_6m, 0)) AS hoechster_einzel_absatz,
                GROUP_CONCAT(DISTINCT a.apotheke) AS apotheken_liste,
                MAX(COALESCE(p.pzn_nmg, '')) AS pzn_nmg_auswertung,
                MAX(COALESCE(p.austauschbar_gegen, '')) AS austausch_auswertung
            FROM tbl_auswertungspositionen p
            JOIN tbl_auswertungen a ON a.id = p.auswertung_id
            WHERE COALESCE(p.absatz_6m, 0) > 0
              AND p.pzn IS NOT NULL AND p.pzn <> ''
              AND (:datenquelle = 'ALLE' OR COALESCE(a.datenquelle, 'NMG') = :datenquelle)
              AND (:auswertung_id IS NULL OR a.id = :auswertung_id)
            GROUP BY p.pzn
            HAVING COUNT(DISTINCT a.apotheke) >= :min_apotheken
        )
        SELECT
            g.*,
            ns.pzn AS nmg_sortiment_pzn,
            ns.artikelname AS nmg_sortiment_artikel,
            aa.nmg_pzn AS austausch_nmg_pzn,
            aa.austauschbar_gegen AS austausch_nmg_artikel,
            aa.bemerkung AS austausch_bemerkung,
            CASE
                WHEN ns.pzn IS NOT NULL THEN 'Bereits im PK-/NMG-Sortiment'
                WHEN NULLIF(COALESCE(aa.nmg_pzn, g.pzn_nmg_auswertung), '') IS NOT NULL
                     OR NULLIF(COALESCE(aa.austauschbar_gegen, g.austausch_auswertung), '') IS NOT NULL
                    THEN 'PK-/NMG-Ersatz vorhanden'
                ELSE 'Neue Produktchance'
            END AS status,
            CASE
                WHEN ns.pzn IS NOT NULL THEN 0
                WHEN NULLIF(COALESCE(aa.nmg_pzn, g.pzn_nmg_auswertung), '') IS NOT NULL
                     OR NULLIF(COALESCE(aa.austauschbar_gegen, g.austausch_auswertung), '') IS NOT NULL
                    THEN 25
                ELSE 100
            END
            + COALESCE(g.gesamt_absatz_6m, 0)
            + (COALESCE(g.apotheken, 0) * 10) AS produktchancen_score
        FROM gruppiert g
        LEFT JOIN tbl_nmg_stamm ns ON ns.pzn = g.pzn
        LEFT JOIN tbl_austauschartikel aa ON aa.original_pzn = g.pzn
    """

    with sqlite3.connect(DB_PATH) as con:
        con.row_factory = sqlite3.Row
        rows_all = con.execute(
            base_sql + " ORDER BY produktchancen_score DESC, gesamt_absatz_6m DESC LIMIT :limit",
            {"min_apotheken": int(min_apotheken), "limit": int(limit), "datenquelle": dq, "auswertung_id": auswertung_id},
        ).fetchall()
        rows_chancen = [r for r in rows_all if r["status"] == "Neue Produktchance"]
        rows_ersatz = [r for r in rows_all if r["status"] == "PK-/NMG-Ersatz vorhanden"]
        rows_sortiment = [r for r in rows_all if r["status"] == "Bereits im PK-/NMG-Sortiment"]

        hersteller = con.execute(
            """
            WITH produktstatus AS (
                """ + base_sql + """
            )
            SELECT
                COALESCE(NULLIF(hersteller, ''), '(unbekannt)') AS hersteller,
                COUNT(*) AS produkte,
                SUM(gesamt_absatz_6m) AS gesamt_absatz_6m,
                SUM(apotheken) AS apotheken_summe,
                SUM(CASE WHEN status='Neue Produktchance' THEN 1 ELSE 0 END) AS neue_chancen,
                SUM(CASE WHEN status='PK-/NMG-Ersatz vorhanden' THEN 1 ELSE 0 END) AS ersatz_vorhanden
            FROM produktstatus
            WHERE status <> 'Bereits im PK-/NMG-Sortiment'
            GROUP BY COALESCE(NULLIF(hersteller, ''), '(unbekannt)')
            ORDER BY gesamt_absatz_6m DESC, produkte DESC
            LIMIT 200
            """,
            {"min_apotheken": int(min_apotheken), "datenquelle": dq, "auswertung_id": auswertung_id},
        ).fetchall()

        summary = con.execute(
            """
            SELECT COUNT(*) AS auswertungen,
                   COALESCE(SUM(anzahl_positionen), 0) AS positionen,
                   COALESCE(SUM(nicht_nmg), 0) AS nicht_nmg,
                   COALESCE(SUM(nmg_treffer), 0) AS nmg_treffer
            FROM tbl_auswertungen
            WHERE (:datenquelle = 'ALLE' OR COALESCE(datenquelle, 'NMG') = :datenquelle)
              AND (:auswertung_id IS NULL OR id = :auswertung_id)
            """, {"datenquelle": dq, "auswertung_id": auswertung_id}
        ).fetchone()

    wb = Workbook()
    wb.remove(wb.active)

    def write_rows(ws, title, rows, note=""):
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:Q1")
        ws["A2"] = f"Basis: {summary['auswertungen']} gespeicherte Auswertungen, {summary['positionen']} Positionen"
        ws.merge_cells("A2:Q2")
        if note:
            ws["A3"] = note
            ws.merge_cells("A3:Q3")
        headers = [
            "Rang", "Status", "Score", "PZN", "Artikelname", "DF", "PCK", "Hersteller",
            "Apotheken", "Vorkommen", "Gesamtabsatz 6M", "Ø Absatz 6M", "Gesamtumsatz",
            "PK-/NMG-PZN/Ersatz", "PK-/NMG-Ersatzartikel", "höchster Einzelabsatz", "Apothekenliste"
        ]
        header_row = 5
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(header_row, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for idx, row in enumerate(rows, start=1):
            r = idx + header_row
            nmg_pzn = row["nmg_sortiment_pzn"] or row["austausch_nmg_pzn"] or row["pzn_nmg_auswertung"] or ""
            nmg_artikel = row["nmg_sortiment_artikel"] or row["austausch_nmg_artikel"] or row["austausch_auswertung"] or ""
            values = [
                idx, row["status"], _safe_num(row["produktchancen_score"]), row["pzn"], row["artikelname"], row["df"], row["pck"], row["hersteller"],
                row["apotheken"], row["vorkommen"], _safe_num(row["gesamt_absatz_6m"]), _safe_num(row["durchschnitt_absatz_6m"]), _safe_num(row["gesamt_umsatz"]),
                nmg_pzn, nmg_artikel, _safe_num(row["hoechster_einzel_absatz"]), row["apotheken_liste"] or ""
            ]
            for c, value in enumerate(values, start=1):
                ws.cell(r, c, value)
        last_row = max(header_row, len(rows) + header_row)
        thin = Side(style="thin", color="B7B7B7")
        for excel_row in ws.iter_rows(min_row=header_row, max_row=last_row, min_col=1, max_col=len(headers)):
            for cell in excel_row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in [3, 11, 12, 13, 16]:
            for r in range(header_row + 1, last_row + 1):
                ws.cell(r, col).number_format = '#,##0.00'
        widths = [8, 24, 14, 12, 42, 8, 10, 14, 10, 10, 16, 14, 14, 16, 42, 16, 55]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A6"
        ws.auto_filter.ref = f"A{header_row}:Q{last_row}"

    write_rows(
        wb.create_sheet("Produktchancen"),
        "Marktanalyse: Neue Produktchancen",
        rows_chancen,
        "Nicht im PK-/NMG-Stamm und kein bekannter PK-/NMG-Austauschartikel. Das ist die wichtigste Liste für Produktentwicklung."
    )
    write_rows(
        wb.create_sheet("Ersatz vorhanden"),
        "Marktanalyse: Nicht-PK-Produkte mit vorhandenem NMG-Ersatz",
        rows_ersatz,
        "Nicht im PK-/NMG-Stamm, aber ein Austauschartikel bzw. Ersatz ist bekannt."
    )
    write_rows(
        wb.create_sheet("Bereits Sortiment"),
        "Marktanalyse: Bereits im NMG-Sortiment",
        rows_sortiment,
        "Zur Kontrolle: Produkte, deren PZN bereits im NMG-Stamm enthalten ist."
    )

    ws_h = wb.create_sheet("Herstelleranalyse")
    ws_h["A1"] = "Herstelleranalyse außerhalb PK-/NMG-Sortiment"
    ws_h["A1"].font = Font(bold=True, size=14)
    headers_h = ["Rang", "Hersteller", "Produkte", "Gesamtabsatz 6M", "Apotheken-Summe", "Neue Produktchancen", "Ersatz vorhanden"]
    for c, h in enumerate(headers_h, start=1):
        cell = ws_h.cell(3, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for idx, row in enumerate(hersteller, start=1):
        vals = [idx, row["hersteller"], row["produkte"], _safe_num(row["gesamt_absatz_6m"]), row["apotheken_summe"], row["neue_chancen"], row["ersatz_vorhanden"]]
        for c, v in enumerate(vals, start=1):
            ws_h.cell(idx + 3, c, v)
    for idx, width in enumerate([8, 18, 12, 16, 16, 20, 18], start=1):
        ws_h.column_dimensions[get_column_letter(idx)].width = width
    ws_h.auto_filter.ref = f"A3:G{max(3, len(hersteller)+3)}"
    ws_h.freeze_panes = "A4"

    ws_s = wb.create_sheet("Übersicht")
    ws_s["A1"] = f"Marktanalyse+ Übersicht ({dq})"
    ws_s["A1"].font = Font(bold=True, size=14)
    info = [
        ("Gespeicherte Auswertungen", summary["auswertungen"]),
        ("Gespeicherte Positionen", summary["positionen"]),
        ("Produkte in Analyse", len(rows_all)),
        ("Neue Produktchancen", len(rows_chancen)),
        ("PK-/NMG-Ersatz vorhanden", len(rows_ersatz)),
        ("Bereits im Sortiment", len(rows_sortiment)),
        ("Mindestanzahl Apotheken", min_apotheken),
        ("Limit je Analyse", limit),
    ]
    for r, (k, v) in enumerate(info, start=3):
        ws_s.cell(r, 1, k).font = Font(bold=True)
        ws_s.cell(r, 2, v)
    ws_s.column_dimensions["A"].width = 32
    ws_s.column_dimensions["B"].width = 20

    wb.save(out)
    return out


def datenbankstatus() -> dict:
    if not DB_PATH.exists():
        init_db(DB_PATH)
    with sqlite3.connect(DB_PATH) as con:
        con.row_factory = sqlite3.Row
        def one(sql):
            return con.execute(sql).fetchone()[0]
        return {
            "auswertungen": one("SELECT COUNT(*) FROM tbl_auswertungen"),
            "positionen": one("SELECT COUNT(*) FROM tbl_auswertungspositionen"),
            "nicht_nmg_positionen": one("SELECT COUNT(*) FROM tbl_auswertungspositionen WHERE COALESCE(ist_nmg_treffer,0)=0"),
            "pk_positionen": one("SELECT COUNT(*) FROM tbl_auswertungspositionen WHERE COALESCE(ist_nmg_treffer,0)=1"),
            "zf_auswertungen": one("SELECT COUNT(*) FROM tbl_auswertungen WHERE COALESCE(datenquelle,'NMG')='ZF'"),
            "pk_auswertungen": one("SELECT COUNT(*) FROM tbl_auswertungen WHERE COALESCE(datenquelle,'NMG')='NMG'"),
        }


# ── SP14: Neue Produktanalyse PK/ZF/PK+ZF ────────────────────────────────────
def _produktanalyse_datenquellen(kundentyp: str) -> list[str]:
    """Mapping zwischen kundentyp ('PK'/'ZF'/'PK+ZF') und den
    tbl_auswertungen.datenquelle-Werten, die wir filtern muessen.
    Altdaten haben datenquelle='NMG' obwohl es PK-Auswertungen sind.
    """
    k = (kundentyp or "").upper().replace(" ", "")
    if k == "PK":
        return ["NMG", "PK"]
    if k == "ZF":
        return ["ZF"]
    return ["NMG", "PK", "ZF"]  # PK+ZF oder ALLE


def _produktanalyse_rows(con: sqlite3.Connection, kundentyp: str, monate: int = 6) -> list[sqlite3.Row]:
    """SP15/SP30: Liefert die Zeilen fuer eine Produktanalyse.

    Filter:
      - Auswertungen der letzten N Monate, ausgewaehlte Datenquellen
      - PZN NICHT im NMG-Stamm (= kein NMG-Artikel)
      - PZN hat KEINEN Austausch-Eintrag (tbl_austauschartikel/tbl_austauschdatenbank)
        -> wenn fuer eine alte PZN bereits eine NMG-Ersatz-PZN hinterlegt ist,
        ist die alte keine "Chance" mehr, sondern bereits substituiert.
      - Biosimilar-Datenbank: technisch identisch zu tbl_austauschdatenbank
        (Schulbank-Biosimilar zeigt dieselbe Tabelle ohne weiteren Filter);
        der Austausch-Filter unten deckt damit beide Faelle ab.

    Stammdaten (Artikelname, DF, PCK, Hersteller) und EK kommen aus
    tbl_artikelstamm. Gesamtumsatz = Gesamtabsatz x Artikelstamm-EK.

    V1.1 SP1: PZNs in allen vier Wissens-Tabellen sind beim Insert bereits
    normalisiert (8-stellig, nur Ziffern). pzn_norm()-UDFs in JOIN-/EXISTS-
    Klauseln zwingen SQLite zum Full-Table-Scan und verhindern die
    vorhandenen Indizes (idx_tbl_auswertungspositionen_pzn,
    idx_austauschdatenbank_pzn_alt, PK auf tbl_nmg_stamm/tbl_artikelstamm).
    Direktvergleich mit `=` nutzt die Indizes und ist ~10-100x schneller.

    SP30 ergaenzt:
      - analyse_anzahl = COUNT(DISTINCT auswertung_id) (= "Vorkommen")
      - erste_sichtung = MIN(tbl_auswertungen.datum)
      - letzte_sichtung = MAX(tbl_auswertungen.datum)
      - kunden_anzahl-Vorbereitung: sobald Auswertungen einem Kunden
        zugeordnet werden (SP31+), kommt zusaetzlich
        COUNT(DISTINCT kunde_id) AS kunden_anzahl.
    """
    datenquellen = _produktanalyse_datenquellen(kundentyp)
    placeholders = ",".join("?" * len(datenquellen))
    sql = f"""
        WITH gruppiert AS (
            SELECT
                p.pzn,
                COUNT(DISTINCT a.apotheke) AS apotheken,
                COUNT(DISTINCT p.auswertung_id) AS analyse_anzahl,
                MIN(a.datum) AS erste_sichtung,
                MAX(a.datum) AS letzte_sichtung,
                SUM(COALESCE(p.absatz_6m, 0)) AS gesamtabsatz_6m,
                AVG(NULLIF(p.absatz_6m, 0)) AS durchschnitt_absatz_6m
            FROM tbl_auswertungspositionen p
            JOIN tbl_auswertungen a ON a.id = p.auswertung_id
            WHERE p.pzn IS NOT NULL AND p.pzn <> ''
              AND COALESCE(p.absatz_6m, 0) > 0
              AND COALESCE(a.datenquelle, 'NMG') IN ({placeholders})
              AND a.datum >= date('now', '-{int(monate)} months')
            GROUP BY p.pzn
        )
        SELECT
            g.pzn,
            COALESCE(ast.artikel, '') AS artikelname,
            COALESCE(ast.df, '') AS df,
            COALESCE(ast.pck, '') AS pck,
            COALESCE(ast.herst, '') AS hersteller_stamm,
            g.apotheken,
            g.gesamtabsatz_6m,
            COALESCE(g.durchschnitt_absatz_6m, 0) AS durchschnitt_absatz_6m,
            COALESCE(ast.ek, 0) AS effektiver_ek,
            g.gesamtabsatz_6m * COALESCE(ast.ek, 0) AS moeglicher_gesamtumsatz,
            g.analyse_anzahl,
            g.erste_sichtung,
            g.letzte_sichtung
        FROM gruppiert g
        JOIN tbl_artikelstamm ast ON ast.pzn = g.pzn
        LEFT JOIN tbl_nmg_stamm nmg ON nmg.pzn = g.pzn
        WHERE nmg.pzn IS NULL
          -- V1.1 SP10: nur Artikel mit Namen im Artikelstamm
          AND COALESCE(ast.artikel, '') <> ''
          AND NOT EXISTS (
              SELECT 1 FROM tbl_austauschartikel aa
              WHERE aa.original_pzn = g.pzn
                AND (COALESCE(aa.nmg_pzn, '') <> '' OR COALESCE(aa.austauschbar_gegen, '') <> '')
          )
          AND NOT EXISTS (
              SELECT 1 FROM tbl_austauschdatenbank ad
              WHERE ad.pzn_alt = g.pzn
                AND (COALESCE(ad.pzn_nmg, '') <> '' OR COALESCE(ad.freitext_austausch, '') <> '')
                AND COALESCE(ad.status, 'aktiv') = 'aktiv'
          )
        ORDER BY moeglicher_gesamtumsatz DESC, g.gesamtabsatz_6m DESC
    """
    con.row_factory = sqlite3.Row
    return con.execute(sql, datenquellen).fetchall()


def _produktanalyse_rows_nmg(con: sqlite3.Connection, kundentyp: str, monate: int = 6) -> list[sqlite3.Row]:
    """V1.1 SP6: Liefert PZNs aus den Auswertungen, die IM NMG-Stamm sind.
    Disjunkt zur Hauptliste, sortiert wie die Hauptliste (Umsatz DESC).
    """
    datenquellen = _produktanalyse_datenquellen(kundentyp)
    placeholders = ",".join("?" * len(datenquellen))
    sql = f"""
        WITH gruppiert AS (
            SELECT
                p.pzn,
                COUNT(DISTINCT a.apotheke) AS apotheken,
                COUNT(DISTINCT p.auswertung_id) AS analyse_anzahl,
                MIN(a.datum) AS erste_sichtung,
                MAX(a.datum) AS letzte_sichtung,
                SUM(COALESCE(p.absatz_6m, 0)) AS gesamtabsatz_6m,
                AVG(NULLIF(p.absatz_6m, 0)) AS durchschnitt_absatz_6m
            FROM tbl_auswertungspositionen p
            JOIN tbl_auswertungen a ON a.id = p.auswertung_id
            WHERE p.pzn IS NOT NULL AND p.pzn <> ''
              AND COALESCE(p.absatz_6m, 0) > 0
              AND COALESCE(a.datenquelle, 'NMG') IN ({placeholders})
              AND a.datum >= date('now', '-{int(monate)} months')
            GROUP BY p.pzn
        )
        SELECT
            g.pzn,
            COALESCE(nmg.artikelname, ast.artikel, '') AS artikelname,
            COALESCE(ast.df, '') AS df,
            COALESCE(ast.pck, '') AS pck,
            COALESCE(nmg.herstellerkuerzel, ast.herst, '') AS hersteller_stamm,
            g.apotheken,
            g.gesamtabsatz_6m,
            COALESCE(g.durchschnitt_absatz_6m, 0) AS durchschnitt_absatz_6m,
            COALESCE(ast.ek, nmg.taxe_ek, 0) AS effektiver_ek,
            g.gesamtabsatz_6m * COALESCE(ast.ek, nmg.taxe_ek, 0) AS moeglicher_gesamtumsatz,
            g.analyse_anzahl,
            g.erste_sichtung,
            g.letzte_sichtung
        FROM gruppiert g
        JOIN tbl_nmg_stamm nmg ON nmg.pzn = g.pzn
        LEFT JOIN tbl_artikelstamm ast ON ast.pzn = g.pzn
        -- V1.1 SP10: nur Treffer mit Artikelnamen (NMG-Stamm oder Artikelstamm)
        WHERE COALESCE(nmg.artikelname, ast.artikel, '') <> ''
        ORDER BY moeglicher_gesamtumsatz DESC, g.gesamtabsatz_6m DESC
    """
    con.row_factory = sqlite3.Row
    return con.execute(sql, datenquellen).fetchall()


def _produktanalyse_rows_austausch(con: sqlite3.Connection, kundentyp: str, monate: int = 6) -> list[sqlite3.Row]:
    """V1.1 SP6: Liefert PZNs aus den Auswertungen, die einen Austausch-
    Eintrag haben aber NICHT im NMG-Stamm sind. Disjunkt zur Hauptliste
    und zur NMG-Liste. Liefert zusaetzlich Ersatz-PZN, Ersatz-Artikel,
    Quelle (aus tbl_austauschdatenbank bevorzugt, sonst tbl_austauschartikel).
    """
    datenquellen = _produktanalyse_datenquellen(kundentyp)
    placeholders = ",".join("?" * len(datenquellen))
    sql = f"""
        WITH gruppiert AS (
            SELECT
                p.pzn,
                COUNT(DISTINCT a.apotheke) AS apotheken,
                COUNT(DISTINCT p.auswertung_id) AS analyse_anzahl,
                MIN(a.datum) AS erste_sichtung,
                MAX(a.datum) AS letzte_sichtung,
                SUM(COALESCE(p.absatz_6m, 0)) AS gesamtabsatz_6m,
                AVG(NULLIF(p.absatz_6m, 0)) AS durchschnitt_absatz_6m
            FROM tbl_auswertungspositionen p
            JOIN tbl_auswertungen a ON a.id = p.auswertung_id
            WHERE p.pzn IS NOT NULL AND p.pzn <> ''
              AND COALESCE(p.absatz_6m, 0) > 0
              AND COALESCE(a.datenquelle, 'NMG') IN ({placeholders})
              AND a.datum >= date('now', '-{int(monate)} months')
            GROUP BY p.pzn
        ),
        ad_best AS (
            -- pro PZN nur EIN Austausch-Eintrag aus der Datenbank
            SELECT pzn_alt,
                   MAX(pzn_nmg)      AS ersatz_pzn,
                   MAX(artikel_nmg)  AS ersatz_artikel,
                   MAX(freitext_austausch) AS freitext,
                   MAX(quelle)       AS quelle
            FROM tbl_austauschdatenbank
            WHERE status = 'aktiv'
              AND (COALESCE(pzn_nmg,'') <> '' OR COALESCE(freitext_austausch,'') <> '')
            GROUP BY pzn_alt
        ),
        aa_best AS (
            SELECT original_pzn,
                   MAX(nmg_pzn)          AS ersatz_pzn,
                   MAX(austauschbar_gegen) AS ersatz_artikel
            FROM tbl_austauschartikel
            WHERE (COALESCE(nmg_pzn,'') <> '' OR COALESCE(austauschbar_gegen,'') <> '')
            GROUP BY original_pzn
        )
        SELECT
            g.pzn,
            COALESCE(ast.artikel, '') AS artikelname,
            COALESCE(ast.df, '') AS df,
            COALESCE(ast.pck, '') AS pck,
            COALESCE(ast.herst, '') AS hersteller_stamm,
            g.apotheken,
            g.gesamtabsatz_6m,
            COALESCE(g.durchschnitt_absatz_6m, 0) AS durchschnitt_absatz_6m,
            COALESCE(ast.ek, 0) AS effektiver_ek,
            g.gesamtabsatz_6m * COALESCE(ast.ek, 0) AS moeglicher_gesamtumsatz,
            g.analyse_anzahl,
            g.erste_sichtung,
            g.letzte_sichtung,
            -- Ersatz: zuerst aus der Datenbank, sonst aus der alten Tabelle
            COALESCE(ad.ersatz_pzn, aa.ersatz_pzn, '') AS ersatz_pzn,
            COALESCE(ad.ersatz_artikel, aa.ersatz_artikel, ad.freitext, '') AS ersatz_artikel,
            CASE
                WHEN ad.pzn_alt IS NOT NULL THEN COALESCE(ad.quelle, 'Austauschdatenbank')
                WHEN aa.original_pzn IS NOT NULL THEN 'Alte Austauschartikel-Tabelle'
                ELSE ''
            END AS quelle
        FROM gruppiert g
        JOIN tbl_artikelstamm ast ON ast.pzn = g.pzn
        LEFT JOIN ad_best ad ON ad.pzn_alt = g.pzn
        LEFT JOIN aa_best aa ON aa.original_pzn = g.pzn
        WHERE NOT EXISTS (SELECT 1 FROM tbl_nmg_stamm nmg WHERE nmg.pzn = g.pzn)
          AND (ad.pzn_alt IS NOT NULL OR aa.original_pzn IS NOT NULL)
          -- V1.1 SP10: nur Artikel mit Namen im Artikelstamm
          AND COALESCE(ast.artikel, '') <> ''
        ORDER BY moeglicher_gesamtumsatz DESC, g.gesamtabsatz_6m DESC
    """
    con.row_factory = sqlite3.Row
    return con.execute(sql, datenquellen).fetchall()


def _produktanalyse_basis_info(con: sqlite3.Connection, kundentyp: str, monate: int = 6) -> tuple[int, int]:
    """Zaehlt Auswertungen + Positionen im gefilterten Zeitraum.
    Wird oben in der Excel als Basis-Information ausgegeben.
    """
    datenquellen = _produktanalyse_datenquellen(kundentyp)
    placeholders = ",".join("?" * len(datenquellen))
    aw = int(con.execute(
        f"""SELECT COUNT(*) FROM tbl_auswertungen
            WHERE COALESCE(datenquelle, 'NMG') IN ({placeholders})
              AND datum >= date('now', '-{int(monate)} months')""",
        datenquellen,
    ).fetchone()[0])
    pos = int(con.execute(
        f"""SELECT COUNT(*) FROM tbl_auswertungspositionen p
            JOIN tbl_auswertungen a ON a.id = p.auswertung_id
            WHERE COALESCE(a.datenquelle, 'NMG') IN ({placeholders})
              AND a.datum >= date('now', '-{int(monate)} months')""",
        datenquellen,
    ).fetchone()[0])
    return aw, pos


_SHEET_KIND_META = {
    "haupt": {
        "title": "Produktanalyse {label}: Neue Produktchancen (Artikel nicht im NMG-Stamm)",
        "desc": (
            "Nicht im NMG-Stamm und kein Austausch-Eintrag. Hersteller "
            "und EK aus dem Artikelstamm. Gesamtumsatz = Gesamtabsatz x EK."
        ),
        "color": "0B4A86",
    },
    "nmg": {
        "title": "Produktanalyse {label}: Artikel bereits im NMG-Sortiment",
        "desc": (
            "PZNs aus den Auswertungen, die im NMG-Stamm liegen. Diese "
            "Artikel sind bereits eigenes Sortiment - sinnvoll als Cross-"
            "Sell-Indikator. Hersteller bevorzugt aus NMG-Stamm."
        ),
        "color": "117a30",
    },
    "austausch": {
        "title": "Produktanalyse {label}: Austausch vorhanden",
        "desc": (
            "PZNs aus den Auswertungen, fuer die ein Austausch-Eintrag "
            "existiert (Schulbank, Biosimilar oder alte Austauschartikel-"
            "Tabelle). Ersatz-PZN und Quelle in den rechten Spalten."
        ),
        "color": "8b5a00",
    },
}


def _produktanalyse_fill_sheet(ws, label: str, rows, basis_auswertungen: int,
                                basis_positionen: int, kind: str = "haupt"):
    """V1.1 SP6: Befuellt ein Worksheet im Vorlagen-Format des Users.

    kind:
      'haupt'     -> Hauptliste (13 Spalten)
      'nmg'       -> Artikel im NMG-Stamm (13 Spalten, gleiches Layout)
      'austausch' -> Austausch vorhanden (16 Spalten: + Ersatz-PZN,
                     Ersatz-Artikel, Quelle)
    """
    meta = _SHEET_KIND_META.get(kind, _SHEET_KIND_META["haupt"])
    is_austausch = kind == "austausch"

    headers = ["Rang", "PZN", "Artikelname", "DF", "PCK", "Hersteller", "Apotheken",
               "Gesamtabsatz 6M", "Ø Absatz 6M", "möglicher Gesamtumsatz",
               "Vorkommen", "Erste Sichtung", "Letzte Sichtung"]
    widths  = [6, 12, 38, 8, 8, 18, 10, 14, 14, 20, 11, 14, 14]
    if is_austausch:
        headers += ["Ersatz-PZN", "Ersatz-Artikel", "Quelle"]
        widths  += [12, 38, 22]
    last_col = len(headers)
    end_col_letter = get_column_letter(last_col)

    ws.cell(1, 1).value = meta["title"].format(label=label)
    ws.cell(1, 1).font = Font(bold=True, size=14, color=meta["color"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)

    ws.cell(2, 1).value = f"Basis: {basis_auswertungen:,} gespeicherte Auswertungen, {basis_positionen:,} Positionen (letzte 6 Monate)".replace(",", ".")
    ws.cell(2, 1).font = Font(italic=True, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)

    ws.cell(3, 1).value = meta["desc"]
    ws.cell(3, 1).font = Font(italic=True, color="555555")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)

    for i, h in enumerate(headers, start=1):
        c = ws.cell(5, i)
        c.value = h
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=meta["color"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for rank, r in enumerate(rows, start=1):
        rw = 5 + rank
        ws.cell(rw, 1).value = rank
        ws.cell(rw, 2).value = r["pzn"]
        ws.cell(rw, 3).value = r["artikelname"]
        ws.cell(rw, 4).value = r["df"]
        ws.cell(rw, 5).value = r["pck"]
        ws.cell(rw, 6).value = r["hersteller_stamm"]
        ws.cell(rw, 7).value = r["apotheken"]
        ws.cell(rw, 8).value = round(_safe_num(r["gesamtabsatz_6m"]), 2)
        ws.cell(rw, 9).value = round(_safe_num(r["durchschnitt_absatz_6m"]), 2)
        ws.cell(rw, 10).value = round(_safe_num(r["moeglicher_gesamtumsatz"]), 2)
        ws.cell(rw, 11).value = int(r["analyse_anzahl"] or 0)
        ws.cell(rw, 12).value = (r["erste_sichtung"] or "")[:10]
        ws.cell(rw, 13).value = (r["letzte_sichtung"] or "")[:10]
        if is_austausch:
            ws.cell(rw, 14).value = r["ersatz_pzn"]
            ws.cell(rw, 15).value = r["ersatz_artikel"]
            ws.cell(rw, 16).value = r["quelle"]

    ws.freeze_panes = "A6"
    if rows:
        ws.auto_filter.ref = f"A5:{end_col_letter}{5 + len(rows)}"


def export_produktanalyse_neu(kundentyp: str = "PK", monate: int = 6) -> Path:
    """SP14/SP15: Produktanalyse nach Spezifikation des Users.

    Args:
        kundentyp: 'PK', 'ZF' oder 'PK+ZF'
        monate: Zeitfilter auf tbl_auswertungen.datum (Default 6)

    Bei 'PK+ZF' werden 3 Sheets in derselben Excel erzeugt: PK, ZF, PK+ZF.
    Sonst eine Excel mit einem Sheet.

    Liefert den Pfad zur erzeugten Excel-Datei.

    SP15: Stammdaten (Artikelname, DF, PCK, Hersteller, EK) kommen aus
    tbl_artikelstamm; Austausch-Artikel werden gefiltert.
    """
    if not DB_PATH.exists():
        init_db(DB_PATH)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # SP15: sicherstellen, dass die referenzierten Tabellen existieren:
    # - tbl_artikelstamm + ek-Spalte (Migration)
    # - tbl_austauschdatenbank (sonst SELECT crasht in der WHERE-Klausel)
    from .artikel_db import ensure_artikelstamm_table
    ensure_artikelstamm_table()
    try:
        from .db_overview import ensure_known_overview_tables
        ensure_known_overview_tables()
    except Exception:
        pass

    label = kundentyp.upper().replace(" ", "")
    if label not in ("PK", "ZF", "PK+ZF"):
        raise ValueError(f"Unbekannter Kundentyp: {kundentyp}")

    out = OUTPUT_DIR / f"Produktanalyse_{label.replace('+', '_und_')}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

    wb = Workbook()
    # Default-Sheet entfernen, wir fuegen unsere selbst an.
    wb.remove(wb.active)

    with sqlite3.connect(DB_PATH) as con:
        sheets_to_make: list[tuple[str, str]] = []  # (sheet_name, kundentyp_filter)
        if label == "PK+ZF":
            sheets_to_make = [("PK", "PK"), ("ZF", "ZF"), ("PK+ZF", "PK+ZF")]
        else:
            sheets_to_make = [(label, label)]

        for sheet_name, ktyp in sheets_to_make:
            aw, pos = _produktanalyse_basis_info(con, ktyp, monate)

            # Hauptliste: neue Produktchancen
            ws = wb.create_sheet(title=sheet_name)
            rows = _produktanalyse_rows(con, ktyp, monate)
            _produktanalyse_fill_sheet(ws, sheet_name, rows, aw, pos, kind="haupt")

            # V1.1 SP6: Im NMG-Sortiment (Excel-Sheet-Name max 31 Zeichen)
            ws_nmg = wb.create_sheet(title=f"{sheet_name} NMG-Sortiment"[:31])
            rows_nmg = _produktanalyse_rows_nmg(con, ktyp, monate)
            _produktanalyse_fill_sheet(ws_nmg, sheet_name, rows_nmg, aw, pos, kind="nmg")

            # V1.1 SP6: Austausch vorhanden
            ws_aus = wb.create_sheet(title=f"{sheet_name} Austausch"[:31])
            rows_aus = _produktanalyse_rows_austausch(con, ktyp, monate)
            _produktanalyse_fill_sheet(ws_aus, sheet_name, rows_aus, aw, pos, kind="austausch")

    wb.save(out)
    return out
