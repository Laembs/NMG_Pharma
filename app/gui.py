import os
import re
import getpass
import shutil
import sys
import sqlite3
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
from pathlib import Path
from datetime import datetime

# Zentrales Design-System (Palette, Schrift, ttk-Styles, Widgets).
from . import theme

# V1.1 SP10: tkcalendar fuer Datum-Picker. Optional - falls nicht vorhanden,
# faellt der Code auf tk.Entry mit YYYY-MM-DD-Format zurueck.
try:
    from tkcalendar import DateEntry as _TkCalDateEntry
    _HAS_TKCAL = True
except Exception:
    _TkCalDateEntry = None
    _HAS_TKCAL = False


# V1.1 SP11: Einheitliche Button-Stile fuer die App.
# Redesign: moderne, flache Optik aus dem zentralen Theme. Signaturen bleiben
# stabil, damit alle ~50 Aufrufstellen unveraendert funktionieren.
_BTN_PRIMARY = {"bg": theme.PRIMARY, "fg": "white", "activebackground": theme.PRIMARY_DARK,
                "relief": "flat", "font": (theme.FONT, 10, "bold"),
                "padx": 16, "pady": 8, "cursor": "hand2"}
_BTN_DANGER = {"bg": theme.DANGER, "fg": "white", "activebackground": theme.DANGER_DARK,
               "relief": "flat", "font": (theme.FONT, 10, "bold"),
               "padx": 16, "pady": 8, "cursor": "hand2"}
_BTN_NEUTRAL = {"bg": "#EDF1F6", "fg": theme.PRIMARY, "activebackground": "#E2E9F1",
                "relief": "flat", "borderwidth": 0,
                "font": (theme.FONT, 10, "bold"), "padx": 14, "pady": 7, "cursor": "hand2"}


def primary_button(parent, text, command, **kwargs):
    """V1.1 SP11: blauer Standard-Button (Haupt-Aktion)."""
    opts = dict(_BTN_PRIMARY); opts.update(kwargs)
    return tk.Button(parent, text=text, command=command, **opts)


def danger_button(parent, text, command, **kwargs):
    """V1.1 SP11: roter Button (destruktive Aktion)."""
    opts = dict(_BTN_DANGER); opts.update(kwargs)
    return tk.Button(parent, text=text, command=command, **opts)


def neutral_button(parent, text, command, **kwargs):
    """V1.1 SP11: heller Button (Abbrechen, Schliessen, Nebensache)."""
    opts = dict(_BTN_NEUTRAL); opts.update(kwargs)
    return tk.Button(parent, text=text, command=command, **opts)


def _close_de_popup_on_click_outside(de, event):
    """V1.1 SP12: Schliesst den DateEntry-Popup-Kalender, wenn der Klick
    ausserhalb des Entries und seines Popups war. tkcalendar schliesst
    sonst nicht automatisch wenn man irgendwohin klickt.
    """
    try:
        top = getattr(de, "_top_cal", None)
        if top is None:
            return
        try:
            if not top.winfo_ismapped():
                return
        except tk.TclError:
            return
        w = event.widget
        # Wenn der Klick innerhalb des Entries oder des Popup-Toplevels war,
        # nichts zu tun. Wandern den Widget-Baum nach oben durch.
        try:
            cur = w
            while cur is not None:
                if cur is de or cur is top:
                    return
                cur = cur.master
        except Exception:
            pass
        try:
            top.withdraw()
        except tk.TclError:
            pass
    except Exception:
        pass


def _make_date_entry(parent, textvariable, width: int = 12, **kwargs):
    """V1.1 SP10: Liefert ein DateEntry-Widget (tkcalendar) oder ein
    klassisches tk.Entry mit Datums-Hint, je nach Verfuegbarkeit.
    Beide schreiben den Wert als 'YYYY-MM-DD' in die textvariable.
    V1.1 SP12: Popup-Kalender schliesst sich, wenn ausserhalb geklickt wird.
    """
    if _HAS_TKCAL:
        de = _TkCalDateEntry(
            parent, textvariable=textvariable, date_pattern="yyyy-mm-dd",
            width=width, background="#0b4a86", foreground="white",
            borderwidth=1, headersbackground="#0b4a86",
            headersforeground="white", **kwargs,
        )
        # Globalen Click-Listener installieren - feuert auch fuer Klicks
        # auf andere Widgets im selben Toplevel. Cleanup nicht noetig, weil
        # bei DateEntry-Destruction der Handler harmlos no-op zurueckkommt.
        try:
            de.winfo_toplevel().bind_all(
                "<Button-1>",
                lambda e, _de=de: _close_de_popup_on_click_outside(_de, e),
                add="+",
            )
        except Exception:
            pass
        return de
    return tk.Entry(parent, textvariable=textvariable, width=width, **kwargs)
from copy import copy as copy_cell_style
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from .db import init_db
from .importer import import_excel
from .learning_db import import_learning_list, import_checked_auswertung
from .manual_analysis_import import import_manual_analysis_files
from .exporter import create_vorlage_export, UnknownInputFormatError
from .market import export_marktanalyse_nicht_nmg, export_marktanalyse_produktchancen, export_produktanalyse_neu
from .compare import export_abweichungsanalyse
from .historical_import import import_historical_market_folder, import_historical_market_file
from .austausch_db import import_austausch_excel, count_austauschdatenbank, add_austausch_entry
from .artikel_db import import_artikelstamm_excel, count_artikelstamm
from .wirkstoff_db import import_wirkstoff_excel, wirkstoff_count
from .vergleichssuche import search_unified, get_pzn_details
from .archiv_db import (
    archiviere_zeitraum, liste_archive, liste_analysen_im_archiv,
    excel_aus_archiv, loesche_archiv, zaehle_zeitraum, ensure_archiv_dir,
)
from .file_loader import SUPPORTED_DATA_FILETYPES, SUPPORTED_DATA_EXTENSIONS
from .db_overview import get_database_overview, format_size
from .update_manager import (
    validate_update_package, install_update_package, open_updates_folder, write_version_file,
    list_rollback_snapshots, restore_rollback_snapshot, find_newest_update,
    find_update_packages, restart_application
)
from .roadmap_db import (
    ensure_roadmap_table,
    seed_default_roadmap_items,
    add_roadmap_item,
    list_roadmap_items,
    update_roadmap_status,
)
from .migrations import run_migrations
from .config import DB_PATH, DATA_DIR, ASSETS_DIR, SAVED_ANALYSES_DIR, IMPORT_DIR, UPDATE_DIR, OUTPUT_DIR, BACKUP_DIR, LOG_DIR, jahr_quartal_pfad
from .i18n import T as _T  # SP11: dict-basierte Uebersetzung
from .backup import backup_erstellen, backup_wiederherstellen, backup_pruefen, versionsinfo, APP_VERSION, APP_VERSION_DISPLAY, backup_auto_taeglich, DB_SCHEMA_VERSION
from .protocol_manager import (
    ensure_protocol_dirs, log_event, log_exception, list_protocol_files, read_protocol_file,
    delete_protocol_file, create_support_package, open_mail_with_attachment, PROTOCOL_ROOT, DEFAULT_RECIPIENT
)


ADMIN_DB_PASSWORD = "Marc&Tino20"
ADMIN_CLEAR_TABLES = [
    ("tbl_austauschdatenbank", "Austauschdatenbank"),
    ("tbl_lernvorschlaege", "Schulbank / Lernvorschläge"),
    ("schulbank_mapping", "Schulbank-Mapping"),
    ("tbl_nmg_stamm", "NMG Stammdaten"),
    ("nmg_rabatte", "PK Rabatte"),
    ("tbl_pzn_basisdaten", "Artikelstamm / PZN-Basis"),
    ("tbl_auswertungen", "Gespeicherte Auswertungen"),
    ("tbl_auswertungspositionen", "Auswertungspositionen"),
    ("tbl_kunden_center", "Kunden"),
    ("tbl_todo_center", "ToDo-Center"),
    ("tbl_mitarbeiter", "Mitarbeiter"),
    ("tbl_mitarbeiterprofil", "Mitarbeiterprofile"),
    ("tbl_import_log", "Import-Protokoll"),
    ("tbl_rohdaten_mapping", "Rohdaten-Mapping"),
]


def _safe_name(name: str) -> str:
    name = re.sub(r"[\\/:*?\"<>|]+", "_", name.strip())
    name = re.sub(r"\s+", " ", name).strip()
    return name or "Analyse"


def _open_folder(path: Path):
    path.mkdir(parents=True, exist_ok=True)
    try:
        if sys.platform.startswith("win"):
            os.startfile(str(path))  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            os.system(f'open "{path}"')
        else:
            os.system(f'xdg-open "{path}"')
    except Exception:
        messagebox.showinfo("Ordner", f"Ordner:\n{path}")


def _open_file(path: Path):
    path = Path(path)
    try:
        if sys.platform.startswith("win"):
            os.startfile(str(path))  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            os.system(f'open "{path}"')
        else:
            os.system(f'xdg-open "{path}"')
    except Exception:
        messagebox.showinfo("Datei", f"Datei:\n{path}")


def _show_output_actions(title: str, path):
    path = Path(path)
    text = f"{title} wurde erstellt:\n{path}\n\nDatei jetzt öffnen?"
    if messagebox.askyesno(title, text):
        _open_file(path)
    elif messagebox.askyesno(title, "Stattdessen den Ordner öffnen?"):
        _open_folder(path.parent)


def datetime_from_mtime(path: Path) -> str:
    try:
        from datetime import datetime
        return datetime.fromtimestamp(path.stat().st_mtime).strftime("%d.%m.%Y %H:%M:%S")
    except Exception:
        return ""


def _dq_label(dq) -> str:
    """Anzeige-Label fuer eine gespeicherte datenquelle.

    Intern ist PK weiterhin 'NMG' (Altdaten); die Zukunftswerk-Datenquelle ist 'ZW'
    (frueher 'ZF', das als Alt-Token noch toleriert wird).
    """
    val = (dq or "NMG")
    if val == "NMG":
        return "PK"
    if val in ("ZW", "ZF"):
        return "ZW"
    return str(val)


class NMGApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"NMGone {APP_VERSION_DISPLAY}")
        self.geometry("1040x640")
        self.minsize(980, 600)
        try:
            self.state("zoomed")
        except Exception:
            try:
                self.attributes("-zoomed", True)
            except Exception:
                pass
        self.input_file = tk.StringVar()
        self.apotheke = tk.StringVar()
        self.status = tk.StringVar(value="Bereit.")
        self.bearbeiter = getpass.getuser()
        try:
            ensure_protocol_dirs()
            log_event("programm", "Programmstart", f"Version: {APP_VERSION} | Datenbank: {DB_PATH}", user=self.bearbeiter)
        except Exception:
            pass
        init_db(DB_PATH)
        try:
            run_migrations(DB_PATH)
            write_version_file()
        except Exception:
            pass

        try:
            ensure_roadmap_table()
            seed_default_roadmap_items()
        except Exception:
            pass
        SAVED_ANALYSES_DIR.mkdir(parents=True, exist_ok=True)
        UPDATE_DIR.mkdir(parents=True, exist_ok=True)
        self.auto_backup_result = backup_auto_taeglich(7)
        self.pending_update = None
        try:
            self.pending_update = find_newest_update()
        except Exception:
            self.pending_update = None
        self._build()
        self._roadmap_mark_v11_erledigt()
        self._roadmap_mark_status_ui_v15_erledigt()
        self._check_mitarbeiterprofil_pflicht()
        self._check_daten_benachrichtigungen()
        self.after(1000, self._check_apotheken_analyse_faellig)
        # SP26: nach Start pruefen, ob wichtige Wissens-Tabellen leer sind.
        self.after(1500, self._check_wissensbasis_leer)

        # Admin-Knoepfe (z.B. in Datenbankuebersicht) standardmaessig versteckt.
        # Ctrl+Alt+A toggelt die Sichtbarkeit. Wer den Shortcut nicht kennt,
        # sieht die destruktiven Aktionen erst gar nicht.
        self._admin_visible = False
        self.bind_all("<Control-Alt-a>", self._toggle_admin_visible)
        self.bind_all("<Control-Alt-A>", self._toggle_admin_visible)

        # SP5: erst lokalen updates/-Ordner pruefen (instant), dann online.
        # So findet die App auch von Admin per OneDrive vorgelegte Setups
        # offline und bietet sie zur Installation an.
        self.after(2000, self._startup_update_check)

    @staticmethod
    def _version_tuple(value):
        """'1.0.2' / 'v1.0.2' / '1.0.2.0' -> (1, 0, 2, 0). Robust gegen Praefix/Suffix."""
        parts = []
        for piece in str(value).lstrip("vV").split("."):
            digits = "".join(c for c in piece if c.isdigit())
            parts.append(int(digits) if digits else 0)
        while len(parts) < 4:
            parts.append(0)
        return tuple(parts[:4])

    @staticmethod
    def _parse_setup_filename(name):
        """'NMGone_Setup_1_0_5.exe' -> (1, 0, 5, 0). None wenn nicht parsbar.
        Wird benutzt um vorhandene Setups in updates/ nach Version zu sortieren.
        """
        import re
        m = re.match(r"NMGone_Setup_(\d+)_(\d+)_(\d+)\.exe$", str(name), re.IGNORECASE)
        if not m:
            return None
        return tuple(int(g) for g in m.groups()) + (0,)

    def _find_pending_local_setup(self):
        """Sucht in UPDATE_DIR nach einem NMGone_Setup_*.exe mit Version > APP_VERSION.
        Liefert (path, version_str) oder None. Mehrere Treffer -> neueste zuerst.
        """
        try:
            UPDATE_DIR.mkdir(parents=True, exist_ok=True)
            candidates = []
            current = self._version_tuple(APP_VERSION)
            for p in UPDATE_DIR.glob("NMGone_Setup_*.exe"):
                v = self._parse_setup_filename(p.name)
                if v is None or v <= current:
                    continue
                candidates.append((p, v))
            if not candidates:
                return None
            candidates.sort(key=lambda x: x[1], reverse=True)
            path, v = candidates[0]
            return (path, ".".join(str(x) for x in v[:3]))
        except Exception:
            return None

    def _launch_setup_and_exit(self, setup_path):
        """Setup im Hintergrund starten und die App beenden. Inno uebernimmt
        das Tauschen der Programmdateien; durch postinstall startet NMGone
        automatisch wieder. Wird vom Online-Download UND vom Local-Pending-Check
        gleichermassen benutzt.
        """
        import subprocess
        try:
            flags = 0
            if sys.platform == "win32":
                flags = (
                    getattr(subprocess, "DETACHED_PROCESS", 0)
                    | getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)
                )
            subprocess.Popen(
                [str(setup_path), "/SILENT", "/SUPPRESSMSGBOXES", "/CLOSEAPPLICATIONS"],
                close_fds=True,
                creationflags=flags,
            )
        except Exception as exc:
            messagebox.showerror(
                "Update",
                f"Setup konnte nicht gestartet werden:\n{exc}\n\n"
                f"Datei liegt unter:\n{setup_path}"
            )
            return False
        self.after(800, self.destroy)
        return True

    def _startup_update_check(self):
        """SP5: erst lokal (UPDATE_DIR), dann online (GitHub).
        Lokal = instant + offline-faehig. Online = falls nichts lokal liegt.
        """
        pending = self._find_pending_local_setup()
        if pending:
            path, version_str = pending
            if messagebox.askyesno(
                "Lokales Update verfuegbar",
                f"Im Update-Ordner liegt ein heruntergeladenes Setup bereit:\n\n"
                f"Version: V{version_str}\n"
                f"Aktuell installiert: V{APP_VERSION}\n"
                f"Datei: {path.name}\n\n"
                f"Jetzt installieren?"
                f"{self.SMARTSCREEN_HINT}"
            ):
                self._launch_setup_and_exit(path)
            # Wer "Nein" sagt: kein Online-Check (gleiche Version).
            return
        # Nichts lokal -> online checken
        self._check_online_update_on_startup()

    @staticmethod
    def _fetch_latest_release_info(timeout=8):
        """Holt latest-release-Info von GitHub. Liefert (tag, html_url, asset_url)
        oder (None, None, None). Wird vom Startup-Check und vom Menue-Knopf
        'Update suchen' gleichermassen genutzt.
        """
        import urllib.request, json
        url = "https://api.github.com/repos/Laembs/NMG_Pharma/releases/latest"
        req = urllib.request.Request(url, headers={
            "Accept": "application/vnd.github+json",
            "User-Agent": f"NMGone/{APP_VERSION}",
        })
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        tag = str(data.get("tag_name", "")).strip()
        html_url = str(data.get("html_url", "")).strip()
        asset_url = None
        for asset in data.get("assets", []) or []:
            name = str(asset.get("name", "")).lower()
            if name.endswith(".exe"):
                asset_url = str(asset.get("browser_download_url", "")).strip()
                break
        return (tag or None, html_url or None, asset_url or None)

    def _check_online_update_on_startup(self):
        """Beim Start asynchron checken. Wenn neuer als APP_VERSION:
        Dialog mit Direkt-Install-Option. Wer 'Nein' sagt, wird beim
        naechsten Start wieder gefragt (kein State gespeichert).
        """
        import threading
        def worker():
            try:
                tag, html_url, asset_url = self._fetch_latest_release_info(timeout=8)
                if not tag:
                    return
                if self._version_tuple(tag) > self._version_tuple(APP_VERSION):
                    self.after(0, lambda: self._show_online_update_dialog(tag, html_url, asset_url))
            except Exception:
                # Offline / Rate Limit -> beim naechsten Start wieder versuchen.
                pass
        threading.Thread(target=worker, daemon=True).start()

    def _show_online_update_dialog(self, tag, html_url, asset_url=None):
        """Bietet Direkt-Install an. Wenn aus irgendeinem Grund kein
        .exe-Asset gefunden wurde (defektes Release), Browser-Fallback.
        """
        if not asset_url:
            import webbrowser
            if messagebox.askyesno(
                "Update verfuegbar",
                f"Neue Version: {tag}\nAktuell: V{APP_VERSION}\n\n"
                f"Konnte das Setup nicht direkt finden. Download-Seite oeffnen?"
            ):
                try:
                    webbrowser.open(html_url or "https://github.com/Laembs/NMG_Pharma/releases")
                except Exception:
                    pass
            return

        if messagebox.askyesno(
            "Update verfuegbar",
            f"Eine neuere Version von NMGone ist verfuegbar: {tag}\n"
            f"Aktuell installiert: V{APP_VERSION}\n\n"
            f"Jetzt herunterladen und installieren?\n\n"
            f"NMGone schliesst sich automatisch, der Setup-Assistent startet,\n"
            f"deine Daten unter C:\\ProgramData\\NMGone bleiben unberuehrt."
            f"{self.SMARTSCREEN_HINT}"
        ):
            self._download_and_install_update(asset_url, tag)

    def _download_and_install_update(self, asset_url, tag):
        """SP5: Laedt Setup nach UPDATE_DIR (nicht mehr %TEMP%) und bietet
        dann Jetzt/Spaeter an. 'Spaeter' laesst die Datei liegen; beim
        naechsten Programmstart wird das Setup vom local-check gefunden.
        """
        import threading
        busy = self._show_busy_modal(f"Update {tag} herunterladen", "Verbindung zu GitHub...")

        def worker():
            try:
                import urllib.request
                UPDATE_DIR.mkdir(parents=True, exist_ok=True)
                safe_tag = tag.lstrip("vV").replace(".", "_")
                target = UPDATE_DIR / f"NMGone_Setup_{safe_tag}.exe"

                req = urllib.request.Request(asset_url, headers={
                    "User-Agent": f"NMGone/{APP_VERSION}",
                    "Accept": "application/octet-stream",
                })
                with urllib.request.urlopen(req, timeout=60) as resp:
                    total = int(resp.headers.get("Content-Length", 0))
                    with open(target, "wb") as fh:
                        downloaded = 0
                        block = 128 * 1024
                        while True:
                            chunk = resp.read(block)
                            if not chunk:
                                break
                            fh.write(chunk)
                            downloaded += len(chunk)
                            if total:
                                pct = int(100 * downloaded / total)
                                mb_done = downloaded / 1024 / 1024
                                mb_total = total / 1024 / 1024
                                self.after(0, lambda p=pct, md=mb_done, mt=mb_total:
                                    self._set_busy_message(busy, f"Download {p}% ({md:.1f} / {mt:.1f} MB)"))
                            else:
                                mb_done = downloaded / 1024 / 1024
                                self.after(0, lambda md=mb_done:
                                    self._set_busy_message(busy, f"Heruntergeladen: {md:.1f} MB"))

                if target.stat().st_size < 1_000_000:
                    raise ValueError(f"Download zu klein ({target.stat().st_size} bytes) - vermutlich fehlgeschlagen")

                final_path = target
                final_size = target.stat().st_size

                def offer_install():
                    self._close_busy_modal(busy)
                    # Yes = Jetzt, No = Spaeter (Datei bleibt), Cancel = loeschen
                    answer = messagebox.askyesnocancel(
                        "Update bereit",
                        f"Download abgeschlossen ({final_size / 1024 / 1024:.1f} MB).\n\n"
                        f"Setup liegt unter:\n{final_path}\n\n"
                        f"JA = jetzt installieren (NMGone beendet sich kurz und oeffnet sich danach wieder)\n"
                        f"NEIN = spaeter installieren (Setup bleibt im updates-Ordner und wird beim naechsten Start angeboten)\n"
                        f"ABBRECHEN = Datei wieder loeschen"
                        f"{self.SMARTSCREEN_HINT}"
                    )
                    if answer is True:
                        self._launch_setup_and_exit(final_path)
                    elif answer is False:
                        self.status.set(f"Update {tag} heruntergeladen. Installation auf spaeter verschoben.")
                    else:
                        try:
                            final_path.unlink()
                            self.status.set("Update-Datei wurde wieder geloescht.")
                        except Exception:
                            pass
                self.after(0, offer_install)

            except Exception as exc:
                error_text = f"{type(exc).__name__}: {exc}"
                def failed():
                    self._close_busy_modal(busy)
                    messagebox.showerror(
                        "Update fehlgeschlagen",
                        f"Konnte das Update nicht herunterladen:\n\n{error_text}\n\n"
                        f"Du kannst das Update auch manuell holen:\n"
                        f"https://github.com/Laembs/NMG_Pharma/releases/tag/{tag}"
                    )
                self.after(0, failed)

        threading.Thread(target=worker, daemon=True).start()

    def _show_busy_modal(self, title, subtitle="Bitte warten..."):
        """Redesign: KEIN modaler Mittendialog mehr - stattdessen die Status-Box
        rechts oben (gleiche Animation wie _run_background). Die UI bleibt waehrend
        des Laufs bedienbar. Rueckgabe ist API-kompatibel: ein Handle mit
        ._msg_var/._pb, damit _set_busy_message/_close_busy_modal (und damit alle
        bestehenden Import-/Export-Aufrufer) unveraendert weiterfunktionieren.
        """
        ctx = self._bg_create_widget(title, subtitle)
        box = ctx["box"]
        box._msg_var = ctx["sub_var"]
        box._pb = ctx["pb"]
        return box

    def _set_busy_message(self, win, text):
        try:
            if win and win.winfo_exists():
                win._msg_var.set(text)
                self.update_idletasks()
        except Exception:
            pass

    def _close_busy_modal(self, win):
        try:
            if win and win.winfo_exists():
                try:
                    win._pb.stop()
                except Exception:
                    pass
                try:
                    win.grab_release()
                except Exception:
                    pass
                win.destroy()
        except Exception:
            pass
        try:
            self.config(cursor="")
        except Exception:
            pass

    # V1.1 SP15: Background-Jobs ---------------------------------------------
    def _bg_status_toast(self, text: str, fg: str = "#0a7d2c", duration_ms: int = 8000):
        """Zeigt eine kurze Erfolgs-/Fehler-Meldung in der Status-Leiste unten.
        Nach `duration_ms` wird der Status auf 'bereit' zurueckgesetzt.
        """
        try:
            self.status.set(text)
        except Exception:
            return
        # Toast-Auto-Reset
        token = getattr(self, "_status_toast_token", 0) + 1
        self._status_toast_token = token

        def _reset():
            if getattr(self, "_status_toast_token", -1) == token:
                try:
                    self.status.set("")
                except Exception:
                    pass
        try:
            self.after(duration_ms, _reset)
        except Exception:
            pass

    def _bg_create_widget(self, title: str, subtitle: str):
        """Erzeugt eine Mini-Status-Box. Liefert dict mit Tk-Variablen.

        SP19: _bg_jobs_panel kann fehlen wenn die rechte Sidebar eingeklappt
        ist oder gerade neu aufgebaut wird (Item-end-Probleme). Vorher knallte
        das hier mit AttributeError/TclError und der Job startete nicht -
        der Bug war seit SP18 verantwortlich dafuer dass "Neue Auswertung"
        scheinbar nichts mehr macht. Jetzt: lazy attach am Hauptfenster.
        """
        panel = getattr(self, "_bg_jobs_panel", None)
        try:
            alive = bool(panel and panel.winfo_exists())
        except Exception:
            alive = False
        if not alive:
            panel = tk.Frame(self, bg=theme.BG)
            panel.place(relx=1.0, x=-18, y=18, anchor="ne")
            self._bg_jobs_panel = panel
        box = tk.Frame(panel, bg=theme.CARD,
                       highlightbackground=theme.BORDER, highlightthickness=1)
        box.pack(side="top", anchor="ne", padx=0, pady=(0, 8))

        title_var = tk.StringVar(value=title)
        sub_var = tk.StringVar(value=subtitle)
        head = tk.Frame(box, bg=theme.CARD)
        head.pack(fill="x", padx=12, pady=(10, 0))
        tk.Label(head, text="⏳", bg=theme.CARD, fg=theme.ACCENT,
                 font=(theme.FONT, 11)).pack(side="left")
        tk.Label(head, textvariable=title_var, font=(theme.FONT, 10, "bold"),
                 fg=theme.INK, bg=theme.CARD, anchor="w").pack(side="left", padx=(6, 0))
        tk.Label(box, textvariable=sub_var, font=(theme.FONT, 9),
                 fg=theme.MUTED, bg=theme.CARD, anchor="w", wraplength=300,
                 justify="left").pack(anchor="w", padx=12, pady=(2, 6))
        pb = ttk.Progressbar(box, mode="indeterminate", length=290,
                             style="NMG.Horizontal.TProgressbar")
        pb.pack(padx=12, pady=(0, 12))
        pb.start(12)
        return {"box": box, "title_var": title_var, "sub_var": sub_var, "pb": pb}

    def _bg_destroy_widget(self, ctx):
        try:
            ctx["pb"].stop()
        except Exception:
            pass
        try:
            ctx["box"].destroy()
        except Exception:
            pass

    def _run_background(self, work_fn, title: str = "NMGone arbeitet",
                         subtitle: str = "Laeuft im Hintergrund ...",
                         progress: bool = True, on_done=None,
                         on_error=None):
        """V1.1 SP15: Startet work_fn in einem Hintergrund-Thread, OHNE Modal-
        Dialog. Stattdessen erscheint rechts oben im Header eine Mini-Status-
        Box mit Titel + Subtitle + indeterminate Progressbar.

        Wenn progress=True, bekommt work_fn ein progress(text)-Callable als
        erstes Argument (thread-safe via self.after).

        on_done(result) wird im UI-Thread aufgerufen wenn fertig.
        on_error(exc) wird im UI-Thread aufgerufen bei Exception.

        UI bleibt waehrend des Jobs klickbar - mehrere Jobs koennen parallel
        laufen und stapeln sich vertikal.
        """
        import threading
        ctx = self._bg_create_widget(title, subtitle)

        def update_progress(text):
            try:
                self.after(0, lambda t=text: ctx["sub_var"].set(t))
            except Exception:
                pass

        def runner():
            try:
                if progress:
                    result = work_fn(update_progress)
                else:
                    result = work_fn()
            except BaseException as exc:
                # SP19: Worker-Exception immer ins Fehler-Log schreiben.
                # Vorher konnten Fehler in work_fn unsichtbar verschwinden.
                # WICHTIG: 'exc' wird am Ende des except-Blocks von Python
                # geloescht. Da _on_err erst spaeter (via after) laeuft, muss die
                # Exception in einer eigenen Variable festgehalten werden - sonst
                # NameError im Handler und der echte Fehler wird verschluckt.
                worker_exc = exc
                try:
                    self._log_error(title, "Background-Worker-Fehler", worker_exc)
                except Exception:
                    pass
                def _on_err(worker_exc=worker_exc):
                    self._bg_destroy_widget(ctx)
                    if on_error:
                        try:
                            on_error(worker_exc)
                        except Exception as cb_exc:
                            # SP19: on_error-Callback-Fehler nicht schlucken.
                            try:
                                self._log_error(title, "on_error-Callback-Fehler", cb_exc)
                            except Exception:
                                pass
                            try:
                                messagebox.showerror(title, f"{worker_exc}\n\n(Folgefehler im on_error-Handler: {cb_exc})")
                            except Exception:
                                pass
                    else:
                        self._bg_status_toast(f"{title}: Fehler - {worker_exc}", fg="#9b1c1c")
                try:
                    self.after(0, _on_err)
                except Exception:
                    pass
                return

            def _on_ok():
                self._bg_destroy_widget(ctx)
                if on_done:
                    try:
                        on_done(result)
                    except Exception as cb_exc:
                        # SP19: on_done-Callback-Fehler nicht schlucken.
                        # Vorher: 'Auswertung erstellt' aber post_export crasht
                        # in shutil.copy2/Vorlage -> User sieht stumm nichts.
                        try:
                            self._log_error(title, "on_done-Callback-Fehler", cb_exc)
                        except Exception:
                            pass
                        try:
                            messagebox.showerror(title, f"Nachverarbeitung fehlgeschlagen:\n{cb_exc}")
                        except Exception:
                            pass
                else:
                    self._bg_status_toast(f"{title}: fertig")
            try:
                self.after(0, _on_ok)
            except Exception:
                pass

        threading.Thread(target=runner, daemon=True).start()

    def _run_busy(self, work_fn, title="NMGone arbeitet", subtitle="Bitte warten ...",
                   progress: bool = False):
        """SP20: Fuehrt work_fn() in einem Hintergrund-Thread aus und zeigt
        waehrend der Ausfuehrung die Status-Box rechts oben mit animierter
        Progressbar (gleiches Design wie _run_background). Die UI bleibt
        bedienbar. Gibt den Rueckgabewert von work_fn zurueck oder wirft die
        Exception weiter. Backend-Funktionen muessen ihre eigenen DB-
        Verbindungen erstellen (sqlite3 standardmaessig thread-bound).

        V1.1 SP14: Wenn progress=True, bekommt work_fn ein progress(text)-
        Callable als erstes Argument. Damit kann das Backend die Subtitle-
        Anzeige aktualisieren (z.B. 'Datei 5 von 96'). Aufrufer:
            self._run_busy(lambda update: my_fn(args, progress=update),
                           title=..., subtitle=..., progress=True)
        """
        import threading
        state = {"value": None, "exc": None}
        done_flag = tk.IntVar(master=self, value=0)
        # Redesign: gleiche Hintergrund-Status-Box rechts oben wie _run_background
        # (statt modalem Mittendialog). Die UI bleibt waehrend des Laufs bedienbar;
        # die synchrone Rueckgabe bleibt erhalten, weil wait_variable den Tk-Event-
        # Loop weiter pumpt. Dadurch bekommen alle bestehenden Aufrufstellen
        # (Importe/Exporte) automatisch dasselbe Design - ohne Einzelumbau.
        ctx = self._bg_create_widget(title, subtitle)

        def update_progress(text):
            # Thread-safe: GUI nur ueber self.after vom Worker aus aendern.
            try:
                self.after(0, lambda t=text: ctx["sub_var"].set(t))
            except Exception:
                pass

        def runner():
            try:
                if progress:
                    state["value"] = work_fn(update_progress)
                else:
                    state["value"] = work_fn()
            except BaseException as exc:
                state["exc"] = exc
            finally:
                try:
                    self.after(0, lambda: done_flag.set(1))
                except Exception:
                    pass

        threading.Thread(target=runner, daemon=True).start()
        try:
            self.wait_variable(done_flag)
        finally:
            self._bg_destroy_widget(ctx)

        if state["exc"] is not None:
            raise state["exc"]
        return state["value"]

    def _toggle_admin_visible(self, event=None):
        self._admin_visible = not self._admin_visible
        zustand = "EIN" if self._admin_visible else "AUS"
        hinweis = "Datenbankuebersicht oeffnen, um Admin-Knoepfe zu sehen." if self._admin_visible else ""
        try:
            self.status.set(f"Admin-Modus {zustand}. {hinweis}".strip())
        except Exception:
            pass

    def _build(self):
        self.configure(bg=theme.BG)
        theme.apply_theme(self)
        theme.apply_widget_defaults(self)
        self.columnconfigure(1, weight=1)
        self.rowconfigure(0, weight=1)

        # Linke Navigation (Redesign: dunkle Sidebar aus dem zentralen Theme).
        self.sidebar = theme.Sidebar(self, width=252, title="NMGone",
                                     subtitle="Pharma-Analyse")
        self.sidebar.grid(row=0, column=0, sticky="ns")
        # NMGone-Logo oben in der Sidebar (auf weisser Karte, weil dunkelblau).
        try:
            _logo_path = ASSETS_DIR / "NMGone.png"
            if _logo_path.exists():
                _raw = tk.PhotoImage(file=str(_logo_path))
                _factor = max(1, round(max(_raw.width() / 210, _raw.height() / 130)))
                self._sidebar_logo = _raw.subsample(_factor, _factor)
                self.sidebar.set_logo(self._sidebar_logo)
        except Exception:
            pass
        self._build_nav_tree()
        self.sidebar.add_footer_note(f"Datenbank:\n{DB_PATH.name}")

        # Hauptbereich
        main = tk.Frame(self, bg=theme.BG)
        main.grid(row=0, column=1, sticky="nsew")
        main.columnconfigure(0, weight=1)
        main.columnconfigure(1, weight=0)
        main.rowconfigure(1, weight=1)

        # Oberer Programmkopf entfernt: Startseite soll direkt mit dem Arbeitsbereich beginnen.
        # V1.1 SP18: _bg_jobs_panel ist jetzt in der rechten Sidebar (ueber dem
        # Backup-Status), nicht mehr im Header. Header bleibt leer.
        header = tk.Frame(main, bg=theme.BG)
        header.grid(row=0, column=0, columnspan=2, sticky="ew", padx=0, pady=0)

        self.page = tk.Frame(main, bg=theme.CARD, highlightbackground=theme.BORDER, highlightthickness=1)
        self.page.grid(row=1, column=0, sticky="nsew", padx=(22, 12), pady=10)
        self.page.columnconfigure(0, weight=1)
        self.page.rowconfigure(0, weight=0)
        self.page.rowconfigure(1, weight=1)

        self._main_frame = main

        self._right_panel = tk.Frame(main, bg=theme.BG, width=280)
        self._right_panel_visible = self._get_meta_value("sidebar_visible", "1") != "0"
        if self._right_panel_visible:
            self._right_panel.grid(row=1, column=1, sticky="ns", padx=(0, 22), pady=10)
            self._status_card(self._right_panel)

        footer = tk.Frame(main, bg=theme.BG)
        footer.grid(row=2, column=0, columnspan=2, sticky="ew", padx=22, pady=(0, 10))
        tk.Label(footer, textvariable=self.status, bg=theme.BG, fg=theme.SUCCESS, anchor="w", justify="left").pack(side="left", fill="x", expand=True)
        self._sidebar_btn_text = tk.StringVar(value=("◀ Info ausblenden" if self._right_panel_visible else "▶ Info einblenden"))
        tk.Button(
            footer,
            textvariable=self._sidebar_btn_text,
            command=self._toggle_sidebar,
            bg="#EDF1F6",
            fg=theme.PRIMARY,
            relief="flat",
            font=(theme.FONT, 9, "bold"),
            cursor="hand2",
            padx=12,
            pady=4,
        ).pack(side="right", padx=(0, 4))

        if self.pending_update:
            self.status.set(f"Update verfügbar: Version {self.pending_update.get('target_version')} ({self.pending_update.get('name')}). Links auf 'Update installieren' klicken.")

        self.navigate("startseite")
        self.show_startseite()

    def _make_window(self, title, geometry="800x600", minsize=None, maximized=False, parent=None):
        """Erstellt ein Toplevel-Fenster mit Windows-Stil (Minimieren/Maximieren/Schließen)."""
        win = tk.Toplevel(parent or self)
        win.resizable(True, True)
        win.title(title)
        win.geometry(geometry)
        if minsize:
            win.minsize(*minsize)
        win.configure(bg="#f5f7fb")
        # Windows-Fensterrahmen mit allen drei Schaltflächen
        win.resizable(True, True)
        if parent:
            win.transient(parent)
        win.grab_set()
        if maximized:
            try:
                win.state("zoomed")
            except Exception:
                pass
        return win

    def _refresh_status_sidebar(self):
        """SP29: Rechte Status-Sidebar neu zeichnen. Wird nach Imports gerufen,
        damit die Counter (z.B. 'PK Rabatte (N Eintr., M mit Wert)') ohne
        App-Neustart aktuelle Zahlen zeigen.
        """
        try:
            if not getattr(self, "_right_panel_visible", False):
                return
            for w in self._right_panel.winfo_children():
                w.destroy()
            self._status_card(self._right_panel)
        except Exception:
            pass

    def _toggle_sidebar(self):
        """Rechte Seitenleiste ein- oder ausblenden. Zustand wird persistiert."""
        if self._right_panel_visible:
            self._right_panel.grid_remove()
            self._right_panel_visible = False
            self._sidebar_btn_text.set("▶ Info einblenden")
            self._set_meta_value("sidebar_visible", "0")
        else:
            for w in self._right_panel.winfo_children():
                w.destroy()
            self._status_card(self._right_panel)
            self._right_panel.grid(row=1, column=1, sticky="ns", padx=(0, 22), pady=10)
            self._right_panel_visible = True
            self._sidebar_btn_text.set("◀ Info ausblenden")
            self._set_meta_value("sidebar_visible", "1")

    # ── ADMIN-LOGINS (kein Pflichtprofil, kein Pflichtdialog) ──────────────────
    ADMIN_LOGINS = {"laemb", "jagdeal", "user"}

    # SP6: Hinweistext wird an alle Update-Dialoge gehaengt, in denen ein
    # Setup gestartet werden koennte. Weil wir kein Code-Signing-Cert haben,
    # warnt SmartScreen jedes Mal "Unbekannter Herausgeber" - der User
    # braucht die Klick-Anleitung.
    SMARTSCREEN_HINT = (
        "\n\nHINWEIS - Windows zeigt beim Setup-Start ggf. die Meldung\n"
        "\"Der Computer wurde durch Windows geschuetzt\". Das ist normal,\n"
        "weil das Setup (noch) kein Code-Signing-Zertifikat hat.\n\n"
        "Schritte zum Fortfahren:\n"
        "  1. Klicke oben auf \"Weitere Informationen\"\n"
        "  2. Klicke unten auf den Button \"Trotzdem ausfuehren\""
    )

    def _is_admin_login(self):
        return (self.bearbeiter or "").strip().lower() in self.ADMIN_LOGINS

    # ── MITARBEITERPROFIL ────────────────────────────────────────────────────────
    def _ensure_mitarbeiterprofil_table(self):
        with sqlite3.connect(DB_PATH) as con:
            con.execute("""\nCREATE TABLE IF NOT EXISTS tbl_mitarbeiterprofil (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,\nlogin TEXT NOT NULL UNIQUE,
                    vorname TEXT NOT NULL DEFAULT '',\nnachname TEXT NOT NULL DEFAULT '',
                    telefon TEXT,\nmobil TEXT,
                    email TEXT,\nabteilung TEXT,
                    position TEXT,\nnotizen TEXT,
                    erstellt_am TEXT DEFAULT CURRENT_TIMESTAMP,\ngeaendert_am TEXT
                )\n""")
            con.commit()

    def _get_mitarbeiterprofil(self, login=None):
        login = (login or self.bearbeiter or "").strip()
        self._ensure_mitarbeiterprofil_table()
        with sqlite3.connect(DB_PATH) as con:
            con.row_factory = sqlite3.Row
            return con.execute("SELECT * FROM tbl_mitarbeiterprofil WHERE login=?", (login,)).fetchone()

    def _check_mitarbeiterprofil_pflicht(self):
        """Beim Start: Wenn kein Admin und kein Profil vorhanden → Pflicht-Dialog."""
        if self._is_admin_login():
            return
        self._ensure_mitarbeiterprofil_table()
        profil = self._get_mitarbeiterprofil()
        if not profil or not str(profil["vorname"]).strip() or not str(profil["nachname"]).strip():
            self.after(300, self._open_mitarbeiterprofil_pflicht_dialog)

    def _open_mitarbeiterprofil_pflicht_dialog(self):
        win = tk.Toplevel(self)
        win.title("Bitte Mitarbeiterprofil ausfüllen")
        win.geometry("480x300")
        win.resizable(False, False)
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()
        win.protocol("WM_DELETE_WINDOW", lambda: None)  # nicht schließbar

        tk.Label(win, text="👤  Mitarbeiterprofil", font=(theme.FONT, 16, "bold"), fg="#0b4a86", bg="#f5f7fb").pack(anchor="w", padx=22, pady=(18, 4))
        tk.Label(win, text=(
            f"Hallo {self.bearbeiter}!\n"
            "Bitte Vorname und Nachname einmalig eintragen.\n"
            "Ohne Profil kann das Programm nicht genutzt werden."
        ), font=(theme.FONT, 10), fg="#444", bg="#f5f7fb", justify="left").pack(anchor="w", padx=22, pady=(0, 12))

        form = tk.Frame(win, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        form.pack(fill="x", padx=22, pady=(0, 12))
        form.columnconfigure(1, weight=1)

        vorname_var = tk.StringVar()
        nachname_var = tk.StringVar()
        existing = self._get_mitarbeiterprofil()
        if existing:
            vorname_var.set(str(existing["vorname"] or ""))
            nachname_var.set(str(existing["nachname"] or ""))

        for r, (label, var) in enumerate([("Vorname *", vorname_var), ("Nachname *", nachname_var)]):
            tk.Label(form, text=label, bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 10, "bold")).grid(row=r, column=0, sticky="w", padx=14, pady=10)
            tk.Entry(form, textvariable=var, font=(theme.FONT, 12)).grid(row=r, column=1, sticky="ew", padx=14, pady=10)

        err_var = tk.StringVar()
        tk.Label(win, textvariable=err_var, fg="#c00", bg="#f5f7fb", font=(theme.FONT, 9)).pack(anchor="w", padx=22)

        def save():
            vn = vorname_var.get().strip()
            nn = nachname_var.get().strip()
            if not vn or not nn:
                err_var.set("Vorname und Nachname sind Pflichtfelder.")
                return
            login = (self.bearbeiter or "").strip()
            with sqlite3.connect(DB_PATH) as con:
                con.execute("""\nINSERT INTO tbl_mitarbeiterprofil(login, vorname, nachname)
                    VALUES(?,?,?)\nON CONFLICT(login) DO UPDATE SET
                        vorname=excluded.vorname,\nnachname=excluded.nachname,
                        geaendert_am=CURRENT_TIMESTAMP\n""", (login, vn, nn))
                con.commit()
            win.destroy()

        tk.Button(win, text="✔  Speichern", command=save, bg="#0b4a86", fg="white",
                  relief="flat", font=(theme.FONT, 12, "bold"), padx=20, pady=8).pack(pady=(4, 18))

    def show_mitarbeiterprofil_dialog(self, login=None, readonly=False):
        """Profil anzeigen/bearbeiten. readonly=True → nur ansehen."""
        self._ensure_mitarbeiterprofil_table()
        ziel_login = (login or self.bearbeiter or "").strip()
        is_own = ziel_login.lower() == (self.bearbeiter or "").strip().lower()
        can_edit = is_own or self._is_admin_login()
        title = f"Mitarbeiterprofil: {ziel_login}"

        win = tk.Toplevel(self)
        win.resizable(True, True)
        win.title(title)
        win.geometry("520x420")
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()

        tk.Label(win, text=f"👤  {title}", font=(theme.FONT, 15, "bold"), fg="#0b4a86", bg="#f5f7fb").pack(anchor="w", padx=22, pady=(18, 4))
        if not can_edit:
            tk.Label(win, text="Nur ansehen — du kannst nur dein eigenes Profil bearbeiten.", font=(theme.FONT, 9), fg="#888", bg="#f5f7fb").pack(anchor="w", padx=22)

        form = tk.Frame(win, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        form.pack(fill="both", expand=True, padx=22, pady=(8, 10))
        form.columnconfigure(1, weight=1)

        profil = self._get_mitarbeiterprofil(ziel_login) or {}
        fields_def = [
            ("vorname", "Vorname *"), ("nachname", "Nachname *"),
            ("telefon", "Telefon"), ("mobil", "Mobil"), ("email", "E-Mail"),
            ("abteilung", "Abteilung"), ("position", "Position"),
        ]
        vars_ = {}
        for r, (key, label) in enumerate(fields_def):
            tk.Label(form, text=label, bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 10, "bold")).grid(row=r, column=0, sticky="w", padx=14, pady=6)
            var = tk.StringVar(value=str(profil[key] if profil and key in profil.keys() and profil[key] else ""))
            vars_[key] = var
            state = "normal" if can_edit else "readonly"
            tk.Entry(form, textvariable=var, state=state).grid(row=r, column=1, sticky="ew", padx=14, pady=6)

        bar = tk.Frame(win, bg="#f5f7fb")
        bar.pack(fill="x", padx=22, pady=(0, 14))
        tk.Button(bar, text="Schließen", command=win.destroy, padx=14, pady=7).pack(side="right", padx=(8, 0))
        if can_edit:
            def save_profil():
                vn = vars_["vorname"].get().strip()
                nn = vars_["nachname"].get().strip()
                if not vn or not nn:
                    messagebox.showinfo(title, "Vorname und Nachname sind Pflichtfelder.")
                    return
                with sqlite3.connect(DB_PATH) as con:
                    data = {k: v.get().strip() for k, v in vars_.items()}
                    data["login"] = ziel_login
                    con.execute("""\nINSERT INTO tbl_mitarbeiterprofil(login,vorname,nachname,telefon,mobil,email,abteilung,position)
                        VALUES(:login,:vorname,:nachname,:telefon,:mobil,:email,:abteilung,:position)\nON CONFLICT(login) DO UPDATE SET
                            vorname=excluded.vorname, nachname=excluded.nachname,\ntelefon=excluded.telefon, mobil=excluded.mobil,
                            email=excluded.email, abteilung=excluded.abteilung,\nposition=excluded.position,
                            geaendert_am=CURRENT_TIMESTAMP\n""", data)
                    con.commit()
                messagebox.showinfo(title, "Profil gespeichert.")
                win.destroy()
            tk.Button(bar, text="✔  Speichern", command=save_profil, bg="#0b4a86", fg="white",
                      relief="flat", font=(theme.FONT, 11, "bold"), padx=16, pady=7).pack(side="right")

    # ── BENACHRICHTIGUNGEN / DATEN-AMPEL ────────────────────────────────────────
    def _check_daten_benachrichtigungen(self):
        """Beim Start: Benachrichtigungen für veraltete Daten prüfen (snooze-fähig)."""
        snooze_key = "benachrichtigung_snooze_bis"
        snooze_until = self._get_meta_value(snooze_key, "")
        if snooze_until:
            try:
                from datetime import date
                if date.today().isoformat() <= snooze_until:
                    return
            except Exception:
                pass

        items = self._get_data_update_status_items()
        rote = [name for name, val, count in items if self._status_ampel_combined(val, count) in ("🔴", "⚫", "⚪")]
        if not rote:
            return

        self.after(600, lambda: self._show_benachrichtigungs_dialog(rote, snooze_key))

    def _show_benachrichtigungs_dialog(self, rote_items, snooze_key):
        from datetime import date, timedelta
        win = tk.Toplevel(self)
        win.title("Daten-Benachrichtigung")
        win.geometry("460x320")
        win.resizable(False, False)
        win.configure(bg="#fff8e1")
        win.transient(self)
        win.grab_set()

        tk.Label(win, text="⚠️  Daten-Aktualität", font=(theme.FONT, 15, "bold"), fg="#8b4513", bg="#fff8e1").pack(anchor="w", padx=20, pady=(16, 4))
        tk.Label(win, text="Folgende Datenquellen sind veraltet (>60 Tage) oder fehlen:", font=(theme.FONT, 10), fg="#555", bg="#fff8e1").pack(anchor="w", padx=20)
        for name in rote_items:
            tk.Label(win, text=f"  🔴  {name}", font=(theme.FONT, 11, "bold"), fg="#c00", bg="#fff8e1").pack(anchor="w", padx=20, pady=2)

        tk.Label(win, text="Wann soll diese Meldung erneut erscheinen?", font=(theme.FONT, 10), fg="#444", bg="#fff8e1").pack(anchor="w", padx=20, pady=(14, 6))

        btn_row = tk.Frame(win, bg="#fff8e1")
        btn_row.pack(anchor="w", padx=20)

        def snooze(tage):
            bis = (date.today() + timedelta(days=tage)).isoformat()
            self._set_meta_value(snooze_key, bis)
            win.destroy()

        def oeffne_daten():
            win.destroy()
            self.show_daten_aktualisieren_page()

        tk.Button(btn_row, text="In 7 Tagen", command=lambda: snooze(7), padx=12, pady=6).pack(side="left", padx=(0, 6))
        tk.Button(btn_row, text="In 15 Tagen", command=lambda: snooze(15), padx=12, pady=6).pack(side="left", padx=(0, 6))
        tk.Button(btn_row, text="In 30 Tagen", command=lambda: snooze(30), padx=12, pady=6).pack(side="left")

        bar = tk.Frame(win, bg="#fff8e1")
        bar.pack(fill="x", padx=20, pady=(16, 14))
        tk.Button(bar, text="Jetzt Daten aktualisieren →", command=oeffne_daten,
                  bg="#0b4a86", fg="white", relief="flat", font=(theme.FONT, 11, "bold"), padx=16, pady=7).pack(side="left")
        tk.Button(bar, text="Schließen", command=win.destroy, padx=14, pady=7).pack(side="right")

    # ── OUTLOOK-KALENDER-WIDGET ──────────────────────────────────────────────────
    def _get_outlook_termine(self, tage=14):
        """Liest Termine aus Outlook – mehrere Fallback-Methoden."""
        from datetime import datetime, timedelta
        termine = []

        # Methode 1: win32com (bevorzugt)
        try:
            import win32com.client
            outlook = win32com.client.Dispatch("Outlook.Application")
            ns = outlook.GetNamespace("MAPI")
            kalender = ns.GetDefaultFolder(9)
            items = kalender.Items
            items.IncludeRecurrences = True
            items.Sort("[Start]")
            start = datetime.now()
            end = start + timedelta(days=tage)
            restricted = items.Restrict(
                f"[Start] >= '{start.strftime('%d/%m/%Y')}' AND [Start] <= '{end.strftime('%d/%m/%Y')}'"
            )
            for item in restricted:
                try:
                    termine.append({
                        "betreff": str(item.Subject or ""),
                        "start": str(item.Start)[:16],
                        "ort": str(item.Location or ""),
                    })
                except Exception:
                    pass
            if termine:
                return termine[:15]
        except Exception:
            pass

        # Methode 2: win32com mit EarlyBind
        try:
            import win32com.client
            outlook = win32com.client.gencache.EnsureDispatch("Outlook.Application")
            ns = outlook.GetNamespace("MAPI")
            kalender = ns.GetDefaultFolder(9)
            items = kalender.Items
            items.IncludeRecurrences = True
            items.Sort("[Start]")
            start = datetime.now()
            end = start + timedelta(days=tage)
            for item in items:
                try:
                    item_start = str(item.Start)[:10]
                    if str(start)[:10] <= item_start <= str(end)[:10]:
                        termine.append({
                            "betreff": str(item.Subject or ""),
                            "start": str(item.Start)[:16],
                            "ort": str(item.Location or ""),
                        })
                    if len(termine) >= 15:
                        break
                except Exception:
                    continue
            if termine:
                return sorted(termine, key=lambda x: x["start"])[:15]
        except Exception:
            pass

        # Methode 3: Outlook via subprocess (prüfen ob Outlook läuft)
        try:
            import subprocess
            result = subprocess.run(
                ["powershell", "-Command",
                 "$ol = New-Object -ComObject Outlook.Application; "
                 "$ns = $ol.GetNamespace('MAPI'); "
                 "$cal = $ns.GetDefaultFolder(9); "
                 "$items = $cal.Items; "
                 "$items.IncludeRecurrences = $true; "
                 "$items.Sort('[Start]'); "
                 f"$start = Get-Date; $end = $start.AddDays({tage}); "
                 "$items | Where-Object {$_.Start -ge $start -and $_.Start -le $end} | "
                 "Select-Object -First 15 Subject, Start, Location | "
                 "ConvertTo-Json"],
                capture_output=True, text=True, timeout=8
            )
            if result.returncode == 0 and result.stdout.strip():
                import json
                data = json.loads(result.stdout)
                if isinstance(data, dict):
                    data = [data]
                for item in data:
                    termine.append({
                        "betreff": str(item.get("Subject") or ""),
                        "start": str(item.get("Start") or "")[:16],
                        "ort": str(item.get("Location") or ""),
                    })
                return termine[:15]
        except Exception:
            pass

        return termine

    def _dashboard_widget_kalender(self, parent):
        """Outlook-Kalender-Widget für das Dashboard."""
        box = tk.Frame(parent, bg="#faf4ff", highlightbackground="#c9b8e8", highlightthickness=1)
        tk.Label(box, text="📅 Termine (Outlook)", font=(theme.FONT, 11, "bold"), fg="#6b4fb3", bg="#faf4ff").pack(anchor="w", padx=10, pady=(10, 4))
        termine = self._get_outlook_termine(14)
        if termine:
            for t in termine:
                start_short = t["start"][5:16].replace("T", "  ").replace("-", ".")
                text = f"📌 {start_short}  –  {t['betreff'][:38]}"
                tk.Label(box, text=text, font=(theme.FONT, 9), fg="#3a2060", bg="#faf4ff", anchor="w").pack(anchor="w", padx=10, pady=1)
        else:
            tk.Label(box, text="Keine Termine gefunden\n(Outlook nicht verfügbar oder keine Termine)", font=(theme.FONT, 9), fg="#888", bg="#faf4ff", justify="left").pack(anchor="w", padx=10, pady=4)
        tk.Label(box, text="Nur-Lese-Ansicht · Termine aus Outlook", font=(theme.FONT, 8), fg="#999", bg="#faf4ff").pack(anchor="w", padx=10, pady=(0, 8))
        return box

    def _check_wissensbasis_leer(self):
        """SP26: Nach Programmstart pruefen, ob die wichtigen Wissens-Tabellen
        befuellt sind. Wenn eine kritische Tabelle leer ist -> Hinweis-Dialog
        mit Direkt-Link auf 'Daten aktualisieren'.

        Kritisch im Sinne der Auswertung:
        - tbl_nmg_stamm (NMG-Artikel) - sonst kein NMG-Treffer.
        - nmg_rabatte (PK-Rabatte) - sonst keine Rabatt-Spalte.
        - tbl_austauschdatenbank (aktive Eintraege) - sonst kein Austausch-Lookup.
        """
        try:
            items = self._get_data_update_status_items()
        except Exception:
            return
        if not items:
            return
        kritische_namen = {"NMG Artikel", "PK Rabatte", "Austauschdatenbank"}
        leere = []
        for name, _value, count in items:
            n = self._count_main(count)
            if name in kritische_namen and n <= 0:
                leere.append(name)
        if not leere:
            return
        zeilen = "\n".join(f"  • {n}" for n in leere)
        msg = (
            "Die folgenden Wissensbasen sind leer oder noch nicht importiert:\n\n"
            f"{zeilen}\n\n"
            "Solange diese Daten fehlen, kommen Auswertungen unvollstaendig zurueck "
            "(z.B. fehlende NMG-Treffer oder Rabatt-Spalte).\n\n"
            "Moechtest du jetzt zur Seite 'Daten aktualisieren' wechseln, um die "
            "fehlenden Daten zu importieren?"
        )
        try:
            if messagebox.askyesno("Wissensbasis unvollstaendig", msg):
                self.show_daten_aktualisieren_page()
        except Exception:
            pass

    def _check_apotheken_analyse_faellig(self):
        """Erstellt ToDos für Apotheken deren letzte Analyse > 5 Monate her ist (außer inaktiven)."""
        try:
            from datetime import datetime, timedelta
            cutoff = (datetime.now() - timedelta(days=150)).strftime("%Y-%m-%d")
            with sqlite3.connect(DB_PATH) as con:
                con.row_factory = sqlite3.Row
                kunden = con.execute(
                    "SELECT kundennummer, kundenname FROM tbl_kunden_center WHERE lower(status)='aktiv'"
                ).fetchall()
                for k in kunden:
                    knr = k["kundennummer"] or ""
                    kname = k["kundenname"] or ""
                    row = con.execute("""\nSELECT MAX(datum) as letzte FROM tbl_auswertungen
                        WHERE (kundennummer=? AND kundennummer<>'') OR (kundenname=? AND kundenname<>'')\n""", (knr, kname)).fetchone()
                    letzte = row["letzte"] if row else None
                    if letzte and str(letzte)[:10] >= cutoff:
                        continue  # aktuell genug
                    # Prüfen ob ToDo schon vorhanden
                    todo_titel = f"Neue Analyse anfordern: {kname or knr}"
                    existing_todo = con.execute(
                        "SELECT id FROM tbl_todo_center WHERE titel=? AND lower(status) NOT IN ('erledigt','done','abgeschlossen')",
                        (todo_titel,)
                    ).fetchone()
                    if existing_todo:
                        continue
                    con.execute("""\nINSERT INTO tbl_todo_center(titel,bereich,status,prioritaet,bearbeiter)
                        VALUES(?,?,?,?,?)\n""", (todo_titel, "Kunden", "offen", "Normal", self.bearbeiter))
                con.commit()
        except Exception:
            pass

    def _build_nav_tree(self):
        """Befuellt die Sidebar. Reihenfolge wie gehabt, in Abschnitte gruppiert.

        SP11: alle Navi-Texte ueber _T() - Sprache wird beim Programmstart aus
        language.json gelesen und gilt bis zum naechsten Start.
        """
        groups = [
            (None, [
                ("startseite", "🏠", _T("Startseite")),
                ("neue_auswertung", "📊", _T("Neue Auswertung")),
            ]),
            ("Arbeiten", [
                ("apps", "🚀", _T("Apps")),
                ("analysen", "📁", _T("Analysen")),
                ("schulbank", "🎓", _T("Schulbank")),
            ]),
            ("Daten", [
                ("daten_aktualisieren", "🗄", _T("Daten aktualisieren")),
                ("datenbankuebersicht", "🗂", _T("Datenbankübersicht")),
                ("update_backup", "🔄", _T("Update / Backup")),
            ]),
            ("Mehr", [
                ("report", "📋", _T("Report")),
                ("roadmap", "📌", _T("Roadmap")),
                ("datenbankpfad", "☁️", _T("Cloud / DB-Pfad")),
                ("hilfe", "❓", _T("Hilfe")),
            ]),
        ]
        for section, entries in groups:
            if section:
                self.sidebar.add_section(section)
            for key, icon, label in entries:
                self.sidebar.add_item(key, icon, label, lambda k=key: self.navigate(k))

    def navigate(self, key):
        """Oeffnet die zum Schluessel gehoerende Seite/Aktion (Sidebar oder
        programmatisch). Ersetzt den alten Treeview-Dispatch."""
        handler = self._nav_handlers().get(key)
        if not handler:
            return
        try:
            self.sidebar.set_active(key)
        except Exception:
            pass
        try:
            log_event("programm", "Navigation", f"Menüpunkt geöffnet: {key}", user=self.bearbeiter)
        except Exception:
            pass
        handler()

    def _nav_handlers(self):
        """Zentrale Zuordnung Navigations-/Aktions-Schluessel -> Methode.
        Einzige Quelle fuer Sidebar-Klicks und programmatische Navigation."""
        return {
            "startseite": self.show_startseite,
            "neue_auswertung": self.show_neue_auswertung_page,

            "analysen": self.show_analysen_page,
            "produktanalyse": self.show_produktanalyse_page,
            # SP8 hotfix: show_marktanalyse_page wurde in SP7 entfernt,
            # die Referenz hier aber vergessen. Beim Programmstart loest Python
            # alle Werte im Dispatch-Dict sofort auf -> AttributeError beim Start.
            # Eintrag komplett raus.
            "gespeicherte_analysen": self.open_saved_analyses,
            "abweichungsanalyse": self.show_abweichungsanalyse_page,

            "schulbank": lambda: self.show_schulbank_page("Schulbank"),
            "lern_neu": lambda: self.show_schulbank_page("Neue Lernvorschläge"),
            "lern_biosimilar": lambda: self.show_schulbank_page("Biosimilar"),
            "lern_uebernommen": lambda: self.show_schulbank_page("Übernommen"),
            "lern_abgelehnt": lambda: self.show_schulbank_page("Abgelehnt"),
            "lern_historie": lambda: self.show_schulbank_page("Historie"),
            "lern_manuelle_pruefung": self.show_schulbank_manuelle_pruefung,

            "daten_aktualisieren": self.show_daten_aktualisieren_page,
            "daten_manuelle_analysen_zw": lambda: self.show_import_page("Manuelle Analysen / Zukunftswerk", self.import_zw_data),
            "daten_partnerkonditionen": lambda: self.show_import_page("Partnerkonditionen", self.import_pk_data),
            "daten_apu_hap": lambda: self.show_import_page("APU/HAP Daten", self.import_apu_data),
            "daten_nmg_artikel": lambda: self.show_import_page("NMG Artikel", self.import_nmg_articles),
            "daten_pk_rabatte": lambda: self.show_import_page("Partnerkonditionen Rabatte", self.import_pk_rabatte),
            "nmg_rabatte": self.show_nmg_rabatte_uebersicht,
            "daten_austauschdatenbank": self.show_austauschdatenbank_page,
            "daten_manuelle_analysen": self.import_manuelle_analysen,
            "daten_artikelstamm": self.show_artikelstamm_page,
            "daten_auswertungsvorlage": self.show_auswertungsvorlage_page,

            "update_backup": self.show_update_backup_page,
            "versionsinfo": self.show_version,
            "update_suchen": self.check_online_update,
            "update_installieren": self.install_update_dialog,
            "backup_erstellen": self.create_backup,
            "backup_wiederherstellen": self.restore_backup,

            "datenbankuebersicht": self.show_datenbankuebersicht_page,

            "apps": self.show_apps_page,
            "kunden_center": self.show_kunden_center,
            "vergleichssuche": self.open_vergleichssuche_window,
            "globale_suche": self.open_globale_suche_window,
            "mitarbeiter_center": self.show_mitarbeiter_center,
            "todo_center": self.show_todo_center,
            "kasse": self.open_kasse_app,
            "bestell_center": self.open_kasse_app,
            "auswertungen": self.open_auswertungen_app,
            "report": self.show_report_page,
            "roadmap": self.show_roadmap_page,
            "datenbankpfad": self.show_datenbankpfad_page,

            "hilfe": self.open_hilfe_app,
        }

    def _tile(self, parent, col, icon, title, desc, button, command, color):
        f = tk.Frame(parent, bg=theme.CARD, highlightbackground=theme.BORDER, highlightthickness=1)
        f.grid(row=2, column=col, sticky="nsew", padx=10, pady=8)
        tk.Label(f, text=icon, font=(theme.FONT, 32), bg=theme.CARD, fg=color).pack(pady=(22, 6))
        tk.Label(f, text=title, font=(theme.FONT, 14, "bold"), bg=theme.CARD, fg=theme.INK).pack()
        tk.Label(f, text=desc, wraplength=180, justify="center", bg=theme.CARD, fg=theme.MUTED, font=(theme.FONT, 10)).pack(padx=14, pady=12)
        tk.Button(f, text=button + "  →", command=command, bg=color, fg="white", activebackground=color, relief="flat", font=(theme.FONT, 11, "bold"), padx=18, pady=8, cursor="hand2").pack(fill="x", padx=18, pady=(8, 18))

    def _parse_status_datetime(self, value):
        """Wandelt verschiedene DB-Datumsformate inkl. Excel-Serienzahl in datetime um."""
        if value in (None, ""):
            return None
        text = str(value).strip()
        if not text:
            return None

        # Excel-Serienzahl, z. B. 45234 statt Datum.
        try:
            if re.fullmatch(r"\d+(\.\d+)?", text):
                number = float(text)
                if 20000 <= number <= 90000:
                    base = datetime(1899, 12, 30)
                    from datetime import timedelta
                    return base + timedelta(days=number)
        except Exception:
            pass

        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%d.%m.%Y %H:%M:%S",
            "%d.%m.%Y %H:%M",
            "%Y-%m-%d",
            "%d.%m.%Y",
        ):
            try:
                return datetime.strptime(text[:19], fmt)
            except Exception:
                pass
        return None

    def _format_status_datetime(self, value):
        """Formatiert Datenbank-Zeitstempel kurz für den rechten Statusbereich."""
        dt = self._parse_status_datetime(value)
        if not dt:
            return "noch nie"
        return dt.strftime("%d.%m.%Y %H:%M")

    def _count_main(self, count):
        """SP28: Aus count (int oder (gesamt, sub)-Tuple) den Haupt-Zaehler.
        Wird fuer Ampel + Start-Warnung verwendet."""
        if isinstance(count, tuple):
            try:
                return int(count[0] or 0)
            except Exception:
                return 0
        try:
            return int(count or 0)
        except Exception:
            return 0

    def _format_count(self, count):
        """SP26/SP28: Anzahl-Eintraege benutzerfreundlich formatieren.
        Wenn count ein (gesamt, sub)-Tuple ist, zeigen beide Werte.
        Beispiel: (105, 39) -> "105 Eintr., 39 mit Wert".
        """
        if isinstance(count, tuple):
            try:
                gesamt = int(count[0] or 0)
                sub = int(count[1] or 0)
            except Exception:
                gesamt, sub = 0, 0
            if gesamt <= 0:
                return "leer"
            haupt = f"{gesamt:,}".replace(",", ".") + " Eintr."
            if sub < gesamt:
                sub_str = f"{sub:,}".replace(",", ".")
                return f"{haupt}, {sub_str} mit Wert"
            return haupt
        try:
            n = int(count or 0)
        except Exception:
            n = 0
        if n <= 0:
            return "leer"
        return f"{n:,}".replace(",", ".") + " Eintr."

    def _status_ampel_combined(self, value, count):
        """SP26: Ampel beruecksichtigt sowohl Datum als auch leere Tabellen.
        count == 0 -> ⚪ leer. Sonst normale Datums-Ampel.
        """
        n = self._count_main(count)
        if n <= 0:
            return "⚪"
        return self._status_ampel(value) or "⚪"

    def _status_ampel_combined_color(self, value, count):
        """SP26: Textfarbe. Leere Tabellen werden rot, sonst Datums-Farbe."""
        n = self._count_main(count)
        if n <= 0:
            return "#b00020"
        return self._status_ampel_color(value)

    def _status_ampel(self, value):
        """Ampel nach Datenalter: grün <30, gelb ab 30, rot ab 60, schwarz ab 90 Tage."""
        dt = self._parse_status_datetime(value)
        if not dt:
            return ""
        try:
            age_days = (datetime.now() - dt).days
        except Exception:
            return ""
        if age_days >= 90:
            return "⚫"
        if age_days >= 60:
            return "🔴"
        if age_days >= 30:
            return "🟡"
        return "🟢"

    def _status_ampel_color(self, value):
        """Textfarbe passend zur Datenampel."""
        dt = self._parse_status_datetime(value)
        if not dt:
            return "#555555"
        try:
            age_days = (datetime.now() - dt).days
        except Exception:
            return "#555555"
        if age_days >= 90:
            return "#000000"
        if age_days >= 60:
            return "#b00020"
        if age_days >= 30:
            return "#b8860b"
        return "#11823b"

    def _get_last_data_update(self, con, queries):
        """Liest das jüngste Datum aus mehreren möglichen Tabellen/Spalten, robust gegen fehlende Tabellen."""
        best_dt = None
        best_raw = ""
        for sql, params in queries:
            try:
                row = con.execute(sql, params).fetchone()
                if row and row[0] not in (None, ""):
                    raw = str(row[0]).strip()
                    dt = self._parse_status_datetime(raw)
                    if dt:
                        if best_dt is None or dt > best_dt:
                            best_dt = dt
                            best_raw = raw
                    elif not best_raw:
                        best_raw = raw
            except Exception:
                continue
        return best_raw

    def _count_table_safe(self, con, query):
        """SP26: Anzahl-Eintraege fuer eine Wissensbasis. Robust gegen fehlende Tabellen."""
        try:
            row = con.execute(query).fetchone()
            return int(row[0]) if row and row[0] is not None else 0
        except Exception:
            return 0

    def _count_pk_rabatte_combo(self, con):
        """SP28: PK Rabatte mit zwei Zahlen: (gesamt, mit echtem Rabatt-Wert).
        Wenn die beiden Werte stark auseinanderlaufen, ist der Import zwar
        gelaufen, der Rabatt-Parser ist aber gescheitert - die Auswertung
        kann dann keine Rabatte schreiben.
        """
        gesamt = self._count_table_safe(con, "SELECT COUNT(*) FROM nmg_rabatte")
        mit_rabatt = self._count_table_safe(con, "SELECT COUNT(*) FROM nmg_rabatte WHERE rabatt IS NOT NULL")
        return (gesamt, mit_rabatt)

    def _get_data_update_status_items(self):
        """Rechter Statusbereich: letzte Aktualisierung + Anzahl Eintraege.
        Rueckgabe je Eintrag: (name, letztes_datum, anzahl).
        """
        try:
            with sqlite3.connect(DB_PATH) as con:
                return [
                    (
                        "APU/HAP Daten",
                        self._get_last_data_update(con, [
                            ("SELECT MAX(importdatum) FROM tbl_nmg_stamm WHERE apu IS NOT NULL OR taxe_ek IS NOT NULL OR taxe_vk IS NOT NULL", ()),
                            ("SELECT MAX(datum) FROM tbl_import_log WHERE lower(typ) IN ('apu_hap','apu','hap')", ()),
                            ("SELECT MAX(datum) FROM tbl_import_log WHERE lower(datei) LIKE '%apu%' OR lower(datei) LIKE '%hap%'", ()),
                        ]),
                        self._count_table_safe(con, "SELECT COUNT(*) FROM tbl_nmg_stamm WHERE apu IS NOT NULL OR taxe_ek IS NOT NULL OR taxe_vk IS NOT NULL"),
                    ),
                    (
                        "NMG Artikel",
                        self._get_last_data_update(con, [
                            ("SELECT MAX(importdatum) FROM tbl_nmg_stamm", ()),
                            ("SELECT MAX(datum) FROM tbl_import_log WHERE lower(typ) IN ('nmg_stamm','nmg_artikel','nmg_articles')", ()),
                            ("SELECT MAX(datum) FROM tbl_import_log WHERE lower(datei) LIKE '%nmg%' AND lower(datei) LIKE '%artikel%'", ()),
                        ]),
                        self._count_table_safe(con, "SELECT COUNT(*) FROM tbl_nmg_stamm"),
                    ),
                    (
                        # SP28: Anzeige zeigt Gesamtgroesse UND wieviele wirklich
                        # einen Rabatt-Wert haben. Wenn beide Zahlen auseinanderlaufen,
                        # ist der Import zwar gelaufen, der Rabatt-Parser ist aber an
                        # den Werten gescheitert.
                        "PK Rabatte",
                        self._get_last_data_update(con, [
                            ("SELECT MAX(letzte_aktualisierung) FROM nmg_rabatte", ()),
                            ("SELECT MAX(datum) FROM tbl_import_log WHERE lower(typ) IN ('pk_rabatte','partnerkonditionen_rabatte','rabatte')", ()),
                            ("SELECT MAX(datum) FROM tbl_import_log WHERE lower(datei) LIKE '%rabatt%'", ()),
                        ]),
                        self._count_pk_rabatte_combo(con),
                    ),
                    (
                        "Artikelstamm",
                        self._get_last_data_update(con, [
                            ("SELECT MAX(letzte_aktualisierung) FROM tbl_pzn_basisdaten", ()),
                            ("SELECT MAX(importdatum) FROM tbl_pzn_basisdaten", ()),
                            ("SELECT MAX(aktualisiert_am) FROM tbl_artikelstamm", ()),
                            ("SELECT MAX(erstellt_am) FROM tbl_artikelstamm", ()),
                            ("SELECT MAX(datum) FROM tbl_import_log WHERE lower(typ) IN ('artikelstamm','pzn_basis','artikel_basis')", ()),
                            ("SELECT MAX(datum) FROM tbl_import_log WHERE lower(datei) LIKE '%artikelstamm%'", ()),
                        ]),
                        self._count_table_safe(con, "SELECT COUNT(*) FROM tbl_artikelstamm")
                            or self._count_table_safe(con, "SELECT COUNT(*) FROM tbl_pzn_basisdaten"),
                    ),
                    (
                        "Austauschdatenbank",
                        self._get_last_data_update(con, [
                            ("SELECT MAX(aktualisiert_am) FROM tbl_austauschdatenbank WHERE COALESCE(status,'aktiv')='aktiv'", ()),
                            ("SELECT MAX(erstellt_am) FROM tbl_austauschdatenbank WHERE COALESCE(status,'aktiv')='aktiv'", ()),
                            ("SELECT MAX(datum) FROM tbl_import_log WHERE lower(typ) IN ('austausch','austauschdatenbank','biosimilar')", ()),
                        ]),
                        self._count_table_safe(con, "SELECT COUNT(*) FROM tbl_austauschdatenbank WHERE COALESCE(status,'aktiv')='aktiv'"),
                    ),
                    # SP29: Lieferfaehigkeit-Kachel ausgeblendet. Kommt mit der
                    # Bestandsfuehrung wieder; vorher gibt es nichts sinnvolles
                    # anzuzeigen.
                ]
        except Exception:
            return []

    def _get_data_update_status_text(self):
        items = self._get_data_update_status_items()
        if not items:
            return "Status nicht verfügbar"
        lines = []
        for name, value, count in items:
            amp = self._status_ampel_combined(value, count)
            lines.append(f"{name} {amp}  ({self._format_count(count)})")
            lines.append(self._format_status_datetime(value))
            lines.append("")
        lines.append("Legende")
        lines.append("🟢 < 30 Tage")
        lines.append("🟡 30–60 Tage")
        lines.append("🔴 > 60 Tage")
        lines.append("⚪ keine Daten")
        return "\n".join(lines).rstrip()

    def _status_card(self, parent):
        auto = self.auto_backup_result

        # V1.1 SP11: 'Datenbank: <Name>' aus Aktiver-Bearbeiter-Box entfernt
        # (Punkt 6) und die separate Datenbankstatus-Box entfernt (Punkt 5,
        # Werte waren nicht real).
        box_bearbeiter = tk.Frame(parent, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        box_bearbeiter.pack(fill="x", pady=(0, 12))
        tk.Label(
            box_bearbeiter,
            text="👤 Aktiver Bearbeiter",
            font=(theme.FONT, 13, "bold"),
            fg="#0b4a86",
            bg="#ffffff"
        ).pack(anchor="w", padx=14, pady=(14, 6))
        tk.Label(
            box_bearbeiter,
            text=(
                f"Windows-Login:\n{self.bearbeiter or 'unbekannt'}\n\n"
                f"Version:\n{APP_VERSION}"
            ),
            justify="left",
            bg="#ffffff",
            fg="#222"
        ).pack(anchor="w", padx=14, pady=(0, 14))

        box_updates = tk.Frame(parent, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        box_updates.pack(fill="x", pady=(0, 12))
        tk.Label(
            box_updates,
            text="🕘 Letzte Datenaktualisierung",
            font=(theme.FONT, 13, "bold"),
            fg="#0b4a86",
            bg="#ffffff"
        ).pack(anchor="w", padx=14, pady=(14, 6))
        try:
            for item_name, item_value, item_count in self._get_data_update_status_items():
                color = self._status_ampel_combined_color(item_value, item_count)
                amp = self._status_ampel_combined(item_value, item_count)
                date_text = self._format_status_datetime(item_value)
                count_text = self._format_count(item_count)
                tk.Label(
                    box_updates,
                    text=f"{amp} {item_name} ({count_text})\n{date_text}",
                    justify="left",
                    bg="#ffffff",
                    fg=color,
                    font=(theme.FONT, 10, "bold")
                ).pack(anchor="w", padx=14, pady=(0, 8))
            tk.Label(
                box_updates,
                text="Legende: Grün <30 Tage | Gelb ab 30 | Rot ab 60 | Schwarz ab 90",
                justify="left",
                bg="#ffffff",
                fg="#555",
                font=(theme.FONT, 8)
            ).pack(anchor="w", padx=14, pady=(0, 14))
        except Exception:
            tk.Label(
                box_updates,
                text=self._get_data_update_status_text(),
                justify="left",
                bg="#ffffff",
                fg="#222"
            ).pack(anchor="w", padx=14, pady=(0, 14))

        filler = tk.Frame(parent, bg="#f5f7fb")
        filler.pack(fill="both", expand=True)

        # V1.1 SP18: Background-Job-Status (frueher rechts oben im Header).
        # Liegt direkt UEBER dem Backup-Status-Block. Wenn kein Job laeuft
        # bleibt es leer (kein Border, kein Header) - somit unsichtbar.
        self._bg_jobs_panel = tk.Frame(parent, bg="#f5f7fb")
        self._bg_jobs_panel.pack(fill="x", side="bottom", pady=(0, 6))

        box_backup = tk.Frame(parent, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        box_backup.pack(fill="x", side="bottom")
        txt = "Backup heute erstellt" if auto.get("created") else "Backup heute vorhanden"
        tk.Label(box_backup, text="✅ Backup-Status", font=(theme.FONT, 13, "bold"), fg="#0b4a86", bg="#ffffff").pack(anchor="w", padx=14, pady=(14, 6))
        tk.Label(box_backup, text=f"{txt}\nAufbewahrung: 7 Tage\nÄltestes Auto-Backup wird zuerst gelöscht.", justify="left", bg="#ffffff", fg="#222").pack(anchor="w", padx=14, pady=(0, 12))
        tk.Button(box_backup, text="Backup-Verwaltung", command=self.show_backup_dialog).pack(fill="x", padx=14, pady=(0, 6))
        tk.Button(box_backup, text="📋 Protokolle", command=self.show_protocol_center).pack(fill="x", padx=14, pady=(0, 14))

    def _get_saved_analysis_rows(self, limit=300):
        try:
            with sqlite3.connect(DB_PATH) as con:
                con.row_factory = sqlite3.Row
                return con.execute("""\nSELECT id, datum, apotheke, quelldatei, ausgabedatei, anzahl_positionen,
                           nmg_treffer, nicht_nmg, COALESCE(datenquelle,'NMG') AS datenquelle\nFROM tbl_auswertungen
                    ORDER BY datetime(datum) DESC, id DESC\nLIMIT ?
                """, (int(limit),)).fetchall()
        except Exception:
            return []

    def _find_output_file_for_analysis_row(self, row):
        val = row["ausgabedatei"] or ""
        candidates = []
        if val:
            candidates.append(Path(val))
            candidates.append(Path(__file__).resolve().parent.parent / val)
        try:
            for f in SAVED_ANALYSES_DIR.rglob("*.xlsx"):
                if row["apotheke"] and _safe_name(str(row["apotheke"])).lower() in str(f).lower():
                    candidates.append(f)
                if row["ausgabedatei"] and Path(str(row["ausgabedatei"])).name == f.name:
                    candidates.append(f)
        except Exception:
            pass
        for c in candidates:
            if c and c.exists():
                return c
        return None

    def _roadmap_mark_abweichung_schulbank_v9_erledigt(self):
        try:
            ensure_roadmap_table()
            titel = "Abweichungsanalyse öffnen und Schulbank-Spalte-H-Logik"
            for row in list_roadmap_items():
                if row["titel"] == titel:
                    update_roadmap_status(int(row["id"]), "Erledigt")
                    return
            add_roadmap_item(
                bereich="Analyse",
                titel=titel,
                beschreibung=(
                    "Abweichungsanalyse-Ergebnis kann nach Erstellung direkt geöffnet oder der Ordner geöffnet werden. "
                    "Lernvorschläge werden aus der manuellen Anpassung erzeugt, sobald Spalte H / im Sortiment gefüllt ist; "
                    "bereits aktiv gelernte PZN alt werden ignoriert."
                ),
                status="Erledigt",
                prioritaet="Hoch"
            )
        except Exception:
            pass

    def _roadmap_mark_abweichung_in_neue_auswertung_erledigt(self):
        try:
            ensure_roadmap_table()
            titel = "Abweichungsanalyse in Neue Auswertung integriert"
            for row in list_roadmap_items():
                if row["titel"] == titel:
                    update_roadmap_status(int(row["id"]), "Erledigt")
                    return
            add_roadmap_item(
                bereich="Analyse",
                titel=titel,
                beschreibung=(
                    "Unter Neue Auswertung wurde ein eigener, getrennter Block für die Abweichungsanalyse ergänzt: "
                    "Programm-Auswertung kann aus gespeicherten Analysen oder als Datei gewählt werden, "
                    "manuelle Anpassung wird separat gewählt und danach die Abweichungsanalyse gestartet."
                ),
                status="Erledigt",
                prioritaet="Hoch"
            )
        except Exception:
            pass

    def neue_auswertung_routine(self):
        self._log_action("neue_auswertung", "Neue Auswertung geöffnet")
        file = filedialog.askopenfilename(title="Rohdaten / Excel-Datei auswählen", filetypes=SUPPORTED_DATA_FILETYPES
        )
        if not file:
            return
        default_name = Path(file).stem.replace("_", " ")
        name = simpledialog.askstring("Analyse-Name", "Bitte Auswertungsname eingeben:", initialvalue=default_name)
        if not name:
            return
        analyse_name = _safe_name(name)
        apotheke = analyse_name
        try:
            self._run_neue_auswertung_export(file, analyse_name)
        except UnknownInputFormatError as exc:
            self.status.set(str(exc))
            if messagebox.askyesno("Format nicht erkannt", f"{exc}\n\nSoll der Rohdaten-Formatassistent geöffnet werden?"):
                mapping = self._open_rohdaten_format_assistent(file, str(exc))
                if mapping:
                    try:
                        self._run_auswertung_after_mapping(file, mapping, analyse_name)
                    except Exception as mapped_exc:
                        self.status.set(f"Auswertung nach Mapping fehlgeschlagen: {mapped_exc}")
                        messagebox.showerror("Auswertungsfehler nach Mapping", str(mapped_exc))
            else:
                messagebox.showwarning("Format nicht erkannt", str(exc))
        except Exception as exc:
            messagebox.showerror("Auswertungsfehler", str(exc))

    def open_saved_analyses(self, start_action=None):
        """Zeigt gespeicherte Analysen direkt im mittleren Arbeitsbereich.

        start_action:
        - None: normale Übersicht
        - "produktanalyse": Nutzer soll zuerst eine Auswertung auswählen
        SP7: marktanalyse-Branch entfernt - Funktion komplett raus.
        """
        self.clear_page()
        if start_action == "produktanalyse":
            self._page_header("Produktanalyse – Auswertung auswählen", "Bitte erst eine gespeicherte Auswertung auswählen. Danach wird die Produktanalyse genau für diese Auswertung erstellt.")
        else:
            self._page_header("Gespeicherte Analysen", "Analyse auswählen und direkt weiter auswerten.")

        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=0)
        body.rowconfigure(1, weight=1)

        # Suchzeile - V1.1 SP8: getrennte Felder fuer feingranulare Suche
        search_bar = tk.Frame(body, bg="#ffffff")
        search_bar.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 6))

        # Volltext (wirkt auf alle Spalten zusammen)
        tk.Label(search_bar, text="Suche:", bg="#ffffff", fg="#0b4a86",
                 font=(theme.FONT, 10, "bold")).pack(side="left")
        _sa_search_var = tk.StringVar()
        tk.Entry(search_bar, textvariable=_sa_search_var, width=22).pack(side="left", padx=(6, 12))

        # Getrennte Felder (V1.1 SP8)
        tk.Label(search_bar, text="Apotheke:", bg="#ffffff", fg="#0b4a86",
                 font=(theme.FONT, 9)).pack(side="left")
        _sa_apo_var = tk.StringVar()
        tk.Entry(search_bar, textvariable=_sa_apo_var, width=14).pack(side="left", padx=(4, 8))

        tk.Label(search_bar, text="Kunden-Nr.:", bg="#ffffff", fg="#0b4a86",
                 font=(theme.FONT, 9)).pack(side="left")
        _sa_knr_var = tk.StringVar()
        tk.Entry(search_bar, textvariable=_sa_knr_var, width=10).pack(side="left", padx=(4, 8))

        tk.Label(search_bar, text="Kunde:", bg="#ffffff", fg="#0b4a86",
                 font=(theme.FONT, 9)).pack(side="left")
        _sa_kname_var = tk.StringVar()
        tk.Entry(search_bar, textvariable=_sa_kname_var, width=14).pack(side="left", padx=(4, 8))

        tk.Label(search_bar, text="Datum von:", bg="#ffffff", fg="#0b4a86",
                 font=(theme.FONT, 9)).pack(side="left")
        _sa_von_var = tk.StringVar()
        _make_date_entry(search_bar, _sa_von_var, width=11).pack(side="left", padx=(4, 4))

        tk.Label(search_bar, text="bis:", bg="#ffffff", fg="#0b4a86",
                 font=(theme.FONT, 9)).pack(side="left")
        _sa_bis_var = tk.StringVar()
        _make_date_entry(search_bar, _sa_bis_var, width=11).pack(side="left", padx=(4, 4))

        # V1.1 SP10: Datum-Felder leer lassen wenn der Picker direkt das
        # heutige Datum schreibt - sonst greift der Filter sofort und schluckt
        # alle Treffer.
        _sa_von_var.set("")
        _sa_bis_var.set("")

        tk.Label(search_bar, text="Doppelklick = Kunden zuordnen",
                 bg="#ffffff", fg="#888", font=(theme.FONT, 9)).pack(side="left", padx=8)

        filter_bar = tk.Frame(body, bg="#ffffff")
        filter_bar.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 8))
        filter_mode = tk.StringVar(value="ALLE")
        tk.Label(filter_bar, text="Anzeigen:", bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 10, "bold")).pack(side="left", padx=(0, 6))
        def set_filter_mode(value):
            filter_mode.set(value)
            _sa_reload()
        # V1.1 SP10: 'IMPORT' = via bemerkung 'historischer Analyseimport%',
        # 'PROGRAMM' = der Rest (create_vorlage_export), 'PA' = Produktanalyse-
        # Excels aus OUTPUT_DIR (Spezialmodus, scannt Dateisystem).
        for label, value in (
            ("Alle", "ALLE"),
            ("PK", "NMG"),
            ("ZW", "ZW"),
            ("Importiert", "IMPORT"),
            ("Programm", "PROGRAMM"),
            ("Produktanalyse", "PA"),
            ("Neue Ausw.", "NEU"),
            ("mit Datei", "DATEI"),
        ):
            tk.Button(filter_bar, text=label,
                      command=lambda v=value: set_filter_mode(v),
                      padx=10, pady=4).pack(side="left", padx=(0, 5))

        # Treeview mit Kundenspalten
        _sa_cols = ("id", "datum", "typ", "apotheke", "kundennummer", "kundenname", "positionen", "treffer")
        _sa_heads = {"id": "ID", "datum": "Datum", "typ": "Typ", "apotheke": "Auswertungsname",
                     "kundennummer": "Kundennr.", "kundenname": "Apothekenname (Kunde)",
                     "positionen": "Pos.", "treffer": "Treffer"}
        _sa_widths = {"id": 50, "datum": 130, "typ": 40, "apotheke": 200,
                      "kundennummer": 90, "kundenname": 180, "positionen": 55, "treffer": 55}
        tree_frame = tk.Frame(body, bg="#ffffff")
        tree_frame.grid(row=2, column=0, sticky="nsew", pady=(0, 16))
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)
        listbox = ttk.Treeview(tree_frame, columns=_sa_cols, show="headings", selectmode="browse")
        for _c in _sa_cols:
            listbox.heading(_c, text=_sa_heads[_c])
            listbox.column(_c, width=_sa_widths.get(_c, 100), anchor="w")
        listbox.grid(row=0, column=0, sticky="nsew")
        scrollbar = tk.Scrollbar(tree_frame, orient="vertical", command=listbox.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        listbox.configure(yscrollcommand=scrollbar.set)
        _sb_h = tk.Scrollbar(tree_frame, orient="horizontal", command=listbox.xview)
        _sb_h.grid(row=1, column=0, sticky="ew")
        listbox.configure(xscrollcommand=_sb_h.set)

        side = tk.Frame(body, bg="#f8fbff", highlightbackground="#d8e2ee", highlightthickness=1, width=270)
        side.grid(row=2, column=1, sticky="ns", padx=(16, 0), pady=(0, 16))
        side.grid_propagate(False)
        detail = tk.StringVar(value="Bitte Analyse auswählen.")
        tk.Label(side, text="Aktionen", font=(theme.FONT, 14, "bold"), fg="#0b4a86", bg="#f8fbff").pack(anchor="w", padx=16, pady=(16, 8))
        tk.Label(side, textvariable=detail, justify="left", wraplength=230, bg="#f8fbff", fg="#333").pack(anchor="w", padx=16, pady=(0, 14))

        rows = []
        try:
            with sqlite3.connect(DB_PATH) as con:
                con.row_factory = sqlite3.Row
                rows = con.execute("""
SELECT id, datum, apotheke, quelldatei, ausgabedatei, anzahl_positionen,
       nmg_treffer, nicht_nmg, COALESCE(datenquelle,'NMG') AS datenquelle,
       COALESCE(bemerkung,'') AS bemerkung
FROM tbl_auswertungen
ORDER BY datetime(datum) DESC, id DESC
LIMIT 500
                """).fetchall()
        except Exception as exc:
            messagebox.showerror("Gespeicherte Analysen", f"Datenbank konnte nicht gelesen werden:\n{exc}")

        # Spalten kundennummer/kundenname ggf. nachrüsten
        _sa_row_map = {}
        try:
            with sqlite3.connect(DB_PATH) as _con:
                _ec = {r[1] for r in _con.execute("PRAGMA table_info(tbl_auswertungen)").fetchall()}
                for _col in ("kundennummer", "kundenname"):
                    if _col not in _ec:
                        _con.execute(f"ALTER TABLE tbl_auswertungen ADD COLUMN {_col} TEXT")
                _con.commit()
        except Exception:
            pass

        def _sa_reload(*_args):
            # V1.1 SP12 Bug-Fix: nach Seitenwechsel kann die alte StringVar
            # noch ihren trace feuern, obwohl die Treeview schon zerstoert ist.
            # Dann wirft listbox.get_children() / .delete() 'Item end not found'.
            # Defensive Guard: Widget-Existenz pruefen.
            try:
                if not listbox.winfo_exists():
                    return
            except tk.TclError:
                return
            try:
                ft = _sa_search_var.get().strip().lower()
                f_apo = _sa_apo_var.get().strip().lower()
                f_knr = _sa_knr_var.get().strip().lower()
                f_kname = _sa_kname_var.get().strip().lower()
                f_von = _sa_von_var.get().strip()
                f_bis = _sa_bis_var.get().strip()
            except tk.TclError:
                return
            try:
                for item in listbox.get_children():
                    listbox.delete(item)
            except tk.TclError:
                return
            _sa_row_map.clear()
            mode = filter_mode.get()
            # V1.1 SP10: Produktanalyse-Modus scannt OUTPUT_DIR statt der DB.
            if mode == "PA":
                _sa_show_produktanalyse_files(ft, f_apo, f_von, f_bis)
                return
            for row in rows:
                dq = _dq_label(row["datenquelle"])
                apo = str(row["apotheke"] or "")
                knr = str(row["kundennummer"] if "kundennummer" in row.keys() else "")
                kname = str(row["kundenname"] if "kundenname" in row.keys() else "")
                vals = (row["id"], str(row["datum"] or "")[:16], dq, apo, knr, kname,
                        row["anzahl_positionen"] or 0, row["nmg_treffer"] or 0)
                if mode in ("NMG", "ZW") and str(row["datenquelle"] or "NMG").replace("ZF", "ZW") != mode:
                    continue
                if mode == "DATEI" and not find_output_file(row):
                    continue
                if mode == "NEU" and apo.lower().startswith(("manuell", "zf import", "zf daten")):
                    continue
                # V1.1 SP10: 'IMPORT' = historischer Analyseimport,
                # 'PROGRAMM' = alles andere (create_vorlage_export etc.).
                bem = str(row["bemerkung"] if "bemerkung" in row.keys() else "")
                is_import = bem.lower().startswith("historischer analyseimport")
                if mode == "IMPORT" and not is_import:
                    continue
                if mode == "PROGRAMM" and is_import:
                    continue
                # V1.1 SP8: feingranulare Felder (UND-Logik)
                if f_apo and f_apo not in apo.lower(): continue
                if f_knr and f_knr not in knr.lower(): continue
                if f_kname and f_kname not in kname.lower(): continue
                if f_von or f_bis:
                    datum_iso = str(row["datum"] or "")[:10]
                    if f_von and datum_iso < f_von: continue
                    if f_bis and datum_iso > f_bis: continue
                # Volltext zuletzt (matched gegen die formatierten Werte)
                if ft:
                    combined = " ".join(str(v) for v in vals).lower()
                    if ft not in combined: continue
                iid = listbox.insert("", "end", values=vals)
                _sa_row_map[iid] = dict(row)

        def _sa_show_produktanalyse_files(ft, f_apo, f_von, f_bis):
            """V1.1 SP10: Listet Produktanalyse-Excels aus OUTPUT_DIR.
            Doppelklick / 'Auswertung oeffnen' oeffnet die Excel direkt.
            """
            try:
                from .config import OUTPUT_DIR as _OUT
                # V1.1 SP12: rglob, weil Produktanalysen jetzt in
                # OUTPUT_DIR/Produktanalyse/<Jahr>/Q<n>/ landen.
                pa_files = sorted(
                    _OUT.rglob("Produktanalyse_*.xlsx"),
                    key=lambda p: p.stat().st_mtime, reverse=True,
                )
            except Exception:
                pa_files = []
            for p in pa_files:
                mtime = datetime.fromtimestamp(p.stat().st_mtime)
                datum_iso = mtime.strftime("%Y-%m-%d")
                datum_disp = mtime.strftime("%Y-%m-%d %H:%M")
                name = p.stem
                vals = (0, datum_disp, "PA", name, "", "", 0, 0)
                if f_apo and f_apo not in name.lower(): continue
                if f_von and datum_iso < f_von: continue
                if f_bis and datum_iso > f_bis: continue
                if ft:
                    combined = " ".join(str(v) for v in vals).lower()
                    if ft not in combined: continue
                iid = listbox.insert("", "end", values=vals)
                # Pseudo-Row mit ausgabedatei fuer "Auswertung oeffnen".
                _sa_row_map[iid] = {
                    "id": 0, "datum": datum_disp, "apotheke": name,
                    "ausgabedatei": str(p), "datenquelle": "PA",
                    "anzahl_positionen": 0, "nmg_treffer": 0,
                    "kundennummer": "", "kundenname": "", "bemerkung": "Produktanalyse-Datei",
                    "_ist_pa_datei": True,
                }
        _sa_reload()
        for _v in (_sa_search_var, _sa_apo_var, _sa_knr_var, _sa_kname_var, _sa_von_var, _sa_bis_var):
            _v.trace_add("write", _sa_reload)

        def selected_row_silent():
            sel = listbox.selection()
            return _sa_row_map.get(sel[0]) if sel else None

        def selected_row():
            row = selected_row_silent()
            if not row:
                messagebox.showinfo("Auswahl", "Bitte zuerst eine Analyse auswählen.")
            return row

        def update_detail(event=None):
            row = selected_row_silent()
            if not row:
                detail.set("Bitte Analyse auswählen.")
                return
            dq = "PK" if row["datenquelle"] == "NMG" else row["datenquelle"]
            detail.set(
                f"Analyse-ID: {row['id']}\n"
                f"Name: {row['apotheke']}\n"
                f"Typ: {dq}\n"
                f"Kundennr.: {row.get('kundennummer') or '–'}\n"
                f"Kunde: {row.get('kundenname') or '–'}\n"
                f"Positionen: {row['anzahl_positionen'] or 0}\n"
                f"Treffer: {row['nmg_treffer'] or 0}"
            )

        def _on_dbl_click(event=None):
            row = selected_row_silent()
            if not row:
                return
            def _refresh():
                try:
                    with sqlite3.connect(DB_PATH) as _c:
                        _c.row_factory = __import__("sqlite3").Row
                        updated = _c.execute("SELECT * FROM tbl_auswertungen WHERE id=?", (row["id"],)).fetchone()
                    if updated:
                        for i, r in enumerate(rows):
                            if r["id"] == row["id"]:
                                rows[i] = updated
                                break
                except Exception:
                    pass
                _sa_reload(_sa_search_var.get().strip())
            self._kunden_zuordnen_dialog(row["id"], apotheke_name=str(row["apotheke"] or ""), callback=_refresh)

        listbox.bind("<<TreeviewSelect>>", update_detail)
        listbox.bind("<Double-1>", _on_dbl_click)

        def find_output_file(row):
            val = row["ausgabedatei"] or ""
            candidates = []
            if val:
                candidates.append(Path(val))
                candidates.append(Path(__file__).resolve().parent.parent / val)
            try:
                for f in SAVED_ANALYSES_DIR.rglob("*.xlsx"):
                    if row["apotheke"] and _safe_name(str(row["apotheke"])).lower() in str(f).lower():
                        candidates.append(f)
                    if row["ausgabedatei"] and Path(str(row["ausgabedatei"])).name == f.name:
                        candidates.append(f)
            except Exception:
                pass
            for c in candidates:
                if c and c.exists():
                    return c
            return None

        def open_selected_output():
            row = selected_row()
            if not row:
                return
            f = find_output_file(row)
            # V1.1 SP8: Datei direkt oeffnen statt nur den Ordner.
            if f:
                try:
                    _open_file(f)
                except Exception as exc:
                    messagebox.showerror("Auswertung oeffnen", str(exc))
                return
            messagebox.showinfo("Auswertung", "Zu dieser Analyse wurde keine Ausgabedatei im Analyseordner gefunden. Der Ordner wird geöffnet.")
            _open_folder(SAVED_ANALYSES_DIR)

        def product_selected():
            row = selected_row()
            if not row:
                return
            dq = row["datenquelle"] or "NMG"
            rid = int(row["id"])

            # Hintergrund-Lauf mit Animation rechts oben, UI bleibt klickbar.
            def on_done(out):
                self.status.set(f"Produktanalyse für Analyse {rid} erzeugt: {out}")
                messagebox.showinfo("Produktanalyse", f"Produktanalyse erstellt:\n{out}")

            self._run_background(
                lambda: export_marktanalyse_produktchancen(limit=500, min_apotheken=1, datenquelle=dq, auswertung_id=rid),
                title="Produktanalyse",
                subtitle="Produktchancen werden berechnet ...",
                progress=False,
                on_done=on_done,
                on_error=lambda exc: messagebox.showerror("Produktanalyse", str(exc)),
            )

        def deviation_selected():
            row = selected_row()
            if not row:
                return
            programm = find_output_file(row)
            if not programm:
                messagebox.showinfo("Abweichungsanalyse", "Keine Programm-Auswertung gefunden. Bitte über die normale Abweichungsanalyse beide Dateien auswählen.")
                self.deviation_analysis()
                return
            manuell = filedialog.askopenfilename(title="Manuelle Vergleichsauswertung auswählen", filetypes=SUPPORTED_DATA_FILETYPES)
            if not manuell:
                return
            try:
                out = self._run_busy(
                    lambda: export_abweichungsanalyse(manuell, str(programm)),
                    title="Abweichungsanalyse",
                    subtitle="Manuelle und Programm-Auswertung vergleichen ...",
                )
                self.status.set(f"Abweichungsanalyse erzeugt: {out}")
                self._roadmap_mark_abweichung_schulbank_v9_erledigt()
                self.show_abweichungs_editor(out)
            except Exception as exc:
                messagebox.showerror("Abweichungsanalyse", str(exc))

        def add_btn(text, cmd, color="#0b4a86"):
            tk.Button(side, text=text, command=cmd, bg=color, fg="white", activebackground=color, relief="flat", font=(theme.FONT, 11, "bold"), padx=10, pady=8).pack(fill="x", padx=16, pady=5)

        add_btn("📄 Auswertung öffnen", open_selected_output, "#0b4a86")
        add_btn("📈 Produktanalyse", product_selected, "#11823b")
        add_btn("🔍 Abweichungsanalyse", deviation_selected, "#8b5a00")
        tk.Button(side, text="Ordner gespeicherte Analysen", command=lambda: _open_folder(SAVED_ANALYSES_DIR), padx=10, pady=7).pack(fill="x", padx=16, pady=(14, 5))
        # V1.1 SP8: Zeitraum-Archivierung + Archiv-Verwaltung
        tk.Button(side, text="📦 Zeitraum archivieren",
                  command=self._archivieren_zeitraum_dialog,
                  bg="#0b6e6e", fg="white", relief="flat",
                  font=(theme.FONT, 10, "bold"), padx=10, pady=7).pack(fill="x", padx=16, pady=(14, 5))
        # V1.1 SP10: Zeitraum-Loeschen
        tk.Button(side, text="🗑  Zeitraum LOESCHEN",
                  command=self._loeschen_zeitraum_dialog,
                  bg="#9b1c1c", fg="white", relief="flat",
                  font=(theme.FONT, 10, "bold"), padx=10, pady=7).pack(fill="x", padx=16, pady=(0, 5))
        tk.Button(side, text="🗂  Archive verwalten",
                  command=self._archive_verwalten_dialog,
                  bg="#3867b7", fg="white", relief="flat",
                  font=(theme.FONT, 10, "bold"), padx=10, pady=7).pack(fill="x", padx=16, pady=(0, 5))
        tk.Button(side, text="🔐 Admin: Auswertungen löschen", command=self.open_admin_auswertungen_loeschen, bg="#9b1c1c", fg="white", relief="flat", font=(theme.FONT, 10, "bold"), padx=10, pady=7).pack(fill="x", padx=16, pady=(8, 5))

        if not rows:
            listbox.insert("end", "Keine gespeicherten Analysen in der Datenbank gefunden.")

    # Bestehende Funktionen / Werkzeuge
    def choose_file(self):
        file = filedialog.askopenfilename(filetypes=SUPPORTED_DATA_FILETYPES)
        if file:
            self.input_file.set(file)
            self.status.set(f"Datei ausgewählt: {Path(file).name}")


    def _ask_data_source(self, title="Datenquelle auswählen"):
        """Schönes Auswahlfenster im Stil der Hauptseite.\n
        Rückgabe bleibt kompatibel: PK wird intern als NMG gefiltert, weil alte Datenbankeinträge\nnoch als NMG gespeichert sind.
        """
        win = tk.Toplevel(self)
        win.resizable(True, True)
        win.title(title)
        win.geometry("560x360")
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()
        win.columnconfigure((0, 1, 2), weight=1)

        tk.Label(win, text=title, font=(theme.FONT, 20, "bold"), fg="#0b4a86", bg="#f5f7fb").grid(row=0, column=0, columnspan=3, sticky="w", padx=24, pady=(22, 4))
        tk.Label(win, text="Welche Daten sollen berücksichtigt werden?", font=(theme.FONT, 11), fg="#333", bg="#f5f7fb").grid(row=1, column=0, columnspan=3, sticky="w", padx=24, pady=(0, 16))

        choice = tk.StringVar(value="ALLE")
        cards = [
            ("PK-Daten", "Partnerkonditionskunden\nAuswertungen, Lernstände, PK-/NMG-Zuordnung", "NMG"),
            ("ZW-Daten", "Zukunftswerk-Kundendaten\nMarkt- und Produktpotenzial ohne Lernstand", "ZW"),
            ("PK + ZW", "Beide Datenwelten gemeinsam\nfür Gesamtmarkt und Produktchancen", "ALLE"),
        ]

        def select(val):
            choice.set(val)
            for frame, value in card_frames:
                selected = choice.get() == value
                frame.configure(bg="#e8f1fb" if selected else "#ffffff", highlightbackground="#0b4a86" if selected else "#d8e2ee")
                for child in frame.winfo_children():
                    child.configure(bg="#e8f1fb" if selected else "#ffffff")

        card_frames = []
        for col, (head, desc, val) in enumerate(cards):
            f = tk.Frame(win, bg="#ffffff", highlightthickness=2, highlightbackground="#d8e2ee", cursor="hand2")
            f.grid(row=2, column=col, sticky="nsew", padx=(24 if col == 0 else 8, 24 if col == 2 else 8), pady=8, ipady=8)
            tk.Label(f, text=head, font=(theme.FONT, 13, "bold"), fg="#0b4a86", bg="#ffffff").pack(pady=(14, 6), padx=10)
            tk.Label(f, text=desc, font=(theme.FONT, 9), fg="#333", bg="#ffffff", justify="center", wraplength=145).pack(padx=10, pady=(0, 12))
            f.bind("<Button-1>", lambda e, v=val: select(v))
            for child in f.winfo_children():
                child.bind("<Button-1>", lambda e, v=val: select(v))
            card_frames.append((f, val))

        result = {"value": None}
        def ok():
            result["value"] = choice.get()
            win.destroy()
        def cancel():
            win.destroy()

        buttonbar = tk.Frame(win, bg="#f5f7fb")
        buttonbar.grid(row=3, column=0, columnspan=3, sticky="e", padx=24, pady=18)
        tk.Button(buttonbar, text="Abbrechen", command=cancel, padx=18, pady=8).pack(side="right", padx=(8, 0))
        tk.Button(buttonbar, text="Analyse starten  →", command=ok, bg="#0b4a86", fg="white", activebackground="#0b4a86", relief="flat", font=(theme.FONT, 11, "bold"), padx=22, pady=9).pack(side="right")
        select("ALLE")
        self.wait_window(win)
        return result["value"]

    def import_zw_data(self):
        files = filedialog.askopenfilenames(title="ZW-Dateien auswählen", filetypes=SUPPORTED_DATA_FILETYPES)
        if not files:
            return
        name = simpledialog.askstring("ZW-Import", "Name/Kommentar für den ZW-Import:", initialvalue="ZW Daten") or "ZW Daten"
        ok = messagebox.askyesno("ZW-Daten importieren", f"{len(files)} Datei(en) als ZW-Daten importieren?\n\nDiese Daten werden NICHT als PK-Lernstand übernommen.")
        if not ok:
            return
        def _do_zw_import():
            imported_local = 0
            rows_local = 0
            errors_local = []
            for file in files:
                try:
                    r = import_historical_market_file(file, datenquelle="ZW", analyse_name=name)
                    imported_local += 1
                    rows_local += int(r.get("rows", 0))
                except Exception as exc:
                    errors_local.append(f"{Path(file).name}: {exc}")
            return imported_local, rows_local, errors_local

        imported, rows, errors = self._run_busy(
            _do_zw_import,
            title="ZW-Daten importieren",
            subtitle=f"Importiere {len(files)} Datei(en) ...",
        )
        msg = f"ZW-Import fertig.\nAusgewählte Dateien: {len(files)}\nErfolgreich importiert: {imported}\nPositionen: {rows}"
        if errors:
            msg += "\n\nFehler:\n" + "\n".join(errors[:10])
        self.status.set(msg)
        messagebox.showinfo("ZW-Import", msg)

    # SP7: market_analysis (Marktanalyse) komplett entfernt - "Produktanalyse"
    # deckt fachlich alles ab, was wir brauchen.

    def _ask_produktanalyse_kundentyp(self):
        """SP14: Fragt PK / ZW / PK+ZW ab. Liefert String oder None bei Abbruch."""
        win = tk.Toplevel(self)
        win.title("Produktanalyse")
        win.transient(self)
        win.grab_set()
        win.configure(bg="#f5f7fb")
        win.resizable(False, False)
        try:
            self.update_idletasks()
            px, py = self.winfo_x(), self.winfo_y()
            pw, ph = self.winfo_width(), self.winfo_height()
            ww, wh = 460, 280
            win.geometry(f"{ww}x{wh}+{px + max(0, (pw - ww) // 2)}+{py + max(0, (ph - wh) // 2)}")
        except Exception:
            win.geometry("460x280")

        tk.Label(win, text="Welche Auswertungen sollen ausgewertet werden?",
                 font=(theme.FONT, 13, "bold"), fg="#0b4a86", bg="#f5f7fb").pack(pady=(20, 6))
        tk.Label(win, text="Filter: nur Auswertungen der letzten 6 Monate.",
                 font=(theme.FONT, 10), fg="#444", bg="#f5f7fb").pack(pady=(0, 18))

        choice = {"value": None}

        def pick(val):
            choice["value"] = val
            win.destroy()

        btn_frame = tk.Frame(win, bg="#f5f7fb")
        btn_frame.pack()
        for label, val, color in [
            ("📊 PK", "PK", "#0b4a86"),
            ("📊 ZW", "ZW", "#3867b7"),
            ("📊 PK + ZW", "PK+ZW", "#11823b"),
        ]:
            tk.Button(btn_frame, text=label, command=lambda v=val: pick(v),
                      bg=color, fg="white", relief="flat",
                      font=(theme.FONT, 12, "bold"), padx=20, pady=10, width=10).pack(side="left", padx=6)

        tk.Button(win, text="Abbrechen", command=win.destroy, padx=12, pady=6).pack(pady=(20, 12))
        win.bind("<Escape>", lambda e: win.destroy())
        self.wait_window(win)
        return choice["value"]

    def market_opportunities(self):
        # SP14: neuer Workflow - User waehlt PK/ZW/PK+ZW, dann erzeuge
        # Produktchancen-Excel pro Kundentyp (bei PK+ZW: 3 Tabs).
        self._log_action("produktanalyse", "Produktanalyse gestartet")
        kundentyp = self._ask_produktanalyse_kundentyp()
        if not kundentyp:
            return

        # Laeuft im Hintergrund mit der Status-Box rechts oben (Animation),
        # die UI bleibt klickbar. Post-Processing im UI-Thread via on_done.
        def on_done(out):
            self._log_action("produktanalyse", "Produktanalyse erzeugt",
                             f"Kundentyp: {kundentyp} | Datei: {out}")
            self.status.set(f"Produktanalyse erzeugt: {out}")
            if messagebox.askyesno("Fertig",
                                   f"Produktanalyse {kundentyp} erstellt:\n{out}\n\nDatei jetzt oeffnen?"):
                _open_file(out)

        def on_error(exc):
            self._log_error("produktanalyse", "Produktanalyse", exc)
            messagebox.showerror("Produktanalyse-Fehler", str(exc))

        self._run_background(
            lambda: export_produktanalyse_neu(kundentyp=kundentyp, monate=6),
            title="Produktanalyse",
            subtitle=f"Erzeuge Produktanalyse {kundentyp} ...",
            progress=False,
            on_done=on_done,
            on_error=on_error,
        )

    def deviation_analysis(self):
        self._log_action("abweichungsanalyse", "Abweichungsanalyse geöffnet")
        manuell = filedialog.askopenfilename(title="Manuelle Auswertung auswählen", filetypes=SUPPORTED_DATA_FILETYPES)
        if not manuell:
            return
        programm = filedialog.askopenfilename(title="Programm-Auswertung auswählen", filetypes=SUPPORTED_DATA_FILETYPES)
        if not programm:
            return
        try:
            out = export_abweichungsanalyse(manuell, programm)
            self._log_action("abweichungsanalyse", "Abweichungsanalyse erzeugt", f"Manuell: {manuell} | Programm: {programm} | Ausgabe: {out}")
            self.status.set(f"Abweichungsanalyse erzeugt: {out}")
            self._roadmap_mark_abweichung_schulbank_v9_erledigt()
            self.show_abweichungs_editor(out)
        except Exception as exc:
            self._log_error("abweichungsanalyse", "Abweichungsanalyse", exc)
            messagebox.showerror("Abweichungsanalyse-Fehler", str(exc))

    def _ensure_abweichungs_editor_table(self):
        """Speichert den Bearbeitungsstand der Abweichungsanalyse, ohne die Excel-Datei zu verändern."""
        with sqlite3.connect(DB_PATH) as con:
            con.execute("""\nCREATE TABLE IF NOT EXISTS tbl_abweichungs_editor (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,\nquelle_datei TEXT,
                    pzn TEXT,\nartikel TEXT,
                    feld TEXT,\nmanuell TEXT,
                    programm TEXT,\nstatus TEXT DEFAULT 'offen',
                    kommentar TEXT,\nbearbeiter TEXT,
                    erstellt_am TEXT DEFAULT CURRENT_TIMESTAMP,\nbearbeitet_am TEXT,
                    UNIQUE(quelle_datei, pzn, feld, manuell, programm)\n)
            """)
            con.commit()

    def _abweichungs_editor_text_clean(self, value):
        """Normiert Freitext für robuste Vergleiche im Abweichungseditor."""
        return re.sub(r"\s+", " ", str(value or "")).strip()

    def _abweichungs_editor_text_equal(self, left, right):
        return self._abweichungs_editor_text_clean(left).lower() == self._abweichungs_editor_text_clean(right).lower()

    def _abweichungs_editor_is_already_learned(self, pzn_alt, pzn_nmg="", austauschtext=""):
        """Prüft, ob genau diese manuelle Korrektur bereits aktiv gelernt ist.\n
        Wichtig: Ein abweichender bereits gelernter Wert darf nicht unterdrückt werden,\nweil das ein echter Konflikt für die manuelle Prüfung sein kann. Unterdrückt
        wird nur, wenn die aktive Austauschdatenbank bereits dieselbe NMG-PZN oder\ndenselben Austauschtext zur Alt-PZN enthält.
        """
        pzn_alt = self._normalize_pzn_input(pzn_alt)
        pzn_nmg = self._normalize_pzn_input(pzn_nmg)
        austauschtext = self._abweichungs_editor_text_clean(austauschtext)
        if not pzn_alt or (not pzn_nmg and not austauschtext):
            return False
        try:
            with sqlite3.connect(DB_PATH) as con:
                con.row_factory = sqlite3.Row
                rows = con.execute("""\nSELECT pzn_alt, pzn_nmg, freitext_austausch
                    FROM tbl_austauschdatenbank\nWHERE COALESCE(status, 'aktiv') = 'aktiv'
                      AND (pzn_alt = ? OR printf('%08d', CAST(pzn_alt AS INTEGER)) = ?)\n""", (pzn_alt, pzn_alt)).fetchall()
            for item in rows:
                learned_nmg = self._normalize_pzn_input(item["pzn_nmg"] or "")
                learned_text = self._abweichungs_editor_text_clean(item["freitext_austausch"] or "")
                if pzn_nmg and learned_nmg and learned_nmg == pzn_nmg:
                    return True
                if austauschtext and learned_text and self._abweichungs_editor_text_equal(learned_text, austauschtext):
                    return True
        except Exception:
            return False
        return False

    def _import_abweichungs_editor_file(self, output_file):
        """Liest nur die echten fachlichen Abweichungsblätter ein.\n
        Der Editor rekonstruiert Abweichungen nicht mehr selbst aus Programm- und\nmanueller Auswertung. Quelle sind ausschließlich die vom Vergleich erzeugten
        Blätter NMG_ZUORDNUNG und AUSTAUSCH. Dadurch erscheinen keine Fälle mehr,\ndie nur gefüllt sind, aber fachlich nicht abweichen.
        """
        self._ensure_abweichungs_editor_table()
        path = Path(output_file)
        if not path.exists():
            raise FileNotFoundError(f"Abweichungsanalyse nicht gefunden: {path}")
        wb = load_workbook(path, read_only=True, data_only=True)
        wanted_sheets = [name for name in ("NMG_ZUORDNUNG", "AUSTAUSCH") if name in wb.sheetnames]
        if not wanted_sheets:
            raise ValueError("Die Datei enthält keine verwertbaren Blätter 'NMG_ZUORDNUNG' oder 'AUSTAUSCH'.")

        def read_sheet(sheet_name):
            ws = wb[sheet_name]
            rows = []
            headers = [str(v or "").strip().lower() for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

            def find_col(*needles):
                for idx, header in enumerate(headers):
                    normalized = header.replace("-", " ").replace("_", " ")
                    if all(str(n).lower() in normalized for n in needles):
                        return idx
                return None

            pzn_idx = find_col("pzn")
            artikel_idx = find_col("artikel")
            diff_idx = find_col("abweichende")
            if sheet_name == "NMG_ZUORDNUNG":
                man_idx = find_col("manuell", "pzn", "nmg")
                prog_idx = find_col("programm", "pzn", "nmg")
                feld = "PZN NMG"
            else:
                man_idx = find_col("manuell", "austausch")
                prog_idx = find_col("programm", "austausch")
                feld = "austauschbar gegen"

            if pzn_idx is None or man_idx is None or prog_idx is None:
                return rows
            for raw in ws.iter_rows(min_row=2, values_only=True):
                values = list(raw)
                pzn = self._normalize_pzn_input(values[pzn_idx] if pzn_idx < len(values) else "")
                if not pzn:
                    continue
                artikel = self._abweichungs_editor_text_clean(values[artikel_idx] if artikel_idx is not None and artikel_idx < len(values) else "")
                manuell = self._abweichungs_editor_text_clean(values[man_idx] if man_idx < len(values) else "")
                programm = self._abweichungs_editor_text_clean(values[prog_idx] if prog_idx < len(values) else "")

                if sheet_name == "NMG_ZUORDNUNG":
                    man_norm = self._normalize_pzn_input(manuell)
                    prog_norm = self._normalize_pzn_input(programm)
                    if not man_norm and not prog_norm:
                        continue
                    if man_norm == prog_norm:
                        continue
                    manuell = man_norm
                    programm = prog_norm
                    already_learned = self._abweichungs_editor_is_already_learned(pzn, pzn_nmg=manuell)
                else:
                    if not manuell and not programm:
                        continue
                    if self._abweichungs_editor_text_equal(manuell, programm):
                        continue
                    already_learned = self._abweichungs_editor_is_already_learned(pzn, austauschtext=manuell)

                status = "bereits_gelernt" if already_learned else "offen"
                abw = self._abweichungs_editor_text_clean(values[diff_idx] if diff_idx is not None and diff_idx < len(values) else feld)
                rows.append((pzn, artikel, feld, manuell, programm, status, abw))
            return rows

        imported_rows = []
        for sheet_name in wanted_sheets:
            imported_rows.extend(read_sheet(sheet_name))

        inserted = 0
        with sqlite3.connect(DB_PATH) as con:
            # Alte, noch offene automatisch eingelesene Zeilen dieser Quelle entfernen, damit
            # frühere fehlerhafte Editor-Importe nicht weiter angezeigt werden. Bewusst
            # bearbeitete Status bleiben erhalten.
            con.execute("""\nDELETE FROM tbl_abweichungs_editor
                WHERE quelle_datei = ?\nAND COALESCE(status, 'offen') IN ('offen', 'bereits_gelernt')
            """, (str(path),))
            for pzn, artikel, feld, manuell, programm, status, abw in imported_rows:
                cur = con.execute("""\nINSERT OR IGNORE INTO tbl_abweichungs_editor (
                        quelle_datei, pzn, artikel, feld, manuell, programm, status, kommentar, bearbeiter\n) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    str(path), pzn, artikel, feld, manuell, programm, status,
                    f"Abweichende Felder: {abw}" if abw else "", self.bearbeiter
                ))
                if cur.rowcount:
                    inserted += 1
            con.commit()
        return inserted

    def _abweichungs_editor_group_values(self, row):
        """Sammelt alle Abweichungsfelder derselben PZN, damit NMG-PZN und Austauschtext zusammen angezeigt/geprüft werden."""
        source = str(row["quelle_datei"] or "")
        pzn = self._normalize_pzn_input(row["pzn"])
        result = {
            "nmg_manuell": "",
            "nmg_programm": "",
            "austausch_manuell": "",
            "austausch_programm": "",
        }
        if not source or not pzn:
            return result
        try:
            with sqlite3.connect(DB_PATH) as con:
                con.row_factory = sqlite3.Row
                rows = con.execute("""\nSELECT feld, manuell, programm
                    FROM tbl_abweichungs_editor\nWHERE quelle_datei = ? AND pzn = ?
                """, (source, pzn)).fetchall()
            for item in rows:
                feld = str(item["feld"] or "").strip().lower()
                manuell = str(item["manuell"] or "").strip()
                programm = str(item["programm"] or "").strip()
                if "pzn" in feld and "nmg" in feld:
                    if manuell and not result["nmg_manuell"]:
                        result["nmg_manuell"] = self._normalize_pzn_input(manuell)
                    if programm and not result["nmg_programm"]:
                        result["nmg_programm"] = self._normalize_pzn_input(programm)
                elif "austausch" in feld:
                    if manuell and not result["austausch_manuell"]:
                        result["austausch_manuell"] = manuell
                    if programm and not result["austausch_programm"]:
                        result["austausch_programm"] = programm
        except Exception:
            pass
        return result

    def _abweichungs_editor_row_is_already_learned(self, row):
        """Prüft eine gruppierte Editor-Zeile gegen den aktuellen Lernstand.\n
        Diese Prüfung läuft zusätzlich beim Anzeigen, damit Fälle verschwinden,\ndie nach dem Öffnen/Import der Abweichungsdatei gelernt wurden.
        """
        try:
            pzn_alt = self._normalize_pzn_input(row["pzn"])
            values = self._abweichungs_editor_group_values(row)
            pzn_nmg = self._normalize_pzn_input(values.get("nmg_manuell")) or self._normalize_pzn_input(values.get("nmg_programm"))
            austauschtext = str(values.get("austausch_manuell") or values.get("austausch_programm") or "").strip()
            return self._abweichungs_editor_is_already_learned(pzn_alt, pzn_nmg=pzn_nmg, austauschtext=austauschtext)
        except Exception:
            return False

    def _is_pruefbare_abweichung_editor_row(self, row):
        """Offen zeigt nur echte fachliche Abweichungen zwischen manueller Korrektur und Programm-Ergebnis.\n
        Geprüft wird pro PZN zusammengefasst:\n- unterschiedliche NMG-PZN -> anzeigen
        - keine NMG-PZN, aber unterschiedlicher/gefüllter Austauschtext -> anzeigen\n- PZN vorhanden, aber NMG-PZN und Austauschtext leer -> nur unter Alle anzeigen
        """
        values = self._abweichungs_editor_group_values(row)
        manuell_pzn = self._normalize_pzn_input(values.get("nmg_manuell"))
        programm_pzn = self._normalize_pzn_input(values.get("nmg_programm"))
        manuell_austausch = re.sub(r"\s+", " ", str(values.get("austausch_manuell") or "")).strip().lower()
        programm_austausch = re.sub(r"\s+", " ", str(values.get("austausch_programm") or "")).strip().lower()

        if not manuell_pzn and not programm_pzn and not manuell_austausch and not programm_austausch:
            return False

        if manuell_pzn or programm_pzn:
            return manuell_pzn != programm_pzn

        return manuell_austausch != programm_austausch

    def _load_abweichungs_editor_rows(self, output_file=None, status_filter=None):
        self._ensure_abweichungs_editor_table()
        sql = """\nSELECT id, quelle_datei, pzn, artikel, feld, manuell, programm,
                   COALESCE(status, 'offen') AS status, COALESCE(kommentar, '') AS kommentar,\nbearbeiter, erstellt_am, bearbeitet_am
            FROM tbl_abweichungs_editor\nWHERE 1=1
        """
        params = []
        if output_file:
            sql += " AND quelle_datei = ?"
            params.append(str(Path(output_file)))
        if status_filter and status_filter != "alle":
            sql += " AND COALESCE(status, 'offen') = ?"
            params.append(status_filter)
        sql += " ORDER BY datetime(COALESCE(bearbeitet_am, erstellt_am)) DESC, id DESC"
        with sqlite3.connect(DB_PATH) as con:
            con.row_factory = sqlite3.Row
            rows = con.execute(sql, params).fetchall()

        if status_filter == "offen":
            rows = [
                row for row in rows
                if self._is_pruefbare_abweichung_editor_row(row)
                and not self._abweichungs_editor_row_is_already_learned(row)
            ]

        # Im Editor darf jede Alt-PZN nur einmal erscheinen.
        # Mehrere Feldabweichungen / Varianten werden in der Detailanzeige zusammengefasst.
        grouped = {}
        for row in rows:
            key = (str(row["quelle_datei"] or ""), self._normalize_pzn_input(row["pzn"]))
            if not key[1]:
                continue
            if key not in grouped:
                grouped[key] = row
        return list(grouped.values())

    def _abweichungs_editor_pzn_summary(self, row):
        """Zählt alle Editor-Zeilen und Varianten zu einer Alt-PZN für Anzeige/Aktionen."""
        source = str(row["quelle_datei"] or "")
        pzn = self._normalize_pzn_input(row["pzn"])
        result = {"ids": [], "anzahl": 0, "nmg_varianten": [], "austausch_varianten": []}
        if not source or not pzn:
            return result
        try:
            with sqlite3.connect(DB_PATH) as con:
                con.row_factory = sqlite3.Row
                rows = con.execute("""\nSELECT id, feld, manuell, programm
                    FROM tbl_abweichungs_editor\nWHERE quelle_datei = ? AND pzn = ?
                    ORDER BY id\n""", (source, pzn)).fetchall()
            nmg = []
            austausch = []
            for item in rows:
                result["ids"].append(int(item["id"]))
                feld = str(item["feld"] or "").strip().lower()
                values = [str(item["manuell"] or "").strip(), str(item["programm"] or "").strip()]
                if "pzn" in feld and "nmg" in feld:
                    for value in values:
                        normalized = self._normalize_pzn_input(value)
                        if normalized and normalized not in nmg:
                            nmg.append(normalized)
                elif "austausch" in feld:
                    for value in values:
                        cleaned = re.sub(r"\s+", " ", value).strip()
                        if cleaned and cleaned not in austausch:
                            austausch.append(cleaned)
            result["anzahl"] = len(rows)
            result["nmg_varianten"] = nmg
            result["austausch_varianten"] = austausch
        except Exception:
            pass
        return result

    def _lookup_editor_artikel_basis(self, pzn):
        """Liest Artikel, DF, PCK, APU und NMG-Rabatt aus den verfügbaren Stammdaten."""
        pzn = self._normalize_pzn_input(pzn)
        result = {"artikel": "", "df": "", "pck": "", "apu": "", "nmg_rabatt": ""}
        if not pzn:
            return result
        candidates = [
            ("tbl_artikelstamm", {
                "artikel": ["artikel", "artikelname", "bezeichnung"],
                "df": ["df", "dar", "darreichungsform"],
                "pck": ["pck", "packung", "pack"],
                "apu": ["apu", "hap", "taxe_ek", "taxe_vk"],
                "nmg_rabatt": ["nmg_rabatt", "rabatt", "rabatt_prozent", "pk_rabatt"],
            }),
            ("tbl_pzn_basisdaten", {
                "artikel": ["artikel", "artikelname", "bezeichnung"],
                "df": ["df", "dar", "darreichungsform"],
                "pck": ["pck", "packung", "pack"],
                "apu": ["apu", "hap", "taxe_ek", "taxe_vk"],
                "nmg_rabatt": ["nmg_rabatt", "rabatt", "rabatt_prozent", "pk_rabatt"],
            }),
            ("tbl_nmg_stamm", {
                "artikel": ["artikel", "artikelname", "bezeichnung"],
                "df": ["df", "dar", "darreichungsform"],
                "pck": ["pck", "packung", "pack"],
                "apu": ["apu", "hap", "taxe_ek", "taxe_vk"],
                "nmg_rabatt": ["nmg_rabatt", "rabatt", "rabatt_prozent", "pk_rabatt"],
            }),
            ("nmg_rabatte", {
                "nmg_rabatt": ["nmg_rabatt", "rabatt", "rabatt_prozent", "pk_rabatt", "wert"],
            }),
        ]
        try:
            with sqlite3.connect(DB_PATH) as con:
                con.row_factory = sqlite3.Row
                for table, fields in candidates:
                    try:
                        cols = {r[1].lower(): r[1] for r in con.execute(f"PRAGMA table_info({table})").fetchall()}
                        if not cols or "pzn" not in cols:
                            continue
                        wanted = []
                        aliases = []
                        for alias, names in fields.items():
                            col = next((cols[n.lower()] for n in names if n.lower() in cols), None)
                            if col:
                                wanted.append(f"{col} AS {alias}")
                                aliases.append(alias)
                        if not wanted:
                            continue
                        row = con.execute(f"SELECT {', '.join(wanted)} FROM {table} WHERE {cols['pzn']} = ? LIMIT 1", (pzn,)).fetchone()
                        if not row:
                            continue
                        for alias in aliases:
                            value = str(row[alias] or "").strip()
                            if value and not result.get(alias):
                                result[alias] = value
                    except Exception:
                        continue
        except Exception:
            pass
        return result

    def _abweichungs_editor_changed_fields_text(self, row):
        """Beschreibt kompakt, welche fachlichen Felder wirklich voneinander abweichen."""
        values = self._abweichungs_editor_group_values(row)
        manuell_pzn = self._normalize_pzn_input(values.get("nmg_manuell"))
        programm_pzn = self._normalize_pzn_input(values.get("nmg_programm"))
        manuell_austausch = re.sub(r"\s+", " ", str(values.get("austausch_manuell") or "")).strip()
        programm_austausch = re.sub(r"\s+", " ", str(values.get("austausch_programm") or "")).strip()
        changed = []
        if manuell_pzn != programm_pzn and (manuell_pzn or programm_pzn):
            changed.append("NMG-PZN")
        if manuell_austausch.lower() != programm_austausch.lower() and (manuell_austausch or programm_austausch):
            changed.append("Austauschbar gegen")
        return ", ".join(changed) if changed else "keine Artikelabweichung"

    def _abweichungs_editor_display_values(self, row):
        """Bereitet die kompakten Anzeigespalten des Abweichungseditors auf."""
        pzn = self._normalize_pzn_input(row["pzn"])
        basis = self._lookup_editor_artikel_basis(pzn)
        artikel = str(row["artikel"] or "").strip() or basis.get("artikel", "")
        df_value = basis.get("df", "")
        pck_value = basis.get("pck", "")

        values = self._abweichungs_editor_group_values(row)
        # Anzeige bevorzugt die manuelle/korrigierte Angabe; falls leer, wird das Programm-Ergebnis gezeigt.
        nmg_pzn = self._normalize_pzn_input(values.get("nmg_manuell")) or self._normalize_pzn_input(values.get("nmg_programm"))
        austausch_value = str(values.get("austausch_manuell") or values.get("austausch_programm") or "").strip()
        nmg_basis = self._lookup_editor_artikel_basis(nmg_pzn) if nmg_pzn else {"artikel": "", "apu": "", "nmg_rabatt": ""}
        nmg_artikel = nmg_basis.get("artikel", "")
        changed_fields = self._abweichungs_editor_changed_fields_text(row)
        return artikel, df_value, pck_value, nmg_pzn, nmg_artikel, austausch_value, changed_fields, nmg_basis.get("nmg_rabatt", ""), nmg_basis.get("apu", "")

    def show_abweichungs_editor(self, output_file=None):
        """Editor für Abweichungsanalyse: Abweichungen prüfen, markieren und als Lernvorschlag übernehmen.\n
        Der Editor läuft bewusst in einem eigenen Fenster, damit die aktuell geöffnete\nHauptseite erhalten bleibt. So kann zwischen Analyse, Schulbank und Editor
        gewechselt werden, ohne dass der Bearbeitungsstand aus der Ansicht verschwindet.\n"""
        if output_file:
            try:
                self._import_abweichungs_editor_file(output_file)
            except Exception as exc:
                messagebox.showwarning("Abweichungsanalyse-Editor", f"Abweichungsdatei konnte nicht vollständig eingelesen werden:\n{exc}")

        win = tk.Toplevel(self)
        win.resizable(True, True)
        win.title("Abweichungsanalyse-Editor")
        win.geometry("1280x760")
        win.minsize(1050, 620)
        win.configure(bg="#f5f7fb")
        # Eigenständiges Fenster: soll in der Taskleiste erscheinen, minimierbar und maximierbar bleiben.
        try:
            win.state("zoomed")
        except Exception:
            pass
        win.columnconfigure(0, weight=1)
        win.rowconfigure(1, weight=1)

        header = tk.Frame(win, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        header.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 8))
        header.columnconfigure(0, weight=1)
        tk.Label(
            header,
            text="Abweichungsanalyse-Editor",
            font=(theme.FONT, 20, "bold"),
            fg="#0b4a86",
            bg="#ffffff"
        ).grid(row=0, column=0, sticky="w", padx=14, pady=(12, 2))
        tk.Label(
            header,
            text="Abweichungen im Programm prüfen. Die Excel-Ausgabe bleibt erhalten; Lernvorschläge entstehen erst durch bewusste Übernahme.",
            font=(theme.FONT, 10),
            fg="#333",
            bg="#ffffff"
        ).grid(row=1, column=0, sticky="w", padx=14, pady=(0, 12))
        win_buttons = tk.Frame(header, bg="#ffffff")
        win_buttons.grid(row=0, column=1, rowspan=2, sticky="e", padx=14, pady=12)
        tk.Button(win_buttons, text="Minimieren", command=win.iconify, padx=10, pady=6).pack(side="left", padx=(0, 6))
        tk.Button(win_buttons, text="Maximieren", command=lambda: win.state("zoomed"), padx=10, pady=6).pack(side="left", padx=(0, 6))
        tk.Button(win_buttons, text="Fenster schließen", command=win.destroy, padx=12, pady=6).pack(side="left")

        body = tk.Frame(win, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        body.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 14))
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=0)
        body.rowconfigure(1, weight=1)

        top = tk.Frame(body, bg="#ffffff")
        top.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 8))
        tk.Label(top, text="Filter:", bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 10, "bold")).pack(side="left", padx=(0, 6))
        status_var = tk.StringVar(value="offen")
        source_label = Path(output_file).name if output_file else "alle Abweichungen"
        tk.Label(top, text=f"Quelle: {source_label}", bg="#ffffff", fg="#333").pack(side="right")

        columns = ("id", "anzahl", "pzn", "artikel", "df", "pck", "nmg_pzn", "nmg_artikel", "austauschbar", "abweichende_felder", "nmg_rabatt", "apu", "status")
        tree = ttk.Treeview(body, columns=columns, show="headings", selectmode="extended")
        tree.grid(row=1, column=0, sticky="nsew")
        headings = {
            "id": "ID",
            "anzahl": "Anzahl",
            "pzn": "PZN",
            "artikel": "Artikel",
            "df": "DF",
            "pck": "PCK (PACK)",
            "nmg_pzn": "NMG-PZN",
            "nmg_artikel": "NMG-Artikel",
            "austauschbar": "Austauschbar gegen",
            "abweichende_felder": "Abweichende Felder",
            "nmg_rabatt": "NMG-Rabatt",
            "apu": "APU",
            "status": "Status",
        }
        widths = {
            "id": 60,
            "anzahl": 70,
            "pzn": 95,
            "artikel": 280,
            "df": 80,
            "pck": 120,
            "nmg_pzn": 105,
            "nmg_artikel": 260,
            "austauschbar": 300,
            "abweichende_felder": 210,
            "nmg_rabatt": 110,
            "apu": 95,
            "status": 120,
        }
        display_columns_without_status = ("id", "anzahl", "pzn", "artikel", "df", "pck", "nmg_pzn", "nmg_artikel", "austauschbar", "abweichende_felder", "nmg_rabatt", "apu")
        display_columns_with_status = ("id", "anzahl", "pzn", "artikel", "df", "pck", "nmg_pzn", "nmg_artikel", "austauschbar", "abweichende_felder", "nmg_rabatt", "apu", "status")
        for col in columns:
            tree.heading(col, text=headings[col])
            tree.column(col, width=widths[col], anchor="w", stretch=(col in {"artikel", "nmg_artikel", "austauschbar"}))
        try:
            self._make_tree_sortable(tree, columns, headings)
        except Exception:
            pass
        scroll = tk.Scrollbar(body, orient="vertical", command=tree.yview)
        scroll.grid(row=1, column=0, sticky="nse")
        hscroll = tk.Scrollbar(body, orient="horizontal", command=tree.xview)
        hscroll.grid(row=2, column=0, sticky="ew", pady=(0, 6))
        tree.configure(yscrollcommand=scroll.set, xscrollcommand=hscroll.set)

        side = tk.Frame(body, bg="#f8fbff", highlightbackground="#d8e2ee", highlightthickness=1, width=285)
        side.grid(row=1, column=1, sticky="ns", padx=(16, 0))
        side.grid_propagate(False)
        detail = tk.StringVar(value="Abweichung auswählen.")
        tk.Label(side, text="Aktionen", font=(theme.FONT, 14, "bold"), fg="#0b4a86", bg="#f8fbff").pack(anchor="w", padx=16, pady=(16, 8))
        tk.Label(side, textvariable=detail, justify="left", wraplength=245, bg="#f8fbff", fg="#333").pack(anchor="w", padx=16, pady=(0, 14))

        row_map = {}

        def reload_rows():
            for iid in tree.get_children(""):
                tree.delete(iid)
            row_map.clear()
            current_filter = status_var.get()
            tree["displaycolumns"] = display_columns_with_status if current_filter == "alle" else display_columns_without_status
            rows = self._load_abweichungs_editor_rows(output_file, current_filter)
            for row in rows:
                artikel, df_value, pck_value, nmg_value, nmg_artikel, austausch_value, changed_fields, nmg_rabatt, apu = self._abweichungs_editor_display_values(row)
                summary = self._abweichungs_editor_pzn_summary(row)
                iid = tree.insert("", "end", values=(
                    row["id"], summary.get("anzahl", 1) or 1, row["pzn"], artikel, df_value, pck_value,
                    nmg_value, nmg_artikel, austausch_value, changed_fields, nmg_rabatt, apu, row["status"]
                ))
                row_map[iid] = {"row": row, "ids": summary.get("ids") or [int(row["id"])], "summary": summary}
            self.status.set(f"Abweichungsanalyse-Editor: {len(rows)} Einträge angezeigt.")

        def selected_ids():
            sel = tree.selection()
            if not sel:
                messagebox.showinfo("Abweichungsanalyse-Editor", "Bitte zuerst einen oder mehrere Einträge auswählen.")
                return []
            ids = []
            for iid in sel:
                data = row_map.get(iid)
                if isinstance(data, dict):
                    ids.extend(int(x) for x in data.get("ids", []) if x)
                else:
                    try:
                        ids.append(int(tree.item(iid, "values")[0]))
                    except Exception:
                        pass
            return sorted(set(ids))

        def selected_row(show_message=True):
            sel = tree.selection()
            if not sel:
                if show_message:
                    messagebox.showinfo("Abweichungsanalyse-Editor", "Bitte zuerst einen Eintrag auswählen.")
                return None
            data = row_map.get(sel[0])
            if isinstance(data, dict):
                return data.get("row")
            return data

        def update_detail(event=None):
            row = selected_row(show_message=False)
            if not row:
                detail.set("Bitte einen Eintrag auswählen.")
                return
            artikel, df_value, pck_value, nmg_value, nmg_artikel, austausch_value, changed_fields, nmg_rabatt, apu = self._abweichungs_editor_display_values(row)
            summary = self._abweichungs_editor_pzn_summary(row)
            nmg_varianten = "; ".join(summary.get("nmg_varianten") or [])
            austausch_varianten = "; ".join(summary.get("austausch_varianten") or [])
            detail.set(
                f"PZN: {row['pzn']}\n"
                f"Artikel: {artikel or ''}\n"
                f"DF: {df_value or ''}\n"
                f"PCK: {pck_value or ''}\n"
                f"NMG-PZN: {nmg_value or ''}\n"
                f"NMG-Artikel: {nmg_artikel or ''}\n"
                f"Abweichende Felder: {changed_fields or ''}\n"
                f"NMG-Rabatt: {nmg_rabatt or ''}\n"
                f"APU: {apu or ''}\n"
                f"Anzahl Abweichungen: {summary.get('anzahl', 1) or 1}\n\n"
                f"NMG-PZN-Varianten:\n{nmg_varianten or '-'}\n\n"
                f"Austausch-Varianten:\n{austausch_varianten or '-'}\n\n"
                f"Status: {row['status']}"
            )

        tree.bind("<<TreeviewSelect>>", update_detail)

        def set_status(status):
            ids = selected_ids()
            if not ids:
                return
            kommentar = ""
            if status in ("falsch", "ignoriert", "abgelehnt"):
                kommentar = simpledialog.askstring("Kommentar", "Kommentar optional:") or ""
            if status == "manuelle_pruefung":
                kommentar = kommentar or "Zur manuellen Prüfung aus Abweichungseditor"
            with sqlite3.connect(DB_PATH) as con:
                con.executemany("""\nUPDATE tbl_abweichungs_editor
                    SET status = ?, kommentar = COALESCE(NULLIF(?, ''), kommentar),\nbearbeitet_am = CURRENT_TIMESTAMP, bearbeiter = ?
                    WHERE id = ?\n""", [(status, kommentar, self.bearbeiter, item_id) for item_id in ids])
                con.commit()
            reload_rows()
            if status == "manuelle_pruefung":
                messagebox.showinfo("Manuelle Prüfung", "Der Fall wurde in die Schulbank → Manuelle Prüfung übergeben.")

        def save_as_lernvorschlag():
            row = selected_row()
            if not row:
                return
            changed_fields = self._abweichungs_editor_changed_fields_text(row)
            if changed_fields == "keine Artikelabweichung":
                messagebox.showinfo("Lernvorschlag", "Dieser Fall enthält keine abweichende NMG-PZN und keinen abweichenden Austauschtext.")
                return
            pzn_alt = self._normalize_pzn_input(row["pzn"])
            values = self._abweichungs_editor_group_values(row)
            pzn_nmg = self._normalize_pzn_input(values.get("nmg_manuell")) or self._normalize_pzn_input(values.get("nmg_programm"))
            freitext = str(values.get("austausch_manuell") or values.get("austausch_programm") or "").strip()
            if not pzn_nmg and not freitext:
                freitext = simpledialog.askstring("Lernvorschlag", "Austauschbar gegen / Freitext:", initialvalue="") or ""
            if not pzn_alt:
                messagebox.showwarning("Lernvorschlag", "Keine gültige PZN vorhanden.")
                return
            if not pzn_nmg and not freitext:
                messagebox.showwarning("Lernvorschlag", "Bitte PZN NMG oder Austauschtext angeben.")
                return
            if pzn_nmg and pzn_alt == pzn_nmg:
                messagebox.showinfo("Lernvorschlag", "PZN alt und PZN NMG sind identisch. Kein Lernvorschlag nötig.")
                return
            if self._insert_or_update_lernvorschlag(pzn_alt, row["artikel"] or "", pzn_nmg, freitext, "Neue Lernvorschläge"):
                with sqlite3.connect(DB_PATH) as con:
                    ids = selected_ids() or [int(row["id"])]
                    con.executemany("""\nUPDATE tbl_abweichungs_editor
                        SET status='lernvorschlag', bearbeitet_am=CURRENT_TIMESTAMP, bearbeiter=?, kommentar='Als Lernvorschlag gespeichert'\nWHERE id=?
                    """, [(self.bearbeiter, item_id) for item_id in ids])
                    con.commit()
                messagebox.showinfo("Lernvorschlag", "Lernvorschlag wurde gespeichert.")
                reload_rows()

        def add_btn(text, cmd, color="#0b4a86"):
            tk.Button(side, text=text, command=cmd, bg=color, fg="white", activebackground=color,
                      relief="flat", font=(theme.FONT, 10, "bold"), padx=10, pady=8).pack(fill="x", padx=16, pady=4)

        for label, val in (("Offen", "offen"), ("Richtig", "richtig"), ("Falsch", "falsch"), ("Lernvorschlag", "lernvorschlag"), ("Abgelehnt", "abgelehnt"), ("Ignoriert", "ignoriert"), ("Alle", "alle")):
            tk.Radiobutton(top, text=label, variable=status_var, value=val, command=reload_rows, bg="#ffffff").pack(side="left", padx=(0, 8))

        add_btn("✓ Richtig erkannt", lambda: set_status("richtig"), "#11823b")
        add_btn("✗ Falsch erkannt", lambda: set_status("falsch"), "#9b1c1c")
        add_btn("🎓 Für Schulbank lernen", save_as_lernvorschlag, "#0b4a86")
        add_btn("🔎 In manuelle Prüfung", lambda: set_status("manuelle_pruefung"), "#6b4fb3")
        add_btn("✗ Ablehnen", lambda: set_status("abgelehnt"), "#9b1c1c")
        add_btn("↷ Ignorieren", lambda: set_status("ignoriert"), "#8b5a00")
        tk.Button(side, text="Excel öffnen", command=lambda: _open_file(Path(output_file)) if output_file else None, padx=10, pady=7).pack(fill="x", padx=16, pady=(14, 4))
        tk.Button(side, text="Aktualisieren", command=reload_rows, padx=10, pady=7).pack(fill="x", padx=16, pady=4)

        reload_rows()

    def report_callback_exception(self, exc, val, tb):
        """Zentrale Fehlerprotokollierung für Tkinter-Callbacks."""
        try:
            import traceback as _traceback
            details = "".join(_traceback.format_exception(exc, val, tb))
            log_event("fehler", "Programmfehler", details, "ERROR", user=getattr(self, "bearbeiter", ""))
        except Exception:
            pass
        try:
            messagebox.showerror("Programmfehler", f"Es ist ein Fehler aufgetreten:\n{val}")
        except Exception:
            pass

    def _log_action(self, category, action, details=""):
        try:
            return log_event(category, action, details, user=getattr(self, "bearbeiter", ""))
        except Exception:
            return None

    def _log_error(self, category, action, exc):
        try:
            return log_exception(category, action, exc, user=getattr(self, "bearbeiter", ""))
        except Exception:
            return None

    def show_protocol_center(self):
        """Eigenes Fenster für Protokolle: anzeigen, mailen, Supportpaket, Admin-Löschen."""
        self._log_action("protokolle", "Protokoll-Center geöffnet")
        win = tk.Toplevel(self)
        win.resizable(True, True)
        win.title("Protokolle")
        win.geometry("1180x720")
        win.minsize(980, 560)
        win.configure(bg="#f5f7fb")
        try:
            win.state("zoomed")
        except Exception:
            pass
        win.columnconfigure(0, weight=1)
        win.rowconfigure(1, weight=1)

        header = tk.Frame(win, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        header.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 8))
        header.columnconfigure(0, weight=1)
        tk.Label(header, text="Protokolle", font=(theme.FONT, 20, "bold"), fg="#0b4a86", bg="#ffffff").grid(row=0, column=0, sticky="w", padx=14, pady=(12, 2))
        tk.Label(header, text=f"Protokollverzeichnis: {PROTOCOL_ROOT}", font=(theme.FONT, 10), fg="#333", bg="#ffffff").grid(row=1, column=0, sticky="w", padx=14, pady=(0, 12))
        tk.Button(header, text="Ordner öffnen", command=lambda: _open_folder(PROTOCOL_ROOT), padx=12, pady=6).grid(row=0, column=1, rowspan=2, sticky="e", padx=14, pady=12)

        body = tk.Frame(win, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        body.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 14))
        body.columnconfigure(0, weight=0)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(1, weight=1)

        filter_frame = tk.Frame(body, bg="#ffffff")
        filter_frame.grid(row=0, column=0, columnspan=2, sticky="ew", padx=12, pady=10)
        tk.Label(filter_frame, text="Kategorie:", bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 10, "bold")).pack(side="left", padx=(0, 8))
        category_var = tk.StringVar(value="Alle")
        categories = ["Alle", "Programm", "Fehler", "Neue_Auswertung", "Schulbank", "Datenaktualisierung", "Produktanalyse", "Marktanalyse", "Abweichungsanalyse", "Update_Backup", "Admin", "Protokolle"]
        category_box = ttk.Combobox(filter_frame, textvariable=category_var, values=categories, state="readonly", width=28)
        category_box.pack(side="left")

        columns = ("category", "name", "mtime", "size")
        tree = ttk.Treeview(body, columns=columns, show="headings", selectmode="browse", height=18)
        tree.grid(row=1, column=0, sticky="nsew", padx=(12, 8), pady=(0, 12))
        for col, label, width in [
            ("category", "Bereich", 155),
            ("name", "Datei", 180),
            ("mtime", "Geändert", 145),
            ("size", "Größe", 80),
        ]:
            tree.heading(col, text=label)
            tree.column(col, width=width, anchor="w")
        yscroll = tk.Scrollbar(body, orient="vertical", command=tree.yview)
        yscroll.grid(row=1, column=0, sticky="nse", padx=(0, 8), pady=(0, 12))
        tree.configure(yscrollcommand=yscroll.set)

        right = tk.Frame(body, bg="#ffffff")
        right.grid(row=1, column=1, sticky="nsew", padx=(0, 12), pady=(0, 12))
        right.columnconfigure(0, weight=1)
        right.rowconfigure(0, weight=1)
        text_box = tk.Text(right, wrap="word", font=("Consolas", 10), bg="#fbfdff")
        text_box.grid(row=0, column=0, sticky="nsew")
        text_scroll = tk.Scrollbar(right, orient="vertical", command=text_box.yview)
        text_scroll.grid(row=0, column=1, sticky="ns")
        text_box.configure(yscrollcommand=text_scroll.set)

        buttonbar = tk.Frame(right, bg="#ffffff")
        buttonbar.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        files = []

        def selected_file():
            sel = tree.selection()
            if not sel:
                messagebox.showinfo("Protokolle", "Bitte zuerst ein Protokoll auswählen.")
                return None
            idx = int(sel[0])
            if idx < 0 or idx >= len(files):
                return None
            return files[idx]["path"]

        def reload_list(*args):
            nonlocal files
            for iid in tree.get_children(""):
                tree.delete(iid)
            selected = category_var.get()
            category = None if selected == "Alle" else selected.lower()
            # Anzeigenamen werden intern als Ordnernamen genutzt.
            files = list_protocol_files(category=None if selected == "Alle" else selected)
            for idx, item in enumerate(files):
                size = f"{int(item['size'] / 1024)} KB" if item["size"] >= 1024 else f"{item['size']} B"
                tree.insert("", "end", iid=str(idx), values=(item["category"], item["name"], item["mtime"], size))
            text_box.delete("1.0", "end")
            text_box.insert("end", f"{len(files)} Protokoll(e) gefunden.\n")

        def show_selected(event=None):
            path = selected_file()
            if not path:
                return
            try:
                text_box.delete("1.0", "end")
                text_box.insert("end", read_protocol_file(path))
                self._log_action("protokolle", "Protokoll angezeigt", str(path))
            except Exception as exc:
                self._log_error("protokolle", "Protokoll anzeigen", exc)
                messagebox.showerror("Protokolle", str(exc))

        def send_selected():
            path = selected_file()
            if not path:
                return
            try:
                attached = open_mail_with_attachment(path, DEFAULT_RECIPIENT, "NMGone Protokoll")
                if attached:
                    messagebox.showinfo("Mail", "Mailentwurf wurde mit Anhang geöffnet.")
                else:
                    messagebox.showinfo("Mail", "Standard-Mailprogramm wurde geöffnet. Falls kein Anhang vorhanden ist, bitte die Datei manuell anhängen.")
                self._log_action("protokolle", "Protokoll per Mail vorbereitet", str(path))
            except Exception as exc:
                self._log_error("protokolle", "Protokoll per Mail senden", exc)
                messagebox.showerror("Mail", str(exc))

        def support_package():
            try:
                package = create_support_package()
                self._log_action("protokolle", "Supportpaket erstellt", str(package))
                if messagebox.askyesno("Supportpaket", f"Supportpaket erstellt:\n{package}\n\nJetzt per Mail senden?"):
                    open_mail_with_attachment(package, DEFAULT_RECIPIENT, "NMGone Supportpaket")
            except Exception as exc:
                self._log_error("protokolle", "Supportpaket erstellen", exc)
                messagebox.showerror("Supportpaket", str(exc))

        def delete_selected_admin():
            path = selected_file()
            if not path:
                return
            password = simpledialog.askstring("Admin", "Admin-Passwort zum Löschen:", show="*")
            if password != ADMIN_DB_PASSWORD:
                messagebox.showwarning("Admin", "Falsches Passwort. Protokoll wurde nicht gelöscht.")
                self._log_action("admin", "Protokoll löschen abgelehnt", str(path))
                return
            if not messagebox.askyesno("Protokoll löschen", f"Dieses Protokoll wirklich löschen?\n\n{path}"):
                return
            try:
                delete_protocol_file(path)
                self._log_action("admin", "Protokoll gelöscht", str(path))
                reload_list()
            except Exception as exc:
                self._log_error("admin", "Protokoll löschen", exc)
                messagebox.showerror("Admin", str(exc))

        tk.Button(buttonbar, text="Aktualisieren", command=reload_list, padx=12, pady=7).pack(side="left", padx=(0, 8))
        tk.Button(buttonbar, text="Per Mail senden", command=send_selected, bg="#0b4a86", fg="white", relief="flat", font=(theme.FONT, 10, "bold"), padx=14, pady=8).pack(side="left", padx=(0, 8))
        tk.Button(buttonbar, text="Supportpaket erstellen", command=support_package, bg="#11823b", fg="white", relief="flat", font=(theme.FONT, 10, "bold"), padx=14, pady=8).pack(side="left", padx=(0, 8))
        tk.Button(buttonbar, text="Admin: Löschen", command=delete_selected_admin, bg="#9b1c1c", fg="white", relief="flat", font=(theme.FONT, 10, "bold"), padx=14, pady=8).pack(side="right")

        category_box.bind("<<ComboboxSelected>>", reload_list)
        tree.bind("<<TreeviewSelect>>", show_selected)
        reload_list()

    def show_backup_dialog(self):
        win = tk.Toplevel(self)
        win.resizable(True, True)
        win.title("Backup & Wiederherstellung")
        win.geometry("480x220")
        tk.Label(win, text="Backup & Wiederherstellung", font=(theme.FONT, 16, "bold")).pack(pady=16)
        tk.Label(win, text="Beim Programmstart wird automatisch ein Tagesbackup erstellt.\nEs werden die letzten 7 Auto-Backups behalten.").pack(pady=4)
        tk.Button(win, text="Backup jetzt manuell erstellen", command=self.create_backup, bg="#0b4a86", fg="white", padx=18, pady=8).pack(pady=6)
        tk.Button(win, text="Backup wiederherstellen", command=self.restore_backup, padx=18, pady=8).pack(pady=6)
        tk.Button(win, text="Backup-Ordner öffnen", command=lambda: _open_folder(BACKUP_DIR)).pack(pady=6)

    def create_backup(self):
        self._log_action("update_backup", "Backup manuell gestartet")
        try:
            out = self._run_busy(
                backup_erstellen,
                title="Backup wird erstellt",
                subtitle="Daten werden gesichert ...",
            )
            self._log_action("update_backup", "Backup erstellt", str(out))
            self.status.set(f"Backup erstellt: {out}")
            messagebox.showinfo("Backup erstellt", f"Backup wurde erstellt:\n{out}")
        except Exception as exc:
            self._log_error("update_backup", "Backup erstellen", exc)
            messagebox.showerror("Backup-Fehler", str(exc))

    def restore_backup(self):
        self._log_action("update_backup", "Backup-Wiederherstellung geöffnet")
        file = filedialog.askopenfilename(title="Backup auswählen", filetypes=[("NMG Backup", "*.zip"), ("Alle Dateien", "*.*")])
        if not file:
            return
        try:
            info = backup_pruefen(file)
            details = (f"Backup-Datei:\n{file}\n\nVersion: {info.get('app_version','unbekannt')}\nErstellt: {info.get('created_at','unbekannt')}\n\nAktuelle Datenbank wird vorher automatisch gesichert. Wiederherstellen?")
            if not messagebox.askyesno("Backup wiederherstellen", details):
                return
            db = self._run_busy(
                lambda: backup_wiederherstellen(file),
                title="Backup wird wiederhergestellt",
                subtitle="Daten werden zurueckgespielt ...",
            )
            self._log_action("update_backup", "Backup wiederhergestellt", f"Quelle: {file} | Ziel: {db}")
            init_db(DB_PATH)
            self.status.set(f"Backup wiederhergestellt: {db}")
            messagebox.showinfo("Wiederherstellung fertig", f"Datenbank wurde wiederhergestellt:\n{db}")
        except Exception as exc:
            self._log_error("update_backup", "Backup wiederherstellen", exc)
            messagebox.showerror("Restore-Fehler", str(exc))


    def install_update_dialog(self):
        # SP5: erst nach lokalem NMGone_Setup_*.exe schauen. Wenn da, anbieten.
        # Sonst alte Logik mit .nmgupdate-Paketen.
        pending = self._find_pending_local_setup()
        if pending:
            path, version_str = pending
            if messagebox.askyesno(
                "Lokales Setup gefunden",
                f"Im Update-Ordner liegt ein Setup bereit:\n\n"
                f"Datei: {path.name}\n"
                f"Version: V{version_str}\n"
                f"Aktuell installiert: V{APP_VERSION}\n\n"
                f"Jetzt installieren?"
                f"{self.SMARTSCREEN_HINT}"
            ):
                self._launch_setup_and_exit(path)
                return
            # Wer "Nein" sagt, geht weiter in den .nmgupdate-Pfad.

        packages = []
        try:
            packages = find_update_packages()
        except Exception:
            packages = []

        chosen_file = None
        if packages:
            newest = next((p for p in packages if p.get("is_newer")), packages[0])
            label = "neues Update" if newest.get("is_newer") else "Updatepaket"
            msg = (
                f"Gefundenes {label} im Update-Ordner:\n\n"
                f"{newest.get('name')}\n"
                f"Zielversion: {newest.get('target_version')}\n"
                f"Aktuelle Version: {APP_VERSION}\n\n"
                f"Dieses Update jetzt installieren?"
            )
            if messagebox.askyesno("Update gefunden", msg):
                chosen_file = str(newest.get("path"))

        if not chosen_file:
            chosen_file = filedialog.askopenfilename(
                title="NMG-Updatepaket auswählen",
                initialdir=str(UPDATE_DIR),
                filetypes=[("NMG Update", "*.nmgupdate"), ("ZIP", "*.zip"), ("Alle Dateien", "*.*")]
            )
        if not chosen_file:
            return
        try:
            manifest = validate_update_package(chosen_file)
            target = manifest.get("target_version", "unbekannt")
            notes = manifest.get("notes", "")
            text = (
                f"Updatepaket:\n{Path(chosen_file).name}\n\n"
                f"Aktuelle Version: {APP_VERSION}\n"
                f"Zielversion: {target}\n\n"
                f"Vor dem Update wird automatisch ein Backup und ein Rücksprungpunkt erstellt.\n"
                f"Datenbank, gespeicherte Analysen und Backups werden nicht überschrieben.\n\n"
                f"Nach erfolgreichem Update wird das Programm automatisch neu gestartet.\n\n"
                f"Hinweise:\n{notes}\n\nUpdate installieren?"
            )
            if not messagebox.askyesno("Update installieren", text):
                return
            result = install_update_package(chosen_file)
            msg = (
                f"Update installiert.\n\n"
                f"Zielversion: {result.get('target_version')}\n"
                f"Kopierte Dateien: {len(result.get('copied', []))}\n"
                f"Backup: {result.get('backup')}\n"
                f"Rücksprungpunkt: {result.get('rollback')}\n\n"
                f"Das Programm wird jetzt neu gestartet."
            )
            self.status.set(msg)
            messagebox.showinfo("Update installiert", msg)
            try:
                restart_application()
                self.destroy()
            except Exception as restart_exc:
                messagebox.showwarning("Neustart", f"Update ist installiert, aber der automatische Neustart ist fehlgeschlagen.\n\n{restart_exc}\n\nBitte Programm manuell neu starten.")
        except Exception as exc:
            messagebox.showerror("Update-Fehler", str(exc))

    def open_update_folder(self):
        try:
            open_updates_folder()
        except Exception as exc:
            messagebox.showerror("Update-Ordner", str(exc))


    def rollback_dialog(self):
        snaps = list_rollback_snapshots()
        if not snaps:
            messagebox.showinfo("Vorherige Version", "Es wurde noch kein Rücksprungpunkt gefunden. Rücksprungpunkte entstehen automatisch vor Updates.")
            return
        win = tk.Toplevel(self)
        win.resizable(True, True)
        win.title("Vorherige Version wiederherstellen")
        win.geometry("720x420")
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()
        tk.Label(win, text="Vorherige Version wiederherstellen", font=(theme.FONT, 18, "bold"), fg="#0b4a86", bg="#f5f7fb").pack(anchor="w", padx=22, pady=(18, 6))
        tk.Label(win, text="Wählen Sie einen Rücksprungpunkt. Vor dem Zurückspielen wird der aktuelle Stand nochmals gesichert.", bg="#f5f7fb", fg="#333").pack(anchor="w", padx=22, pady=(0, 12))
        lb = tk.Listbox(win, font=("Consolas", 10), height=12)
        lb.pack(fill="both", expand=True, padx=22, pady=(0, 12))
        for s in snaps:
            lb.insert("end", f"{s.name}   ({datetime_from_mtime(s)})")
        def do_restore():
            sel = lb.curselection()
            if not sel:
                messagebox.showinfo("Auswahl", "Bitte Rücksprungpunkt auswählen.")
                return
            snap = snaps[sel[0]]
            if not messagebox.askyesno("Rollback", f"Diesen Rücksprungpunkt wiederherstellen?\n\n{snap.name}\n\nDanach Programm neu starten."):
                return
            try:
                res = restore_rollback_snapshot(snap)
                messagebox.showinfo("Rollback fertig", f"Vorherige Version wurde wiederhergestellt.\n\nSicherheitskopie aktueller Stand:\n{res.get('safety')}\n\nBitte Programm schließen und neu starten.")
                win.destroy()
            except Exception as exc:
                messagebox.showerror("Rollback-Fehler", str(exc))
        bar = tk.Frame(win, bg="#f5f7fb")
        bar.pack(fill="x", padx=22, pady=(0, 18))
        tk.Button(bar, text="Abbrechen", command=win.destroy, padx=16, pady=8).pack(side="right", padx=(8,0))
        tk.Button(bar, text="Wiederherstellen", command=do_restore, bg="#0b4a86", fg="white", padx=18, pady=8).pack(side="right")

    def show_version(self):
        """V1.1 SP11: Versionsinfo als Seite mit Versionsliste links und
        Changelog rechts (statt nur Messagebox).
        """
        from .changelog import get_changelog
        self.clear_page()
        self._page_header(
            "Versionsinfo",
            f"Aktuelle Version: {APP_VERSION_DISPLAY} ({APP_VERSION})  ·  DB-Schema: {DB_SCHEMA_VERSION}  ·  Update-Ordner: {UPDATE_DIR}",
        )

        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure(0, weight=0)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        # Links: Versionsliste
        left = tk.LabelFrame(body, text=" Versionen ",
                             bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 10, "bold"))
        left.grid(row=0, column=0, sticky="ns", padx=(0, 8))
        left.rowconfigure(0, weight=1)
        left.columnconfigure(0, weight=1)
        version_tree = ttk.Treeview(left, columns=("name", "datum"),
                                    show="headings", selectmode="browse",
                                    height=22)
        version_tree.heading("name", text="Version")
        version_tree.heading("datum", text="Datum")
        version_tree.column("name", width=140, anchor="w")
        version_tree.column("datum", width=100, anchor="w")
        version_tree.grid(row=0, column=0, sticky="nsew", padx=4, pady=4)
        vsb = tk.Scrollbar(left, orient="vertical", command=version_tree.yview)
        vsb.grid(row=0, column=1, sticky="ns")
        version_tree.configure(yscrollcommand=vsb.set)

        # Rechts: Changelog
        right = tk.LabelFrame(body, text=" Aenderungen ",
                              bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 10, "bold"))
        right.grid(row=0, column=1, sticky="nsew")
        right.rowconfigure(0, weight=1)
        right.columnconfigure(0, weight=1)
        text_widget = tk.Text(right, wrap="word", font=(theme.FONT, 10),
                              bg="#fcfdff", fg="#222", relief="flat",
                              padx=12, pady=10)
        text_widget.grid(row=0, column=0, sticky="nsew", padx=4, pady=4)
        tsb = tk.Scrollbar(right, orient="vertical", command=text_widget.yview)
        tsb.grid(row=0, column=1, sticky="ns")
        text_widget.configure(yscrollcommand=tsb.set, state="disabled")

        changelog = get_changelog()
        id_to_entry: dict[str, tuple[str, str, list[str]]] = {}
        for entry in changelog:
            name, datum, _lines = entry
            iid = version_tree.insert("", "end", values=(name, datum))
            id_to_entry[iid] = entry

        def on_select(_e=None):
            sel = version_tree.selection()
            if not sel:
                return
            entry = id_to_entry.get(sel[0])
            if not entry:
                return
            name, datum, lines = entry
            text_widget.configure(state="normal")
            text_widget.delete("1.0", "end")
            text_widget.insert("end", f"{name}  ({datum})\n\n", "head")
            for line in lines:
                text_widget.insert("end", f"  • {line}\n\n")
            text_widget.tag_configure("head", font=(theme.FONT, 12, "bold"),
                                      foreground="#0b4a86")
            text_widget.configure(state="disabled")

        version_tree.bind("<<TreeviewSelect>>", on_select)
        # erste Version vorselektieren
        first = version_tree.get_children()
        if first:
            version_tree.selection_set(first[0])
            version_tree.focus(first[0])
            on_select()

        self.status.set(f"Versionsinfo: {len(changelog)} Versionen.")

    def clear_page(self):
        for widget in self.page.winfo_children():
            widget.destroy()
        self.page.columnconfigure(0, weight=1)
        self.page.rowconfigure(0, weight=1)
        self.page.rowconfigure(1, weight=1)

    def _page_header(self, title, subtitle="", icon=None):
        header = tk.Frame(self.page, bg=theme.CARD)
        header.grid(row=0, column=0, sticky="ew", padx=20, pady=(16, 8))
        if icon is not None:
            tk.Label(header, image=icon, bg=theme.CARD).pack(side="left", padx=(0, 12))
            self._page_header_icon = icon  # Referenz halten (sonst GC)
        textcol = tk.Frame(header, bg=theme.CARD)
        textcol.pack(side="left", anchor="w")
        tk.Label(textcol, text=title, font=(theme.FONT, 20, "bold"), fg=theme.INK, bg=theme.CARD).pack(anchor="w")
        if subtitle:
            tk.Label(textcol, text=subtitle, font=(theme.FONT, 11), fg=theme.MUTED, bg=theme.CARD).pack(anchor="w", pady=(2, 0))

    def _action_page(self, title, subtitle, button_text, command, color=theme.PRIMARY):
        self.clear_page()
        self._page_header(title, subtitle)
        body = tk.Frame(self.page, bg=theme.CARD)
        body.grid(row=1, column=0, sticky="nsew", padx=20, pady=(0, 18))
        tk.Button(
            body,
            text=button_text,
            command=command,
            bg=color,
            fg="white",
            activebackground=color,
            relief="flat",
            font=(theme.FONT, 12, "bold"),
            cursor="hand2",
            padx=20,
            pady=10
        ).pack(anchor="w", pady=8)


    def _sort_key_for_tree(self, value):
        """Robuste Sortierung für Treeview-Spalten: Zahlen, Datumswerte und Text."""
        text = str(value or "").strip()
        if not text:
            return (0, "")
        normalized = text.replace(".", "").replace(",", ".")
        try:
            return (1, float(normalized))
        except Exception:
            pass
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y %H:%M:%S", "%d.%m.%Y"):
            try:
                from datetime import datetime
                return (2, datetime.strptime(text[:19], fmt))
            except Exception:
                pass
        return (3, text.lower())

    def _make_tree_sortable(self, tree, columns, headings=None):
        """Macht eine ttk.Treeview per Klick auf die Spaltenüberschrift sortierbar."""
        headings = headings or {}
        sort_state = {}

        def sort_by(col):
            reverse = sort_state.get(col, False)
            items = []
            for iid in tree.get_children(""):
                try:
                    items.append((self._sort_key_for_tree(tree.set(iid, col)), iid))
                except Exception:
                    items.append(((3, ""), iid))
            items.sort(key=lambda item: item[0], reverse=reverse)
            for index, (_, iid) in enumerate(items):
                tree.move(iid, "", index)
            sort_state[col] = not reverse

            for c in columns:
                label = headings.get(c, c)
                if c == col:
                    label += "  ▼" if reverse else "  ▲"
                tree.heading(c, text=label, command=lambda cc=c: sort_by(cc))

        for col in columns:
            tree.heading(col, text=headings.get(col, col), command=lambda c=col: sort_by(c))

    def _apply_output_period_header(self, output_file, zeitraum_monate=None):
        """Benennt in der fertigen Auswertung die Absatzspalte passend zum Mapping-Zeitraum um.\n
        Intern bleibt die Mapping-Standarddatei kompatibel mit dem Exporter.\nSichtbar in der fertigen Auswertung steht danach z. B.
        "Abverkäufe 12 Monate" statt "Abverkäufe 6 Monate".\n"""
        try:
            zeitraum = int(zeitraum_monate or 0)
        except Exception:
            zeitraum = 0
        if zeitraum not in (6, 12):
            return

        target_text = f"Abverkäufe {zeitraum} Monate"
        try:
            path = Path(output_file)
            if not path.exists():
                return
            wb = load_workbook(path)
            changed = False
            for ws in wb.worksheets:
                max_scan_row = min(ws.max_row, 10)
                for row in ws.iter_rows(min_row=1, max_row=max_scan_row):
                    for cell in row:
                        value = str(cell.value or "").strip()
                        normalized = value.lower().replace("ae", "ä")
                        if value == "Abverkäufe 6 Monate" or normalized in {"abverkäufe 6 monate", "abverkaeufe 6 monate"}:
                            cell.value = target_text
                            changed = True
                        elif "verkaufsmenge" in normalized and "6 monate" in normalized:
                            cell.value = f"Verkaufsmenge der letzten {zeitraum} Monate"
                            changed = True
            if changed:
                wb.save(path)
        except Exception:
            # Die Auswertung darf wegen einer reinen Überschriftenkorrektur nicht fehlschlagen.
            pass


    def _auswertungsvorlagen_dir(self):
        path = DATA_DIR / "vorlagen" / "auswertungsvorlagen"
        path.mkdir(parents=True, exist_ok=True)
        return path

    def _template_meta_key(self, slot, suffix):
        return f"auswertungsvorlage_slot_{int(slot)}_{suffix}"

    def _list_auswertungsvorlagen(self):
        """Liest bis zu drei hinterlegte Auswertungsvorlagen aus meta."""
        templates = []
        with sqlite3.connect(DB_PATH) as con:
            con.execute("CREATE TABLE IF NOT EXISTS meta(key TEXT PRIMARY KEY, value TEXT NOT NULL)")
            for slot in (1, 2, 3):
                path = ""
                name = f"Vorlage {slot}"
                source = ""
                updated = ""
                row = con.execute("SELECT value FROM meta WHERE key=?", (self._template_meta_key(slot, "path"),)).fetchone()
                if row:
                    path = row[0]
                row = con.execute("SELECT value FROM meta WHERE key=?", (self._template_meta_key(slot, "name"),)).fetchone()
                if row and row[0]:
                    name = row[0]
                row = con.execute("SELECT value FROM meta WHERE key=?", (self._template_meta_key(slot, "source"),)).fetchone()
                if row:
                    source = row[0]
                row = con.execute("SELECT value FROM meta WHERE key=?", (self._template_meta_key(slot, "updated"),)).fetchone()
                if row:
                    updated = row[0]
                if path and Path(path).exists():
                    templates.append({"slot": slot, "name": name, "path": path, "source": source, "updated": updated})
        return templates

    def _get_selected_auswertungsvorlage_slot(self):
        value = self._get_meta_value("auswertungsvorlage_selected_slot", "")
        try:
            slot = int(value)
            return slot if slot in (1, 2, 3) else None
        except Exception:
            return None

    def _set_selected_auswertungsvorlage_slot(self, slot):
        if slot in (1, 2, 3):
            self._set_meta_value("auswertungsvorlage_selected_slot", str(slot))
        else:
            self._set_meta_value("auswertungsvorlage_selected_slot", "")

    def _template_combo_values(self):
        values = ["Keine Vorlage / Standard"]
        for item in self._list_auswertungsvorlagen():
            values.append(f"{item['slot']}: {item['name']}")
        return values

    def _template_label_for_selected_slot(self):
        selected = self._get_selected_auswertungsvorlage_slot()
        if selected:
            for item in self._list_auswertungsvorlagen():
                if item["slot"] == selected:
                    return f"{item['slot']}: {item['name']}"
        return "Keine Vorlage / Standard"

    def _slot_from_template_label(self, label):
        text = str(label or "").strip()
        if len(text) >= 2 and text[0].isdigit() and text[1] == ":":
            slot = int(text[0])
            return slot if slot in (1, 2, 3) else None
        return None

    def _get_auswertungsvorlage_path(self, slot=None):
        slot = slot or self._get_selected_auswertungsvorlage_slot()
        if slot not in (1, 2, 3):
            return None
        with sqlite3.connect(DB_PATH) as con:
            row = con.execute("SELECT value FROM meta WHERE key=?", (self._template_meta_key(slot, "path"),)).fetchone()
            if not row:
                return None
        path = Path(row[0])
        return path if path.exists() else None

    def _preview_auswertungsvorlage_text(self, path):
        try:
            wb = load_workbook(path, read_only=False, data_only=True)
            ws = wb.active
            headers = []
            for c in range(1, min(ws.max_column, 22) + 1):
                value = ws.cell(1, c).value
                if value not in (None, ""):
                    headers.append(str(value))
            header_text = "; ".join(headers[:18]) or "keine Kopfzeile erkannt"
            return (
                f"Datei: {Path(path).name}\n"
                f"Blatt: {ws.title}\n"
                f"Spalten: {ws.max_column} | Zeilen: {ws.max_row}\n"
                f"Kopfzeile: {header_text}"
            )
        except Exception as exc:
            return f"Vorschau konnte nicht gelesen werden:\n{exc}"

    def _copy_cell_format(self, src, dst, copy_value=False):
        if copy_value:
            dst.value = src.value
        if src.has_style:
            dst.font = copy_cell_style(src.font)
            dst.fill = copy_cell_style(src.fill)
            dst.border = copy_cell_style(src.border)
            dst.alignment = copy_cell_style(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy_cell_style(src.protection)

    def _apply_auswertungsvorlage(self, output_file, slot=None, zeitraum_monate=None):
        """Übernimmt aus der gewählten Vorlage nur Layout, Spaltenbreiten und Kopfzeile; keine Datenzeilen."""
        template_path = self._get_auswertungsvorlage_path(slot)
        if not template_path:
            return
        try:
            out_path = Path(output_file)
            if not out_path.exists():
                return
            tpl_wb = load_workbook(template_path)
            tpl_ws = tpl_wb.active
            out_wb = load_workbook(out_path)
            out_ws = out_wb.active

            max_col = max(tpl_ws.max_column, out_ws.max_column)

            # Spaltenüberschriften und deren Formatierung übernehmen.
            for c in range(1, max_col + 1):
                src = tpl_ws.cell(1, c)
                dst = out_ws.cell(1, c)
                if src.value not in (None, ""):
                    dst.value = src.value
                self._copy_cell_format(src, dst, copy_value=False)
                letter = get_column_letter(c)
                if tpl_ws.column_dimensions[letter].width:
                    out_ws.column_dimensions[letter].width = tpl_ws.column_dimensions[letter].width

            # Formatierung der ersten Datenzeile als Muster auf vorhandene Datenzeilen anwenden.
            sample_row = 2 if tpl_ws.max_row >= 2 else 1
            for r in range(2, out_ws.max_row + 1):
                for c in range(1, max_col + 1):
                    self._copy_cell_format(tpl_ws.cell(sample_row, c), out_ws.cell(r, c), copy_value=False)

            # Zeilenhöhe, Filter, Freeze und Druck-/Ansichtseinstellungen übernehmen, soweit vorhanden.
            for r in range(1, min(tpl_ws.max_row, 5) + 1):
                if tpl_ws.row_dimensions[r].height:
                    out_ws.row_dimensions[r].height = tpl_ws.row_dimensions[r].height
            out_ws.freeze_panes = tpl_ws.freeze_panes or out_ws.freeze_panes
            if tpl_ws.auto_filter and tpl_ws.auto_filter.ref:
                out_ws.auto_filter.ref = out_ws.dimensions
            try:
                out_ws.sheet_view.showGridLines = tpl_ws.sheet_view.showGridLines
            except Exception:
                pass

            out_wb.save(out_path)
        except Exception as exc:
            # Layout darf die fachliche Auswertung nicht blockieren.
            # status.set thread-sicher (Methode laeuft jetzt im Hintergrund-Thread).
            try:
                msg = f"Auswertung erstellt, aber Vorlage konnte nicht angewendet werden: {exc}"
                self.after(0, lambda m=msg: self.status.set(m))
            except Exception:
                pass

    def _run_neue_auswertung_export(self, file, analyse_name, kundentyp=None, kundennummer="", kundenname="", zeitraum_monate=None, vorlage_slot=None, on_done=None):
        """V1.1 SP18: Background-Variante. Auswertung laeuft im Hintergrund,
        UI bleibt klickbar. Post-Processing (Datei verschieben, Vorlage,
        messagebox) im UI-Thread nach Abschluss via _run_background-Callback.

        on_done(copied_out) wird nach erfolgreichem Abschluss + Post-Processing
        im UI-Thread aufgerufen. Bei UnknownInputFormatError kommt der Format-
        Assistent-Dialog (interaktiv im UI-Thread).
        """
        # SP19: Doppelklick-Schutz. Verhindert dass der "Auswertung starten"-Button
        # waehrend einer laufenden Auswertung wieder feuert (Log zeigte 90 Aufrufe
        # in einer Sekunde -> stille Mehrfachstarts mit DB-Locks/Konflikten).
        if getattr(self, "_auswertung_running", False):
            try:
                messagebox.showinfo(
                    "Auswertung läuft",
                    "Es läuft bereits eine Auswertung. Bitte warten, bis sie fertig ist."
                )
            except Exception:
                pass
            return
        self._auswertung_running = True
        self._log_action("neue_auswertung", "Auswertungserstellung gestartet", f"Datei: {file} | Analyse: {analyse_name}")

        def post_export(out_raw):
            self._auswertung_running = False
            out = Path(out_raw)
            try:
                if kundentyp is not None:
                    self._mark_latest_auswertung_customer(kundentyp, kundennummer, kundenname or analyse_name)
            except Exception:
                pass
            # SP7: Auswertungen werden je nach Kundentyp in einen Unterordner sortiert.
            _kt = "ZW" if kundentyp == "ZF" else kundentyp
            sub = _kt if _kt in ("PK", "ZW") else None
            base_dir = SAVED_ANALYSES_DIR / sub if sub else SAVED_ANALYSES_DIR
            analyse_dir = base_dir / f"{analyse_name}_{out.stem[-15:] if len(out.stem) >= 15 else ''}".strip("_")
            analyse_dir.mkdir(parents=True, exist_ok=True)
            copied_out = analyse_dir / out.name
            shutil.copy2(out, copied_out)
            try:
                shutil.copy2(file, analyse_dir / f"Rohdaten_{Path(file).name}")
            except Exception:
                pass
            self._roadmap_mark_neue_auswertung_form_erledigt()

            # SP31: Die gesamte SCHWERE Nachbearbeitung (Vorlage anwenden =
            # zehntausende Zell-Format-Kopien, Zeitraum-Kopf, Kurzbericht) lief
            # bisher im UI-Thread -> Oberflaeche fror nach jeder Auswertung ein.
            # Jetzt komplett im Hintergrund-Thread; nur die Abschluss-Meldung +
            # Ordner-Oeffnen laufen via after() wieder im UI-Thread.
            self.status.set("Auswertung wird nachbearbeitet ...")

            def _heavy_and_finish():
                try:
                    self._apply_auswertungsvorlage(copied_out, vorlage_slot, zeitraum_monate)
                except Exception:
                    pass
                try:
                    self._apply_output_period_header(copied_out, zeitraum_monate)
                except Exception:
                    pass

                def _finished_ui():
                    self._log_action("neue_auswertung", "Auswertung gespeichert", str(copied_out))
                    self.status.set(f"Auswertung gespeichert: {copied_out}")
                    messagebox.showinfo(
                        "Auswertung fertig",
                        f"Auswertung wurde erstellt und gespeichert:\n{copied_out}\n\n"
                        "Der Kurzbericht (Excel + PDF) wird noch erstellt und liegt gleich im selben Ordner.\n\n"
                        "Der Ordner wird jetzt geöffnet."
                    )
                    _open_folder(analyse_dir)
                    if on_done:
                        try:
                            on_done(copied_out)
                        except Exception:
                            pass
                try:
                    self.after(0, _finished_ui)
                except Exception:
                    pass

                # Kurzbericht zuletzt - er ist optional und darf nichts blockieren.
                try:
                    from .kurzbericht import create_kurzbericht
                    kb = create_kurzbericht(
                        copied_out, analyse_name, analyse_dir,
                        zeitraum_monate=zeitraum_monate,
                        apotheke=kundenname or analyse_name,
                    )
                    erzeugt = [fmt for fmt in ("excel", "pdf") if kb.get(fmt)]
                    if erzeugt:
                        self._log_action("neue_auswertung", "Kurzbericht erstellt",
                                         "; ".join(str(kb[f]) for f in erzeugt))
                except Exception as kb_exc:
                    self._log_action("neue_auswertung", "Kurzbericht fehlgeschlagen", str(kb_exc))

            threading.Thread(target=_heavy_and_finish, daemon=True).start()

        def on_export_error(exc):
            self._auswertung_running = False
            # UnknownInputFormatError -> interaktiver Format-Assistent.
            if isinstance(exc, UnknownInputFormatError):
                self.status.set(str(exc))
                if messagebox.askyesno("Format nicht erkannt",
                                        f"{exc}\n\nSoll der Rohdaten-Formatassistent geöffnet werden?"):
                    mapping = self._open_rohdaten_format_assistent(file, str(exc))
                    if mapping:
                        try:
                            self._run_auswertung_after_mapping(
                                file, mapping, analyse_name, kundentyp,
                                kundennummer, kundenname or analyse_name,
                                vorlage_slot=vorlage_slot,
                            )
                        except Exception as mapped_exc:
                            self.status.set(f"Auswertung nach Mapping fehlgeschlagen: {mapped_exc}")
                            messagebox.showerror("Auswertungsfehler nach Mapping", str(mapped_exc))
                else:
                    messagebox.showwarning("Format nicht erkannt", str(exc))
                return
            messagebox.showerror("Auswertungsfehler", str(exc))

        def on_duplicate_prompt(info):
            # Wird aus dem Worker-Thread aufgerufen. Die Abfrage muss aber im
            # UI-Thread laufen, also via after(0, ...) marshallen und den Worker
            # bis zur Antwort blockieren.
            holder = {}
            answered = threading.Event()

            def ask():
                beispiele = ", ".join(info["pzns"][:10])
                if info["count"] > 10:
                    beispiele += " …"
                msg = (
                    f"{info['count']} PZN kommen in der Rohdatei mehrfach vor"
                    + (" – teils mit negativen Mengen (z. B. Retouren)" if info["has_negative"] else "")
                    + ".\n\n"
                    f"Betroffen: {beispiele}\n\n"
                    "Sollen die Mengen je PZN summiert werden (eine Zeile pro PZN)?\n\n"
                    "Ja = zusammenfassen und summieren\n"
                    "Nein = jede Zeile einzeln lassen (wie bisher)"
                )
                holder["v"] = messagebox.askyesno("Mehrfache PZN gefunden", msg)
                answered.set()

            self.after(0, ask)
            answered.wait()
            return holder.get("v", False)

        self._run_background(
            lambda: create_vorlage_export(file, analyse_name, on_duplicate_prompt=on_duplicate_prompt),
            title="Auswertung wird erstellt",
            subtitle=f"NMGone wertet {Path(file).name} aus ...",
            progress=False,
            on_done=post_export,
            on_error=on_export_error,
        )

    def _create_standard_rohdaten_from_mapping(self, file_path, mapping):
        """Erstellt aus einer unbekannten Rohdatei eine Standard-Rohdatei, die der Exporter lesen kann."""
        path = Path(file_path)
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        def to_int(value, default=0):
            try:
                return int(str(value).strip())
            except Exception:
                return default

        pzn_col = to_int(mapping.get("pzn_spalte"))
        menge_col = to_int(mapping.get("absatz_spalte"))
        artikel_col = to_int(mapping.get("artikel_spalte"))
        hersteller_col = to_int(mapping.get("hersteller_spalte"))
        ek_col = to_int(mapping.get("ek_spalte"))
        header_zeile = to_int(mapping.get("header_zeile"), 1)
        zeitraum = to_int(mapping.get("zeitraum_monate"), 6)

        if not pzn_col or not menge_col:
            raise ValueError("Mapping unvollständig: PZN-Spalte und Menge-Spalte sind Pflicht.")

        src_wb = load_workbook(path, read_only=True, data_only=True)
        src_ws = src_wb.active
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = "Rohdaten_Mapping"
        out_ws.append([
            "PZN",
            "Artikelname",
            "DF",
            "Pck",
            "Herst",
            "EK",
            "Abverkäufe 6 Monate",
            "Zeitraum Monate",
        ])

        def val(row, col):
            if not col:
                return None
            return row[col - 1] if len(row) >= col else None

        written = 0
        for row in src_ws.iter_rows(min_row=header_zeile + 1, values_only=True):
            pzn = val(row, pzn_col)
            menge = val(row, menge_col)
            if pzn in (None, "") and menge in (None, ""):
                continue
            if pzn in (None, ""):
                continue
            out_ws.append([
                pzn,
                val(row, artikel_col) or "",
                "",
                "",
                val(row, hersteller_col) or "",
                val(row, ek_col) or "",
                menge,
                zeitraum,
            ])
            written += 1

        if written == 0:
            raise ValueError("Mit dem gewählten Mapping wurden keine verwertbaren Zeilen gefunden.")

        safe = re.sub(r"[^A-Za-z0-9_-]+", "_", path.stem)[:45]
        out_path = OUTPUT_DIR / f"Rohdaten_Mapping_{safe}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        out_wb.save(out_path)
        return out_path

    def _run_auswertung_after_mapping(self, original_file, mapping, analyse_name, kundentyp=None, kundennummer="", kundenname="", vorlage_slot=None):
        """Speichert Mapping, baut Standarddatei und startet die Auswertung direkt danach."""
        mapped_file = self._create_standard_rohdaten_from_mapping(original_file, mapping)
        self.status.set(f"Mapping angewendet. Starte Auswertung mit Standarddatei: {mapped_file.name}")
        return self._run_neue_auswertung_export(str(mapped_file), analyse_name, kundentyp, kundennummer, kundenname, mapping.get("zeitraum_monate"))

    def _ensure_rohdaten_mapping_table(self):
        with sqlite3.connect(DB_PATH) as con:
            con.execute("""\nCREATE TABLE IF NOT EXISTS tbl_rohdaten_mapping (
                    dateiname TEXT PRIMARY KEY,\npzn_spalte TEXT,
                    hersteller_spalte TEXT,\nek_spalte TEXT,
                    absatz_spalte TEXT,\nheader_zeile INTEGER DEFAULT 1,
                    format_typ TEXT DEFAULT 'standard',\nzeitraum_monate INTEGER DEFAULT 6,
                    artikel_spalte TEXT,\nbearbeiter TEXT,
                    letzte_aktualisierung TEXT DEFAULT CURRENT_TIMESTAMP\n)
            """)
            cols = {row[1] for row in con.execute("PRAGMA table_info(tbl_rohdaten_mapping)").fetchall()}
            if "zeitraum_monate" not in cols:
                con.execute("ALTER TABLE tbl_rohdaten_mapping ADD COLUMN zeitraum_monate INTEGER DEFAULT 6")
            if "artikel_spalte" not in cols:
                con.execute("ALTER TABLE tbl_rohdaten_mapping ADD COLUMN artikel_spalte TEXT")
            if "bearbeiter" not in cols:
                con.execute("ALTER TABLE tbl_rohdaten_mapping ADD COLUMN bearbeiter TEXT")
            con.commit()

    def _open_rohdaten_format_assistent(self, file_path, fehlertext=""):
        """Fragt Spalten für unbekannte Rohdaten ab und speichert das Mapping für spätere Importe."""
        path = Path(file_path)
        self._ensure_rohdaten_mapping_table()
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = []
            for col in range(1, ws.max_column + 1):
                value = ws.cell(1, col).value
                label = str(value).strip() if value not in (None, "") else f"Spalte {col}"
                headers.append(f"{col}: {label}")
        except Exception as exc:
            messagebox.showerror("Formatassistent", f"Datei konnte nicht gelesen werden:\n{exc}")
            return False

        if not headers:
            messagebox.showinfo("Formatassistent", "Keine Spalten in der Datei erkannt.")
            return False

        win = tk.Toplevel(self)
        win.resizable(True, True)
        win.title("Rohdaten-Formatassistent")
        win.geometry("720x520")
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()

        tk.Label(win, text="Rohdaten-Formatassistent", font=(theme.FONT, 18, "bold"), fg="#0b4a86", bg="#f5f7fb").pack(anchor="w", padx=22, pady=(18, 4))
        hinweis = f"Datei: {path.name}\n"
        if fehlertext:
            hinweis += f"\nAutomatische Erkennung fehlgeschlagen:\n{fehlertext}\n"
        hinweis += "\nBitte die Spalten zuordnen. Pflichtfelder sind PZN, Menge/Absatz und Zeitraum."
        tk.Label(win, text=hinweis, justify="left", bg="#f5f7fb", fg="#333", wraplength=660).pack(anchor="w", padx=22, pady=(0, 12))

        form = tk.Frame(win, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        form.pack(fill="both", expand=True, padx=22, pady=(0, 12))
        form.columnconfigure(1, weight=1)

        pzn_var = tk.StringVar()
        menge_var = tk.StringVar()
        zeitraum_var = tk.StringVar(value="6")
        artikel_var = tk.StringVar(value="")
        hersteller_var = tk.StringVar(value="")
        ek_var = tk.StringVar(value="")

        combo_values = [""] + headers
        rows = [
            ("PZN-Spalte *", pzn_var, headers),
            ("Menge / Absatz-Spalte *", menge_var, headers),
            ("Artikelname-Spalte optional", artikel_var, combo_values),
            ("Hersteller-Spalte optional", hersteller_var, combo_values),
            ("EK-Spalte optional", ek_var, combo_values),
        ]
        for r, (label, var, values) in enumerate(rows):
            tk.Label(form, text=label, bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 11, "bold")).grid(row=r, column=0, sticky="w", padx=14, pady=9)
            cb = ttk.Combobox(form, textvariable=var, values=values, state="readonly", width=58)
            cb.grid(row=r, column=1, sticky="ew", padx=14, pady=9)

        tk.Label(form, text="Zeitraum *", bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 11, "bold")).grid(row=5, column=0, sticky="w", padx=14, pady=9)
        period = tk.Frame(form, bg="#ffffff")
        period.grid(row=5, column=1, sticky="w", padx=14, pady=9)
        tk.Radiobutton(period, text="6 Monate", variable=zeitraum_var, value="6", bg="#ffffff").pack(side="left", padx=(0, 18))
        tk.Radiobutton(period, text="12 Monate", variable=zeitraum_var, value="12", bg="#ffffff").pack(side="left")

        result = {"mapping": None}

        def clean_col(value):
            if not value:
                return ""
            return value.split(":", 1)[0].strip()

        def save_mapping():
            pzn_col = clean_col(pzn_var.get())
            menge_col = clean_col(menge_var.get())
            if not pzn_col or not menge_col:
                messagebox.showinfo("Formatassistent", "Bitte mindestens PZN-Spalte und Menge/Absatz-Spalte auswählen.")
                return
            with sqlite3.connect(DB_PATH) as con:
                con.execute("""\nINSERT INTO tbl_rohdaten_mapping (
                        dateiname, pzn_spalte, absatz_spalte, zeitraum_monate,\nartikel_spalte, hersteller_spalte, ek_spalte,
                        header_zeile, format_typ, bearbeiter, letzte_aktualisierung\n) VALUES (?, ?, ?, ?, ?, ?, ?, 1, 'manuell', ?, CURRENT_TIMESTAMP)
                    ON CONFLICT(dateiname) DO UPDATE SET\npzn_spalte=excluded.pzn_spalte,
                        absatz_spalte=excluded.absatz_spalte,\nzeitraum_monate=excluded.zeitraum_monate,
                        artikel_spalte=excluded.artikel_spalte,\nhersteller_spalte=excluded.hersteller_spalte,
                        ek_spalte=excluded.ek_spalte,\nformat_typ='manuell',
                        bearbeiter=excluded.bearbeiter,\nletzte_aktualisierung=CURRENT_TIMESTAMP
                """, (
                    path.name, pzn_col, menge_col, int(zeitraum_var.get()),
                    clean_col(artikel_var.get()), clean_col(hersteller_var.get()), clean_col(ek_var.get()),
                    getpass.getuser()
                ))
                con.commit()
            mapping = {
                "dateiname": path.name,
                "pzn_spalte": pzn_col,
                "absatz_spalte": menge_col,
                "zeitraum_monate": int(zeitraum_var.get()),
                "artikel_spalte": clean_col(artikel_var.get()),
                "hersteller_spalte": clean_col(hersteller_var.get()),
                "ek_spalte": clean_col(ek_var.get()),
                "header_zeile": 1,
            }
            result["mapping"] = mapping
            self.status.set(f"Rohdaten-Mapping gespeichert: {path.name}")
            try:
                add_roadmap_item(
                    bereich="Import",
                    titel="Rohdaten-Formatassistent / Mapping gespeichert",
                    beschreibung=f"Mapping für {path.name}: PZN-Spalte {pzn_col}, Menge-Spalte {menge_col}, Zeitraum {zeitraum_var.get()} Monate.",
                    status="Erledigt",
                    prioritaet="Normal",
                )
            except Exception:
                pass
            messagebox.showinfo("Formatassistent", "Mapping gespeichert. Die Auswertung wird jetzt automatisch gestartet.")
            win.destroy()

        buttons = tk.Frame(win, bg="#f5f7fb")
        buttons.pack(fill="x", padx=22, pady=(0, 18))
        tk.Button(buttons, text="Abbrechen", command=win.destroy, padx=16, pady=8).pack(side="right", padx=(8, 0))
        tk.Button(buttons, text="Mapping speichern und Auswertung starten", command=save_mapping, bg="#0b4a86", fg="white", relief="flat", font=(theme.FONT, 11, "bold"), padx=20, pady=9).pack(side="right")
        self.wait_window(win)
        return result["mapping"]

    def _dashboard_all_tiles(self):
        """Alle verfügbaren Kacheln für das Dashboard (Schnellzugriff + Info-Widgets)."""
        return [
            # --- Schnellzugriff ---
            ("neue_auswertung",  "📊", "Bedarfsanalyse",      "PK-/ZW-Auswertung starten.",               self.show_neue_auswertung_page,              "#0b4a86"),
            ("gespeicherte",     "📁", "Ges. Analysen",        "Vorhandene Analysen öffnen.",               self.open_saved_analyses,                    "#3867b7"),
            ("schulbank",        "🎓", "Schulbank",            "Lernvorschläge bearbeiten.",                lambda: self.show_schulbank_page("Schulbank"),"#11823b"),
            ("kunden",           "👥", "Kunden",           "Kundenstamm und -history.",                 self.show_kunden_center,                     "#0b4a86"),
            ("kasse",            "🛒", "Kasse",                "Verkauf an Apotheken + Wareneingang.",      self.open_kasse_app,                         "#8b5a00"),
            ("auswertungen",     "📑", "Auswertungen",         "Verkäufe, Kunden, Artikel frei auswerten.", self.open_auswertungen_app,                  "#0b6e6e"),
            ("todo",             "✅", "ToDo",                 "Aufgaben und offene Punkte.",               self.show_todo_center,                        "#11823b"),
            ("mitarbeiter",      "👥", "Mitarbeiter",          "Zuständigkeiten und Datenpfade.",           self.show_mitarbeiter_center,                "#6b4fb3"),
            ("produktanalyse",   "📈", "Produktanalyse",       "Produktchancen erstellen.",                 self.market_opportunities,                   "#11823b"),
            # SP7: Marktanalyse-Tile entfernt.
            ("abweichung",       "🔍", "Abweichungsanalyse",   "Manuelle vs. Programm-Auswertung.",         self.deviation_analysis,                     "#8b5a00"),
            ("datenupdates",     "🗄", "Daten aktualisieren",  "Stammdaten und Importe pflegen.",           self.show_daten_aktualisieren_page,          "#555"),
            ("roadmap",          "📌", "Roadmap",              "Offene und erledigte Punkte.",              self.show_roadmap_page,                      "#6b4fb3"),
            ("vergleichssuche",  "🔍", "Vergleichs-Suche",     "PZN/Artikelname schnell finden.",           self.open_vergleichssuche_window,            "#0b6e6e"),
            ("globale_suche",    "🔍", "Globale Suche",        "Kunde, Analyse oder Artikel finden.",       self.open_globale_suche_window,              "#0b4a86"),
            ("nmg_rabatte",      "💰", "NMG-Rabatte",          "Aktuelle Rabatte, Statistik, Diff, Verlauf.", self.show_nmg_rabatte_uebersicht,         "#6b4fb3"),
            ("hilfe",            "❓", "Hilfe",                "Bebildertes Handbuch: was wie funktioniert.", self.open_hilfe_app,                      "#208acd"),
        ]

    def _dashboard_all_info_widgets(self):
        """Alle wählbaren Info-Widget-IDs mit Beschreibung."""
        return [
            ("info_ampel",    "🗄 Daten-Aktualität",     "Zeigt APU/HAP, NMG, PK Rabatte, Artikelstamm mit Ampel-Farben."),
            ("info_todos",    "✅ Offene ToDos",          "Zeigt die nächsten offenen Aufgaben mit Priorität und Fälligkeitsdatum."),
            ("info_analysen", "📊 Letzte Auswertungen",  "Zeigt die 5 neuesten gespeicherten Analysen."),

        ]

    def _dashboard_get_active_tiles(self):
        # V1.1 SP12 Bug-Fix: Sentinel '_NONE_' unterscheidet "nie gesetzt"
        # (Default zurueckgeben) von "explizit leer abgewaehlt" (leeres Set).
        # '_EMPTY_' = explizit leer. Vorher: leerer String fiel auf Default
        # zurueck -> 'Alles abwaehlen' hat nie funktioniert.
        saved = self._get_meta_value("dashboard_aktive_kacheln", "_NONE_")
        if saved == "_NONE_":
            return {t[0] for t in self._dashboard_all_tiles()[:7]}
        if not saved or saved == "_EMPTY_":
            return set()
        return set(saved.split(","))

    def _dashboard_set_active_tiles(self, active_set):
        if not active_set:
            self._set_meta_value("dashboard_aktive_kacheln", "_EMPTY_")
        else:
            self._set_meta_value("dashboard_aktive_kacheln", ",".join(sorted(active_set)))

    def _dashboard_get_active_info(self):
        saved = self._get_meta_value("dashboard_aktive_info", "_NONE_")
        if saved == "_NONE_":
            return {"info_ampel", "info_todos", "info_analysen"}
        if not saved or saved == "_EMPTY_":
            return set()
        return set(saved.split(","))

    def _dashboard_set_active_info(self, active_set):
        if not active_set:
            self._set_meta_value("dashboard_aktive_info", "_EMPTY_")
        else:
            self._set_meta_value("dashboard_aktive_info", ",".join(sorted(active_set)))

    def _dashboard_info_panel(self, parent, active_info):
        """Kompakter Info-Streifen oben auf der Startseite."""
        if not active_info:
            return

        info_row = tk.Frame(parent, bg="#f5f7fb")
        info_row.pack(fill="x", padx=14, pady=(6, 4))
        col_count = len(active_info)
        for i in range(col_count):
            info_row.columnconfigure(i, weight=1)
        col_idx = 0

        if "info_ampel" in active_info:
            amp_box = tk.Frame(info_row, bg="#f0f4fa", highlightbackground="#c5d3e8", highlightthickness=1)
            amp_box.grid(row=0, column=col_idx, sticky="nsew", padx=(0, 4), ipady=2)
            tk.Label(amp_box, text="🗄 Daten-Aktualität", font=(theme.FONT, 9, "bold"), fg="#0b4a86", bg="#f0f4fa").pack(anchor="w", padx=8, pady=(5,2))
            items = self._get_data_update_status_items()
            if items:
                for name, val, count in items:
                    amp = self._status_ampel_combined(val, count)
                    color = self._status_ampel_combined_color(val, count)
                    date_str = self._format_status_datetime(val)
                    count_str = self._format_count(count)
                    tk.Label(amp_box, text=f"{amp} {name} ({count_str}): {date_str}", font=(theme.FONT, 8), fg=color, bg="#f0f4fa").pack(anchor="w", padx=8)
            else:
                tk.Label(amp_box, text="–", font=(theme.FONT, 8), fg="#888", bg="#f0f4fa").pack(anchor="w", padx=8)
            tk.Label(amp_box, text="🟢<30T 🟡30-60T 🔴>60T ⚪leer/kein Datum", font=(theme.FONT, 7), fg="#888", bg="#f0f4fa").pack(anchor="w", padx=8, pady=(0,4))
            col_idx += 1

        if "info_todos" in active_info:
            todo_box = tk.Frame(info_row, bg="#f0faf4", highlightbackground="#b2d8c0", highlightthickness=1)
            todo_box.grid(row=0, column=col_idx, sticky="nsew", padx=(4, 4), ipady=2)
            tk.Label(todo_box, text="✅ Offene ToDos", font=(theme.FONT, 9, "bold"), fg="#11823b", bg="#f0faf4").pack(anchor="w", padx=8, pady=(5,2))
            try:
                with sqlite3.connect(DB_PATH) as con:
                    con.row_factory = sqlite3.Row
                    todos = con.execute(
                        "SELECT titel, prioritaet FROM tbl_todo_center WHERE lower(status) NOT IN ('erledigt','done','abgeschlossen') ORDER BY faellig_am ASC LIMIT 4"
                    ).fetchall()
                    gesamt = con.execute("SELECT COUNT(*) FROM tbl_todo_center WHERE lower(status) NOT IN ('erledigt','done','abgeschlossen')").fetchone()[0]
                if todos:
                    for t in todos:
                        prio = str(t["prioritaet"] or "")
                        icon = "🔴" if "hoch" in prio.lower() else "🟡"
                        tk.Label(todo_box, text=f"{icon} {(t['titel'] or '')[:38]}", font=(theme.FONT, 8), fg="#145c2e", bg="#f0faf4").pack(anchor="w", padx=8)
                    if gesamt > 4:
                        tk.Label(todo_box, text=f"… +{gesamt-4} weitere", font=(theme.FONT, 7), fg="#666", bg="#f0faf4").pack(anchor="w", padx=8)
                else:
                    tk.Label(todo_box, text="Keine offenen ToDos ✔", font=(theme.FONT, 8), fg="#11823b", bg="#f0faf4").pack(anchor="w", padx=8)
            except Exception:
                tk.Label(todo_box, text="–", font=(theme.FONT, 8), fg="#888", bg="#f0faf4").pack(anchor="w", padx=8)
            tk.Button(todo_box, text="ToDo →", command=self.show_todo_center, bg="#11823b", fg="white",
                      relief="flat", font=(theme.FONT, 8), padx=6, pady=2).pack(anchor="w", padx=8, pady=(2,4))
            col_idx += 1

        if "info_analysen" in active_info:
            ana_box = tk.Frame(info_row, bg="#fafafa", highlightbackground="#d8e2ee", highlightthickness=1)
            ana_box.grid(row=0, column=col_idx, sticky="nsew", padx=(4, 0), ipady=2)
            tk.Label(ana_box, text="📊 Letzte Auswertungen", font=(theme.FONT, 9, "bold"), fg="#0b4a86", bg="#fafafa").pack(anchor="w", padx=8, pady=(5,2))
            try:
                rows = self._get_saved_analysis_rows(limit=4)
                if rows:
                    for r in rows:
                        dq = _dq_label(r["datenquelle"])
                        datum_str = str(r["datum"] or "")[:10]
                        name = str(r["apotheke"] or "–")[:28]
                        tk.Label(ana_box, text=f"[{dq}] {datum_str} – {name}", font=(theme.FONT, 8), fg="#333", bg="#fafafa").pack(anchor="w", padx=8)
                else:
                    tk.Label(ana_box, text="Noch keine Auswertungen.", font=(theme.FONT, 8), fg="#888", bg="#fafafa").pack(anchor="w", padx=8)
            except Exception:
                tk.Label(ana_box, text="–", font=(theme.FONT, 8), fg="#888", bg="#fafafa").pack(anchor="w", padx=8)
            tk.Button(ana_box, text="Alle →", command=self.open_saved_analyses, bg="#3867b7", fg="white",
                      relief="flat", font=(theme.FONT, 8), padx=6, pady=2).pack(anchor="w", padx=8, pady=(2,4))



    def _dashboard_open_kachel_editor(self):
        """Dialog: Mitarbeiter wählt welche Kacheln und Info-Widgets angezeigt werden."""
        active_tiles = self._dashboard_get_active_tiles()
        active_info = self._dashboard_get_active_info()
        all_tiles = self._dashboard_all_tiles()
        all_info = self._dashboard_all_info_widgets()

        win = tk.Toplevel(self)
        win.resizable(True, True)
        win.title("Dashboard anpassen")
        # SP31: vorher 860x680 + win.state("zoomed") -> der Editor blockierte den
        # ganzen Bildschirm. Jetzt um ~20% kleiner und ohne Zoom-State.
        # V1.1 SP12: Hoehe wieder vergroessert (720x780), sonst war der
        # Speicher-Button unten abgeschnitten und musste manuell aufgezogen
        # werden. minsize hochgesetzt damit man den Button immer sieht.
        win.geometry("720x780")
        win.minsize(640, 600)
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()

        tk.Label(win, text="⚙️  Dashboard anpassen", font=(theme.FONT, 16, "bold"), fg="#0b4a86", bg="#f5f7fb").pack(anchor="w", padx=22, pady=(18, 2))
        tk.Label(win, text="Wähle welche Info-Bereiche und Schnellzugriff-Kacheln auf der Startseite erscheinen.\nDie Auswahl wird für dich gespeichert.", font=(theme.FONT, 10), fg="#444", bg="#f5f7fb", justify="left").pack(anchor="w", padx=22, pady=(0, 10))

        # --- Info-Widgets ---
        tk.Label(win, text="Info-Bereiche (obere Zeile):", font=(theme.FONT, 11, "bold"), fg="#0b4a86", bg="#f5f7fb").pack(anchor="w", padx=22, pady=(4, 2))
        info_frame = tk.Frame(win, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        info_frame.pack(fill="x", padx=22, pady=(0, 10))
        info_checks = {}
        for key, label, desc in all_info:
            row_f = tk.Frame(info_frame, bg="#ffffff")
            row_f.pack(fill="x", padx=10, pady=4)
            var = tk.BooleanVar(value=(key in active_info))
            info_checks[key] = var
            tk.Checkbutton(row_f, variable=var, bg="#ffffff", activebackground="#ffffff", cursor="hand2").pack(side="left")
            tk.Label(row_f, text=label, font=(theme.FONT, 11, "bold"), fg="#123", bg="#ffffff", width=26, anchor="w").pack(side="left")
            tk.Label(row_f, text=desc, font=(theme.FONT, 9), fg="#555", bg="#ffffff", anchor="w", wraplength=280).pack(side="left", padx=(4, 0))

        # --- Schnellzugriff-Kacheln ---
        tk.Label(win, text="Schnellzugriff-Kacheln:", font=(theme.FONT, 11, "bold"), fg="#0b4a86", bg="#f5f7fb").pack(anchor="w", padx=22, pady=(4, 2))
        tile_container = tk.Frame(win, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        tile_container.pack(fill="both", expand=True, padx=22, pady=(0, 10))

        canvas = tk.Canvas(tile_container, bg="#ffffff", highlightthickness=0)
        scrollbar = tk.Scrollbar(tile_container, orient="vertical", command=canvas.yview)
        scroll_inner = tk.Frame(canvas, bg="#ffffff")
        scroll_inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        tile_checks = {}
        for key, icon, title, desc, _, color in all_tiles:
            row_f = tk.Frame(scroll_inner, bg="#ffffff")
            row_f.pack(fill="x", padx=10, pady=3)
            var = tk.BooleanVar(value=(key in active_tiles))
            tile_checks[key] = var
            tk.Checkbutton(row_f, variable=var, bg="#ffffff", activebackground="#ffffff", cursor="hand2").pack(side="left")
            tk.Label(row_f, text=icon, font=(theme.FONT, 13), fg=color, bg="#ffffff", width=3).pack(side="left")
            tk.Label(row_f, text=title, font=(theme.FONT, 11, "bold"), fg="#123", bg="#ffffff", width=20, anchor="w").pack(side="left")
            tk.Label(row_f, text=desc, font=(theme.FONT, 9), fg="#555", bg="#ffffff", anchor="w").pack(side="left", padx=(4, 0))

        def save():
            new_tiles = {k for k, v in tile_checks.items() if v.get()}
            new_info = {k for k, v in info_checks.items() if v.get()}
            # V1.1 SP11: leeres Set ist OK - Dashboard wird auf der Startseite
            # einfach kleiner dargestellt (kein leerer Platzhalter).
            if len(new_tiles) > 8:
                messagebox.showinfo("Dashboard", f"Maximal 8 Schnellzugriff-Kacheln erlaubt. Aktuell gewählt: {len(new_tiles)}.\nBitte {len(new_tiles)-8} Kachel(n) abwählen.")
                return
            self._dashboard_set_active_tiles(new_tiles)
            self._dashboard_set_active_info(new_info)
            win.destroy()
            self.show_startseite()

        def select_all():
            for v in tile_checks.values():
                v.set(True)
            for v in info_checks.values():
                v.set(True)

        def deselect_all():
            for v in tile_checks.values():
                v.set(False)
            for v in info_checks.values():
                v.set(False)

        btn_row = tk.Frame(win, bg="#f5f7fb")
        btn_row.pack(fill="x", padx=22, pady=(0, 14))
        tk.Button(btn_row, text="Alle auswählen", command=select_all, padx=10, pady=6).pack(side="left", padx=(0, 6))
        tk.Button(btn_row, text="Alle abwählen", command=deselect_all, padx=10, pady=6).pack(side="left")
        tk.Button(btn_row, text="Abbrechen", command=win.destroy, padx=14, pady=6).pack(side="right", padx=(8, 0))
        tk.Button(btn_row, text="✔  Speichern", command=save, bg="#0b4a86", fg="white", relief="flat", font=(theme.FONT, 11, "bold"), padx=16, pady=6).pack(side="right")

    def show_globale_suche(self, start_query=""):
        """Globale Suche über Kunden, Auswertungen, ToDos und Mitarbeiter."""
        win = tk.Toplevel(self)
        win.resizable(True, True)
        win.title("Globale Suche")
        win.geometry("860x600")
        win.minsize(640, 460)
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()

        tk.Label(win, text="🔍  Globale Suche", font=(theme.FONT, 14, "bold"), fg="#0b4a86", bg="#f5f7fb").pack(anchor="w", padx=20, pady=(16, 4))
        tk.Label(win, text="Suche in Kunden (Name, Kundennummer, PLZ, Inhaber), Auswertungen und ToDos.", font=(theme.FONT, 9), fg="#666", bg="#f5f7fb").pack(anchor="w", padx=20, pady=(0, 8))

        search_frame = tk.Frame(win, bg="#f5f7fb")
        search_frame.pack(fill="x", padx=20, pady=(0, 8))
        search_var = tk.StringVar(value=start_query)
        search_entry = tk.Entry(search_frame, textvariable=search_var, font=(theme.FONT, 12), width=50)
        search_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        search_entry.focus_set()

        # Ergebnis-Notebook (Tabs)
        nb = ttk.Notebook(win)
        nb.pack(fill="both", expand=True, padx=20, pady=(0, 8))

        # Tab 1: Kunden
        kunden_frame = tk.Frame(nb, bg="#ffffff")
        nb.add(kunden_frame, text="👥 Kunden")
        kunden_frame.columnconfigure(0, weight=1)
        kunden_frame.rowconfigure(0, weight=1)
        k_cols = ("kundennummer", "kundenname", "plz", "ort", "inhaber", "email", "status")
        k_tree = ttk.Treeview(kunden_frame, columns=k_cols, show="headings", selectmode="browse")
        k_heads = {"kundennummer":"Kundennr.", "kundenname":"Apothekenname", "plz":"PLZ",
                   "ort":"Ort", "inhaber":"Inhaber", "email":"E-Mail", "status":"Status"}
        for c in k_cols:
            k_tree.heading(c, text=k_heads[c])
            k_tree.column(c, width=110 if c != "kundenname" else 200, anchor="w")
        k_tree.grid(row=0, column=0, sticky="nsew")
        k_sb = tk.Scrollbar(kunden_frame, orient="vertical", command=k_tree.yview)
        k_sb.grid(row=0, column=1, sticky="ns")
        k_tree.configure(yscrollcommand=k_sb.set)
        k_row_map = {}

        def open_kunde(event=None):
            sel = k_tree.selection()
            if sel:
                self._kunden_detail_dialog(k_row_map.get(sel[0]))
        k_tree.bind("<Double-1>", open_kunde)
        tk.Label(kunden_frame, text="Doppelklick = Kundendaten öffnen", font=(theme.FONT, 8), fg="#888", bg="#ffffff").grid(row=1, column=0, sticky="w", padx=4)

        # Tab 2: Auswertungen
        ana_frame = tk.Frame(nb, bg="#ffffff")
        nb.add(ana_frame, text="📊 Auswertungen")
        ana_frame.columnconfigure(0, weight=1)
        ana_frame.rowconfigure(0, weight=1)
        a_cols = ("datum", "apotheke", "kundennummer", "typ", "treffer")
        a_tree = ttk.Treeview(ana_frame, columns=a_cols, show="headings", selectmode="browse")
        a_heads = {"datum":"Datum", "apotheke":"Name", "kundennummer":"Kundennr.", "typ":"Typ", "treffer":"Treffer"}
        for c in a_cols:
            a_tree.heading(c, text=a_heads[c])
            a_tree.column(c, width=100 if c != "apotheke" else 250, anchor="w")
        a_tree.grid(row=0, column=0, sticky="nsew")
        a_sb = tk.Scrollbar(ana_frame, orient="vertical", command=a_tree.yview)
        a_sb.grid(row=0, column=1, sticky="ns")
        a_tree.configure(yscrollcommand=a_sb.set)
        a_row_map = {}

        # Tab 3: ToDos
        todo_frame = tk.Frame(nb, bg="#ffffff")
        nb.add(todo_frame, text="✅ ToDos")
        todo_frame.columnconfigure(0, weight=1)
        todo_frame.rowconfigure(0, weight=1)
        t_cols = ("titel", "bereich", "status", "prioritaet", "faellig_am")
        t_tree = ttk.Treeview(todo_frame, columns=t_cols, show="headings", selectmode="browse")
        t_heads = {"titel":"ToDo", "bereich":"Bereich", "status":"Status", "prioritaet":"Priorität", "faellig_am":"Fällig"}
        for c in t_cols:
            t_tree.heading(c, text=t_heads[c])
            t_tree.column(c, width=100 if c != "titel" else 280, anchor="w")
        t_tree.grid(row=0, column=0, sticky="nsew")
        t_sb = tk.Scrollbar(todo_frame, orient="vertical", command=t_tree.yview)
        t_sb.grid(row=0, column=1, sticky="ns")
        t_tree.configure(yscrollcommand=t_sb.set)

        def do_search(*args):
            q = search_var.get().strip().lower()
            # Kunden
            for item in k_tree.get_children(): k_tree.delete(item)
            k_row_map.clear()
            try:
                with sqlite3.connect(DB_PATH) as con:
                    con.row_factory = sqlite3.Row
                    kunden = con.execute("SELECT * FROM tbl_kunden_center ORDER BY kundenname").fetchall()
                for k in kunden:
                    search_str = " ".join(str(k[c] or "") for c in k.keys()).lower()
                    if not q or q in search_str:
                        vals = tuple(str(k[c] or "") for c in k_cols if c in k.keys())
                        # pad fehlende Spalten
                        vals = tuple(str(k[c] if c in k.keys() else "") for c in k_cols)
                        iid = k_tree.insert("", "end", values=vals)
                        k_row_map[iid] = dict(k)
            except Exception: pass

            # Auswertungen
            for item in a_tree.get_children(): a_tree.delete(item)
            a_row_map.clear()
            try:
                with sqlite3.connect(DB_PATH) as con:
                    con.row_factory = sqlite3.Row
                    rows = con.execute("""
                        SELECT id, datum, apotheke, COALESCE(kundennummer,'') as kundennummer,
                               COALESCE(datenquelle,'NMG') as datenquelle, nmg_treffer
                        FROM tbl_auswertungen ORDER BY datetime(datum) DESC LIMIT 300
                    """).fetchall()
                for r in rows:
                    search_str = f"{r['apotheke'] or ''} {r['kundennummer'] or ''}".lower()
                    if not q or q in search_str:
                        dq = _dq_label(r["datenquelle"])
                        iid = a_tree.insert("", "end", values=(
                            str(r["datum"] or "")[:10], str(r["apotheke"] or ""),
                            str(r["kundennummer"] or ""), dq, r["nmg_treffer"] or 0
                        ))
                        a_row_map[iid] = dict(r)
            except Exception: pass

            # ToDos
            for item in t_tree.get_children(): t_tree.delete(item)
            try:
                with sqlite3.connect(DB_PATH) as con:
                    con.row_factory = sqlite3.Row
                    todos = con.execute("SELECT * FROM tbl_todo_center ORDER BY faellig_am").fetchall()
                for t in todos:
                    search_str = f"{t['titel'] or ''} {t['bereich'] or ''}".lower()
                    if not q or q in search_str:
                        t_tree.insert("", "end", values=tuple(str(t[c] or "") for c in t_cols))
            except Exception: pass

            # Tab-Labels aktualisieren
            nb.tab(0, text=f"👥 Kunden ({len(k_tree.get_children())})")
            nb.tab(1, text=f"📊 Auswertungen ({len(a_tree.get_children())})")
            nb.tab(2, text=f"✅ ToDos ({len(t_tree.get_children())})")

        search_var.trace_add("write", do_search)
        do_search()

        bar = tk.Frame(win, bg="#f5f7fb")
        bar.pack(fill="x", padx=20, pady=(0, 14))
        tk.Button(bar, text="Schließen", command=win.destroy, padx=14, pady=7).pack(side="right")

    def _schnellnotiz_widget(self, parent):
        """Schnell-Notiz-Eingabe für die Startseite."""
        box = tk.Frame(parent, bg="#fffef0", highlightbackground="#e8d99a", highlightthickness=1)
        box.pack(fill="x", padx=14, pady=(0, 6))
        box.columnconfigure(1, weight=1)

        tk.Label(box, text="📝 Schnellnotiz", font=(theme.FONT, 9, "bold"), fg="#7a6000", bg="#fffef0").grid(row=0, column=0, sticky="w", padx=8, pady=(5,2))

        notiz_var = tk.StringVar()
        entry = tk.Entry(box, textvariable=notiz_var, font=(theme.FONT, 10), bg="#fffff8",
                         relief="solid", bd=1)
        entry.grid(row=0, column=1, sticky="ew", padx=(4, 4), pady=(5,2))

        def speichern_notiz():
            text = notiz_var.get().strip()
            if not text:
                return

            # Auswahl: ToDo oder Kunde
            win = tk.Toplevel(self)
            win.resizable(True, True)
            win.title("Notiz speichern")
            win.geometry("400x220")
            win.configure(bg="#f5f7fb")
            win.transient(self)
            win.grab_set()

            tk.Label(win, text=f'Notiz: "{text[:50]}"', font=(theme.FONT, 10), fg="#333", bg="#f5f7fb",
                     wraplength=360).pack(anchor="w", padx=20, pady=(14,8))
            tk.Label(win, text="Wohin soll die Notiz?", font=(theme.FONT, 11, "bold"), fg="#0b4a86", bg="#f5f7fb").pack(anchor="w", padx=20)

            btn_row = tk.Frame(win, bg="#f5f7fb")
            btn_row.pack(fill="x", padx=20, pady=(12,0))

            def als_todo():
                try:
                    with sqlite3.connect(DB_PATH) as con:
                        con.execute("CREATE TABLE IF NOT EXISTS tbl_todo_center (id INTEGER PRIMARY KEY AUTOINCREMENT, titel TEXT NOT NULL, bereich TEXT, status TEXT DEFAULT 'offen', prioritaet TEXT DEFAULT 'Normal', bearbeiter TEXT, erstellt_am TEXT DEFAULT CURRENT_TIMESTAMP, geaendert_am TEXT, notizen TEXT, verantwortlich TEXT, faellig_am TEXT)")
                        con.execute("INSERT INTO tbl_todo_center(titel,bereich,status,prioritaet,bearbeiter) VALUES(?,?,?,?,?)",
                                    (text, "Schnellnotiz", "offen", "Normal", self.bearbeiter))
                        con.commit()
                except Exception as exc:
                    messagebox.showerror("Fehler", str(exc))
                    return
                notiz_var.set("")
                win.destroy()

            def als_kundennotiz():
                win.destroy()
                self.show_globale_suche(start_query="")
                # Hinweis: Benutzer wählt dann den Kunden via globale Suche

            tk.Button(btn_row, text="✅  Als ToDo speichern", command=als_todo,
                      bg="#11823b", fg="white", relief="flat", font=(theme.FONT, 10, "bold"),
                      padx=14, pady=6).pack(side="left", padx=(0,8))
            tk.Button(btn_row, text="👥  Zu Kunde (Suche öffnen)", command=als_kundennotiz,
                      bg="#0b4a86", fg="white", relief="flat", font=(theme.FONT, 10),
                      padx=14, pady=6).pack(side="left")
            tk.Button(btn_row, text="Abbrechen", command=win.destroy, padx=10, pady=6).pack(side="right")

        tk.Button(box, text="💾", command=speichern_notiz, bg="#e8d99a", fg="#7a6000",
                  relief="flat", font=(theme.FONT, 11, "bold"), padx=6, pady=2).grid(row=0, column=2, padx=(0,4), pady=(5,2))

        # Suche-Button
        tk.Button(box, text="🔍 Suche", command=self.show_globale_suche, bg="#d8e2ee", fg="#0b4a86",
                  relief="flat", font=(theme.FONT, 9), padx=8, pady=2).grid(row=0, column=3, padx=(0,8), pady=(5,2))

        tk.Label(box, text="Enter = Todo, 🔍 = Globale Suche", font=(theme.FONT, 7), fg="#aaa", bg="#fffef0").grid(
            row=1, column=0, columnspan=4, sticky="w", padx=8, pady=(0,3))
        entry.bind("<Return>", lambda e: speichern_notiz())

    # V1.1 SP7: Globale Suche auf der Startseite ──────────────────────────────
    def _global_search(self, q: str, limit_per_typ: int = 25) -> list[dict]:
        """V1.1 SP7: Sucht parallel in Kunden, gespeicherten Analysen und
        Artikeln. Liefert Liste von dicts mit:
            typ:     'kunde' | 'analyse' | 'artikel'
            label:   Hauptbezeichnung fuer die Treeview
            detail:  Kontextinfo (PZN, Datum, etc.)
            payload: dict mit Aktion-Info (kundennummer, datei, pzn, ...)
        """
        q = (q or "").strip()
        if len(q) < 2:
            return []
        like = f"%{q.lower()}%"
        out: list[dict] = []

        try:
            with sqlite3.connect(DB_PATH) as con:
                con.row_factory = sqlite3.Row

                # --- Kunden (PK + ZW zusammen) ---
                for tbl, typlabel in (("tbl_pk_kunden", "PK"), ("tbl_zw_kunden", "ZW")):
                    try:
                        rows = con.execute(f"""
                            SELECT kundennummer, kundenname, apotheke, status
                            FROM {tbl}
                            WHERE LOWER(COALESCE(kundennummer,'')) LIKE ?
                               OR LOWER(COALESCE(kundenname,'')) LIKE ?
                               OR LOWER(COALESCE(apotheke,'')) LIKE ?
                            LIMIT ?
                        """, (like, like, like, limit_per_typ)).fetchall()
                        for r in rows:
                            label = r["kundenname"] or r["apotheke"] or r["kundennummer"] or "(ohne Name)"
                            details = []
                            if r["kundennummer"]:
                                details.append(f"#{r['kundennummer']}")
                            if r["apotheke"] and r["apotheke"] != label:
                                details.append(r["apotheke"])
                            if r["status"]:
                                details.append(r["status"])
                            out.append({
                                "typ": "kunde",
                                "typlabel": f"Kunde {typlabel}",
                                "label": label,
                                "detail": " · ".join(details),
                                "payload": {
                                    "kundennummer": r["kundennummer"] or "",
                                    "kundenname": r["kundenname"] or "",
                                    "apotheke": r["apotheke"] or "",
                                    "search_value": r["kundennummer"] or r["kundenname"] or r["apotheke"] or "",
                                },
                            })
                    except sqlite3.OperationalError:
                        pass

                # --- Analysen ---
                try:
                    rows = con.execute("""
                        SELECT id, apotheke, kundennummer, kundenname, datum,
                               ausgabedatei, anzahl_positionen, COALESCE(datenquelle,'NMG') AS datenquelle
                        FROM tbl_auswertungen
                        WHERE LOWER(COALESCE(apotheke,'')) LIKE ?
                           OR LOWER(COALESCE(kundennummer,'')) LIKE ?
                           OR LOWER(COALESCE(kundenname,'')) LIKE ?
                        ORDER BY datetime(datum) DESC
                        LIMIT ?
                    """, (like, like, like, limit_per_typ)).fetchall()
                    for r in rows:
                        label = r["apotheke"] or r["kundenname"] or f"Analyse #{r['id']}"
                        details = []
                        if r["datum"]:
                            details.append(str(r["datum"])[:10])
                        if r["datenquelle"]:
                            details.append(_dq_label(r["datenquelle"]))
                        if r["anzahl_positionen"]:
                            details.append(f"{r['anzahl_positionen']} Pos.")
                        out.append({
                            "typ": "analyse",
                            "typlabel": "Analyse",
                            "label": label,
                            "detail": " · ".join(details),
                            "payload": {
                                "id": r["id"],
                                "ausgabedatei": r["ausgabedatei"] or "",
                            },
                        })
                except sqlite3.OperationalError:
                    pass
        except Exception:
            pass

        # --- Artikel (via Vergleichs-Suche, nur wenn Begriff >= 3 Zeichen) ---
        if len(q) >= 3:
            try:
                hits = search_unified(q, limit=limit_per_typ)
                for r in hits:
                    label = r.get("artikel") or r.get("pzn") or "(ohne Name)"
                    details = []
                    if r.get("pzn"):
                        details.append(f"PZN {r['pzn']}")
                    if r.get("herst"):
                        details.append(r["herst"])
                    if r.get("ist_nmg"):
                        details.append("NMG")
                    out.append({
                        "typ": "artikel",
                        "typlabel": "Artikel",
                        "label": label,
                        "detail": " · ".join(details),
                        "payload": {"pzn": r.get("pzn", ""), "query": q},
                    })
            except Exception:
                pass

        return out

    def _global_search_dispatch(self, entry: dict) -> None:
        """V1.1 SP7: Verzweigt die Aktion abhaengig vom Typ des Treffers."""
        typ = entry.get("typ")
        payload = entry.get("payload") or {}
        if typ == "kunde":
            # Such-Wert fuer die Kunden vormerken, dann Seite oeffnen.
            self._kunden_center_pre_search = payload.get("search_value", "")
            self.show_kunden_center()
        elif typ == "analyse":
            datei = payload.get("ausgabedatei", "")
            if datei and Path(datei).exists():
                try:
                    _open_file(Path(datei))
                except Exception as exc:
                    messagebox.showerror("Analyse oeffnen", str(exc))
            else:
                self.open_saved_analyses()
        elif typ == "artikel":
            # Vergleichs-Suche oeffnen, dann das Such-Wort eintragen.
            self._vergleichssuche_pre_query = payload.get("query", "")
            self.open_vergleichssuche_window()

    def _global_suche_widget(self, parent, *, initial_limit: int = 3,
                              tree_height: int = 6, hide_when_empty: bool = False):
        """V1.1 SP7 / SP9 / SP12: Globales Such-Widget.
        Sucht parallel in Kunden, gespeicherten Analysen und Artikeln.
        Doppelklick oder Button 'Anzeigen' springt zum passenden Bereich.

        V1.1 SP9: zeigt nur die ersten `initial_limit` Treffer; bei mehr
        Treffern erscheint ein 'Mehr anzeigen'-Button am unteren Rand.
        V1.1 SP12: hide_when_empty=True versteckt die Treffer-Tabelle so
        lange, bis es mindestens einen Treffer gibt - damit das Widget
        auf der Startseite kompakt bleibt.
        """
        box = tk.Frame(parent, bg="#e8f1fb", highlightbackground="#a8c5e8", highlightthickness=1)
        box.pack(fill="x", padx=18, pady=(10, 0))

        head = tk.Frame(box, bg="#e8f1fb")
        head.pack(fill="x", padx=10, pady=(8, 4))
        tk.Label(head, text="🔍  Globale Suche",
                 font=(theme.FONT, 12, "bold"), fg="#0b4a86", bg="#e8f1fb").pack(side="left")
        tk.Label(head, text="Kunde, Analyse oder Artikel suchen (ab 2 Zeichen) - Doppelklick oeffnet",
                 font=(theme.FONT, 9), fg="#555", bg="#e8f1fb").pack(side="left", padx=(10, 0))

        entry_row = tk.Frame(box, bg="#e8f1fb")
        entry_row.pack(fill="x", padx=10, pady=(0, 6))

        query_var = tk.StringVar()
        entry = tk.Entry(entry_row, textvariable=query_var, font=(theme.FONT, 11), width=50)
        entry.pack(side="left", padx=(0, 8))

        status_lbl = tk.Label(entry_row, text="", bg="#e8f1fb", fg="#555", font=(theme.FONT, 9))
        status_lbl.pack(side="left")

        # Treeview fuer Treffer. V1.1 SP12: bei hide_when_empty erst sichtbar
        # wenn Treffer da sind.
        tree_frame = tk.Frame(box, bg="#e8f1fb")
        if not hide_when_empty:
            tree_frame.pack(fill="x", padx=10, pady=(0, 4))
        cols = ("typ", "label", "detail")
        heads = {"typ": "Typ", "label": "Bezeichnung", "detail": "Details"}
        widths = {"typ": 90, "label": 360, "detail": 260}
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings",
                            selectmode="browse", height=tree_height)
        for c in cols:
            tree.heading(c, text=heads[c])
            tree.column(c, width=widths[c], anchor="w", stretch=(c == "label"))
        tree.pack(side="left", fill="x", expand=True)
        sb = tk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        sb.pack(side="right", fill="y")
        tree.configure(yscrollcommand=sb.set)

        # V1.1 SP9: 'Mehr anzeigen'-Button-Zeile, default versteckt.
        more_row = tk.Frame(box, bg="#e8f1fb")
        more_btn = tk.Button(more_row, text="Mehr anzeigen",
                             bg="#0b4a86", fg="white", relief="flat",
                             font=(theme.FONT, 10, "bold"), padx=12, pady=4)
        more_btn.pack()

        row_map: dict[str, dict] = {}
        pending = {"id": None}
        seq = {"current": 0}
        all_results: list[dict] = []
        state = {"shown": initial_limit}

        def _render_subset(subset: list[dict]):
            for iid in tree.get_children():
                tree.delete(iid)
            row_map.clear()
            for entry in subset:
                iid = tree.insert("", "end", values=(
                    entry["typlabel"], entry["label"], entry["detail"],
                ))
                row_map[iid] = entry

        def _update_status_and_button():
            total = len(all_results)
            shown = min(state["shown"], total)
            if total == 0:
                status_lbl.configure(text="" if not query_var.get().strip() else "0 Treffer")
                more_row.pack_forget()
                # V1.1 SP12: keine Treffer -> Tabelle verstecken (Startseite kompakt).
                if hide_when_empty:
                    tree_frame.pack_forget()
                return
            # V1.1 SP12: jetzt zeigen sobald es Treffer gibt.
            if hide_when_empty and not tree_frame.winfo_ismapped():
                tree_frame.pack(fill="x", padx=10, pady=(0, 4))
            if shown < total:
                status_lbl.configure(text=f"{shown} von {total} Treffern")
                more_btn.configure(text=f"Mehr anzeigen ({total - shown})")
                more_row.pack(fill="x", padx=10, pady=(0, 6))
            else:
                status_lbl.configure(text=f"{total} Treffer")
                more_row.pack_forget()

        def render(rows: list[dict]):
            nonlocal all_results
            all_results = list(rows)
            state["shown"] = initial_limit
            _render_subset(all_results[:state["shown"]])
            _update_status_and_button()

        def show_more():
            state["shown"] = len(all_results)
            _render_subset(all_results)
            _update_status_and_button()
        more_btn.configure(command=show_more)

        def do_search():
            pending["id"] = None
            q = query_var.get().strip()
            if len(q) < 2:
                render([])
                status_lbl.configure(text="Mindestens 2 Zeichen.")
                return
            status_lbl.configure(text="Suche ...")
            seq["current"] += 1
            my_id = seq["current"]

            def worker():
                try:
                    result = self._global_search(q, limit_per_typ=25)
                except Exception as exc:
                    result = exc

                def apply():
                    if my_id != seq["current"]:
                        return
                    if isinstance(result, Exception):
                        status_lbl.configure(text=f"Fehler: {result}")
                        return
                    render(result)

                try:
                    parent.after(0, apply)
                except Exception:
                    pass

            threading.Thread(target=worker, daemon=True).start()

        def on_key(_e=None):
            if pending["id"]:
                try:
                    parent.after_cancel(pending["id"])
                except Exception:
                    pass
            pending["id"] = parent.after(250, do_search)

        def open_selected(_e=None):
            sel = tree.selection()
            if not sel:
                return
            entry = row_map.get(sel[0])
            if entry:
                self._global_search_dispatch(entry)

        entry.bind("<KeyRelease>", on_key)
        entry.bind("<Return>", lambda _e: do_search())
        tree.bind("<Double-1>", open_selected)
        tree.bind("<Return>", open_selected)

        tk.Button(entry_row, text="🔍  Anzeigen", command=open_selected,
                  bg="#0b4a86", fg="white", relief="flat",
                  font=(theme.FONT, 10, "bold"), padx=14, pady=4).pack(side="left", padx=(0, 6))

        def _clear():
            query_var.set("")
            render([])
            entry.focus_set()
        tk.Button(entry_row, text="Leeren", command=_clear, padx=10, pady=4).pack(side="left")

        return {"entry": entry, "query_var": query_var, "do_search": do_search}

    def open_globale_suche_window(self):
        """V1.1 SP9: Globale Suche als eigenes Toplevel-Fenster.
        Singleton: zweiter Aufruf bringt das vorhandene Fenster nach vorn.
        """
        existing = getattr(self, "_globale_suche_window", None)
        if existing is not None:
            try:
                if existing.winfo_exists():
                    existing.deiconify()
                    existing.lift()
                    existing.focus_force()
                    return
            except Exception:
                pass

        win = tk.Toplevel(self)
        win.title("NMGone — Globale Suche")
        win.geometry("950x520")
        win.minsize(680, 380)
        win.configure(bg="#ffffff")
        self._globale_suche_window = win

        def on_close():
            self._globale_suche_window = None
            win.destroy()
        win.protocol("WM_DELETE_WINDOW", on_close)

        header = tk.Frame(win, bg="#ffffff")
        header.pack(fill="x", padx=14, pady=(12, 4))
        tk.Label(header, text="Globale Suche",
                 font=(theme.FONT, 15, "bold"), fg="#0b4a86", bg="#ffffff").pack(anchor="w")
        tk.Label(header,
                 text=("Sucht parallel in Kunden, gespeicherten Analysen und Artikeln. "
                       "Doppelklick auf einen Treffer oeffnet die passende Seite."),
                 font=(theme.FONT, 9), fg="#666", bg="#ffffff").pack(anchor="w", pady=(2, 6))

        body = tk.Frame(win, bg="#ffffff")
        body.pack(fill="both", expand=True, padx=4, pady=(0, 14))

        # Widget mit groesserer Treeview (10 Zeilen statt 6).
        ctx = self._global_suche_widget(body, initial_limit=3, tree_height=12)

        # V1.1 SP9: Pre-Search-Mechanismus analog Vergleichs-Suche.
        pre = getattr(self, "_globale_suche_pre_query", "") or ""
        if pre:
            ctx["query_var"].set(pre)
            self._globale_suche_pre_query = ""
            win.after(50, ctx["do_search"])

        win.lift()
        win.focus_force()
        try:
            ctx["entry"].focus_set()
        except Exception:
            pass

    def show_startseite(self):
        self.clear_page()

        active_info = self._dashboard_get_active_info()
        active_keys = self._dashboard_get_active_tiles()
        all_tiles = self._dashboard_all_tiles()
        active_tiles = [t for t in all_tiles if t[0] in active_keys]

        # Scrollbares Hauptframe – füllt den gesamten self.page-Bereich
        outer = tk.Frame(self.page, bg="#ffffff")
        outer.grid(row=0, column=0, rowspan=2, sticky="nsew")
        self.page.columnconfigure(0, weight=1)
        self.page.rowconfigure(0, weight=1)
        self.page.rowconfigure(1, weight=1)
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(0, weight=1)

        # Canvas + Scrollbar für vollständige Scrollunterstützung
        canvas = tk.Canvas(outer, bg="#ffffff", highlightthickness=0)
        vsb = tk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        outer.rowconfigure(0, weight=1)
        outer.columnconfigure(0, weight=1)

        inner = tk.Frame(canvas, bg="#ffffff")
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def on_configure(event=None):
            bbox = canvas.bbox("all")
            if not bbox: return
            canvas.configure(scrollregion=bbox)
            try: canvas.itemconfig(win_id, width=canvas.winfo_width())
            except Exception: pass
            content_h = bbox[3] - bbox[1]
            canvas_h = canvas.winfo_height()
            if content_h > canvas_h + 10:
                vsb.grid(row=0, column=1, sticky="ns")
            else:
                vsb.grid_remove()
        inner.bind("<Configure>", on_configure)
        canvas.bind("<Configure>", lambda e: (on_configure(), canvas.itemconfig(win_id, width=e.width)))

        def on_mousewheel(event):
            try:
                bbox = canvas.bbox("all")
                if bbox and (bbox[3] - bbox[1]) > canvas.winfo_height() + 10:
                    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass
        # Nur auf diesen Canvas binden, nicht global – verhindert Fehler auf anderen Seiten
        canvas.bind("<MouseWheel>", on_mousewheel)
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", on_mousewheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        def _unbind_wheel(event=None):
            try:
                canvas.unbind_all("<MouseWheel>")
            except Exception:
                pass
        canvas.bind("<Destroy>", _unbind_wheel)
        outer.bind("<Destroy>", _unbind_wheel)

        # --- Kopfzeile ---
        header_bar = tk.Frame(inner, bg="#f0f4fa", highlightbackground="#c5d3e8", highlightthickness=1)
        header_bar.pack(fill="x", padx=18, pady=(14, 0))
        begruessung = f"🏠  Guten Tag, {self.bearbeiter or 'Bearbeiter'}!"
        tk.Label(header_bar, text=begruessung, font=(theme.FONT, 15, "bold"), fg="#0b4a86", bg="#f0f4fa").pack(side="left", padx=16, pady=10)
        tk.Button(
            header_bar,
            text="⚙️  Dashboard anpassen",
            command=self._dashboard_open_kachel_editor,
            bg="#0b4a86",
            fg="white",
            activebackground="#0b4a86",
            relief="flat",
            font=(theme.FONT, 10, "bold"),
            padx=14,
            pady=6,
        ).pack(side="right", padx=14, pady=8)

        # --- V1.1 SP7: Globale Suche (Kunde / Analyse / Artikel) ---
        # ersetzt das alte Austausch-Suche-Widget (SP16); die Vergleichs-Suche
        # ist jetzt ein eigenes Tool und uebernimmt die Artikel-Recherche.
        # V1.1 SP12: kompakter Modus - Tabelle nur sichtbar wenn Treffer da,
        # Hoehe 3 Zeilen (Toplevel-Tool ist die ausfuehrliche Variante).
        self._global_suche_widget(inner, initial_limit=3, tree_height=3,
                                  hide_when_empty=True)

        # --- Schnellnotiz + Suche ---
        self._schnellnotiz_widget(inner)

        # --- Info-Panel ---
        if active_info:
            self._dashboard_info_panel(inner, active_info)
            sep = tk.Frame(inner, bg="#d8e2ee", height=1)
            sep.pack(fill="x", padx=14, pady=(2, 0))

        # --- Schnellzugriff ---
        if active_tiles:
            tk.Label(inner, text="Schnellzugriff", font=(theme.FONT, 11, "bold"), fg="#555", bg="#ffffff").pack(anchor="w", padx=22, pady=(10, 2))
            tile_frame = tk.Frame(inner, bg="#ffffff")
            tile_frame.pack(fill="x", padx=14, pady=(0, 18))
            cols_count = 4
            for i in range(cols_count):
                tile_frame.columnconfigure(i, weight=1)

            # V1.1 SP11: alle Kacheln gleich hoch ueber rowconfigure + grid.
            rows_needed = (len(active_tiles) + cols_count - 1) // cols_count
            for ri in range(rows_needed):
                tile_frame.rowconfigure(ri, weight=1, minsize=210)
            for idx, (key, icon, title, desc, cmd, color) in enumerate(active_tiles):
                row = idx // cols_count
                col = idx % cols_count
                f = tk.Frame(tile_frame, bg="#f8fbff", highlightbackground="#d8e2ee", highlightthickness=1)
                f.grid(row=row, column=col, sticky="nsew", padx=8, pady=8)
                f.rowconfigure(2, weight=1)
                f.columnconfigure(0, weight=1)
                tk.Label(f, text=icon, font=(theme.FONT, 26), bg="#f8fbff", fg=color
                         ).grid(row=0, column=0, pady=(12, 4))
                tk.Label(f, text=title, font=(theme.FONT, 11, "bold"), bg="#f8fbff",
                         fg="#123").grid(row=1, column=0)
                tk.Label(f, text=desc, wraplength=180, justify="center",
                         bg="#f8fbff", fg="#555", font=(theme.FONT, 9)
                         ).grid(row=2, column=0, padx=10, pady=6, sticky="n")
                tk.Button(f, text="Öffnen  →", command=cmd,
                          bg=color, fg="white", activebackground=color,
                          relief="flat", font=(theme.FONT, 10, "bold"),
                          padx=14, pady=7
                          ).grid(row=3, column=0, sticky="ew", padx=14, pady=(4, 12))
        else:
            tk.Label(inner, text="Keine Kacheln aktiv. Bitte '⚙️ Dashboard anpassen' nutzen.", fg="#888", bg="#ffffff", font=(theme.FONT, 10)).pack(pady=20)

        self.status.set(f"Startseite bereit.  {len(active_tiles)} Schnellzugriff-Kacheln  |  {len(active_info)} Info-Bereiche aktiv.")

    # ── SHAREPOINT / ONEDRIVE VORBEREITUNG ──────────────────────────────────────
    def show_datenbankpfad_page(self):
        """Datenbankpfad konfigurieren – vorbereitet für OneDrive/SharePoint."""
        self.clear_page()
        self._page_header(
            "Datenbankpfad / Cloud-Synchronisation",
            "Lokaler Pfad, OneDrive-Ordner oder SharePoint-Pfad für die Datenbank."
        )
        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure(1, weight=1)

        current_path = str(DB_PATH)
        override = self._get_meta_value("db_path_override", "")

        tk.Label(body, text="Aktiver Datenbankpfad:", bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 11, "bold")).grid(row=0, column=0, sticky="nw", padx=0, pady=(0, 4))
        tk.Label(body, text=current_path, bg="#ffffff", fg="#333", wraplength=700, justify="left").grid(row=1, column=0, columnspan=2, sticky="w", pady=(0, 16))

        tk.Label(body, text="Status:", bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 11, "bold")).grid(row=2, column=0, sticky="w", pady=(0, 12))
        if override:
            status_text = f"Override aktiv: {override}"
            status_color = "#11823b"
        else:
            status_text = "Standard-Pfad (kein Override gesetzt)"
            status_color = "#666"
        tk.Label(body, text=status_text, bg="#ffffff", fg=status_color, font=(theme.FONT, 10)).grid(row=2, column=1, sticky="w", pady=(0, 12))

        sep = tk.Frame(body, bg="#d8e2ee", height=1)
        sep.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(0, 16))

        info_box = tk.Frame(body, bg="#fff8e1", highlightbackground="#efd39a", highlightthickness=1)
        info_box.grid(row=4, column=0, columnspan=2, sticky="ew", pady=(0, 16))
        tk.Label(info_box, text=(
            "\U0001f512  Diese Funktion ist vorbereitet und wird mit einem zuk\u00fcnftigen Update aktiviert.\n\n"
            "Geplant: Datenbankpfad auf einen synchronisierten OneDrive-Ordner oder SharePoint-Pfad legen.\n"
            "Alle Mitarbeiter die auf denselben Ordner zugreifen teilen automatisch dieselbe Datenbank.\n\n"
            "Voraussetzungen:\n"
            "  \u2022  OneDrive/SharePoint muss auf dem PC synchronisiert sein\n"
            "  \u2022  Alle Nutzer ben\u00f6tigen Schreibrechte auf den Ordner\n"
            "  \u2022  Gleichzeitiger Zugriff wird durch SQLite-Locking gehandhabt\n\n"
            "Zum Aktivieren: Update einspielen und hier den Zielordner ausw\u00e4hlen."
        ), bg="#fff8e1", fg="#5a3800", justify="left", anchor="w", padx=14, pady=12, font=(theme.FONT, 10)).pack(fill="x")

        # Vorschau: Pfad wählen (noch deaktiviert)
        path_var = tk.StringVar(value=override or current_path)
        tk.Label(body, text="Zielordner (Vorschau):", bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 11, "bold")).grid(row=5, column=0, sticky="w", pady=(8, 4))
        pf = tk.Frame(body, bg="#ffffff")
        pf.grid(row=6, column=0, columnspan=2, sticky="ew", pady=(0, 8))
        pf.columnconfigure(0, weight=1)
        path_entry = tk.Entry(pf, textvariable=path_var, state="disabled", fg="#888")
        path_entry.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        tk.Button(pf, text="Ordner auswählen (inaktiv)", state="disabled", padx=12, pady=5).grid(row=0, column=1)

        tk.Label(body, text="Diese Einstellung ist mit dem aktuellen Programmstand noch nicht aktiv.\nSie wird durch ein zuk\u00fcnftiges Update freigeschaltet.",
                 bg="#ffffff", fg="#888", font=(theme.FONT, 9), justify="left").grid(row=7, column=0, columnspan=2, sticky="w", pady=(4, 0))

        self.status.set("Datenbankpfad-Konfiguration – bereit für zukünftiges Cloud-Update.")

    def show_report_page(self):
        """Report-Seite: Kunden-Ampel-Auswertung und Duplikat-Erkennung."""
        self.clear_page()
        self._page_header("Report", "Kunden-Ampel, Duplikate und Auswertungsübersicht.")

        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0,14))
        body.columnconfigure(0, weight=1)
        body.rowconfigure(1, weight=1)

        # Tab-Auswahl
        nb = ttk.Notebook(body)
        nb.grid(row=0, column=0, sticky="nsew", rowspan=2)
        body.rowconfigure(0, weight=1)

        # ── Tab 1: Kunden-Ampel ─────────────────────────────────────────────
        ampel_frame = tk.Frame(nb, bg="#ffffff")
        nb.add(ampel_frame, text="🚦 Kunden-Ampel")
        ampel_frame.columnconfigure(0, weight=1)
        ampel_frame.rowconfigure(1, weight=1)

        af_top = tk.Frame(ampel_frame, bg="#ffffff")
        af_top.grid(row=0, column=0, sticky="ew", pady=(8,4))
        filter_var = tk.StringVar(value="alle")
        tk.Label(af_top, text="Filter:", bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 10, "bold")).pack(side="left")
        for val, label, color in [("alle","Alle","#555"),("gruen","🟢 Grün (<6M)","#11823b"),("gelb","🟡 Gelb (6-9M)","#8b6914"),("rot","🔴 Rot (>9M / nie)","#c00")]:
            tk.Radiobutton(af_top, text=label, variable=filter_var, value=val,
                           bg="#ffffff", fg=color, command=lambda: reload_ampel(filter_var.get())).pack(side="left", padx=6)
        tk.Button(af_top, text="📄 Export als PDF vorbereiten", bg="#3867b7", fg="white", relief="flat",
                  padx=10, pady=4, command=lambda: messagebox.showinfo("Report", "PDF-Export wird in einer zukünftigen Version implementiert.")).pack(side="right", padx=8)

        a_cols = ("ampel", "kundennummer", "kundenname", "plz", "status", "letzte_analyse", "tage")
        a_heads = {"ampel":"🚦","kundennummer":"Kundennr.","kundenname":"Apotheke","plz":"PLZ",
                   "status":"Status","letzte_analyse":"Letzte Analyse","tage":"Tage"}
        a_tree = ttk.Treeview(ampel_frame, columns=a_cols, show="headings", selectmode="browse")
        for c in a_cols:
            a_tree.heading(c, text=a_heads[c])
            a_tree.column(c, width=30 if c=="ampel" else (180 if c=="kundenname" else 80), anchor="w")
        a_tree.grid(row=1, column=0, sticky="nsew", padx=8, pady=(0,8))
        a_sb = tk.Scrollbar(ampel_frame, orient="vertical", command=a_tree.yview)
        a_sb.grid(row=1, column=1, sticky="ns", pady=(0,8))
        a_tree.configure(yscrollcommand=a_sb.set)
        a_row_map = {}

        def reload_ampel(filt="alle"):
            from datetime import datetime
            for item in a_tree.get_children(): a_tree.delete(item)
            a_row_map.clear()
            try:
                with sqlite3.connect(DB_PATH) as con:
                    con.row_factory = sqlite3.Row
                    kunden = con.execute("SELECT * FROM tbl_kunden_center ORDER BY kundenname").fetchall()
                for k in kunden:
                    if str(k["status"] or "").lower() == "inaktiv":
                        continue
                    knr = str(k["kundennummer"] or "")
                    kname = str(k["kundenname"] or "")
                    with sqlite3.connect(DB_PATH) as con:
                        row = con.execute("""
                            SELECT MAX(datum) as letzte FROM tbl_auswertungen
                            WHERE (kundennummer=? AND kundennummer<>'') OR (kundenname=? AND kundenname<>'')
                        """, (knr, kname)).fetchone()
                    letzte = row[0] if row and row[0] else None
                    if letzte:
                        try:
                            dt = datetime.strptime(str(letzte)[:10], "%Y-%m-%d")
                            tage = (datetime.now() - dt).days
                            datum_str = dt.strftime("%d.%m.%Y")
                        except Exception:
                            tage = 9999; datum_str = str(letzte)[:10]
                    else:
                        tage = 9999; datum_str = "–"

                    if tage < 180: amp_key = "gruen"; amp = "🟢"
                    elif tage < 270: amp_key = "gelb"; amp = "🟡"
                    else: amp_key = "rot"; amp = "🔴"

                    if filt != "alle" and filt != amp_key:
                        continue

                    iid = a_tree.insert("", "end", values=(
                        amp, knr, kname,
                        str(k["plz"] if "plz" in k.keys() else ""),
                        str(k["status"] or ""),
                        datum_str, f"{tage}d" if tage < 9999 else "–"
                    ))
                    a_row_map[iid] = dict(k)
            except Exception as exc:
                tk.Label(ampel_frame, text=f"Fehler: {exc}", fg="#c00", bg="#ffffff").grid(row=2, column=0)
            nb.tab(0, text=f"🚦 Kunden-Ampel ({len(a_tree.get_children())})")

        reload_ampel()
        filter_var.trace_add("write", lambda *_: reload_ampel(filter_var.get()))

        # ── Tab 2: Duplikate ─────────────────────────────────────────────────
        dup_frame = tk.Frame(nb, bg="#ffffff")
        nb.add(dup_frame, text="🔍 Duplikate")
        dup_frame.columnconfigure(0, weight=1)
        dup_frame.rowconfigure(1, weight=1)

        df_top = tk.Frame(dup_frame, bg="#ffffff")
        df_top.grid(row=0, column=0, sticky="ew", pady=(8,4))
        tk.Label(df_top, text="Duplikate nach Namen, PLZ oder Kundennummer.", bg="#ffffff", fg="#555", font=(theme.FONT, 9)).pack(side="left", padx=8)
        tk.Button(df_top, text="🔄 Aktualisieren", command=lambda: reload_duplikate(), padx=10, pady=4).pack(side="right", padx=8)

        d_cols = ("id1","name1","plz1","id2","name2","plz2","grund")
        d_heads = {"id1":"ID 1","name1":"Apotheke 1","plz1":"PLZ 1","id2":"ID 2","name2":"Apotheke 2","plz2":"PLZ 2","grund":"Grund"}
        d_tree = ttk.Treeview(dup_frame, columns=d_cols, show="headings", selectmode="browse")
        for c in d_cols:
            d_tree.heading(c, text=d_heads[c])
            d_tree.column(c, width=50 if c.startswith("id") else (160 if "name" in c else 60), anchor="w")
        d_tree.grid(row=1, column=0, sticky="nsew", padx=8, pady=(0,4))
        d_sb = tk.Scrollbar(dup_frame, orient="vertical", command=d_tree.yview)
        d_sb.grid(row=1, column=1, sticky="ns", pady=(0,4))
        d_tree.configure(yscrollcommand=d_sb.set)
        d_row_map = {}

        def reload_duplikate():
            for item in d_tree.get_children(): d_tree.delete(item)
            d_row_map.clear()
            try:
                with sqlite3.connect(DB_PATH) as con:
                    con.row_factory = sqlite3.Row
                    kunden = con.execute("SELECT * FROM tbl_kunden_center ORDER BY kundenname").fetchall()
                kunden_list = [dict(k) for k in kunden]
                seen = set()
                for i, k1 in enumerate(kunden_list):
                    for k2 in kunden_list[i+1:]:
                        pair = (min(k1["id"],k2["id"]), max(k1["id"],k2["id"]))
                        if pair in seen: continue
                        grund = None
                        n1 = str(k1["kundenname"] or "").lower().strip()
                        n2 = str(k2["kundenname"] or "").lower().strip()
                        p1 = str(k1["plz"] if "plz" in k1 else "").strip()
                        p2 = str(k2["plz"] if "plz" in k2 else "").strip()
                        nr1 = str(k1["kundennummer"] or "").strip()
                        nr2 = str(k2["kundennummer"] or "").strip()
                        if nr1 and nr2 and nr1 == nr2:
                            grund = "Gleiche Kundennummer"
                        elif n1 and n2 and (n1 == n2 or (len(n1)>4 and n1 in n2) or (len(n2)>4 and n2 in n1)):
                            grund = "Ähnlicher Name"
                        elif p1 and p2 and p1 == p2 and n1 and n2 and n1[:4] == n2[:4]:
                            grund = "Gleiche PLZ + ähnlicher Name"
                        if grund:
                            seen.add(pair)
                            iid = d_tree.insert("", "end", values=(
                                k1["id"], k1["kundenname"] or "–", p1,
                                k2["id"], k2["kundenname"] or "–", p2, grund
                            ))
                            d_row_map[iid] = (k1, k2)
            except Exception as exc:
                pass
            nb.tab(1, text=f"🔍 Duplikate ({len(d_tree.get_children())})")

        reload_duplikate()

        def zusammenfuehren():
            sel = d_tree.selection()
            if not sel:
                messagebox.showinfo("Duplikate", "Bitte ein Duplikat-Paar auswählen.")
                return
            k1, k2 = d_row_map.get(sel[0], (None, None))
            if not k1 or not k2:
                return
            self._kunden_zusammenfuehren_dialog(k1, k2, callback=reload_duplikate)

        tk.Button(dup_frame, text="🔀 Zusammenführen", command=zusammenfuehren,
                  bg="#8b5a00", fg="white", relief="flat", font=(theme.FONT, 10, "bold"),
                  padx=12, pady=5).grid(row=2, column=0, sticky="w", padx=8, pady=(0,8))

        # Nav-Eintrag: Report in Nav-Tree (falls noch nicht vorhanden)
        self.status.set("Report bereit.")

    def _kunden_zusammenfuehren_dialog(self, k1, k2, callback=None):
        """Dialog: Zwei Kunden zusammenführen – Felder auswählen."""
        win = tk.Toplevel(self)
        win.resizable(True, True)
        win.title("Kunden zusammenführen")
        win.geometry("860x580")
        win.minsize(700, 480)
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()

        tk.Label(win, text="🔀  Kunden zusammenführen", font=(theme.FONT, 14, "bold"), fg="#0b4a86", bg="#f5f7fb").pack(anchor="w", padx=20, pady=(16,4))
        tk.Label(win, text="Wähle für jedes Feld welcher Wert behalten werden soll. Der nicht gewählte Eintrag wird gelöscht.", font=(theme.FONT, 9), fg="#666", bg="#f5f7fb", justify="left").pack(anchor="w", padx=20, pady=(0,10))

        felder = ["kundennummer","kundenname","plz","ort","strasse","kundentyp","inhaber",
                  "ansprechpartner","ansprechpartner2","telefon","email","status","notizen"]
        f_labels = {"kundennummer":"Kundennummer","kundenname":"Apothekenname","plz":"PLZ",
                    "ort":"Ort","strasse":"Straße","kundentyp":"Typ","inhaber":"Inhaber",
                    "ansprechpartner":"Ansprechpartner","ansprechpartner2":"AP 2",
                    "telefon":"Telefon","email":"E-Mail","status":"Status","notizen":"Notizen"}

        form = tk.Frame(win, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        form.pack(fill="both", expand=True, padx=20, pady=(0,10))
        form.columnconfigure(1, weight=1)
        form.columnconfigure(2, weight=1)

        tk.Label(form, text="Feld", bg="#e8edf5", font=(theme.FONT,9,"bold")).grid(row=0,column=0,sticky="ew",padx=6,pady=4)
        tk.Label(form, text=f"Eintrag 1 (ID {k1['id']})", bg="#e8edf5", font=(theme.FONT,9,"bold")).grid(row=0,column=1,sticky="ew",padx=6,pady=4)
        tk.Label(form, text=f"Eintrag 2 (ID {k2['id']})", bg="#e8edf5", font=(theme.FONT,9,"bold")).grid(row=0,column=2,sticky="ew",padx=6,pady=4)

        choices = {}
        for r, f in enumerate(felder, start=1):
            v1 = str(k1.get(f,"") or "")
            v2 = str(k2.get(f,"") or "")
            tk.Label(form, text=f_labels.get(f,f), bg="#ffffff", fg="#0b4a86", font=(theme.FONT,9,"bold")).grid(row=r,column=0,sticky="w",padx=8,pady=3)
            var = tk.StringVar(value="1")
            choices[f] = var
            rb1 = tk.Radiobutton(form, text=v1[:50] or "–", variable=var, value="1", bg="#ffffff", anchor="w", wraplength=280)
            rb1.grid(row=r, column=1, sticky="ew", padx=6, pady=2)
            rb2 = tk.Radiobutton(form, text=v2[:50] or "–", variable=var, value="2", bg="#ffffff", anchor="w", wraplength=280)
            rb2.grid(row=r, column=2, sticky="ew", padx=6, pady=2)
            # Bevorzuge nicht-leeren Wert
            if not v1 and v2: var.set("2")

        bar = tk.Frame(win, bg="#f5f7fb")
        bar.pack(fill="x", padx=20, pady=(0,14))

        def do_merge():
            merged = {}
            for f, var in choices.items():
                merged[f] = str(k1.get(f,"") or "") if var.get() == "1" else str(k2.get(f,"") or "")
            merged["bearbeiter"] = self.bearbeiter
            keep_id = k1["id"]
            del_id = k2["id"]
            try:
                with sqlite3.connect(DB_PATH) as con:
                    # Auswertungen umhängen
                    for col in ("kundennummer","kundenname"):
                        try:
                            con.execute(f"UPDATE tbl_auswertungen SET {col}=? WHERE kundennummer=? OR kundenname=?",
                                        (merged.get(col,""), str(k2.get("kundennummer","") or ""), str(k2.get("kundenname","") or "")))
                        except Exception:
                            pass
                    # Ziel-Datensatz aktualisieren
                    set_sql = ", ".join(f"{f}=:{f}" for f in felder)
                    merged["id"] = keep_id
                    con.execute(f"UPDATE tbl_kunden_center SET {set_sql}, geaendert_am=CURRENT_TIMESTAMP, bearbeiter=:bearbeiter WHERE id=:id", merged)
                    # Duplikat löschen
                    con.execute("DELETE FROM tbl_kunden_center WHERE id=?", (del_id,))
                    con.commit()
                messagebox.showinfo("Zusammenführen", f"Zusammenführung abgeschlossen. Eintrag {del_id} gelöscht, Eintrag {keep_id} behalten.")
                win.destroy()
                if callback: callback()
            except Exception as exc:
                messagebox.showerror("Fehler", str(exc))

        tk.Button(bar, text="Abbrechen", command=win.destroy, padx=14, pady=7).pack(side="right", padx=(8,0))
        tk.Button(bar, text="🔀  Zusammenführen", command=do_merge, bg="#8b5a00", fg="white",
                  relief="flat", font=(theme.FONT,11,"bold"), padx=16, pady=7).pack(side="right")

    def show_placeholder_page(self, title):
        self.clear_page()
        self._page_header(title, "Dieser Bereich ist vorbereitet und wird später fachlich ausgebaut.")
        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        tk.Label(
            body,
            text=(
                f"{title} ist als Startseitenbereich angelegt.\n\n"
                "Die Fachlogik wird später separat umgesetzt, damit die aktuelle Auswertungs-, Schulbank- und Vorlagenlogik nicht verändert wird."
            ),
            bg="#ffffff",
            fg="#333",
            justify="left",
            font=(theme.FONT, 11),
        ).pack(anchor="w", pady=(4, 16))
        try:
            add_roadmap_item(
                bereich="Startseite",
                titel=f"Startseitenbereich {title}",
                beschreibung=f"Der Bereich {title} wurde als Startseiten-Kachel vorbereitet und wird später fachlich ausgebaut.",
                status="Offen",
                prioritaet="Normal",
            )
        except Exception:
            pass

    def _ensure_center_tables(self):
        """Bereitet Mitarbeiter-, Kunden-, Bestell- und ToDo-Tabellen vor."""
        with sqlite3.connect(DB_PATH) as con:
            con.execute("""\nCREATE TABLE IF NOT EXISTS tbl_mitarbeiter (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,\nvorname TEXT,
                    name TEXT,\nabteilung TEXT,
                    position TEXT,\ntelefon TEXT,
                    mobil TEXT,\nemail TEXT,
                    vertretung_1 TEXT,\nvertretung_2 TEXT,
                    aufgaben TEXT,\nonedrive_pfad TEXT,
                    notizen TEXT,\naktiv INTEGER DEFAULT 1,
                    erstellt_am TEXT DEFAULT CURRENT_TIMESTAMP,\ngeaendert_am TEXT,
                    bearbeiter TEXT\n)
            """)
            con.execute("""\nCREATE TABLE IF NOT EXISTS tbl_kunden_center (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,\nkundennummer TEXT,
                    kundenname TEXT,\nkundentyp TEXT,
                    ansprechpartner TEXT,\ntelefon TEXT,
                    email TEXT,\nstatus TEXT DEFAULT 'aktiv',
                    notizen TEXT,\nerstellt_am TEXT DEFAULT CURRENT_TIMESTAMP,
                    geaendert_am TEXT,\nbearbeiter TEXT
                )\n""")
            con.execute("""\nCREATE TABLE IF NOT EXISTS tbl_bestellungen_center (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,\nkundennummer TEXT,
                    kunde TEXT,\nbestellnummer TEXT,
                    datum TEXT,\nstatus TEXT DEFAULT 'offen',
                    quelle TEXT,\nnotizen TEXT,
                    erstellt_am TEXT DEFAULT CURRENT_TIMESTAMP,\ngeaendert_am TEXT,
                    bearbeiter TEXT\n)
            """)
            con.execute("""\nCREATE TABLE IF NOT EXISTS tbl_todo_center (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,\ntitel TEXT NOT NULL,
                    bereich TEXT,\nverantwortlich TEXT,
                    faellig_am TEXT,\nstatus TEXT DEFAULT 'offen',
                    prioritaet TEXT DEFAULT 'normal',\nnotizen TEXT,
                    erstellt_am TEXT DEFAULT CURRENT_TIMESTAMP,\ngeaendert_am TEXT,
                    bearbeiter TEXT\n)
            """)
            # Board-/Post-it-Ansicht vorbereiten: Position der Mitarbeiterkarten speichern.
            cols_m = {row[1] for row in con.execute("PRAGMA table_info(tbl_mitarbeiter)").fetchall()}
            if "board_x" not in cols_m:
                con.execute("ALTER TABLE tbl_mitarbeiter ADD COLUMN board_x INTEGER DEFAULT 40")
            if "board_y" not in cols_m:
                con.execute("ALTER TABLE tbl_mitarbeiter ADD COLUMN board_y INTEGER DEFAULT 40")
            con.commit()
        self._ensure_mitarbeiter_phase1_tables()
        self._roadmap_mark_center_erledigt()

    # Phase 1 (2026-06-23): frei definierbare Zusatzfelder + Vorgesetzten-Matrix
    # als Grundlage fuer das spaetere Organigramm. Bewusst an tbl_mitarbeiter
    # (id-basiert) gehaengt; die login-basierten Profile bleiben unberuehrt.
    ARTEN_VORGESETZTER = ("disziplinarisch", "fachlich", "Vertretung")

    def _ensure_mitarbeiter_phase1_tables(self):
        """Legt die Zusatzfeld- und Vorgesetzten-Tabellen an (idempotent)."""
        with sqlite3.connect(DB_PATH) as con:
            # Frei definierbare Felddefinitionen (EAV-Pattern, kein DB-Umbau pro Wunsch)
            con.execute("""\nCREATE TABLE IF NOT EXISTS tbl_mitarbeiter_feld (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,\nfeld_name TEXT NOT NULL,
                    feld_typ TEXT DEFAULT 'Text',\noptionen TEXT,
                    reihenfolge INTEGER DEFAULT 0,\naktiv INTEGER DEFAULT 1,
                    erstellt_am TEXT DEFAULT CURRENT_TIMESTAMP\n)
            """)
            # Werte je Mitarbeiter (id aus tbl_mitarbeiter) und Feld
            con.execute("""\nCREATE TABLE IF NOT EXISTS tbl_mitarbeiter_wert (
                    mitarbeiter_id INTEGER NOT NULL,\nfeld_id INTEGER NOT NULL,
                    wert TEXT,\nPRIMARY KEY (mitarbeiter_id, feld_id)\n)
            """)
            # Vorgesetzten-Matrix: mehrere Vorgesetzte je Person, je mit Art.
            # ist_primaer markiert die EINE Beziehung, die den Hauptbaum bildet.
            con.execute("""\nCREATE TABLE IF NOT EXISTS tbl_mitarbeiter_vorgesetzter (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,\nmitarbeiter_id INTEGER NOT NULL,
                    vorgesetzter_id INTEGER NOT NULL,\nart TEXT DEFAULT 'disziplinarisch',
                    ist_primaer INTEGER DEFAULT 0,\nerstellt_am TEXT DEFAULT CURRENT_TIMESTAMP\n)
            """)
            con.commit()

    def _roadmap_mark_center_erledigt(self):
        try:
            titel = "Mitarbeiter-, Kunden-, Bestell- und ToDo-Center vorbereitet"
            for row in list_roadmap_items():
                if str(row["titel"]).strip().lower() == titel.lower():
                    update_roadmap_status(int(row["id"]), "Erledigt")
                    return
            add_roadmap_item(
                bereich="Organisation",
                titel=titel,
                beschreibung=(
                    "Mitarbeiterdatenbank mit Vertretung 1 und Vertretung 2, Kunden mit einem Ansprechpartner, "
                    "Bestell-Center als Vorbereitung und ToDo-Center wurden als Grundstruktur angelegt. "
                    "Mitarbeiter-Center zeigt Aufgaben, Vertretungen, OneDrive-Pfade und bereitet eine Board-/Post-it-Ansicht vor."
                ),
                status="Erledigt",
                prioritaet="Normal",
            )
        except Exception:
            pass

    def _center_text_dialog(self, title, fields, initial=None):
        """Einfaches Formularfenster für Center-Datensätze."""
        initial = initial or {}
        win = tk.Toplevel(self)
        win.resizable(True, True)
        win.title(title)
        win.geometry("760x620")
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()
        win.columnconfigure(0, weight=1)
        tk.Label(win, text=title, font=(theme.FONT, 18, "bold"), fg="#0b4a86", bg="#f5f7fb").pack(anchor="w", padx=22, pady=(18, 8))
        form = tk.Frame(win, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        form.pack(fill="both", expand=True, padx=22, pady=(0, 12))
        form.columnconfigure(1, weight=1)
        vars_ = {}
        for r, (key, label, kind) in enumerate(fields):
            tk.Label(form, text=label, bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 10, "bold")).grid(row=r, column=0, sticky="nw", padx=14, pady=8)
            var = tk.StringVar(value=str(initial.get(key, "") or ""))
            vars_[key] = var
            if kind == "text":
                txt = tk.Text(form, height=4, wrap="word")
                txt.insert("1.0", var.get())
                txt.grid(row=r, column=1, sticky="ew", padx=14, pady=8)
                vars_[key] = txt
            else:
                tk.Entry(form, textvariable=var).grid(row=r, column=1, sticky="ew", padx=14, pady=8)
        result = {"data": None}
        def save():
            data = {}
            for key, widget in vars_.items():
                if isinstance(widget, tk.Text):
                    data[key] = widget.get("1.0", "end").strip()
                else:
                    data[key] = widget.get().strip()
            result["data"] = data
            win.destroy()
        bar = tk.Frame(win, bg="#f5f7fb")
        bar.pack(fill="x", padx=22, pady=(0, 18))
        tk.Button(bar, text="Abbrechen", command=win.destroy, padx=16, pady=8).pack(side="right", padx=(8,0))
        tk.Button(bar, text="Speichern", command=save, bg="#0b4a86", fg="white", relief="flat", padx=20, pady=9).pack(side="right")
        self.wait_window(win)
        return result["data"]

    def _show_center_table(self, title, subtitle, table, columns, headings, fields, insert_sql, update_sql):
        self._ensure_center_tables()
        self.clear_page()
        self._page_header(title, subtitle)
        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure(0, weight=1)
        body.rowconfigure(1, weight=1)
        toolbar = tk.Frame(body, bg="#ffffff")
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        tree = ttk.Treeview(body, columns=columns, show="headings", selectmode="browse")
        tree.grid(row=1, column=0, sticky="nsew")
        for col in columns:
            tree.heading(col, text=headings.get(col, col))
            tree.column(col, width=120 if col != "notizen" else 240, anchor="w")
        try:
            self._make_tree_sortable(tree, columns, headings)
        except Exception:
            pass
        sb = tk.Scrollbar(body, orient="vertical", command=tree.yview)
        sb.grid(row=1, column=0, sticky="nse")
        tree.configure(yscrollcommand=sb.set)
        detail = tk.StringVar(value="Datensatz auswählen.")
        tk.Label(body, textvariable=detail, bg="#ffffff", fg="#333", justify="left", anchor="w", wraplength=900).grid(row=2, column=0, sticky="ew", pady=(10,0))
        row_map = {}
        def reload():
            for iid in tree.get_children(""):
                tree.delete(iid)
            row_map.clear()
            with sqlite3.connect(DB_PATH) as con:
                con.row_factory = sqlite3.Row
                rows = con.execute(f"SELECT * FROM {table} ORDER BY id DESC LIMIT 1000").fetchall()
            for row in rows:
                vals = [row[c] if c in row.keys() else "" for c in columns]
                iid = tree.insert("", "end", values=vals)
                row_map[iid] = dict(row)
            self.status.set(f"{title}: {len(rows)} Datensätze angezeigt.")
        def selected_row():
            sel = tree.selection()
            if not sel:
                messagebox.showinfo(title, "Bitte zuerst einen Datensatz auswählen.")
                return None
            return row_map.get(sel[0])
        def on_select(event=None):
            row = selected_row()
            if not row:
                return
            text = "\n".join(f"{headings.get(k,k)}: {row.get(k,'') or ''}" for k in columns if k != "id")
            detail.set(text)
        tree.bind("<<TreeviewSelect>>", on_select)
        def add_record():
            data = self._center_text_dialog(f"{title} – Neu", fields)
            if not data:
                return
            data["bearbeiter"] = self.bearbeiter
            with sqlite3.connect(DB_PATH) as con:
                con.execute(insert_sql, data)
                con.commit()
            reload()
        def edit_record():
            row = selected_row()
            if not row:
                return
            data = self._center_text_dialog(f"{title} – Bearbeiten", fields, row)
            if not data:
                return
            data["id"] = row["id"]
            data["bearbeiter"] = self.bearbeiter
            with sqlite3.connect(DB_PATH) as con:
                con.execute(update_sql, data)
                con.commit()
            reload()
        def delete_record():
            row = selected_row()
            if not row:
                return
            if not messagebox.askyesno(title, "Datensatz wirklich löschen?"):
                return
            with sqlite3.connect(DB_PATH) as con:
                con.execute(f"DELETE FROM {table} WHERE id=?", (row["id"],))
                con.commit()
            reload()
        tk.Button(toolbar, text="Neu", command=add_record, bg="#0b4a86", fg="white", relief="flat", padx=14, pady=7).pack(side="left", padx=(0,8))
        tk.Button(toolbar, text="Bearbeiten", command=edit_record, padx=14, pady=7).pack(side="left", padx=8)
        tk.Button(toolbar, text="Löschen", command=delete_record, padx=14, pady=7).pack(side="left", padx=8)
        tk.Button(toolbar, text="Aktualisieren", command=reload, padx=14, pady=7).pack(side="left", padx=8)
        reload()


    def _show_mitarbeiter_center_page(self, columns, headings, fields, insert_sql, update_sql):
        """Mitarbeiter-Center mit Tabellenansicht, Board-Ansicht und Mitarbeiterprofil-Bereich."""
        self._ensure_center_tables()
        self._ensure_mitarbeiterprofil_table()
        self.clear_page()
        self._page_header(
            "Mitarbeiter-Center",
            "Mitarbeiter, Zuständigkeiten, Vertretungen, Aufgaben und Datenpfade."
        )

        # ── Profil-Leiste oben ──────────────────────────────────────────────────
        profil_bar = tk.Frame(self.page, bg="#f0f4fa", highlightbackground="#c5d3e8", highlightthickness=1)
        profil_bar.grid(row=0, column=0, sticky="ew", padx=18, pady=(0, 8))
        profil_bar.columnconfigure(0, weight=1)

        tk.Label(profil_bar, text="👤  Mitarbeiterprofile", font=(theme.FONT, 11, "bold"),
                 fg="#0b4a86", bg="#f0f4fa").grid(row=0, column=0, sticky="w", padx=14, pady=(10, 4))

        cards_frame = tk.Frame(profil_bar, bg="#f0f4fa")
        cards_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 10))

        def reload_profil_bar():
            for w in cards_frame.winfo_children():
                w.destroy()
            try:
                with sqlite3.connect(DB_PATH) as con:
                    con.row_factory = sqlite3.Row
                    profiles = con.execute(
                        "SELECT login, vorname, nachname, abteilung, position FROM tbl_mitarbeiterprofil ORDER BY nachname, vorname"
                    ).fetchall()
            except Exception:
                profiles = []

            own_login = (self.bearbeiter or "").strip().lower()
            # Eigenes Profil zuerst
            sorted_profiles = sorted(profiles, key=lambda p: (0 if str(p["login"]).lower() == own_login else 1, str(p["nachname"]), str(p["vorname"])))

            for col_idx, p in enumerate(sorted_profiles):
                is_own = str(p["login"]).lower() == own_login
                bg = "#e8f1fb" if is_own else "#ffffff"
                border = "#0b4a86" if is_own else "#d8e2ee"
                card = tk.Frame(cards_frame, bg=bg, highlightbackground=border, highlightthickness=2, cursor="hand2")
                card.grid(row=0, column=col_idx, sticky="ns", padx=6, pady=4, ipadx=8, ipady=6)
                name = f"{p['vorname']} {p['nachname']}".strip() or str(p["login"])
                badge = " ✏️ (Ich)" if is_own else " 👁"
                tk.Label(card, text=f"👤  {name}{badge}", font=(theme.FONT, 10, "bold"),
                         fg="#0b4a86" if is_own else "#333", bg=bg).pack(anchor="w")
                if p["abteilung"] or p["position"]:
                    info = " · ".join(filter(None, [str(p["abteilung"] or ""), str(p["position"] or "")]))
                    tk.Label(card, text=info, font=(theme.FONT, 8), fg="#666", bg=bg).pack(anchor="w")
                login_val = str(p["login"])
                readonly = not is_own
                card.bind("<Button-1>", lambda e, l=login_val, r=readonly: self.show_mitarbeiterprofil_dialog(l, r))
                for child in card.winfo_children():
                    child.bind("<Button-1>", lambda e, l=login_val, r=readonly: self.show_mitarbeiterprofil_dialog(l, r))

            # Eigenes Profil fehlt noch? → "Profil anlegen"-Button
            own_exists = any(str(p["login"]).lower() == own_login for p in sorted_profiles)
            if not own_exists and not self._is_admin_login():
                card = tk.Frame(cards_frame, bg="#fff8e1", highlightbackground="#efd39a", highlightthickness=2, cursor="hand2")
                card.grid(row=0, column=len(sorted_profiles), sticky="ns", padx=6, pady=4, ipadx=8, ipady=6)
                tk.Label(card, text="➕  Mein Profil anlegen", font=(theme.FONT, 10, "bold"), fg="#8b5a00", bg="#fff8e1").pack(anchor="w")
                card.bind("<Button-1>", lambda e: self._open_mitarbeiterprofil_pflicht_dialog())

        reload_profil_bar()
        

        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure(0, weight=1)
        body.rowconfigure(1, weight=1)

        modebar = tk.Frame(body, bg="#ffffff")
        modebar.grid(row=0, column=0, sticky="ew", pady=(0, 8))

        content = tk.Frame(body, bg="#ffffff")
        content.grid(row=1, column=0, sticky="nsew")
        content.columnconfigure(0, weight=1)
        content.rowconfigure(0, weight=1)

        def clear_content():
            for w in content.winfo_children():
                w.destroy()

        def show_table():
            clear_content()
            # Die bestehende Tabellenlogik in den Content-Bereich eingebettet.
            frame = tk.Frame(content, bg="#ffffff")
            frame.grid(row=0, column=0, sticky="nsew")
            frame.columnconfigure(0, weight=1)
            frame.rowconfigure(1, weight=1)

            toolbar = tk.Frame(frame, bg="#ffffff")
            toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 8))
            tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode="browse")
            tree.grid(row=1, column=0, sticky="nsew")
            for col in columns:
                tree.heading(col, text=headings.get(col, col))
                tree.column(col, width=120 if col != "onedrive_pfad" else 260, anchor="w")
            try:
                self._make_tree_sortable(tree, columns, headings)
            except Exception:
                pass
            sb = tk.Scrollbar(frame, orient="vertical", command=tree.yview)
            sb.grid(row=1, column=0, sticky="nse")
            tree.configure(yscrollcommand=sb.set)
            detail = tk.StringVar(value="Datensatz auswählen.")
            tk.Label(frame, textvariable=detail, bg="#ffffff", fg="#333", justify="left", anchor="w", wraplength=900).grid(row=2, column=0, sticky="ew", pady=(10,0))
            row_map = {}
            def reload():
                for iid in tree.get_children(""):
                    tree.delete(iid)
                row_map.clear()
                with sqlite3.connect(DB_PATH) as con:
                    con.row_factory = sqlite3.Row
                    rows = con.execute("SELECT * FROM tbl_mitarbeiter ORDER BY id DESC LIMIT 1000").fetchall()
                for row in rows:
                    vals = [row[c] if c in row.keys() else "" for c in columns]
                    iid = tree.insert("", "end", values=vals)
                    row_map[iid] = dict(row)
                self.status.set(f"Mitarbeiter-Center: {len(rows)} Datensätze angezeigt.")
            def selected_row():
                sel = tree.selection()
                if not sel:
                    messagebox.showinfo("Mitarbeiter-Center", "Bitte zuerst einen Datensatz auswählen.")
                    return None
                return row_map.get(sel[0])
            def on_select(event=None):
                row = selected_row()
                if not row:
                    return
                text = "\n".join(f"{headings.get(k,k)}: {row.get(k,'') or ''}" for k in columns if k != "id")
                if row.get("aufgaben"):
                    text += f"\n\nAufgaben:\n{row.get('aufgaben')}"
                if row.get("notizen"):
                    text += f"\n\nNotizen:\n{row.get('notizen')}"
                detail.set(text)
            tree.bind("<<TreeviewSelect>>", on_select)
            def add_record():
                data = self._center_text_dialog("Mitarbeiter-Center – Neu", fields)
                if not data:
                    return
                data["bearbeiter"] = self.bearbeiter
                with sqlite3.connect(DB_PATH) as con:
                    con.execute(insert_sql, data)
                    con.commit()
                reload()
            def edit_record():
                row = selected_row()
                if not row:
                    return
                data = self._center_text_dialog("Mitarbeiter-Center – Bearbeiten", fields, row)
                if not data:
                    return
                data["id"] = row["id"]
                data["bearbeiter"] = self.bearbeiter
                with sqlite3.connect(DB_PATH) as con:
                    con.execute(update_sql, data)
                    con.commit()
                reload()
            def delete_record():
                row = selected_row()
                if not row:
                    return
                if not messagebox.askyesno("Mitarbeiter-Center", "Mitarbeiter wirklich löschen?"):
                    return
                with sqlite3.connect(DB_PATH) as con:
                    con.execute("DELETE FROM tbl_mitarbeiter WHERE id=?", (row["id"],))
                    con.commit()
                reload()
            def open_path():
                row = selected_row()
                if not row:
                    return
                path = str(row.get("onedrive_pfad") or "").strip()
                if not path:
                    messagebox.showinfo("Mitarbeiter-Center", "Bei diesem Mitarbeiter ist kein OneDrive-/Datenpfad hinterlegt.")
                    return
                _open_folder(Path(path))
            tk.Button(toolbar, text="Neu", command=add_record, bg="#0b4a86", fg="white", relief="flat", padx=14, pady=7).pack(side="left", padx=(0,8))
            tk.Button(toolbar, text="Bearbeiten", command=edit_record, padx=14, pady=7).pack(side="left", padx=8)
            tk.Button(toolbar, text="Löschen", command=delete_record, padx=14, pady=7).pack(side="left", padx=8)
            tk.Button(toolbar, text="Datenpfad öffnen", command=open_path, padx=14, pady=7).pack(side="left", padx=8)
            tk.Button(toolbar, text="Aktualisieren", command=reload, padx=14, pady=7).pack(side="left", padx=8)
            reload()

        def show_board():
            clear_content()
            board_frame = tk.Frame(content, bg="#ffffff")
            board_frame.grid(row=0, column=0, sticky="nsew")
            board_frame.columnconfigure(0, weight=1)
            board_frame.rowconfigure(1, weight=1)
            info = tk.Label(
                board_frame,
                text="Post-it-Ansicht: Karten können mit gedrückter Maustaste verschoben werden. Positionen werden gespeichert. Doppelklick öffnet den Datenpfad.",
                bg="#ffffff",
                fg="#333",
                anchor="w",
            )
            info.grid(row=0, column=0, sticky="ew", pady=(0, 8))
            canvas = tk.Canvas(board_frame, bg="#f8fbff", highlightbackground="#d8e2ee", highlightthickness=1)
            canvas.grid(row=1, column=0, sticky="nsew")
            hbar = tk.Scrollbar(board_frame, orient="horizontal", command=canvas.xview)
            hbar.grid(row=2, column=0, sticky="ew")
            vbar = tk.Scrollbar(board_frame, orient="vertical", command=canvas.yview)
            vbar.grid(row=1, column=1, sticky="ns")
            canvas.configure(xscrollcommand=hbar.set, yscrollcommand=vbar.set, scrollregion=(0, 0, 1800, 1200))

            with sqlite3.connect(DB_PATH) as con:
                con.row_factory = sqlite3.Row
                rows = con.execute("SELECT * FROM tbl_mitarbeiter WHERE COALESCE(aktiv,1)=1 ORDER BY id").fetchall()

            card_items = {}
            drag = {"id": None, "x": 0, "y": 0}

            def card_text(row):
                name = " ".join([str(row["vorname"] or "").strip(), str(row["name"] or "").strip()]).strip() or f"Mitarbeiter {row['id']}"
                lines = [name]
                if row["position"]:
                    lines.append(str(row["position"]))
                if row["abteilung"]:
                    lines.append(str(row["abteilung"]))
                if row["vertretung_1"]:
                    lines.append(f"Vertretung: {row['vertretung_1']}")
                if row["vertretung_2"]:
                    lines.append(f"2. Vertretung: {row['vertretung_2']}")
                if row["aufgaben"]:
                    aufg = str(row["aufgaben"]).replace("\n", "; ")[:70]
                    lines.append(f"Aufgaben: {aufg}")
                if row["onedrive_pfad"]:
                    lines.append("📂 Datenpfad")
                return "\n".join(lines)

            def save_position(emp_id, x, y):
                try:
                    with sqlite3.connect(DB_PATH) as con:
                        con.execute("UPDATE tbl_mitarbeiter SET board_x=?, board_y=?, geaendert_am=CURRENT_TIMESTAMP, bearbeiter=? WHERE id=?", (int(x), int(y), self.bearbeiter, int(emp_id)))
                        con.commit()
                except Exception:
                    pass

            def make_card(row, index):
                emp_id = int(row["id"])
                x = row["board_x"] if "board_x" in row.keys() and row["board_x"] is not None else 40 + (index % 4) * 280
                y = row["board_y"] if "board_y" in row.keys() and row["board_y"] is not None else 40 + (index // 4) * 190
                w, h = 240, 145
                rect = canvas.create_rectangle(x, y, x+w, y+h, fill="#fff7c2", outline="#d6bd48", width=2)
                txt = canvas.create_text(x+12, y+12, text=card_text(row), anchor="nw", width=w-24, font=(theme.FONT, 10))
                card_items[emp_id] = (rect, txt, row)
                for item in (rect, txt):
                    canvas.tag_bind(item, "<ButtonPress-1>", lambda e, eid=emp_id: start_drag(e, eid))
                    canvas.tag_bind(item, "<B1-Motion>", lambda e, eid=emp_id: move_drag(e, eid))
                    canvas.tag_bind(item, "<ButtonRelease-1>", lambda e, eid=emp_id: end_drag(e, eid))
                    canvas.tag_bind(item, "<Double-Button-1>", lambda e, eid=emp_id: open_card_path(eid))

            def start_drag(event, emp_id):
                drag["id"] = emp_id
                drag["x"] = canvas.canvasx(event.x)
                drag["y"] = canvas.canvasy(event.y)

            def move_drag(event, emp_id):
                if drag["id"] != emp_id:
                    return
                x = canvas.canvasx(event.x)
                y = canvas.canvasy(event.y)
                dx = x - drag["x"]
                dy = y - drag["y"]
                rect, txt, _row = card_items[emp_id]
                canvas.move(rect, dx, dy)
                canvas.move(txt, dx, dy)
                drag["x"] = x
                drag["y"] = y

            def end_drag(event, emp_id):
                if drag["id"] != emp_id:
                    return
                rect, _txt, _row = card_items[emp_id]
                x1, y1, _x2, _y2 = canvas.coords(rect)
                save_position(emp_id, x1, y1)
                drag["id"] = None
                self.status.set("Mitarbeiter-Board: Position gespeichert.")

            def open_card_path(emp_id):
                row = card_items[emp_id][2]
                path = str(row["onedrive_pfad"] or "").strip()
                if path:
                    _open_folder(Path(path))
                else:
                    messagebox.showinfo("Mitarbeiter-Board", "Für diese Karte ist kein Datenpfad hinterlegt.")

            for idx, row in enumerate(rows):
                make_card(row, idx)
            self.status.set(f"Mitarbeiter-Board: {len(rows)} Karten angezeigt.")

        tk.Button(modebar, text="Tabelle", command=show_table, bg="#0b4a86", fg="white", relief="flat", padx=16, pady=8).pack(side="left", padx=(0,8))
        tk.Button(modebar, text="Board / Post-it", command=show_board, bg="#6b4fb3", fg="white", relief="flat", padx=16, pady=8).pack(side="left", padx=8)
        show_table()

    def show_mitarbeiter_center(self):
        """Mitarbeiter-Center: alle Mitarbeiter als Kacheln. Eigene Karte bearbeitbar, fremde nur ansehen."""
        self._ensure_center_tables()
        self._ensure_mitarbeiterprofil_table()
        self.clear_page()
        self._mitarbeiter_icon = theme.load_icon(ASSETS_DIR / "Kunden.ico", 40)
        self._page_header("Mitarbeiter", "Alle Mitarbeiter auf einen Blick. Eigene Karte bearbeitbar.",
                          icon=self._mitarbeiter_icon)

        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure(0, weight=1)
        body.rowconfigure(1, weight=1)

        # Toolbar
        toolbar = tk.Frame(body, bg="#ffffff")
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 10))

        own_login = (self.bearbeiter or "").strip().lower()

        def reload():
            for w in cards_outer.winfo_children():
                w.destroy()
            _render_cards()

        # Karten-Bereich scrollbar
        canvas_m = tk.Canvas(body, bg="#ffffff", highlightthickness=0)
        vsb_m = tk.Scrollbar(body, orient="vertical", command=canvas_m.yview)
        canvas_m.configure(yscrollcommand=vsb_m.set)
        canvas_m.grid(row=1, column=0, sticky="nsew")
        vsb_m.grid(row=1, column=1, sticky="ns")

        cards_outer = tk.Frame(canvas_m, bg="#ffffff")
        win_id_m = canvas_m.create_window((0, 0), window=cards_outer, anchor="nw")
        cards_outer.bind("<Configure>", lambda e: canvas_m.configure(scrollregion=canvas_m.bbox("all")))
        canvas_m.bind("<Configure>", lambda e: canvas_m.itemconfig(win_id_m, width=e.width))

        def _render_cards():
            # Lade Mitarbeiter: erst tbl_mitarbeiterprofil, dann tbl_mitarbeiter als Fallback
            mitarbeiter = []
            try:
                with sqlite3.connect(DB_PATH) as con:
                    con.row_factory = sqlite3.Row
                    # Profile aus tbl_mitarbeiterprofil
                    profil_rows = con.execute(
                        "SELECT login, vorname, nachname, telefon, mobil, email, abteilung, position FROM tbl_mitarbeiterprofil ORDER BY nachname, vorname"
                    ).fetchall()
                    for p in profil_rows:
                        name = f"{p['vorname']} {p['nachname']}".strip()
                        mitarbeiter.append({
                            "name": name or str(p["login"]),
                            "login": str(p["login"]),
                            "abteilung": str(p["abteilung"] or ""),
                            "position": str(p["position"] or ""),
                            "telefon": str(p["telefon"] or ""),
                            "mobil": str(p["mobil"] or ""),
                            "email": str(p["email"] or ""),
                            "quelle": "profil",
                        })
                    # Ergänze aus tbl_mitarbeiter (wenn kein Profil-Duplikat)
                    ma_rows = con.execute("SELECT * FROM tbl_mitarbeiter ORDER BY name, vorname").fetchall()
                    existing_names = {m["name"].lower() for m in mitarbeiter}
                    for row in ma_rows:
                        full = f"{row['vorname'] or ''} {row['name'] or ''}".strip()
                        if full.lower() not in existing_names:
                            mitarbeiter.append({
                                "name": full or f"MA #{row['id']}",
                                "login": "",
                                "abteilung": str(row["abteilung"] or ""),
                                "position": str(row["position"] or ""),
                                "telefon": str(row["telefon"] or ""),
                                "mobil": str(row["mobil"] or ""),
                                "email": str(row["email"] or ""),
                                "vertretung": str(row["vertretung_1"] or ""),
                                "aufgaben": str(row["aufgaben"] or ""),
                                "quelle": "mitarbeiter",
                                "id": row["id"],
                            })
            except Exception:
                pass

            # Eigenen Login ganz vorne
            mitarbeiter.sort(key=lambda m: (0 if m.get("login","").lower() == own_login else 1, m["name"]))

            if not mitarbeiter:
                tk.Label(cards_outer, text="Noch keine Mitarbeiterdaten vorhanden.\nMitarbeiterprofile k\u00f6nnen unter 'Mein Profil' angelegt werden.",
                         bg="#ffffff", fg="#888", font=(theme.FONT, 10), justify="center").pack(pady=40)
                return

            cols_count = 4
            for idx, m in enumerate(mitarbeiter):
                r = idx // cols_count
                c = idx % cols_count
                is_own = m.get("login", "").lower() == own_login
                bg = "#e8f1fb" if is_own else "#f8fbff"
                border = "#0b4a86" if is_own else "#d8e2ee"
                card = tk.Frame(cards_outer, bg=bg, highlightbackground=border, highlightthickness=2)
                card.grid(row=r, column=c, sticky="nsew", padx=8, pady=8, ipadx=8, ipady=8)
                cards_outer.columnconfigure(c, weight=1)

                # Initialen-Avatar
                initials = "".join(p[0].upper() for p in m["name"].split()[:2] if p)
                avatar_bg = "#0b4a86" if is_own else "#6b4fb3"
                av = tk.Label(card, text=initials or "?", font=(theme.FONT, 20, "bold"),
                              fg="white", bg=avatar_bg, width=3, relief="flat")
                av.pack(pady=(12, 6))

                tk.Label(card, text=m["name"], font=(theme.FONT, 11, "bold"), fg="#123", bg=bg).pack()
                if is_own:
                    tk.Label(card, text="(Ich)", font=(theme.FONT, 8), fg="#0b4a86", bg=bg).pack()

                # Daten anzeigen
                info_lines = []
                if m.get("position"):   info_lines.append(f"🏷  {m['position']}")
                if m.get("abteilung"):  info_lines.append(f"🏢  {m['abteilung']}")
                if m.get("telefon"):    info_lines.append(f"📞  {m['telefon']}")
                if m.get("mobil"):      info_lines.append(f"📱  {m['mobil']}")
                if m.get("email"):      info_lines.append(f"✉  {m['email']}")
                if m.get("vertretung"):info_lines.append(f"🔄  {m['vertretung']}")

                for line in info_lines[:5]:
                    tk.Label(card, text=line, font=(theme.FONT, 9), fg="#444", bg=bg, anchor="w").pack(anchor="w", padx=4)

                # Button
                if m.get("quelle") == "mitarbeiter" and m.get("id"):
                    # Manuell gepflegter Mitarbeiter (id-basiert): Detail-Dialog mit
                    # Zusatzfeldern + Vorgesetzten (Phase 1, Organigramm-Grundlage).
                    emp_id = int(m["id"])
                    tk.Button(card, text="✏️  Bearbeiten",
                              command=lambda eid=emp_id: self._mitarbeiter_detail_dialog(eid, on_saved=reload),
                              bg="#0b4a86", fg="white", relief="flat", font=(theme.FONT, 9, "bold"),
                              padx=10, pady=4).pack(fill="x", padx=8, pady=(8, 4))
                elif is_own or self._is_admin_login():
                    btn_text = "✏️  Bearbeiten" if is_own else "✏️  Admin"
                    login_val = m.get("login") or ""
                    tk.Button(card, text=btn_text,
                              command=lambda l=login_val: self.show_mitarbeiterprofil_dialog(l, readonly=False),
                              bg="#0b4a86", fg="white", relief="flat", font=(theme.FONT, 9, "bold"),
                              padx=10, pady=4).pack(fill="x", padx=8, pady=(8, 4))
                else:
                    login_val = m.get("login") or ""
                    tk.Button(card, text="👁  Ansehen",
                              command=lambda l=login_val: self.show_mitarbeiterprofil_dialog(l, readonly=True),
                              bg="#888", fg="white", relief="flat", font=(theme.FONT, 9),
                              padx=10, pady=4).pack(fill="x", padx=8, pady=(8, 4))

        _render_cards()

        # Toolbar-Buttons
        def neuer_eintrag():
            # Mitarbeiter in tbl_mitarbeiter anlegen (für Nicht-Profilnutzer)
            fields = [("vorname","Vorname","entry"),("name","Nachname","entry"),
                      ("abteilung","Abteilung","entry"),("position","Position","entry"),
                      ("telefon","Telefon","entry"),("mobil","Mobil","entry"),
                      ("email","E-Mail","entry"),("vertretung_1","Vertretung","entry")]
            data = self._center_text_dialog("Neuer Mitarbeiter", fields)
            if not data:
                return
            data["bearbeiter"] = self.bearbeiter
            with sqlite3.connect(DB_PATH) as con:
                cur = con.execute("""INSERT INTO tbl_mitarbeiter(vorname,name,abteilung,position,telefon,mobil,email,vertretung_1,bearbeiter)
                    VALUES(:vorname,:name,:abteilung,:position,:telefon,:mobil,:email,:vertretung_1,:bearbeiter)""", data)
                new_id = cur.lastrowid
                con.commit()
            # Direkt in den Detail-Dialog: Zusatzfelder + Vorgesetzte pflegen.
            self._mitarbeiter_detail_dialog(int(new_id), on_saved=reload)
            reload()

        tk.Button(toolbar, text="➕ Mitarbeiter anlegen", command=neuer_eintrag,
                  bg="#0b4a86", fg="white", relief="flat", padx=14, pady=6).pack(side="left", padx=(0, 8))
        tk.Button(toolbar, text="🗂 Felder verwalten",
                  command=lambda: self._mitarbeiter_felder_dialog(on_saved=reload),
                  padx=12, pady=6).pack(side="left", padx=(0, 8))
        tk.Button(toolbar, text="🔄 Aktualisieren", command=reload, padx=12, pady=6).pack(side="left")
        self.status.set("Mitarbeiter-Center bereit.")

    # ── Phase 1: Zusatzfelder-Verwaltung + Mitarbeiter-Detail (Vorgesetzte) ──
    def _mitarbeiter_felder_dialog(self, on_saved=None):
        """Frei definierbare Zusatzfelder anlegen/bearbeiten/loeschen."""
        self._ensure_mitarbeiter_phase1_tables()
        win = tk.Toplevel(self)
        win.title("Zusatzfelder verwalten")
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()
        win.geometry("640x520")
        win.columnconfigure(0, weight=1)
        win.rowconfigure(1, weight=1)
        tk.Label(win, text="🗂  Zusatzfelder", font=(theme.FONT, 16, "bold"),
                 fg="#0b4a86", bg="#f5f7fb").grid(row=0, column=0, sticky="w", padx=18, pady=(16, 8))

        cols = ("feld_name", "feld_typ", "optionen", "reihenfolge")
        heads = {"feld_name": "Feldname", "feld_typ": "Typ", "optionen": "Auswahl-Optionen", "reihenfolge": "Reihenfolge"}
        tree = ttk.Treeview(win, columns=cols, show="headings", selectmode="browse")
        for c in cols:
            tree.heading(c, text=heads[c])
            tree.column(c, width=140, anchor="w")
        tree.grid(row=1, column=0, sticky="nsew", padx=18)
        row_map = {}

        def reload_fields():
            for iid in tree.get_children(""):
                tree.delete(iid)
            row_map.clear()
            with sqlite3.connect(DB_PATH) as con:
                con.row_factory = sqlite3.Row
                rows = con.execute(
                    "SELECT * FROM tbl_mitarbeiter_feld WHERE COALESCE(aktiv,1)=1 ORDER BY reihenfolge, id"
                ).fetchall()
            for row in rows:
                iid = tree.insert("", "end", values=[row[c] for c in cols])
                row_map[iid] = dict(row)

        def feld_dialog(initial=None):
            initial = initial or {}
            d = tk.Toplevel(win)
            d.title("Feld bearbeiten" if initial else "Neues Feld")
            d.configure(bg="#f5f7fb")
            d.transient(win)
            d.grab_set()
            d.geometry("440x340")
            d.columnconfigure(1, weight=1)
            tk.Label(d, text="Feldname", bg="#f5f7fb", fg="#0b4a86", font=(theme.FONT, 10, "bold")).grid(row=0, column=0, sticky="w", padx=14, pady=10)
            name_var = tk.StringVar(value=str(initial.get("feld_name", "") or ""))
            tk.Entry(d, textvariable=name_var).grid(row=0, column=1, sticky="ew", padx=14, pady=10)
            tk.Label(d, text="Typ", bg="#f5f7fb", fg="#0b4a86", font=(theme.FONT, 10, "bold")).grid(row=1, column=0, sticky="w", padx=14, pady=10)
            typ_var = tk.StringVar(value=str(initial.get("feld_typ", "Text") or "Text"))
            ttk.Combobox(d, textvariable=typ_var, values=("Text", "Zahl", "Datum", "Auswahl"),
                         state="readonly").grid(row=1, column=1, sticky="ew", padx=14, pady=10)
            tk.Label(d, text="Optionen (nur bei\nAuswahl, mit ; trennen)", bg="#f5f7fb", fg="#0b4a86", font=(theme.FONT, 9)).grid(row=2, column=0, sticky="w", padx=14, pady=10)
            opt_var = tk.StringVar(value=str(initial.get("optionen", "") or ""))
            tk.Entry(d, textvariable=opt_var).grid(row=2, column=1, sticky="ew", padx=14, pady=10)
            tk.Label(d, text="Reihenfolge", bg="#f5f7fb", fg="#0b4a86", font=(theme.FONT, 10, "bold")).grid(row=3, column=0, sticky="w", padx=14, pady=10)
            reihen_var = tk.StringVar(value=str(initial.get("reihenfolge", "0") or "0"))
            tk.Entry(d, textvariable=reihen_var).grid(row=3, column=1, sticky="ew", padx=14, pady=10)

            def save():
                name = name_var.get().strip()
                if not name:
                    messagebox.showinfo("Zusatzfeld", "Bitte einen Feldnamen eingeben.")
                    return
                try:
                    reihen = int(reihen_var.get().strip() or "0")
                except ValueError:
                    reihen = 0
                with sqlite3.connect(DB_PATH) as con:
                    if initial.get("id"):
                        con.execute("UPDATE tbl_mitarbeiter_feld SET feld_name=?, feld_typ=?, optionen=?, reihenfolge=? WHERE id=?",
                                    (name, typ_var.get(), opt_var.get().strip(), reihen, initial["id"]))
                    else:
                        con.execute("INSERT INTO tbl_mitarbeiter_feld(feld_name, feld_typ, optionen, reihenfolge) VALUES(?,?,?,?)",
                                    (name, typ_var.get(), opt_var.get().strip(), reihen))
                    con.commit()
                d.destroy()
                reload_fields()

            bar = tk.Frame(d, bg="#f5f7fb")
            bar.grid(row=4, column=0, columnspan=2, sticky="ew", padx=14, pady=14)
            tk.Button(bar, text="Abbrechen", command=d.destroy, padx=14, pady=7).pack(side="right", padx=(8, 0))
            tk.Button(bar, text="Speichern", command=save, bg="#0b4a86", fg="white", relief="flat", padx=18, pady=8).pack(side="right")
            win.wait_window(d)

        def add_field():
            feld_dialog()

        def edit_field():
            sel = tree.selection()
            if not sel:
                messagebox.showinfo("Zusatzfeld", "Bitte zuerst ein Feld auswählen.")
                return
            feld_dialog(row_map.get(sel[0]))

        def delete_field():
            sel = tree.selection()
            if not sel:
                messagebox.showinfo("Zusatzfeld", "Bitte zuerst ein Feld auswählen.")
                return
            row = row_map.get(sel[0])
            if not messagebox.askyesno("Zusatzfeld", f"Feld '{row.get('feld_name')}' löschen?\n(Bereits erfasste Werte bleiben in der Datenbank, werden aber nicht mehr angezeigt.)"):
                return
            with sqlite3.connect(DB_PATH) as con:
                con.execute("UPDATE tbl_mitarbeiter_feld SET aktiv=0 WHERE id=?", (row["id"],))
                con.commit()
            reload_fields()

        bar = tk.Frame(win, bg="#f5f7fb")
        bar.grid(row=2, column=0, sticky="ew", padx=18, pady=12)
        tk.Button(bar, text="➕ Neu", command=add_field, bg="#0b4a86", fg="white", relief="flat", padx=14, pady=7).pack(side="left", padx=(0, 8))
        tk.Button(bar, text="Bearbeiten", command=edit_field, padx=14, pady=7).pack(side="left", padx=8)
        tk.Button(bar, text="Löschen", command=delete_field, padx=14, pady=7).pack(side="left", padx=8)
        tk.Button(bar, text="Schließen", command=win.destroy, padx=14, pady=7).pack(side="right")
        reload_fields()
        self.wait_window(win)
        if callable(on_saved):
            on_saved()

    def _mitarbeiter_detail_dialog(self, emp_id, on_saved=None):
        """Detail-Dialog für einen tbl_mitarbeiter-Datensatz mit Reitern:
        Stammdaten · Zusatzfelder · Vorgesetzte (Organigramm-Grundlage)."""
        self._ensure_mitarbeiter_phase1_tables()
        with sqlite3.connect(DB_PATH) as con:
            con.row_factory = sqlite3.Row
            row = con.execute("SELECT * FROM tbl_mitarbeiter WHERE id=?", (emp_id,)).fetchone()
        if not row:
            messagebox.showinfo("Mitarbeiter", "Datensatz nicht gefunden.")
            return
        row = dict(row)
        name_disp = f"{row.get('vorname','') or ''} {row.get('name','') or ''}".strip() or f"Mitarbeiter #{emp_id}"

        win = tk.Toplevel(self)
        win.title(f"Mitarbeiter: {name_disp}")
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()
        win.geometry("720x600")
        win.columnconfigure(0, weight=1)
        win.rowconfigure(1, weight=1)
        tk.Label(win, text=f"👤  {name_disp}", font=(theme.FONT, 16, "bold"),
                 fg="#0b4a86", bg="#f5f7fb").grid(row=0, column=0, sticky="w", padx=18, pady=(16, 8))

        nb = ttk.Notebook(win)
        nb.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 8))

        # ── Reiter 1: Stammdaten ────────────────────────────────────────────
        tab_stamm = tk.Frame(nb, bg="#ffffff")
        nb.add(tab_stamm, text="Stammdaten")
        tab_stamm.columnconfigure(1, weight=1)
        stamm_fields = [("vorname", "Vorname"), ("name", "Nachname"), ("abteilung", "Abteilung"),
                        ("position", "Position"), ("telefon", "Telefon"), ("mobil", "Mobil"),
                        ("email", "E-Mail"), ("onedrive_pfad", "Datenpfad")]
        stamm_vars = {}
        for r, (key, label) in enumerate(stamm_fields):
            tk.Label(tab_stamm, text=label, bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 10, "bold")).grid(row=r, column=0, sticky="w", padx=14, pady=7)
            var = tk.StringVar(value=str(row.get(key, "") or ""))
            stamm_vars[key] = var
            tk.Entry(tab_stamm, textvariable=var).grid(row=r, column=1, sticky="ew", padx=14, pady=7)

        # ── Reiter 2: Zusatzfelder ──────────────────────────────────────────
        tab_extra = tk.Frame(nb, bg="#ffffff")
        nb.add(tab_extra, text="Zusatzfelder")
        tab_extra.columnconfigure(1, weight=1)
        with sqlite3.connect(DB_PATH) as con:
            con.row_factory = sqlite3.Row
            felder = con.execute("SELECT * FROM tbl_mitarbeiter_feld WHERE COALESCE(aktiv,1)=1 ORDER BY reihenfolge, id").fetchall()
            werte = {w["feld_id"]: w["wert"] for w in con.execute(
                "SELECT feld_id, wert FROM tbl_mitarbeiter_wert WHERE mitarbeiter_id=?", (emp_id,)).fetchall()}
        extra_vars = {}
        if not felder:
            tk.Label(tab_extra, text="Noch keine Zusatzfelder definiert.\nÜber 'Felder verwalten' im Mitarbeiter-Center anlegen.",
                     bg="#ffffff", fg="#888", font=(theme.FONT, 10), justify="left").grid(row=0, column=0, columnspan=2, sticky="w", padx=14, pady=20)
        else:
            for r, f in enumerate(felder):
                fid = f["id"]
                tk.Label(tab_extra, text=str(f["feld_name"]), bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 10, "bold")).grid(row=r, column=0, sticky="w", padx=14, pady=7)
                var = tk.StringVar(value=str(werte.get(fid, "") or ""))
                extra_vars[fid] = var
                if f["feld_typ"] == "Auswahl":
                    opts = [o.strip() for o in str(f["optionen"] or "").split(";") if o.strip()]
                    ttk.Combobox(tab_extra, textvariable=var, values=opts, state="readonly").grid(row=r, column=1, sticky="ew", padx=14, pady=7)
                else:
                    tk.Entry(tab_extra, textvariable=var).grid(row=r, column=1, sticky="ew", padx=14, pady=7)

        # ── Reiter 3: Vorgesetzte ───────────────────────────────────────────
        tab_vg = tk.Frame(nb, bg="#ffffff")
        nb.add(tab_vg, text="Vorgesetzte")
        tab_vg.columnconfigure(0, weight=1)
        tab_vg.rowconfigure(1, weight=1)
        tk.Label(tab_vg, text="Mehrere Vorgesetzte möglich. Genau einer sollte als „primär“ markiert sein – daraus entsteht das Organigramm.",
                 bg="#ffffff", fg="#555", font=(theme.FONT, 9), wraplength=640, justify="left").grid(row=0, column=0, sticky="ew", padx=14, pady=(10, 6))
        vg_cols = ("vorgesetzter", "art", "primaer")
        vg_heads = {"vorgesetzter": "Vorgesetzter", "art": "Art", "primaer": "Primär"}
        vg_tree = ttk.Treeview(tab_vg, columns=vg_cols, show="headings", selectmode="browse", height=6)
        for c in vg_cols:
            vg_tree.heading(c, text=vg_heads[c])
            vg_tree.column(c, width=180 if c == "vorgesetzter" else 110, anchor="w")
        vg_tree.grid(row=1, column=0, sticky="nsew", padx=14)
        vg_map = {}

        def load_kandidaten():
            with sqlite3.connect(DB_PATH) as con:
                con.row_factory = sqlite3.Row
                rows = con.execute("SELECT id, vorname, name FROM tbl_mitarbeiter WHERE id<>? ORDER BY name, vorname", (emp_id,)).fetchall()
            return {f"{(r['vorname'] or '').strip()} {(r['name'] or '').strip()}".strip() or f"#{r['id']}": r["id"] for r in rows}

        def reload_vg():
            for iid in vg_tree.get_children(""):
                vg_tree.delete(iid)
            vg_map.clear()
            with sqlite3.connect(DB_PATH) as con:
                con.row_factory = sqlite3.Row
                rows = con.execute("""
                    SELECT v.id, v.art, v.ist_primaer, v.vorgesetzter_id,
                           m.vorname, m.name
                    FROM tbl_mitarbeiter_vorgesetzter v
                    LEFT JOIN tbl_mitarbeiter m ON m.id = v.vorgesetzter_id
                    WHERE v.mitarbeiter_id=? ORDER BY v.ist_primaer DESC, v.id""", (emp_id,)).fetchall()
            for row_v in rows:
                vg_name = f"{(row_v['vorname'] or '').strip()} {(row_v['name'] or '').strip()}".strip() or f"#{row_v['vorgesetzter_id']}"
                iid = vg_tree.insert("", "end", values=[vg_name, row_v["art"], "★" if row_v["ist_primaer"] else ""])
                vg_map[iid] = dict(row_v)

        def vg_add():
            kandidaten = load_kandidaten()
            if not kandidaten:
                messagebox.showinfo("Vorgesetzte", "Es gibt noch keine anderen Mitarbeiter, die als Vorgesetzte zugeordnet werden könnten.")
                return
            d = tk.Toplevel(win)
            d.title("Vorgesetzten zuordnen")
            d.configure(bg="#f5f7fb")
            d.transient(win)
            d.grab_set()
            d.geometry("440x280")
            d.columnconfigure(1, weight=1)
            tk.Label(d, text="Vorgesetzter", bg="#f5f7fb", fg="#0b4a86", font=(theme.FONT, 10, "bold")).grid(row=0, column=0, sticky="w", padx=14, pady=10)
            vg_var = tk.StringVar()
            ttk.Combobox(d, textvariable=vg_var, values=list(kandidaten.keys()), state="readonly").grid(row=0, column=1, sticky="ew", padx=14, pady=10)
            tk.Label(d, text="Art", bg="#f5f7fb", fg="#0b4a86", font=(theme.FONT, 10, "bold")).grid(row=1, column=0, sticky="w", padx=14, pady=10)
            art_var = tk.StringVar(value=self.ARTEN_VORGESETZTER[0])
            ttk.Combobox(d, textvariable=art_var, values=list(self.ARTEN_VORGESETZTER), state="readonly").grid(row=1, column=1, sticky="ew", padx=14, pady=10)
            primaer_var = tk.BooleanVar(value=False)
            tk.Checkbutton(d, text="Als primär markieren (bildet das Organigramm)", variable=primaer_var, bg="#f5f7fb").grid(row=2, column=0, columnspan=2, sticky="w", padx=14, pady=10)

            def save():
                sel_name = vg_var.get().strip()
                if not sel_name or sel_name not in kandidaten:
                    messagebox.showinfo("Vorgesetzte", "Bitte einen Vorgesetzten auswählen.")
                    return
                vid = kandidaten[sel_name]
                with sqlite3.connect(DB_PATH) as con:
                    if primaer_var.get():
                        con.execute("UPDATE tbl_mitarbeiter_vorgesetzter SET ist_primaer=0 WHERE mitarbeiter_id=?", (emp_id,))
                    con.execute("INSERT INTO tbl_mitarbeiter_vorgesetzter(mitarbeiter_id, vorgesetzter_id, art, ist_primaer) VALUES(?,?,?,?)",
                                (emp_id, vid, art_var.get(), 1 if primaer_var.get() else 0))
                    con.commit()
                d.destroy()
                reload_vg()

            bar = tk.Frame(d, bg="#f5f7fb")
            bar.grid(row=3, column=0, columnspan=2, sticky="ew", padx=14, pady=14)
            tk.Button(bar, text="Abbrechen", command=d.destroy, padx=14, pady=7).pack(side="right", padx=(8, 0))
            tk.Button(bar, text="Speichern", command=save, bg="#0b4a86", fg="white", relief="flat", padx=18, pady=8).pack(side="right")
            win.wait_window(d)

        def vg_set_primaer():
            sel = vg_tree.selection()
            if not sel:
                messagebox.showinfo("Vorgesetzte", "Bitte zuerst eine Zuordnung auswählen.")
                return
            rel = vg_map.get(sel[0])
            with sqlite3.connect(DB_PATH) as con:
                con.execute("UPDATE tbl_mitarbeiter_vorgesetzter SET ist_primaer=0 WHERE mitarbeiter_id=?", (emp_id,))
                con.execute("UPDATE tbl_mitarbeiter_vorgesetzter SET ist_primaer=1 WHERE id=?", (rel["id"],))
                con.commit()
            reload_vg()

        def vg_delete():
            sel = vg_tree.selection()
            if not sel:
                messagebox.showinfo("Vorgesetzte", "Bitte zuerst eine Zuordnung auswählen.")
                return
            rel = vg_map.get(sel[0])
            with sqlite3.connect(DB_PATH) as con:
                con.execute("DELETE FROM tbl_mitarbeiter_vorgesetzter WHERE id=?", (rel["id"],))
                con.commit()
            reload_vg()

        vg_bar = tk.Frame(tab_vg, bg="#ffffff")
        vg_bar.grid(row=2, column=0, sticky="ew", padx=14, pady=10)
        tk.Button(vg_bar, text="➕ Zuordnen", command=vg_add, bg="#0b4a86", fg="white", relief="flat", padx=12, pady=6).pack(side="left", padx=(0, 8))
        tk.Button(vg_bar, text="★ Als primär", command=vg_set_primaer, padx=12, pady=6).pack(side="left", padx=8)
        tk.Button(vg_bar, text="Entfernen", command=vg_delete, padx=12, pady=6).pack(side="left", padx=8)
        reload_vg()

        def save_all():
            with sqlite3.connect(DB_PATH) as con:
                con.execute("""UPDATE tbl_mitarbeiter SET vorname=?, name=?, abteilung=?, position=?,
                               telefon=?, mobil=?, email=?, onedrive_pfad=?, geaendert_am=CURRENT_TIMESTAMP, bearbeiter=?
                               WHERE id=?""",
                            (stamm_vars["vorname"].get().strip(), stamm_vars["name"].get().strip(),
                             stamm_vars["abteilung"].get().strip(), stamm_vars["position"].get().strip(),
                             stamm_vars["telefon"].get().strip(), stamm_vars["mobil"].get().strip(),
                             stamm_vars["email"].get().strip(), stamm_vars["onedrive_pfad"].get().strip(),
                             self.bearbeiter, emp_id))
                for fid, var in extra_vars.items():
                    con.execute("""INSERT INTO tbl_mitarbeiter_wert(mitarbeiter_id, feld_id, wert) VALUES(?,?,?)
                                   ON CONFLICT(mitarbeiter_id, feld_id) DO UPDATE SET wert=excluded.wert""",
                                (emp_id, fid, var.get().strip()))
                con.commit()
            win.destroy()
            if callable(on_saved):
                on_saved()

        footer = tk.Frame(win, bg="#f5f7fb")
        footer.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 16))
        tk.Button(footer, text="Abbrechen", command=win.destroy, padx=16, pady=8).pack(side="right", padx=(8, 0))
        tk.Button(footer, text="Speichern", command=save_all, bg="#0b4a86", fg="white", relief="flat", padx=22, pady=9).pack(side="right")
        self.wait_window(win)

    def _ensure_kunden_center_extended(self):
        """Erweitert tbl_kunden_center um neue Pflichtfelder falls noch nicht vorhanden."""
        with sqlite3.connect(DB_PATH) as con:
            existing = {r[1] for r in con.execute("PRAGMA table_info(tbl_kunden_center)").fetchall()}
            for col, typedef in [
                ("plz", "TEXT"), ("ort", "TEXT"), ("strasse", "TEXT"),
                ("inhaber", "TEXT"), ("ansprechpartner2", "TEXT"),
                ("kundentyp", "TEXT"), ("ansprechpartner", "TEXT"),
                # Kunden-App-Erweiterung (Spec 2026-06-22)
                ("msk_kundennummer", "TEXT"), ("hausnummer", "TEXT"),
                ("inhaber_titel", "TEXT"), ("inhaber_anrede", "TEXT"),
                ("inhaber_vorname", "TEXT"), ("inhaber_zuname", "TEXT"),
                ("besteller_name", "TEXT"), ("besteller_durchwahl", "TEXT"),
                ("besteller_email", "TEXT"), ("rechnungsemail", "TEXT"),
                ("rechnungsart", "TEXT"), ("quartalsverguetung", "TEXT"),
            ]:
                if col not in existing:
                    con.execute(f"ALTER TABLE tbl_kunden_center ADD COLUMN {col} {typedef}")
            # Kunden-Typ vereinheitlichen: altes 'ZF' -> 'ZW'. (Die Analyse-Datenquelle
            # wird parallel in db.ensure_runtime_migrations ebenfalls auf 'ZW' migriert.)
            con.execute("UPDATE tbl_kunden_center SET kundentyp='ZW' WHERE kundentyp='ZF'")
            con.commit()

    def _kunden_detail_dialog(self, kunden_row=None):
        """Kompaktes Kunden-Formular mit Reitern: Stammdaten, Artikel-Rabatte,
        Analysen, Verkaufte Artikel, Notizen."""
        self._ensure_kunden_center_extended()
        is_new = kunden_row is None
        initial = dict(kunden_row) if kunden_row else {}
        title = "Neuer Kunde" if is_new else f"Kunde: {initial.get('kundenname','') or initial.get('kundennummer','')}"

        win = tk.Toplevel(self)
        win.title(title)
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.resizable(True, True)
        win.minsize(720, 500)
        # Kompaktes, mittig platziertes Fenster (kein Vollbild).
        w, h = 820, 600
        try:
            self.update_idletasks()
            x = self.winfo_rootx() + max(0, (self.winfo_width() - w) // 2)
            y = self.winfo_rooty() + max(0, (self.winfo_height() - h) // 2)
            win.geometry(f"{w}x{h}+{x}+{y}")
        except Exception:
            win.geometry(f"{w}x{h}")
        win.grab_set()

        win.columnconfigure(0, weight=1)
        win.rowconfigure(2, weight=1)

        vars_ = {}

        # ── Kopf: Typ-Auswahl (PK/ZW) + Kundennummer + MSK ────────────────────
        head = tk.Frame(win, bg="#f5f7fb")
        head.grid(row=0, column=0, sticky="ew", padx=18, pady=(14, 8))

        typ_var = tk.StringVar(value=str(initial.get("kundentyp", "") or ""))
        vars_["kundentyp"] = typ_var

        tk.Label(head, text="Typ:", bg="#f5f7fb", fg="#0b4a86",
                 font=(theme.FONT, 11, "bold")).pack(side="left")
        typ_btns = {}

        def _set_typ(val):
            typ_var.set(val)
            for v, b in typ_btns.items():
                on = (v == val)
                b.configure(bg=("#0b4a86" if on else "#e8eef5"),
                            fg=("white" if on else "#11304d"))
            _apply_typ_defaults(val)

        for val in ("PK", "ZW"):
            b = tk.Button(head, text=val, width=6, relief="flat", cursor="hand2",
                          bg="#e8eef5", fg="#11304d", font=(theme.FONT, 11, "bold"),
                          activebackground="#d8e2ee", command=lambda v=val: _set_typ(v))
            b.pack(side="left", padx=(8, 0))
            typ_btns[val] = b

        tk.Label(head, text="Kundennummer *", bg="#f5f7fb", fg="#c00",
                 font=(theme.FONT, 11, "bold")).pack(side="left", padx=(22, 6))
        knr_var = tk.StringVar(value=str(initial.get("kundennummer", "") or ""))
        vars_["kundennummer"] = knr_var
        tk.Entry(head, textvariable=knr_var, width=16, font=(theme.FONT, 11)).pack(side="left")

        tk.Label(head, text="MSK:", bg="#f5f7fb", fg="#0b4a86",
                 font=(theme.FONT, 11, "bold")).pack(side="left", padx=(16, 6))
        msk_var = tk.StringVar()
        vars_["msk_kundennummer"] = msk_var
        tk.Entry(head, textvariable=msk_var, width=16, state="readonly",
                 readonlybackground="#eef2f8", fg="#555").pack(side="left")

        def _sync_msk(*_a):
            k = knr_var.get().strip()
            msk_var.set(("216" + k) if k else "")
        knr_var.trace_add("write", _sync_msk)
        _sync_msk()

        tk.Frame(win, bg="#d8e2ee", height=1).grid(row=1, column=0, sticky="ew", padx=18)

        # ── Notebook mit Reitern ──────────────────────────────────────────────
        nb = ttk.Notebook(win)
        nb.grid(row=2, column=0, sticky="nsew", padx=18, pady=(8, 6))

        def _mk_section(parent, rowref, text, top=10):
            tk.Label(parent, text=text, bg="#ffffff", fg="#0b4a86",
                     font=(theme.FONT, 11, "bold")).grid(row=rowref[0], column=0, columnspan=2,
                                                      sticky="w", pady=(top, 2))
            rowref[0] += 1

        def _mk_field(parent, rowref, key, label, required=False):
            fg = "#c00" if required else "#11304d"
            tk.Label(parent, text=label, bg="#ffffff", fg=fg,
                     font=(theme.FONT, 9, "bold")).grid(row=rowref[0], column=0, sticky="w", pady=3, padx=(0, 8))
            var = tk.StringVar(value=str(initial.get(key, "") or ""))
            vars_[key] = var
            tk.Entry(parent, textvariable=var).grid(row=rowref[0], column=1, sticky="ew", pady=3)
            rowref[0] += 1
            return var

        def _mk_combo(parent, rowref, key, label, values, default=""):
            tk.Label(parent, text=label, bg="#ffffff", fg="#11304d",
                     font=(theme.FONT, 9, "bold")).grid(row=rowref[0], column=0, sticky="w", pady=3, padx=(0, 8))
            var = tk.StringVar(value=(str(initial.get(key, "") or "") or default))
            vars_[key] = var
            ttk.Combobox(parent, textvariable=var, values=values, state="readonly").grid(
                row=rowref[0], column=1, sticky="ew", pady=3)
            rowref[0] += 1
            return var

        # ---- Reiter: Stammdaten (zweispaltig) ----
        tab_stamm = tk.Frame(nb, bg="#ffffff")
        nb.add(tab_stamm, text="  Stammdaten  ")
        tab_stamm.columnconfigure(0, weight=1, uniform="cols")
        tab_stamm.columnconfigure(1, weight=1, uniform="cols")

        colL = tk.Frame(tab_stamm, bg="#ffffff")
        colL.grid(row=0, column=0, sticky="nsew", padx=(10, 12), pady=8)
        colL.columnconfigure(1, weight=1)
        colR = tk.Frame(tab_stamm, bg="#ffffff")
        colR.grid(row=0, column=1, sticky="nsew", padx=(12, 10), pady=8)
        colR.columnconfigure(1, weight=1)

        # Linke Spalte: Adresse (inkl. Apothekenname) + Inhaber
        rL = [0]
        _mk_section(colL, rL, "Adresse", top=0)
        _mk_field(colL, rL, "kundenname", "Apothekenname *", required=True)
        _mk_field(colL, rL, "strasse", "Straße")
        _mk_field(colL, rL, "hausnummer", "Hausnummer")
        tk.Label(colL, text="PLZ *", bg="#ffffff", fg="#c00",
                 font=(theme.FONT, 9, "bold")).grid(row=rL[0], column=0, sticky="w", pady=3, padx=(0, 8))
        _plz_box = tk.Frame(colL, bg="#ffffff")
        _plz_box.grid(row=rL[0], column=1, sticky="ew", pady=3)
        plz_var = tk.StringVar(value=str(initial.get("plz", "") or ""))
        vars_["plz"] = plz_var
        tk.Entry(_plz_box, textvariable=plz_var, width=10).pack(side="left")
        _plz_status = tk.Label(_plz_box, text="", bg="#ffffff", font=(theme.FONT, 8))
        _plz_status.pack(side="left", padx=(6, 0))
        rL[0] += 1
        _mk_field(colL, rL, "ort", "Ort")

        _mk_section(colL, rL, "Inhaber")
        _mk_field(colL, rL, "inhaber_titel", "Titel")
        _mk_combo(colL, rL, "inhaber_anrede", "Anrede", ["", "Frau", "Herr", "Divers"])
        _mk_field(colL, rL, "inhaber_vorname", "Vorname")
        _mk_field(colL, rL, "inhaber_zuname", "Zuname")

        def _check_plz(*_a):
            try:
                from .plz_lookup import is_valid_plz, lookup_ort
            except Exception:
                return
            p = plz_var.get().strip()
            if not p:
                _plz_status.config(text="", fg="#555")
                return
            if is_valid_plz(p):
                _plz_status.config(text="✓ gültig", fg="#127a2e")
                ort = lookup_ort(p)
                cur = vars_["ort"].get().strip()
                if ort and (not cur or cur == _check_plz.last):
                    vars_["ort"].set(ort)
                    _check_plz.last = ort
            else:
                _plz_status.config(text="✗ nicht gefunden", fg="#c00")
        _check_plz.last = ""
        try:
            from .plz_lookup import lookup_ort as _lo0
            _ip = str(initial.get("plz", "") or "").strip()
            if _ip and _lo0(_ip) and _lo0(_ip) == str(initial.get("ort", "") or "").strip():
                _check_plz.last = _lo0(_ip)
        except Exception:
            pass
        plz_var.trace_add("write", _check_plz)
        _check_plz()

        # Rechte Spalte: Kontakt + Besteller + Abrechnung
        rR = [0]
        _mk_section(colR, rR, "Kontakt", top=0)
        _mk_field(colR, rR, "telefon", "Telefon")
        _mk_field(colR, rR, "email", "E-Mail")
        _mk_field(colR, rR, "rechnungsemail", "Rechnungs-E-Mail")

        _mk_section(colR, rR, "Verantw. Besteller (Rückfragen)")
        _mk_field(colR, rR, "besteller_name", "Name (leer = Inhaber)")
        _mk_field(colR, rR, "besteller_durchwahl", "Durchwahl")
        _mk_field(colR, rR, "besteller_email", "E-Mail")

        _mk_section(colR, rR, "Abrechnung")
        _mk_combo(colR, rR, "rechnungsart", "Rechnungsart",
                  ["Sofortige Rechnung", "Monatlich", "Quartalsrechnung"],
                  default="Sofortige Rechnung")
        _mk_combo(colR, rR, "status", "Status", ["aktiv", "inaktiv"], default="aktiv")

        def _truthy(v):
            return str(v or "").strip().lower() in ("ja", "yes", "1", "true", "x")
        quart_touched = [False]
        quart_bv = tk.BooleanVar(value=_truthy(initial.get("quartalsverguetung")))
        tk.Checkbutton(colR, text="Quartalsvergütung Partnerprogramm", variable=quart_bv,
                       bg="#ffffff", fg="#11304d", activebackground="#ffffff",
                       font=(theme.FONT, 9, "bold"), anchor="w",
                       command=lambda: quart_touched.__setitem__(0, True)).grid(
            row=rR[0], column=0, columnspan=2, sticky="w", pady=(6, 3))
        rR[0] += 1

        def _apply_typ_defaults(val):
            # Vorbelegung nur bei NEUEM Kunden und solange der Nutzer den Haken
            # nicht selbst gesetzt hat: ZW -> nein, PK -> ja.
            if is_new and not quart_touched[0]:
                quart_bv.set(val == "PK")

        # Bestehenden Typ markieren (stylt die Buttons); bei neuem Kunden bleibt
        # bewusst nichts vorbelegt -> Typ muss aktiv gewählt werden.
        if str(initial.get("kundentyp", "") or "") in ("PK", "ZW"):
            _set_typ(initial["kundentyp"])

        # ---- Reiter: Artikel-Rabatte ----
        tab_rab = tk.Frame(nb, bg="#ffffff")
        nb.add(tab_rab, text="  Artikel-Rabatte  ")
        tab_rab.columnconfigure(0, weight=1)
        tab_rab.rowconfigure(0, weight=1)
        rab_tree = ttk.Treeview(tab_rab, columns=("pzn", "artikel", "rabatt"),
                                show="headings", selectmode="browse")
        for col, head_t, wdt in [("pzn", "PZN", 90), ("artikel", "Artikel", 260), ("rabatt", "Rabatt %", 80)]:
            rab_tree.heading(col, text=head_t)
            rab_tree.column(col, width=wdt, anchor=("e" if col == "rabatt" else "w"))
        rab_tree.grid(row=0, column=0, sticky="nsew", padx=(8, 0), pady=8)
        rab_sb = tk.Scrollbar(tab_rab, orient="vertical", command=rab_tree.yview)
        rab_sb.grid(row=0, column=1, sticky="ns", pady=8)
        rab_tree.configure(yscrollcommand=rab_sb.set)

        nmg_artikel = []
        try:
            with sqlite3.connect(DB_PATH) as con:
                nmg_artikel = con.execute(
                    "SELECT pzn, artikelname FROM tbl_nmg_stamm ORDER BY artikelname").fetchall()
        except Exception:
            nmg_artikel = []
        nmg_by_pzn = {str(p): (a or "") for p, a in nmg_artikel}

        def _rab_load():
            for it in rab_tree.get_children():
                rab_tree.delete(it)
            knr = knr_var.get().strip()
            if not knr:
                return
            try:
                with sqlite3.connect(DB_PATH) as con:
                    if not con.execute("SELECT 1 FROM sqlite_master WHERE type='table' "
                                       "AND name='tbl_pk_konditionen'").fetchone():
                        return
                    for pzn, rab in con.execute(
                            "SELECT pzn, rabatt_prozent FROM tbl_pk_konditionen "
                            "WHERE kundennummer=? ORDER BY pzn", (knr,)).fetchall():
                        art = nmg_by_pzn.get(str(pzn), "")
                        rab_tree.insert("", "end",
                                        values=(pzn, art, f"{rab:g}" if rab is not None else ""))
            except Exception:
                pass
        _rab_load()

        def _rab_add():
            if not knr_var.get().strip():
                messagebox.showinfo(title, "Bitte zuerst die Kundennummer eingeben.")
                return
            if not nmg_artikel:
                messagebox.showinfo(title, "Keine NMG-Artikel verfügbar.")
                return
            dlg = tk.Toplevel(win)
            dlg.title("Artikel-Rabatt hinzufügen")
            dlg.configure(bg="#f5f7fb")
            dlg.transient(win)
            dlg.grab_set()
            tk.Label(dlg, text="Artikel", bg="#f5f7fb", fg="#0b4a86",
                     font=(theme.FONT, 10, "bold")).grid(row=0, column=0, sticky="w", padx=12, pady=(12, 4))
            disp = [f"{a}  ({p})" for p, a in nmg_artikel]
            art_var = tk.StringVar()
            cb = ttk.Combobox(dlg, textvariable=art_var, values=disp, width=44, state="readonly")
            cb.grid(row=0, column=1, padx=12, pady=(12, 4))
            tk.Label(dlg, text="Rabatt %", bg="#f5f7fb", fg="#0b4a86",
                     font=(theme.FONT, 10, "bold")).grid(row=1, column=0, sticky="w", padx=12, pady=4)
            rab_var = tk.StringVar()
            tk.Entry(dlg, textvariable=rab_var, width=10).grid(row=1, column=1, sticky="w", padx=12, pady=4)

            def _ok():
                idx = cb.current()
                if idx < 0:
                    messagebox.showinfo(title, "Bitte einen Artikel wählen.")
                    return
                pzn = str(nmg_artikel[idx][0])
                art = nmg_artikel[idx][1] or ""
                try:
                    rab = float(rab_var.get().strip().replace(",", "."))
                except ValueError:
                    messagebox.showinfo(title, "Rabatt muss eine Zahl sein.")
                    return
                for it in rab_tree.get_children():
                    if str(rab_tree.set(it, "pzn")) == pzn:
                        rab_tree.set(it, "rabatt", f"{rab:g}")
                        dlg.destroy()
                        return
                rab_tree.insert("", "end", values=(pzn, art, f"{rab:g}"))
                dlg.destroy()
            tk.Button(dlg, text="Übernehmen", command=_ok, bg="#0b4a86", fg="white",
                      relief="flat", padx=14, pady=6).grid(row=2, column=1, sticky="e", padx=12, pady=12)

        def _rab_del():
            for it in rab_tree.selection():
                rab_tree.delete(it)

        rab_btns = tk.Frame(tab_rab, bg="#ffffff")
        rab_btns.grid(row=1, column=0, columnspan=2, sticky="ew", padx=8, pady=(0, 8))
        tk.Button(rab_btns, text="➕ Rabatt", command=_rab_add, bg="#3867b7", fg="white",
                  relief="flat", padx=12, pady=6).pack(side="left")
        tk.Button(rab_btns, text="➖ Entfernen", command=_rab_del, padx=10, pady=6).pack(side="left", padx=(8, 0))

        # ---- Reiter: Analysen ----
        tab_ana = tk.Frame(nb, bg="#ffffff")
        nb.add(tab_ana, text="  Analysen  ")
        tab_ana.columnconfigure(0, weight=1)
        tab_ana.rowconfigure(0, weight=1)
        ana_tree = ttk.Treeview(tab_ana, columns=("datum", "typ", "apotheke", "treffer"),
                                show="headings", selectmode="browse")
        for col, head_t, wdt in [("datum", "Datum", 90), ("typ", "Typ", 50),
                                 ("apotheke", "Name", 200), ("treffer", "Treffer", 60)]:
            ana_tree.heading(col, text=head_t)
            ana_tree.column(col, width=wdt, anchor="w")
        ana_tree.grid(row=0, column=0, sticky="nsew", padx=(8, 0), pady=8)
        ana_sb = tk.Scrollbar(tab_ana, orient="vertical", command=ana_tree.yview)
        ana_sb.grid(row=0, column=1, sticky="ns", pady=8)
        ana_tree.configure(yscrollcommand=ana_sb.set)

        ana_rows = {}

        def load_analysen():
            for item in ana_tree.get_children():
                ana_tree.delete(item)
            ana_rows.clear()
            knr = knr_var.get().strip()
            kname = vars_["kundenname"].get().strip()
            if not knr and not kname:
                return
            try:
                with sqlite3.connect(DB_PATH) as con:
                    con.row_factory = sqlite3.Row
                    rows = con.execute(
                        "SELECT id, datum, COALESCE(datenquelle,'PK') as datenquelle, "
                        "apotheke, quelldatei, ausgabedatei, nmg_treffer, kundennummer, kundenname "
                        "FROM tbl_auswertungen "
                        "WHERE (kundennummer=? AND kundennummer<>'') "
                        "OR (kundenname=? AND kundenname<>'') "
                        "ORDER BY datetime(datum) DESC LIMIT 30", (knr, kname)).fetchall()
                for row in rows:
                    dq = _dq_label(row["datenquelle"])
                    datum = str(row["datum"] or "")[:10]
                    iid = ana_tree.insert("", "end", values=(datum, dq, row["apotheke"] or "", row["nmg_treffer"] or 0))
                    ana_rows[iid] = dict(row)
            except Exception:
                pass

        load_analysen()

        def send_email_analyse():
            sel = ana_tree.selection()
            if not sel:
                messagebox.showinfo(title, "Bitte zuerst eine Analyse auswählen.")
                return
            row = ana_rows.get(sel[0], {})
            email_addr = vars_["email"].get().strip()
            if not email_addr:
                if messagebox.askyesno(title, "Keine E-Mail-Adresse hinterlegt.\nJetzt E-Mail-Adresse eingeben?"):
                    new_mail = simpledialog.askstring(title, "E-Mail-Adresse eingeben:")
                    if new_mail:
                        vars_["email"].set(new_mail.strip())
                        email_addr = new_mail.strip()
                    else:
                        return
                else:
                    return
            ausgabe = row.get("ausgabedatei", "") or ""
            anhang = ""
            if ausgabe and Path(ausgabe).exists():
                anhang = ausgabe
            else:
                try:
                    for f in (Path(__file__).resolve().parent.parent / "gespeicherte_analysen").rglob("*.xlsx"):
                        if row.get("apotheke") and str(row["apotheke"]).lower() in str(f).lower():
                            anhang = str(f)
                            break
                except Exception:
                    pass
            try:
                import sys
                if sys.platform.startswith("win"):
                    import win32com.client
                    outlook = win32com.client.Dispatch("Outlook.Application")
                    mail = outlook.CreateItem(0)
                    mail.To = email_addr
                    mail.Subject = f"Analyse – {row.get('apotheke','')}"
                    mail.Body = f"Anbei die Auswertung vom {str(row.get('datum',''))[:10]}."
                    if anhang:
                        mail.Attachments.Add(anhang)
                    mail.Display(True)
                else:
                    messagebox.showinfo(title, "Outlook-Integration nur unter Windows verfügbar.")
            except Exception as exc:
                messagebox.showerror(title, f"Outlook konnte nicht geöffnet werden:\n{exc}")

        ana_btns = tk.Frame(tab_ana, bg="#ffffff")
        ana_btns.grid(row=1, column=0, columnspan=2, sticky="ew", padx=8, pady=(0, 8))
        tk.Button(ana_btns, text="📧 Analyse per E-Mail senden", command=send_email_analyse,
                  bg="#3867b7", fg="white", relief="flat", padx=12, pady=6).pack(side="left")
        tk.Button(ana_btns, text="🔄 Aktualisieren", command=load_analysen, padx=10, pady=6).pack(side="left", padx=(8, 0))

        # ---- Reiter: Verkaufte Artikel ----
        tab_vk = tk.Frame(nb, bg="#ffffff")
        nb.add(tab_vk, text="  Verkaufte Artikel  ")
        tab_vk.columnconfigure(0, weight=1)
        tab_vk.rowconfigure(0, weight=1)
        vk_tree = ttk.Treeview(tab_vk, columns=("datum", "pzn", "artikel", "menge", "rabatt", "status"),
                               show="headings", selectmode="browse")
        for col, head_t, wdt, anc in [("datum", "Datum", 90, "w"), ("pzn", "PZN", 80, "w"),
                                      ("artikel", "Artikel", 210, "w"), ("menge", "Menge", 60, "e"),
                                      ("rabatt", "Rabatt %", 70, "e"), ("status", "Status", 90, "w")]:
            vk_tree.heading(col, text=head_t)
            vk_tree.column(col, width=wdt, anchor=anc)
        vk_tree.grid(row=0, column=0, sticky="nsew", padx=(8, 0), pady=8)
        vk_sb = tk.Scrollbar(tab_vk, orient="vertical", command=vk_tree.yview)
        vk_sb.grid(row=0, column=1, sticky="ns", pady=8)
        vk_tree.configure(yscrollcommand=vk_sb.set)
        vk_info = tk.Label(tab_vk, text="", bg="#ffffff", fg="#555", font=(theme.FONT, 9))
        vk_info.grid(row=1, column=0, columnspan=2, sticky="w", padx=10, pady=(0, 6))

        def load_verkaufte():
            for it in vk_tree.get_children():
                vk_tree.delete(it)
            knr = knr_var.get().strip()
            if not knr:
                vk_info.config(text="Keine Kundennummer angegeben.")
                return
            try:
                with sqlite3.connect(DB_PATH) as con:
                    if not con.execute("SELECT 1 FROM sqlite_master WHERE type='table' "
                                       "AND name='tbl_bestellpositionen'").fetchone():
                        vk_info.config(text="Noch keine Verkaufsdaten vorhanden.")
                        return
                    rows = con.execute(
                        "SELECT b.datum, p.pzn, p.artikelname, p.menge, p.rabatt_prozent, b.status "
                        "FROM tbl_bestellpositionen p "
                        "JOIN tbl_bestellungen b ON b.id = p.bestell_id "
                        "WHERE b.kundennummer = ? "
                        "ORDER BY datetime(b.datum) DESC, p.id DESC LIMIT 300", (knr,)).fetchall()
                gesamt = 0
                for datum, pzn, art, menge, rab, status in rows:
                    vk_tree.insert("", "end", values=(
                        str(datum or "")[:10], pzn or "", art or "",
                        menge if menge is not None else "",
                        f"{rab:g}" if rab is not None else "", status or ""))
                    gesamt += (menge or 0)
                vk_info.config(text=(f"{len(rows)} Positionen · Gesamtmenge {gesamt}"
                                     if rows else "Für diesen Kunden sind keine Verkäufe erfasst."))
            except Exception as exc:
                vk_info.config(text=f"Fehler beim Laden: {exc}")
        load_verkaufte()

        # ---- Reiter: Notizen ----
        tab_notiz = tk.Frame(nb, bg="#ffffff")
        nb.add(tab_notiz, text="  Notizen  ")
        tab_notiz.columnconfigure(0, weight=1)
        tab_notiz.rowconfigure(0, weight=1)
        notiz_txt = tk.Text(tab_notiz, wrap="word", relief="flat",
                            highlightbackground="#d8e2ee", highlightthickness=1)
        notiz_txt.insert("1.0", str(initial.get("notizen", "") or ""))
        notiz_txt.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        notiz_sb = tk.Scrollbar(tab_notiz, orient="vertical", command=notiz_txt.yview)
        notiz_sb.grid(row=0, column=1, sticky="ns", pady=10)
        notiz_txt.configure(yscrollcommand=notiz_sb.set)

        # ── Button-Leiste ─────────────────────────────────────────────────────
        bar = tk.Frame(win, bg="#e8edf5", highlightbackground="#c5d3e8", highlightthickness=1)
        bar.grid(row=3, column=0, sticky="ew")

        def save_kunde():
            if typ_var.get() not in ("PK", "ZW"):
                messagebox.showinfo(title, "Bitte oben den Typ (PK oder ZW) wählen.")
                return
            knr = knr_var.get().strip()
            kname = vars_["kundenname"].get().strip()
            plz = vars_["plz"].get().strip()
            if not knr:
                messagebox.showinfo(title, "Kundennummer ist ein Pflichtfeld.")
                return
            if not kname:
                messagebox.showinfo(title, "Apothekenname ist ein Pflichtfeld.")
                return
            if not plz:
                messagebox.showinfo(title, "PLZ ist ein Pflichtfeld.")
                return
            try:
                from .plz_lookup import is_valid_plz
                if not is_valid_plz(plz) and not messagebox.askyesno(
                        title, f"Die PLZ {plz} wurde nicht gefunden.\nTrotzdem speichern?"):
                    return
            except Exception:
                pass
            from datetime import datetime as _dt
            data = {k: v.get().strip() for k, v in vars_.items()}
            data["msk_kundennummer"] = ("216" + knr) if knr else ""
            _teile = [data.get("inhaber_titel", ""), data.get("inhaber_anrede", ""),
                      data.get("inhaber_vorname", ""), data.get("inhaber_zuname", "")]
            data["inhaber"] = " ".join(t for t in (x.strip() for x in _teile) if t)
            data["quartalsverguetung"] = "ja" if quart_bv.get() else "nein"
            data["notizen"] = notiz_txt.get("1.0", "end").strip()
            data["bearbeiter"] = self.bearbeiter

            cols = ["kundennummer", "msk_kundennummer", "kundenname", "kundentyp",
                    "strasse", "hausnummer", "plz", "ort",
                    "inhaber_titel", "inhaber_anrede", "inhaber_vorname", "inhaber_zuname", "inhaber",
                    "telefon", "email", "rechnungsemail",
                    "besteller_name", "besteller_durchwahl", "besteller_email",
                    "rechnungsart", "quartalsverguetung", "status", "notizen", "bearbeiter"]
            with sqlite3.connect(DB_PATH) as con:
                if is_new:
                    con.execute(
                        f"INSERT INTO tbl_kunden_center({','.join(cols)}) "
                        f"VALUES({','.join(':' + c for c in cols)})", data)
                else:
                    data["id"] = kunden_row["id"]
                    _sets = ",".join(f"{c}=:{c}" for c in cols)
                    con.execute(
                        f"UPDATE tbl_kunden_center SET {_sets},"
                        f"geaendert_am=CURRENT_TIMESTAMP WHERE id=:id", data)
                # Artikel-Rabatte: die Kunden-Maske ist alleiniger Editor je Kunde,
                # daher alle Konditionen dieses Kunden neu setzen.
                con.execute("""CREATE TABLE IF NOT EXISTS tbl_pk_konditionen(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    kundennummer TEXT, kundenname TEXT, pzn TEXT, rabatt_prozent REAL,
                    gueltigkeit TEXT, quelle TEXT, importdatum TEXT,
                    letzte_aktualisierung TEXT DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(kundennummer, pzn))""")
                con.execute("DELETE FROM tbl_pk_konditionen WHERE kundennummer=?", (knr,))
                _jetzt = _dt.now().isoformat(timespec="seconds")
                for it in rab_tree.get_children():
                    pzn = str(rab_tree.set(it, "pzn")).strip()
                    _raw = str(rab_tree.set(it, "rabatt")).strip().replace(",", ".")
                    if not pzn:
                        continue
                    try:
                        rabv = float(_raw) if _raw else None
                    except ValueError:
                        rabv = None
                    con.execute(
                        "INSERT OR REPLACE INTO tbl_pk_konditionen"
                        "(kundennummer,kundenname,pzn,rabatt_prozent,quelle,letzte_aktualisierung) "
                        "VALUES(?,?,?,?,?,?)", (knr, kname, pzn, rabv, "Kunden", _jetzt))
                con.commit()
            try:
                with sqlite3.connect(DB_PATH) as con:
                    cols_a = {r[1] for r in con.execute("PRAGMA table_info(tbl_auswertungen)").fetchall()}
                    for col in ("kundennummer", "kundenname"):
                        if col not in cols_a:
                            con.execute(f"ALTER TABLE tbl_auswertungen ADD COLUMN {col} TEXT")
                    con.execute(
                        "UPDATE tbl_auswertungen SET kundennummer=?, kundenname=? "
                        "WHERE (apotheke=? OR kundenname=?) AND (kundennummer IS NULL OR kundennummer='')",
                        (data["kundennummer"], data["kundenname"], data["kundenname"], data["kundenname"]))
                    con.commit()
            except Exception:
                pass
            win.destroy()

        tk.Button(bar, text="Abbrechen", command=win.destroy, padx=14, pady=7).pack(side="right", padx=(8, 12), pady=6)
        tk.Button(bar, text="✔  Speichern", command=save_kunde, bg="#0b4a86", fg="white",
                  relief="flat", font=(theme.FONT, 11, "bold"), padx=18, pady=7).pack(side="right", pady=6)


    def show_nmg_rabatte_uebersicht(self):
        """V1.1 SP19: Uebersicht ueber nmg_rabatte mit 4 Reitern.

        - Tabelle: alle aktuellen Rabatte sortier-/filterbar
        - Statistik: Anzahl, Min/Max/Avg, letztes Aktualisierungsdatum
        - Diff: gegen den letzten Snapshot (vor dem letzten Import)
        - Verlauf: PZN eingeben -> Historie ueber alle Snapshots
        """
        from .nmg_rabatte_history import (
            current_stats, diff_against_snapshot, history_for_pzn,
            list_snapshots,
        )
        self.clear_page()
        self._page_header("NMG-Rabatte", "Aktuelle Rabatte, Statistik, Diff zum letzten Stand und Verlauf pro PZN.")

        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=1)

        nb = ttk.Notebook(body)
        nb.grid(row=0, column=0, sticky="nsew")

        # --- Tab 1: Tabelle ---
        tab_tab = tk.Frame(nb, bg="#ffffff")
        nb.add(tab_tab, text="Alle Rabatte")
        tab_tab.columnconfigure(0, weight=1)
        tab_tab.rowconfigure(1, weight=1)

        filter_row = tk.Frame(tab_tab, bg="#ffffff")
        filter_row.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 6))
        tk.Label(filter_row, text="Filter (PZN, Artikel, Hersteller):", bg="#ffffff", fg="#0b4a86").pack(side="left", padx=(0, 6))
        filter_var = tk.StringVar()
        filter_entry = tk.Entry(filter_row, textvariable=filter_var, width=42)
        filter_entry.pack(side="left", padx=4)

        cols = ("nmg_pzn", "artikel", "rabatt", "quelle", "letzte_aktualisierung")
        col_headers = {"nmg_pzn": "PZN", "artikel": "Artikel", "rabatt": "Rabatt %", "quelle": "Quelle", "letzte_aktualisierung": "Aktualisiert"}
        tv = ttk.Treeview(tab_tab, columns=cols, show="headings", height=20)
        for c in cols:
            tv.heading(c, text=col_headers[c])
            tv.column(c, width=180 if c == "artikel" else 110, anchor="w")
        tv.column("rabatt", width=80, anchor="e")
        tv.grid(row=1, column=0, sticky="nsew", padx=10, pady=4)
        sb = ttk.Scrollbar(tab_tab, orient="vertical", command=tv.yview)
        sb.grid(row=1, column=1, sticky="ns")
        tv.configure(yscrollcommand=sb.set)

        info_var = tk.StringVar(value="")
        tk.Label(tab_tab, textvariable=info_var, bg="#ffffff", fg="#555", font=(theme.FONT, 9)).grid(row=2, column=0, sticky="w", padx=10, pady=(0, 8))

        def reload_table():
            for iid in tv.get_children():
                tv.delete(iid)
            q = filter_var.get().strip().lower()
            try:
                with sqlite3.connect(DB_PATH) as con:
                    rows = con.execute(
                        """SELECT nmg_pzn, artikel, rabatt, quelle, letzte_aktualisierung
                           FROM nmg_rabatte ORDER BY rabatt DESC, nmg_pzn"""
                    ).fetchall()
            except Exception:
                rows = []
            shown = 0
            for r in rows:
                if q:
                    hay = " ".join(str(v or "").lower() for v in r)
                    if q not in hay:
                        continue
                pct = f"{float(r[2]) * 100:.1f} %" if r[2] is not None else ""
                tv.insert("", "end", values=(r[0], r[1] or "", pct, r[3] or "", r[4] or ""))
                shown += 1
            info_var.set(f"{shown} von {len(rows)} Eintraegen")

        filter_var.trace_add("write", lambda *a: reload_table())
        reload_table()

        # --- Tab 2: Statistik ---
        tab_stats = tk.Frame(nb, bg="#ffffff")
        nb.add(tab_stats, text="Statistik")
        stats_inner = tk.Frame(tab_stats, bg="#f8fbff", highlightbackground="#d8e2ee", highlightthickness=1)
        stats_inner.pack(fill="both", expand=True, padx=14, pady=14)

        def render_stats():
            for w in stats_inner.winfo_children():
                w.destroy()
            with sqlite3.connect(DB_PATH) as con:
                stats = current_stats(con)
                snaps = list_snapshots(con, limit=10)
            tk.Label(stats_inner, text="Aktueller Stand", font=(theme.FONT, 13, "bold"), bg="#f8fbff", fg="#0b4a86").pack(anchor="w", padx=18, pady=(16, 6))
            grid = tk.Frame(stats_inner, bg="#f8fbff")
            grid.pack(anchor="w", padx=18, pady=4)
            def row(lbl, val):
                r = len(grid.grid_slaves(column=0))
                tk.Label(grid, text=lbl, bg="#f8fbff", fg="#444", font=(theme.FONT, 11)).grid(row=r, column=0, sticky="w", padx=(0, 18), pady=2)
                tk.Label(grid, text=val, bg="#f8fbff", fg="#123", font=(theme.FONT, 11, "bold")).grid(row=r, column=1, sticky="w", pady=2)
            row("Anzahl Eintraege:", str(stats["anzahl"]))
            row("Hoechster Rabatt:", f"{(stats['max'] or 0) * 100:.1f} %" if stats["max"] is not None else "-")
            row("Niedrigster Rabatt:", f"{(stats['min'] or 0) * 100:.1f} %" if stats["min"] is not None else "-")
            row("Durchschnitt:", f"{(stats['avg'] or 0) * 100:.1f} %" if stats["avg"] is not None else "-")
            row("Letzte Aktualisierung:", str(stats["letzte_aktualisierung"] or "-"))

            tk.Label(stats_inner, text=f"Snapshots ({len(snaps)})", font=(theme.FONT, 13, "bold"), bg="#f8fbff", fg="#0b4a86").pack(anchor="w", padx=18, pady=(24, 6))
            if not snaps:
                tk.Label(stats_inner, text="Noch keine Snapshots vorhanden. Der erste entsteht beim naechsten PK-Rabatte-Import.", bg="#f8fbff", fg="#666", justify="left", wraplength=600).pack(anchor="w", padx=18, pady=4)
            else:
                snap_tv = ttk.Treeview(stats_inner, columns=("id", "erstellt_am", "anzahl", "quelle"), show="headings", height=8)
                for c, h, w in [("id", "ID", 50), ("erstellt_am", "Erstellt am", 160), ("anzahl", "Eintraege", 80), ("quelle", "Quelle", 360)]:
                    snap_tv.heading(c, text=h)
                    snap_tv.column(c, width=w, anchor="w")
                snap_tv.pack(fill="x", padx=18, pady=8)
                for s in snaps:
                    snap_tv.insert("", "end", values=(s["id"], s["erstellt_am"], s["anzahl_eintraege"], s["quelle"] or ""))
        render_stats()

        # --- Tab 3: Diff ---
        tab_diff = tk.Frame(nb, bg="#ffffff")
        nb.add(tab_diff, text="Diff zum letzten Stand")
        diff_top = tk.Frame(tab_diff, bg="#ffffff")
        diff_top.pack(fill="x", padx=10, pady=(10, 4))
        diff_info = tk.StringVar(value="")
        tk.Label(diff_top, textvariable=diff_info, bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 11, "bold")).pack(anchor="w")
        tk.Button(diff_top, text="Neu laden", command=lambda: render_diff()).pack(anchor="w", pady=(4, 0))

        diff_lists = tk.Frame(tab_diff, bg="#ffffff")
        diff_lists.pack(fill="both", expand=True, padx=10, pady=4)
        diff_lists.columnconfigure((0, 1, 2), weight=1)
        diff_lists.rowconfigure(1, weight=1)

        def make_list(parent, col, title, headers):
            tk.Label(parent, text=title, font=(theme.FONT, 11, "bold"), bg="#ffffff", fg="#0b4a86").grid(row=0, column=col, sticky="w", padx=4, pady=(0, 4))
            t = ttk.Treeview(parent, columns=headers, show="headings", height=15)
            for h in headers:
                t.heading(h, text=h)
                t.column(h, width=110, anchor="w")
            t.grid(row=1, column=col, sticky="nsew", padx=4)
            return t

        diff_neu_tv = make_list(diff_lists, 0, "Neu hinzugekommen", ("PZN", "Artikel", "Rabatt %"))
        diff_chg_tv = make_list(diff_lists, 1, "Geaendert", ("PZN", "Artikel", "Vorher", "Nachher"))
        diff_ent_tv = make_list(diff_lists, 2, "Entfernt", ("PZN", "Artikel", "Rabatt %"))

        def render_diff():
            for tvw in (diff_neu_tv, diff_chg_tv, diff_ent_tv):
                for iid in tvw.get_children():
                    tvw.delete(iid)
            with sqlite3.connect(DB_PATH) as con:
                diff = diff_against_snapshot(con)
            if diff["snapshot_id"] is None:
                diff_info.set("Noch kein Snapshot vorhanden - der erste entsteht beim naechsten PK-Rabatte-Import.")
                return
            for d in diff["geaendert"]:
                # Cosmetic-Diff in Treeview-Spalten
                a = f"{(d['rabatt_alt'] or 0) * 100:.1f} %" if d['rabatt_alt'] is not None else "-"
                b = f"{(d['rabatt_neu'] or 0) * 100:.1f} %" if d['rabatt_neu'] is not None else "-"
                diff_chg_tv.column("Vorher", width=80)
                diff_chg_tv.column("Nachher", width=80)
                diff_chg_tv.insert("", "end", values=(d["nmg_pzn"], d.get("artikel") or "", a, b))
            for d in diff["neu"]:
                pct = f"{(d['rabatt'] or 0) * 100:.1f} %" if d.get("rabatt") is not None else "-"
                diff_neu_tv.insert("", "end", values=(d["nmg_pzn"], d.get("artikel") or "", pct))
            for d in diff["entfernt"]:
                pct = f"{(d['rabatt'] or 0) * 100:.1f} %" if d.get("rabatt") is not None else "-"
                diff_ent_tv.insert("", "end", values=(d["nmg_pzn"], d.get("artikel") or "", pct))
            diff_info.set(
                f"Snapshot #{diff['snapshot_id']}: "
                f"{len(diff['neu'])} neu, {len(diff['geaendert'])} geaendert, "
                f"{len(diff['entfernt'])} entfernt, {diff['unveraendert_count']} unveraendert."
            )
        render_diff()

        # --- Tab 4: Verlauf pro PZN ---
        tab_hist = tk.Frame(nb, bg="#ffffff")
        nb.add(tab_hist, text="Verlauf pro PZN")
        hist_top = tk.Frame(tab_hist, bg="#ffffff")
        hist_top.pack(fill="x", padx=10, pady=(10, 4))
        tk.Label(hist_top, text="PZN:", bg="#ffffff", fg="#0b4a86").pack(side="left", padx=(0, 6))
        hist_pzn_var = tk.StringVar()
        hist_entry = tk.Entry(hist_top, textvariable=hist_pzn_var, width=18)
        hist_entry.pack(side="left", padx=4)
        hist_info = tk.StringVar(value="PZN eingeben und Enter druecken.")
        tk.Label(hist_top, textvariable=hist_info, bg="#ffffff", fg="#555").pack(side="left", padx=12)

        hist_cols = ("snapshot", "erstellt_am", "artikel", "rabatt", "quelle")
        hist_tv = ttk.Treeview(tab_hist, columns=hist_cols, show="headings", height=18)
        for c, h, w in [("snapshot", "Snapshot", 90), ("erstellt_am", "Erstellt am", 160), ("artikel", "Artikel", 240), ("rabatt", "Rabatt %", 90), ("quelle", "Quelle", 240)]:
            hist_tv.heading(c, text=h)
            hist_tv.column(c, width=w, anchor="w")
        hist_tv.pack(fill="both", expand=True, padx=10, pady=4)

        def render_history(*_a):
            for iid in hist_tv.get_children():
                hist_tv.delete(iid)
            pzn = hist_pzn_var.get().strip()
            if not pzn:
                hist_info.set("PZN eingeben und Enter druecken.")
                return
            # PZN normalisieren: zfill 8 falls reine Ziffern
            digits = "".join(ch for ch in pzn if ch.isdigit())
            pzn_norm = digits.zfill(8) if digits else pzn
            with sqlite3.connect(DB_PATH) as con:
                verlauf = history_for_pzn(con, pzn_norm)
            if not verlauf:
                hist_info.set(f"Kein Verlauf fuer PZN {pzn_norm} gefunden.")
                return
            for h in verlauf:
                snap_label = f"#{h['snapshot_id']}" if h["snapshot_id"] else "aktuell"
                pct = f"{(h['rabatt'] or 0) * 100:.1f} %" if h.get("rabatt") is not None else "-"
                hist_tv.insert("", "end", values=(snap_label, h["erstellt_am"], h.get("artikel") or "", pct, h.get("snapshot_quelle") or ""))
            hist_info.set(f"{len(verlauf)} Eintraege fuer PZN {pzn_norm}.")
        hist_entry.bind("<Return>", render_history)
        tk.Button(hist_top, text="Anzeigen", command=render_history).pack(side="left", padx=(8, 0))

        self.status.set("NMG-Rabatte-Uebersicht geladen.")

    def show_apps_page(self):
        """Apps-Übersicht: alle Center als Schnellstart-Kacheln."""
        self.clear_page()
        self._page_header("Apps", "Schnellzugriff auf alle Bereiche.")
        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure((0, 1, 2, 3), weight=1)

        app_tiles = [
            ("\U0001f465", "Kunden", "Kundenstamm, Analysen, E-Mail-Versand.", self.show_kunden_center, "#0b4a86"),
            ("\U0001f464", "Mitarbeiter", "Mitarbeiterdaten, Profile, Vertretungen.", self.show_mitarbeiter_center, "#6b4fb3"),
            ("\u2705", "ToDo", "Aufgaben, offene Punkte und Notizen.", self.show_todo_center, "#11823b"),
            ("\U0001f6d2", "Kasse", "Verkauf an Apotheken + Wareneingang.", self.open_kasse_app, "#8b5a00"),
            ("\U0001f4d1", "Auswertungen", "Verkäufe, Kunden, Artikel frei auswerten und exportieren.", self.open_auswertungen_app, "#0b6e6e"),
            ("\U0001f50d", _T("Vergleichs-Suche"), _T("PZN oder Artikelname schnell in allen Wissens-Tabellen finden."), self.open_vergleichssuche_window, "#0b6e6e"),
            # V1.1 SP9: Globale Suche als App-Kachel.
            ("\U0001f50d", "Globale Suche", "Kunden, Analysen und Artikel uebergreifend finden.", self.open_globale_suche_window, "#0b4a86"),
            # V1.1 SP19: NMG-Rabatte-Uebersicht.
            ("\U0001f4b0", "NMG-Rabatte", "Aktuelle Rabatte sehen, Diff zum letzten Import, Verlauf pro PZN.", self.show_nmg_rabatte_uebersicht, "#6b4fb3"),
            ("❓", "Hilfe", "Bebildertes Handbuch: was wie funktioniert.", self.open_hilfe_app, "#208acd"),
        ]
        # V1.1 SP11: gleichgrosse Kacheln. rowconfigure(weight=1) + minsize,
        # grid statt pack innerhalb der Kachel, Beschreibung mit wraplength.
        cols_per_row = 4
        rows_needed = (len(app_tiles) + cols_per_row - 1) // cols_per_row
        for ri in range(rows_needed):
            body.rowconfigure(ri, weight=1, minsize=260)

        for idx, (icon, title, desc, cmd, color) in enumerate(app_tiles):
            r, c = divmod(idx, cols_per_row)
            f = tk.Frame(body, bg="#f8fbff", highlightbackground="#d8e2ee",
                         highlightthickness=1)
            f.grid(row=r, column=c, sticky="nsew", padx=10, pady=10)
            f.rowconfigure(2, weight=1)   # Beschreibung zieht sich, Rest fix
            f.columnconfigure(0, weight=1)
            tk.Label(f, text=icon, font=(theme.FONT, 32), bg="#f8fbff", fg=color
                     ).grid(row=0, column=0, pady=(18, 4))
            tk.Label(f, text=title, font=(theme.FONT, 13, "bold"), bg="#f8fbff",
                     fg="#123").grid(row=1, column=0)
            tk.Label(f, text=desc, justify="center", bg="#f8fbff", fg="#555",
                     font=(theme.FONT, 9), wraplength=190
                     ).grid(row=2, column=0, padx=14, pady=8, sticky="n")
            # Einheitlicher Oeffnen-Button (gleiche Hoehe + Schrift in allen Kacheln).
            tk.Button(f, text="Öffnen  →", command=cmd,
                      bg=color, fg="white", activebackground=color,
                      relief="flat", font=(theme.FONT, 10, "bold"),
                      padx=14, pady=7
                      ).grid(row=3, column=0, sticky="ew", padx=18, pady=(4, 16))
        self.status.set("Apps bereit.")

    def open_vergleichssuche_window(self):
        """SP31: Live-Suche in eigenem Toplevel-Fenster, parallel zum Hauptfenster nutzbar.
        Singleton: zweiter Aufruf bringt das vorhandene Fenster nach vorn.
        """
        existing = getattr(self, "_vergleichssuche_window", None)
        if existing is not None:
            try:
                if existing.winfo_exists():
                    existing.deiconify()
                    existing.lift()
                    existing.focus_force()
                    return
            except Exception:
                pass

        win = tk.Toplevel(self)
        win.title("NMGone — Vergleichs-Suche")
        win.geometry("1100x650")
        win.minsize(720, 420)
        win.configure(bg="#ffffff")
        self._vergleichssuche_window = win

        def on_close():
            self._vergleichssuche_window = None
            win.destroy()
        win.protocol("WM_DELETE_WINDOW", on_close)

        header = tk.Frame(win, bg="#ffffff")
        header.pack(fill="x", padx=14, pady=(10, 4))
        tk.Label(header, text=_T("Vergleichs-Suche"),
                 font=(theme.FONT, 15, "bold"), fg="#0b4a86", bg="#ffffff").pack(anchor="w")
        tk.Label(header,
                 text=_T("PZN, Artikelname, Wirkstoff, Hersteller oder Staerke eingeben (ab 3 Zeichen)."),
                 font=(theme.FONT, 9), fg="#666", bg="#ffffff").pack(anchor="w", pady=(2, 0))
        tk.Label(header,
                 text=_T("Filter mit # ergaenzen: #nmg  #austausch  #schulbank  #wirkstoff   ·   #<hersteller> (z.B. #hexal, #ratio)   ·   UND-Logik"),
                 font=(theme.FONT, 9), fg="#0b6e6e", bg="#ffffff").pack(anchor="w", pady=(1, 0))

        body = tk.Frame(win, bg="#ffffff")
        body.pack(fill="both", expand=True, padx=18, pady=(4, 14))
        body.columnconfigure(0, weight=1)
        body.rowconfigure(1, weight=1)

        toolbar = tk.Frame(body, bg="#ffffff")
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        tk.Label(toolbar, text=_T("Suche") + ":", bg="#ffffff",
                 fg="#0b4a86", font=(theme.FONT, 10, "bold")).pack(side="left")
        search_var = tk.StringVar()
        # V1.1 SP7: Pre-Search uebernehmen, falls aus der globalen Suche
        # heraus geoeffnet (do_search wird nach Aufbau noch unten getriggert).
        _vergleich_pre = getattr(self, "_vergleichssuche_pre_query", "") or ""
        if _vergleich_pre:
            search_var.set(_vergleich_pre)
            self._vergleichssuche_pre_query = ""
        search_entry = tk.Entry(toolbar, textvariable=search_var, width=42, font=(theme.FONT, 11))
        search_entry.pack(side="left", padx=(6, 12))
        status_label = tk.Label(toolbar, text="", bg="#ffffff", fg="#666", font=(theme.FONT, 9))
        status_label.pack(side="left")
        search_entry.focus_set()

        paned = ttk.Panedwindow(body, orient="horizontal")
        paned.grid(row=1, column=0, sticky="nsew")

        # --- Linke Seite: Treffer-Liste ---
        left = tk.Frame(paned, bg="#ffffff")
        left.columnconfigure(0, weight=1)
        left.rowconfigure(0, weight=1)
        paned.add(left, weight=2)

        cols = ("nmg", "austausch", "pzn", "artikel", "df", "pck", "herst")
        heads = {
            "nmg": "NMG",
            "austausch": _T("Austausch"),
            "pzn": "PZN",
            "artikel": _T("Artikel"),
            "df": "DF",
            "pck": "PCK",
            "herst": _T("Hersteller"),
        }
        widths = {"nmg": 40, "austausch": 70, "pzn": 90, "artikel": 320, "df": 50, "pck": 90, "herst": 110}

        tree = ttk.Treeview(left, columns=cols, show="headings", selectmode="browse")
        for c in cols:
            tree.heading(c, text=heads[c])
            anchor = "center" if c in ("nmg", "austausch") else "w"
            tree.column(c, width=widths[c], anchor=anchor, stretch=(c == "artikel"))
        tree.grid(row=0, column=0, sticky="nsew")
        sb_left = tk.Scrollbar(left, orient="vertical", command=tree.yview)
        sb_left.grid(row=0, column=1, sticky="ns")
        tree.configure(yscrollcommand=sb_left.set)

        # --- Rechte Seite: Detail-Panel ---
        right = tk.Frame(paned, bg="#ffffff")
        right.columnconfigure(0, weight=1)
        paned.add(right, weight=3)

        # Kopfzeile mit selektierter PZN + Artikelname.
        detail_head = tk.Label(
            right, text=_T("Bitte Treffer auswaehlen."),
            bg="#f0f5fa", fg="#0b4a86", font=(theme.FONT, 11, "bold"),
            anchor="w", padx=10, pady=8,
        )
        detail_head.grid(row=0, column=0, columnspan=2, sticky="ew")

        # Kompakte Stammdaten-Karte (Label/Wert in zwei Spalten).
        info_card = tk.Frame(right, bg="#fcfdff", padx=10, pady=8,
                             highlightbackground="#e0e8f0", highlightthickness=1)
        info_card.grid(row=1, column=0, columnspan=2, sticky="ew", padx=0, pady=(0, 6))
        info_card.columnconfigure(1, weight=1)
        info_card.columnconfigure(3, weight=1)
        info_labels = {}
        info_fields = [
            ("artikel", _T("Artikel"), 0, 0),
            ("herst", _T("Hersteller"), 0, 2),
            ("df", "DF", 1, 0),
            ("pck", "PCK", 1, 2),
            ("wirkstoff", _T("Wirkstoff"), 2, 0),
            ("staerke", _T("Staerke"), 2, 2),
            ("apu", "APU", 3, 0),
            ("ek", "EK", 3, 2),
            ("nmg", "NMG-Stamm", 4, 0),
            ("rabatt", _T("PK-Rabatt"), 4, 2),
        ]
        for key, label, row_i, col_i in info_fields:
            tk.Label(info_card, text=label + ":", bg="#fcfdff", fg="#666",
                     font=(theme.FONT, 9), anchor="w").grid(
                row=row_i, column=col_i, sticky="w", padx=(0, 6), pady=2)
            val = tk.Label(info_card, text="–", bg="#fcfdff", fg="#123",
                           font=(theme.FONT, 10, "bold"), anchor="w")
            val.grid(row=row_i, column=col_i + 1, sticky="ew", padx=(0, 14), pady=2)
            info_labels[key] = val

        # Treeview fuer Austauschartikel (PZN | Artikel | DF | PCK | Hersteller | Quelle).
        alt_label = tk.Label(right, text=_T("Austauschartikel"),
                             bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 11, "bold"),
                             anchor="w", padx=10, pady=4)
        alt_label.grid(row=2, column=0, columnspan=2, sticky="nw", pady=(6, 0))

        alt_frame = tk.Frame(right, bg="#ffffff")
        alt_frame.grid(row=3, column=0, columnspan=2, sticky="nsew")
        right.rowconfigure(3, weight=1)
        alt_frame.columnconfigure(0, weight=1)
        alt_frame.rowconfigure(0, weight=1)

        alt_cols = ("pzn", "artikel", "df", "pck", "herst", "quelle")
        alt_heads = {
            "pzn": "PZN",
            "artikel": _T("Artikelname"),
            "df": "DF",
            "pck": "PCK",
            "herst": _T("Hersteller"),
            "quelle": _T("Quelle"),
        }
        alt_widths = {"pzn": 80, "artikel": 240, "df": 50, "pck": 80, "herst": 90, "quelle": 100}
        alt_tree = ttk.Treeview(alt_frame, columns=alt_cols, show="headings", selectmode="browse", height=10)
        for c in alt_cols:
            alt_tree.heading(c, text=alt_heads[c])
            alt_tree.column(c, width=alt_widths[c], anchor="w", stretch=(c == "artikel"))
        alt_tree.grid(row=0, column=0, sticky="nsew")
        sb_alt = tk.Scrollbar(alt_frame, orient="vertical", command=alt_tree.yview)
        sb_alt.grid(row=0, column=1, sticky="ns")
        alt_tree.configure(yscrollcommand=sb_alt.set)

        alt_status = tk.Label(right, text="", bg="#ffffff", fg="#666",
                              font=(theme.FONT, 9), anchor="w", padx=10)
        alt_status.grid(row=4, column=0, columnspan=2, sticky="ew", pady=(4, 0))

        row_map: dict[str, dict] = {}
        pending_job = {"id": None}
        search_seq = {"current": 0}  # nur die neueste Suche darf die UI updaten
        detail_seq = {"current": 0}  # dito fuer Detail-Lookups beim Reihen-Wechsel

        def render_results(rows):
            for item in tree.get_children():
                tree.delete(item)
            row_map.clear()
            for r in rows:
                nmg_badge = "✓" if r["ist_nmg"] else ""
                aus_badge = "✓" if r["hat_austausch"] else ""
                iid = tree.insert("", "end", values=(
                    nmg_badge,
                    aus_badge,
                    r["pzn"],
                    r["artikel"] or "",
                    r.get("df", "") or "",
                    r.get("pck", "") or "",
                    r["herst"] or "",
                ))
                row_map[iid] = r

        def _fmt(value, default="–"):
            if value in (None, ""):
                return default
            return str(value)

        def render_details(pzn):
            """Laedt die Details im Worker-Thread, damit Klicks die UI nicht einfrieren."""
            detail_seq["current"] += 1
            my_id = detail_seq["current"]
            detail_head.configure(text=f"PZN {pzn}  ·  {_T('Lade ...')}")

            def worker():
                try:
                    d = get_pzn_details(pzn)
                except Exception as exc:
                    d = exc

                def apply():
                    if my_id != detail_seq["current"]:
                        return
                    if isinstance(d, Exception):
                        detail_head.configure(text=_T("Fehler beim Laden der Details"))
                        alt_status.configure(text=str(d))
                        return
                    _apply_details(d)

                try:
                    win.after(0, apply)
                except Exception:
                    pass

            threading.Thread(target=worker, daemon=True).start()

        def _apply_details(d):
            info = d.get("info") or {}
            detail_head.configure(
                text=f"PZN {d['pzn']}  ·  {info.get('artikel') or _T('Kein Artikelname bekannt')}"
            )

            info_labels["artikel"].configure(text=_fmt(info.get("artikel")))
            info_labels["herst"].configure(text=_fmt(info.get("herst")))
            info_labels["df"].configure(text=_fmt(info.get("df")))
            info_labels["pck"].configure(text=_fmt(info.get("pck")))
            info_labels["wirkstoff"].configure(text=_fmt(info.get("wirkstoff")))
            info_labels["staerke"].configure(text=_fmt(info.get("staerke")))
            apu = info.get("apu")
            info_labels["apu"].configure(
                text=f"{apu:.2f} €" if isinstance(apu, (int, float)) else "–"
            )
            ek = info.get("ek")
            info_labels["ek"].configure(
                text=f"{ek:.2f} €" if isinstance(ek, (int, float)) else "–"
            )
            info_labels["nmg"].configure(
                text="Ja" if d.get("ist_nmg") else "Nein",
                fg="#11823b" if d.get("ist_nmg") else "#888",
            )
            rabatt = d.get("nmg_rabatt")
            if isinstance(rabatt, (int, float)):
                info_labels["rabatt"].configure(text=f"{rabatt * 100:.2f} %")
            else:
                info_labels["rabatt"].configure(text="–")

            # Austauschartikel-Tabelle befuellen.
            for item in alt_tree.get_children():
                alt_tree.delete(item)
            alternativen = d.get("alternativen") or []
            for a in alternativen:
                pzn_disp = a["pzn"] or "–"
                artikel = a["artikel"] or a.get("freitext") or "–"
                alt_tree.insert("", "end", values=(
                    pzn_disp,
                    artikel,
                    a["df"] or "",
                    a["pck"] or "",
                    a["herst"] or "",
                    a.get("quelle") or "–",
                ))

            if alternativen:
                alt_label.configure(text=f"{_T('Austauschartikel')} ({len(alternativen)})")
                alt_status.configure(text="")
            else:
                alt_label.configure(text=_T("Austauschartikel"))
                alt_status.configure(text=_T("Keine Austauscheintraege fuer diese PZN."))

        def on_select(_event=None):
            sel = tree.selection()
            if not sel:
                return
            row = row_map.get(sel[0])
            if row:
                render_details(row["pzn"])

        tree.bind("<<TreeviewSelect>>", on_select)

        def do_search():
            """Triggert eine neue Suche im Worker-Thread. UI bleibt responsiv."""
            pending_job["id"] = None
            q = search_var.get().strip()
            if len(q) < 3:
                render_results([])
                status_label.configure(text=_T("Mindestens 3 Zeichen eingeben."))
                return
            status_label.configure(text=_T("Suche ..."))

            search_seq["current"] += 1
            my_id = search_seq["current"]

            def worker():
                try:
                    result = search_unified(q, limit=200)
                except Exception as exc:
                    result = exc

                def apply():
                    # Wenn waehrenddessen eine neue Suche gestartet wurde,
                    # alte Ergebnisse verwerfen.
                    if my_id != search_seq["current"]:
                        return
                    if isinstance(result, Exception):
                        status_label.configure(text=f"{_T('Fehler')}: {result}")
                        return
                    render_results(result)
                    status_label.configure(
                        text=_T("{n} Treffer").format(n=len(result))
                        if "{n}" in _T("{n} Treffer") else f"{len(result)} Treffer"
                    )

                try:
                    win.after(0, apply)
                except Exception:
                    pass

            threading.Thread(target=worker, daemon=True).start()

        def on_key(_event=None):
            if pending_job["id"]:
                try:
                    win.after_cancel(pending_job["id"])
                except Exception:
                    pass
            # 250 ms Debounce gegen Suche bei jedem einzelnen Tastendruck.
            pending_job["id"] = win.after(250, do_search)

        search_entry.bind("<KeyRelease>", on_key)
        search_entry.bind("<Return>", lambda _e: do_search())

        win.lift()
        win.focus_force()

        # V1.1 SP7: wenn ein Pre-Search aus der globalen Suche reinkam, direkt suchen.
        if _vergleich_pre:
            win.after(50, do_search)

    def show_kunden_center(self):
        """Kunden: Tabelle + Detail-Dialog mit Analysen und E-Mail."""
        self._ensure_center_tables()
        self._ensure_kunden_center_extended()
        self.clear_page()
        self._kunden_icon = theme.load_icon(ASSETS_DIR / "Kunden.ico", 40)
        self._page_header("Kunden", "Apotheken verwalten · Analysen einsehen · per E-Mail versenden.",
                          icon=self._kunden_icon)

        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure(0, weight=1)
        body.rowconfigure(1, weight=1)

        toolbar = tk.Frame(body, bg="#ffffff")
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 8))

        search_var = tk.StringVar()
        # V1.1 SP7: Pre-Search aus der globalen Startseiten-Suche uebernehmen.
        pre = getattr(self, "_kunden_center_pre_search", "") or ""
        if pre:
            search_var.set(pre)
            self._kunden_center_pre_search = ""
        tk.Label(toolbar, text="Suche:", bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 10, "bold")).pack(side="left")
        search_entry = tk.Entry(toolbar, textvariable=search_var, width=28)
        search_entry.pack(side="left", padx=(6, 12))

        cols = ("ampel", "kundennummer", "kundenname", "plz", "ort", "status", "letzte_analyse", "email")
        heads = {"ampel": "🚦", "kundennummer": "Kundennummer", "kundenname": "Apothekenname",
                 "plz": "PLZ", "ort": "Ort", "status": "Status",
                 "letzte_analyse": "Letzte Analyse", "email": "E-Mail"}

        tree_frame = tk.Frame(body, bg="#ffffff")
        tree_frame.grid(row=1, column=0, sticky="nsew")
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)

        tree = ttk.Treeview(tree_frame, columns=cols, show="headings", selectmode="browse")
        for col in cols:
            tree.heading(col, text=heads[col])
            if col == "ampel":
                tree.column(col, width=30, anchor="center", stretch=False)
            elif col == "kundenname":
                tree.column(col, width=200, anchor="w")
            elif col == "letzte_analyse":
                tree.column(col, width=90, anchor="w")
            else:
                tree.column(col, width=100, anchor="w")
        tree.grid(row=0, column=0, sticky="nsew")
        sb = tk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        sb.grid(row=0, column=1, sticky="ns")
        tree.configure(yscrollcommand=sb.set)

        row_map = {}

        def _kunden_ampel(kundennummer, kundenname):
            """Ampel: grün < 6 Monate, gelb 6-9, rot > 9 Monate oder nie."""
            from datetime import datetime, timedelta
            try:
                with sqlite3.connect(DB_PATH) as con:
                    knr = kundennummer or ""
                    kname = kundenname or ""
                    row = con.execute("""
                        SELECT MAX(datum) as letzte FROM tbl_auswertungen
                        WHERE (kundennummer=? AND kundennummer<>'') OR (kundenname=? AND kundenname<>'')
                    """, (knr, kname)).fetchone()
                letzte = row[0] if row and row[0] else None
                if not letzte:
                    return "🔴", "–"
                dt = datetime.strptime(str(letzte)[:10], "%Y-%m-%d")
                tage = (datetime.now() - dt).days
                datum_str = dt.strftime("%d.%m.%Y")
                if tage < 180:
                    return "🟢", datum_str
                elif tage < 270:
                    return "🟡", datum_str
                else:
                    return "🔴", datum_str
            except Exception:
                return "⚪", "–"

        def reload(filter_text=""):
            for item in tree.get_children():
                tree.delete(item)
            row_map.clear()
            try:
                with sqlite3.connect(DB_PATH) as con:
                    con.row_factory = sqlite3.Row
                    rows = con.execute("SELECT * FROM tbl_kunden_center ORDER BY kundenname").fetchall()
                for row in rows:
                    amp, letzte = _kunden_ampel(str(row["kundennummer"] or ""), str(row["kundenname"] or ""))
                    # vals mit Ampel und letzte_analyse
                    col_vals = {"ampel": amp, "kundennummer": str(row["kundennummer"] or ""),
                                "kundenname": str(row["kundenname"] or ""), "plz": str(row["plz"] if "plz" in row.keys() else ""),
                                "ort": str(row["ort"] if "ort" in row.keys() else ""),
                                "status": str(row["status"] or ""), "letzte_analyse": letzte,
                                "email": str(row["email"] if "email" in row.keys() else "")}
                    vals = tuple(col_vals.get(c, "") for c in cols)
                    if filter_text and filter_text.lower() not in " ".join(str(v) for v in vals).lower():
                        continue
                    iid = tree.insert("", "end", values=vals)
                    row_map[iid] = dict(row)
                    row_map[iid]["_letzte_analyse"] = letzte
                    row_map[iid]["_ampel"] = amp
            except Exception:
                pass

        reload()
        search_var.trace_add("write", lambda *_: reload(search_var.get().strip()))

        def open_detail(event=None):
            sel = tree.selection()
            if sel:
                self._kunden_detail_dialog(row_map.get(sel[0]))
                reload(search_var.get().strip())

        def new_kunde():
            self._kunden_detail_dialog(None)
            reload(search_var.get().strip())

        def delete_kunde():
            sel = tree.selection()
            if not sel:
                return
            row = row_map.get(sel[0])
            if not row:
                return
            if not messagebox.askyesno("Kunden", f"Kunde '{row.get('kundenname','')}' wirklich löschen?"):
                return
            with sqlite3.connect(DB_PATH) as con:
                con.execute("DELETE FROM tbl_kunden_center WHERE id=?", (row["id"],))
                con.commit()
            reload(search_var.get().strip())

        tree.bind("<Double-1>", open_detail)

        tk.Button(toolbar, text="Neu", command=new_kunde, bg="#0b4a86", fg="white", relief="flat", padx=14, pady=7).pack(side="left", padx=(0, 6))
        tk.Button(toolbar, text="Öffnen / Bearbeiten", command=open_detail, padx=14, pady=7).pack(side="left", padx=(0, 6))
        tk.Button(toolbar, text="Löschen", command=delete_kunde, padx=14, pady=7).pack(side="left", padx=(0, 6))
        tk.Button(toolbar, text="Aktualisieren", command=lambda: reload(search_var.get().strip()), padx=14, pady=7).pack(side="left")

        self.status.set("Kunden bereit.")

    def open_kasse_app(self):
        """Kasse als EIGENEN Prozess starten (NMGone.exe --kasse bzw. start.py
        --kasse). Nur so bekommt die Kasse ein eigenes Taskleisten-Icon (eigene
        AppUserModelID NMG.Kasse) getrennt von NMGone. DB wird via WAL geteilt.
        Zweiter Aufruf bringt keinen weiteren Start, solange der Prozess laeuft.
        """
        import subprocess
        proc = getattr(self, "_kasse_proc", None)
        if proc is not None and proc.poll() is None:
            messagebox.showinfo(
                "NMG Kasse",
                "Die Kasse läuft bereits in einem eigenen Fenster.\n"
                "Bitte über die Taskleiste nach vorn holen.", parent=self)
            return
        try:
            flags = 0
            if sys.platform == "win32":
                flags = getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)
            if getattr(sys, "frozen", False):
                cmd = [sys.executable, "--kasse"]
            else:
                start_py = Path(__file__).resolve().parent.parent / "start.py"
                cmd = [sys.executable, str(start_py), "--kasse"]
            self._kasse_proc = subprocess.Popen(cmd, close_fds=True, creationflags=flags)
        except Exception as exc:
            messagebox.showerror("NMG Kasse", f"Kasse konnte nicht gestartet werden:\n{exc}",
                                 parent=self)

    def open_auswertungen_app(self):
        """Auswertungs-/Report-Modul als EIGENEN Prozess starten (NMGone.exe
        --report bzw. start.py --report). Eigenes Taskleisten-Icon (AUMID
        NMG.Report), DB wird via WAL geteilt. Liest Kundendaten nur."""
        import subprocess
        proc = getattr(self, "_report_proc", None)
        if proc is not None and proc.poll() is None:
            messagebox.showinfo(
                "NMG Auswertungen",
                "Das Auswertungsmodul läuft bereits in einem eigenen Fenster.\n"
                "Bitte über die Taskleiste nach vorn holen.", parent=self)
            return
        try:
            flags = 0
            if sys.platform == "win32":
                flags = getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)
            if getattr(sys, "frozen", False):
                cmd = [sys.executable, "--report"]
            else:
                start_py = Path(__file__).resolve().parent.parent / "start.py"
                cmd = [sys.executable, str(start_py), "--report"]
            self._report_proc = subprocess.Popen(cmd, close_fds=True, creationflags=flags)
        except Exception as exc:
            messagebox.showerror("NMG Auswertungen",
                                 f"Auswertungsmodul konnte nicht gestartet werden:\n{exc}",
                                 parent=self)

    def open_hilfe_app(self):
        """Hilfe-/Handbuch-Modul als EIGENEN Prozess starten (NMGone.exe --hilfe
        bzw. start.py --hilfe). Eigenes Taskleisten-Icon (AUMID NMG.Hilfe). Das
        Modul zeigt nur das bebilderte Handbuch und aendert keine Daten."""
        import subprocess
        proc = getattr(self, "_hilfe_proc", None)
        if proc is not None and proc.poll() is None:
            messagebox.showinfo(
                "NMGone Hilfe",
                "Die Hilfe läuft bereits in einem eigenen Fenster.\n"
                "Bitte über die Taskleiste nach vorn holen.", parent=self)
            return
        try:
            flags = 0
            if sys.platform == "win32":
                flags = getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)
            if getattr(sys, "frozen", False):
                cmd = [sys.executable, "--hilfe"]
            else:
                start_py = Path(__file__).resolve().parent.parent / "start.py"
                cmd = [sys.executable, str(start_py), "--hilfe"]
            self._hilfe_proc = subprocess.Popen(cmd, close_fds=True, creationflags=flags)
        except Exception as exc:
            messagebox.showerror("NMGone Hilfe",
                                 f"Hilfe konnte nicht gestartet werden:\n{exc}",
                                 parent=self)

    def show_todo_center(self):
        cols = ("id", "titel", "bereich", "verantwortlich", "faellig_am", "status", "prioritaet")
        heads = {"id":"ID", "titel":"ToDo", "bereich":"Bereich", "verantwortlich":"Verantwortlich", "faellig_am":"Fällig", "status":"Status", "prioritaet":"Priorität"}
        fields = [("titel","ToDo / Aufgabe","entry"),("bereich","Bereich","entry"),("verantwortlich","Verantwortlich","entry"),("faellig_am","Fällig am","entry"),("status","Status","entry"),("prioritaet","Priorität","entry"),("notizen","Notizen","text")]
        insert = """INSERT INTO tbl_todo_center(titel,bereich,verantwortlich,faellig_am,status,prioritaet,notizen,bearbeiter) VALUES (:titel,:bereich,:verantwortlich,:faellig_am,:status,:prioritaet,:notizen,:bearbeiter)"""
        update = """UPDATE tbl_todo_center SET titel=:titel,bereich=:bereich,verantwortlich=:verantwortlich,faellig_am=:faellig_am,status=:status,prioritaet=:prioritaet,notizen=:notizen,geaendert_am=CURRENT_TIMESTAMP,bearbeiter=:bearbeiter WHERE id=:id"""
        self._show_center_table("ToDo-Center", "Aufgaben und offene Punkte intern vorbereiten.", "tbl_todo_center", cols, heads, fields, insert, update)

    def _get_meta_value(self, key, default=""):
        try:
            with sqlite3.connect(DB_PATH) as con:
                row = con.execute("SELECT value FROM meta WHERE key = ?", (key,)).fetchone()
                return row[0] if row else default
        except Exception:
            return default

    def _set_meta_value(self, key, value):
        try:
            with sqlite3.connect(DB_PATH) as con:
                con.execute("INSERT OR REPLACE INTO meta(key, value) VALUES (?, ?)", (key, value))
                con.commit()
        except Exception:
            pass

    def _roadmap_mark_v11_erledigt(self):
        """Roadmap-Eintrag für v11 nur einmal anlegen."""
        if self._get_meta_value("roadmap_v11_sortierung_formatassistent", "") == "erledigt":
            return
        try:
            add_roadmap_item(
                bereich="GUI / Import",
                titel="v11: Tabellen sortierbar und Rohdaten-Formatassistent vorbereitet",
                beschreibung=(
                    "Alle Treeview-Tabellen sind per Spaltenkopf sortierbar. "
                    "Bei unbekanntem Rohdatenformat kann ein Formatassistent PZN-Spalte, Menge/Absatz-Spalte und Zeitraum 6/12 Monate erfassen. "
                    "Das Mapping wird in tbl_rohdaten_mapping gespeichert."
                ),
                status="Erledigt",
                prioritaet="Hoch",
            )
            self._set_meta_value("roadmap_v11_sortierung_formatassistent", "erledigt")
        except Exception:
            pass

    def _roadmap_mark_status_ui_v15_erledigt(self):
        """Roadmap-Eintrag für rechten Statusbereich, Ampel, Logo und Vollbild."""
        title = "Rechter Statusbereich mit Daten-Ampel und Vollbildstart"
        try:
            for row in list_roadmap_items():
                if str(row["titel"]).strip().lower() == title.lower():
                    if row["status"] != "Erledigt":
                        update_roadmap_status(int(row["id"]), "Erledigt")
                    return
            add_roadmap_item(
                bereich="GUI",
                titel=title,
                beschreibung=(
                    "Aktiver Bearbeiter zeigt zusätzlich Version und Datenbankname. "
                    "Letzte Datenaktualisierung für APU/HAP, NMG Artikel, PK Rabatte und Artikelstamm "
                    "wird mit Datum, Leerzeilen, Ampel und Legende angezeigt. Logo wird proportional dargestellt "
                    "und das Programm startet maximiert."
                ),
                status="Erledigt",
                prioritaet="Normal",
            )
        except Exception:
            pass

    def _ensure_kunden_tables(self):
        with sqlite3.connect(DB_PATH) as con:
            for table in ("tbl_pk_kunden", "tbl_zw_kunden"):
                con.execute(f"""\nCREATE TABLE IF NOT EXISTS {table} (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,\nkundennummer TEXT,
                        kundenname TEXT,\napotheke TEXT,
                        status TEXT NOT NULL DEFAULT 'aktiv',\nerstellt_am TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        aktualisiert_am TEXT,\nbearbeiter TEXT
                    )\n""")
                con.execute(f"""\nCREATE UNIQUE INDEX IF NOT EXISTS idx_{table}_kundennummer
                    ON {table}(kundennummer)\nWHERE kundennummer IS NOT NULL AND kundennummer <> ''
                """)
            con.execute("INSERT OR IGNORE INTO meta(key,value) VALUES('neue_auswertung_kundentyp','PK')")
            con.commit()

    def _save_kunde_for_auswertung(self, kundentyp, kundennummer, kundenname, apotheke):
        self._ensure_kunden_tables()
        table = "tbl_zw_kunden" if kundentyp in ("ZW", "ZF") else "tbl_pk_kunden"
        nummer = (kundennummer or "").strip()
        name = (kundenname or "").strip()
        apo = (apotheke or name or nummer or "").strip()
        if not any([nummer, name, apo]):
            return
        bearbeiter = getpass.getuser()
        with sqlite3.connect(DB_PATH) as con:
            if nummer:
                existing = con.execute(f"SELECT id FROM {table} WHERE kundennummer = ? LIMIT 1", (nummer,)).fetchone()
            else:
                existing = con.execute(f"SELECT id FROM {table} WHERE COALESCE(kundenname,'') = ? AND COALESCE(apotheke,'') = ? LIMIT 1", (name, apo)).fetchone()
            if existing:
                con.execute(f"""\nUPDATE {table}
                    SET kundenname = ?,\napotheke = ?,
                        status = 'aktiv',\naktualisiert_am = CURRENT_TIMESTAMP,
                        bearbeiter = ?\nWHERE id = ?
                """, (name, apo, bearbeiter, existing[0]))
            else:
                con.execute(f"""\nINSERT INTO {table} (
                        kundennummer, kundenname, apotheke, status, erstellt_am, bearbeiter\n) VALUES (?, ?, ?, 'aktiv', CURRENT_TIMESTAMP, ?)
                """, (nummer, name, apo, bearbeiter))
            con.commit()

    def _mark_latest_auswertung_customer(self, kundentyp, kundennummer, kundenname):
        """Ergänzt Kundenangaben an der gerade erstellten Auswertung, falls die Spalten vorhanden sind."""
        try:
            with sqlite3.connect(DB_PATH) as con:
                cols = {r[1] for r in con.execute("PRAGMA table_info(tbl_auswertungen)").fetchall()}
                for col_name in ("kundentyp", "kundennummer", "kundenname"):
                    if col_name not in cols:
                        con.execute(f"ALTER TABLE tbl_auswertungen ADD COLUMN {col_name} TEXT")
                datenquelle = "ZW" if kundentyp in ("ZW", "ZF") else "NMG"
                row = con.execute("SELECT id FROM tbl_auswertungen ORDER BY id DESC LIMIT 1").fetchone()
                if row:
                    con.execute("""\nUPDATE tbl_auswertungen
                        SET kundentyp = ?,\nkundennummer = ?,
                            kundenname = ?,\ndatenquelle = ?
                        WHERE id = ?\n""", (kundentyp, kundennummer, kundenname, datenquelle, row[0]))
                con.commit()
        except Exception:
            pass

    def _roadmap_mark_neue_auswertung_form_erledigt(self):
        try:
            ensure_roadmap_table()
            titel = "Neue Auswertung Formular mit Kundentyp und Kundendaten"
            for row in list_roadmap_items():
                if row["titel"] == titel:
                    update_roadmap_status(int(row["id"]), "Erledigt")
                    return
            add_roadmap_item(
                bereich="Analyse",
                titel=titel,
                beschreibung=(
                    "Neue Auswertung startet jetzt über eine Formularansicht mit Kundentyp "
                    "Partnerkondition/Zukunftswerk, gespeicherter letzter Auswahl, Kundennummer, "
                    "Kundenname und Rohdaten-Dateiauswahl. Kundentabellen für PK und ZW sind vorbereitet."
                ),
                status="Erledigt",
                prioritaet="Hoch"
            )
        except Exception:
            pass

    def _kunden_zuordnen_dialog(self, auswertung_id, apotheke_name="", callback=None):
        """Dialog um eine Auswertung nachträglich einem Kunden zuzuordnen."""
        win = tk.Toplevel(self)
        win.resizable(True, True)
        win.title("Auswertung einem Kunden zuordnen")
        win.geometry("680x480")
        win.minsize(500, 380)
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()
        tk.Label(win, text="Auswertung einem Kunden zuordnen", font=(theme.FONT, 14, "bold"),
                 fg="#0b4a86", bg="#f5f7fb").pack(anchor="w", padx=20, pady=(16, 4))
        tk.Label(win, text=f"Auswertung: {apotheke_name}", font=(theme.FONT, 10), fg="#555",
                 bg="#f5f7fb").pack(anchor="w", padx=20, pady=(0, 8))
        tk.Label(win, text="Kunden suchen und zuordnen – oder Abbrechen für keine Zuordnung.",
                 font=(theme.FONT, 9), fg="#888", bg="#f5f7fb").pack(anchor="w", padx=20, pady=(0, 10))
        sf = tk.Frame(win, bg="#f5f7fb")
        sf.pack(fill="x", padx=20, pady=(0, 6))
        sv = tk.StringVar()
        tk.Label(sf, text="Suche:", bg="#f5f7fb", font=(theme.FONT, 10, "bold")).pack(side="left")
        tk.Entry(sf, textvariable=sv, width=35).pack(side="left", padx=(6, 8))
        cols = ("kundennummer", "kundenname", "plz", "ort", "status")
        heads = {"kundennummer": "Kundennr.", "kundenname": "Apothekenname", "plz": "PLZ", "ort": "Ort", "status": "Status"}
        tf = tk.Frame(win, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        tf.pack(fill="both", expand=True, padx=20, pady=(0, 8))
        tf.columnconfigure(0, weight=1)
        tf.rowconfigure(0, weight=1)
        tree = ttk.Treeview(tf, columns=cols, show="headings", selectmode="browse", height=10)
        for c in cols:
            tree.heading(c, text=heads[c])
            tree.column(c, width=120 if c != "kundenname" else 200, anchor="w")
        tree.grid(row=0, column=0, sticky="nsew")
        sb = tk.Scrollbar(tf, orient="vertical", command=tree.yview)
        sb.grid(row=0, column=1, sticky="ns")
        tree.configure(yscrollcommand=sb.set)
        row_map = {}
        def reload(ft=""):
            for item in tree.get_children(): tree.delete(item)
            row_map.clear()
            try:
                with sqlite3.connect(DB_PATH) as con:
                    con.row_factory = sqlite3.Row
                    rows = con.execute("SELECT * FROM tbl_kunden_center ORDER BY kundenname").fetchall()
                for row in rows:
                    vals = tuple(str(row[c] or "") for c in cols)
                    if ft and ft.lower() not in " ".join(vals).lower(): continue
                    iid = tree.insert("", "end", values=vals)
                    row_map[iid] = dict(row)
            except Exception: pass
        reload()
        if apotheke_name: sv.set(apotheke_name[:20]); reload(apotheke_name[:20])
        sv.trace_add("write", lambda *_: reload(sv.get().strip()))
        bar = tk.Frame(win, bg="#f5f7fb")
        bar.pack(fill="x", padx=20, pady=(0, 14))
        def uebernehmen():
            sel = tree.selection()
            if not sel:
                messagebox.showinfo("Zuordnen", "Bitte einen Kunden auswählen.")
                return
            k = row_map.get(sel[0], {})
            try:
                with sqlite3.connect(DB_PATH) as con:
                    for col in ("kundennummer", "kundenname"):
                        if col not in {r[1] for r in con.execute("PRAGMA table_info(tbl_auswertungen)").fetchall()}:
                            con.execute(f"ALTER TABLE tbl_auswertungen ADD COLUMN {col} TEXT")
                    con.execute("UPDATE tbl_auswertungen SET kundennummer=?, kundenname=? WHERE id=?",
                                (k.get("kundennummer",""), k.get("kundenname",""), auswertung_id))
                    con.commit()
            except Exception as exc:
                messagebox.showerror("Fehler", str(exc)); return
            win.destroy()
            if callback: callback()
        tk.Button(bar, text="Abbrechen", command=win.destroy, padx=14, pady=7).pack(side="right", padx=(8,0))
        tk.Button(bar, text="Übernehmen", command=uebernehmen, bg="#0b4a86", fg="white",
                  relief="flat", font=(theme.FONT, 11, "bold"), padx=16, pady=7).pack(side="right")

    def show_neue_auswertung_page(self):
        self._ensure_kunden_tables()
        self.clear_page()
        self._page_header(
            "Bedarfsanalyse",
            "Kundentyp, Kundendaten und Rohdaten in einer Ansicht erfassen."
        )

        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure(1, weight=1)

        last_type = self._get_meta_value("neue_auswertung_kundentyp", "PK")
        if last_type == "ZF":      # Alt-Token -> ZW
            last_type = "ZW"
        if last_type not in ("PK", "ZW"):
            last_type = "PK"

        kundentyp_var = tk.StringVar(value=last_type)
        kundennummer_var = tk.StringVar()
        kundenname_var = tk.StringVar()
        apotheke_var = tk.StringVar()
        rohdatei_var = tk.StringVar()
        vorlage_var = tk.StringVar(value=self._template_label_for_selected_slot())

        form = tk.Frame(body, bg="#f8fbff", highlightbackground="#d8e2ee", highlightthickness=1)
        form.grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 14), ipady=8)
        form.columnconfigure(1, weight=1)

        tk.Label(form, text="1. Kundentyp", bg="#f8fbff", fg="#0b4a86", font=(theme.FONT, 12, "bold")).grid(row=0, column=0, sticky="w", padx=14, pady=(14, 6))
        type_frame = tk.Frame(form, bg="#f8fbff")
        type_frame.grid(row=0, column=1, sticky="w", padx=14, pady=(14, 6))
        tk.Radiobutton(type_frame, text="Partnerkondition", variable=kundentyp_var, value="PK", bg="#f8fbff").pack(side="left", padx=(0, 18))
        tk.Radiobutton(type_frame, text="Zukunftswerk", variable=kundentyp_var, value="ZW", bg="#f8fbff").pack(side="left")

        # Kundennummer optional – für spätere Zuordnung
        tk.Label(form, text="2. Kundennummer (optional)", bg="#f8fbff", fg="#0b4a86", font=(theme.FONT, 12, "bold")).grid(row=1, column=0, sticky="w", padx=14, pady=7)
        entry_knr = tk.Entry(form, textvariable=kundennummer_var, width=28)
        entry_knr.grid(row=1, column=1, sticky="ew", padx=14, pady=7)
        tk.Label(form, text="Kann leer bleiben – Zuordnung nach Auswertung möglich", bg="#f8fbff", fg="#888", font=(theme.FONT, 9)).grid(row=1, column=2, sticky="w", padx=(0,14), pady=7)
        labels = [
            ("3. Kundenname / Apotheke", kundenname_var, "z. B. Rosen Apotheke Forst"),
            ("4. Auswertungsname", apotheke_var, "Name für Ausgabeordner und Analyse"),
        ]
        for idx, (label, var, hint) in enumerate(labels, start=2):
            tk.Label(form, text=label, bg="#f8fbff", fg="#0b4a86", font=(theme.FONT, 12, "bold")).grid(row=idx, column=0, sticky="w", padx=14, pady=7)
            entry = tk.Entry(form, textvariable=var, width=54)
            entry.grid(row=idx, column=1, sticky="ew", padx=14, pady=7)
            tk.Label(form, text=hint, bg="#f8fbff", fg="#666", font=(theme.FONT, 9)).grid(row=idx, column=2, sticky="w", padx=(0, 14), pady=7)

        def choose_raw_file():
            file = filedialog.askopenfilename(title="Rohdaten auswählen", filetypes=SUPPORTED_DATA_FILETYPES)
            if not file:
                return
            # Sofortige, klare Rueckmeldung bei unbekanntem Format - statt erst
            # nach dem Start im Hintergrund auf einen Fehler zu laufen.
            suffix = Path(file).suffix.lower()
            if suffix not in SUPPORTED_DATA_EXTENSIONS:
                erlaubt = ", ".join(sorted(SUPPORTED_DATA_EXTENSIONS))
                messagebox.showwarning(
                    "Format nicht unterstützt",
                    f"Die Datei „{Path(file).name}“ hat das Format „{suffix or '(kein)'}“, "
                    f"das nicht eingelesen werden kann.\n\nErlaubte Formate: {erlaubt}."
                )
                self.status.set(f"Format nicht unterstützt: {suffix or Path(file).name}")
                return
            rohdatei_var.set(file)
            if not apotheke_var.get().strip():
                apotheke_var.set(Path(file).stem.replace("_", " "))
            if not kundenname_var.get().strip():
                kundenname_var.set(Path(file).stem.replace("_", " "))
            self.status.set(f"Rohdaten ausgewählt: {Path(file).name}")

        tk.Label(form, text="5. Rohdaten", bg="#f8fbff", fg="#0b4a86", font=(theme.FONT, 12, "bold")).grid(row=4, column=0, sticky="w", padx=14, pady=7)
        file_frame = tk.Frame(form, bg="#f8fbff")
        file_frame.grid(row=4, column=1, columnspan=2, sticky="ew", padx=14, pady=7)
        file_frame.columnconfigure(0, weight=1)
        tk.Entry(file_frame, textvariable=rohdatei_var, state="readonly").grid(row=0, column=0, sticky="ew", padx=(0, 8))
        tk.Button(file_frame, text="Datei auswählen", command=choose_raw_file, padx=12, pady=5).grid(row=0, column=1)

        tk.Label(form, text="6. Auswertungsvorlage", bg="#f8fbff", fg="#0b4a86", font=(theme.FONT, 12, "bold")).grid(row=5, column=0, sticky="w", padx=14, pady=7)
        template_frame = tk.Frame(form, bg="#f8fbff")
        template_frame.grid(row=5, column=1, columnspan=2, sticky="ew", padx=14, pady=7)
        template_frame.columnconfigure(0, weight=1)
        ttk.Combobox(template_frame, textvariable=vorlage_var, values=self._template_combo_values(), state="readonly").grid(row=0, column=0, sticky="ew", padx=(0, 8))
        tk.Button(template_frame, text="Vorlagen verwalten", command=self.show_auswertungsvorlage_page, padx=12, pady=5).grid(row=0, column=1)

        info = (
            "Hinweis: Der Kundentyp und die gewählte Auswertungsvorlage werden gespeichert und beim nächsten Öffnen wieder vorgeschlagen.\n"
            "Die Kundendaten werden bereits in PK-/ZW-Kundentabellen vorbereitet."
        )
        tk.Label(body, text=info, justify="left", bg="#ffffff", fg="#333").grid(row=1, column=0, columnspan=3, sticky="w", pady=(0, 12))

        def start_analysis():
            file = rohdatei_var.get().strip()
            if not file:
                messagebox.showinfo("Bedarfsanalyse", "Bitte zuerst eine Rohdaten-Datei auswählen.")
                return
            kundentyp = kundentyp_var.get()
            kundennummer = kundennummer_var.get().strip()
            kundenname = kundenname_var.get().strip()
            analyse_name = _safe_name(apotheke_var.get().strip() or kundenname or Path(file).stem.replace("_", " "))
            if not analyse_name:
                messagebox.showinfo("Bedarfsanalyse", "Bitte Kundenname oder Auswertungsname eingeben.")
                return

            self._set_meta_value("neue_auswertung_kundentyp", kundentyp)
            self._save_kunde_for_auswertung(kundentyp, kundennummer, kundenname, analyse_name)
            vorlage_slot = self._slot_from_template_label(vorlage_var.get())
            self._set_selected_auswertungsvorlage_slot(vorlage_slot)

            # V1.1 SP18: Auswertung laeuft im Hintergrund. UnknownInputFormatError
            # + andere Exceptions werden in on_error innerhalb von
            # _run_neue_auswertung_export behandelt. _nach_auswertung_zuordnen
            # kommt jetzt als on_done-Callback (UI-Thread).
            self._run_neue_auswertung_export(
                file, analyse_name, kundentyp, kundennummer, kundenname or analyse_name,
                vorlage_slot=vorlage_slot,
                on_done=lambda _out: self.after(100, _nach_auswertung_zuordnen),
            )

        def _nach_auswertung_zuordnen():
            if kundennummer_var.get().strip():
                return
            try:
                with sqlite3.connect(DB_PATH) as con:
                    row = con.execute("SELECT id, kundennummer FROM tbl_auswertungen ORDER BY id DESC LIMIT 1").fetchone()
                    if row and not row[1]:
                        ana_id = row[0]
                        if messagebox.askyesno("Kundenzuordnung",
                                "Soll diese Auswertung einem Kunden zugeordnet werden?\n\n"
                                "Nein lässt die Auswertung ohne Kundenzuordnung gespeichert."):
                            self._kunden_zuordnen_dialog(ana_id, apotheke_name=apotheke_var.get().strip() or kundenname_var.get().strip())
            except Exception:
                pass

        buttonbar = tk.Frame(body, bg="#ffffff")
        buttonbar.grid(row=2, column=0, columnspan=3, sticky="w", pady=(0, 18))
        tk.Button(
            buttonbar,
            text="Auswertung starten  →",
            command=start_analysis,
            bg="#0b4a86",
            fg="white",
            activebackground="#0b4a86",
            relief="flat",
            font=(theme.FONT, 12, "bold"),
            padx=20,
            pady=10
        ).pack(side="left")
        tk.Button(
            buttonbar,
            text="Zurücksetzen",
            command=lambda: [kundennummer_var.set(""), kundenname_var.set(""), apotheke_var.set(""), rohdatei_var.set("")],
            padx=16,
            pady=9
        ).pack(side="left", padx=10)

        # Eigener Block: Abweichungsanalyse. Bewusst getrennt von Neue Auswertung.
        abw = tk.Frame(body, bg="#f8fbff", highlightbackground="#d8e2ee", highlightthickness=1)
        abw.grid(row=3, column=0, columnspan=3, sticky="ew", pady=(4, 0), ipady=8)
        abw.columnconfigure(1, weight=1)

        programm_auswertung_var = tk.StringVar()
        manuelle_anpassung_var = tk.StringVar()
        selected_programm_info = tk.StringVar(value="Keine Programm-Auswertung ausgewählt.")

        tk.Label(
            abw,
            text="Abweichungsanalyse",
            bg="#f8fbff",
            fg="#0b4a86",
            font=(theme.FONT, 16, "bold")
        ).grid(row=0, column=0, columnspan=3, sticky="w", padx=14, pady=(14, 4))
        tk.Label(
            abw,
            text="Getrennter Bereich: Programm-Auswertung auswählen, manuelle Anpassung laden und Abweichungsanalyse starten.",
            bg="#f8fbff",
            fg="#333"
        ).grid(row=1, column=0, columnspan=3, sticky="w", padx=14, pady=(0, 10))

        def choose_programm_file():
            file = filedialog.askopenfilename(title="Programm-Auswertung auswählen", filetypes=SUPPORTED_DATA_FILETYPES)
            if not file:
                return
            programm_auswertung_var.set(file)
            selected_programm_info.set(f"Datei: {Path(file).name}")
            self.status.set(f"Programm-Auswertung ausgewählt: {Path(file).name}")

        def choose_saved_programm():
            rows = self._get_saved_analysis_rows()
            if not rows:
                messagebox.showinfo("Gespeicherte Auswertungen", "Es wurden keine gespeicherten Auswertungen gefunden.")
                return

            win = tk.Toplevel(self)
            win.resizable(True, True)
            win.title("Gespeicherte Auswertung auswählen")
            win.geometry("920x460")
            win.configure(bg="#f5f7fb")
            win.transient(self)
            win.grab_set()

            tk.Label(win, text="Programm-Auswertung aus Datenbank wählen", font=(theme.FONT, 18, "bold"), fg="#0b4a86", bg="#f5f7fb").pack(anchor="w", padx=22, pady=(18, 6))
            tk.Label(win, text="Die Ausgabedatei der gewählten gespeicherten Analyse wird für die Abweichungsanalyse verwendet.", bg="#f5f7fb", fg="#333").pack(anchor="w", padx=22, pady=(0, 10))

            frame = tk.Frame(win, bg="#f5f7fb")
            frame.pack(fill="both", expand=True, padx=22, pady=(0, 12))
            frame.columnconfigure(0, weight=1)
            frame.rowconfigure(0, weight=1)

            lb = tk.Listbox(frame, font=("Consolas", 10), height=15)
            lb.grid(row=0, column=0, sticky="nsew")
            scroll = tk.Scrollbar(frame, orient="vertical", command=lb.yview)
            scroll.grid(row=0, column=0, sticky="nse")
            lb.configure(yscrollcommand=scroll.set)

            for row in rows:
                dq = _dq_label(row["datenquelle"])
                datum = str(row["datum"] or "")[:19]
                name = str(row["apotheke"] or "")[:38]
                lb.insert("end", f"{row['id']:>5} | {datum:<19} | {dq:<3} | {name:<38} | Pos: {row['anzahl_positionen'] or 0:>5}")

            def use_selected():
                sel = lb.curselection()
                if not sel:
                    messagebox.showinfo("Auswahl", "Bitte zuerst eine gespeicherte Auswertung auswählen.")
                    return
                row = rows[sel[0]]
                f = self._find_output_file_for_analysis_row(row)
                if not f:
                    messagebox.showwarning(
                        "Ausgabedatei fehlt",
                        "Zu dieser gespeicherten Auswertung wurde keine Ausgabedatei gefunden. Bitte die Programm-Auswertung als Datei auswählen."
                    )
                    return
                programm_auswertung_var.set(str(f))
                selected_programm_info.set(f"Gespeichert: ID {row['id']} | {row['apotheke']} | {f.name}")
                self.status.set(f"Gespeicherte Programm-Auswertung ausgewählt: {f.name}")
                win.destroy()

            btns = tk.Frame(win, bg="#f5f7fb")
            btns.pack(fill="x", padx=22, pady=(0, 18))
            tk.Button(btns, text="Abbrechen", command=win.destroy, padx=14, pady=8).pack(side="right", padx=(8, 0))
            tk.Button(btns, text="Auswertung übernehmen", command=use_selected, bg="#0b4a86", fg="white", relief="flat", padx=18, pady=9).pack(side="right")
            lb.bind("<Double-Button-1>", lambda e: use_selected())

        def choose_manual_file():
            file = filedialog.askopenfilename(title="Manuelle Anpassung auswählen", filetypes=SUPPORTED_DATA_FILETYPES)
            if not file:
                return
            manuelle_anpassung_var.set(file)
            self.status.set(f"Manuelle Anpassung ausgewählt: {Path(file).name}")

        def start_deviation_from_page():
            programm = programm_auswertung_var.get().strip()
            manuell = manuelle_anpassung_var.get().strip()
            if not programm:
                messagebox.showinfo("Abweichungsanalyse", "Bitte zuerst eine Programm-Auswertung auswählen.")
                return
            if not manuell:
                messagebox.showinfo("Abweichungsanalyse", "Bitte zuerst die manuelle Anpassung auswählen.")
                return
            try:
                out = export_abweichungsanalyse(manuell, programm)
                self._roadmap_mark_abweichung_in_neue_auswertung_erledigt()
                self.status.set(f"Abweichungsanalyse erzeugt: {out}")
                self._roadmap_mark_abweichung_schulbank_v9_erledigt()
                self.show_abweichungs_editor(out)
            except Exception as exc:
                messagebox.showerror("Abweichungsanalyse", str(exc))

        tk.Label(abw, text="1. Programm-Auswertung", bg="#f8fbff", fg="#0b4a86", font=(theme.FONT, 12, "bold")).grid(row=2, column=0, sticky="w", padx=14, pady=7)
        prog_frame = tk.Frame(abw, bg="#f8fbff")
        prog_frame.grid(row=2, column=1, columnspan=2, sticky="ew", padx=14, pady=7)
        prog_frame.columnconfigure(0, weight=1)
        tk.Entry(prog_frame, textvariable=programm_auswertung_var, state="readonly").grid(row=0, column=0, sticky="ew", padx=(0, 8))
        tk.Button(prog_frame, text="Gespeicherte auswählen", command=choose_saved_programm, padx=10, pady=5).grid(row=0, column=1, padx=(0, 6))
        tk.Button(prog_frame, text="Datei auswählen", command=choose_programm_file, padx=10, pady=5).grid(row=0, column=2)
        tk.Label(abw, textvariable=selected_programm_info, bg="#f8fbff", fg="#666", font=(theme.FONT, 9)).grid(row=3, column=1, columnspan=2, sticky="w", padx=14, pady=(0, 6))

        tk.Label(abw, text="2. Manuelle Anpassung", bg="#f8fbff", fg="#0b4a86", font=(theme.FONT, 12, "bold")).grid(row=4, column=0, sticky="w", padx=14, pady=7)
        man_frame = tk.Frame(abw, bg="#f8fbff")
        man_frame.grid(row=4, column=1, columnspan=2, sticky="ew", padx=14, pady=7)
        man_frame.columnconfigure(0, weight=1)
        tk.Entry(man_frame, textvariable=manuelle_anpassung_var, state="readonly").grid(row=0, column=0, sticky="ew", padx=(0, 8))
        tk.Button(man_frame, text="Datei auswählen", command=choose_manual_file, padx=10, pady=5).grid(row=0, column=1)

        abw_buttons = tk.Frame(abw, bg="#f8fbff")
        abw_buttons.grid(row=5, column=0, columnspan=3, sticky="w", padx=14, pady=(10, 14))
        tk.Button(
            abw_buttons,
            text="Abweichungsanalyse starten  →",
            command=start_deviation_from_page,
            bg="#8b5a00",
            fg="white",
            activebackground="#8b5a00",
            relief="flat",
            font=(theme.FONT, 12, "bold"),
            padx=18,
            pady=9
        ).pack(side="left")
        tk.Button(
            abw_buttons,
            text="Abweichung zurücksetzen",
            command=lambda: [programm_auswertung_var.set(""), manuelle_anpassung_var.set(""), selected_programm_info.set("Keine Programm-Auswertung ausgewählt.")],
            padx=14,
            pady=8
        ).pack(side="left", padx=10)

    def show_analysen_page(self):
        # SP7: Marktanalyse-Kachel raus. Produktanalyse-Kachel ruft jetzt
        # market_opportunities direkt (wie der Startseite-Button), statt
        # erst auf gespeicherte Auswertungen umzuleiten.
        self.clear_page()
        self._page_header("Analysen", "Produktanalyse, gespeicherte Analysen, Abweichungsanalyse oder manuelle Analysen importieren.")
        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure((0, 1, 2, 3), weight=1)
        self._tile(body, 0, "📈", "Produktanalyse", "Produktchancen erstellen.", "Starten", self.market_opportunities, "#11823b")
        self._tile(body, 1, "📁", "Gespeichert", "Gespeicherte Analysen öffnen.", "Öffnen", self.open_saved_analyses, "#0b4a86")
        self._tile(body, 2, "🔍", "Abweichung", "Manuelle und Programm-Auswertung vergleichen.", "Starten", self.deviation_analysis, "#8b5a00")
        self._tile(body, 3, "📥", "Manuelle Analysen", "Mitarbeiter-Auswertungen importieren und Schulbank/Analysen füttern.", "Import", self.import_manuelle_analysen, "#6b4fb3")

    def show_produktanalyse_page(self):
        # SP7: Verhalten an Startseite-Button angeglichen - kein Umweg ueber
        # "gespeicherte Auswertung waehlen", direkt market_opportunities.
        self.market_opportunities()

    # SP7: show_marktanalyse_page entfernt.

    def show_abweichungsanalyse_page(self):
        self._action_page(
            "Abweichungsanalyse",
            "Manuelle Auswertung und Programm-Auswertung auswählen und vergleichen.",
            "Abweichungsanalyse starten  →",
            self.deviation_analysis,
            "#8b5a00"
        )

    def show_schulbank_page(self, section="Schulbank"):
        self._ensure_lernvorschlaege_table()
        self._roadmap_mark_schulbank_bulk_erledigt()
        self.clear_page()

        if section == "Schulbank":
            self._page_header(
                "Schulbank",
                "Bitte einen Bereich auswählen. Daten werden erst in den Unterbereichen angezeigt."
            )

            body = tk.Frame(self.page, bg="#ffffff")
            body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
            body.columnconfigure((0, 1), weight=1)

            cards = [
                ("🆕", "Neue Lernvorschläge", "Neue, noch nicht bearbeitete Vorschläge anzeigen.", lambda: self.show_schulbank_page("Neue Lernvorschläge"), "#0b4a86"),
                ("🧬", "Biosimilar", "100% Austauscheinträge aus der Austauschdatenbank.", lambda: self.show_schulbank_page("Biosimilar"), "#0a7d8a"),
                ("✅", "Übernommen", "Manuell gelernte Vorschläge der letzten 4 Wochen anzeigen.", lambda: self.show_schulbank_page("Übernommen"), "#11823b"),
                ("❌", "Abgelehnt", "Abgelehnte Vorschläge der letzten 4 Wochen anzeigen.", lambda: self.show_schulbank_page("Abgelehnt"), "#9b1c1c"),
                ("🕘", "Historie", "Bearbeitbare Historie von 4 Wochen bis 6 Monate anzeigen.", lambda: self.show_schulbank_page("Historie"), "#8b5a00"),
                ("🔎", "Manuelle Prüfung", "Mehrdeutige gelernte Austausche prüfen und bereinigen.", self.show_schulbank_manuelle_pruefung, "#6b4fb3"),
            ]

            for idx, (icon, title, desc, cmd, color) in enumerate(cards):
                row = idx // 2
                col = idx % 2
                f = tk.Frame(body, bg="#f8fbff", highlightbackground="#d8e2ee", highlightthickness=1)
                f.grid(row=row, column=col, sticky="nsew", padx=10, pady=10, ipadx=8, ipady=8)
                tk.Label(f, text=icon, font=(theme.FONT, 34), bg="#f8fbff", fg=color).pack(pady=(20, 6))
                tk.Label(f, text=title, font=(theme.FONT, 15, "bold"), bg="#f8fbff", fg="#123").pack()
                tk.Label(f, text=desc, wraplength=260, justify="center", bg="#f8fbff", fg="#333").pack(padx=18, pady=12)
                tk.Button(
                    f,
                    text="Öffnen  →",
                    command=cmd,
                    bg=color,
                    fg="white",
                    activebackground=color,
                    relief="flat",
                    font=(theme.FONT, 12, "bold"),
                    padx=18,
                    pady=8
                ).pack(fill="x", padx=24, pady=(8, 18))

            self.status.set("Schulbank: Bereich auswählen.")
            return

        # SP19: Biosimilar als eigener Bereich. Inhalt kommt direkt aus
        # tbl_austauschdatenbank (Importdaten), separat von manuell gelernten
        # Faellen in tbl_lernvorschlaege.
        if section == "Biosimilar":
            self.show_schulbank_biosimilar_page()
            return

        self._page_header(
            f"AKTIVE SCHULBANK – {section}",
            "Lernen schreibt sofort in die Austauschdatenbank. Einträge bleiben dauerhaft nachvollziehbar."
        )

        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure(0, weight=1)
        body.rowconfigure(1, weight=1)

        status_map = {
            "Neue Lernvorschläge": "neu",
            "Übernommen": "uebernommen",
            "Abgelehnt": "abgelehnt",
            "Historie": getattr(self, "_schulbank_historie_filter", "historie_uebernommen"),
        }
        current_status = status_map.get(section, None)

        top = tk.Frame(body, bg="#ffffff")
        top.grid(row=0, column=0, sticky="ew", pady=(0, 10))

        if section == "Historie":
            def set_history_filter(value):
                self._schulbank_historie_filter = value
                self.show_schulbank_page("Historie")
            active_filter = getattr(self, "_schulbank_historie_filter", "historie_uebernommen")
            tk.Button(
                top,
                text="Übernommene",
                command=lambda: set_history_filter("historie_uebernommen"),
                bg="#11823b" if active_filter == "historie_uebernommen" else "#e9eef5",
                fg="white" if active_filter == "historie_uebernommen" else "#123",
                relief="flat",
                padx=14,
                pady=8
            ).pack(side="left", padx=(0, 8))
            tk.Button(
                top,
                text="Abgelehnte",
                command=lambda: set_history_filter("historie_abgelehnt"),
                bg="#9b1c1c" if active_filter == "historie_abgelehnt" else "#e9eef5",
                fg="white" if active_filter == "historie_abgelehnt" else "#123",
                relief="flat",
                padx=14,
                pady=8
            ).pack(side="left", padx=(0, 8))
        else:
            tk.Button(
                top,
                text="+ Lernvorschlag hinzufügen",
                command=lambda: self.add_lernvorschlag(section),
                bg="#0b4a86",
                fg="white",
                relief="flat",
                padx=14,
                pady=8
            ).pack(side="left")

        tk.Label(
            top,
            text="Übernommen und Abgelehnt bleiben 7 Tage sichtbar und wandern danach in die Historie.",
            bg="#ffffff",
            fg="#333",
            font=(theme.FONT, 10)
        ).pack(side="left", padx=14)

        columns = ("id", "pzn_alt", "artikel_alt", "pzn_nmg", "freitext_austausch", "status", "bearbeiter", "erstellt_am", "bearbeitet_am")
        tree = ttk.Treeview(body, columns=columns, show="headings", selectmode="extended")
        tree.grid(row=1, column=0, sticky="nsew")

        headings = {
            "id": "ID",
            "pzn_alt": "PZN alt",
            "artikel_alt": "Artikel alt",
            "pzn_nmg": "PZN NMG",
            "freitext_austausch": "Austauschbar gegen",
            "status": "Status",
            "bearbeiter": "Bearbeiter",
            "erstellt_am": "Erstellt",
            "bearbeitet_am": "Bearbeitet",
        }
        widths = {
            "id": 50,
            "pzn_alt": 90,
            "artikel_alt": 210,
            "pzn_nmg": 90,
            "freitext_austausch": 230,
            "status": 105,
            "bearbeiter": 110,
            "erstellt_am": 135,
            "bearbeitet_am": 135,
        }

        for col in columns:
            tree.heading(col, text=headings[col])
            tree.column(col, width=widths[col], anchor="w")
        self._make_tree_sortable(tree, columns, headings)

        scrollbar = tk.Scrollbar(body, orient="vertical", command=tree.yview)
        scrollbar.grid(row=1, column=0, sticky="nse")
        tree.configure(yscrollcommand=scrollbar.set)

        actionbar = tk.Frame(body, bg="#ffffff")
        actionbar.grid(row=2, column=0, sticky="ew", pady=(12, 0))

        def selected_ids():
            sel = tree.selection()
            if not sel:
                messagebox.showinfo("Schulbank", "Bitte zuerst einen oder mehrere Lernvorschläge auswählen.")
                return []
            ids = []
            for iid in sel:
                try:
                    ids.append(int(tree.item(iid, "values")[0]))
                except Exception:
                    pass
            return ids

        def selected_id():
            ids = selected_ids()
            return ids[0] if ids else None

        rows = self.get_lernvorschlaege(current_status)
        visible_ids = []
        # SP18: Mapping iid -> (source_table, id) damit delete_selected
        # in die richtige Tabelle loescht.
        row_source_map = {}
        for row in rows:
            visible_ids.append(int(row["id"]))
            iid = tree.insert("", "end", values=(
                row["id"],
                row["pzn_alt"] or "",
                row["artikel_alt"] or row["produkt_alt"] or "",
                row["pzn_nmg"] or "",
                row["freitext_austausch"] or row["produkt_neu"] or "",
                row["status"],
                row["bearbeiter"] or "",
                row["erstellt_am"],
                row["bearbeitet_am"] or "",
            ))
            try:
                row_source_map[iid] = (
                    row["source_table"] if "source_table" in row.keys() else "tbl_lernvorschlaege",
                    int(row["id"]),
                )
            except Exception:
                row_source_map[iid] = ("tbl_lernvorschlaege", int(row["id"]))

        def update_many(item_ids, status, label, confirm_all=False):
            if not item_ids:
                messagebox.showinfo("Schulbank", "Keine Einträge ausgewählt.")
                return
            if confirm_all:
                if not messagebox.askyesno("Schulbank Sammelaktion", f"{len(item_ids)} sichtbare Vorschläge werden jetzt {label}.\n\nFortfahren?"):
                    return
            elif len(item_ids) > 1:
                if not messagebox.askyesno("Schulbank Sammelaktion", f"{len(item_ids)} markierte Vorschläge werden jetzt {label}.\n\nFortfahren?"):
                    return

            done = 0
            errors = []
            for item_id in item_ids:
                try:
                    self.update_lernvorschlag_status(item_id, status)
                    done += 1
                except Exception as exc:
                    errors.append(f"ID {item_id}: {exc}")

            if errors:
                messagebox.showwarning("Schulbank", f"{done} Einträge verarbeitet.\n\nFehler:\n" + "\n".join(errors[:8]))
            else:
                self.status.set(f"Schulbank: {done} Einträge {label}.")
            self.show_schulbank_page(section)

        def set_status(status):
            ids = selected_ids()
            if not ids:
                return
            # SP19: vor dem Uebernehmen pruefen, ob die PZN_alt bereits in
            # der Biosimilar-Datenbank (tbl_austauschdatenbank) liegt.
            # User entscheidet, ob er trotzdem uebernehmen will.
            if status == "uebernommen":
                pzns_to_check = []
                for iid in tree.selection():
                    try:
                        vals = tree.item(iid, "values")
                        pzn_alt = self._normalize_pzn_input(vals[1] if len(vals) > 1 else "")
                        item_id = int(vals[0])
                        if pzn_alt:
                            pzns_to_check.append((item_id, pzn_alt))
                    except Exception:
                        pass
                bereits_biosimilar = []
                try:
                    with sqlite3.connect(DB_PATH) as con:
                        for item_id, pzn_alt in pzns_to_check:
                            row = con.execute(
                                "SELECT pzn_nmg, freitext_austausch FROM tbl_austauschdatenbank "
                                "WHERE COALESCE(status,'aktiv')='aktiv' AND pzn_alt=? LIMIT 1",
                                (pzn_alt,),
                            ).fetchone()
                            if row:
                                bereits_biosimilar.append((pzn_alt, row[0] or "", row[1] or ""))
                except Exception:
                    pass
                if bereits_biosimilar:
                    summary = "\n".join(
                        f"  - PZN {p}  ->  NMG {n or '-'}  ({t[:40]})" for p, n, t in bereits_biosimilar[:8]
                    )
                    if not messagebox.askyesno(
                        "Schulbank: bereits als Biosimilar erfasst",
                        f"{len(bereits_biosimilar)} der markierten PZN ist/sind bereits in der Biosimilar-Datenbank vorhanden:\n\n{summary}\n\nTrotzdem zusaetzlich als manuell uebernommen speichern?"
                    ):
                        return
            label = "gelernt" if status == "uebernommen" else ("abgelehnt" if status == "abgelehnt" else "zurückgesetzt")
            update_many(ids, status, label)

        def delete_selected():
            # SP18: nach source_table aufteilen. Frueher wurde immer aus
            # tbl_lernvorschlaege geloescht, was bei "Uebernommen"-Eintraegen
            # (die aus tbl_austauschdatenbank kommen) keine Wirkung hatte.
            sel = tree.selection()
            if not sel:
                messagebox.showinfo("Schulbank", "Bitte zuerst einen oder mehrere Lernvorschläge auswählen.")
                return
            to_delete = []  # list of (source_table, id)
            for iid in sel:
                if iid in row_source_map:
                    to_delete.append(row_source_map[iid])
            if not to_delete:
                return
            # Zaehler pro Tabelle in der Bestaetigungsmeldung
            from_lern = sum(1 for t, _ in to_delete if t == "tbl_lernvorschlaege")
            from_aust = sum(1 for t, _ in to_delete if t == "tbl_austauschdatenbank")
            info = f"{len(to_delete)} Eintrag/Einträge werden gelöscht:"
            if from_lern:
                info += f"\n  • aus Lernvorschlägen: {from_lern}"
            if from_aust:
                info += f"\n  • aus Austauschdatenbank: {from_aust}"
            info += "\n\nFortfahren?"
            if not messagebox.askyesno("Schulbank löschen", info):
                return
            done = 0
            errors = []
            for source_table, item_id in to_delete:
                try:
                    if source_table == "tbl_austauschdatenbank":
                        with sqlite3.connect(DB_PATH) as con:
                            con.execute("DELETE FROM tbl_austauschdatenbank WHERE id = ?", (item_id,))
                            con.commit()
                    else:
                        self.delete_lernvorschlag(item_id)
                    done += 1
                except Exception as exc:
                    errors.append(f"{source_table}#{item_id}: {exc}")
            msg = f"{done} von {len(to_delete)} Eintrag/Einträgen gelöscht."
            if errors:
                msg += "\n\nFehler:\n" + "\n".join(errors[:8])
                messagebox.showwarning("Schulbank", msg)
            else:
                self.status.set(msg)
                messagebox.showinfo("Schulbank", msg)
            self.show_schulbank_page(section)

        tk.Button(actionbar, text="✓ Markierte lernen", command=lambda: set_status("uebernommen"), bg="#11823b", fg="white", relief="flat", padx=14, pady=8).pack(side="left", padx=(0, 8))
        tk.Button(actionbar, text="✗ Markierte ablehnen", command=lambda: set_status("abgelehnt"), bg="#9b1c1c", fg="white", relief="flat", padx=14, pady=8).pack(side="left", padx=8)
        tk.Button(actionbar, text="✓ Alle sichtbaren lernen", command=lambda: update_many(visible_ids, "uebernommen", "gelernt", True), bg="#11823b", fg="white", relief="flat", padx=14, pady=8).pack(side="left", padx=8)
        tk.Button(actionbar, text="✗ Alle sichtbaren ablehnen", command=lambda: update_many(visible_ids, "abgelehnt", "abgelehnt", True), bg="#9b1c1c", fg="white", relief="flat", padx=14, pady=8).pack(side="left", padx=8)
        tk.Button(actionbar, text="↩ Rückgängig", command=lambda: set_status("neu"), bg="#8b5a00", fg="white", relief="flat", padx=14, pady=8).pack(side="left", padx=8)
        tk.Button(actionbar, text="Löschen", command=delete_selected, padx=14, pady=8).pack(side="left", padx=8)

        self.status.set(f"Schulbank aktiv: {len(rows)} Einträge angezeigt. Mehrfachauswahl ist möglich.")


    def show_schulbank_biosimilar_page(self):
        """SP19: Biosimilar-Bereich. Zeigt alle aktiven Eintraege aus
        tbl_austauschdatenbank (100% Austausch aus Importdatei).
        Loeschen und Doppelklick zum Aendern.
        """
        self.clear_page()
        self._page_header(
            "Schulbank – Biosimilar",
            "100% Austauscheintraege aus der Austauschdatenbank. Daten kommen aus dem Import (Austauschdatenbank aktualisieren). Doppelklick zum Aendern."
        )

        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure(0, weight=1)
        body.rowconfigure(2, weight=1)

        info_var = tk.StringVar(value="Biosimilar-Eintraege werden geladen ...")
        tk.Label(body, textvariable=info_var, bg="#ffffff", fg="#333", anchor="w", justify="left").grid(row=0, column=0, sticky="ew", pady=(0, 8))

        search_var = tk.StringVar()
        search_row = tk.Frame(body, bg="#ffffff")
        search_row.grid(row=1, column=0, sticky="ew", pady=(0, 6))
        tk.Label(search_row, text="Suche (PZN oder Artikel):", bg="#ffffff").pack(side="left")
        tk.Entry(search_row, textvariable=search_var, width=40).pack(side="left", padx=8)

        columns = ("id", "pzn_alt", "artikel_alt", "pzn_nmg", "artikel_nmg", "freitext_austausch", "quelle", "bearbeiter", "erstellt_am")
        headings = {
            "id": "ID",
            "pzn_alt": "PZN alt",
            "artikel_alt": "Artikel alt",
            "pzn_nmg": "PZN NMG",
            "artikel_nmg": "Artikel NMG",
            "freitext_austausch": "Austauschbar gegen",
            "quelle": "Quelle",
            "bearbeiter": "Bearbeiter",
            "erstellt_am": "Erstellt",
        }
        widths = {
            "id": 55,
            "pzn_alt": 90,
            "artikel_alt": 200,
            "pzn_nmg": 90,
            "artikel_nmg": 200,
            "freitext_austausch": 220,
            "quelle": 130,
            "bearbeiter": 110,
            "erstellt_am": 135,
        }
        tree = ttk.Treeview(body, columns=columns, show="headings", selectmode="extended")
        tree.grid(row=2, column=0, sticky="nsew")
        for col in columns:
            tree.heading(col, text=headings[col])
            tree.column(col, width=widths[col], anchor="w")
        self._make_tree_sortable(tree, columns, headings)
        scrollbar = tk.Scrollbar(body, orient="vertical", command=tree.yview)
        scrollbar.grid(row=2, column=0, sticky="nse")
        tree.configure(yscrollcommand=scrollbar.set)

        row_map = {}

        def load_rows(filter_text=""):
            for iid in tree.get_children(""):
                tree.delete(iid)
            row_map.clear()
            data = []
            try:
                with sqlite3.connect(DB_PATH) as con:
                    con.row_factory = sqlite3.Row
                    cols = {r[1] for r in con.execute("PRAGMA table_info(tbl_austauschdatenbank)").fetchall()}
                    if "pzn_alt" not in cols:
                        info_var.set("tbl_austauschdatenbank existiert nicht oder hat keine Spalte pzn_alt.")
                        return
                    def expr(name, fallback="''"):
                        return name if name in cols else fallback
                    sql = f"""
                        SELECT id,
                               COALESCE(pzn_alt, '') AS pzn_alt,
                               COALESCE({expr('artikel_alt')}, '') AS artikel_alt,
                               COALESCE({expr('pzn_nmg')}, '') AS pzn_nmg,
                               COALESCE({expr('artikel_nmg')}, '') AS artikel_nmg,
                               COALESCE({expr('freitext_austausch')}, '') AS freitext_austausch,
                               COALESCE({expr('quelle', "'Austauschdatenbank'")}, '') AS quelle,
                               COALESCE({expr('bearbeiter')}, '') AS bearbeiter,
                               COALESCE({expr('erstellt_am', "''")}, '') AS erstellt_am
                          FROM tbl_austauschdatenbank
                         WHERE COALESCE(status,'aktiv') = 'aktiv'
                         ORDER BY id DESC
                    """
                    rows = con.execute(sql).fetchall()
                    for r in rows:
                        d = dict(r)
                        if filter_text:
                            blob = " ".join(str(v or "") for v in d.values()).lower()
                            if filter_text.lower() not in blob:
                                continue
                        data.append(d)
            except Exception as exc:
                info_var.set(f"Biosimilar-Eintraege konnten nicht geladen werden: {exc}")
                return
            for d in data:
                iid = tree.insert("", "end", values=tuple(d.get(c, "") for c in columns))
                row_map[iid] = d
            info_var.set(f"{len(data)} Biosimilar-Eintraege angezeigt.")
            self.status.set(f"Biosimilar: {len(data)} Eintraege.")

        def on_search(*_):
            load_rows(search_var.get().strip())
        search_var.trace_add("write", on_search)

        def edit_row(iid):
            d = row_map.get(iid)
            if not d:
                return
            dlg = tk.Toplevel(self)
            dlg.title("Biosimilar bearbeiten")
            dlg.transient(self)
            dlg.grab_set()
            frm = tk.Frame(dlg, padx=14, pady=14)
            frm.pack(fill="both", expand=True)
            fields = [
                ("PZN alt", "pzn_alt"),
                ("Artikel alt", "artikel_alt"),
                ("PZN NMG", "pzn_nmg"),
                ("Artikel NMG", "artikel_nmg"),
                ("Austauschbar gegen (Freitext)", "freitext_austausch"),
            ]
            entries = {}
            for i, (label, key) in enumerate(fields):
                tk.Label(frm, text=label + ":", anchor="w").grid(row=i, column=0, sticky="w", pady=4)
                e = tk.Entry(frm, width=46)
                e.insert(0, str(d.get(key) or ""))
                e.grid(row=i, column=1, padx=8, pady=4)
                entries[key] = e

            def save():
                new_values = {k: entries[k].get().strip() for k in entries}
                pzn_alt = self._normalize_pzn_input(new_values["pzn_alt"])
                pzn_nmg = self._normalize_pzn_input(new_values["pzn_nmg"])
                if not pzn_alt:
                    messagebox.showwarning("Biosimilar", "PZN alt darf nicht leer sein.")
                    return
                if not pzn_nmg and not new_values["freitext_austausch"]:
                    messagebox.showwarning("Biosimilar", "Bitte PZN NMG oder Austauschbar-gegen-Freitext angeben.")
                    return
                try:
                    with sqlite3.connect(DB_PATH) as con:
                        cols = {r[1] for r in con.execute("PRAGMA table_info(tbl_austauschdatenbank)").fetchall()}
                        assignments = []
                        params = []
                        col_map = {
                            "pzn_alt": pzn_alt,
                            "artikel_alt": new_values["artikel_alt"],
                            "pzn_nmg": pzn_nmg,
                            "artikel_nmg": new_values["artikel_nmg"],
                            "freitext_austausch": new_values["freitext_austausch"],
                        }
                        for col, val in col_map.items():
                            if col in cols:
                                assignments.append(f"{col}=?")
                                params.append(val)
                        if "bearbeiter" in cols:
                            assignments.append("bearbeiter=?")
                            params.append(self.bearbeiter)
                        if "aktualisiert_am" in cols:
                            assignments.append("aktualisiert_am=CURRENT_TIMESTAMP")
                        params.append(int(d["id"]))
                        con.execute(
                            f"UPDATE tbl_austauschdatenbank SET {', '.join(assignments)} WHERE id=?",
                            params,
                        )
                        con.commit()
                    dlg.destroy()
                    load_rows(search_var.get().strip())
                except Exception as exc:
                    messagebox.showerror("Biosimilar", f"Speichern fehlgeschlagen:\n{exc}")

            btns = tk.Frame(frm)
            btns.grid(row=len(fields), column=0, columnspan=2, pady=(12, 0), sticky="e")
            tk.Button(btns, text="Speichern", command=save, bg="#11823b", fg="white", relief="flat", padx=14, pady=6).pack(side="right", padx=(8, 0))
            tk.Button(btns, text="Abbrechen", command=dlg.destroy, padx=14, pady=6).pack(side="right")

        def on_double_click(event):
            iid = tree.identify_row(event.y)
            if iid:
                edit_row(iid)
        tree.bind("<Double-1>", on_double_click)

        def delete_selected():
            sel = tree.selection()
            if not sel:
                messagebox.showinfo("Biosimilar", "Bitte zuerst einen oder mehrere Eintraege auswaehlen.")
                return
            ids = [int(row_map[iid]["id"]) for iid in sel if iid in row_map]
            if not ids:
                return
            if not messagebox.askyesno("Biosimilar loeschen", f"{len(ids)} Eintrag/Eintraege wirklich loeschen?\n\nDas entfernt sie aus der Austauschdatenbank."):
                return
            try:
                with sqlite3.connect(DB_PATH) as con:
                    placeholders = ",".join("?" for _ in ids)
                    con.execute(f"DELETE FROM tbl_austauschdatenbank WHERE id IN ({placeholders})", ids)
                    con.commit()
                load_rows(search_var.get().strip())
            except Exception as exc:
                messagebox.showerror("Biosimilar", f"Loeschen fehlgeschlagen:\n{exc}")

        actionbar = tk.Frame(body, bg="#ffffff")
        actionbar.grid(row=3, column=0, sticky="ew", pady=(12, 0))
        tk.Button(actionbar, text="Aendern (oder Doppelklick)", command=lambda: edit_row(tree.selection()[0]) if tree.selection() else messagebox.showinfo("Biosimilar", "Bitte einen Eintrag waehlen."), padx=14, pady=8).pack(side="left", padx=(0, 8))
        tk.Button(actionbar, text="Loeschen", command=delete_selected, padx=14, pady=8).pack(side="left", padx=8)

        load_rows()


    def show_schulbank_manuelle_pruefung(self):
        """Prüft gelernte Austauschfälle mit mehreren möglichen NMG-PZN/Austauschtexten je PZN alt."""
        self.clear_page()
        self._page_header(
            "Schulbank – Manuelle Prüfung",
            "Zeigt PZN alt, für die mehrere gelernte NMG-PZN oder Austauschtexte vorhanden sind. Jeder darf hier bereinigen."
        )

        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure(0, weight=1)
        body.rowconfigure(1, weight=1)

        info_var = tk.StringVar(value="Mehrdeutige Austausche werden geladen...")
        tk.Label(body, textvariable=info_var, bg="#ffffff", fg="#333", anchor="w", justify="left").grid(row=0, column=0, sticky="ew", pady=(0, 8))

        columns = ("id", "pzn_alt", "artikel_alt", "pzn_nmg", "freitext_austausch", "quelle", "bearbeiter", "erstellt_am")
        headings = {
            "id": "ID",
            "pzn_alt": "PZN alt",
            "artikel_alt": "Artikel alt",
            "pzn_nmg": "PZN NMG",
            "freitext_austausch": "Austauschbar gegen",
            "quelle": "Quelle",
            "bearbeiter": "Bearbeiter",
            "erstellt_am": "Erstellt",
        }
        widths = {
            "id": 55,
            "pzn_alt": 95,
            "artikel_alt": 230,
            "pzn_nmg": 95,
            "freitext_austausch": 260,
            "quelle": 120,
            "bearbeiter": 120,
            "erstellt_am": 145,
        }
        tree = ttk.Treeview(body, columns=columns, show="headings", selectmode="extended")
        tree.grid(row=1, column=0, sticky="nsew")
        for col in columns:
            tree.heading(col, text=headings[col])
            tree.column(col, width=widths[col], anchor="w")
        self._make_tree_sortable(tree, columns, headings)
        scrollbar = tk.Scrollbar(body, orient="vertical", command=tree.yview)
        scrollbar.grid(row=1, column=0, sticky="nse")
        tree.configure(yscrollcommand=scrollbar.set)

        def normalize_pzn_sql_expr(col):
            return f"printf('%08d', CAST(REPLACE(REPLACE(COALESCE({col},''),'.0',''),' ','') AS INTEGER))"

        def load_rows():
            rows = []
            try:
                with sqlite3.connect(DB_PATH) as con:
                    con.row_factory = sqlite3.Row
                    # 1) Mehrdeutige, bereits gelernte Austauschdaten aus der Austauschdatenbank.
                    # Gruppierung erfolgt bewusst in Python über normalisierte PZN, damit 1234567,
                    # 01234567 und Excel-Werte wie 1234567.0 als dieselbe PZN alt gelten.
                    db_rows = con.execute("""\nSELECT
                            id,\nCOALESCE(pzn_alt, '') AS pzn_alt,
                            COALESCE(artikel_alt, '') AS artikel_alt,\nCOALESCE(pzn_nmg, '') AS pzn_nmg,
                            COALESCE(freitext_austausch, '') AS freitext_austausch,\nCOALESCE(quelle, '') AS quelle,
                            COALESCE(bearbeiter, '') AS bearbeiter,\nCOALESCE(erstellt_am, gueltig_ab, '') AS erstellt_am,
                            COALESCE(status, 'aktiv') AS status\nFROM tbl_austauschdatenbank
                        WHERE COALESCE(status, 'aktiv') = 'aktiv'\nAND COALESCE(pzn_alt, '') <> ''
                          AND (COALESCE(pzn_nmg, '') <> '' OR COALESCE(freitext_austausch, '') <> '')\nORDER BY pzn_alt, id DESC
                    """).fetchall()

                    def _norm_text(value):
                        return " ".join(str(value or "").strip().lower().split())

                    grouped = {}
                    for r in db_rows:
                        item = dict(r)
                        key = self._normalize_pzn_input(item.get("pzn_alt"))
                        if not key:
                            continue
                        item["_source_table"] = "tbl_austauschdatenbank"
                        item["_norm_pzn_alt"] = key
                        grouped.setdefault(key, []).append(item)

                    for key, items in grouped.items():
                        variants = {
                            (
                                self._normalize_pzn_input(item.get("pzn_nmg")),
                                _norm_text(item.get("freitext_austausch")),
                            )
                            for item in items
                            if self._normalize_pzn_input(item.get("pzn_nmg")) or _norm_text(item.get("freitext_austausch"))
                        }
                        if len(variants) > 1:
                            rows.extend(items)

                    # 2) Fälle, die im Abweichungsanalyse-Editor bewusst zur manuellen Prüfung geschickt wurden.
                    # Diese sind noch nicht zwingend gelernt und dürfen deshalb nicht nur in der Austauschdatenbank gesucht werden.
                    try:
                        editor_rows = con.execute("""\nSELECT
                                id,\nCOALESCE(pzn, '') AS pzn_alt,
                                COALESCE(artikel, '') AS artikel_alt,\nCOALESCE(MAX(CASE WHEN lower(feld) LIKE '%nmg%' AND lower(feld) LIKE '%pzn%' THEN NULLIF(manuell, '') END),
                                         MAX(CASE WHEN lower(feld) LIKE '%nmg%' AND lower(feld) LIKE '%pzn%' THEN NULLIF(programm, '') END),\n'') AS pzn_nmg,
                                COALESCE(MAX(CASE WHEN lower(feld) LIKE '%austausch%' THEN NULLIF(manuell, '') END),\nMAX(CASE WHEN lower(feld) LIKE '%austausch%' THEN NULLIF(programm, '') END),
                                         '') AS freitext_austausch,\n'Abweichungseditor' AS quelle,
                                COALESCE(bearbeiter, '') AS bearbeiter,\nCOALESCE(bearbeitet_am, erstellt_am, '') AS erstellt_am,
                                COALESCE(status, '') AS status\nFROM tbl_abweichungs_editor
                            WHERE COALESCE(status, '') = 'manuelle_pruefung'\nGROUP BY quelle_datei, pzn
                            ORDER BY datetime(COALESCE(bearbeitet_am, erstellt_am)) DESC, id DESC\n""").fetchall()
                        for r in editor_rows:
                            item = dict(r)
                            item["_source_table"] = "tbl_abweichungs_editor"
                            # Nur anzeigen, wenn wenigstens eine verwertbare Zielangabe vorhanden ist.
                            if str(item.get("pzn_nmg") or "").strip() or str(item.get("freitext_austausch") or "").strip():
                                rows.append(item)
                    except Exception:
                        pass
            except Exception as exc:
                messagebox.showerror("Manuelle Prüfung", f"Mehrdeutige Austausche konnten nicht geladen werden:\n{exc}")
            return rows

        row_map = {}
        def reload():
            for iid in tree.get_children(""):
                tree.delete(iid)
            row_map.clear()
            rows = load_rows()
            for row in rows:
                iid = tree.insert("", "end", values=tuple(row[col] if col in row.keys() else "" for col in columns))
                row_map[iid] = dict(row)
            pzn_count = len({str(r["pzn_alt"]) for r in rows})
            info_var.set(f"{len(rows)} gelernte Austausch-Einträge zu {pzn_count} PZN alt mit mehreren Varianten gefunden.")
            self.status.set(f"Manuelle Prüfung: {len(rows)} Einträge angezeigt.")

        def selected_rows():
            selected = []
            for iid in tree.selection():
                row = row_map.get(iid)
                if row:
                    selected.append(row)
            if not selected:
                messagebox.showinfo("Manuelle Prüfung", "Bitte zuerst einen Eintrag auswählen.")
            return selected

        def keep_selected_variant():
            selected = selected_rows()
            if not selected:
                return
            # Pro PZN alt darf genau ein Eintrag gewählt werden, damit nicht versehentlich mehrere Gruppen bereinigt werden.
            pzn_values = {str(r.get("pzn_alt") or "").strip() for r in selected}
            if len(pzn_values) != 1:
                messagebox.showinfo("Manuelle Prüfung", "Bitte nur Einträge zu einer PZN alt auswählen.")
                return
            if len(selected) != 1:
                messagebox.showinfo("Manuelle Prüfung", "Bitte genau den Eintrag markieren, der aktiv bleiben soll.")
                return
            keep = selected[0]
            if keep.get("_source_table") != "tbl_austauschdatenbank":
                messagebox.showinfo("Manuelle Prüfung", "Diese Zeile kommt aus dem Abweichungseditor und ist noch kein gelernter Austausch. Bitte im Abweichungseditor als Lernvorschlag speichern oder dort weiter bearbeiten.")
                return
            pzn_alt = str(keep.get("pzn_alt") or "").strip()
            keep_id = int(keep.get("id"))
            if not messagebox.askyesno(
                "Austausch bereinigen",
                f"Für PZN alt {pzn_alt} bleibt Eintrag {keep_id} aktiv.\n"
                "Alle anderen aktiven Varianten zu dieser PZN werden auf inaktiv gesetzt.\n\nFortfahren?"
            ):
                return
            try:
                with sqlite3.connect(DB_PATH) as con:
                    cols = {r[1] for r in con.execute("PRAGMA table_info(tbl_austauschdatenbank)").fetchall()}
                    assignments = "status='inaktiv'"
                    params = []
                    if "gueltig_bis" in cols:
                        assignments += ", gueltig_bis=CURRENT_TIMESTAMP"
                    if "ersetzt_durch_id" in cols:
                        assignments += ", ersetzt_durch_id=?"
                        params.append(keep_id)
                    if "bearbeiter" in cols:
                        assignments += ", bearbeiter=?"
                        params.append(self.bearbeiter)
                    params.extend([pzn_alt, keep_id])
                    con.execute(
                        f"UPDATE tbl_austauschdatenbank SET {assignments} WHERE COALESCE(status,'aktiv')='aktiv' AND pzn_alt=? AND id<>?",
                        params,
                    )
                    con.commit()
                self.status.set(f"Manuelle Prüfung: PZN {pzn_alt} bereinigt.")
                reload()
            except Exception as exc:
                messagebox.showerror("Manuelle Prüfung", f"Bereinigung fehlgeschlagen:\n{exc}")

        def open_schulbank_uebernommen():
            self.show_schulbank_page("Übernommen")

        def set_rows_inactive(rows_to_update, label):
            if not rows_to_update:
                messagebox.showinfo("Manuelle Prüfung", "Keine Einträge ausgewählt.")
                return
            austausch_ids = []
            editor_ids = []
            for row in rows_to_update:
                try:
                    if row.get("_source_table") == "tbl_abweichungs_editor":
                        editor_ids.append(int(row.get("id")))
                    else:
                        austausch_ids.append(int(row.get("id")))
                except Exception:
                    pass
            if not austausch_ids and not editor_ids:
                messagebox.showinfo("Manuelle Prüfung", "Keine gültigen Einträge ausgewählt.")
                return
            if not messagebox.askyesno("Manuelle Prüfung", f"{len(austausch_ids) + len(editor_ids)} Eintrag/Einträge wirklich {label}?"):
                return
            try:
                with sqlite3.connect(DB_PATH) as con:
                    if austausch_ids:
                        cols = {r[1] for r in con.execute("PRAGMA table_info(tbl_austauschdatenbank)").fetchall()}
                        assignments = "status='inaktiv'"
                        params = []
                        if "gueltig_bis" in cols:
                            assignments += ", gueltig_bis=CURRENT_TIMESTAMP"
                        if "bearbeiter" in cols:
                            assignments += ", bearbeiter=?"
                            params.append(self.bearbeiter)
                        placeholders = ",".join("?" for _ in austausch_ids)
                        con.execute(f"UPDATE tbl_austauschdatenbank SET {assignments} WHERE id IN ({placeholders})", params + austausch_ids)
                    if editor_ids:
                        placeholders = ",".join("?" for _ in editor_ids)
                        con.execute(f"UPDATE tbl_abweichungs_editor SET status='ignoriert', bearbeitet_am=CURRENT_TIMESTAMP, bearbeiter=? WHERE id IN ({placeholders})", [self.bearbeiter] + editor_ids)
                    con.commit()
                self.status.set(f"Manuelle Prüfung: {len(austausch_ids) + len(editor_ids)} Einträge {label}.")
                reload()
            except Exception as exc:
                messagebox.showerror("Manuelle Prüfung", f"Aktion fehlgeschlagen:\n{exc}")

        def reject_selected_variants():
            set_rows_inactive(selected_rows(), "abgelehnt/inaktiv gesetzt")

        def reject_all_visible_variants():
            set_rows_inactive(list(row_map.values()), "abgelehnt/inaktiv gesetzt")

        actionbar = tk.Frame(body, bg="#ffffff")
        actionbar.grid(row=2, column=0, sticky="ew", pady=(12, 0))
        tk.Button(actionbar, text="✓ Gewählten behalten", command=keep_selected_variant, bg="#11823b", fg="white", relief="flat", padx=14, pady=8).pack(side="left", padx=(0, 8))
        tk.Button(actionbar, text="✗ Markierte ablehnen", command=reject_selected_variants, bg="#9b1c1c", fg="white", relief="flat", padx=14, pady=8).pack(side="left", padx=8)
        tk.Button(actionbar, text="✗ Alle sichtbaren ablehnen", command=reject_all_visible_variants, bg="#9b1c1c", fg="white", relief="flat", padx=14, pady=8).pack(side="left", padx=8)
        tk.Button(actionbar, text="Aktualisieren", command=reload, padx=14, pady=8).pack(side="left", padx=8)
        tk.Button(actionbar, text="Zu Übernommen", command=open_schulbank_uebernommen, padx=14, pady=8).pack(side="left", padx=8)
        reload()

    def _ensure_lernvorschlaege_table(self):
        with sqlite3.connect(DB_PATH) as con:
            con.execute("""\nCREATE TABLE IF NOT EXISTS tbl_lernvorschlaege (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,\nprodukt_alt TEXT NOT NULL DEFAULT '',
                    produkt_neu TEXT NOT NULL DEFAULT '',\npzn_alt TEXT,
                    artikel_alt TEXT,\npzn_nmg TEXT,
                    freitext_austausch TEXT,\nquelle_datei TEXT,
                    status TEXT NOT NULL DEFAULT 'neu',\nerstellt_am TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    bearbeitet_am TEXT,\nbearbeiter TEXT,
                    historie_ab TEXT,\nhistorie_bis TEXT,
                    austausch_id INTEGER\n)
            """)
            cols = {row[1] for row in con.execute("PRAGMA table_info(tbl_lernvorschlaege)").fetchall()}
            additions = {
                "pzn_alt": "TEXT",
                "artikel_alt": "TEXT",
                "pzn_nmg": "TEXT",
                "freitext_austausch": "TEXT",
                "quelle_datei": "TEXT",
                "bearbeiter": "TEXT",
                "historie_ab": "TEXT",
                "historie_bis": "TEXT",
                "austausch_id": "INTEGER",
            }
            for col, definition in additions.items():
                if col not in cols:
                    con.execute(f"ALTER TABLE tbl_lernvorschlaege ADD COLUMN {col} {definition}")
            con.execute("UPDATE tbl_lernvorschlaege SET artikel_alt = COALESCE(NULLIF(artikel_alt, ''), produkt_alt) WHERE artikel_alt IS NULL OR artikel_alt = ''")
            con.execute("UPDATE tbl_lernvorschlaege SET freitext_austausch = COALESCE(NULLIF(freitext_austausch, ''), produkt_neu) WHERE freitext_austausch IS NULL OR freitext_austausch = ''")
            con.commit()

        try:
            self._roadmap_mark_change_erledigt()
        except Exception:
            pass

    def _roadmap_mark_change_erledigt(self):
        """Hält die Roadmap zur aktuellen Schulbank-Änderung aktuell, ohne Duplikate zu erzeugen."""
        title = "Schulbank Entscheidungsspeicher und Austauschdatenbank-Lernen"
        rows = list_roadmap_items()
        for row in rows:
            if str(row["titel"]).strip().lower() == title.lower():
                if row["status"] != "Erledigt":
                    update_roadmap_status(int(row["id"]), "Erledigt")
                return
        add_roadmap_item(
            bereich="Schulbank",
            titel=title,
            beschreibung=(
                "tbl_lernvorschlaege erweitert, Bearbeiter gespeichert, Historie-Regel vorbereitet, "
                "Lernen übernimmt direkt in tbl_austauschdatenbank und alte Austauschdatensätze werden nicht gelöscht."
            ),
            status="Erledigt",
            prioritaet="Hoch",
        )

    def _roadmap_mark_schulbank_bulk_erledigt(self):
        """Dokumentiert Mehrfachauswahl, Sammelaktionen und Dublettenschutz in der Roadmap."""
        title = "Schulbank Mehrfachauswahl und Dublettenschutz"
        try:
            rows = list_roadmap_items()
            for row in rows:
                if str(row["titel"]).strip().lower() == title.lower():
                    if row["status"] != "Erledigt":
                        update_roadmap_status(int(row["id"]), "Erledigt")
                    return
            add_roadmap_item(
                bereich="Schulbank",
                titel=title,
                beschreibung=(
                    "Schulbank-Tabellen erlauben Mehrfachauswahl. Markierte oder alle sichtbaren Vorschläge "
                    "können gesammelt gelernt oder abgelehnt werden. Abweichungsanalyse erzeugt keine doppelten "
                    "Lernvorschläge mehr für bereits vorhandene Fälle."
                ),
                status="Erledigt",
                prioritaet="Hoch",
            )
        except Exception:
            pass

    def get_lernvorschlaege(self, status_filter=None):
        """Liest Schulbank-Einträge nach der aktuellen Fachlogik.\n
        Neue Lernvorschläge: nur echte neue Fälle, nicht übernommen, nicht abgelehnt,\nnicht bereits aktiv gelernt und nicht in der manuellen Prüfung.
\nÜbernommen/Abgelehnt: bleiben 7 Tage sichtbar. Danach erscheinen sie in der Historie.
        Übernommene Fälle werden zusätzlich robust aus der Austauschdatenbank gelesen, weil\nLernen fachlich dort gespeichert wird.
        """
        self._ensure_lernvorschlaege_table()

        def _rows(sql, params=()):
            try:
                with sqlite3.connect(DB_PATH) as con:
                    con.row_factory = sqlite3.Row
                    return con.execute(sql, params).fetchall()
            except Exception:
                return []

        def _table_cols(table):
            try:
                with sqlite3.connect(DB_PATH) as con:
                    return {r[1] for r in con.execute(f"PRAGMA table_info({table})").fetchall()}
            except Exception:
                return set()

        # SP18: source_table mitliefern, damit delete_selected weiss, aus
        # welcher Tabelle der angezeigte Eintrag wirklich kommt.
        base_select = """\nSELECT id,
                   COALESCE(produkt_alt, '') AS produkt_alt,\nCOALESCE(produkt_neu, '') AS produkt_neu,
                   COALESCE(pzn_alt, '') AS pzn_alt,\nCOALESCE(artikel_alt, produkt_alt, '') AS artikel_alt,
                   COALESCE(pzn_nmg, '') AS pzn_nmg,\nCOALESCE(freitext_austausch, produkt_neu, '') AS freitext_austausch,
                   COALESCE(quelle_datei, '') AS quelle_datei,\nstatus,
                   COALESCE(bearbeiter, '') AS bearbeiter,\nerstellt_am,
                   bearbeitet_am,\nhistorie_ab,
                   historie_bis,\naustausch_id,
                   'tbl_lernvorschlaege' AS source_table
            FROM tbl_lernvorschlaege\n"""
        valid_filter = """\nCOALESCE(pzn_alt, '') <> ''
            AND (COALESCE(pzn_nmg, '') <> '' OR COALESCE(freitext_austausch, produkt_neu, '') <> '')\nAND (COALESCE(pzn_nmg, '') = '' OR COALESCE(pzn_alt, '') <> COALESCE(pzn_nmg, ''))
        """
        recent_7 = "datetime(COALESCE(bearbeitet_am, erstellt_am)) > datetime('now', '-7 days')"
        older_7 = "datetime(COALESCE(bearbeitet_am, erstellt_am)) <= datetime('now', '-7 days')"

        def _austausch_rows(recent=None):
            cols = _table_cols("tbl_austauschdatenbank")
            if not cols or "pzn_alt" not in cols:
                return []
            def expr(col, fallback="''"):
                return f"COALESCE({col}, '')" if col in cols else fallback
            produkt_alt = expr("artikel_alt")
            artikel_alt = expr("artikel_alt")
            pzn_nmg = expr("pzn_nmg")
            artikel_nmg = expr("artikel_nmg")
            freitext = expr("freitext_austausch", artikel_nmg)
            quelle = expr("quelle", "'Austauschdatenbank'")
            bearbeiter = expr("bearbeiter")
            date_candidates = [c for c in ("aktualisiert_am", "bearbeitet_am", "gueltig_ab", "erstellt_am") if c in cols]
            date_expr = "COALESCE(" + ", ".join(date_candidates + ["''"]) + ")" if date_candidates else "''"
            status_expr = "COALESCE(status, 'aktiv')" if "status" in cols else "'aktiv'"
            date_where = ""
            if recent is True:
                date_where = f"AND datetime({date_expr}) > datetime('now', '-7 days')"
            elif recent is False:
                date_where = f"AND ({date_expr} = '' OR datetime({date_expr}) <= datetime('now', '-7 days'))"
            # SP18: source_table-Marker, damit der Loeschen-Knopf weiss
            # aus welcher Tabelle der Eintrag wirklich kommt.
            sql = f"""\nSELECT id,
                       {produkt_alt} AS produkt_alt,\nCOALESCE({freitext}, {artikel_nmg}, '') AS produkt_neu,
                       COALESCE(pzn_alt, '') AS pzn_alt,\n{artikel_alt} AS artikel_alt,
                       {pzn_nmg} AS pzn_nmg,\nCOALESCE({freitext}, {artikel_nmg}, '') AS freitext_austausch,
                       {quelle} AS quelle_datei,\n'uebernommen' AS status,
                       {bearbeiter} AS bearbeiter,\n{date_expr} AS erstellt_am,
                       {date_expr} AS bearbeitet_am,\nNULL AS historie_ab,
                       NULL AS historie_bis,\nid AS austausch_id,
                       'tbl_austauschdatenbank' AS source_table
                FROM tbl_austauschdatenbank\nWHERE {status_expr} = 'aktiv'
                  AND COALESCE(pzn_alt, '') <> ''\nAND ({pzn_nmg} <> '' OR COALESCE({freitext}, {artikel_nmg}, '') <> '')
                  AND ({pzn_nmg} = '' OR COALESCE(pzn_alt, '') <> {pzn_nmg})\n{date_where}
                ORDER BY id DESC\n"""
            return _rows(sql)

        def _combine_unique(*row_lists):
            combined = []
            seen = set()
            for row_list in row_lists:
                for row in row_list:
                    key = (
                        str(row["status"] or "").strip(),
                        str(row["pzn_alt"] or "").strip(),
                        str(row["pzn_nmg"] or "").strip(),
                        str(row["freitext_austausch"] or "").strip().lower(),
                    )
                    if key in seen:
                        continue
                    seen.add(key)
                    combined.append(row)
            return combined

        if status_filter == "uebernommen":
            # SP19: Uebernommen zeigt nur noch manuelle Lernvorschlaege,
            # die Biosimilar-/Austauschdatenbank-Eintraege bekommen ihren
            # eigenen Bereich.
            return _rows(base_select + f"""\nWHERE status = 'uebernommen'
                  AND {recent_7}\nAND {valid_filter}
                ORDER BY datetime(COALESCE(bearbeitet_am, erstellt_am)) DESC, id DESC\n""")

        if status_filter == "abgelehnt":
            # Abgelehnte Fälle sollen auch dann sichtbar bleiben, wenn sie fachlich unvollständig waren.
            return _rows(base_select + f"""\nWHERE status = 'abgelehnt'
                  AND {recent_7}\nORDER BY datetime(COALESCE(bearbeitet_am, erstellt_am)) DESC, id DESC
            """)

        if status_filter in ("historie", "historie_uebernommen"):
            # SP19: Historie der manuell gelernten Vorschlaege bleibt sortenrein.
            return _rows(base_select + f"""\nWHERE status = 'uebernommen'
                  AND {older_7}\nAND {valid_filter}
                ORDER BY datetime(COALESCE(bearbeitet_am, erstellt_am)) DESC, id DESC\n""")

        if status_filter == "historie_abgelehnt":
            return _rows(base_select + f"""\nWHERE status = 'abgelehnt'
                  AND {older_7}\nORDER BY datetime(COALESCE(bearbeitet_am, erstellt_am)) DESC, id DESC
            """)

        if status_filter == "neu":
            return _rows(base_select + f"""\nWHERE status = 'neu'
                  AND {valid_filter}\nAND NOT EXISTS (
                      SELECT 1\nFROM tbl_austauschdatenbank a
                      WHERE COALESCE(a.status, 'aktiv') = 'aktiv'\nAND COALESCE(a.pzn_alt, '') = COALESCE(tbl_lernvorschlaege.pzn_alt, '')
                  )\nAND NOT EXISTS (
                      WITH basis AS (\nSELECT pzn_alt,
                                 TRIM(COALESCE(pzn_nmg, '') || '|' || COALESCE(freitext_austausch, '')) AS kombi\nFROM tbl_austauschdatenbank
                          WHERE COALESCE(status, 'aktiv') = 'aktiv'\nAND COALESCE(pzn_alt, '') <> ''
                            AND (COALESCE(pzn_nmg, '') <> '' OR COALESCE(freitext_austausch, '') <> '')\n), mehrfach AS (
                          SELECT pzn_alt\nFROM basis
                          GROUP BY pzn_alt\nHAVING COUNT(DISTINCT kombi) > 1
                      )\nSELECT 1 FROM mehrfach m
                      WHERE m.pzn_alt = COALESCE(tbl_lernvorschlaege.pzn_alt, '')\n)
                ORDER BY datetime(COALESCE(bearbeitet_am, erstellt_am)) DESC, id DESC\n""")

        if status_filter:
            return _rows(base_select + f"""\nWHERE status = ?
                  AND {valid_filter}\nORDER BY datetime(COALESCE(bearbeitet_am, erstellt_am)) DESC, id DESC
            """, (status_filter,))

        return _rows(base_select + f"""\nWHERE status = 'neu'
              AND {valid_filter}\nORDER BY datetime(COALESCE(bearbeitet_am, erstellt_am)) DESC, id DESC
        """)

    def _normalize_pzn_input(self, value):
        text = str(value or "").strip()
        # SP29: SK-Importdateien tragen ein Lagervariante-Suffix wie ' /1' an
        # der PZN ('12457880 /1'). Vorher wurde die Suffix-Ziffer mit ueber den
        # isdigit()-Filter in die PZN gezogen, was zu 9-stelligen Phantom-PZNs
        # fuehrte und alle PK-Rabatt-Lookups stumm scheitern liess.
        if "/" in text:
            text = text.split("/", 1)[0].strip()
        if text.endswith(".0") and text[:-2].isdigit():
            text = text[:-2]
        digits = "".join(ch for ch in text if ch.isdigit())
        return digits.zfill(8) if digits else ""

    def _find_artikelname_by_pzn(self, pzn):
        pzn = self._normalize_pzn_input(pzn)
        if not pzn:
            return ""
        with sqlite3.connect(DB_PATH) as con:
            con.row_factory = sqlite3.Row
            candidates = [
                ("tbl_artikelstamm", "artikel"),
                ("tbl_artikelstamm", "artikelname"),
                ("tbl_pzn_basisdaten", "artikelname"),
                ("tbl_nmg_stamm", "artikelname"),
            ]
            for table, col in candidates:
                try:
                    row = con.execute(f"SELECT {col} AS artikel FROM {table} WHERE pzn = ? LIMIT 1", (pzn,)).fetchone()
                    if row and row["artikel"]:
                        return str(row["artikel"]).strip()
                except Exception:
                    continue
        return ""

    def _roadmap_mark_lernvorschlag_form_erledigt(self):
        """Dokumentiert die Formular-Umstellung in der Roadmap ohne doppelte Einträge."""
        title = "Schulbank Lernvorschlag-Formular im Arbeitsbereich"
        try:
            rows = list_roadmap_items()
            for row in rows:
                if str(row["titel"]).strip().lower() == title.lower():
                    if row["status"] != "Erledigt":
                        update_roadmap_status(int(row["id"]), "Erledigt")
                    return
            add_roadmap_item(
                bereich="Schulbank",
                titel=title,
                beschreibung=(
                    "Lernvorschlag hinzufügen wurde von mehreren Einzelabfragen auf ein Formular umgestellt. "
                    "Abgegebene PZN und PZN NMG füllen Artikelnamen automatisch aus; Austauschbar gegen ist optional, "
                    "wenn PZN NMG vorhanden ist. Vorhandene Fälle werden vor dem Speichern geprüft."
                ),
                status="Erledigt",
                prioritaet="Hoch",
            )
        except Exception:
            pass

    def _insert_or_update_lernvorschlag(self, pzn_alt, artikel_alt, pzn_nmg, freitext, section="Neue Lernvorschläge"):
        self._ensure_lernvorschlaege_table()
        pzn_alt = self._normalize_pzn_input(pzn_alt)
        pzn_nmg = self._normalize_pzn_input(pzn_nmg)
        artikel_alt = str(artikel_alt or "").strip()
        freitext = str(freitext or "").strip()

        if not artikel_alt and pzn_alt:
            artikel_alt = self._find_artikelname_by_pzn(pzn_alt)
        if not freitext and pzn_nmg:
            freitext = self._find_artikelname_by_pzn(pzn_nmg) or f"PZN NMG: {pzn_nmg}"

        if not pzn_alt:
            messagebox.showwarning("Lernvorschlag", "Bitte eine abgegebene PZN eingeben.")
            return False
        if not pzn_nmg and not freitext:
            messagebox.showwarning("Lernvorschlag", "Bitte entweder eine PZN NMG oder einen Eintrag bei 'Austauschbar gegen' eingeben.")
            return False

        with sqlite3.connect(DB_PATH) as con:
            con.row_factory = sqlite3.Row
            existing_suggestion = con.execute("""\nSELECT id, status, pzn_alt, artikel_alt, pzn_nmg, freitext_austausch
                FROM tbl_lernvorschlaege\nWHERE COALESCE(pzn_alt, '') = ?
                  AND COALESCE(pzn_nmg, '') = ?\nORDER BY id DESC
                LIMIT 1\n""", (pzn_alt, pzn_nmg)).fetchone()

            if existing_suggestion:
                msg = (
                    "Dieser Lernfall existiert bereits in der Schulbank.\n\n"
                    f"Status: {existing_suggestion['status']}\n"
                    f"PZN alt: {existing_suggestion['pzn_alt'] or ''}\n"
                    f"PZN NMG: {existing_suggestion['pzn_nmg'] or ''}\n\n"
                    "Soll der vorhandene Lernvorschlag überschrieben und wieder auf 'neu' gesetzt werden?"
                )
                if not messagebox.askyesno("Fall existiert bereits", msg):
                    return False
                con.execute("""\nUPDATE tbl_lernvorschlaege
                    SET produkt_alt = ?, produkt_neu = ?, artikel_alt = ?, freitext_austausch = ?,\nquelle_datei = 'Manuell', status = 'neu', bearbeiter = ?, bearbeitet_am = CURRENT_TIMESTAMP
                    WHERE id = ?\n""", (artikel_alt, freitext, artikel_alt, freitext, self.bearbeiter, existing_suggestion["id"]))
                con.commit()
                return True

            existing_active = None
            try:
                existing_active = con.execute("""\nSELECT id, pzn_alt, pzn_nmg, freitext_austausch, quelle
                    FROM tbl_austauschdatenbank\nWHERE status = 'aktiv'
                      AND COALESCE(pzn_alt, '') = ?\nORDER BY id DESC
                    LIMIT 1\n""", (pzn_alt,)).fetchone()
            except Exception:
                existing_active = None

            if existing_active:
                msg = (
                    "Für diese abgegebene PZN gibt es bereits einen aktiven Eintrag in der Austauschdatenbank.\n\n"
                    f"Aktuell: {existing_active['freitext_austausch'] or ''}\n"
                    f"PZN NMG: {existing_active['pzn_nmg'] or ''}\n\n"
                    "Soll trotzdem ein neuer Lernvorschlag angelegt werden? Beim späteren Lernen wird der alte Eintrag nicht gelöscht, sondern inaktiv gesetzt."
                )
                if not messagebox.askyesno("Austausch bereits gelernt", msg):
                    return False

            con.execute("""\nINSERT INTO tbl_lernvorschlaege (
                    produkt_alt, produkt_neu, pzn_alt, artikel_alt, pzn_nmg,\nfreitext_austausch, quelle_datei, status, erstellt_am, bearbeiter
                )\nVALUES (?, ?, ?, ?, ?, ?, 'Manuell', 'neu', CURRENT_TIMESTAMP, ?)
            """, (artikel_alt, freitext, pzn_alt, artikel_alt, pzn_nmg, freitext, self.bearbeiter))
            con.commit()
            return True

    def add_lernvorschlag(self, section="Neue Lernvorschläge"):
        self._ensure_lernvorschlaege_table()
        self._roadmap_mark_lernvorschlag_form_erledigt()

        win = tk.Toplevel(self)
        win.resizable(True, True)
        win.title("Lernvorschlag hinzufügen")
        win.geometry("720x430")
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()

        tk.Label(win, text="Lernvorschlag hinzufügen", font=(theme.FONT, 20, "bold"), fg="#0b4a86", bg="#f5f7fb").pack(anchor="w", padx=24, pady=(22, 4))
        tk.Label(
            win,
            text="Abgegebene PZN erfassen. Artikelnamen werden automatisch aus der Datenbank gefüllt, wenn vorhanden.",
            font=(theme.FONT, 10), fg="#333", bg="#f5f7fb"
        ).pack(anchor="w", padx=24, pady=(0, 14))

        form = tk.Frame(win, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        form.pack(fill="both", expand=True, padx=24, pady=(0, 14))
        form.columnconfigure(1, weight=1)

        pzn_alt_var = tk.StringVar()
        artikel_alt_var = tk.StringVar()
        pzn_nmg_var = tk.StringVar()
        artikel_nmg_var = tk.StringVar()
        freitext_var = tk.StringVar()
        hint_var = tk.StringVar(value="Austauschbar gegen ist optional, sobald eine PZN NMG eingetragen ist.")

        def add_row(row, label, var, readonly=False):
            tk.Label(form, text=label, bg="#ffffff", fg="#0b4a86", font=(theme.FONT, 10, "bold")).grid(row=row, column=0, sticky="w", padx=18, pady=(14 if row == 0 else 8, 4))
            state = "readonly" if readonly else "normal"
            entry = tk.Entry(form, textvariable=var, state=state, font=(theme.FONT, 11))
            entry.grid(row=row, column=1, sticky="ew", padx=18, pady=(14 if row == 0 else 8, 4))
            return entry

        e_pzn_alt = add_row(0, "Abgegebene PZN", pzn_alt_var)
        add_row(1, "Artikelname abgegeben", artikel_alt_var, readonly=True)
        e_pzn_nmg = add_row(2, "Austauschbar gegen PZN NMG", pzn_nmg_var)
        add_row(3, "Artikelname PZN NMG", artikel_nmg_var, readonly=True)
        e_freitext = add_row(4, "Austauschbar gegen", freitext_var)

        tk.Label(form, textvariable=hint_var, bg="#ffffff", fg="#666", justify="left", wraplength=610).grid(row=5, column=0, columnspan=2, sticky="w", padx=18, pady=(10, 14))

        def set_readonly_var(var, value):
            var.set(value or "")

        def refresh_alt(*_):
            pzn = self._normalize_pzn_input(pzn_alt_var.get())
            set_readonly_var(artikel_alt_var, self._find_artikelname_by_pzn(pzn) if pzn else "")

        def refresh_nmg(*_):
            pzn = self._normalize_pzn_input(pzn_nmg_var.get())
            artikel = self._find_artikelname_by_pzn(pzn) if pzn else ""
            set_readonly_var(artikel_nmg_var, artikel)
            if pzn and not freitext_var.get().strip() and artikel:
                freitext_var.set(artikel)
            if pzn:
                hint_var.set("PZN NMG ist gefüllt. 'Austauschbar gegen' darf leer bleiben; dann wird der Artikelname der PZN NMG übernommen.")
            else:
                hint_var.set("Keine PZN NMG eingetragen. Dann muss 'Austauschbar gegen' ausgefüllt werden.")

        pzn_alt_var.trace_add("write", refresh_alt)
        pzn_nmg_var.trace_add("write", refresh_nmg)

        def save():
            pzn_alt = self._normalize_pzn_input(pzn_alt_var.get())
            pzn_nmg = self._normalize_pzn_input(pzn_nmg_var.get())
            artikel_alt = artikel_alt_var.get().strip()
            freitext = freitext_var.get().strip()
            if not freitext and pzn_nmg:
                freitext = artikel_nmg_var.get().strip() or self._find_artikelname_by_pzn(pzn_nmg) or f"PZN NMG: {pzn_nmg}"

            if self._insert_or_update_lernvorschlag(pzn_alt, artikel_alt, pzn_nmg, freitext, section):
                win.destroy()
                self.show_schulbank_page(section if section != "Schulbank" else "Neue Lernvorschläge")

        buttons = tk.Frame(win, bg="#f5f7fb")
        buttons.pack(fill="x", padx=24, pady=(0, 20))
        tk.Button(buttons, text="Abbrechen", command=win.destroy, padx=16, pady=8).pack(side="right", padx=(8, 0))
        tk.Button(buttons, text="Speichern", command=save, bg="#0b4a86", fg="white", activebackground="#0b4a86", relief="flat", font=(theme.FONT, 11, "bold"), padx=22, pady=9).pack(side="right")

        e_pzn_alt.focus_set()
        win.bind("<Return>", lambda _event: save())
        self.wait_window(win)

    def edit_lernvorschlag(self, item_id, section="Schulbank"):
        if item_id is None:
            return
        self._ensure_lernvorschlaege_table()
        with sqlite3.connect(DB_PATH) as con:
            con.row_factory = sqlite3.Row
            row = con.execute("SELECT * FROM tbl_lernvorschlaege WHERE id = ?", (item_id,)).fetchone()
            if not row:
                messagebox.showinfo("Schulbank", "Eintrag wurde nicht gefunden.")
                return

        pzn_alt = simpledialog.askstring("Bearbeiten", "PZN alt:", initialvalue=row["pzn_alt"] or "")
        if pzn_alt is None:
            return
        artikel_alt = simpledialog.askstring("Bearbeiten", "Artikel alt:", initialvalue=row["artikel_alt"] or row["produkt_alt"] or "")
        if artikel_alt is None:
            return
        pzn_nmg = simpledialog.askstring("Bearbeiten", "PZN NMG:", initialvalue=row["pzn_nmg"] or "")
        if pzn_nmg is None:
            return
        freitext = simpledialog.askstring("Bearbeiten", "Austauschbar gegen:", initialvalue=row["freitext_austausch"] or row["produkt_neu"] or "")
        if freitext is None:
            return
        if not freitext and pzn_nmg:
            freitext = self._find_artikelname_by_pzn(pzn_nmg) or f"PZN NMG: {pzn_nmg}"

        with sqlite3.connect(DB_PATH) as con:
            con.execute("""\nUPDATE tbl_lernvorschlaege
                SET pzn_alt = ?, artikel_alt = ?, pzn_nmg = ?, freitext_austausch = ?,\nprodukt_alt = ?, produkt_neu = ?, bearbeiter = ?, bearbeitet_am = CURRENT_TIMESTAMP
                WHERE id = ?\n""", (pzn_alt.strip(), artikel_alt.strip(), pzn_nmg.strip(), freitext.strip(), artikel_alt.strip(), freitext.strip(), self.bearbeiter, item_id))
            con.commit()
        self.show_schulbank_page(section)

    def update_lernvorschlag_status(self, item_id, status):
        self._ensure_lernvorschlaege_table()
        result = None

        with sqlite3.connect(DB_PATH) as con:
            con.row_factory = sqlite3.Row
            row = con.execute("SELECT * FROM tbl_lernvorschlaege WHERE id = ?", (item_id,)).fetchone()
            if not row:
                raise ValueError("Lernvorschlag wurde nicht gefunden.")

        pzn_alt = row["pzn_alt"] or ""
        artikel_alt = row["artikel_alt"] or row["produkt_alt"] or ""
        pzn_nmg = row["pzn_nmg"] or ""
        freitext = row["freitext_austausch"] or row["produkt_neu"] or ""
        artikel_nmg = self._find_artikelname_by_pzn(pzn_nmg) if pzn_nmg else ""
        if status == "uebernommen":
            if not freitext and pzn_nmg:
                freitext = artikel_nmg or f"PZN NMG: {pzn_nmg}"
            result = add_austausch_entry(
                pzn_alt=pzn_alt,
                pzn_nmg=pzn_nmg,
                freitext_austausch=freitext,
                quelle="Schulbank",
                artikel_alt=artikel_alt,
                artikel_nmg=artikel_nmg,
                bearbeiter=self.bearbeiter,
                bemerkung=f"Übernommen aus Schulbank-Vorschlag {item_id}",
            )

        with sqlite3.connect(DB_PATH) as con:
            con.execute("""\nUPDATE tbl_lernvorschlaege
                SET status = ?,\nbearbeitet_am = CURRENT_TIMESTAMP,
                    bearbeiter = ?,\nfreitext_austausch = ?,
                    produkt_neu = ?,\nhistorie_ab = CASE WHEN ? IN ('uebernommen', 'abgelehnt') THEN datetime('now', '+28 days') ELSE NULL END,
                    historie_bis = CASE WHEN ? IN ('uebernommen', 'abgelehnt') THEN datetime('now', '+6 months') ELSE NULL END,\naustausch_id = COALESCE(?, austausch_id)
                WHERE id = ?\n""", (status, self.bearbeiter, freitext, freitext, status, status, result.get("id") if result else None, item_id))
            con.commit()
        if result:
            return {"austausch_id": result.get("id"), "created": result.get("created")}
        return None

    def delete_lernvorschlag(self, item_id):
        self._ensure_lernvorschlaege_table()
        with sqlite3.connect(DB_PATH) as con:
            con.execute("DELETE FROM tbl_lernvorschlaege WHERE id = ?", (item_id,))
            con.commit()

    def open_schulbank(self):
        self.show_schulbank_page("Schulbank")

    def show_daten_aktualisieren_page(self):
        # SP17: Layout auf 3x3 erweitert, Austauschdatenbank-Kachel ergaenzt
        # (die fehlte seit SP7 und der Import war von hier nicht erreichbar).
        self.clear_page()
        self._page_header("Daten aktualisieren", "Importbereiche für APU/HAP, NMG Artikel, Rabatte, Artikelstamm und Austauschdatenbank.")
        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure((0, 1, 2), weight=1)
        self._tile(body, 0, "💊", "APU/HAP Daten", "APU/HAP Daten importieren.", "Import", self.import_apu_data, "#11823b")
        self._tile(body, 1, "📋", "NMG Artikel", "NMG Artikel importieren.", "Import", self.import_nmg_articles, "#8b5a00")
        self._tile(body, 2, "💰", "PK Rabatte", "Partnerkonditionen-Rabatte importieren.", "Import", self.import_pk_rabatte, "#6b4fb3")

        row2 = tk.Frame(body, bg="#ffffff")
        row2.grid(row=3, column=0, columnspan=3, sticky="ew")
        row2.columnconfigure((0, 1, 2), weight=1)
        self._tile(row2, 0, "🔎", "Artikelstamm", "PZN-Artikelbasis importieren.", "Import", self.show_artikelstamm_page, "#0b4a86")
        self._tile(row2, 1, "🔄", "Austauschdatenbank", "PZN -> NMG-Austausch importieren.", "Import", self.show_austauschdatenbank_page, "#8b5a00")
        self._tile(row2, 2, "📄", "Auswertungsvorlage", "Vorlage der Ausgabe/Auswertung aktualisieren.", "Öffnen", self.show_auswertungsvorlage_page, "#3867b7")

        row3 = tk.Frame(body, bg="#ffffff")
        row3.grid(row=4, column=0, columnspan=3, sticky="ew")
        row3.columnconfigure((0, 1, 2), weight=1)
        self._tile(row3, 0, "📥", "Manuelle Analysen", "Manuelle PK-/ZW-Analysen importieren.", "Import", self.import_manuelle_analysen, "#11823b")

    def _ask_manual_analysis_type(self):
        """Fragt gezielt, ob manuelle Analysen als PK oder ZW importiert werden sollen."""
        win = tk.Toplevel(self)
        win.resizable(True, True)
        win.title("Manuelle Analysen importieren")
        win.geometry("460x260")
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()
        choice = tk.StringVar(value="PK")

        tk.Label(win, text="Analyseart auswählen", font=(theme.FONT, 18, "bold"), fg="#0b4a86", bg="#f5f7fb").pack(anchor="w", padx=22, pady=(20, 6))
        tk.Label(win, text="Sind die manuellen Analysen PK- oder ZW-Analysen?", bg="#f5f7fb", fg="#333").pack(anchor="w", padx=22, pady=(0, 14))

        box = tk.Frame(win, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        box.pack(fill="x", padx=22, pady=(0, 14))
        tk.Radiobutton(box, text="PK / Partnerkondition", variable=choice, value="PK", bg="#ffffff", font=(theme.FONT, 11)).pack(anchor="w", padx=16, pady=(14, 6))
        tk.Radiobutton(box, text="ZW / Zukunftswerk", variable=choice, value="ZW", bg="#ffffff", font=(theme.FONT, 11)).pack(anchor="w", padx=16, pady=(6, 14))

        result = {"value": None}
        def ok():
            result["value"] = choice.get()
            win.destroy()
        def cancel():
            win.destroy()

        buttons = tk.Frame(win, bg="#f5f7fb")
        buttons.pack(fill="x", padx=22, pady=(0, 18))
        tk.Button(buttons, text="Abbrechen", command=cancel, padx=14, pady=8).pack(side="right", padx=(8, 0))
        tk.Button(buttons, text="Weiter", command=ok, bg="#0b4a86", fg="white", relief="flat", padx=18, pady=9).pack(side="right")
        self.wait_window(win)
        return result["value"]

    def import_manuelle_analysen(self):
        self._log_action("datenaktualisierung", "Manuelle Analysen importieren geöffnet")
        analyse_typ = self._ask_manual_analysis_type()
        if not analyse_typ:
            return
        files = filedialog.askopenfilenames(
            title=f"Manuelle {analyse_typ}-Analysen auswählen",
            filetypes=SUPPORTED_DATA_FILETYPES
        )
        if not files:
            return
        if not messagebox.askyesno(
            "Manuelle Analysen importieren",
            f"{len(files)} Datei(en) als {analyse_typ}-Analysen importieren?\n\n"
            "Die Dateien werden nicht neu ausgewertet. Sie werden als bereits geprüfte manuelle Analysen übernommen, "
            "für Produkt-/Marktanalyse nutzbar gemacht und vorhandene NMG-/Austauschentscheidungen werden als Schulbank-Lernvorschläge angelegt.\n\n"
            "Leere Lernfälle ohne PZN NMG und ohne Austauschtext werden nicht übernommen."
        ):
            return

        # V1.1 SP15: Import laeuft im Hintergrund. Status oben rechts.
        # UI bleibt klickbar. Nach Abschluss laeuft on_main_done im UI-Thread
        # (dort sind messagebox + interaktiver Retry-Pfad wieder OK).

        def on_main_done(stats):
            self._handle_manual_import_done(files, analyse_typ, stats)

        def on_main_error(exc):
            messagebox.showerror("Manuelle Analysen", str(exc))

        self._run_background(
            lambda update: import_manual_analysis_files(
                files, analyse_typ=analyse_typ, bearbeiter=self.bearbeiter,
                progress_callback=update),
            title=f"Manuelle {analyse_typ}-Analysen ({len(files)} Dateien)",
            subtitle=f"Datei 1 von {len(files)} ...",
            progress=True,
            on_done=on_main_done,
            on_error=on_main_error,
        )

    def _handle_manual_import_done(self, files, analyse_typ, stats):
        """V1.1 SP15: Post-Import-Logik (lief frueher synchron nach _run_busy).
        Wird im UI-Thread als on_done-Callback aus _run_background aufgerufen.

        V1.1 SP17: nicht-erkannte Dateien werden zusammengesammelt in
        einer TXT-Liste; pro-Datei-Mapping-Dialog ist raus. Stattdessen
        EINE Sammel-Frage am Ende, ob alle nicht-erkannten Dateien
        nacheinander manuell gemapped werden sollen.
        """
        nicht_erkannt_pfad = None
        if stats.get("failed"):
            nicht_erkannt_pfad = self._write_nicht_erkannte_dateien(analyse_typ, stats)
        self._show_manual_import_stats(analyse_typ, stats, nicht_erkannt_pfad)

        # Einmalige Sammel-Frage am Ende.
        if not stats.get("failed"):
            return
        retry_errors = list(stats.get("errors", []))
        failed_names = {e.split(":", 1)[0] for e in retry_errors}
        if not failed_names:
            return
        if not messagebox.askyesno(
            "Nicht erkannte Dateien",
            f"{stats.get('failed', 0)} Datei(en) wurden nicht automatisch erkannt.\n\n"
            f"Eine Liste wurde gespeichert unter:\n{nicht_erkannt_pfad or '(konnte nicht geschrieben werden)'}\n\n"
            "Sollen die nicht-erkannten Dateien jetzt nacheinander\n"
            "mit dem Format-Assistenten manuell zugeordnet werden?"
        ):
            return

        mapped_temp_files = []
        for file in files:
            if Path(file).name not in failed_names:
                continue
            mapping = self._open_rohdaten_format_assistent(file, "Manuelle Analyse konnte nicht automatisch erkannt werden.")
            if not mapping:
                continue
            try:
                mapped = self._create_standard_rohdaten_from_mapping(file, mapping)
                mapped_temp_files.append(str(mapped))
            except Exception as mapped_exc:
                stats.setdefault("errors", []).append(f"{Path(file).name} Mapping: {mapped_exc}")
        if mapped_temp_files:
            def on_retry_done(retry_stats):
                for key in ("selected", "imported", "duplicates", "failed", "positions", "learning_suggestions"):
                    stats[key] = int(stats.get(key, 0) or 0) + int(retry_stats.get(key, 0) or 0)
                stats.setdefault("duplicate_files", []).extend(retry_stats.get("duplicate_files", []))
                stats.setdefault("errors", []).extend(retry_stats.get("errors", []))
                self._show_manual_import_stats(analyse_typ, stats, None)
            self._run_background(
                lambda update: import_manual_analysis_files(
                    mapped_temp_files, analyse_typ=analyse_typ,
                    bearbeiter=self.bearbeiter, progress_callback=update),
                title=f"Manuelle {analyse_typ}-Analysen (Retry, {len(mapped_temp_files)} Dateien)",
                subtitle=f"Datei 1 von {len(mapped_temp_files)} ...",
                progress=True,
                on_done=on_retry_done,
            )

    def _write_nicht_erkannte_dateien(self, analyse_typ, stats):
        """V1.1 SP17: Schreibt eine TXT-Sammeldatei mit den nicht-erkannten
        Imports unter IMPORT_DIR/<typ>/<Jahr>/Q<n>/. Liefert den Pfad oder
        None, falls das Schreiben fehlschlug.
        """
        try:
            target_dir = jahr_quartal_pfad(IMPORT_DIR / analyse_typ)
        except Exception:
            return None
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pfad = target_dir / f"nicht_erkannte_dateien_{stamp}.txt"
        lines = [
            f"Nicht erkannte Dateien aus dem Manuellen Import vom {datetime.now():%Y-%m-%d %H:%M:%S}",
            f"Analyseart: {analyse_typ}",
            f"Anzahl: {stats.get('failed', 0)}",
            "=" * 70,
            "",
        ]
        for err in stats.get("errors", []):
            lines.append(err)
        try:
            pfad.write_text("\n".join(lines), encoding="utf-8")
        except Exception:
            return None
        return pfad

    def _show_manual_import_stats(self, analyse_typ, stats, nicht_erkannt_pfad=None):
        """V1.1 SP15 / SP17: Statistik-Messagebox + Status-Toast nach Abschluss.
        Wenn nicht-erkannte Dateien aufgelistet sind, wird der TXT-Pfad mit
        angezeigt.
        """
        msg = (
            f"Manuelle {analyse_typ}-Analysen importiert.\n\n"
            f"Ausgewählt: {stats.get('selected', 0)}\n"
            f"Neu übernommen: {stats.get('imported', 0)}\n"
            f"Bereits verarbeitet: {stats.get('duplicates', 0)}\n"
            f"Fehler / nicht erkannt: {stats.get('failed', 0)}\n\n"
            f"Positionen für Produkt-/Marktanalyse: {stats.get('positions', 0)}\n"
            f"Neue Schulbank-Lernvorschläge: {stats.get('learning_suggestions', 0)}"
        )
        if stats.get("duplicates"):
            msg += "\n\nHinweis: " + str(stats.get("duplicates")) + " Analyse(n) wurden schon einmal verarbeitet."
        if nicht_erkannt_pfad:
            msg += f"\n\nNicht-erkannte Dateien:\n{nicht_erkannt_pfad}"
        try:
            add_roadmap_item(
                bereich="Import / Analyse",
                titel="Manuelle Analysen importieren für Schulbank und Markt-/Produktanalyse",
                beschreibung=(
                    "Manuelle PK-/ZW-Analysen können per Mehrfachauswahl importiert werden. "
                    "Dubletten werden per SHA256 erkannt; vorhandene NMG-/Austauschentscheidungen werden "
                    "als Schulbank-Lernvorschläge angelegt. Leere Lernfälle werden nicht übernommen."
                ),
                status="Erledigt",
                prioritaet="Hoch",
            )
        except Exception:
            pass
        # Toast in der Status-Leiste + EINE Statistik-Messagebox (vorher doppelt).
        self._bg_status_toast(
            f"Manuelle {analyse_typ}-Analysen: {stats.get('imported',0)} neu, "
            f"{stats.get('duplicates',0)} doppelt, {stats.get('failed',0)} nicht erkannt."
        )
        messagebox.showinfo(f"Manuelle {analyse_typ}-Analysen", msg)


    def open_admin_auswertungen_loeschen(self):
        """Admin-Funktion: einzelne oder alle gespeicherten Auswertungen aus der DB löschen.

        Es werden nur Datenbankeinträge gelöscht: tbl_auswertungspositionen, tbl_auswertungen
        und optionale Import-Verweise. Ausgabedateien bleiben erhalten, damit nichts außerhalb
        der DB versehentlich verloren geht.
        """
        password = simpledialog.askstring("Admin – Auswertungen löschen", "Admin-Passwort eingeben:", show="*")
        if password is None:
            return
        if password != ADMIN_DB_PASSWORD:
            messagebox.showerror("Admin", "Passwort ist falsch.")
            return

        try:
            with sqlite3.connect(DB_PATH) as con:
                con.row_factory = sqlite3.Row
                rows = con.execute("""
                    SELECT id, datum, apotheke, quelldatei, ausgabedatei,
                           anzahl_positionen, COALESCE(datenquelle,'NMG') AS datenquelle
                    FROM tbl_auswertungen
                    ORDER BY datetime(datum) DESC, id DESC
                """).fetchall()
        except Exception as exc:
            messagebox.showerror("Admin", f"Auswertungen konnten nicht gelesen werden:\n{exc}")
            return

        if not rows:
            messagebox.showinfo("Admin", "Keine gespeicherten Auswertungen vorhanden.")
            return

        win = tk.Toplevel(self)
        win.title("Admin – Auswertungen löschen")
        win.geometry("980x620")
        win.minsize(850, 500)
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()
        win.columnconfigure(0, weight=1)
        win.rowconfigure(2, weight=1)

        tk.Label(win, text="Admin – einzelne oder alle Auswertungen löschen", font=(theme.FONT, 17, "bold"), fg="#0b4a86", bg="#f5f7fb").grid(row=0, column=0, sticky="w", padx=18, pady=(16, 4))
        tk.Label(win, text="Es werden nur Datenbankeinträge gelöscht. Ausgabedateien bleiben erhalten. Vor dem Löschen wird automatisch ein Backup erstellt.", bg="#f5f7fb", fg="#333", justify="left").grid(row=1, column=0, sticky="w", padx=18, pady=(0, 10))

        frame = tk.Frame(win, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        frame.grid(row=2, column=0, sticky="nsew", padx=18, pady=(0, 10))
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)

        cols = ("id", "datum", "typ", "name", "positionen", "datei")
        tree = ttk.Treeview(frame, columns=cols, show="headings", selectmode="extended")
        heads = {"id":"ID", "datum":"Datum", "typ":"Typ", "name":"Auswertung", "positionen":"Pos.", "datei":"Datei"}
        widths = {"id":70, "datum":135, "typ":60, "name":280, "positionen":70, "datei":320}
        for c in cols:
            tree.heading(c, text=heads[c])
            tree.column(c, width=widths[c], anchor="w")
        tree.grid(row=0, column=0, sticky="nsew")
        sb = tk.Scrollbar(frame, orient="vertical", command=tree.yview)
        sb.grid(row=0, column=1, sticky="ns")
        tree.configure(yscrollcommand=sb.set)

        id_map = {}
        for row in rows:
            dq = _dq_label(row["datenquelle"])
            iid = tree.insert("", "end", values=(row["id"], str(row["datum"] or "")[:16], dq, str(row["apotheke"] or ""), row["anzahl_positionen"] or 0, Path(str(row["ausgabedatei"] or "")).name))
            id_map[iid] = int(row["id"])

        buttons = tk.Frame(win, bg="#f5f7fb")
        buttons.grid(row=3, column=0, sticky="ew", padx=18, pady=(0, 14))

        def select_all():
            tree.selection_set(tree.get_children())

        def clear_selection():
            tree.selection_remove(tree.selection())

        def delete_selected():
            selected = list(tree.selection())
            if not selected:
                messagebox.showinfo("Admin", "Bitte mindestens eine Auswertung auswählen.")
                return
            ids = [id_map[iid] for iid in selected if iid in id_map]
            if not ids:
                return
            ok = messagebox.askyesno(
                "Auswertungen löschen?",
                f"{len(ids)} Auswertung(en) werden aus der Datenbank gelöscht.\n\n"
                "Vorher wird automatisch ein Backup erstellt. Ausgabedateien bleiben erhalten. Fortfahren?"
            )
            if not ok:
                return
            confirm = simpledialog.askstring("Sicherheitsabfrage", "Zum endgültigen Löschen bitte AUSWERTUNGEN LÖSCHEN eingeben:")
            if confirm != "AUSWERTUNGEN LÖSCHEN":
                messagebox.showinfo("Admin", "Löschvorgang abgebrochen.")
                return
            try:
                backup_path = backup_erstellen()
                placeholders = ",".join("?" for _ in ids)
                with sqlite3.connect(DB_PATH) as con:
                    con.execute("PRAGMA foreign_keys = ON")
                    pos_count = con.execute(f"SELECT COUNT(*) FROM tbl_auswertungspositionen WHERE auswertung_id IN ({placeholders})", ids).fetchone()[0]
                    con.execute(f"DELETE FROM tbl_auswertungspositionen WHERE auswertung_id IN ({placeholders})", ids)
                    try:
                        con.execute(f"DELETE FROM tbl_importierte_analysen WHERE auswertung_id IN ({placeholders})", ids)
                    except Exception:
                        pass
                    con.execute(f"DELETE FROM tbl_auswertungen WHERE id IN ({placeholders})", ids)
                    con.commit()
                for iid in selected:
                    try:
                        tree.delete(iid)
                    except Exception:
                        pass
                self.status.set(f"Admin: {len(ids)} Auswertung(en) und {pos_count} Positionen gelöscht. Backup: {backup_path}")
                messagebox.showinfo("Admin", f"Gelöscht: {len(ids)} Auswertung(en)\nPositionen: {pos_count}\nBackup:\n{backup_path}")
                # V1.1 SP12 Bug-Fix: dahinterliegende Seite refreshen, sonst
                # zeigt 'Gespeicherte Analysen' weiter die geloeschten Items.
                try:
                    win.destroy()
                    self.open_saved_analyses()
                except Exception:
                    pass
            except Exception as exc:
                messagebox.showerror("Admin", f"Auswertungen konnten nicht gelöscht werden:\n{exc}")

        tk.Button(buttons, text="Alle markieren", command=select_all, padx=12, pady=7).pack(side="left")
        tk.Button(buttons, text="Auswahl aufheben", command=clear_selection, padx=12, pady=7).pack(side="left", padx=(8, 0))
        tk.Button(buttons, text="Ausgewählte löschen", command=delete_selected, bg="#9b1c1c", fg="white", relief="flat", font=(theme.FONT, 10, "bold"), padx=16, pady=8).pack(side="right")
        tk.Button(buttons, text="Schließen", command=win.destroy, padx=12, pady=7).pack(side="right", padx=(0, 8))

    def show_import_page(self, title, command):
        self._action_page(
            title,
            "Dieser Importbereich wird im mittleren Arbeitsbereich angezeigt. Der eigentliche Import startet über den Button.",
            f"{title} importieren  →",
            command,
            "#0b4a86"
        )

    def show_update_backup_page(self):
        self.clear_page()
        self._page_header("Update / Backup", "Versionsinfo, Updates, Backup erstellen und Backup wiederherstellen.")
        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure((0, 1, 2), weight=1)
        self._tile(body, 0, "ℹ️", "Versionsinfo", "Aktuelle Programm- und Datenbankversion anzeigen.", "Anzeigen", self.show_version, "#0b4a86")
        self._tile(body, 1, "⬆️", "Update suchen", "Online-Updateprüfung vorbereiteter Bereich.", "Suchen", self.check_online_update, "#3867b7")
        self._tile(body, 2, "📦", "Update installieren", "Lokales .nmgupdate-Paket installieren.", "Installieren", self.install_update_dialog, "#8b5a00")

        row2 = tk.Frame(body, bg="#ffffff")
        row2.grid(row=3, column=0, columnspan=3, sticky="ew")
        row2.columnconfigure((0, 1), weight=1)
        self._tile(row2, 0, "💾", "Backup erstellen", "Manuelles Backup erstellen.", "Erstellen", self.create_backup, "#11823b")
        self._tile(row2, 1, "♻️", "Backup wiederherstellen", "Backup auswählen und zurückspielen.", "Wiederherstellen", self.restore_backup, "#6b4fb3")
    
    def show_datenbankuebersicht_page(self):
        self.clear_page()
        self._page_header(
            "Datenbankübersicht",
            "Alle Tabellen der NMG-Datenbank mit Datensatzanzahl, Zweck und Inhalt."
        )

        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure(0, weight=1)
        body.rowconfigure(2, weight=1)

        try:
            overview = get_database_overview()
        except Exception as exc:
            tk.Label(
                body,
                text=f"Datenbankübersicht konnte nicht geladen werden:\n{exc}",
                bg="#ffffff",
                fg="#a00000",
                justify="left"
            ).grid(row=0, column=0, sticky="w")
            return

        summary = (
            f"Datenbank: {overview.get('db_path')}\n"
            f"Größe: {format_size(overview.get('db_size_bytes', 0))}\n"
            f"Tabellen: {overview.get('total_tables', 0)} | Datensätze gesamt: {overview.get('total_rows', 0):,}".replace(",", ".")
        )
        tk.Label(
            body,
            text=summary,
            bg="#f8fbff",
            fg="#0b4a86",
            justify="left",
            anchor="w",
            font=(theme.FONT, 10, "bold"),
            padx=12,
            pady=10,
            highlightbackground="#d8e2ee",
            highlightthickness=1
        ).grid(row=0, column=0, sticky="ew", pady=(0, 10))

        columns = ("table", "rows", "purpose", "content")
        tree = ttk.Treeview(body, columns=columns, show="headings", selectmode="browse")
        tree.grid(row=2, column=0, sticky="nsew")

        overview_headings = {"table": "Tabelle", "rows": "Datensätze", "purpose": "Zweck", "content": "Inhalt"}
        tree.heading("table", text="Tabelle")
        tree.heading("rows", text="Datensätze")
        tree.heading("purpose", text="Zweck")
        tree.heading("content", text="Inhalt")
        self._make_tree_sortable(tree, columns, overview_headings)

        tree.column("table", width=190, anchor="w")
        tree.column("rows", width=95, anchor="e")
        tree.column("purpose", width=300, anchor="w")
        tree.column("content", width=360, anchor="w")

        scrollbar = tk.Scrollbar(body, orient="vertical", command=tree.yview)
        scrollbar.grid(row=2, column=0, sticky="nse")
        tree.configure(yscrollcommand=scrollbar.set)

        detail = tk.StringVar(value="Tabelle auswählen, um Spalten anzuzeigen.")
        tk.Label(
            body,
            textvariable=detail,
            bg="#ffffff",
            fg="#333",
            justify="left",
            anchor="w",
            wraplength=850
        ).grid(row=3, column=0, sticky="ew", pady=(10, 0))

        table_rows = overview.get("tables", [])
        item_to_table = {}
        for item in table_rows:
            row_count = f"{int(item.get('rows', 0)):,}".replace(",", ".")
            iid = tree.insert(
                "",
                "end",
                values=(
                    item.get("table", ""),
                    row_count,
                    item.get("purpose", ""),
                    item.get("content", ""),
                )
            )
            item_to_table[iid] = item

        def on_select(event=None):
            sel = tree.selection()
            if not sel:
                detail.set("Tabelle auswählen, um Spalten anzuzeigen.")
                return
            item = item_to_table.get(sel[0], {})
            cols = item.get("columns", [])
            detail.set(
                f"{item.get('display_name', item.get('table', ''))}\n"
                f"Spalten: {', '.join(cols) if cols else 'keine Spalten erkannt'}"
            )

        tree.bind("<<TreeviewSelect>>", on_select)

        toolbar = tk.Frame(body, bg="#ffffff")
        toolbar.grid(row=1, column=0, sticky="w", pady=(0, 8))

        tk.Button(
            toolbar,
            text="Aktualisieren",
            command=self.show_datenbankuebersicht_page,
            bg="#0b4a86",
            fg="white",
            relief="flat",
            font=(theme.FONT, 10, "bold"),
            padx=14,
            pady=7
        ).pack(side="left")

        tk.Button(
            toolbar,
            text="Wirkstoff/Stärke importieren",
            command=self._import_wirkstoff_staerke_dialog,
            bg="#0b6e6e",
            fg="white",
            relief="flat",
            font=(theme.FONT, 10, "bold"),
            padx=14,
            pady=7
        ).pack(side="left", padx=(8, 0))

        # V1.1 SP3: Cleanup direkt erreichbar - ohne Ctrl+Alt+A-Toggle.
        # Schutz liegt auf dem Admin-Passwort + LEEREN-Confirm im Dialog
        # selbst, ein Backup wird vorher automatisch erstellt.
        tk.Button(
            toolbar,
            text="🗑 Datenbankinhalte leeren",
            command=self.open_admin_database_clear,
            bg="#8b5a00",
            fg="white",
            relief="flat",
            font=(theme.FONT, 10, "bold"),
            padx=14,
            pady=7
        ).pack(side="left", padx=(8, 0))

        # 'Auswertungen loeschen' bleibt versteckt - das ist die wirklich
        # destruktive Aktion (loescht Auswertungs-Historie inkl. Positionen)
        # und sollte nicht versehentlich gefunden werden. Sichtbar nur im
        # Admin-Modus (Ctrl+Alt+A) bei Admin-Windows-Login.
        if getattr(self, "_admin_visible", False) and self._is_admin_login():
            tk.Button(
                toolbar,
                text="🔐 Auswertungen löschen",
                command=self.open_admin_auswertungen_loeschen,
                bg="#9b1c1c",
                fg="white",
                relief="flat",
                font=(theme.FONT, 10, "bold"),
                padx=14,
                pady=7
            ).pack(side="left", padx=(8, 0))

    def _import_wirkstoff_staerke_dialog(self):
        """V1.1 SP2: Import der Wirkstoff/Staerke-Excel ueber Datenbankuebersicht.
        Excel-Datei waehlen → Worker-Thread-Import → Status + Logging + Refresh.
        """
        self._log_action("datenaktualisierung", "Wirkstoff/Staerke-Import geoeffnet")
        file = filedialog.askopenfilename(
            title="Wirkstoff/Staerke-Excel auswaehlen",
            filetypes=SUPPORTED_DATA_FILETYPES,
        )
        if not file:
            return
        try:
            stats = self._run_busy(
                lambda: import_wirkstoff_excel(file),
                title="Wirkstoff/Staerke-Import",
                subtitle="Importiere Wirkstoffe und Staerken ...",
            )
        except Exception as exc:
            self._log_error("datenaktualisierung", "Wirkstoff/Staerke-Import", exc)
            messagebox.showerror("Wirkstoff/Staerke-Import", f"Fehler:\n{exc}")
            return

        total = wirkstoff_count()
        msg = (
            f"Datei: {Path(file).name}\n"
            f"Gelesen: {stats['gelesen']:,}\n"
            f"Neu importiert: {stats['importiert']:,}\n"
            f"Aktualisiert: {stats['aktualisiert']:,}\n"
            f"Uebersprungen: {stats['uebersprungen']:,}\n"
            f"Fehler: {stats['fehler']:,}\n\n"
            f"Tabelle tbl_wirkstoff_staerke enthaelt jetzt {total:,} Eintraege."
        ).replace(",", ".")
        self._log_action("datenaktualisierung", "Wirkstoff/Staerke importiert", msg.replace("\n", " | "))
        self.status.set(f"Wirkstoff/Staerke-Import: {stats['importiert']} neu, {stats['aktualisiert']} aktualisiert.")
        messagebox.showinfo("Wirkstoff/Staerke-Import", msg)
        # Tabellen-Ansicht neu laden, damit die neue Zeilenzahl sichtbar ist.
        self.show_datenbankuebersicht_page()

    # V1.1 SP8: Zeitraum-Archivierung + Archiv-Verwaltung ─────────────────────
    def _loeschen_zeitraum_dialog(self):
        """V1.1 SP10: Endgueltig loeschen aller Auswertungen im Zeitraum.
        Im Unterschied zu _archivieren_zeitraum_dialog: kein ZIP-Backup, die
        Daten sind nach LOESCHEN-Confirm + Admin-Passwort unwiderruflich weg.
        """
        from .archiv_db import _normalize_zeitraum, zaehle_zeitraum  # lokaler Import vermeidet Modul-Zyklus

        win = tk.Toplevel(self)
        win.title("Auswertungen im Zeitraum LOESCHEN")
        win.geometry("520x380")
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()
        win.columnconfigure(0, weight=1)

        tk.Label(win, text="Auswertungen im Zeitraum endgueltig loeschen",
                 font=(theme.FONT, 14, "bold"), fg="#9b1c1c", bg="#f5f7fb").pack(
            anchor="w", padx=20, pady=(18, 4))
        tk.Label(win,
                 text=("Alle Auswertungen im Datumsbereich werden aus der Datenbank "
                       "GELOESCHT. Kein ZIP-Backup, keine Wiederherstellung. "
                       "Vorher wird automatisch ein voller DB-Backup erstellt "
                       "(ueber backup_erstellen)."),
                 bg="#f5f7fb", fg="#9b1c1c", wraplength=480, justify="left").pack(
            anchor="w", padx=20, pady=(0, 14))

        form = tk.Frame(win, bg="#f5f7fb")
        form.pack(padx=20, pady=(0, 10), anchor="w")
        tk.Label(form, text="Datum von:",
                 bg="#f5f7fb", fg="#0b4a86", font=(theme.FONT, 10, "bold")).grid(row=0, column=0, sticky="w", pady=4)
        von_var = tk.StringVar()
        _make_date_entry(form, von_var, width=14, font=(theme.FONT, 11)).grid(row=0, column=1, padx=(8, 0))
        tk.Label(form, text="Datum bis:",
                 bg="#f5f7fb", fg="#0b4a86", font=(theme.FONT, 10, "bold")).grid(row=1, column=0, sticky="w", pady=4)
        bis_var = tk.StringVar()
        _make_date_entry(form, bis_var, width=14, font=(theme.FONT, 11)).grid(row=1, column=1, padx=(8, 0))

        vorschau_lbl = tk.Label(win, text="", bg="#f5f7fb", fg="#0b4a86",
                                font=(theme.FONT, 10, "bold"))
        vorschau_lbl.pack(anchor="w", padx=20, pady=(8, 0))

        def vorschau():
            try:
                c = zaehle_zeitraum(von_var.get(), bis_var.get())
                vorschau_lbl.configure(text=f"{c} Auswertungen im Zeitraum.",
                                        fg="#0b4a86")
            except Exception as exc:
                vorschau_lbl.configure(text=f"Fehler: {exc}", fg="#9b1c1c")

        def do_delete():
            try:
                anzahl = zaehle_zeitraum(von_var.get(), bis_var.get())
                von_iso, bis_iso = _normalize_zeitraum(von_var.get(), bis_var.get())
            except Exception as exc:
                messagebox.showerror("Loeschen", f"Datumsformat unklar:\n{exc}")
                return
            if anzahl == 0:
                messagebox.showinfo("Loeschen", "Keine Auswertungen im Zeitraum.")
                return
            pwd = simpledialog.askstring("Admin-Bereich",
                                          "Admin-Passwort eingeben:",
                                          show="*")
            if pwd is None:
                return
            if pwd != ADMIN_DB_PASSWORD:
                messagebox.showerror("Admin-Bereich", "Passwort falsch.")
                return
            confirm = simpledialog.askstring(
                "Sicherheitsabfrage",
                f"{anzahl} Auswertungen werden ENDGUELTIG geloescht. "
                "Zum Bestaetigen bitte LOESCHEN eingeben:")
            if confirm != "LOESCHEN":
                messagebox.showinfo("Loeschen", "Vorgang abgebrochen.")
                return
            try:
                backup_path = backup_erstellen()
            except Exception as exc:
                messagebox.showerror("Backup", f"Backup fehlgeschlagen, Loeschvorgang abgebrochen:\n{exc}")
                return
            try:
                with sqlite3.connect(DB_PATH) as con:
                    con.execute("PRAGMA foreign_keys = ON")
                    aw_ids = [r[0] for r in con.execute(
                        "SELECT id FROM tbl_auswertungen WHERE date(datum) >= date(?) AND date(datum) <= date(?)",
                        (von_iso, bis_iso),
                    ).fetchall()]
                    if not aw_ids:
                        messagebox.showinfo("Loeschen", "Keine Auswertungen mehr im Zeitraum.")
                        return
                    placeholders = ",".join("?" for _ in aw_ids)
                    pos_count = con.execute(
                        f"SELECT COUNT(*) FROM tbl_auswertungspositionen WHERE auswertung_id IN ({placeholders})",
                        aw_ids,
                    ).fetchone()[0]
                    con.execute(
                        f"DELETE FROM tbl_auswertungspositionen WHERE auswertung_id IN ({placeholders})",
                        aw_ids,
                    )
                    try:
                        con.execute(
                            f"DELETE FROM tbl_importierte_analysen WHERE auswertung_id IN ({placeholders})",
                            aw_ids,
                        )
                    except Exception:
                        pass
                    con.execute(
                        f"DELETE FROM tbl_auswertungen WHERE id IN ({placeholders})",
                        aw_ids,
                    )
                    con.commit()
            except Exception as exc:
                messagebox.showerror("Loeschen", f"Loeschvorgang fehlgeschlagen:\n{exc}")
                return
            msg = (f"Geloescht: {len(aw_ids)} Auswertungen, {pos_count} Positionen.\n"
                   f"Backup vorher: {backup_path}")
            self._log_action("auswertungen", "Zeitraum geloescht", msg.replace("\n", " | "))
            self.status.set(f"{len(aw_ids)} Auswertungen geloescht.")
            messagebox.showinfo("Loeschen abgeschlossen", msg)
            win.destroy()
            self.open_saved_analyses()

        btn_bar = tk.Frame(win, bg="#f5f7fb")
        btn_bar.pack(side="bottom", fill="x", padx=20, pady=16)
        tk.Button(btn_bar, text="Abbrechen", command=win.destroy,
                  padx=14, pady=6).pack(side="right")
        tk.Button(btn_bar, text="🔐 LOESCHEN  →", command=do_delete,
                  bg="#9b1c1c", fg="white", relief="flat",
                  font=(theme.FONT, 10, "bold"), padx=16, pady=7).pack(side="right", padx=(0, 8))
        tk.Button(btn_bar, text="Vorschau zaehlen", command=vorschau,
                  bg="#0b4a86", fg="white", relief="flat",
                  font=(theme.FONT, 10, "bold"), padx=14, pady=6).pack(side="right", padx=(0, 8))

    def _archivieren_zeitraum_dialog(self):
        """Dialog: von/bis-Datum -> Vorschau -> Confirm -> ZIP erstellen,
        DB-Zeilen loeschen.
        """
        win = tk.Toplevel(self)
        win.title("Auswertungen im Zeitraum archivieren")
        win.geometry("520x360")
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()
        win.columnconfigure(0, weight=1)

        tk.Label(win, text="Auswertungen im Zeitraum archivieren",
                 font=(theme.FONT, 14, "bold"), fg="#0b4a86", bg="#f5f7fb").pack(
            anchor="w", padx=20, pady=(18, 4))
        tk.Label(win,
                 text=("Alle Auswertungen im Datumsbereich werden in ein ZIP-Backup "
                       "ausgelagert und dann aus der aktiven DB entfernt. Sie tauchen "
                       "weder in Suchen noch in Produktanalysen auf, bleiben aber im "
                       "Archiv-Ordner verfuegbar."),
                 bg="#f5f7fb", fg="#333", wraplength=480, justify="left").pack(
            anchor="w", padx=20, pady=(0, 14))

        form = tk.Frame(win, bg="#f5f7fb")
        form.pack(padx=20, pady=(0, 10), anchor="w")
        tk.Label(form, text="Datum von:",
                 bg="#f5f7fb", fg="#0b4a86", font=(theme.FONT, 10, "bold")).grid(row=0, column=0, sticky="w", pady=4)
        von_var = tk.StringVar()
        _make_date_entry(form, von_var, width=14, font=(theme.FONT, 11)).grid(row=0, column=1, padx=(8, 0))
        tk.Label(form, text="Datum bis:",
                 bg="#f5f7fb", fg="#0b4a86", font=(theme.FONT, 10, "bold")).grid(row=1, column=0, sticky="w", pady=4)
        bis_var = tk.StringVar()
        _make_date_entry(form, bis_var, width=14, font=(theme.FONT, 11)).grid(row=1, column=1, padx=(8, 0))

        vorschau_lbl = tk.Label(win, text="", bg="#f5f7fb", fg="#0b4a86",
                                font=(theme.FONT, 10, "bold"))
        vorschau_lbl.pack(anchor="w", padx=20, pady=(8, 0))

        def vorschau():
            try:
                c = zaehle_zeitraum(von_var.get(), bis_var.get())
                vorschau_lbl.configure(text=f"{c} Auswertungen im Zeitraum.",
                                        fg="#0b4a86")
            except Exception as exc:
                vorschau_lbl.configure(text=f"Fehler: {exc}", fg="#9b1c1c")

        def do_archive():
            try:
                anzahl = zaehle_zeitraum(von_var.get(), bis_var.get())
            except Exception as exc:
                messagebox.showerror("Archivieren", f"Datumsformat unklar:\n{exc}")
                return
            if anzahl == 0:
                messagebox.showinfo("Archivieren", "Keine Auswertungen im Zeitraum.")
                return
            ok = messagebox.askyesno(
                "Archivieren bestaetigen",
                f"{anzahl} Auswertungen werden ins Archiv verschoben und "
                "danach aus der aktiven DB geloescht.\n\nFortfahren?")
            if not ok:
                return
            try:
                stats = self._run_busy(
                    lambda: archiviere_zeitraum(von_var.get(), bis_var.get()),
                    title="Auswertungen archivieren",
                    subtitle=f"Schreibe ZIP mit {anzahl} Auswertungen ...",
                )
            except Exception as exc:
                messagebox.showerror("Archivieren", str(exc))
                return
            msg = (
                f"Archiviert: {stats['archiviert']} Auswertungen, "
                f"{stats['excel_anzahl']} Excel-Dateien.\n"
                f"ZIP: {stats['zip_pfad']}\n"
                f"Groesse: {stats['groesse_bytes']:,} Bytes".replace(",", ".")
            )
            self._log_action("archiv", "Zeitraum archiviert", msg.replace("\n", " | "))
            self.status.set(f"{stats['archiviert']} Auswertungen archiviert.")
            messagebox.showinfo("Archivieren fertig", msg)
            win.destroy()
            self.open_saved_analyses()

        btn_bar = tk.Frame(win, bg="#f5f7fb")
        btn_bar.pack(side="bottom", fill="x", padx=20, pady=16)
        tk.Button(btn_bar, text="Abbrechen", command=win.destroy,
                  padx=14, pady=6).pack(side="right")
        tk.Button(btn_bar, text="Archivieren  →", command=do_archive,
                  bg="#9b1c1c", fg="white", relief="flat",
                  font=(theme.FONT, 10, "bold"), padx=16, pady=7).pack(side="right", padx=(0, 8))
        tk.Button(btn_bar, text="Vorschau zaehlen", command=vorschau,
                  bg="#0b4a86", fg="white", relief="flat",
                  font=(theme.FONT, 10, "bold"), padx=14, pady=6).pack(side="right", padx=(0, 8))

    def _archive_verwalten_dialog(self):
        """Dialog: Liste aller ZIP-Archive. Pro ZIP eine zweite Liste mit
        den enthaltenen Auswertungen. Doppelklick = Excel oeffnen.
        Loeschen nur per Admin-Passwort.
        """
        ensure_archiv_dir()
        win = tk.Toplevel(self)
        win.title("Archive verwalten")
        win.geometry("900x600")
        win.configure(bg="#ffffff")
        win.transient(self)
        win.columnconfigure(0, weight=1)
        win.columnconfigure(1, weight=2)
        win.rowconfigure(1, weight=1)

        tk.Label(win, text="Archive verwalten",
                 font=(theme.FONT, 14, "bold"), fg="#0b4a86", bg="#ffffff").grid(
            row=0, column=0, columnspan=2, sticky="w", padx=14, pady=(14, 6))

        # Linke Liste: ZIPs
        zip_frame = tk.LabelFrame(win, text=" Archiv-ZIPs ",
                                  bg="#ffffff", fg="#0b4a86",
                                  font=(theme.FONT, 10, "bold"))
        zip_frame.grid(row=1, column=0, sticky="nsew", padx=(14, 6), pady=(0, 6))
        zip_frame.rowconfigure(0, weight=1)
        zip_frame.columnconfigure(0, weight=1)

        zip_cols = ("name", "zeitraum", "count", "groesse")
        zip_heads = {"name": "Name", "zeitraum": "Zeitraum",
                     "count": "Anz.", "groesse": "Groesse"}
        zip_widths = {"name": 200, "zeitraum": 130, "count": 50, "groesse": 80}
        zip_tree = ttk.Treeview(zip_frame, columns=zip_cols, show="headings",
                                selectmode="browse")
        for c in zip_cols:
            zip_tree.heading(c, text=zip_heads[c])
            zip_tree.column(c, width=zip_widths[c], anchor="w")
        zip_tree.grid(row=0, column=0, sticky="nsew")

        # Rechte Liste: Inhalt des selektierten ZIPs
        inh_frame = tk.LabelFrame(win, text=" Inhalt (Doppelklick = Excel oeffnen) ",
                                  bg="#ffffff", fg="#0b4a86",
                                  font=(theme.FONT, 10, "bold"))
        inh_frame.grid(row=1, column=1, sticky="nsew", padx=(6, 14), pady=(0, 6))
        inh_frame.rowconfigure(0, weight=1)
        inh_frame.columnconfigure(0, weight=1)
        inh_cols = ("id", "datum", "apotheke", "kunde", "pos")
        inh_heads = {"id": "ID", "datum": "Datum", "apotheke": "Apotheke",
                     "kunde": "Kunde", "pos": "Pos."}
        inh_widths = {"id": 50, "datum": 110, "apotheke": 220, "kunde": 140, "pos": 50}
        inh_tree = ttk.Treeview(inh_frame, columns=inh_cols, show="headings",
                                selectmode="browse")
        for c in inh_cols:
            inh_tree.heading(c, text=inh_heads[c])
            inh_tree.column(c, width=inh_widths[c], anchor="w")
        inh_tree.grid(row=0, column=0, sticky="nsew")

        zip_map: dict[str, dict] = {}
        inh_map: dict[str, dict] = {}
        current_zip = {"pfad": None}

        def reload_zips():
            for iid in zip_tree.get_children():
                zip_tree.delete(iid)
            zip_map.clear()
            for a in liste_archive():
                groesse_kb = a["groesse_bytes"] / 1024
                groesse_txt = f"{groesse_kb:,.0f} KB".replace(",", ".") \
                    if groesse_kb < 1024 else f"{groesse_kb/1024:,.1f} MB".replace(",", ".")
                zeitraum = f"{a['zeitraum_von']} - {a['zeitraum_bis']}" if a["zeitraum_von"] else ""
                iid = zip_tree.insert("", "end", values=(
                    a["name"], zeitraum, a["auswertungen_count"], groesse_txt
                ))
                zip_map[iid] = a
            for iid in inh_tree.get_children():
                inh_tree.delete(iid)
            inh_map.clear()
            current_zip["pfad"] = None

        def on_zip_select(_e=None):
            sel = zip_tree.selection()
            if not sel:
                return
            a = zip_map.get(sel[0])
            if not a:
                return
            current_zip["pfad"] = a["pfad"]
            for iid in inh_tree.get_children():
                inh_tree.delete(iid)
            inh_map.clear()
            try:
                analysen = liste_analysen_im_archiv(a["pfad"])
            except Exception as exc:
                messagebox.showerror("Archiv lesen", str(exc))
                return
            for r in analysen:
                iid = inh_tree.insert("", "end", values=(
                    r.get("id", ""), str(r.get("datum", ""))[:10],
                    r.get("apotheke", ""), r.get("kundenname", ""),
                    r.get("anzahl_positionen", 0)
                ))
                inh_map[iid] = r

        def on_inh_dbl(_e=None):
            sel = inh_tree.selection()
            if not sel or not current_zip["pfad"]:
                return
            r = inh_map.get(sel[0])
            if not r:
                return
            try:
                ziel = excel_aus_archiv(current_zip["pfad"], int(r["id"]))
            except Exception as exc:
                messagebox.showerror("Archiv-Datei", str(exc))
                return
            if not ziel:
                messagebox.showinfo("Archiv-Datei",
                                    "Im ZIP ist fuer diese Auswertung keine Excel hinterlegt.")
                return
            try:
                _open_file(ziel)
            except Exception as exc:
                messagebox.showerror("Datei oeffnen", str(exc))

        def loeschen_zip():
            sel = zip_tree.selection()
            if not sel:
                messagebox.showinfo("Loeschen", "Bitte zuerst ein Archiv auswaehlen.")
                return
            a = zip_map.get(sel[0])
            if not a:
                return
            pwd = simpledialog.askstring("Admin-Bereich",
                                          "Admin-Passwort eingeben (Archiv wird endgueltig geloescht):",
                                          show="*")
            if pwd is None:
                return
            if pwd != ADMIN_DB_PASSWORD:
                messagebox.showerror("Admin-Bereich", "Passwort falsch.")
                return
            ok = messagebox.askyesno("Archiv loeschen",
                                      f"Archiv endgueltig loeschen?\n\n{a['name']}\n\n"
                                      "Die enthaltenen Auswertungen sind danach unwiderruflich weg.")
            if not ok:
                return
            try:
                loesche_archiv(a["pfad"])
            except Exception as exc:
                messagebox.showerror("Loeschen", str(exc))
                return
            self._log_action("archiv", "Archiv geloescht", a["name"])
            reload_zips()
            messagebox.showinfo("Loeschen", f"Archiv {a['name']} geloescht.")

        zip_tree.bind("<<TreeviewSelect>>", on_zip_select)
        inh_tree.bind("<Double-1>", on_inh_dbl)

        btn_bar = tk.Frame(win, bg="#ffffff")
        btn_bar.grid(row=2, column=0, columnspan=2, sticky="ew", padx=14, pady=(0, 14))
        tk.Button(btn_bar, text="Aktualisieren", command=reload_zips,
                  bg="#0b4a86", fg="white", relief="flat",
                  font=(theme.FONT, 10, "bold"), padx=12, pady=6).pack(side="left")
        tk.Button(btn_bar, text="Archiv-Ordner oeffnen",
                  command=lambda: _open_folder(ensure_archiv_dir()),
                  padx=12, pady=6).pack(side="left", padx=(8, 0))
        tk.Button(btn_bar, text="🔐 Archiv endgueltig loeschen",
                  command=loeschen_zip,
                  bg="#9b1c1c", fg="white", relief="flat",
                  font=(theme.FONT, 10, "bold"), padx=12, pady=6).pack(side="right")
        tk.Button(btn_bar, text="Schliessen", command=win.destroy,
                  padx=12, pady=6).pack(side="right", padx=(0, 8))

        reload_zips()

    def _table_exists(self, table_name):
        with sqlite3.connect(DB_PATH) as con:
            row = con.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                (table_name,)
            ).fetchone()
            return row is not None

    def _table_count(self, table_name):
        if not self._table_exists(table_name):
            return None
        with sqlite3.connect(DB_PATH) as con:
            return con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]

    def _mark_roadmap_admin_clear(self, beschreibung):
        """Roadmap ohne Pflichtabhängigkeit aktualisieren: keine Fehlermeldung, falls Roadmap nicht erreichbar ist."""
        try:
            ensure_roadmap_table()
            titel = "Admin-Funktion Datenbankinhalte gezielt leeren"
            existing = None
            for row in list_roadmap_items():
                if row["titel"] == titel:
                    existing = row
                    break
            if existing:
                update_roadmap_status(existing["id"], "Erledigt")
            else:
                add_roadmap_item(
                    bereich="Datenbank",
                    titel=titel,
                    beschreibung=beschreibung,
                    status="Erledigt",
                    prioritaet="Hoch"
                )
        except Exception:
            pass

    def open_admin_database_clear(self):
        password = simpledialog.askstring(
            "Admin-Bereich",
            "Admin-Passwort eingeben:",
            show="*"
        )
        if password is None:
            return
        if password != ADMIN_DB_PASSWORD:
            messagebox.showerror("Admin-Bereich", "Passwort ist falsch.")
            return

        # SP3: alle existierenden Tabellen dynamisch ermitteln. Vorher war
        # die Liste hartcodiert in ADMIN_CLEAR_TABLES, viele Bereiche
        # waren unsichtbar. Schutz-Tabellen (App-Konfig + Logs) bleiben
        # weiter ausgeschlossen, sonst wuerde das Programm danach nicht
        # mehr starten oder die Historie waere weg.
        from .db_overview import TABLE_DESCRIPTIONS as _TABLE_DESCS
        _SKIP = {"meta", "tbl_update_log", "tbl_system_log", "sqlite_sequence"}
        available = []
        try:
            with sqlite3.connect(DB_PATH) as con:
                rows = con.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' "
                    "AND name NOT LIKE 'sqlite_%' ORDER BY name"
                ).fetchall()
            for (table_name,) in rows:
                if table_name in _SKIP:
                    continue
                count = self._table_count(table_name)
                if count is None:
                    continue
                desc = _TABLE_DESCS.get(table_name, {})
                label = desc.get("name", table_name)
                available.append((table_name, label, count))
        except Exception:
            pass

        if not available:
            messagebox.showinfo("Admin-Bereich", "Keine Tabellen in der Datenbank gefunden.")
            return

        win = tk.Toplevel(self)
        win.resizable(True, True)
        win.title("Admin – Datenbankinhalte leeren")
        win.geometry("660x470")
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()
        win.columnconfigure(0, weight=1)

        tk.Label(
            win,
            text="Admin – Datenbankinhalte leeren",
            font=(theme.FONT, 18, "bold"),
            fg="#0b4a86",
            bg="#f5f7fb"
        ).grid(row=0, column=0, sticky="w", padx=22, pady=(20, 6))

        tk.Label(
            win,
            text="Es werden nur Inhalte gelöscht, nicht die Tabellenstruktur. Vor dem Löschen wird automatisch ein Backup erstellt.",
            bg="#f5f7fb",
            fg="#333",
            wraplength=600,
            justify="left"
        ).grid(row=1, column=0, sticky="w", padx=22, pady=(0, 14))

        # SP3: scrollbarer Container, weil die Tabellenliste jetzt deutlich
        # laenger sein kann. Default: alle Haekchen AUS, damit man nicht
        # versehentlich die halbe DB leert.
        win.rowconfigure(2, weight=1)
        scroll_outer = tk.Frame(win, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        scroll_outer.grid(row=2, column=0, sticky="nsew", padx=22, pady=(0, 14))
        scroll_outer.columnconfigure(0, weight=1)
        scroll_outer.rowconfigure(0, weight=1)

        canvas = tk.Canvas(scroll_outer, bg="#ffffff", highlightthickness=0)
        canvas.grid(row=0, column=0, sticky="nsew")
        sb = ttk.Scrollbar(scroll_outer, orient="vertical", command=canvas.yview)
        sb.grid(row=0, column=1, sticky="ns")
        canvas.configure(yscrollcommand=sb.set)

        box = tk.Frame(canvas, bg="#ffffff")
        canvas.create_window((0, 0), window=box, anchor="nw", tags="box")
        box.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure("box", width=e.width))

        selections = {}
        for idx, (table_name, label, count) in enumerate(available):
            var = tk.BooleanVar(value=False)
            selections[table_name] = var
            text = f"{label}  |  {table_name}  |  {count:,} Datensätze".replace(",", ".")
            tk.Checkbutton(
                box,
                text=text,
                variable=var,
                bg="#ffffff",
                anchor="w",
                justify="left",
                font=(theme.FONT, 10)
            ).grid(row=idx, column=0, sticky="ew", padx=14, pady=(10 if idx == 0 else 4, 4))

        warn = tk.Label(
            win,
            text="Wichtig: Jede Löschung erstellt vorher automatisch ein Backup. Roadmap-Einträge und System-Logs werden nicht gelöscht.",
            bg="#fff7e6",
            fg="#8b5a00",
            justify="left",
            anchor="w",
            padx=12,
            pady=10,
            wraplength=600,
            highlightbackground="#efd39a",
            highlightthickness=1
        )
        warn.grid(row=3, column=0, sticky="ew", padx=22, pady=(0, 14))

        buttons = tk.Frame(win, bg="#f5f7fb")
        buttons.grid(row=4, column=0, sticky="e", padx=22, pady=(0, 18))

        def do_clear():
            chosen = [table for table, var in selections.items() if var.get()]
            if not chosen:
                messagebox.showinfo("Admin-Bereich", "Bitte mindestens eine Tabelle auswählen.")
                return
            names = "\n".join(f"- {table}" for table in chosen)
            ok = messagebox.askyesno(
                "Inhalte wirklich löschen?",
                "Folgende Tabelleninhalte werden geleert:\n\n"
                f"{names}\n\n"
                "Die Tabellen bleiben bestehen. Vorher wird automatisch ein Backup erstellt. Fortfahren?"
            )
            if not ok:
                return
            confirm = simpledialog.askstring(
                "Sicherheitsabfrage",
                "Zum endgültigen Löschen bitte LEEREN eingeben:"
            )
            if confirm != "LEEREN":
                messagebox.showinfo("Admin-Bereich", "Löschvorgang abgebrochen.")
                return
            try:
                backup_path = backup_erstellen()
                deleted = {}
                with sqlite3.connect(DB_PATH) as con:
                    for table in chosen:
                        count = con.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
                        con.execute(f"DELETE FROM {table}")
                        deleted[table] = count
                    for table in chosen:
                        con.execute("DELETE FROM sqlite_sequence WHERE name=?", (table,))
                    con.commit()
                details = "\n".join(
                    f"{table}: {count:,} gelöscht".replace(",", ".")
                    for table, count in deleted.items()
                )
                self._mark_roadmap_admin_clear(
                    "Admin-Schaltfläche in der Datenbankübersicht ergänzt. "
                    "Ausgewählte Tabelleninhalte können nach Passwortprüfung und automatischem Backup geleert werden."
                )
                self.status.set(f"Admin-Bereinigung abgeschlossen. Backup: {backup_path}")
                messagebox.showinfo(
                    "Admin-Bereinigung abgeschlossen",
                    f"Backup erstellt:\n{backup_path}\n\nGelöschte Inhalte:\n{details}"
                )
                win.destroy()
                self.show_datenbankuebersicht_page()
            except Exception as exc:
                messagebox.showerror("Admin-Bereich", f"Daten konnten nicht geleert werden:\n{exc}")

        def _select_all():
            for v in selections.values():
                v.set(True)
        def _select_none():
            for v in selections.values():
                v.set(False)
        tk.Button(buttons, text="Alle markieren", command=_select_all, padx=12, pady=7).pack(side="left", padx=(0, 8))
        tk.Button(buttons, text="Keine markieren", command=_select_none, padx=12, pady=7).pack(side="left", padx=(0, 8))
        tk.Button(buttons, text="Abbrechen", command=win.destroy, padx=16, pady=8).pack(side="right", padx=(8, 0))
        tk.Button(
            buttons,
            text="Backup erstellen und Inhalte leeren",
            command=do_clear,
            bg="#9b1c1c",
            fg="white",
            relief="flat",
            font=(theme.FONT, 10, "bold"),
            padx=18,
            pady=9
        ).pack(side="right")

    def _auswertungsvorlage_paths(self):
        vorlagen_dir = DATA_DIR / "vorlagen"
        target = vorlagen_dir / "Auswertungsvorlage.xlsx"
        backup_dir = vorlagen_dir / "backups"
        return vorlagen_dir, target, backup_dir

    def _roadmap_mark_auswertungsvorlage_erledigt(self):
        try:
            ensure_roadmap_table()
            titel = "Auswertungsvorlagen-Verwaltung mit Auswahl und Vorschau"
            for row in list_roadmap_items():
                if row["titel"] == titel:
                    update_roadmap_status(int(row["id"]), "Erledigt")
                    return
            add_roadmap_item(
                bereich="Daten aktualisieren",
                titel=titel,
                beschreibung=(
                    "Auswertungsvorlagen können jetzt in bis zu drei Slots hinterlegt werden. "
                    "In Neue Auswertung kann eine Vorlage ausgewählt und als Standard gespeichert werden. "
                    "Beim Export werden Layout, Farben, Spaltenbreiten und Kopfzeile übernommen; Datenzeilen der Vorlage werden nicht übernommen."
                ),
                status="Erledigt",
                prioritaet="Hoch"
            )
        except Exception:
            pass

    def _selected_template_from_tree(self, tree):
        sel = tree.selection()
        if not sel:
            messagebox.showinfo("Auswertungsvorlage", "Bitte zuerst eine Vorlage auswählen.")
            return None
        values = tree.item(sel[0], "values")
        try:
            return int(values[0])
        except Exception:
            return None

    def show_auswertungsvorlage_page(self):
        self.clear_page()
        self._page_header(
            "Auswertungsvorlagen verwalten",
            "Bis zu 3 Vorlagen hochladen. Übernommen werden Layout, Farben, Spaltenbreiten und Spaltenüberschriften – keine Dateninhalte."
        )

        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure(0, weight=1)
        body.rowconfigure(1, weight=1)

        info = (
            "Regel: Maximal 3 Auswertungsvorlagen. In 'Bedarfsanalyse' kann die Vorlage ausgewählt werden.\n"
            "Die Auswahl wird gespeichert. Daten aus der Vorlage werden nicht übernommen; nur Optik, Kopfzeile und Tabellenformat."
        )
        tk.Label(body, text=info, justify="left", bg="#ffffff", fg="#333").grid(row=0, column=0, sticky="w", pady=(0, 10))

        columns = ("slot", "name", "datei", "status", "aktualisiert")
        tree = ttk.Treeview(body, columns=columns, show="headings", selectmode="browse", height=7)
        tree.grid(row=1, column=0, sticky="nsew")
        headings = {"slot": "Slot", "name": "Name", "datei": "Datei", "status": "Status", "aktualisiert": "Aktualisiert"}
        widths = {"slot": 55, "name": 220, "datei": 280, "status": 130, "aktualisiert": 150}
        for col in columns:
            tree.heading(col, text=headings[col])
            tree.column(col, width=widths[col], anchor="w")
        self._make_tree_sortable(tree, columns, headings)

        selected_slot = self._get_selected_auswertungsvorlage_slot()
        templates_by_slot = {t["slot"]: t for t in self._list_auswertungsvorlagen()}
        for slot in (1, 2, 3):
            item = templates_by_slot.get(slot)
            if item:
                status = "Standard" if selected_slot == slot else "hinterlegt"
                tree.insert("", "end", values=(slot, item["name"], Path(item["path"]).name, status, item.get("updated", "")))
            else:
                tree.insert("", "end", values=(slot, f"Vorlage {slot}", "", "frei", ""))

        preview = tk.StringVar(value="Vorlage auswählen, um die Vorschau zu sehen.")
        tk.Label(
            body,
            textvariable=preview,
            justify="left",
            bg="#f8fbff",
            fg="#222",
            anchor="w",
            padx=12,
            pady=10,
            wraplength=850,
            highlightbackground="#d8e2ee",
            highlightthickness=1,
        ).grid(row=2, column=0, sticky="ew", pady=(10, 10))

        def refresh_preview(event=None):
            slot = self._selected_template_from_tree(tree) if tree.selection() else None
            if not slot:
                preview.set("Vorlage auswählen, um die Vorschau zu sehen.")
                return
            path = self._get_auswertungsvorlage_path(slot)
            if not path:
                preview.set(f"Slot {slot} ist frei.")
                return
            preview.set(self._preview_auswertungsvorlage_text(path))

        tree.bind("<<TreeviewSelect>>", refresh_preview)

        actions = tk.Frame(body, bg="#ffffff")
        actions.grid(row=3, column=0, sticky="w", pady=(4, 0))

        def upload_template():
            slot = self._selected_template_from_tree(tree)
            if not slot:
                return
            file = filedialog.askopenfilename(
                title=f"Auswertungsvorlage für Slot {slot} auswählen",
                filetypes=[("Excel-Vorlagen", "*.xlsx *.xlsm"), ("Excel", "*.xlsx"), ("Excel mit Makros", "*.xlsm"), ("Alle Dateien", "*.*")]
            )
            if not file:
                return
            source = Path(file)
            if source.suffix.lower() not in {".xlsx", ".xlsm"}:
                messagebox.showwarning("Auswertungsvorlage", "Bitte eine Excel-Datei (*.xlsx oder *.xlsm) auswählen.")
                return
            name = simpledialog.askstring("Vorlagenname", "Name für diese Auswertungsvorlage:", initialvalue=source.stem)
            if not name:
                return
            target_dir = self._auswertungsvorlagen_dir()
            target = target_dir / f"Auswertungsvorlage_{slot}{source.suffix.lower()}"
            # Altes Slot-File sichern, wenn vorhanden.
            old_path = self._get_auswertungsvorlage_path(slot)
            if old_path:
                backup_dir = target_dir / "backups"
                backup_dir.mkdir(parents=True, exist_ok=True)
                backup = backup_dir / f"Auswertungsvorlage_{slot}_{datetime.now():%Y%m%d_%H%M%S}{old_path.suffix}"
                try:
                    shutil.copy2(old_path, backup)
                except Exception:
                    pass
            # Falls vorher ein anderer Dateityp im Slot lag, entfernen.
            for suffix in (".xlsx", ".xlsm"):
                other = target_dir / f"Auswertungsvorlage_{slot}{suffix}"
                if other.exists() and other != target:
                    try:
                        other.unlink()
                    except Exception:
                        pass
            shutil.copy2(source, target)
            with sqlite3.connect(DB_PATH) as con:
                con.execute("CREATE TABLE IF NOT EXISTS meta(key TEXT PRIMARY KEY, value TEXT NOT NULL)")
                con.execute("INSERT OR REPLACE INTO meta(key,value) VALUES(?, ?)", (self._template_meta_key(slot, "path"), str(target)))
                con.execute("INSERT OR REPLACE INTO meta(key,value) VALUES(?, ?)", (self._template_meta_key(slot, "name"), name.strip()))
                con.execute("INSERT OR REPLACE INTO meta(key,value) VALUES(?, ?)", (self._template_meta_key(slot, "source"), source.name))
                con.execute("INSERT OR REPLACE INTO meta(key,value) VALUES(?, CURRENT_TIMESTAMP)", (self._template_meta_key(slot, "updated"),))
                con.execute("INSERT OR REPLACE INTO meta(key,value) VALUES('auswertungsvorlage_selected_slot', ?)", (str(slot),))
                con.commit()
            self._roadmap_mark_auswertungsvorlage_erledigt()
            self.status.set(f"Auswertungsvorlage Slot {slot} gespeichert: {name}")
            messagebox.showinfo("Auswertungsvorlage", f"Vorlage gespeichert und als Standard gesetzt:\n{name}")
            self.show_auswertungsvorlage_page()

        def set_default():
            slot = self._selected_template_from_tree(tree)
            if not slot:
                return
            if not self._get_auswertungsvorlage_path(slot):
                messagebox.showinfo("Auswertungsvorlage", "Dieser Slot ist leer und kann nicht als Standard gesetzt werden.")
                return
            self._set_selected_auswertungsvorlage_slot(slot)
            self.status.set(f"Auswertungsvorlage Slot {slot} als Standard gespeichert.")
            self.show_auswertungsvorlage_page()

        def delete_template():
            slot = self._selected_template_from_tree(tree)
            if not slot:
                return
            path = self._get_auswertungsvorlage_path(slot)
            if not path:
                messagebox.showinfo("Auswertungsvorlage", "Dieser Slot ist bereits frei.")
                return
            if not messagebox.askyesno("Auswertungsvorlage löschen", f"Vorlage in Slot {slot} entfernen?\n\nDie Datei wird aus der Vorlagenverwaltung entfernt, Auswertungen bleiben unverändert."):
                return
            try:
                path.unlink()
            except Exception:
                pass
            with sqlite3.connect(DB_PATH) as con:
                for suffix in ("path", "name", "source", "updated"):
                    con.execute("DELETE FROM meta WHERE key=?", (self._template_meta_key(slot, suffix),))
                if self._get_selected_auswertungsvorlage_slot() == slot:
                    con.execute("INSERT OR REPLACE INTO meta(key,value) VALUES('auswertungsvorlage_selected_slot','')")
                con.commit()
            self.status.set(f"Auswertungsvorlage Slot {slot} entfernt.")
            self.show_auswertungsvorlage_page()

        tk.Button(actions, text="Vorlage hochladen/ersetzen", command=upload_template, bg="#0b4a86", fg="white", relief="flat", font=(theme.FONT, 11, "bold"), padx=16, pady=8).pack(side="left", padx=(0, 8))
        tk.Button(actions, text="Als Standard verwenden", command=set_default, padx=14, pady=8).pack(side="left", padx=8)
        tk.Button(actions, text="Vorlage entfernen", command=delete_template, padx=14, pady=8).pack(side="left", padx=8)
        tk.Button(actions, text="Vorlagenordner öffnen", command=lambda: _open_folder(self._auswertungsvorlagen_dir()), padx=14, pady=8).pack(side="left", padx=8)

    def update_auswertungsvorlage(self):
        """Kompatibilitätsfunktion: alter Button ruft neue Vorlagenverwaltung auf."""
        self.show_auswertungsvorlage_page()

    def show_artikelstamm_page(self):
        self.clear_page()
        self._page_header(
            "Artikelstamm / PZN-Basis",
            "Zentrale Artikelbasis für alle Dateien, in denen eine PZN vorkommt."
        )

        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))

        try:
            count = count_artikelstamm()
        except Exception:
            count = 0

        info = (
            f"Gespeicherte Artikel: {count}\n\n"
            "Erwartete Excel-Spalten:\n"
            "- PZN / Pharmazentralnummer\n"
            "- Artikel / Artikelname / Artikelbez.\n"
            "- DF / DAR\n"
            "- PCK / Packung\n"
            "- Herst / Hersteller\n\n"
            "PZN ist der feste Schlüssel. Bestehende PZN werden aktualisiert."
        )

        tk.Label(
            body,
            text=info,
            justify="left",
            bg="#ffffff",
            fg="#222",
            font=(theme.FONT, 11)
        ).pack(anchor="w", pady=(0, 16))

        tk.Button(
            body,
            text="Artikelstamm importieren  →",
            command=self.import_artikelstamm,
            bg="#0b4a86",
            fg="white",
            relief="flat",
            font=(theme.FONT, 12, "bold"),
            padx=18,
            pady=9
        ).pack(anchor="w")

    def import_artikelstamm(self):
        file = filedialog.askopenfilename(
            title="Artikelstamm / PZN-Basis auswählen",
            filetypes=SUPPORTED_DATA_FILETYPES
        )
        if not file:
            return

        self.status.set("Artikelstamm-Import gestartet. Bitte Programm geöffnet lassen...")
        busy = self._show_busy_modal("Artikelstamm-Import", "Lese Datei und schreibe in Datenbank...")

        def progress(done, total):
            def update_status():
                if total:
                    pct = int((done / total) * 100)
                    self.status.set(f"Artikelstamm-Import läuft: {done:,} / {total:,} Zeilen ({pct} %)".replace(",", "."))
                    self._set_busy_message(busy, f"{done:,} / {total:,} Zeilen ({pct} %)".replace(",", "."))
                else:
                    self.status.set(f"Artikelstamm-Import läuft: {done:,} Zeilen".replace(",", "."))
                    self._set_busy_message(busy, f"{done:,} Zeilen verarbeitet".replace(",", "."))
            self.after(0, update_status)

        def worker():
            try:
                result = import_artikelstamm_excel(file, quelle="GUI-Import", progress_callback=progress)
                total = count_artikelstamm()
                msg = (
    "Artikelstamm importiert.\n\n"
    f"Importiert/aktualisiert: {result.get('imported_or_updated', result.get('inserted', 0))}\n"
    f"Übersprungen: {result.get('skipped', 0)}\n\n"
    f"Artikel gesamt: {total}"
)
                def done():
                    self._close_busy_modal(busy)
                    self.status.set(msg)
                    messagebox.showinfo("Artikelstamm", msg)
                    self.show_artikelstamm_page()
                self.after(0, done)
            except Exception as exc:
                # exc wird am Ende des except-Blocks geloescht; vor dem Scheduling
                # in eine normale lokale Variable kopieren, sonst sieht failed()
                # nur eine "free variable" und crasht spaeter im Tk-Loop.
                error_text = str(exc)
                def failed():
                    self._close_busy_modal(busy)
                    self.status.set(f"Artikelstamm-Import abgebrochen: {error_text}")
                    messagebox.showerror("Artikelstamm", error_text)
                self.after(0, failed)

        threading.Thread(target=worker, daemon=True).start()

    def show_austauschdatenbank_page(self):
        self.clear_page()
        self._page_header(
            "Austauschdatenbank",
            "Bekannte Austauschbeziehungen importieren und später über die Schulbank erweitern."
        )

        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))

        try:
            count = count_austauschdatenbank()
        except Exception:
            count = 0

        info = (
            f"Aktive Einträge: {count}\n\n"
            "Erwartete Excel-Spalten:\n"
            "- PZN\n"
            "- PZN NMG\n"
            "- austauschbar gegen\n\n"
            "Hinweis:\n"
            "Wenn keine PZN NMG vorhanden ist, wird der Freitext trotzdem gespeichert."
        )

        tk.Label(
            body,
            text=info,
            justify="left",
            bg="#ffffff",
            fg="#222",
            font=(theme.FONT, 11)
        ).pack(anchor="w", pady=(0, 16))

        tk.Button(
            body,
            text="Austauschdatenbank importieren  →",
            command=self.import_austauschdatenbank,
            bg="#0b4a86",
            fg="white",
            relief="flat",
            font=(theme.FONT, 12, "bold"),
            padx=18,
            pady=9
        ).pack(anchor="w")

    def import_austauschdatenbank(self):
        file = filedialog.askopenfilename(
            title="Austauschdatenbank auswählen",
            filetypes=SUPPORTED_DATA_FILETYPES
        )
        if not file:
            return

        try:
            result = self._run_busy(
                lambda: import_austausch_excel(file, quelle="GUI-Import"),
                title="Austauschdatenbank importieren",
                subtitle=f"Lese {Path(file).name} ...",
            )
            msg = (
                "Austauschdatenbank importiert.\n\n"
                f"Neu angelegt: {result.get('inserted', 0)}\n"
                f"Aktualisiert: {result.get('updated', 0)}\n"
                f"Übersprungen: {result.get('skipped', 0)}\n\n"
                f"Aktive Einträge gesamt: {count_austauschdatenbank()}"
            )
            self.status.set(msg)
            messagebox.showinfo("Austauschdatenbank", msg)
            self.show_austauschdatenbank_page()
        except Exception as exc:
            messagebox.showerror("Austauschdatenbank", str(exc))

    @staticmethod
    def _parse_decimal_or_none(value):
        """Konvertiert Excel-Werte ('5,99' / '5.99' / 5.99 / '') zu float oder None.
        Wird fuer numerische Spalten beim Import gebraucht, sonst kippt sqlite
        ueber Strings wie '5,99' rein. None bei leeren oder unparsbaren Werten.
        """
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip().replace(",", ".")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None

    @staticmethod
    def _parse_rabatt(value):
        """SP28: Spezial-Parser fuer Rabatt-Werte aus Excel.

        Akzeptiert:
        - '15%'   -> 0.15  (Text mit Prozent)
        - '15,5%' -> 0.155
        - '0,15'  -> 0.15  (deutscher Dezimalbruch)
        - 0.15    -> 0.15  (Excel-Zahl mit Prozentformat)
        - 15      -> 0.15  (Heuristik: Wert > 1 = als ganze Prozent gemeint)
        - 1.0     -> 1.0   (= 100%, wird belassen)
        Vorher: '15%' lief in float() -> ValueError -> None.
        Dadurch waren importierte Rabatte alle NULL, die Auswertung
        konnte keinen Rabatt anzeigen.
        """
        if value is None:
            return None
        is_percent_text = False
        if isinstance(value, (int, float)):
            v = float(value)
        else:
            s = str(value).strip().replace(",", ".")
            if not s:
                return None
            if s.endswith("%"):
                is_percent_text = True
                s = s[:-1].strip()
                if not s:
                    return None
            try:
                v = float(s)
            except ValueError:
                return None
        if is_percent_text:
            return v / 100.0
        # Heuristik: ein nackter Wert > 1 ist als Prozentpunkt-Zahl gemeint.
        if v > 1.0:
            return v / 100.0
        return v

    # ── GENERISCHER IMPORT-ASSISTENT ────────────────────────────────────────────
    def _import_assistent(self, title, ziel_tabelle, pflicht_spalten, optionale_spalten,
                           insert_sql, mapping_key, beschreibung="", numeric_fields=None,
                           rabatt_fields=None):
        """Generischer Import-Assistent: xlsx/csv/txt → Datenbank.
        Erkennt Spalten automatisch, bietet Mapping-Dialog bei Unklarheit.

        numeric_fields: Liste der Spaltennamen, die als Zahl (REAL) in die DB
        sollen. Diese werden ueber _parse_decimal_or_none normalisiert
        (Komma -> Punkt, leer -> None). Default: keine.
        rabatt_fields: SP28 - Spaltennamen, die als Rabatt-Prozent interpretiert
        werden sollen ('15%', '15,5%', 15 -> 0.15). Eine echte Untermenge von
        numeric_fields. Default: keine.
        """
        numeric_fields = set(numeric_fields or [])
        rabatt_fields = set(rabatt_fields or [])
        from .file_loader import load_table, SUPPORTED_DATA_FILETYPES, find_column, normalize_header
        from tkinter import filedialog

        files = filedialog.askopenfilenames(
            title=f"{title} – Datei(en) auswählen",
            filetypes=SUPPORTED_DATA_FILETYPES + [("PDF", "*.pdf"), ("Alle Dateien", "*.*")]
        )
        if not files:
            return

        gesamt_neu = 0
        gesamt_aktualisiert = 0
        fehler = []
        # SP13: Diagnose pro Datei sammeln, damit der User SIEHT was die App
        # gemacht hat. Vorher gab es nur "Neu importiert: 0" ohne Hinweis warum.
        diagnostik = []  # list of dicts mit Pro-Datei-Infos

        # Tabellen-Stand VOR dem Import festhalten (fuer Vorher/Nachher).
        count_vorher = None
        try:
            with sqlite3.connect(DB_PATH) as con_pre:
                cur = con_pre.execute(f'SELECT COUNT(*) FROM "{ziel_tabelle}"')
                count_vorher = int(cur.fetchone()[0])
        except Exception:
            count_vorher = None

        # V1.1 SP19: Vor jedem nmg_rabatte-Import einen Snapshot anlegen,
        # damit Diff/Verlauf in der NMG-Rabatte-Uebersicht moeglich sind.
        if ziel_tabelle == "nmg_rabatte":
            try:
                from .nmg_rabatte_history import take_snapshot
                with sqlite3.connect(DB_PATH) as con_snap:
                    quellen = ", ".join(Path(f).name for f in files[:3])
                    if len(files) > 3:
                        quellen += f" (+{len(files)-3} weitere)"
                    take_snapshot(con_snap, quelle=quellen, bemerkung=f"vor Import von {len(files)} Datei(en)")
            except Exception:
                pass

        busy = self._show_busy_modal(title, f"Lese {len(files)} Datei(en)...")
        try:
            for fi, filepath in enumerate(files, start=1):
                self._set_busy_message(busy, f"Datei {fi}/{len(files)}: {Path(filepath).name}")
                file_diag = {
                    "name": Path(filepath).name,
                    "status": "",
                    "rows_seen": 0,
                    "rows_inserted": 0,
                    "rows_updated": 0,
                    "rows_skipped_empty_pzn": 0,
                    "rows_consolidated": 0,
                    "columns_mapped": [],
                    "missing_pflicht": [],
                }
                try:
                    table = load_table(filepath)
                except Exception as exc:
                    file_diag["status"] = f"Datei nicht lesbar: {type(exc).__name__}: {exc}"
                    fehler.append(f"{Path(filepath).name}: {exc}")
                    diagnostik.append(file_diag)
                    continue

                headers = table.headers
                header_norm = [normalize_header(h) for h in headers]

                mapping = {}
                for ziel_col, aliases, contains in (pflicht_spalten + optionale_spalten):
                    idx = find_column(headers, aliases, contains)
                    if idx is not None:
                        mapping[ziel_col] = idx

                missing = [col for col, *_ in pflicht_spalten if col not in mapping]
                if missing:
                    saved = self._get_meta_value(f"{mapping_key}_{Path(filepath).stem[:20]}", "")
                    if saved:
                        try:
                            import json
                            mapping = json.loads(saved)
                            missing = [col for col, *_ in pflicht_spalten if col not in mapping]
                        except Exception:
                            pass

                if missing:
                    self._close_busy_modal(busy)
                    mapping = self._mapping_dialog(
                        title=f"{title} – Spalten zuordnen",
                        filepath=filepath,
                        headers=headers,
                        pflicht=pflicht_spalten,
                        optional=optionale_spalten,
                        current_mapping=mapping,
                        beschreibung=beschreibung,
                    )
                    if not mapping:
                        file_diag["status"] = "Spalten-Zuordnung abgebrochen → Datei uebersprungen"
                        file_diag["missing_pflicht"] = missing
                        diagnostik.append(file_diag)
                        busy = self._show_busy_modal(title, "Bitte warten...")
                        continue
                    try:
                        import json
                        self._set_meta_value(f"{mapping_key}_{Path(filepath).stem[:20]}", json.dumps(mapping))
                    except Exception:
                        pass
                    busy = self._show_busy_modal(title, f"Importiere {Path(filepath).name}...")
                else:
                    self._set_busy_message(busy, f"Importiere {Path(filepath).name}...")

                file_diag["columns_mapped"] = sorted(mapping.keys())

                neu = 0
                aktualisiert = 0
                rows_seen = 0
                rows_skipped_empty_pzn = 0
                rows_consolidated = 0  # SP29: Mehrfach-Rabatte pro PZN zusammengefasst
                try:
                    all_columns = [col for col, *_ in (pflicht_spalten + optionale_spalten)]

                    # SP29: Phase 1 - alle Zeilen zu Records bauen. Wenn rabatt_fields
                    # gesetzt sind (z.B. PK-Rabatte), pro PZN den hoechsten Wert pro
                    # Rabatt-Feld behalten, statt zufaellig die letzte Zeile gewinnen
                    # zu lassen. Quell-Dateien (z.B. SK-Rabatt-Listen) listen pro
                    # Artikel oft mehrere Staffeln; der hoechste Wert ist der fuer
                    # die Kundin/den Kunden vorteilhafteste.
                    pending = []
                    by_pzn = {}  # nur befuellt bei rabatt_fields
                    consolidate = bool(rabatt_fields)

                    for ri, row_data in enumerate(table.rows, start=1):
                        rows_seen += 1
                        record = {col: None for col in all_columns}
                        for col_name, col_idx in mapping.items():
                            if isinstance(col_idx, int) and col_idx < len(row_data):
                                raw = row_data[col_idx]
                                if col_name in rabatt_fields:
                                    # SP28: Rabatt-Spalten via _parse_rabatt - handhabt "15%",
                                    # "15,5%" und nackte Zahl 15 (-> 0.15).
                                    record[col_name] = self._parse_rabatt(raw)
                                elif col_name in numeric_fields:
                                    record[col_name] = self._parse_decimal_or_none(raw)
                                else:
                                    record[col_name] = raw if raw not in ("", None) else None
                        record["importdatum"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        record["quelle"] = Path(filepath).name

                        # SP13: Leere PZN bewusst ueberspringen statt NULL-Eintraege zu produzieren.
                        if "pzn" in record and (record["pzn"] is None or str(record["pzn"]).strip() == ""):
                            rows_skipped_empty_pzn += 1
                            continue

                        # SP24/SP29: PZN konsistent auf 8 Stellen normalisieren - inkl.
                        # Abschneiden von SK-Lagervariante-Suffix ' /N'.
                        if "pzn" in record and record["pzn"] not in (None, ""):
                            normalized = self._normalize_pzn_input(record["pzn"])
                            if normalized:
                                record["pzn"] = normalized

                        if ri % 500 == 0:
                            self._set_busy_message(busy, f"{Path(filepath).name}: {ri:,} Zeilen verarbeitet".replace(",", "."))

                        if consolidate and record.get("pzn"):
                            pzn_key = record["pzn"]
                            if pzn_key in by_pzn:
                                existing = pending[by_pzn[pzn_key]]
                                # Pro Rabatt-Feld Max() behalten.
                                for rf in rabatt_fields:
                                    new_val = record.get(rf)
                                    old_val = existing.get(rf)
                                    if new_val is not None and (old_val is None or new_val > old_val):
                                        existing[rf] = new_val
                                # Andere Felder nur fuellen, wenn bisher leer.
                                for col in all_columns:
                                    if col in rabatt_fields:
                                        continue
                                    if existing.get(col) in (None, "") and record.get(col) not in (None, ""):
                                        existing[col] = record[col]
                                rows_consolidated += 1
                                continue
                            by_pzn[pzn_key] = len(pending)
                        pending.append(record)

                    # SP29: Phase 2 - konsolidierte Records in die DB schreiben.
                    with sqlite3.connect(DB_PATH) as con:
                        for record in pending:
                            # SP28: Pre-Check, ob der Datensatz bereits existiert. Bei
                            # 'INSERT ... ON CONFLICT DO UPDATE' wirft SQLite KEIN
                            # IntegrityError - der UPSERT verhaelt sich nach aussen wie
                            # ein normaler INSERT. Ohne Pre-Check wurde 'neu += 1' fuer
                            # jeden UPSERT hochgezaehlt.
                            pk_col = "pzn" if "pzn" in record else "kundennummer"
                            pk_value = record.get(pk_col)
                            already_exists = False
                            if pk_value:
                                try:
                                    cur_chk = con.execute(
                                        f'SELECT 1 FROM "{ziel_tabelle}" WHERE {pk_col}=? LIMIT 1',
                                        (pk_value,),
                                    )
                                    already_exists = cur_chk.fetchone() is not None
                                except Exception:
                                    already_exists = False
                            try:
                                con.execute(insert_sql, record)
                                if already_exists:
                                    aktualisiert += 1
                                else:
                                    neu += 1
                            except sqlite3.IntegrityError:
                                set_parts = ", ".join(f"{k}=:{k}" for k in record if k not in ("pzn", "kundennummer"))
                                try:
                                    con.execute(f"UPDATE {ziel_tabelle} SET {set_parts} WHERE {pk_col}=:{pk_col}", record)
                                    aktualisiert += 1
                                except Exception as exc2:
                                    if len([f for f in fehler if "[update-fallback]" in f]) < 3:
                                        fehler.append(f"{Path(filepath).name} [update-fallback]: {type(exc2).__name__}: {exc2}")
                            except Exception as exc1:
                                if len([f for f in fehler if "[insert]" in f]) < 3:
                                    fehler.append(f"{Path(filepath).name} [insert]: {type(exc1).__name__}: {exc1}")
                        con.commit()
                    gesamt_neu += neu
                    gesamt_aktualisiert += aktualisiert
                    file_diag["status"] = "OK"
                    file_diag["rows_seen"] = rows_seen
                    file_diag["rows_inserted"] = neu
                    file_diag["rows_updated"] = aktualisiert
                    file_diag["rows_skipped_empty_pzn"] = rows_skipped_empty_pzn
                    file_diag["rows_consolidated"] = rows_consolidated
                except Exception as exc:
                    file_diag["status"] = f"Importfehler: {type(exc).__name__}: {exc}"
                    fehler.append(f"{Path(filepath).name}: {exc}")
                diagnostik.append(file_diag)
        finally:
            self._close_busy_modal(busy)

        # Tabellen-Stand NACH dem Import.
        count_nachher = None
        try:
            with sqlite3.connect(DB_PATH) as con_post:
                cur = con_post.execute(f'SELECT COUNT(*) FROM "{ziel_tabelle}"')
                count_nachher = int(cur.fetchone()[0])
        except Exception:
            count_nachher = None

        # SP13: Diagnose-Block fuer den Abschluss-Dialog bauen.
        msg_lines = [f"{title} – Diagnose", ""]
        msg_lines.append(f"Zieltabelle: {ziel_tabelle}")
        if count_vorher is not None and count_nachher is not None:
            delta = count_nachher - count_vorher
            sign = "+" if delta >= 0 else ""
            msg_lines.append(f"Datensätze vorher: {count_vorher:,}  →  nachher: {count_nachher:,}  ({sign}{delta:,})".replace(",", "."))
        msg_lines.append("")

        for i, d in enumerate(diagnostik, start=1):
            msg_lines.append(f"Datei {i}: {d['name']}")
            msg_lines.append(f"  Status: {d['status']}")
            if d['columns_mapped']:
                msg_lines.append(f"  Erkannte Spalten: {', '.join(d['columns_mapped'])}")
            if d['missing_pflicht']:
                msg_lines.append(f"  Fehlende Pflichtspalten: {', '.join(d['missing_pflicht'])}")
            if d['rows_seen']:
                msg_lines.append(f"  Zeilen gelesen: {d['rows_seen']:,}".replace(",", "."))
                msg_lines.append(f"  Neu eingefuegt: {d['rows_inserted']:,}".replace(",", "."))
                msg_lines.append(f"  Aktualisiert: {d['rows_updated']:,}".replace(",", "."))
                if d['rows_skipped_empty_pzn']:
                    msg_lines.append(f"  Uebersprungen (leere PZN): {d['rows_skipped_empty_pzn']:,}".replace(",", "."))
                if d.get('rows_consolidated'):
                    msg_lines.append(f"  Mehrfach-Eintraege zusammengefasst (hoechster Rabatt gewinnt): {d['rows_consolidated']:,}".replace(",", "."))
            msg_lines.append("")

        msg_lines.append(f"Zusammenfassung: {gesamt_neu} neu, {gesamt_aktualisiert} aktualisiert.")
        if fehler:
            msg_lines.append("")
            msg_lines.append(f"Fehler ({len(fehler)}):")
            for f in fehler[:5]:
                msg_lines.append(f"  {f}")

        msg = "\n".join(msg_lines)
        messagebox.showinfo(title, msg)
        self.status.set(f"{title}: {gesamt_neu} neu, {gesamt_aktualisiert} aktualisiert.")
        try:
            log_event("import", title, msg, user=self.bearbeiter)
        except Exception:
            pass
        # SP29: Statuszeile rechts (Counter, letzte Aktualisierung) sofort
        # auffrischen, damit der User nicht neu starten muss.
        self._refresh_status_sidebar()

    def _mapping_dialog(self, title, filepath, headers, pflicht, optional, current_mapping, beschreibung=""):
        """Dialog: Benutzer ordnet Spalten manuell zu. Gibt mapping-dict zurück oder None."""
        result = {"mapping": None}
        win = tk.Toplevel(self)
        win.resizable(True, True)
        win.title(title)
        win.geometry("700x560")
        win.minsize(600, 460)
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()

        tk.Label(win, text=title, font=(theme.FONT, 14, "bold"), fg="#0b4a86", bg="#f5f7fb").pack(anchor="w", padx=20, pady=(16, 4))
        tk.Label(win, text=f"Datei: {Path(filepath).name}", font=(theme.FONT, 9), fg="#555", bg="#f5f7fb").pack(anchor="w", padx=20)
        if beschreibung:
            tk.Label(win, text=beschreibung, font=(theme.FONT, 9), fg="#888", bg="#f5f7fb", justify="left", wraplength=640).pack(anchor="w", padx=20, pady=(2, 0))
        tk.Label(win, text="Spalten aus der Datei den Feldern zuordnen. Pflichtfelder (*) müssen belegt sein.",
                 font=(theme.FONT, 9), fg="#666", bg="#f5f7fb").pack(anchor="w", padx=20, pady=(4, 10))

        form = tk.Frame(win, bg="#ffffff", highlightbackground="#d8e2ee", highlightthickness=1)
        form.pack(fill="both", expand=True, padx=20, pady=(0, 10))
        form.columnconfigure(1, weight=1)

        header_choices = ["– nicht zuordnen –"] + [f"[{i}] {h}" for i, h in enumerate(headers)]
        combos = {}

        all_fields = [(col, aliases, contains, True) for col, aliases, contains in pflicht] +                      [(col, aliases, contains, False) for col, aliases, contains in optional]

        for r, (col_name, aliases, _, required) in enumerate(all_fields):
            label_text = f"{aliases[0]} {'*' if required else ''}".strip()
            color = "#c00" if required else "#0b4a86"
            tk.Label(form, text=label_text, bg="#ffffff", fg=color, font=(theme.FONT, 10, "bold"),
                     anchor="w").grid(row=r, column=0, sticky="w", padx=12, pady=5)
            var = tk.StringVar()
            if col_name in current_mapping:
                idx = current_mapping[col_name]
                if isinstance(idx, int) and idx < len(headers):
                    var.set(f"[{idx}] {headers[idx]}")
                else:
                    var.set(header_choices[0])
            else:
                var.set(header_choices[0])
            combos[col_name] = var
            cb = ttk.Combobox(form, textvariable=var, values=header_choices, state="readonly")
            cb.grid(row=r, column=1, sticky="ew", padx=12, pady=5)

        def save():
            mapping = {}
            for col_name, var in combos.items():
                val = var.get()
                if val.startswith("["):
                    try:
                        idx = int(val.split("]")[0][1:])
                        mapping[col_name] = idx
                    except Exception:
                        pass
            # Pflichtfelder prüfen
            missing = [col for col, *_ in pflicht if col not in mapping]
            if missing:
                messagebox.showinfo(title, f"Pflichtfelder nicht belegt: {', '.join(missing)}")
                return
            result["mapping"] = mapping
            win.destroy()

        bar = tk.Frame(win, bg="#f5f7fb")
        bar.pack(fill="x", padx=20, pady=(0, 14))
        tk.Button(bar, text="Abbrechen", command=win.destroy, padx=14, pady=7).pack(side="right", padx=(8, 0))
        tk.Button(bar, text="✔  Importieren", command=save, bg="#0b4a86", fg="white",
                  relief="flat", font=(theme.FONT, 11, "bold"), padx=16, pady=7).pack(side="right")
        self.wait_window(win)
        return result["mapping"]

    def _ensure_import_tables(self):
        """Erstellt Import-Zieltabellen falls noch nicht vorhanden.

        Hinweis: Die alten tbl_*_import-Tabellen bleiben als Schema-Skelett
        erhalten (falls externe Tools darauf zugreifen), werden aber seit
        V1.0 SP1 nicht mehr beschrieben. Wahrheits-Tabellen sind:
          NMG-Artikel + APU/HAP -> tbl_nmg_stamm
          PK-Rabatte            -> nmg_rabatte
          Partnerkonditionen    -> tbl_pk_konditionen (siehe unten, neu)
        """
        with sqlite3.connect(DB_PATH) as con:
            # Neue Wahrheits-Tabelle fuer Partnerkonditionen, vorher fehlte sie.
            con.execute("""CREATE TABLE IF NOT EXISTS tbl_pk_konditionen (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                kundennummer TEXT NOT NULL,
                kundenname TEXT,
                pzn TEXT,
                rabatt_prozent REAL,
                gueltigkeit TEXT,
                quelle TEXT,
                importdatum TEXT,
                letzte_aktualisierung TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(kundennummer, pzn)
            )""")
            con.execute("""CREATE TABLE IF NOT EXISTS tbl_nmg_stamm_import (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pzn TEXT UNIQUE,
                artikelname TEXT, hersteller TEXT, darreichungsform TEXT,
                packungsgroesse TEXT, apu REAL, taxe_ek REAL, taxe_vk REAL,
                importdatum TEXT, quelle TEXT
            )""")
            con.execute("""CREATE TABLE IF NOT EXISTS tbl_pk_rabatte_import (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pzn TEXT UNIQUE,
                artikelname TEXT, hersteller TEXT, rabatt_prozent REAL,
                rabatt_euro REAL, gueltigkeit TEXT,
                importdatum TEXT, quelle TEXT
            )""")
            con.execute("""CREATE TABLE IF NOT EXISTS tbl_apu_hap_import (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pzn TEXT UNIQUE,
                artikelname TEXT, apu REAL, hap REAL, hersteller TEXT,
                importdatum TEXT, quelle TEXT
            )""")
            con.execute("""CREATE TABLE IF NOT EXISTS tbl_pk_konditionen_import (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                kundennummer TEXT,
                kundenname TEXT, pzn TEXT, rabatt_prozent REAL,
                gueltigkeit TEXT,
                importdatum TEXT, quelle TEXT
            )""")
            con.commit()

    def import_nmg_articles(self):
        self._ensure_import_tables()
        self._import_assistent(
            title="NMG Artikel importieren",
            ziel_tabelle="tbl_nmg_stamm",
            pflicht_spalten=[
                ("pzn", ["PZN", "PZN-Code", "Artikel-PZN"], ["pzn"]),
            ],
            optionale_spalten=[
                ("artikelname", ["Artikelname", "Bezeichnung", "Name", "Artikel"], ["artikel", "name", "bezeichn"]),
                ("hersteller", ["Hersteller", "Lieferant", "Firma"], ["herst", "liefer"]),
                ("darreichungsform", ["Darreichungsform", "DF", "Form"], ["darreich", " df "]),
                ("packungsgroesse", ["Packungsgröße", "Packung", "PG", "Menge"], ["packung", " pg ", " menge"]),
                ("apu", ["APU", "Apotheken-PU", "Einkaufspreis"], ["apu", "ek", "einkauf"]),
                ("taxe_ek", ["Taxe EK", "HAP", "Großhandelspreis"], ["taxe", "hap", "gross"]),
                ("taxe_vk", ["Taxe VK", "Apothekenverkaufspreis", "AVP"], ["avp", "taxe vk", "vk"]),
            ],
            # Direkt in die Wahrheits-Tabelle (V1.0 SP1). Spalte 'hersteller'
            # wandert in herstellerkuerzel. DF/PCK werden gemappt aber nicht
            # gespeichert, da tbl_nmg_stamm sie nicht hat (Artikelstamm-Import
            # ist der richtige Pfad fuer DF/PCK).
            insert_sql="""INSERT INTO tbl_nmg_stamm(pzn,artikelname,herstellerkuerzel,apu,taxe_ek,taxe_vk,quelle,importdatum)
                VALUES(:pzn,:artikelname,:hersteller,:apu,:taxe_ek,:taxe_vk,:quelle,:importdatum)
                ON CONFLICT(pzn) DO UPDATE SET
                    artikelname=COALESCE(NULLIF(excluded.artikelname,''),tbl_nmg_stamm.artikelname),
                    herstellerkuerzel=COALESCE(NULLIF(excluded.herstellerkuerzel,''),tbl_nmg_stamm.herstellerkuerzel),
                    apu=COALESCE(excluded.apu,tbl_nmg_stamm.apu),
                    taxe_ek=COALESCE(excluded.taxe_ek,tbl_nmg_stamm.taxe_ek),
                    taxe_vk=COALESCE(excluded.taxe_vk,tbl_nmg_stamm.taxe_vk),
                    quelle=excluded.quelle,
                    importdatum=excluded.importdatum""",
            mapping_key="nmg_artikel",
            beschreibung="NMG-Artikelstamm importieren. Erwartet mindestens eine PZN-Spalte.",
            numeric_fields=["apu", "taxe_ek", "taxe_vk"],
        )

    def import_pk_rabatte(self):
        self._ensure_import_tables()
        # SP29: Slowakische Spalten-Header werden ergaenzt, weil SK-Quelldateien
        # (z.B. Apotheken-Lagerbestand) statt "Artikel" / "Rabatt" oft "tovar" /
        # "nazov" / "zlava" benutzen. Die Aliase werden case-insensitiv und ohne
        # Diakritika gegen die Datei-Header gematcht.
        self._import_assistent(
            title="PK Rabatte importieren",
            ziel_tabelle="nmg_rabatte",
            pflicht_spalten=[
                ("pzn", ["PZN", "PZN-Code", "Artikel-PZN", "ŠÚKL", "SUKL", "Kód"], ["pzn", "sukl", "kod"]),
            ],
            optionale_spalten=[
                ("artikelname", ["Artikelname", "Bezeichnung", "Name", "Artikel", "Tovar", "Názov", "Položka"], ["artikel", "name", "bezeichn", "tovar", "nazov", "polozka"]),
                ("hersteller", ["Hersteller", "Lieferant", "Výrobca", "Dodávateľ"], ["herst", "liefer", "vyrobca", "dodavat"]),
                ("rabatt_prozent", ["Rabatt %", "Rabatt Prozent", "Prozent", "PK-Rabatt", "Rabatt", "Zľava", "Zlava"], ["rabatt", "prozent", " %", "zlava", "zľava"]),
                ("rabatt_euro", ["Rabatt €", "Rabatt Euro", "Abschlag", "Zľava EUR"], ["euro", "rabatt eur", "abschlag"]),
                ("gueltigkeit", ["Gültigkeit", "Gültig bis", "Gültig ab", "Laufzeit", "Platnosť"], ["gueltig", "laufzeit", "platnost"]),
            ],
            # Direkt in die Wahrheits-Tabelle nmg_rabatte (V1.0 SP1).
            # Spaltenmapping: pzn -> nmg_pzn, artikelname -> artikel,
            # rabatt_prozent -> rabatt, importdatum -> letzte_aktualisierung.
            # rabatt_euro, hersteller, gueltigkeit landen (noch) nirgends -
            # Wahrheits-Tabelle hat sie nicht.
            insert_sql="""INSERT INTO nmg_rabatte(nmg_pzn,artikel,rabatt,quelle,letzte_aktualisierung)
                VALUES(:pzn,:artikelname,:rabatt_prozent,:quelle,:importdatum)
                ON CONFLICT(nmg_pzn) DO UPDATE SET
                    artikel=COALESCE(NULLIF(excluded.artikel,''),nmg_rabatte.artikel),
                    rabatt=COALESCE(excluded.rabatt,nmg_rabatte.rabatt),
                    quelle=excluded.quelle,
                    letzte_aktualisierung=excluded.letzte_aktualisierung""",
            mapping_key="pk_rabatte",
            beschreibung="Partnerkonditionen-Rabatte importieren. Erwartet PZN und Rabattspalte(n).",
            numeric_fields=["rabatt_prozent", "rabatt_euro"],
            rabatt_fields=["rabatt_prozent"],
        )

    def import_apu_data(self):
        self._ensure_import_tables()
        self._import_assistent(
            title="APU/HAP Daten importieren",
            ziel_tabelle="tbl_nmg_stamm",
            pflicht_spalten=[
                ("pzn", ["PZN", "PZN-Code", "Artikel-PZN"], ["pzn"]),
            ],
            optionale_spalten=[
                ("artikelname", ["Artikelname", "Bezeichnung", "Name", "Artikel"], ["artikel", "name"]),
                ("apu", ["APU", "Apotheken-PU", "APU-Preis"], ["apu"]),
                ("hap", ["HAP", "Großhandelspreis", "Handelspreis", "HEP"], ["hap", "gross", " hep"]),
                ("hersteller", ["Hersteller", "Lieferant"], ["herst", "liefer"]),
            ],
            # HAP -> taxe_ek (Grosshandelspreis = Taxe EK im DE-Pharma-Kontext).
            # Update-Modus: bestehende Stammdaten werden ergaenzt, nicht
            # ueberschrieben.
            insert_sql="""INSERT INTO tbl_nmg_stamm(pzn,artikelname,herstellerkuerzel,apu,taxe_ek,quelle,importdatum)
                VALUES(:pzn,:artikelname,:hersteller,:apu,:hap,:quelle,:importdatum)
                ON CONFLICT(pzn) DO UPDATE SET
                    artikelname=COALESCE(NULLIF(excluded.artikelname,''),tbl_nmg_stamm.artikelname),
                    herstellerkuerzel=COALESCE(NULLIF(excluded.herstellerkuerzel,''),tbl_nmg_stamm.herstellerkuerzel),
                    apu=COALESCE(excluded.apu,tbl_nmg_stamm.apu),
                    taxe_ek=COALESCE(excluded.taxe_ek,tbl_nmg_stamm.taxe_ek),
                    quelle=excluded.quelle,
                    importdatum=excluded.importdatum""",
            mapping_key="apu_hap",
            beschreibung="APU/HAP-Preisdaten importieren. Erwartet PZN, APU- und/oder HAP-Spalten.",
            numeric_fields=["apu", "hap"],
        )

    def import_pk_data(self):
        self._ensure_import_tables()
        self._import_assistent(
            title="Partnerkonditionen importieren",
            ziel_tabelle="tbl_pk_konditionen",
            pflicht_spalten=[
                ("kundennummer", ["Kundennummer", "KD-Nr", "Kunden-ID", "KundNr"], ["kundennr", "kd-nr", "kd nr"]),
            ],
            optionale_spalten=[
                ("kundenname", ["Kundenname", "Apotheke", "Name"], ["kundenname", "apotheke"]),
                ("pzn", ["PZN", "Artikel-PZN"], ["pzn"]),
                ("rabatt_prozent", ["Rabatt %", "Rabatt Prozent", "PK-Rabatt"], ["rabatt", "prozent"]),
                ("gueltigkeit", ["Gültigkeit", "Gültig bis", "Laufzeit"], ["gueltig", "laufzeit"]),
            ],
            # Wahrheits-Tabelle tbl_pk_konditionen wird in _ensure_import_tables
            # angelegt. UNIQUE(kundennummer,pzn) erlaubt mehrere Eintraege pro
            # Kunde (verschiedene Artikel) und ON CONFLICT pflegt sie.
            insert_sql="""INSERT INTO tbl_pk_konditionen(kundennummer,kundenname,pzn,rabatt_prozent,gueltigkeit,quelle,importdatum,letzte_aktualisierung)
                VALUES(:kundennummer,:kundenname,:pzn,:rabatt_prozent,:gueltigkeit,:quelle,:importdatum,:importdatum)
                ON CONFLICT(kundennummer,pzn) DO UPDATE SET
                    kundenname=COALESCE(NULLIF(excluded.kundenname,''),tbl_pk_konditionen.kundenname),
                    rabatt_prozent=COALESCE(excluded.rabatt_prozent,tbl_pk_konditionen.rabatt_prozent),
                    gueltigkeit=COALESCE(NULLIF(excluded.gueltigkeit,''),tbl_pk_konditionen.gueltigkeit),
                    quelle=excluded.quelle,
                    importdatum=excluded.importdatum,
                    letzte_aktualisierung=excluded.importdatum""",
            mapping_key="pk_konditionen",
            beschreibung="Partnerkonditionen importieren. Erwartet Kundennummer-Spalte.",
            numeric_fields=["rabatt_prozent"],
        )

    def check_online_update(self):
        """Manueller Check via 'Update / Backup' -> 'Update suchen'-Kachel.
        Anders als der Startup-Check zeigt diese Variante immer eine Rueckmeldung,
        auch wenn die aktuelle Version schon die neueste ist.
        """
        import threading
        busy = self._show_busy_modal("Update suchen", "Frage GitHub nach neuester Version...")

        def worker():
            try:
                tag, html_url, asset_url = self._fetch_latest_release_info(timeout=10)
                def on_main():
                    self._close_busy_modal(busy)
                    if not tag:
                        messagebox.showinfo(
                            "Update suchen",
                            "Konnte keine Release-Info von GitHub holen.\n"
                            "Pruefe deine Internetverbindung und versuche es erneut."
                        )
                        return
                    if self._version_tuple(tag) > self._version_tuple(APP_VERSION):
                        self._show_online_update_dialog(tag, html_url, asset_url)
                    else:
                        messagebox.showinfo(
                            "Update suchen",
                            f"Du hast bereits die aktuellste Version installiert.\n\n"
                            f"Installiert: V{APP_VERSION}\n"
                            f"Neueste auf GitHub: {tag}"
                        )
                self.after(0, on_main)
            except Exception as exc:
                error_text = f"{type(exc).__name__}: {exc}"
                def failed():
                    self._close_busy_modal(busy)
                    messagebox.showerror(
                        "Update suchen",
                        f"Konnte GitHub nicht erreichen:\n\n{error_text}\n\n"
                        f"Pruefe deine Internetverbindung."
                    )
                self.after(0, failed)

        threading.Thread(target=worker, daemon=True).start()

    def show_roadmap_page(self):
        self.clear_page()
        self._page_header(
            "Roadmap",
            "Wünsche, Ideen, offene Punkte, begonnene Arbeiten und erledigte Aufgaben."
        )

        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        body.columnconfigure(0, weight=1)
        body.rowconfigure(1, weight=1)

        top = tk.Frame(body, bg="#ffffff")
        top.grid(row=0, column=0, sticky="ew", pady=(0, 10))

        tk.Button(
            top,
            text="➕ Neuer Wunsch",
            command=self.add_roadmap_wunsch,
            bg="#0b4a86",
            fg="white",
            relief="flat",
            padx=14,
            pady=8
        ).pack(side="left", padx=(0, 8))

        tk.Button(
            top,
            text="Aktualisieren",
            command=self.show_roadmap_page,
            padx=14,
            pady=8
        ).pack(side="left")

        columns = ("id", "status", "bereich", "titel", "prioritaet", "erstellt")
        tree = ttk.Treeview(body, columns=columns, show="headings", selectmode="browse")
        tree.grid(row=1, column=0, sticky="nsew")

        headings = {
            "id": "ID",
            "status": "Status",
            "bereich": "Bereich",
            "titel": "Titel",
            "prioritaet": "Priorität",
            "erstellt": "Erstellt",
        }
    
        widths = {
            "id": 55,
            "status": 100,
            "bereich": 150,
            "titel": 360,
            "prioritaet": 90,
            "erstellt": 150,
        }

        for col in columns:
            tree.heading(col, text=headings[col])
            tree.column(col, width=widths[col], anchor="w")
        self._make_tree_sortable(tree, columns, headings)

        scrollbar = tk.Scrollbar(body, orient="vertical", command=tree.yview)
        scrollbar.grid(row=1, column=0, sticky="nse")
        tree.configure(yscrollcommand=scrollbar.set)

        detail = tk.StringVar(value="Eintrag auswählen, um Beschreibung zu sehen.")
        tk.Label(
            body,
            textvariable=detail,
            bg="#ffffff",
            fg="#333",
            justify="left",
            anchor="w",
            wraplength=850
        ).grid(row=2, column=0, sticky="ew", pady=(10, 0))

        rows = list_roadmap_items()
        item_map = {}

        for row in rows:
            iid = tree.insert(
                "",
                "end",
                values=(
                    row["id"],
                    row["status"],
                    row["bereich"],
                    row["titel"],
                    row["prioritaet"],
                    row["erstellt_am"],
                )
            )
            item_map[iid] = row

        def selected_id():
            sel = tree.selection()
            if not sel:
                messagebox.showinfo("Roadmap", "Bitte zuerst einen Roadmap-Eintrag auswählen.")
                return None
            return int(tree.item(sel[0], "values")[0])

        def on_select(event=None):
            sel = tree.selection()
            if not sel:
                detail.set("Eintrag auswählen, um Beschreibung zu sehen.")
                return
            row = item_map.get(sel[0])
            if not row:
                return
            text = (
                f"{row['titel']}\n"
                f"Bereich: {row['bereich']} | Status: {row['status']} | Priorität: {row['prioritaet']}\n\n"
                f"{row['beschreibung'] or ''}"
            )
            detail.set(text)

        tree.bind("<<TreeviewSelect>>", on_select)

        actionbar = tk.Frame(body, bg="#ffffff")
        actionbar.grid(row=3, column=0, sticky="ew", pady=(12, 0))

        def set_status(status):
            item_id = selected_id()
            if item_id is None:
                return
            update_roadmap_status(item_id, status)
            self.show_roadmap_page()

        tk.Button(actionbar, text="💡 Idee", command=lambda: set_status("Idee"), padx=12, pady=7).pack(side="left", padx=(0, 6))
        tk.Button(actionbar, text="⬜ Offen", command=lambda: set_status("Offen"), padx=12, pady=7).pack(side="left", padx=6)
        tk.Button(actionbar, text="🔄 Begonnen", command=lambda: set_status("Begonnen"), padx=12, pady=7).pack(side="left", padx=6)
        tk.Button(actionbar, text="✅ Erledigt", command=lambda: set_status("Erledigt"), padx=12, pady=7).pack(side="left", padx=6)

        self.status.set(f"Roadmap: {len(rows)} Einträge angezeigt.")


    def add_roadmap_wunsch(self):
        win = tk.Toplevel(self)
        win.resizable(True, True)
        win.title("Neuer Wunsch")
        win.geometry("560x430")
        win.configure(bg="#f5f7fb")
        win.transient(self)
        win.grab_set()

        tk.Label(win, text="Neuer Wunsch", font=(theme.FONT, 18, "bold"), fg="#0b4a86", bg="#f5f7fb").pack(anchor="w", padx=22, pady=(18, 8))

        form = tk.Frame(win, bg="#f5f7fb")
        form.pack(fill="both", expand=True, padx=22, pady=8)

        tk.Label(form, text="Bereich", bg="#f5f7fb").grid(row=0, column=0, sticky="w")
        bereich_var = tk.StringVar(value="Sonstiges")
        bereich = ttk.Combobox(
            form,
            textvariable=bereich_var,
            values=["GUI", "Analyse", "Schulbank", "Datenbank", "Import", "Export", "Kundenverwaltung", "E-Mail", "Roadmap", "Sonstiges"],
            state="readonly",
            width=35
        )
        bereich.grid(row=1, column=0, sticky="ew", pady=(0, 10))

        tk.Label(form, text="Titel / Headline", bg="#f5f7fb").grid(row=2, column=0, sticky="w")
        titel_var = tk.StringVar()
        tk.Entry(form, textvariable=titel_var, width=54).grid(row=3, column=0, sticky="ew", pady=(0, 10))
    
        tk.Label(form, text="Beschreibung / Freitext", bg="#f5f7fb").grid(row=4, column=0, sticky="w")
        text = tk.Text(form, height=9, width=54)
        text.grid(row=5, column=0, sticky="nsew", pady=(0, 10))
    
        form.columnconfigure(0, weight=1)
        form.rowconfigure(5, weight=1)
    
        def save():
            try:
                add_roadmap_item(
                    bereich=bereich_var.get(),
                    titel=titel_var.get(),
                    beschreibung=text.get("1.0", "end").strip(),
                    status="Idee",
                    prioritaet="Normal"
                )
                win.destroy()
                self.show_roadmap_page()
            except Exception as exc:
                messagebox.showerror("Roadmap", str(exc))

        buttons = tk.Frame(win, bg="#f5f7fb")
        buttons.pack(fill="x", padx=22, pady=(0, 18))
        tk.Button(buttons, text="Abbrechen", command=win.destroy, padx=14, pady=8).pack(side="right", padx=(8, 0))
        tk.Button(buttons, text="Speichern", command=save, bg="#0b4a86", fg="white", relief="flat", padx=18, pady=9).pack(side="right")

    def show_help_center(self):
        """Die Hilfe ist ein eigenes Fenster (bebildertes Handbuch). Diese
        eingebettete Seite startet sie und verweist darauf."""
        self.open_hilfe_app()
        self.clear_page()
        self._page_header("Hilfe", "Das bebilderte Handbuch öffnet sich in einem eigenen Fenster.")

        body = tk.Frame(self.page, bg="#ffffff")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        tk.Label(body, text="Die Hilfe wurde in einem eigenen Fenster geöffnet.\n"
                            "Hol sie bei Bedarf über die Taskleiste nach vorn.",
                 justify="left", bg="#ffffff", fg="#222", font=(theme.FONT, 11)).pack(anchor="w", pady=(0, 12))
        tk.Button(body, text="❓  Hilfe öffnen", command=self.open_hilfe_app,
                  bg="#208acd", fg="white", relief="flat", font=(theme.FONT, 11, "bold"),
                  padx=18, pady=8, cursor="hand2").pack(anchor="w")



def _show_splash_screen():
    """SP12: Splash mit Hintergrundbild (NMGone_Splash.png) + Sprach-Auswahl
    + Start-Button. Wenn das Splash-Bild fehlt, fallback auf Logo-only.
    Liefert True wenn der User Start geklickt hat, False bei Abbruch.
    """
    from .i18n import LANGUAGES, load_language, save_language, get_language, T

    load_language()

    splash_root = tk.Tk()
    splash_root.title("NMGone")
    splash_root.configure(bg="#ffffff")
    splash_root.resizable(False, False)
    try:
        splash_root.iconbitmap(str(ASSETS_DIR / "NMGone.ico"))
    except Exception:
        pass

    # Splash-Bild laden, falls vorhanden. Bestimmt die Fenstergroesse.
    splash_img = None
    img_w, img_h = 0, 0
    try:
        bg_path = ASSETS_DIR / "NMGone_Splash.png"
        if bg_path.exists():
            raw = tk.PhotoImage(file=str(bg_path))
            # Auf hoechstens 720 px Breite herunter skalieren.
            factor = max(1, int(raw.width() / 720 + 0.999))
            splash_img = raw.subsample(factor, factor)
            img_w, img_h = splash_img.width(), splash_img.height()
            splash_root._bg_img = splash_img
    except Exception:
        splash_img = None

    if splash_img is not None:
        # Unter dem Bild ein Steuerbereich von 130 px Hoehe.
        ww, wh = img_w, img_h + 130
    else:
        ww, wh = 520, 420

    sw = splash_root.winfo_screenwidth()
    sh = splash_root.winfo_screenheight()
    splash_root.geometry(f"{ww}x{wh}+{(sw - ww) // 2}+{max(20, (sh - wh) // 3)}")

    # Layout: Bild oben, Steuerbereich unten
    if splash_img is not None:
        tk.Label(splash_root, image=splash_img, bg="#ffffff", bd=0).pack(side="top")
    else:
        # Fallback: nur Logo
        try:
            logo_path = ASSETS_DIR / "NMGone.png"
            if logo_path.exists():
                raw = tk.PhotoImage(file=str(logo_path))
                factor = max(1, int(max(raw.width() / 260, raw.height() / 110) + 0.999))
                splash_root._logo_img = raw.subsample(factor, factor)
                tk.Label(splash_root, image=splash_root._logo_img, bg="#ffffff").pack(pady=(28, 12))
        except Exception:
            pass
        welcome_var = tk.StringVar(value=T("Willkommen bei NMGone"))
        tk.Label(splash_root, textvariable=welcome_var, font=(theme.FONT, 16, "bold"),
                 fg="#0b4a86", bg="#ffffff").pack(pady=(0, 8))

    # Steuerbereich
    controls = tk.Frame(splash_root, bg="#ffffff")
    controls.pack(side="top", fill="x", pady=(8, 0))

    lang_label_var = tk.StringVar(value=T("Sprache auswählen"))
    tk.Label(controls, textvariable=lang_label_var, font=(theme.FONT, 10),
             fg="#444", bg="#ffffff").pack()

    lang_codes = list(LANGUAGES.keys())
    lang_display = [f"{LANGUAGES[c]} ({c})" for c in lang_codes]
    current_idx = lang_codes.index(get_language()) if get_language() in lang_codes else 0
    lang_var = tk.StringVar(value=lang_display[current_idx])

    combo = ttk.Combobox(controls, textvariable=lang_var, values=lang_display,
                         state="readonly", font=(theme.FONT, 11), width=22, justify="center")
    combo.pack(pady=(4, 12))

    def _on_lang_change(_event=None):
        sel = lang_var.get()
        for c in lang_codes:
            if sel.endswith(f"({c})"):
                save_language(c)
                lang_label_var.set(T("Sprache auswählen"))
                start_btn.config(text=T("Starten"))
                break

    combo.bind("<<ComboboxSelected>>", _on_lang_change)

    started = {"ok": False}

    def _on_start():
        _on_lang_change()
        started["ok"] = True
        splash_root.destroy()

    start_btn = tk.Button(
        controls, text=T("Starten"), command=_on_start,
        bg="#0b4a86", fg="white", relief="flat",
        font=(theme.FONT, 12, "bold"), padx=32, pady=8, cursor="hand2",
    )
    start_btn.pack(pady=(0, 10))

    from .backup import APP_VERSION_DISPLAY
    tk.Label(splash_root, text=f"NMGone {APP_VERSION_DISPLAY}",
             font=(theme.FONT, 8), fg="#aaa", bg="#ffffff").pack(side="bottom", pady=(0, 6))

    splash_root.bind("<Return>", lambda e: _on_start())
    splash_root.bind("<Escape>", lambda e: splash_root.destroy())

    splash_root.mainloop()
    return started["ok"]


def main():
    # Eigene AppUserModelID -> NMGone und die Kasse (NMG.Kasse) bilden getrennte
    # Taskleisten-Gruppen mit eigenem Icon. Muss VOR dem ersten Fenster gesetzt werden.
    if os.name == "nt":
        try:
            import ctypes
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("NMG.NMGone")
        except Exception:
            pass
    # SP11: erst Splash, dann Hauptprogramm. Wenn der User die Splash
    # ueber X/Escape schliesst, wird das Hauptprogramm nicht gestartet.
    if not _show_splash_screen():
        return
    app = NMGApp()
    app.mainloop()