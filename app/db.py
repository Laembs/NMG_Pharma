import csv
import re
import sqlite3
from pathlib import Path
from openpyxl import load_workbook
from .config import DB_PATH, DATA_DIR, REFERENCE_XLSX, LINDEN_REFERENCE_XLSX, ROSEN_REFERENCE_XLSX, SONNEN_REFERENCE_XLSX, HANDCHECKED_DIR, HISTORICAL_ANALYSIS_DIR

NMG_ARTIKELLISTE_CSV = DATA_DIR / "NMG_ARTIKELLISTE_APU_HAP.csv"
REPORT_HERSTELLER_CSV = DATA_DIR / "REPORT_HERSTELLER.csv"

SCHEMA_SQL = """
PRAGMA foreign_keys = ON;
DROP TABLE IF EXISTS meta;
DROP TABLE IF EXISTS tbl_nmg_stamm;
DROP TABLE IF EXISTS tbl_hersteller_lern;
DROP TABLE IF EXISTS tbl_hersteller_stimmen;
DROP TABLE IF EXISTS tbl_pzn_basisdaten;
DROP TABLE IF EXISTS tbl_pzn_basis_stimmen;
DROP TABLE IF EXISTS tbl_pzn_ek_rohdaten;
DROP TABLE IF EXISTS tbl_austauschartikel;
DROP TABLE IF EXISTS tbl_lieferfaehigkeit;
DROP TABLE IF EXISTS tbl_lernhistorie;
DROP TABLE IF EXISTS tbl_referenz_h_o;
DROP TABLE IF EXISTS tbl_import_log;
DROP TABLE IF EXISTS tbl_auswertungspositionen;
DROP TABLE IF EXISTS tbl_auswertungen;
DROP TABLE IF EXISTS tbl_rohdaten_mapping;
DROP TABLE IF EXISTS nmg_rabatte;

CREATE TABLE meta(key TEXT PRIMARY KEY, value TEXT NOT NULL);

CREATE TABLE tbl_nmg_stamm(
    pzn TEXT PRIMARY KEY,
    artikelname TEXT,
    herstellerkuerzel TEXT,
    apu REAL,
    taxe_ek REAL,
    taxe_vk REAL,
    menge TEXT,
    einheit TEXT,
    wirkstoffe TEXT,
    quelle TEXT,
    importdatum TEXT DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX idx_tbl_nmg_stamm_herst ON tbl_nmg_stamm(herstellerkuerzel);

CREATE TABLE tbl_hersteller_lern(
    pzn TEXT PRIMARY KEY,
    herstellerkuerzel TEXT NOT NULL,
    herstellername TEXT DEFAULT '',
    treffer INTEGER DEFAULT 1,
    quelle TEXT,
    letzte_aktualisierung TEXT DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE tbl_hersteller_stimmen(
    pzn TEXT NOT NULL,
    herstellerkuerzel TEXT NOT NULL,
    anzahl INTEGER DEFAULT 1,
    quelle TEXT,
    letzte_aktualisierung TEXT DEFAULT CURRENT_TIMESTAMP,
    PRIMARY KEY(pzn, herstellerkuerzel)
);

CREATE TABLE tbl_pzn_basisdaten(
    pzn TEXT PRIMARY KEY,
    artikelname TEXT,
    herstellerkuerzel TEXT,
    df TEXT,
    pck TEXT,
    treffer INTEGER DEFAULT 1,
    quelle TEXT,
    letzte_aktualisierung TEXT DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE tbl_pzn_basis_stimmen(
    pzn TEXT NOT NULL,
    feld TEXT NOT NULL,
    wert TEXT NOT NULL,
    anzahl INTEGER DEFAULT 1,
    quelle TEXT,
    letzte_aktualisierung TEXT DEFAULT CURRENT_TIMESTAMP,
    PRIMARY KEY(pzn, feld, wert)
);
CREATE TABLE tbl_pzn_ek_rohdaten(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    pzn TEXT NOT NULL,
    ek REAL NOT NULL,
    quelle_datei TEXT,
    ek_spalte TEXT,
    importdatum TEXT DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX idx_tbl_pzn_ek_rohdaten_pzn_datum ON tbl_pzn_ek_rohdaten(pzn, importdatum);

CREATE TABLE nmg_rabatte(
    nmg_pzn TEXT PRIMARY KEY,
    artikel TEXT,
    rabatt REAL,
    quelle TEXT,
    letzte_aktualisierung TEXT,
    gueltig_ab TEXT
);
CREATE TABLE tbl_austauschartikel(
    original_pzn TEXT PRIMARY KEY,
    original_artikel TEXT,
    nmg_pzn TEXT,
    austauschbar_gegen TEXT,
    quelle TEXT,
    letzte_aktualisierung TEXT,
    treffer_anzahl INTEGER,
    bemerkung TEXT
);
CREATE TABLE tbl_lieferfaehigkeit(
    nmg_pzn TEXT PRIMARY KEY,
    lieferbar TEXT,
    bevorratung_angeraten TEXT,
    liefervorschlag TEXT,
    quelle TEXT,
    letzte_aktualisierung TEXT
);
CREATE TABLE tbl_referenz_h_o(
    original_pzn TEXT PRIMARY KEY,
    im_sortiment TEXT,
    nmg_pzn TEXT,
    apu_nmg REAL,
    rabatt REAL,
    lieferbar TEXT,
    bevorratung_angeraten TEXT,
    liefervorschlag TEXT,
    austauschbar_gegen TEXT
);
CREATE TABLE tbl_lernhistorie(
    zeitpunkt TEXT,
    quelle TEXT,
    original_pzn TEXT,
    original_artikel TEXT,
    nmg_pzn TEXT,
    austauschbar_gegen TEXT,
    aktion TEXT,
    status TEXT
);

CREATE TABLE tbl_auswertungen(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    apotheke TEXT,
    quelldatei TEXT,
    ausgabedatei TEXT,
    datum TEXT DEFAULT CURRENT_TIMESTAMP,
    anzahl_positionen INTEGER DEFAULT 0,
    nmg_treffer INTEGER DEFAULT 0,
    nicht_nmg INTEGER DEFAULT 0,
    gesamt_absatz REAL DEFAULT 0,
    bemerkung TEXT,
    datenquelle TEXT DEFAULT 'NMG'
);
CREATE TABLE tbl_auswertungspositionen(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    auswertung_id INTEGER NOT NULL,
    pzn TEXT,
    artikelname TEXT,
    df TEXT,
    pck TEXT,
    herstellerkuerzel TEXT,
    ek REAL,
    absatz_6m REAL,
    im_sortiment TEXT,
    pzn_nmg TEXT,
    apu_nmg REAL,
    nmg_rabatt REAL,
    lieferbar TEXT,
    bevorratung_angeraten TEXT,
    liefervorschlag TEXT,
    austauschbar_gegen TEXT,
    nmg_rabatt_euro REAL,
    nmg_rabatt_gesamt REAL,
    umsatz REAL,
    ist_nmg_treffer INTEGER DEFAULT 0,
    quelle TEXT,
    datenquelle TEXT DEFAULT 'NMG',
    FOREIGN KEY(auswertung_id) REFERENCES tbl_auswertungen(id) ON DELETE CASCADE
);
CREATE INDEX idx_tbl_auswertungspositionen_pzn ON tbl_auswertungspositionen(pzn);
CREATE INDEX idx_tbl_auswertungspositionen_nmg_absatz ON tbl_auswertungspositionen(ist_nmg_treffer, absatz_6m);

CREATE TABLE tbl_import_log(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    datei TEXT,
    typ TEXT,
    datensaetze INTEGER,
    meldung TEXT,
    datum TEXT DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE tbl_rohdaten_mapping(
    dateiname TEXT PRIMARY KEY,
    pzn_spalte TEXT,
    hersteller_spalte TEXT,
    ek_spalte TEXT,
    absatz_spalte TEXT,
    header_zeile INTEGER,
    format_typ TEXT DEFAULT 'standard',
    zeitraum_monate INTEGER DEFAULT 6,
    artikel_spalte TEXT,
    bearbeiter TEXT,
    letzte_aktualisierung TEXT DEFAULT CURRENT_TIMESTAMP
);
"""

def _pzn(value):
    if value is None:
        return ""
    text = str(value).strip().replace(".0", "")
    text = re.sub(r"\D", "", text) if re.search(r"\d", text) else text
    return text.zfill(8) if text.isdigit() else ""

def _num(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("€", "").replace("%", "").replace(" ", "")
    try:
        return float(s.replace(".", "").replace(",", "."))
    except Exception:
        try:
            return float(s)
        except Exception:
            return None

def _norm_header(value):
    text = str(value or "").strip().lower()
    text = text.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    return re.sub(r"[^a-z0-9]+", "", text)

def connect(db_path: Path = DB_PATH) -> sqlite3.Connection:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys = ON")
    return con

def _read_rows(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    return [row for row in ws.iter_rows(min_row=2, values_only=True) if any(v not in (None, "") for v in row)]

def _clean_hersteller(value):
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return ""
    lines = [line.strip() for line in text.split("\n")]
    i = 0
    saw_numeric_prefix = False
    while i < len(lines) and re.fullmatch(r"\d+", lines[i] or ""):
        saw_numeric_prefix = True
        i += 1
    blank_count = 0
    while i < len(lines) and lines[i] == "":
        blank_count += 1
        i += 1
    if saw_numeric_prefix and blank_count >= 1 and i < len(lines):
        return " ".join(line for line in lines[i:] if line).strip()
    return " ".join(line for line in lines if line).strip()


def register_basisdaten_simple(con, pzn, artikelname=None, herstellerkuerzel=None, df=None, pck=None, quelle=""):
    """Kleine DB-interne Variante für Startdaten. Die volle Lernlogik liegt in learning_db.py."""
    pzn = _pzn(pzn)
    if not pzn:
        return False
    def clean(v):
        return str(v).strip() if v not in (None, "") else None
    artikelname = clean(artikelname)
    herstellerkuerzel = _clean_hersteller(herstellerkuerzel) or None
    df = clean(df)
    pck = clean(pck)
    if not any([artikelname, herstellerkuerzel, df, pck]):
        return False
    con.execute("""
        INSERT INTO tbl_pzn_basisdaten(pzn, artikelname, herstellerkuerzel, df, pck, treffer, quelle, letzte_aktualisierung)
        VALUES (?, ?, ?, ?, ?, 1, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(pzn) DO UPDATE SET
            artikelname=COALESCE(excluded.artikelname, tbl_pzn_basisdaten.artikelname),
            herstellerkuerzel=COALESCE(excluded.herstellerkuerzel, tbl_pzn_basisdaten.herstellerkuerzel),
            df=COALESCE(excluded.df, tbl_pzn_basisdaten.df),
            pck=COALESCE(excluded.pck, tbl_pzn_basisdaten.pck),
            treffer=tbl_pzn_basisdaten.treffer+1,
            quelle=excluded.quelle,
            letzte_aktualisierung=CURRENT_TIMESTAMP
    """, (pzn, artikelname, herstellerkuerzel, df, pck, quelle))
    for feld, wert in [("herstellerkuerzel", herstellerkuerzel), ("df", df), ("pck", pck)]:
        if wert:
            con.execute("""
                INSERT INTO tbl_pzn_basis_stimmen(pzn, feld, wert, anzahl, quelle, letzte_aktualisierung)
                VALUES (?, ?, ?, 1, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(pzn, feld, wert) DO UPDATE SET
                    anzahl=anzahl+1, quelle=excluded.quelle, letzte_aktualisierung=CURRENT_TIMESTAMP
            """, (pzn, feld, wert, quelle))
    return True

def register_hersteller(con, pzn, kuerzel, quelle):
    pzn = _pzn(pzn)
    kuerzel = _clean_hersteller(kuerzel)
    if not pzn or not kuerzel:
        return False
    con.execute("""
        INSERT INTO tbl_hersteller_stimmen(pzn, herstellerkuerzel, anzahl, quelle, letzte_aktualisierung)
        VALUES (?, ?, 1, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(pzn, herstellerkuerzel) DO UPDATE SET
            anzahl=anzahl+1, quelle=excluded.quelle, letzte_aktualisierung=CURRENT_TIMESTAMP
    """, (pzn, kuerzel, quelle))
    winner = con.execute("""
        SELECT herstellerkuerzel, anzahl FROM tbl_hersteller_stimmen
        WHERE pzn=? ORDER BY anzahl DESC, letzte_aktualisierung DESC LIMIT 1
    """, (pzn,)).fetchone()
    if winner:
        register_basisdaten_simple(con, pzn, herstellerkuerzel=winner[0], quelle=quelle)
        con.execute("""
            INSERT INTO tbl_hersteller_lern(pzn, herstellerkuerzel, herstellername, treffer, quelle, letzte_aktualisierung)
            VALUES (?, ?, '', ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(pzn) DO UPDATE SET
                herstellerkuerzel=excluded.herstellerkuerzel,
                treffer=excluded.treffer,
                quelle=excluded.quelle,
                letzte_aktualisierung=CURRENT_TIMESTAMP
        """, (pzn, winner[0], winner[1], quelle))
    return True

def import_nmg_artikelliste(path: str | Path, con=None) -> dict:
    path = Path(path)
    own = con is None
    con = con or connect()
    stats = {"imported": 0, "skipped": 0}
    try:
        if path.suffix.lower() == ".csv":
            raw = None
            for enc in ("utf-8-sig", "cp1252", "latin1"):
                try:
                    raw = path.read_text(encoding=enc)
                    break
                except UnicodeDecodeError:
                    continue
            if raw is None:
                raw = path.read_text(errors="replace")
            try:
                dialect = csv.Sniffer().sniff(raw[:4096], delimiters=";,	,")
            except Exception:
                dialect = csv.excel; dialect.delimiter = ";"
            rows = list(csv.reader(raw.splitlines(), dialect))
            if not rows:
                return stats
            headers = [_norm_header(h) for h in rows[0]]
            def col(*names):
                wanted = {_norm_header(n) for n in names}
                for i, h in enumerate(headers):
                    if h in wanted:
                        return i
                for i, h in enumerate(headers):
                    if any(w and w in h for w in wanted):
                        return i
                return None
            idx = {
                "pzn": col("PZN"),
                "artikel": col("Artikelname", "Artikel"),
                "herst": col("Anbieter", "Hersteller", "Herst"),
                "apu": col("APU / HAP", "APU", "HAP"),
                "taxe_ek": col("Taxe-EK", "TAX-EK"),
                "taxe_vk": col("Taxe-VK"),
                "menge": col("Menge"),
                "einheit": col("Einh.", "Einheit"),
                "wirkstoffe": col("Wirkstoffe und Mengen", "Wirkstoffe"),
            }
            for row in rows[1:]:
                pzn_i = idx["pzn"]
                pzn = _pzn(row[pzn_i] if pzn_i is not None and len(row) > pzn_i else None)
                if not pzn:
                    stats["skipped"] += 1; continue
                def val(key):
                    i = idx[key]
                    return row[i] if i is not None and len(row) > i else None
                con.execute("""
                    INSERT INTO tbl_nmg_stamm(pzn, artikelname, herstellerkuerzel, apu, taxe_ek, taxe_vk, menge, einheit, wirkstoffe, quelle, importdatum)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                    ON CONFLICT(pzn) DO UPDATE SET
                        artikelname=excluded.artikelname,
                        herstellerkuerzel=excluded.herstellerkuerzel,
                        apu=excluded.apu,
                        taxe_ek=excluded.taxe_ek,
                        taxe_vk=excluded.taxe_vk,
                        menge=excluded.menge,
                        einheit=excluded.einheit,
                        wirkstoffe=excluded.wirkstoffe,
                        quelle=excluded.quelle,
                        importdatum=CURRENT_TIMESTAMP
                """, (pzn, val("artikel"), _clean_hersteller(val("herst")), _num(val("apu")), _num(val("taxe_ek")), _num(val("taxe_vk")), val("menge"), val("einheit"), val("wirkstoffe"), path.name))
                register_basisdaten_simple(con, pzn, val("artikel"), val("herst"), None, None, f"NMG-Stamm:{path.name}")
                if val("herst"):
                    register_hersteller(con, pzn, val("herst"), f"NMG-Stamm:{path.name}")
                stats["imported"] += 1
        else:
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return stats
            headers = [_norm_header(h) for h in rows[0]]
            # minimal Excel support same as CSV
            def col(*names):
                wanted = {_norm_header(n) for n in names}
                return next((i for i, h in enumerate(headers) if h in wanted or any(w and w in h for w in wanted)), None)
            idx = {"pzn": col("PZN"), "artikel": col("Artikelname", "Artikel"), "herst": col("Anbieter", "Hersteller", "Herst"), "apu": col("APU / HAP", "APU", "HAP"), "taxe_ek": col("Taxe-EK", "TAX-EK"), "taxe_vk": col("Taxe-VK"), "menge": col("Menge"), "einheit": col("Einh.", "Einheit"), "wirkstoffe": col("Wirkstoffe und Mengen", "Wirkstoffe")}
            for row in rows[1:]:
                pzn = _pzn(row[idx["pzn"]] if idx["pzn"] is not None and len(row) > idx["pzn"] else None)
                if not pzn:
                    stats["skipped"] += 1; continue
                def val(key):
                    i = idx[key]
                    return row[i] if i is not None and len(row) > i else None
                con.execute("""
                    INSERT INTO tbl_nmg_stamm(pzn, artikelname, herstellerkuerzel, apu, taxe_ek, taxe_vk, menge, einheit, wirkstoffe, quelle, importdatum)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                    ON CONFLICT(pzn) DO UPDATE SET artikelname=excluded.artikelname, herstellerkuerzel=excluded.herstellerkuerzel, apu=excluded.apu, taxe_ek=excluded.taxe_ek, taxe_vk=excluded.taxe_vk, menge=excluded.menge, einheit=excluded.einheit, wirkstoffe=excluded.wirkstoffe, quelle=excluded.quelle, importdatum=CURRENT_TIMESTAMP
                """, (pzn, val("artikel"), _clean_hersteller(val("herst")), _num(val("apu")), _num(val("taxe_ek")), _num(val("taxe_vk")), val("menge"), val("einheit"), val("wirkstoffe"), path.name))
                register_basisdaten_simple(con, pzn, val("artikel"), val("herst"), None, None, f"NMG-Stamm:{path.name}")
                if val("herst"):
                    register_hersteller(con, pzn, val("herst"), f"NMG-Stamm:{path.name}")
                stats["imported"] += 1
        con.execute("INSERT INTO tbl_import_log(datei, typ, datensaetze, meldung) VALUES (?, 'nmg_stamm', ?, ?)", (path.name, stats["imported"], f"{stats['imported']} NMG-Stammdaten importiert"))
        if own:
            con.commit()
        return stats
    finally:
        if own:
            con.close()


def ensure_runtime_migrations(db_path: Path = DB_PATH) -> None:
    """Ergänzt neue Spalten, ohne vorhandene Lern- und Analysedaten zu löschen."""
    try:
        from .migrations import run_migrations
        run_migrations(db_path)
    except Exception:
        pass
    if not Path(db_path).exists():
        return
    with connect(db_path) as con:
        def cols(table):
            try:
                return {r[1] for r in con.execute(f"PRAGMA table_info({table})").fetchall()}
            except Exception:
                return set()
        if "datenquelle" not in cols("tbl_auswertungen"):
            con.execute("ALTER TABLE tbl_auswertungen ADD COLUMN datenquelle TEXT DEFAULT 'NMG'")
        if "datenquelle" not in cols("tbl_auswertungspositionen"):
            con.execute("ALTER TABLE tbl_auswertungspositionen ADD COLUMN datenquelle TEXT DEFAULT 'NMG'")
        for col_name in ("kundentyp", "kundennummer", "kundenname"):
            if col_name not in cols("tbl_auswertungen"):
                con.execute(f"ALTER TABLE tbl_auswertungen ADD COLUMN {col_name} TEXT")
        mapping_cols = cols("tbl_rohdaten_mapping")
        if mapping_cols:
            if "zeitraum_monate" not in mapping_cols:
                con.execute("ALTER TABLE tbl_rohdaten_mapping ADD COLUMN zeitraum_monate INTEGER DEFAULT 6")
            if "artikel_spalte" not in mapping_cols:
                con.execute("ALTER TABLE tbl_rohdaten_mapping ADD COLUMN artikel_spalte TEXT")
            if "bearbeiter" not in mapping_cols:
                con.execute("ALTER TABLE tbl_rohdaten_mapping ADD COLUMN bearbeiter TEXT")
        con.execute("UPDATE tbl_auswertungen SET datenquelle='NMG' WHERE datenquelle IS NULL OR datenquelle='' ")
        con.execute("UPDATE tbl_auswertungspositionen SET datenquelle='NMG' WHERE datenquelle IS NULL OR datenquelle='' ")
        con.execute("INSERT OR REPLACE INTO meta(key,value) VALUES('app_feature_zf_import','2.5.0')")

        # Vorbereitung Kundenstamm für neue Auswertung: PK und Zukunftswerk getrennt.
        con.execute("""
            CREATE TABLE IF NOT EXISTS tbl_pk_kunden (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                kundennummer TEXT,
                kundenname TEXT,
                apotheke TEXT,
                status TEXT NOT NULL DEFAULT 'aktiv',
                erstellt_am TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                aktualisiert_am TEXT,
                bearbeiter TEXT
            )
        """)
        con.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_tbl_pk_kunden_kundennummer
            ON tbl_pk_kunden(kundennummer)
            WHERE kundennummer IS NOT NULL AND kundennummer <> ''
        """)
        con.execute("""
            CREATE TABLE IF NOT EXISTS tbl_zf_kunden (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                kundennummer TEXT,
                kundenname TEXT,
                apotheke TEXT,
                status TEXT NOT NULL DEFAULT 'aktiv',
                erstellt_am TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                aktualisiert_am TEXT,
                bearbeiter TEXT
            )
        """)
        con.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_tbl_zf_kunden_kundennummer
            ON tbl_zf_kunden(kundennummer)
            WHERE kundennummer IS NOT NULL AND kundennummer <> ''
        """)
        con.execute("INSERT OR IGNORE INTO meta(key,value) VALUES('neue_auswertung_kundentyp','PK')")
        con.execute("INSERT OR REPLACE INTO meta(key,value) VALUES('app_feature_neue_auswertung_form','4.0')")
        con.commit()

def init_db(db_path: Path = DB_PATH) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    # Version 1.9 wird mit Startdatenbank ausgeliefert. Wenn sie vorhanden ist,
    # nicht bei jedem Programmstart langsam neu aufbauen.
    if Path(db_path).exists():
        try:
            test_con = connect(db_path)
            row = test_con.execute("SELECT value FROM meta WHERE key='db_version'").fetchone()
            test_con.close()
            if row and str(row[0]) == '1.9':
                ensure_runtime_migrations(db_path)
                return
        except Exception:
            pass
    with connect(db_path) as con:
        con.executescript(SCHEMA_SQL)
        con.execute("INSERT INTO meta(key,value) VALUES('db_version','1.9')")
        con.execute("INSERT INTO meta(key,value) VALUES('basis','2.6: Installations- und Update-System + PK/ZF-Datenquellen')")
        con.execute("INSERT INTO meta(key,value) VALUES('db_schema_version','1.1')")

        # NMG-Stamm aus offizieller APU/HAP-Liste zuerst befüllen.
        if NMG_ARTIKELLISTE_CSV.exists():
            import_nmg_artikelliste(NMG_ARTIKELLISTE_CSV, con)

        # Vollversion 1.3 bleibt Quelle für Rabatte, Austausch, Lieferfähigkeit und Lernhistorie.
        if REFERENCE_XLSX.exists():
            wb = load_workbook(REFERENCE_XLSX, data_only=True, read_only=True)
            for r in _read_rows(wb, 'DB_NMG_Rabatte'):
                con.execute("INSERT OR REPLACE INTO nmg_rabatte(nmg_pzn,artikel,rabatt,quelle,letzte_aktualisierung) VALUES (?,?,?,?,?)", (_pzn(r[0]), r[1], _num(r[2]), r[3], str(r[4] or '')))
            for r in _read_rows(wb, 'DB_Austausch_Artikel'):
                con.execute("INSERT OR REPLACE INTO tbl_austauschartikel VALUES (?,?,?,?,?,?,?,?)", (_pzn(r[0]), r[1], _pzn(r[2]), r[3], r[4], str(r[5] or ''), r[6], r[7]))
            for r in _read_rows(wb, 'DB_Lieferfaehigkeit'):
                con.execute("INSERT OR REPLACE INTO tbl_lieferfaehigkeit VALUES (?,?,?,?,?,?)", (_pzn(r[0]), r[1], r[2], r[3], r[4], str(r[5] or '')))
            for r in _read_rows(wb, 'DB_Lernhistorie'):
                con.execute("INSERT INTO tbl_lernhistorie VALUES (?,?,?,?,?,?,?,?)", tuple(str(v or '') for v in r[:8]))

        # Geprüfte Auswertungen als H-O-Lernstand und Herstellerkürzel-Lernquelle.
        for ref_path, quelle in [(LINDEN_REFERENCE_XLSX, 'Linden'), (ROSEN_REFERENCE_XLSX, 'Rosen'), (SONNEN_REFERENCE_XLSX, 'Sonnen')]:
            if not ref_path.exists():
                continue
            wb = load_workbook(ref_path, data_only=True, read_only=True)
            ws = wb.active
            for r in ws.iter_rows(min_row=2, max_col=15, values_only=True):
                pzn = _pzn(r[0])
                if not pzn:
                    continue
                register_basisdaten_simple(con, pzn, r[1] if len(r)>1 else None, r[4] if len(r)>4 else None, r[2] if len(r)>2 else None, r[3] if len(r)>3 else None, f"{quelle}_Referenz")
                if len(r) > 4 and r[4]:
                    register_hersteller(con, pzn, r[4], f"{quelle}_Referenz")
                h_to_o = r[7:15]
                if ref_path in (LINDEN_REFERENCE_XLSX, ROSEN_REFERENCE_XLSX):
                    con.execute("INSERT OR REPLACE INTO tbl_referenz_h_o VALUES (?,?,?,?,?,?,?,?,?)",
                        (pzn, h_to_o[0], _pzn(h_to_o[1]) if h_to_o[1] not in (None, '') else None,
                         _num(h_to_o[2]), _num(h_to_o[3]), h_to_o[4], h_to_o[5], h_to_o[6], h_to_o[7]))

        # Große Herstellerkürzel-Liste einlesen.
        if REPORT_HERSTELLER_CSV.exists():
            raw = REPORT_HERSTELLER_CSV.read_text(encoding='utf-8-sig', errors='replace')
            try:
                dialect = csv.Sniffer().sniff(raw[:4096], delimiters=';,\t,')
            except Exception:
                dialect = csv.excel; dialect.delimiter = ';'
            rows = list(csv.reader(raw.splitlines(), dialect))
            if rows:
                headers = [_norm_header(h) for h in rows[0]]
                pzn_i = next((i for i,h in enumerate(headers) if 'pzn' in h), 0)
                herst_i = next((i for i,h in enumerate(headers) if 'herst' in h or 'anbieter' in h), 1 if len(headers)>1 else 0)
                count = 0
                for row in rows[1:]:
                    if len(row) <= max(pzn_i, herst_i):
                        continue
                    if register_hersteller(con, row[pzn_i], row[herst_i], 'REPORT_HERSTELLER.csv'):
                        count += 1
                con.execute("INSERT INTO tbl_import_log(datei, typ, datensaetze, meldung) VALUES ('REPORT_HERSTELLER.csv','hersteller_lern',?,?)", (count, f'{count} Herstellerkürzel importiert'))


        con.commit()

        # Historische/Lernimporte werden in 1.9 über die mitgelieferte Startdatenbank
        # und über die Programm-Buttons ausgeführt, nicht während init_db mit offener Verbindung.
        con.commit()
