from pathlib import Path
from datetime import datetime
import sqlite3
import re
import csv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from .config import OUTPUT_DIR, DB_PATH
from .db import init_db
from .file_loader import load_worksheet, SUPPORTED_DATA_EXTENSIONS
from .learning_db import clean_hersteller, find_columns, register_hersteller, lookup_hersteller, parse_number, register_basisdaten, lookup_basisdaten, register_ek, lookup_latest_ek
from .abgleich_protocol import build_trace_index, trace_lookup_source, write_abgleichartikel_protocol

LINDEN_HEADERS = [
    "PZN", "Artikelname", "DF", "Pck", "Herst", "EK",
    "Abverkäufe 6 Monate", "im Sortiment", "PZN NMG", "APU NMG",
    "NMG Rabatt", "lieferbar", "Bevorratung angeraten", "Liefervor- schlag",
    "austauschbar gegen", "NMG Rabatt in Euro", "NMG Rabatt Gesamt nach Absatz", "Umsatz"
]


class UnknownInputFormatError(Exception):
    """Wird ausgelöst, wenn eine Rohdatei nicht sicher erkannt wurde."""
    def __init__(self, message: str, report_path: Path | None = None):
        super().__init__(message)
        self.report_path = report_path


def _col_letter(idx):
    try:
        return get_column_letter(idx) if idx else ""
    except Exception:
        return ""


def _diagnose_format(input_file: Path, ws, mapping):
    """Erzeugt einen Diagnosebericht, wenn die Datei nicht sicher auswertbar ist."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    safe = ''.join(ch if ch.isalnum() or ch in '-_' else '_' for ch in input_file.stem)[:60]
    report = OUTPUT_DIR / f"Diagnose_Format_nicht_erkannt_{safe}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    wb = Workbook()
    sh = wb.active
    sh.title = "Diagnose"
    sh.append(["Datei", input_file.name])
    sh.append(["Status", "Format nicht sicher erkannt – Datei wurde nicht ausgewertet."])
    sh.append([])

    mapping = mapping or {}
    checks = [
        ("PZN", mapping.get("pzn"), True),
        ("Artikelname", mapping.get("artikel"), False),
        ("Hersteller/Herst", mapping.get("hersteller"), False),
        ("Absatz/Verbrauch", mapping.get("absatz"), True),
        ("EK", mapping.get("ek"), False),
    ]
    sh.append(["Pflicht-/Zielfeld", "Gefunden", "Spalte", "Hinweis"])
    for name, col, required in checks:
        sh.append([name, "Ja" if col else "Nein", _col_letter(col), "Pflichtfeld" if required else "optional"])
    sh.append([])
    sh.append(["Gefundene Überschriften / erste Zeilen"])

    # Zeilen 1-15 als Diagnose ausgeben.
    max_col = min(ws.max_column, 40)
    for r in range(1, min(ws.max_row, 15) + 1):
        values = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        sh.append([f"Zeile {r}"] + values)

    sh2 = wb.create_sheet("Manuelle Zuordnung")
    sh2.append(["Feld", "Trage hier Spaltenbuchstabe ein", "Beispiel", "Pflicht"])
    rows = [
        ("PZN", "", "A", "Ja"),
        ("Artikelname", "", "B oder N", "Nein, aber empfohlen"),
        ("Hersteller", "", "E oder O", "Nein"),
        ("EK", "", "F / Apo-EK / Taxe-EK", "Nein"),
        ("Absatz 6 Monate", "", "G oder P", "Ja"),
        ("DF", "", "C", "Nein"),
        ("Pck", "", "D", "Nein"),
    ]
    for row in rows:
        sh2.append(row)
    sh2.append([])
    sh2.append(["Hinweis", "Diese Version erzeugt die Vorlage. Die automatische Nutzung der manuellen Zuordnung kommt im nächsten Schritt."])

    for sheet in wb.worksheets:
        for col in range(1, min(sheet.max_column, 12) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 22
    wb.save(report)
    return report


def _is_mapping_usable(mapping):
    if not mapping:
        return False
    # Für eine sichere Auswertung müssen PZN und Absatz vorhanden sein.
    # EK ist optional, Hersteller und Artikel können aus DB/leer ergänzt werden.
    return bool(mapping.get("pzn") and mapping.get("absatz"))


def _pzn(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith('.0'):
        text = text[:-2]
    return text.zfill(8) if text.isdigit() else text


def _to_number(value, default=0):
    if value in (None, ""):
        return default
    if isinstance(value, (int, float)):
        return value
    try:
        return float(str(value).strip().replace(".", "").replace(",", "."))
    except Exception:
        return default


def _rowdict(row):
    return dict(row) if row else None


def _lookup(con: sqlite3.Connection, original_pzn: str):
    """Liest die gelernte Zuordnung für die Neue Auswertung.

    Wichtig für die Schulbank-/Editor-Logik:
    Die zentrale Wissensquelle ist tbl_austauschdatenbank. Sie muss vor alten
    Referenz-/Importtabellen gewinnen, sonst erscheinen bereits gelernte Fälle
    nach einer neuen Auswertung wieder als Abweichung.
    """
    original_pzn = _pzn(original_pzn)
    if not original_pzn:
        return {}

    # SP24: pzn_norm() als SQLite-Funktion verfuegbar machen, damit JOINs und
    # WHERE-Klauseln Format-Unterschiede tolerieren (Excel-Int, fuehrende Nullen,
    # ".0"-Suffix). Idempotent - mehrfaches Aufrufen schadet nicht.
    try:
        con.create_function("pzn_norm", 1, _pzn)
    except Exception:
        pass

    def _norm_db_pzn(value):
        return _pzn(value)

    def _article_name_by_pzn(pzn):
        pzn = _pzn(pzn)
        if not pzn:
            return ""
        for table, col in (
            ("tbl_nmg_stamm", "artikelname"),
            ("tbl_artikelstamm", "artikel"),
            ("tbl_artikelstamm", "artikelname"),
            ("tbl_pzn_basisdaten", "artikelname"),
        ):
            try:
                row = con.execute(f"SELECT {col} FROM {table} WHERE pzn_norm(pzn)=pzn_norm(?) LIMIT 1", (pzn,)).fetchone()
                if row and row[0]:
                    return str(row[0]).strip()
            except Exception:
                continue
        return ""

    # 1) Höchste Priorität: neue Austauschdatenbank / Schulbank-Lernstand.
    # SP20: Mehrere aktive Treffer = mehrere Austausch-Moeglichkeiten; erster Treffer
    # wird primaer, alle weiteren werden im Freitext "austauschbar gegen" zusaetzlich gelistet.
    # SP21/SP23: Sortier-Logik fuer den Haupttreffer:
    # 1) Kandidaten MIT einem echten Rabatt-Eintrag (nmg_rabatte.rabatt IS NOT NULL)
    #    haben Vorrang gegenueber Kandidaten ohne Eintrag. Damit landet ein Artikel
    #    ohne Rabatt-Eintrag nicht mehr in der NMG-PZN-Spalte, solange es einen mit
    #    Rabatt gibt.
    # 2) Innerhalb der "mit Rabatt"-Gruppe gewinnt der hoechste Rabatt.
    # 3) Als zusaetzliches Sicherheitsnetz: Kandidaten, die im NMG-Stamm existieren,
    #    werden bevorzugt (sonst veraltete Schulbank-Eintraege gewinnen).
    # 4) Tie-Breaker: aktualisiert_am DESC, id DESC.
    austausch = None
    weitere = []
    try:
        rows = con.execute("""
            SELECT a.*,
                   r.rabatt AS _rabatt_pk,
                   ns.pzn   AS _nmg_stamm_pzn
            FROM tbl_austauschdatenbank a
            LEFT JOIN nmg_rabatte   r  ON pzn_norm(r.nmg_pzn) = pzn_norm(a.pzn_nmg)
            LEFT JOIN tbl_nmg_stamm ns ON pzn_norm(ns.pzn)    = pzn_norm(a.pzn_nmg)
            WHERE COALESCE(a.status, 'aktiv') = 'aktiv'
              AND COALESCE(a.pzn_alt, '') <> ''
            ORDER BY
                CASE WHEN r.rabatt IS NOT NULL THEN 0 ELSE 1 END,
                COALESCE(r.rabatt, 0) DESC,
                CASE WHEN ns.pzn IS NOT NULL THEN 0 ELSE 1 END,
                datetime(COALESCE(a.aktualisiert_am, a.erstellt_am, a.gueltig_ab, '1970-01-01')) DESC,
                a.id DESC
        """).fetchall()
        matches = []
        for candidate in rows:
            data = _rowdict(candidate)
            if _norm_db_pzn(data.get("pzn_alt")) == original_pzn:
                matches.append(data)

        # SP20: Mehrere aktive Treffer werden nicht mehr unterdrueckt. Der neueste
        # Treffer wird als Haupt-Austausch verwendet (NMG-PZN, Rabatt, Lieferbarkeit),
        # alle weiteren werden im Freitext "austauschbar gegen" zusaetzlich aufgefuehrt.
        if matches:
            austausch = matches[0]
            weitere = matches[1:]
    except Exception:
        austausch = None
        weitere = []

    if austausch:
        nmg_pzn = _pzn(austausch.get("pzn_nmg"))
        freitext = str(austausch.get("freitext_austausch") or "").strip()
        artikel_nmg = str(austausch.get("artikel_nmg") or "").strip()
        if nmg_pzn and (not freitext or freitext.lower().startswith("pzn nmg:")):
            freitext = artikel_nmg or _article_name_by_pzn(nmg_pzn) or freitext

        # SP20: Weitere aktive Austausch-Optionen anhaengen.
        # SP22: In der Spalte "austauschbar gegen" duerfen nur Artikel auftauchen,
        # die tatsaechlich im NMG-Stamm vorhanden sind. Freitext-Eintraege ohne
        # verifizierbare NMG-PZN werden ausgeblendet, ebenso PZNs, die der NMG-
        # Stamm nicht kennt (veraltet oder Tippfehler).
        def _pzn_in_nmg_stamm(pzn):
            if not pzn:
                return False
            try:
                return con.execute(
                    "SELECT 1 FROM tbl_nmg_stamm WHERE pzn_norm(pzn)=pzn_norm(?) LIMIT 1", (pzn,)
                ).fetchone() is not None
            except Exception:
                return False

        weitere_texte = []
        seen_pzns = {nmg_pzn} if nmg_pzn else set()
        for extra in weitere or []:
            extra_pzn = _pzn(extra.get("pzn_nmg"))
            if not extra_pzn or extra_pzn in seen_pzns:
                continue
            if not _pzn_in_nmg_stamm(extra_pzn):
                continue
            seen_pzns.add(extra_pzn)
            extra_name = str(extra.get("artikel_nmg") or "").strip() or _article_name_by_pzn(extra_pzn)
            weitere_texte.append(f"PZN {extra_pzn}" + (f" – {extra_name}" if extra_name else ""))

        if weitere_texte:
            zusatz = "weitere: " + " | ".join(weitere_texte)
            freitext = f"{freitext} | {zusatz}" if freitext else zusatz

        stamm = _rowdict(con.execute("SELECT * FROM tbl_nmg_stamm WHERE pzn_norm(pzn)=pzn_norm(?)", (nmg_pzn or '',)).fetchone()) if nmg_pzn else None
        rabatt = _rowdict(con.execute("SELECT * FROM nmg_rabatte WHERE pzn_norm(nmg_pzn)=pzn_norm(?)", (nmg_pzn or '',)).fetchone()) if nmg_pzn else None
        liefer = _rowdict(con.execute("SELECT * FROM tbl_lieferfaehigkeit WHERE pzn_norm(nmg_pzn)=pzn_norm(?)", (nmg_pzn or '',)).fetchone()) if nmg_pzn else None

        return {
            "im_sortiment": "X" if nmg_pzn else "X Austausch mögl",
            "nmg_pzn": nmg_pzn or None,
            "apu_nmg": stamm.get("apu") if stamm and stamm.get("apu") is not None else None,
            "rabatt": rabatt.get("rabatt") if rabatt else None,
            "lieferbar": liefer.get("lieferbar") if liefer else None,
            "bevorratung": liefer.get("bevorratung_angeraten") if liefer else None,
            "liefervorschlag": liefer.get("liefervorschlag") if liefer else None,
            "austauschbar_gegen": freitext or None,
        }

    # 2) Sicherheitsnetz: bereits übernommene Lernvorschläge, falls ältere Datenbanken
    # noch nicht vollständig in tbl_austauschdatenbank gespiegelt wurden.
    try:
        rows = con.execute("""
            SELECT *
            FROM tbl_lernvorschlaege
            WHERE status = 'uebernommen'
              AND COALESCE(pzn_alt, '') <> ''
            ORDER BY datetime(COALESCE(bearbeitet_am, erstellt_am, '1970-01-01')) DESC, id DESC
        """).fetchall()
        for candidate in rows:
            data = _rowdict(candidate)
            if _norm_db_pzn(data.get("pzn_alt")) != original_pzn:
                continue
            nmg_pzn = _pzn(data.get("pzn_nmg"))
            freitext = str(data.get("freitext_austausch") or data.get("produkt_neu") or "").strip()
            if nmg_pzn and (not freitext or freitext.lower().startswith("pzn nmg:")):
                freitext = _article_name_by_pzn(nmg_pzn) or freitext
            stamm = _rowdict(con.execute("SELECT * FROM tbl_nmg_stamm WHERE pzn_norm(pzn)=pzn_norm(?)", (nmg_pzn or '',)).fetchone()) if nmg_pzn else None
            rabatt = _rowdict(con.execute("SELECT * FROM nmg_rabatte WHERE pzn_norm(nmg_pzn)=pzn_norm(?)", (nmg_pzn or '',)).fetchone()) if nmg_pzn else None
            liefer = _rowdict(con.execute("SELECT * FROM tbl_lieferfaehigkeit WHERE pzn_norm(nmg_pzn)=pzn_norm(?)", (nmg_pzn or '',)).fetchone()) if nmg_pzn else None
            return {
                "im_sortiment": "X" if nmg_pzn else "X Austausch mögl",
                "nmg_pzn": nmg_pzn or None,
                "apu_nmg": stamm.get("apu") if stamm and stamm.get("apu") is not None else None,
                "rabatt": rabatt.get("rabatt") if rabatt else None,
                "lieferbar": liefer.get("lieferbar") if liefer else None,
                "bevorratung": liefer.get("bevorratung_angeraten") if liefer else None,
                "liefervorschlag": liefer.get("liefervorschlag") if liefer else None,
                "austauschbar_gegen": freitext or None,
            }
    except Exception:
        pass

    # 3) Danach erst alte geprüfte Referenz, damit neue Schulbank-Korrekturen nicht
    # durch veraltete Referenzwerte überschrieben werden.
    ref = _rowdict(con.execute(
        "SELECT * FROM tbl_referenz_h_o WHERE original_pzn=?", (original_pzn,)
    ).fetchone())
    if ref:
        nmg_pzn_ref = _pzn(ref.get("nmg_pzn"))
        stamm_ref = _rowdict(con.execute("SELECT * FROM tbl_nmg_stamm WHERE pzn_norm(pzn)=pzn_norm(?)", (nmg_pzn_ref or '',)).fetchone())
        return {
            "im_sortiment": ref.get("im_sortiment"),
            "nmg_pzn": nmg_pzn_ref,
            "apu_nmg": stamm_ref.get("apu") if stamm_ref and stamm_ref.get("apu") is not None else ref.get("apu_nmg"),
            "rabatt": ref.get("rabatt"),
            "lieferbar": ref.get("lieferbar"),
            "bevorratung": ref.get("bevorratung_angeraten"),
            "liefervorschlag": ref.get("liefervorschlag"),
            "austauschbar_gegen": ref.get("austauschbar_gegen"),
        }

    # 4) Falls die abgegebene PZN selbst NMG-Artikel ist oder Rabatt-/Lieferdaten hat.
    nmg_pzn = original_pzn
    stamm = _rowdict(con.execute("SELECT * FROM tbl_nmg_stamm WHERE pzn_norm(pzn)=pzn_norm(?)", (nmg_pzn or '',)).fetchone())
    rabatt = _rowdict(con.execute("SELECT * FROM nmg_rabatte WHERE pzn_norm(nmg_pzn)=pzn_norm(?)", (nmg_pzn or '',)).fetchone())
    liefer = _rowdict(con.execute("SELECT * FROM tbl_lieferfaehigkeit WHERE pzn_norm(nmg_pzn)=pzn_norm(?)", (nmg_pzn or '',)).fetchone())

    if not any([stamm, rabatt, liefer]):
        return {}

    return {
        "im_sortiment": "X",
        "nmg_pzn": nmg_pzn,
        "apu_nmg": stamm.get("apu") if stamm and stamm.get("apu") is not None else None,
        "rabatt": rabatt.get("rabatt") if rabatt else None,
        "lieferbar": liefer.get("lieferbar") if liefer else None,
        "bevorratung": liefer.get("bevorratung_angeraten") if liefer else None,
        "liefervorschlag": liefer.get("liefervorschlag") if liefer else None,
        "austauschbar_gegen": None,
    }

def _lookup_hersteller(con: sqlite3.Connection, pzn: str) -> str:
    return lookup_hersteller(con, pzn)


def _lookup_stammdaten(con: sqlite3.Connection, pzn: str) -> dict:
    """SP19: Fallback fuer Artikelname/DF/Pck/Herst/EK zur abgegebenen PZN.
    Wird benutzt, wenn weder die Rohdatei noch tbl_pzn_basisdaten (selbstlernend)
    Werte liefern. Liest aus tbl_artikelstamm, tbl_pzn_basisdaten, tbl_nmg_stamm.
    """
    pzn = _pzn(pzn)
    result = {"artikelname": None, "df": None, "pck": None, "hersteller": None, "ek": None}
    if not pzn:
        return result

    # SP24: pzn_norm() bereitstellen fuer Format-Toleranz beim Lookup.
    try:
        con.create_function("pzn_norm", 1, _pzn)
    except Exception:
        pass

    def _set(key, value):
        if result[key] in (None, "") and value not in (None, ""):
            result[key] = value

    try:
        row = con.execute(
            "SELECT artikel, df, pck, herst, ek FROM tbl_artikelstamm WHERE pzn_norm(pzn)=pzn_norm(?) LIMIT 1",
            (pzn,),
        ).fetchone()
        if row:
            data = dict(row) if hasattr(row, "keys") else {"artikel": row[0], "df": row[1], "pck": row[2], "herst": row[3], "ek": row[4]}
            _set("artikelname", data.get("artikel"))
            _set("df", data.get("df"))
            _set("pck", data.get("pck"))
            _set("hersteller", data.get("herst"))
            _set("ek", data.get("ek"))
    except Exception:
        pass

    try:
        row = con.execute(
            "SELECT artikelname, df, pck, herstellerkuerzel FROM tbl_pzn_basisdaten WHERE pzn_norm(pzn)=pzn_norm(?) LIMIT 1",
            (pzn,),
        ).fetchone()
        if row:
            data = dict(row) if hasattr(row, "keys") else {"artikelname": row[0], "df": row[1], "pck": row[2], "herstellerkuerzel": row[3]}
            _set("artikelname", data.get("artikelname"))
            _set("df", data.get("df"))
            _set("pck", data.get("pck"))
            _set("hersteller", data.get("herstellerkuerzel"))
    except Exception:
        pass

    try:
        row = con.execute(
            "SELECT artikelname, herstellerkuerzel, taxe_ek FROM tbl_nmg_stamm WHERE pzn_norm(pzn)=pzn_norm(?) LIMIT 1",
            (pzn,),
        ).fetchone()
        if row:
            data = dict(row) if hasattr(row, "keys") else {"artikelname": row[0], "herstellerkuerzel": row[1], "taxe_ek": row[2]}
            _set("artikelname", data.get("artikelname"))
            _set("hersteller", data.get("herstellerkuerzel"))
            _set("ek", data.get("taxe_ek"))
    except Exception:
        pass

    return result


# ---------------------------------------------------------------------------
# SP25: In-Memory-Caches fuer die Auswertungs-Schleife.
#
# Vor SP25 hat _lookup(), _lookup_stammdaten(), lookup_hersteller(),
# lookup_basisdaten() und lookup_latest_ek() pro Rohdaten-Zeile zwischen 5 und
# 10 SQL-Queries abgesetzt. Mit SP24's pzn_norm()-Funktion auf beiden Seiten
# der JOIN-Klauseln konnten die Indizes nicht mehr genutzt werden, jede Query
# wurde zum full table scan. Bei 8000 Zeilen × 8 Queries × N k Tabellen-Rows
# kostete eine Auswertung viele Minuten.
#
# SP25 laedt die Lookup-Tabellen genau einmal pro Auswertung als Python-Dicts
# (PZN bereits auf 8 Stellen normalisiert). _lookup_fast() und Co. arbeiten
# danach mit O(1) Dict-Zugriffen. Schreiboperationen (register_basisdaten,
# register_ek) bleiben SQL - die werden in einer kuenftigen SP gebatcht.
# ---------------------------------------------------------------------------

def _load_lookup_caches(con: sqlite3.Connection):
    caches = {
        "rabatte": {},
        "nmg_stamm": {},
        "lieferfaehigkeit": {},
        "artikelstamm": {},
        "pzn_basisdaten": {},
        "hersteller_lern": {},
        "ek_rohdaten_latest": {},
        "austausch": {},
        "lernvorschlaege": {},
        "referenz_h_o": {},
    }

    def _safe_exec(sql):
        try:
            return con.execute(sql).fetchall()
        except Exception:
            return []

    for row in _safe_exec("SELECT nmg_pzn, rabatt FROM nmg_rabatte"):
        pzn = _pzn(row[0])
        if pzn and row[1] is not None:
            caches["rabatte"][pzn] = row[1]

    for row in _safe_exec("SELECT pzn, artikelname, apu, herstellerkuerzel, taxe_ek FROM tbl_nmg_stamm"):
        pzn = _pzn(row[0])
        if pzn:
            caches["nmg_stamm"][pzn] = {
                "artikelname": row[1], "apu": row[2],
                "herstellerkuerzel": row[3], "taxe_ek": row[4],
            }

    for row in _safe_exec("SELECT nmg_pzn, lieferbar, bevorratung_angeraten, liefervorschlag FROM tbl_lieferfaehigkeit"):
        pzn = _pzn(row[0])
        if pzn:
            caches["lieferfaehigkeit"][pzn] = {
                "lieferbar": row[1], "bevorratung_angeraten": row[2], "liefervorschlag": row[3],
            }

    for row in _safe_exec("SELECT pzn, artikel, df, pck, herst, ek FROM tbl_artikelstamm"):
        pzn = _pzn(row[0])
        if pzn:
            caches["artikelstamm"][pzn] = {
                "artikel": row[1], "df": row[2], "pck": row[3],
                "herst": row[4], "ek": row[5],
            }

    for row in _safe_exec("SELECT pzn, artikelname, df, pck, herstellerkuerzel FROM tbl_pzn_basisdaten"):
        pzn = _pzn(row[0])
        if pzn:
            caches["pzn_basisdaten"][pzn] = {
                "artikelname": row[1], "df": row[2], "pck": row[3],
                "herstellerkuerzel": row[4],
            }

    for row in _safe_exec("SELECT pzn, herstellerkuerzel FROM tbl_hersteller_lern"):
        pzn = _pzn(row[0])
        if pzn and row[1]:
            caches["hersteller_lern"][pzn] = row[1]

    # tbl_pzn_ek_rohdaten: pro PZN den juengsten Eintrag (ORDER BY importdatum DESC, id DESC).
    for row in _safe_exec("""
        SELECT pzn, ek
        FROM tbl_pzn_ek_rohdaten
        ORDER BY importdatum DESC, id DESC
    """):
        pzn = _pzn(row[0])
        if pzn and pzn not in caches["ek_rohdaten_latest"] and row[1] is not None:
            caches["ek_rohdaten_latest"][pzn] = row[1]

    # Austausch: aktive Eintraege, sortiert mit SP23-Priorisierung.
    austausch_rows = _safe_exec("""
        SELECT id, pzn_alt, pzn_nmg, artikel_alt, artikel_nmg,
               freitext_austausch, aktualisiert_am, erstellt_am, gueltig_ab
        FROM tbl_austauschdatenbank
        WHERE COALESCE(status, 'aktiv') = 'aktiv'
          AND COALESCE(pzn_alt, '') <> ''
    """)

    def _dt_str(*candidates):
        for c in candidates:
            if c:
                return str(c)
        return "1970-01-01"

    def _sort_key(r):
        pzn_nmg = _pzn(r[2])
        rabatt = caches["rabatte"].get(pzn_nmg)
        in_stamm = pzn_nmg in caches["nmg_stamm"]
        return (
            0 if rabatt is not None else 1,
            -(rabatt if rabatt is not None else 0),
            0 if in_stamm else 1,
            _dt_str(r[6], r[7], r[8]),
            r[0] or 0,
        )

    # Reverse sortieren fuer DESC bei datetime + id, ASC fuer die Flags - trick: sortieren
    # mit (flag_asc, -rabatt, flag_asc, -datetime_lex, -id). Aber Datums-String DESC = "ZZZZ" zuerst.
    # Einfacher: zwei Pass-Sortierung.
    decorated = []
    for r in austausch_rows:
        pzn_nmg = _pzn(r[2])
        rabatt = caches["rabatte"].get(pzn_nmg)
        in_stamm = pzn_nmg in caches["nmg_stamm"]
        decorated.append((
            0 if rabatt is not None else 1,
            -(rabatt if rabatt is not None else 0),
            0 if in_stamm else 1,
            r,
        ))
    # Stabil sortieren, danach in einer zweiten Sortierung nach Datum DESC.
    austausch_rows_sorted = [d[3] for d in sorted(decorated, key=lambda d: (d[0], d[1], d[2]))]
    # Innerhalb gleicher Schluessel ist die ursprueliche SQL-Reihenfolge erhalten - reicht.
    # Wir holen aber lieber nach datetime DESC sortieren in einer Voraus-Sortierung der SQL:
    # Geht einfacher per SECOND-Sort - aber Tie-Breaker (Datum DESC) ist meist nicht entscheidend.
    # Pragmatisch: nach Datum DESC vorsortieren, dann obigen stabilen sort.
    pre_sorted = sorted(austausch_rows, key=lambda r: (_dt_str(r[6], r[7], r[8]), r[0] or 0), reverse=True)
    decorated = []
    for r in pre_sorted:
        pzn_nmg = _pzn(r[2])
        rabatt = caches["rabatte"].get(pzn_nmg)
        in_stamm = pzn_nmg in caches["nmg_stamm"]
        decorated.append((
            0 if rabatt is not None else 1,
            -(rabatt if rabatt is not None else 0),
            0 if in_stamm else 1,
            r,
        ))
    austausch_rows_sorted = [d[3] for d in sorted(decorated, key=lambda d: (d[0], d[1], d[2]))]

    for r in austausch_rows_sorted:
        key = _pzn(r[1])
        if not key:
            continue
        caches["austausch"].setdefault(key, []).append({
            "id": r[0], "pzn_alt": r[1], "pzn_nmg": r[2],
            "artikel_alt": r[3], "artikel_nmg": r[4],
            "freitext_austausch": r[5],
            "aktualisiert_am": r[6], "erstellt_am": r[7], "gueltig_ab": r[8],
        })

    for row in _safe_exec("""
        SELECT id, pzn_alt, pzn_nmg, freitext_austausch, produkt_neu, bearbeitet_am, erstellt_am
        FROM tbl_lernvorschlaege
        WHERE status = 'uebernommen'
          AND COALESCE(pzn_alt, '') <> ''
        ORDER BY datetime(COALESCE(bearbeitet_am, erstellt_am, '1970-01-01')) DESC, id DESC
    """):
        key = _pzn(row[1])
        if not key:
            continue
        caches["lernvorschlaege"].setdefault(key, []).append({
            "id": row[0], "pzn_alt": row[1], "pzn_nmg": row[2],
            "freitext_austausch": row[3], "produkt_neu": row[4],
            "bearbeitet_am": row[5], "erstellt_am": row[6],
        })

    for row in _safe_exec("SELECT * FROM tbl_referenz_h_o"):
        try:
            data = dict(row) if hasattr(row, "keys") else None
        except Exception:
            data = None
        if not data:
            continue
        key = _pzn(data.get("original_pzn"))
        if key:
            caches["referenz_h_o"][key] = data

    return caches


def _lookup_fast(caches, original_pzn):
    """SP25 In-Memory-Variante von _lookup. Identisches Verhalten, aber alle
    Tabellen-Zugriffe gehen ueber caches statt SQL.
    """
    original_pzn = _pzn(original_pzn)
    if not original_pzn:
        return {}

    def _article_name_by_pzn(pzn):
        if not pzn:
            return ""
        for store_key in ("nmg_stamm", "artikelstamm", "pzn_basisdaten"):
            entry = caches[store_key].get(pzn)
            if entry:
                name = entry.get("artikelname") or entry.get("artikel")
                if name:
                    return str(name).strip()
        return ""

    # 1) Austauschdatenbank
    matches = caches["austausch"].get(original_pzn, [])
    if matches:
        austausch = matches[0]
        weitere = matches[1:]

        nmg_pzn = _pzn(austausch.get("pzn_nmg"))
        freitext = str(austausch.get("freitext_austausch") or "").strip()
        artikel_nmg = str(austausch.get("artikel_nmg") or "").strip()
        if nmg_pzn and (not freitext or freitext.lower().startswith("pzn nmg:")):
            freitext = artikel_nmg or _article_name_by_pzn(nmg_pzn) or freitext

        weitere_texte = []
        seen_pzns = {nmg_pzn} if nmg_pzn else set()
        for extra in weitere:
            extra_pzn = _pzn(extra.get("pzn_nmg"))
            if not extra_pzn or extra_pzn in seen_pzns:
                continue
            if extra_pzn not in caches["nmg_stamm"]:
                continue
            seen_pzns.add(extra_pzn)
            extra_name = str(extra.get("artikel_nmg") or "").strip() or _article_name_by_pzn(extra_pzn)
            weitere_texte.append(f"PZN {extra_pzn}" + (f" – {extra_name}" if extra_name else ""))

        if weitere_texte:
            zusatz = "weitere: " + " | ".join(weitere_texte)
            freitext = f"{freitext} | {zusatz}" if freitext else zusatz

        stamm = caches["nmg_stamm"].get(nmg_pzn)
        rabatt = caches["rabatte"].get(nmg_pzn)
        liefer = caches["lieferfaehigkeit"].get(nmg_pzn)
        return {
            "im_sortiment": "X" if nmg_pzn else "X Austausch mögl",
            "nmg_pzn": nmg_pzn or None,
            "apu_nmg": stamm.get("apu") if stamm and stamm.get("apu") is not None else None,
            "rabatt": rabatt,
            "lieferbar": liefer.get("lieferbar") if liefer else None,
            "bevorratung": liefer.get("bevorratung_angeraten") if liefer else None,
            "liefervorschlag": liefer.get("liefervorschlag") if liefer else None,
            "austauschbar_gegen": freitext or None,
        }

    # 2) Lernvorschlaege (uebernommen)
    for data in caches["lernvorschlaege"].get(original_pzn, []):
        nmg_pzn = _pzn(data.get("pzn_nmg"))
        freitext = str(data.get("freitext_austausch") or data.get("produkt_neu") or "").strip()
        if nmg_pzn and (not freitext or freitext.lower().startswith("pzn nmg:")):
            freitext = _article_name_by_pzn(nmg_pzn) or freitext
        stamm = caches["nmg_stamm"].get(nmg_pzn)
        rabatt = caches["rabatte"].get(nmg_pzn)
        liefer = caches["lieferfaehigkeit"].get(nmg_pzn)
        return {
            "im_sortiment": "X" if nmg_pzn else "X Austausch mögl",
            "nmg_pzn": nmg_pzn or None,
            "apu_nmg": stamm.get("apu") if stamm and stamm.get("apu") is not None else None,
            "rabatt": rabatt,
            "lieferbar": liefer.get("lieferbar") if liefer else None,
            "bevorratung": liefer.get("bevorratung_angeraten") if liefer else None,
            "liefervorschlag": liefer.get("liefervorschlag") if liefer else None,
            "austauschbar_gegen": freitext or None,
        }

    # 3) tbl_referenz_h_o
    ref = caches["referenz_h_o"].get(original_pzn)
    if ref:
        nmg_pzn_ref = _pzn(ref.get("nmg_pzn"))
        stamm_ref = caches["nmg_stamm"].get(nmg_pzn_ref)
        return {
            "im_sortiment": ref.get("im_sortiment"),
            "nmg_pzn": nmg_pzn_ref,
            "apu_nmg": stamm_ref.get("apu") if stamm_ref and stamm_ref.get("apu") is not None else ref.get("apu_nmg"),
            "rabatt": ref.get("rabatt"),
            "lieferbar": ref.get("lieferbar"),
            "bevorratung": ref.get("bevorratung_angeraten"),
            "liefervorschlag": ref.get("liefervorschlag"),
            "austauschbar_gegen": ref.get("austauschbar_gegen"),
        }

    # 4) Original-PZN selbst NMG?
    nmg_pzn = original_pzn
    stamm = caches["nmg_stamm"].get(nmg_pzn)
    rabatt = caches["rabatte"].get(nmg_pzn)
    liefer = caches["lieferfaehigkeit"].get(nmg_pzn)
    if not any([stamm, rabatt is not None, liefer]):
        return {}
    return {
        "im_sortiment": "X",
        "nmg_pzn": nmg_pzn,
        "apu_nmg": stamm.get("apu") if stamm and stamm.get("apu") is not None else None,
        "rabatt": rabatt,
        "lieferbar": liefer.get("lieferbar") if liefer else None,
        "bevorratung": liefer.get("bevorratung_angeraten") if liefer else None,
        "liefervorschlag": liefer.get("liefervorschlag") if liefer else None,
        "austauschbar_gegen": None,
    }


def _lookup_stammdaten_fast(caches, pzn):
    pzn = _pzn(pzn)
    result = {"artikelname": None, "df": None, "pck": None, "hersteller": None, "ek": None}
    if not pzn:
        return result
    art = caches["artikelstamm"].get(pzn)
    if art:
        if result["artikelname"] in (None, ""): result["artikelname"] = art.get("artikel")
        if result["df"] in (None, ""): result["df"] = art.get("df")
        if result["pck"] in (None, ""): result["pck"] = art.get("pck")
        if result["hersteller"] in (None, ""): result["hersteller"] = art.get("herst")
        if result["ek"] in (None, "") and art.get("ek") is not None: result["ek"] = art.get("ek")
    basis = caches["pzn_basisdaten"].get(pzn)
    if basis:
        if result["artikelname"] in (None, ""): result["artikelname"] = basis.get("artikelname")
        if result["df"] in (None, ""): result["df"] = basis.get("df")
        if result["pck"] in (None, ""): result["pck"] = basis.get("pck")
        if result["hersteller"] in (None, ""): result["hersteller"] = basis.get("herstellerkuerzel")
    stamm = caches["nmg_stamm"].get(pzn)
    if stamm:
        if result["artikelname"] in (None, ""): result["artikelname"] = stamm.get("artikelname")
        if result["hersteller"] in (None, ""): result["hersteller"] = stamm.get("herstellerkuerzel")
        if result["ek"] in (None, "") and stamm.get("taxe_ek") is not None: result["ek"] = stamm.get("taxe_ek")
    return result


def _lookup_basisdaten_fast(caches, pzn):
    pzn = _pzn(pzn)
    return caches["pzn_basisdaten"].get(pzn, {}) if pzn else {}


def _lookup_hersteller_fast(caches, pzn):
    pzn = _pzn(pzn)
    return caches["hersteller_lern"].get(pzn, "") if pzn else ""


def _lookup_latest_ek_fast(caches, pzn):
    pzn = _pzn(pzn)
    return caches["ek_rohdaten_latest"].get(pzn) if pzn else None


def _strip_leading_manufacturer_number(value):
    """Entfernt EKM-artige numerische Anbieter-Codes vor dem Herstellertext.
    Beispiel: "17760 Aristo Pha" -> "Aristo Pha".
    Alphanumerische Hersteller wie "1A Pharma" bleiben erhalten, weil nach der 1 kein Leerzeichen folgt.
    """
    if value is None:
        return None
    text = str(value).strip()
    m = re.match(r"^\d+\s+(.+)$", text)
    return m.group(1).strip() if m else value



class _CSVCell:
    def __init__(self, value):
        self.value = value

class _CSVWorksheet:
    """Kleine openpyxl-kompatible Hülle für CSV-Dateien.
    Sie stellt genau die Eigenschaften bereit, die find_columns(), Diagnose und
    die Export-Routine benötigen: max_row, max_column, ws[row], cell() und iter_rows().
    """
    def __init__(self, rows):
        self._rows = rows
        self.max_row = len(rows)
        self.max_column = max((len(r) for r in rows), default=0)

    def __getitem__(self, row_idx):
        row = self._rows[row_idx - 1] if 1 <= row_idx <= self.max_row else []
        padded = list(row) + [None] * max(0, self.max_column - len(row))
        return [_CSVCell(v) for v in padded]

    def cell(self, row, column):
        try:
            return _CSVCell(self._rows[row - 1][column - 1])
        except Exception:
            return _CSVCell(None)

    def iter_rows(self, min_row=1, values_only=False, **kwargs):
        max_row = kwargs.get('max_row') or self.max_row
        min_col = kwargs.get('min_col') or 1
        max_col = kwargs.get('max_col') or self.max_column
        for r in range(min_row, max_row + 1):
            source = self._rows[r - 1] if 1 <= r <= self.max_row else []
            vals = []
            for c in range(min_col, max_col + 1):
                vals.append(source[c - 1] if c - 1 < len(source) else None)
            if values_only:
                yield tuple(vals)
            else:
                yield tuple(_CSVCell(v) for v in vals)


def _read_csv_rows(path: Path):
    """Liest CSV robust mit typischen deutschen Export-Einstellungen.
    Unterstützt Semikolon, Komma und Tab sowie UTF-8/UTF-8-SIG/CP1252.
    """
    encodings = ['utf-8-sig', 'utf-8', 'cp1252', 'latin1']
    last_error = None
    for enc in encodings:
        try:
            text = path.read_text(encoding=enc)
            sample = text[:4096]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=';,	|')
                delimiter = dialect.delimiter
            except Exception:
                delimiter = ';' if sample.count(';') >= sample.count(',') else ','
            rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
            return [[cell.strip() if isinstance(cell, str) else cell for cell in row] for row in rows]
        except Exception as exc:
            last_error = exc
    raise ValueError(f'CSV-Datei konnte nicht gelesen werden: {last_error}')


def _load_input_worksheet(input_file: Path):
    # Erlaubte Formate kommen aus file_loader (eine Wahrheit), damit neue
    # Formate wie .xls nicht erneut hier vergessen werden.
    suffix = input_file.suffix.lower()
    if suffix in SUPPORTED_DATA_EXTENSIONS:
        return load_worksheet(input_file)
    erlaubt = ", ".join(sorted(SUPPORTED_DATA_EXTENSIONS))
    raise UnknownInputFormatError(
        f'Nicht unterstütztes Dateiformat. Erlaubt sind: {erlaubt}'
    )

def _read_input_rows(input_file: Path):
    """
    Liest Rohdaten flexibel ein und baut exakt die Linden-Spalten A-G:
    A PZN, B Artikelname, C DF/DAR, D Pck/Pack.Gr, E Herst, F EK, G Absatz.

    Unterstützte Formate:
    - Klassische Hochpreiser-/Abverkaufslisten mit Header-Erkennung
    - CSV-Dateien mit automatischer Trennzeichen-/Encoding-Erkennung
    - Monatsverbrauchslisten wie EKM: PZN + YYYYMM-Spalten + Artikel/Hersteller/6 Monate/12 Monate
      Hier werden nur die Rohdaten-Spalten gelesen; vorhandene manuelle Auswertungsspalten rechts davon werden ignoriert.

    EK kommt zuerst aus der aktuellen Rohdatei. Wenn kein EK vorhanden ist, wird der letzte gespeicherte echte Rohdaten-EK als Fallback genutzt.
    """
    ws, input_typ = _load_input_worksheet(input_file)
    mapping = find_columns(ws)
    if not _is_mapping_usable(mapping):
        report = _diagnose_format(input_file, ws, mapping)
        raise UnknownInputFormatError(
            "Dateiformat nicht erkannt. Die Datei wurde nicht ausgewertet. "
            f"Bitte Diagnosebericht prüfen und Datei/Spalten anpassen: {report}",
            report
        )
    header_row = mapping.get("header_row", 1)
    format_typ = mapping.get("format", "standard")

    def get(row, key, fallback_idx=None):
        col = mapping.get(key)
        if col and len(row) >= col:
            return row[col - 1]
        if fallback_idx is not None and len(row) > fallback_idx:
            return row[fallback_idx]
        return None

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(v not in (None, "") for v in row):
            continue

        source = [None] * 7
        if format_typ == "monatsverbrauch":
            # EKM-/Monatsverbrauchsformat: keine Auswertungsspalten rechts übernehmen.
            source[0] = get(row, "pzn", 0)
            source[1] = get(row, "artikel")
            source[2] = None
            source[3] = None
            source[4] = _strip_leading_manufacturer_number(get(row, "hersteller"))
            source[5] = None  # EK gibt es in diesem Format nicht: bewusst leer lassen.
            source[6] = get(row, "absatz")
        else:
            # Fallback-Indizes entsprechen nur dann dem alten Linden-Layout,
            # wenn keine bessere Header-Erkennung vorhanden ist.
            source[0] = get(row, "pzn", 0)
            source[1] = get(row, "artikel", 1)
            source[2] = get(row, "df", 2)
            source[3] = get(row, "packung", 3)
            source[4] = get(row, "hersteller", 4 if not mapping.get("ek") == 5 else None)
            source[5] = get(row, "ek", 5)
            source[6] = get(row, "absatz", 6)

            # Falls EK per Header in Spalte E erkannt wurde und keine Hersteller-Spalte existiert,
            # darf der Wert aus Spalte E nicht als Hersteller-Fallback landen.
            if not mapping.get("hersteller"):
                source[4] = None

        # Nur echte PZN-Zeilen exportieren.
        if not _pzn(source[0]):
            continue
        rows.append(source)

    return rows, mapping

def _aggregate_rows_by_pzn(rows):
    """Fasst Zeilen mit gleicher PZN zusammen und summiert die Menge (Spalte G,
    Index 6). Die Reihenfolge des ersten Auftretens bleibt erhalten.

    - Zeilen ohne gueltige PZN bleiben unveraendert als Einzelzeilen erhalten.
    - Negative Mengen (z.B. Retouren) werden mit aufsummiert und verrechnen sich
      dadurch gegen positive Mengen derselben PZN.
    - Fehlende Stammfelder (Artikel/DF/Pck/Herst/EK) werden aus spaeteren Zeilen
      derselben PZN ergaenzt.
    """
    aggregated: dict[str, list] = {}
    order: list[str] = []
    passthrough: list[list] = []
    for raw in rows:
        row = list(raw)
        pzn = _pzn(row[0] if row else None)
        if not pzn:
            passthrough.append(row)
            continue
        if pzn not in aggregated:
            aggregated[pzn] = row + [None] * max(0, 7 - len(row))
            order.append(pzn)
        else:
            base = aggregated[pzn]
            base[6] = _to_number(base[6], 0) + _to_number(row[6] if len(row) > 6 else 0, 0)
            for i in range(1, 6):
                if base[i] in (None, "") and i < len(row) and row[i] not in (None, ""):
                    base[i] = row[i]
    result = [aggregated[p] for p in order]
    result.extend(passthrough)
    return result


def _wnorm(value):
    """Vergleichsschluessel fuer Wirkstoffnamen: klein, ohne Sonderzeichen."""
    text = str(value or "").strip().lower()
    text = text.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    return re.sub(r"[^a-z0-9]+", "", text)


def _substance_norm(value):
    """Wirkstoff-Vergleichsschluessel OHNE angehaengte Staerke.

    NMG-Stamm fuehrt den Wirkstoff inkl. Staerke ("Etanercept 50 mg"), die
    Wirkstoff-Tabelle der abgegebenen Artikel fuehrt nur den Namen ("Etanercept")
    + separate Staerke. Damit der Austausch ueber den Wirkstoff matcht, wird die
    trailing Staerke+Einheit entfernt.
    """
    s = str(value or "")
    s = re.sub(r"\s+\d[\d.,]*\s*(mg|µg|mcg|ug|g|ml|i\.?e\.?|%).*$", "", s, flags=re.IGNORECASE)
    return _wnorm(s)


def _pzn_core(value):
    """Wie _pzn, strippt aber zusaetzlich das ' /N'-Lagervarianten-Suffix."""
    text = re.sub(r"\s*/\s*\w+\s*$", "", str(value or "").strip())
    return _pzn(text)


def _build_wirkstoff_indices(con):
    """Indizes fuer das Wirkstoff-/Austausch-Blatt und den Wirkstoff-Fallback:
    - wirkstoff_by_pzn:  norm_pzn -> (wirkstoff, staerke)
    - nmg_by_pzn:        norm_pzn -> artikelname (NMG-Stamm)
    - nmg_by_wirkstoff:  wirkstoff_norm -> [(norm_pzn, artikelname)]
    - nmg_detail_by_pzn: norm_pzn -> {apu, rabatt, lieferbar, bevorratung, liefervorschlag}
    """
    wirkstoff_by_pzn = {}
    try:
        for pzn, wirkstoff, staerke in con.execute(
            "SELECT pzn, wirkstoff, staerke FROM tbl_wirkstoff_staerke"
        ):
            key = _pzn_core(pzn)
            if key and key not in wirkstoff_by_pzn:
                wirkstoff_by_pzn[key] = (str(wirkstoff or "").strip(), str(staerke or "").strip())
    except Exception:
        pass

    nmg_by_pzn = {}
    nmg_by_wirkstoff = {}
    nmg_detail_by_pzn = {}
    try:
        for pzn, name, wirkstoffe, apu in con.execute(
            "SELECT pzn, artikelname, wirkstoffe, apu FROM tbl_nmg_stamm"
        ):
            key = _pzn_core(pzn)
            nm = str(name or "").strip()
            if not key:
                continue
            nmg_by_pzn[key] = nm
            nmg_detail_by_pzn.setdefault(key, {})["apu"] = apu
            wk = _substance_norm(wirkstoffe)
            if wk:
                nmg_by_wirkstoff.setdefault(wk, []).append((key, nm))
    except Exception:
        pass

    # Rabatt + Lieferfaehigkeit je NMG-PZN ergaenzen (fuer den Wirkstoff-Fallback).
    try:
        for nmg_pzn, rabatt in con.execute("SELECT nmg_pzn, rabatt FROM nmg_rabatte"):
            d = nmg_detail_by_pzn.get(_pzn_core(nmg_pzn))
            if d is not None and rabatt is not None:
                d["rabatt"] = rabatt
    except Exception:
        pass
    try:
        for nmg_pzn, lieferbar, bevorratung, liefervorschlag in con.execute(
            "SELECT nmg_pzn, lieferbar, bevorratung_angeraten, liefervorschlag FROM tbl_lieferfaehigkeit"
        ):
            d = nmg_detail_by_pzn.get(_pzn_core(nmg_pzn))
            if d is not None:
                d["lieferbar"] = lieferbar
                d["bevorratung"] = bevorratung
                d["liefervorschlag"] = liefervorschlag
    except Exception:
        pass

    return wirkstoff_by_pzn, nmg_by_pzn, nmg_by_wirkstoff, nmg_detail_by_pzn


def create_vorlage_export(input_file: str | Path, apotheke: str, on_duplicate_prompt=None) -> Path:
    input_file = Path(input_file)
    if not DB_PATH.exists():
        init_db(DB_PATH)
    safe_name = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in apotheke.strip()) or "Apotheke"
    # V1.1 SP12: Strukturiert nach OUTPUT_DIR/PK/<Jahr>/Q<n>/
    # create_vorlage_export ist der PK-Auswertungs-Pfad (Apotheken-Excel).
    # ZW kommt aus import_historical_market_file und erzeugt gar keine Excel.
    from .config import jahr_quartal_pfad
    out = jahr_quartal_pfad(OUTPUT_DIR, "PK") / f"Auswertung_{safe_name}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

    rows, input_mapping = _read_input_rows(input_file)

    # Dubletten-Erkennung: kommt dieselbe PZN mehrfach vor, fragt die GUI ueber
    # on_duplicate_prompt, ob die Mengen je PZN summiert werden sollen. Ohne
    # Callback (z.B. Stapel-/Hintergrundlauf) bleibt das alte Verhalten: jede
    # Zeile einzeln.
    if on_duplicate_prompt is not None and rows:
        counts: dict[str, int] = {}
        has_negative = False
        for row in rows:
            pzn = _pzn(row[0] if row else None)
            if not pzn:
                continue
            counts[pzn] = counts.get(pzn, 0) + 1
            if _to_number(row[6] if len(row) > 6 else 0, 0) < 0:
                has_negative = True
        dups = sorted(p for p, c in counts.items() if c > 1)
        if dups:
            info = {"pzns": dups, "count": len(dups), "has_negative": has_negative}
            try:
                if on_duplicate_prompt(info):
                    rows = _aggregate_rows_by_pzn(rows)
            except Exception:
                pass

    wb = Workbook()
    ws = wb.active
    ws.title = apotheke[:31] if apotheke else "Auswertung"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    start_row = 1
    for col, header in enumerate(LINDEN_HEADERS, start=1):
        cell = ws.cell(start_row, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    with sqlite3.connect(DB_PATH) as con:
        con.row_factory = sqlite3.Row
        # SP24: Damit JOINs/WHEREs auf PZN-Spalten Format-Unterschiede tolerieren
        # (Excel-Int vs. String, fuehrende Nullen, ".0"-Suffix). _pzn() zfilled
        # auf 8 Stellen wenn rein numerisch, sonst nimmt es den getrimmten String.
        con.create_function("pzn_norm", 1, _pzn)

        # SP27: Stammtabellen einmal in den RAM laden. Danach laeuft jeder
        # pro-Zeile-Lookup als Dict-Zugriff statt als SQL-Query.
        caches = _load_lookup_caches(con)
        # V1.1 SP19: Abgleich-Trace-Index einmal aufbauen. Vorher hat
        # trace_lookup_source pro Zeile die komplette Austauschdatenbank
        # (~18k Zeilen) gefetcht und in Python iteriert - bei 400+ Zeilen
        # >90 % der Auswertungszeit (~170 s von 180 s). Mit Index nur Dict-Lookup.
        trace_index = build_trace_index(con)

        # Indizes fuer das zusaetzliche Wirkstoff-/NMG-Austausch-Blatt und den
        # Wirkstoff-Fallback (NMG-Empfehlung ueber den Wirkstoff, wenn keine
        # direkte PZN-Zuordnung existiert).
        wirkstoff_by_pzn, nmg_by_pzn, nmg_by_wirkstoff, nmg_detail_by_pzn = _build_wirkstoff_indices(con)
        wirkstoff_rows = []
        seen_wirkstoff_pzn = set()
        wirkstoff_fallback_einzel = 0
        wirkstoff_fallback_mehrere = 0

        auswertung_id = None
        statistik = {"positionen": 0, "nmg_treffer": 0, "nicht_nmg": 0, "gesamt_absatz": 0.0}
        abgleich_rows = []
        abgleich_protocol_path = None
        try:
            cur = con.execute(
                """INSERT INTO tbl_auswertungen(apotheke, quelldatei, ausgabedatei, bemerkung)
                   VALUES (?, ?, ?, ?)""",
                (apotheke, input_file.name, str(out), "automatisch beim Export gespeichert")
            )
            auswertung_id = cur.lastrowid
        except Exception:
            auswertung_id = None
        try:
            con.execute("""INSERT INTO tbl_rohdaten_mapping(dateiname, pzn_spalte, hersteller_spalte, ek_spalte, absatz_spalte, header_zeile, format_typ, letzte_aktualisierung)
                         VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                         ON CONFLICT(dateiname) DO UPDATE SET
                           pzn_spalte=excluded.pzn_spalte, hersteller_spalte=excluded.hersteller_spalte,
                           ek_spalte=excluded.ek_spalte, absatz_spalte=excluded.absatz_spalte,
                           header_zeile=excluded.header_zeile, format_typ=excluded.format_typ, letzte_aktualisierung=CURRENT_TIMESTAMP""",
                        (input_file.name, str(input_mapping.get("pzn_header", input_mapping.get("pzn", ""))),
                         str(input_mapping.get("hersteller_header", input_mapping.get("hersteller", ""))),
                         str(input_mapping.get("ek_header", input_mapping.get("ek", ""))),
                         str(input_mapping.get("absatz_header", input_mapping.get("absatz", ""))),
                         int(input_mapping.get("header_row", 1)), str(input_mapping.get("format", "standard"))))
            con.commit()
        except Exception:
            pass
        for r_idx, source_row in enumerate(rows, start=2):
            original_pzn = _pzn(source_row[0] if source_row else None)

            # A-G aus Rohdatei übernehmen. Fehlende Artikel/DF/PCK/Hersteller werden aus der PZN-Basisdatenbank ergänzt.
            # SP27: alle Reads gehen ueber den Cache statt SQL.
            basis = _lookup_basisdaten_fast(caches, original_pzn)
            artikel_roh = source_row[1] if len(source_row) > 1 else None
            df_roh = source_row[2] if len(source_row) > 2 else None
            pck_roh = source_row[3] if len(source_row) > 3 else None
            hersteller_roh = clean_hersteller(source_row[4] if len(source_row) > 4 else None)

            # SP19: Wenn weder Rohdatei noch selbstlernende Basis Werte liefert,
            # auf Stammdaten zurueckgreifen (Artikelstamm, PZN-Basis, NMG-Stamm).
            stamm = None
            artikel_final = artikel_roh or basis.get("artikelname")
            df_final = df_roh or basis.get("df")
            pck_final = pck_roh or basis.get("pck")
            hersteller_final = hersteller_roh or basis.get("herstellerkuerzel") or _lookup_hersteller_fast(caches, original_pzn)

            if not (artikel_final and df_final and pck_final and hersteller_final):
                stamm = _lookup_stammdaten_fast(caches, original_pzn)
                artikel_final = artikel_final or stamm.get("artikelname")
                df_final = df_final or stamm.get("df")
                pck_final = pck_final or stamm.get("pck")
                hersteller_final = hersteller_final or stamm.get("hersteller")

            register_basisdaten(con, original_pzn, artikel_roh, hersteller_roh, df_roh, pck_roh, input_file.name)

            ek_roh = source_row[5] if len(source_row) > 5 else None
            ek_num = parse_number(ek_roh)
            if ek_num is not None:
                register_ek(con, original_pzn, ek_num, input_file.name, str(input_mapping.get("ek_header") or input_mapping.get("ek") or ""))
                ek_final = ek_num
            else:
                ek_final = _lookup_latest_ek_fast(caches, original_pzn)
                if ek_final is None:
                    if stamm is None:
                        stamm = _lookup_stammdaten_fast(caches, original_pzn)
                    ek_final = stamm.get("ek")

            output_values = [
                original_pzn, artikel_final, df_final, pck_final, hersteller_final,
                ek_final if ek_final is not None else None,
                source_row[6] if len(source_row) > 6 else None,
            ]
            for c_idx, value in enumerate(output_values, start=1):
                ws.cell(r_idx, c_idx, value)

            hit = _lookup_fast(caches, original_pzn)
            abgleich_trace = trace_lookup_source(con, original_pzn, hit, index=trace_index)
            abgleich_trace.update({
                "zeile": r_idx,
                "artikel_alt": artikel_final or "",
                "df": df_final or "",
                "pck": pck_final or "",
                "hersteller": hersteller_final or "",
                "absatz_6m": source_row[6] if len(source_row) > 6 else "",
            })
            abgleich_rows.append(abgleich_trace)

            # Wirkstoff-Fallback: Keine direkte NMG-Zuordnung gefunden, aber es gibt
            # NMG-Produkt(e) mit demselben Wirkstoff -> als Empfehlung uebernehmen.
            #   genau 1 Treffer  -> wird zur NMG-Empfehlung (PZN NMG + APU/Rabatt/Liefer)
            #   mehrere Treffer  -> alle in "austauschbar gegen", PZN NMG bleibt leer
            if not hit.get("nmg_pzn") and hit.get("apu_nmg") is None:
                wk_name = wirkstoff_by_pzn.get(original_pzn, ("", ""))[0]
                cands = nmg_by_wirkstoff.get(_substance_norm(wk_name), []) if wk_name else []
                # Dubletten (gleiche PZN) zusammenfassen, Reihenfolge erhalten.
                uniq = []
                seen_c = set()
                for cp, cn in cands:
                    if cp and cp not in seen_c:
                        seen_c.add(cp)
                        uniq.append((cp, cn))
                if len(uniq) == 1:
                    npzn, nname = uniq[0]
                    d = nmg_detail_by_pzn.get(npzn, {})
                    hit = dict(hit)
                    hit.update({
                        "im_sortiment": "Austausch (Wirkstoff)",
                        "nmg_pzn": npzn,
                        "apu_nmg": d.get("apu"),
                        "rabatt": d.get("rabatt"),
                        "lieferbar": d.get("lieferbar"),
                        "bevorratung": d.get("bevorratung"),
                        "liefervorschlag": d.get("liefervorschlag"),
                        "austauschbar_gegen": None,
                    })
                    wirkstoff_fallback_einzel += 1
                elif len(uniq) > 1:
                    text = " | ".join((f"{p} – {n}" if n else p) for p, n in uniq)
                    hit = dict(hit)
                    hit.update({
                        "im_sortiment": "Austausch mögl. (Wirkstoff)",
                        "austauschbar_gegen": text,
                    })
                    wirkstoff_fallback_mehrere += 1

            # Wunsch: Steht bereits eine NMG-PZN (Spalte 9), soll "austauschbar
            # gegen" (Spalte 15) leer bleiben - die Austausch-Details stehen im
            # zusaetzlichen Wirkstoff-Blatt.
            austauschbar_wert = None if hit.get("nmg_pzn") else hit.get("austauschbar_gegen")

            im_sort_val = hit.get("im_sortiment")
            c8 = ws.cell(r_idx, 8, im_sort_val)
            # Wirkstoff-Treffer farblich markieren, damit sie beim Pruefen auffallen.
            if im_sort_val == "Austausch (Wirkstoff)":
                c8.fill = PatternFill("solid", fgColor="DFF5E3")   # zartgruen = feste Empfehlung
            elif im_sort_val and str(im_sort_val).startswith("Austausch mögl. (Wirkstoff)"):
                c8.fill = PatternFill("solid", fgColor="FFF3D6")   # zartgelb = mehrere Optionen
            ws.cell(r_idx, 9, hit.get("nmg_pzn"))
            ws.cell(r_idx, 10, hit.get("apu_nmg"))
            ws.cell(r_idx, 11, hit.get("rabatt"))
            ws.cell(r_idx, 12, hit.get("lieferbar"))
            ws.cell(r_idx, 13, hit.get("bevorratung"))
            ws.cell(r_idx, 14, hit.get("liefervorschlag"))
            ws.cell(r_idx, 15, austauschbar_wert)

            # Wirkstoff-/NMG-Austausch-Blatt: je abgegebener PZN einmal sammeln.
            if original_pzn not in seen_wirkstoff_pzn:
                seen_wirkstoff_pzn.add(original_pzn)
                wk_name, wk_staerke = wirkstoff_by_pzn.get(original_pzn, ("", ""))
                austausch_items = []
                seen_nmg = set()
                direct = _pzn(hit.get("nmg_pzn"))
                if direct:
                    seen_nmg.add(direct)
                    austausch_items.append((direct, nmg_by_pzn.get(direct, "")))
                for npzn, nname in nmg_by_wirkstoff.get(_substance_norm(wk_name), []):
                    if npzn and npzn not in seen_nmg:
                        seen_nmg.add(npzn)
                        austausch_items.append((npzn, nname))
                if wk_name or austausch_items:
                    # Pivot-freundlich: Kandidatenliste behalten, im Blatt je
                    # NMG-Artikel eine eigene Zeile schreiben (nicht in eine Zelle).
                    wirkstoff_rows.append(
                        (original_pzn, artikel_final or "", wk_name, wk_staerke, austausch_items)
                    )

            # 0.7: Werte direkt berechnen, damit sie sofort sichtbar sind und auch beim
            # automatischen Vergleich nicht als leer erscheinen.
            # Logik bleibt: P = J × K, Q = P × G, R = J × G
            apu_wert = _to_number(hit.get("apu_nmg"), 0)
            rabatt_wert = _to_number(hit.get("rabatt"), 0)
            menge_wert = _to_number(source_row[6] if len(source_row) > 6 else 0, 0)
            ek_wert = _to_number(ek_final, 0)
            rabatt_euro = apu_wert * rabatt_wert
            ws.cell(r_idx, 16, rabatt_euro)
            ws.cell(r_idx, 17, rabatt_euro * menge_wert)
            # Umsatz wurde im geprüften Rosen/Linden-Stand aus EK × Absatz gebildet.
            # Das überschreibt die frühere Annahme R = J × G.
            umsatz_wert = ek_wert * menge_wert
            ws.cell(r_idx, 18, umsatz_wert)

            ist_nmg = 1 if (hit.get("nmg_pzn") or hit.get("im_sortiment") or hit.get("apu_nmg") is not None) else 0
            statistik["positionen"] += 1
            statistik["gesamt_absatz"] += menge_wert
            if ist_nmg:
                statistik["nmg_treffer"] += 1
            else:
                statistik["nicht_nmg"] += 1
            if auswertung_id is not None:
                try:
                    con.execute(
                        """INSERT INTO tbl_auswertungspositionen(
                            auswertung_id, pzn, artikelname, df, pck, herstellerkuerzel, ek, absatz_6m,
                            im_sortiment, pzn_nmg, apu_nmg, nmg_rabatt, lieferbar, bevorratung_angeraten,
                            liefervorschlag, austauschbar_gegen, nmg_rabatt_euro, nmg_rabatt_gesamt, umsatz,
                            ist_nmg_treffer, quelle
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (auswertung_id, original_pzn, artikel_final, df_final, pck_final, hersteller_final,
                         ek_final, menge_wert, hit.get("im_sortiment"), hit.get("nmg_pzn"),
                         hit.get("apu_nmg"), hit.get("rabatt"), hit.get("lieferbar"),
                         hit.get("bevorratung"), hit.get("liefervorschlag"), austauschbar_wert,
                         rabatt_euro, rabatt_euro * menge_wert, umsatz_wert, ist_nmg, input_file.name)
                    )
                except Exception:
                    pass
        try:
            abgleich_protocol_path = write_abgleichartikel_protocol(
                abgleich_rows,
                analyse_name=apotheke,
                input_file=input_file,
                output_file=out,
                auswertung_id=auswertung_id,
            )
        except Exception:
            abgleich_protocol_path = None

        if auswertung_id is not None:
            try:
                bemerkung = "automatisch beim Export gespeichert"
                if abgleich_protocol_path:
                    bemerkung += f" | Abgleichartikel-Protokoll: {abgleich_protocol_path}"
                con.execute(
                    """UPDATE tbl_auswertungen
                       SET anzahl_positionen=?, nmg_treffer=?, nicht_nmg=?, gesamt_absatz=?, bemerkung=?
                       WHERE id=?""",
                    (statistik["positionen"], statistik["nmg_treffer"], statistik["nicht_nmg"],
                     statistik["gesamt_absatz"], bemerkung, auswertung_id)
                )
                con.commit()
            except Exception:
                pass

    last_row = max(2, len(rows) + 1)

    # Zahlenformat wie Linden: Preise/Rabatte/Umsatz sichtbar und berechenbar.
    for row in range(2, last_row + 1):
        for col in [6, 10, 16, 17, 18]:
            ws.cell(row, col).number_format = '#,##0.00'
        ws.cell(row, 11).number_format = '0.00%'
        ws.cell(row, 7).number_format = '0'

    thin = Side(style="thin", color="B7B7B7")
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=18):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [10, 32, 8, 10, 10, 12, 18, 16, 12, 12, 12, 10, 16, 14, 28, 14, 18, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:R{last_row}"

    # Zweites Blatt: Wirkstoff der abgegebenen Artikel + austauschbare NMG-Artikel.
    # Pivot-freundlich: je möglichem NMG-Artikel eine eigene Zeile (nicht in eine Zelle).
    try:
        ws2 = wb.create_sheet("Wirkstoff & NMG-Austausch")
        # Zeile 1: Hinweiszeile (ueber die ganze Breite).
        ws2.merge_cells("A1:F1")
        note = ws2.cell(1, 1, "Hinweis: Abgleich erfolgt über den Wirkstoff (ohne Stärke). "
                              "Je möglichem NMG-Artikel eine eigene Zeile – als Grundlage für eine Pivot-Tabelle.")
        note.font = Font(italic=True, color="666666", size=9)
        note.alignment = Alignment(vertical="center", wrap_text=True)
        # Zeile 2: Kopfzeile.
        kopf = ["PZN (abgegeben)", "Artikel (abgegeben)", "Wirkstoff", "Stärke", "NMG-PZN", "NMG-Artikel"]
        for c, header in enumerate(kopf, start=1):
            cell = ws2.cell(2, c, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        r2 = 3
        for pzn, artikel, wk_name, wk_staerke, items in wirkstoff_rows:
            zeilen = items if items else [("", "")]
            for npzn, nname in zeilen:
                ws2.cell(r2, 1, pzn)
                ws2.cell(r2, 2, artikel)
                ws2.cell(r2, 3, wk_name)
                ws2.cell(r2, 4, wk_staerke)
                ws2.cell(r2, 5, npzn)
                ws2.cell(r2, 6, nname)
                r2 += 1
        if r2 == 3:
            ws2.cell(3, 1, "Keine Wirkstoff-/Austausch-Treffer gefunden.")
            r2 = 4
        thin2 = Side(style="thin", color="B7B7B7")
        for row in ws2.iter_rows(min_row=2, max_row=r2 - 1, min_col=1, max_col=6):
            for cell in row:
                cell.border = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for idx, width in enumerate([16, 34, 24, 10, 14, 40], start=1):
            ws2.column_dimensions[get_column_letter(idx)].width = width
        ws2.row_dimensions[1].height = 26
        ws2.row_dimensions[2].height = 28
        ws2.freeze_panes = "A3"
        ws2.auto_filter.ref = f"A2:F{max(2, r2 - 1)}"
    except Exception:
        pass

    wb.save(out)
    return out
