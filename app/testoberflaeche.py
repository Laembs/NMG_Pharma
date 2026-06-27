"""Testoberflaeche (Preview) fuer NMGone – moderne UI, isoliert vom Original.

Ziel: ein professioneller, leicht verstaendlicher Entwurf der Neuen Auswertung,
der die echte Engine (create_vorlage_export) nutzt, aber das Original-GUI NICHT
veraendert. Vier Schwerpunkte:

  1. Sichtbarer Fortschritt – animierter Balken, Stoppuhr, Stufen-Checkliste,
     echte Positionszahl aus der DB (man sieht, dass gearbeitet wird).
  2. Modernes, professionelles Design – Karten, klare Typografie, Sidebar.
  3. Hilfe mit Bildern und Erklaerungen.
  4. .xls-Unterstuetzung – alte Excel-Dateien werden automatisch nach .xlsx
     konvertiert (xlrd -> openpyxl) und dann von der echten Engine verarbeitet.

Start:  python start_testoberflaeche.py
"""
from __future__ import annotations

import threading
from .i18n import T as _T
import time
import sqlite3
from pathlib import Path

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from .config import DB_PATH
from .testui_assets import ensure_help_images

# ── Palette ──────────────────────────────────────────────────────────────────
PRIMARY = "#0B4A86"
PRIMARY_DARK = "#083A6B"
ACCENT = "#208ACD"
SUCCESS = "#11823B"
WARNING = "#C88200"
DANGER = "#B3261E"
BG = "#F4F6F9"
CARD = "#FFFFFF"
INK = "#1F2933"
MUTED = "#626E7D"
BORDER = "#DCE2E8"
SIDEBAR = "#0B2C4A"
SIDEBAR_ACTIVE = "#13456E"

FONT = "Segoe UI"

SUPPORTED = {".xlsx", ".xlsm", ".xls", ".csv", ".txt"}


# ── .xls-Bruecke ─────────────────────────────────────────────────────────────
def konvertiere_xls_zu_xlsx(xls_path: Path) -> Path:
    """Wandelt eine alte .xls-Datei in eine temporaere .xlsx, damit die
    bestehende Engine (die nur xlsx/xlsm/csv/txt liest) sie verarbeiten kann.
    Liest mit xlrd, schreibt mit openpyxl. Liefert den Pfad der neuen Datei.
    """
    import xlrd
    from openpyxl import Workbook

    import tempfile

    book = xlrd.open_workbook(str(xls_path))
    sheet = book.sheet_by_index(0)
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet.name or "Tabelle1")[:31]
    for r in range(sheet.nrows):
        ws.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
    out = Path(tempfile.gettempdir()) / f"_xls_konvertiert_{xls_path.stem[:40]}_{int(time.time())}.xlsx"
    wb.save(out)
    return out


def zeilen_schaetzen(path: Path) -> int:
    """Grobe Zeilenzahl der Quelldatei fuer die Fortschrittsschaetzung."""
    suffix = path.suffix.lower()
    try:
        if suffix in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            return max(0, wb.active.max_row - 1)
        if suffix == ".xls":
            import xlrd
            return max(0, xlrd.open_workbook(str(path)).sheet_by_index(0).nrows - 1)
        with open(path, "r", encoding="utf-8", errors="replace") as fh:
            return max(0, sum(1 for _ in fh) - 1)
    except Exception:
        return 0


# ── Datenzugriff (read-only) ─────────────────────────────────────────────────
def _norm_pzn(value) -> str:
    """8-stellige PZN aus beliebiger Eingabe (nur Ziffern, links mit 0 gefuellt)."""
    import re
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\D", "", text)
    return text.zfill(8) if text else ""


def _db():
    """Read-only Verbindung zur aktuellen NMGone-Datenbank."""
    con = sqlite3.connect(f"file:{DB_PATH}?mode=ro", uri=True)
    con.row_factory = sqlite3.Row
    return con


def _table_exists(con, name: str) -> bool:
    return con.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (name,)
    ).fetchone() is not None


def artikel_360(pzn: str) -> dict:
    """Sammelt alle verfuegbaren Informationen zu einer PZN aus allen Tabellen."""
    p = _norm_pzn(pzn)
    out: dict = {"pzn": p, "gefunden": False}
    if not p:
        return out
    con = _db()
    try:
        r = con.execute(
            "SELECT artikelname,herstellerkuerzel,apu,taxe_ek,taxe_vk,menge,einheit,wirkstoffe "
            "FROM tbl_nmg_stamm WHERE pzn=?", (p,)).fetchone()
        out["nmg"] = dict(r) if r else None

        r = con.execute(
            "SELECT artikel,df,pck,herst,ek,quelle FROM tbl_artikelstamm WHERE pzn=?", (p,)).fetchone()
        out["stamm"] = dict(r) if r else None

        r = con.execute("SELECT rabatt,quelle FROM nmg_rabatte WHERE nmg_pzn=?", (p,)).fetchone()
        out["rabatt"] = dict(r) if r else None

        if _table_exists(con, "tbl_lieferfaehigkeit"):
            r = con.execute(
                "SELECT lieferbar,bevorratung_angeraten,liefervorschlag "
                "FROM tbl_lieferfaehigkeit WHERE nmg_pzn=?", (p,)).fetchone()
            out["liefer"] = dict(r) if r else None

        # bester verfuegbarer EK (Stamm-EK oder Roh-EK)
        ek = (out.get("stamm") or {}).get("ek")
        if ek is None and _table_exists(con, "tbl_pzn_ek_rohdaten"):
            r = con.execute(
                "SELECT ek FROM tbl_pzn_ek_rohdaten WHERE pzn=? AND ek IS NOT NULL "
                "ORDER BY importdatum DESC LIMIT 1", (p,)).fetchone()
            if r:
                ek = r["ek"]
        out["ek_effektiv"] = ek

        # Biosimilar-Gruppe + Alternativen
        out["biosimilar"] = None
        if _table_exists(con, "tbl_biosimilar_pzn"):
            link = con.execute(
                "SELECT produkt_id FROM tbl_biosimilar_pzn WHERE pzn=?", (p,)).fetchone()
            if link:
                prod = con.execute(
                    "SELECT gruppe_id,name,rolle FROM tbl_biosimilar_produkt WHERE id=?",
                    (link["produkt_id"],)).fetchone()
                if prod:
                    grp = con.execute(
                        "SELECT wirkstoff,referenzprodukt FROM tbl_biosimilar_gruppe WHERE id=?",
                        (prod["gruppe_id"],)).fetchone()
                    alts = con.execute(
                        "SELECT DISTINCT bp.name,bp.rolle FROM tbl_biosimilar_produkt bp "
                        "WHERE bp.gruppe_id=? AND bp.id<>? ORDER BY bp.rolle,bp.name",
                        (prod["gruppe_id"], link["produkt_id"])).fetchall()
                    out["biosimilar"] = {
                        "wirkstoff": grp["wirkstoff"] if grp else None,
                        "referenz": grp["referenzprodukt"] if grp else None,
                        "rolle": prod["rolle"],
                        "name": prod["name"],
                        "alternativen": [dict(a) for a in alts],
                    }

        # Austausch-Eintraege (diese PZN wird ersetzt durch NMG-PZN)
        aus = con.execute(
            "SELECT pzn_nmg,artikel_nmg,status FROM tbl_austauschdatenbank "
            "WHERE pzn_alt=? AND COALESCE(status,'aktiv')<>'inaktiv' LIMIT 12", (p,)).fetchall()
        out["austausch_zu_nmg"] = [dict(a) for a in aus]

        out["gefunden"] = any(out.get(k) for k in ("nmg", "stamm", "rabatt"))
    finally:
        con.close()
    return out


def suche_artikel(begriff: str, limit: int = 25) -> list[dict]:
    """Sucht nach PZN-Teil oder Artikelnamen in NMG-Stamm und Artikelstamm."""
    begriff = (begriff or "").strip()
    if not begriff:
        return []
    con = _db()
    try:
        if begriff.replace(".", "").isdigit():
            like = f"%{_norm_pzn(begriff).lstrip('0')}%"
            rows = con.execute(
                "SELECT pzn,artikelname AS name,herstellerkuerzel AS herst FROM tbl_nmg_stamm "
                "WHERE pzn LIKE ? LIMIT ?", (f"%{begriff}%", limit)).fetchall()
        else:
            like = f"%{begriff}%"
            rows = con.execute(
                "SELECT pzn,artikelname AS name,herstellerkuerzel AS herst FROM tbl_nmg_stamm "
                "WHERE artikelname LIKE ? ORDER BY artikelname LIMIT ?", (like, limit)).fetchall()
        return [dict(r) for r in rows]
    finally:
        con.close()


def db_kennzahlen() -> dict:
    """Liefert Kennzahlen fuer das Dashboard."""
    con = _db()
    k: dict = {}
    try:
        def count(t):
            try:
                return con.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
            except Exception:
                return 0
        k["artikel"] = count("tbl_artikelstamm")
        k["nmg_stamm"] = count("tbl_nmg_stamm")
        k["rabatte"] = count("nmg_rabatte")
        k["austausch"] = count("tbl_austauschdatenbank")
        k["biosimilar_gruppen"] = count("tbl_biosimilar_gruppe")
        k["biosimilar_pzn"] = count("tbl_biosimilar_pzn")
        k["auswertungen"] = count("tbl_auswertungen")
        k["positionen"] = count("tbl_auswertungspositionen")
        # Rabatt-Verteilung
        try:
            row = con.execute(
                "SELECT AVG(rabatt),MIN(rabatt),MAX(rabatt) FROM nmg_rabatte WHERE rabatt IS NOT NULL").fetchone()
            k["rabatt_avg"], k["rabatt_min"], k["rabatt_max"] = row[0], row[1], row[2]
        except Exception:
            k["rabatt_avg"] = k["rabatt_min"] = k["rabatt_max"] = None
        # Top-Hersteller im Artikelstamm (echte Lieferanten-/Anbietervielfalt)
        try:
            k["top_herst"] = [dict(r) for r in con.execute(
                "SELECT herst, COUNT(*) AS n FROM tbl_artikelstamm "
                "WHERE herst IS NOT NULL AND herst<>'' "
                "GROUP BY herst ORDER BY n DESC LIMIT 8").fetchall()]
        except Exception:
            k["top_herst"] = []
        # Lieferfaehigkeit
        if _table_exists(con, "tbl_lieferfaehigkeit"):
            k["lieferbar"] = con.execute(
                "SELECT COUNT(*) FROM tbl_lieferfaehigkeit WHERE lieferbar='X'").fetchone()[0]
            k["bevorratung"] = con.execute(
                "SELECT COUNT(*) FROM tbl_lieferfaehigkeit WHERE bevorratung_angeraten='X'").fetchone()[0]
        # letzte Auswertung
        try:
            r = con.execute(
                "SELECT datum,kundenname,apotheke FROM tbl_auswertungen ORDER BY id DESC LIMIT 1").fetchone()
            k["letzte"] = dict(r) if r else None
        except Exception:
            k["letzte"] = None
    finally:
        con.close()
    return k


def markt_top(modus: str = "nmg", limit: int = 30) -> list[dict]:
    """Aggregiert ueber ALLE Auswertungspositionen, was die Apotheken-Kunden
    tatsaechlich nachfragen.

    modus='nmg'    -> nur NMG-Treffer, sortiert nach Rabatt-Potenzial (Einkaufs-Sicht)
    modus='bedarf' -> alle Produkte, sortiert nach Gesamt-Absatz (Markt-Sicht)
    """
    con = _db()
    try:
        if modus == "nmg":
            sql = (
                "SELECT pzn, MAX(artikelname) AS name, "
                "COUNT(DISTINCT auswertung_id) AS apotheken, "
                "SUM(COALESCE(absatz_6m,0)) AS absatz, "
                "SUM(COALESCE(nmg_rabatt_gesamt,0)) AS rabatt_pot, "
                "AVG(nmg_rabatt) AS rabatt_satz "
                "FROM tbl_auswertungspositionen "
                "WHERE ist_nmg_treffer=1 AND pzn IS NOT NULL AND pzn<>'' "
                "GROUP BY pzn ORDER BY rabatt_pot DESC LIMIT ?")
        else:
            sql = (
                "SELECT pzn, MAX(artikelname) AS name, "
                "COUNT(DISTINCT auswertung_id) AS apotheken, "
                "SUM(COALESCE(absatz_6m,0)) AS absatz, "
                "SUM(COALESCE(nmg_rabatt_gesamt,0)) AS rabatt_pot, "
                "MAX(ist_nmg_treffer) AS ist_nmg "
                "FROM tbl_auswertungspositionen "
                "WHERE pzn IS NOT NULL AND pzn<>'' "
                "GROUP BY pzn ORDER BY absatz DESC LIMIT ?")
        return [dict(r) for r in con.execute(sql, (limit,)).fetchall()]
    finally:
        con.close()


def auswertungen_liste(limit: int = 60) -> list[dict]:
    con = _db()
    try:
        rows = con.execute(
            "SELECT id,datum,COALESCE(NULLIF(kundenname,''),apotheke) AS kunde,"
            "anzahl_positionen,nmg_treffer,nicht_nmg,gesamt_absatz,ausgabedatei,datenquelle "
            "FROM tbl_auswertungen ORDER BY id DESC LIMIT ?", (limit,)).fetchall()
        return [dict(r) for r in rows]
    finally:
        con.close()


# ── Kunden / ABC (read-only) ─────────────────────────────────────────────────
def _norm_name(s) -> str:
    """Vereinheitlicht Apotheken-/Kundennamen fuer den Abgleich."""
    import re
    t = str(s or "").lower()
    t = t.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    return re.sub(r"[^a-z0-9]+", " ", t).strip()


_STOP = {"apotheke", "apo", "die", "der", "das", "am", "an", "zur", "zum", "und",
         "monate", "bench", "bench2", "hr", "herr", "frau", "12", "gmbh"}


def _name_tokens(s) -> set[str]:
    return {w for w in _norm_name(s).split() if len(w) >= 4 and w not in _STOP}


def abc_ranking() -> list[dict]:
    """ABC-Analyse aller Apotheken aus den Auswertungen, nach Rabatt-Potenzial.

    Mehrfach-Auswertungen derselben Apotheke werden zusammengefasst (juengste
    zaehlt). Klassen nach Pareto: A bis 80 %, B bis 95 %, C Rest des kumulierten
    Potenzials.
    """
    con = _db()
    try:
        rows = con.execute(
            "SELECT a.id, a.apotheke, a.datum, a.gesamt_absatz, a.nmg_treffer, "
            "a.anzahl_positionen, COALESCE(SUM(p.nmg_rabatt_gesamt),0) AS pot "
            "FROM tbl_auswertungen a "
            "LEFT JOIN tbl_auswertungspositionen p ON p.auswertung_id=a.id "
            "GROUP BY a.id ORDER BY a.id DESC").fetchall()
    finally:
        con.close()
    # je Apotheke (normalisiert) die juengste Auswertung behalten
    best: dict[str, dict] = {}
    for r in rows:
        key = _norm_name(r["apotheke"])
        if not key or key in best:
            continue
        best[key] = {"apotheke": r["apotheke"], "datum": r["datum"],
                     "absatz": r["gesamt_absatz"] or 0, "treffer": r["nmg_treffer"] or 0,
                     "positionen": r["anzahl_positionen"] or 0, "pot": r["pot"] or 0.0}
    items = sorted(best.values(), key=lambda d: d["pot"], reverse=True)
    total = sum(d["pot"] for d in items) or 1.0
    kum = 0.0
    for d in items:
        kum += d["pot"]
        anteil = kum / total
        d["klasse"] = "A" if anteil <= 0.80 else ("B" if anteil <= 0.95 else "C")
        d["anteil_kum"] = anteil
    return items


def kunden_master() -> list[dict]:
    """Stammkunden aus dem Kunden-Center inkl. zugeordneter ABC-Klasse (Fuzzy)."""
    con = _db()
    try:
        if not _table_exists(con, "tbl_kunden_center"):
            return []
        rows = con.execute(
            "SELECT kundennummer,kundenname,kundentyp,ansprechpartner,telefon,email,"
            "status,notizen,plz,ort,strasse,hausnummer,inhaber,inhaber_vorname,"
            "inhaber_zuname,besteller_name,besteller_email,rechnungsemail "
            "FROM tbl_kunden_center ORDER BY kundenname").fetchall()
        kunden = [dict(r) for r in rows]
    finally:
        con.close()
    # ABC zuordnen ueber Token-Match gegen die Auswertungs-Apotheken
    ranking = abc_ranking()
    for k in kunden:
        ktok = _name_tokens(k.get("kundenname"))
        treffer = None
        for a in ranking:
            atok = _name_tokens(a["apotheke"])
            if ktok & atok:
                treffer = a
                break
        if treffer:
            k["abc"] = treffer["klasse"]
            k["abc_pot"] = treffer["pot"]
            k["abc_quelle"] = treffer["apotheke"]
        else:
            k["abc"] = None
            k["abc_pot"] = None
        from app import geo_de
        k["latlon"] = geo_de.plz_to_latlon(k.get("plz"))
    return kunden


def _zeitraum_cutoff(monate: int | None) -> str | None:
    """Liefert das Startdatum (YYYY-MM-DD) fuer 'letzte N Monate' oder None=alles."""
    if not monate:
        return None
    from datetime import date
    import calendar
    t = date.today()
    m = t.month - monate
    y = t.year
    while m <= 0:
        m += 12
        y -= 1
    d = min(t.day, calendar.monthrange(y, m)[1])
    return f"{y:04d}-{m:02d}-{d:02d}"


def _bestell_basis(con):
    """True, wenn Bestelltabellen vorhanden sind."""
    return _table_exists(con, "tbl_bestellungen") and _table_exists(con, "tbl_bestellpositionen")


def kunde_umsatz(kundennummer: str, monate: int | None = None) -> dict:
    """Umsatz (apu*menge*(1-rabatt)) eines Kunden im Zeitraum. Storno ausgenommen."""
    out = {"umsatz": 0.0, "bestellungen": 0, "positionen": 0, "von": None, "bis": None}
    if not kundennummer:
        return out
    con = _db()
    try:
        if not _bestell_basis(con):
            return out
        cut = _zeitraum_cutoff(monate)
        where = "b.kundennummer=? AND COALESCE(b.status,'')<>'storniert'"
        args = [str(kundennummer)]
        if cut:
            where += " AND b.datum>=?"
            args.append(cut)
        row = con.execute(
            f"SELECT COALESCE(SUM(p.apu*p.menge*(1-COALESCE(p.rabatt_prozent,0))),0) umsatz, "
            f"COUNT(DISTINCT b.id) best, COUNT(p.id) pos, MIN(b.datum) von, MAX(b.datum) bis "
            f"FROM tbl_bestellungen b JOIN tbl_bestellpositionen p ON p.bestell_id=b.id "
            f"WHERE {where}", args).fetchone()
        out.update(umsatz=row["umsatz"] or 0.0, bestellungen=row["best"] or 0,
                   positionen=row["pos"] or 0, von=row["von"], bis=row["bis"])
    finally:
        con.close()
    return out


def kunde_top_artikel(kundennummer: str, monate: int | None = None,
                      limit: int | None = None) -> list[dict]:
    """Gekaufte Artikel eines Kunden, nach Umsatz. limit=None -> alle."""
    if not kundennummer:
        return []
    con = _db()
    try:
        if not _bestell_basis(con):
            return []
        cut = _zeitraum_cutoff(monate)
        where = "b.kundennummer=? AND COALESCE(b.status,'')<>'storniert'"
        args: list = [str(kundennummer)]
        if cut:
            where += " AND b.datum>=?"
            args.append(cut)
        sql = (
            "SELECT p.pzn, MAX(p.artikelname) name, SUM(p.menge) menge, "
            "SUM(p.apu*p.menge*(1-COALESCE(p.rabatt_prozent,0))) umsatz, "
            "COUNT(DISTINCT b.id) bestellungen "
            "FROM tbl_bestellungen b JOIN tbl_bestellpositionen p ON p.bestell_id=b.id "
            f"WHERE {where} GROUP BY p.pzn ORDER BY umsatz DESC")
        if limit:
            sql += " LIMIT ?"
            args.append(limit)
        return [dict(r) for r in con.execute(sql, args).fetchall()]
    finally:
        con.close()


def kunde_vorbestellungen(kundennummer: str) -> list[dict]:
    """Offene Vorbestellungen eines Kunden (bestellpositionen.bestellart='Vorbestellung')."""
    if not kundennummer:
        return []
    con = _db()
    try:
        if not _bestell_basis(con):
            return []
        return [dict(r) for r in con.execute(
            "SELECT b.datum, b.status, p.pzn, p.artikelname, p.menge, p.apu, "
            "p.liefertermin, p.lieferzeit "
            "FROM tbl_bestellungen b JOIN tbl_bestellpositionen p ON p.bestell_id=b.id "
            "WHERE b.kundennummer=? AND p.bestellart='Vorbestellung' "
            "AND COALESCE(b.status,'')<>'storniert' "
            "ORDER BY b.datum DESC", (str(kundennummer),)).fetchall()]
    finally:
        con.close()


# ── Wiederverwendbare UI-Bausteine ───────────────────────────────────────────
class Card(tk.Frame):
    """Weisse Karte mit dezentem Rahmen."""
    def __init__(self, master, padding=20, **kw):
        super().__init__(master, bg=CARD, highlightbackground=BORDER,
                         highlightthickness=1, bd=0, **kw)
        self.inner = tk.Frame(self, bg=CARD)
        self.inner.pack(fill="both", expand=True, padx=padding, pady=padding)


class PillButton(tk.Label):
    """Flacher, moderner Button mit Hover-Effekt."""
    def __init__(self, master, text, command, kind="primary", **kw):
        colors = {
            "primary": (PRIMARY, "#0D5596", "#FFFFFF"),
            "accent": (ACCENT, "#2A97DA", "#FFFFFF"),
            "success": (SUCCESS, "#159247", "#FFFFFF"),
            "ghost": (CARD, "#EEF3F8", PRIMARY),
        }
        self._base, self._hover, fg = colors.get(kind, colors["primary"])
        super().__init__(master, text=text, bg=self._base, fg=fg,
                         font=(FONT, 11, "bold"), padx=22, pady=11, cursor="hand2", **kw)
        self._command = command
        self._enabled = True
        if kind == "ghost":
            self.config(highlightbackground=BORDER, highlightthickness=1)
        self.bind("<Button-1>", self._click)
        self.bind("<Enter>", lambda e: self._enabled and self.config(bg=self._hover))
        self.bind("<Leave>", lambda e: self._enabled and self.config(bg=self._base))

    def _click(self, _e):
        if self._enabled and self._command:
            self._command()

    def set_enabled(self, on: bool):
        self._enabled = on
        self.config(bg=self._base if on else "#C3CCD6",
                    fg="#FFFFFF" if on else "#8A97A5",
                    cursor="hand2" if on else "arrow")


# ── Hauptfenster ─────────────────────────────────────────────────────────────
class TestOberflaeche(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("NMGone – Testoberfläche (Vorschau)")
        self.geometry("1180x760")
        self.minsize(1040, 680)
        self.configure(bg=BG)
        self._busy = False
        self._selected_file: Path | None = None
        self._start_ts = 0.0
        self._erwartete_zeilen = 0

        self._assets = ensure_help_images(Path(__file__).resolve().parent.parent / "testui_assets")
        self._imgcache: dict[str, tk.PhotoImage] = {}

        self._build_sidebar()
        self._build_content()
        self.show_page("auswertung")

    # ----- Layout -----------------------------------------------------------
    def _build_sidebar(self):
        bar = tk.Frame(self, bg=SIDEBAR, width=248)
        bar.pack(side="left", fill="y")
        bar.pack_propagate(False)

        head = tk.Frame(bar, bg=SIDEBAR)
        head.pack(fill="x", pady=(26, 8), padx=22)
        tk.Label(head, text="NMGone", bg=SIDEBAR, fg="#FFFFFF",
                 font=(FONT, 21, "bold")).pack(anchor="w")
        tk.Label(head, text="Testoberfläche · Vorschau", bg=SIDEBAR, fg="#8FB4D6",
                 font=(FONT, 10)).pack(anchor="w")

        tk.Frame(bar, bg="#16395C", height=1).pack(fill="x", padx=18, pady=(14, 10))

        self._nav_buttons: dict[str, tk.Label] = {}
        items = [
            ("auswertung", "📊", "Neue Auswertung"),
            ("kunden", "👥", "Kunden & ABC"),
            ("auskunft", "🔎", "PZN-Schnellauskunft"),
            ("markt", "📦", "Markt-Insights"),
            ("dashboard", "📈", "Datenbank-Dashboard"),
            ("verlauf", "🗂️", "Auswertungs-Verlauf"),
            ("hilfe", "❓", "Hilfe & Anleitung"),
            ("info", "ℹ️", "Über diese Vorschau"),
        ]
        for key, icon, label in items:
            b = tk.Label(bar, text=f"   {icon}   {label}", bg=SIDEBAR, fg="#D7E6F4",
                         font=(FONT, 12), anchor="w", padx=14, pady=12, cursor="hand2")
            b.pack(fill="x", padx=12, pady=2)
            b.bind("<Button-1>", lambda e, k=key: self.show_page(k))
            b.bind("<Enter>", lambda e, bb=b, k=key: bb.config(bg=SIDEBAR_ACTIVE) if self._active != k else None)
            b.bind("<Leave>", lambda e, bb=b, k=key: bb.config(bg=SIDEBAR if self._active != k else SIDEBAR_ACTIVE))
            self._nav_buttons[key] = b

        foot = tk.Label(bar, text="Original bleibt unberührt.\nNur eine Vorschau.",
                        bg=SIDEBAR, fg="#6E92B4", font=(FONT, 9), justify="left")
        foot.pack(side="bottom", anchor="w", padx=24, pady=18)
        self._active = "auswertung"

    def _build_content(self):
        self.content = tk.Frame(self, bg=BG)
        self.content.pack(side="left", fill="both", expand=True)
        self._pages: dict[str, tk.Frame] = {}

    def show_page(self, key: str):
        if self._busy and key != "auswertung":
            # Waehrend einer laufenden Auswertung Navigation erlauben, aber Seite neu bauen ist ok.
            pass
        self._active = key
        for k, b in self._nav_buttons.items():
            b.config(bg=SIDEBAR_ACTIVE if k == key else SIDEBAR,
                     fg="#FFFFFF" if k == key else "#D7E6F4")
        for child in self.content.winfo_children():
            child.destroy()
        builder = {"auswertung": self._page_auswertung,
                   "kunden": self._page_kunden,
                   "auskunft": self._page_auskunft,
                   "markt": self._page_markt,
                   "dashboard": self._page_dashboard,
                   "verlauf": self._page_verlauf,
                   "hilfe": self._page_hilfe,
                   "info": self._page_info}.get(key, self._page_auswertung)
        builder()

    def _page_header(self, parent, title, subtitle):
        head = tk.Frame(parent, bg=BG)
        head.pack(fill="x", padx=36, pady=(30, 10))
        tk.Label(head, text=title, bg=BG, fg=INK, font=(FONT, 26, "bold")).pack(anchor="w")
        tk.Label(head, text=subtitle, bg=BG, fg=MUTED, font=(FONT, 13)).pack(anchor="w", pady=(2, 0))

    def _img(self, name: str, max_w: int | None = None) -> tk.PhotoImage | None:
        p = self._assets.get(name)
        if not p or not Path(p).exists():
            return None
        if name in self._imgcache:
            return self._imgcache[name]
        try:
            img = tk.PhotoImage(file=str(p))
            if max_w and img.width() > max_w:
                factor = max(1, round(img.width() / max_w))
                img = img.subsample(factor, factor)
            self._imgcache[name] = img
            return img
        except Exception:
            return None

    # ----- Seite: Neue Auswertung ------------------------------------------
    def _page_auswertung(self):
        page = tk.Frame(self.content, bg=BG)
        page.pack(fill="both", expand=True)
        self._page_header(page, "Neue Auswertung",
                          "Datei wählen, starten – der Fortschritt ist jederzeit sichtbar.")

        body = tk.Frame(page, bg=BG)
        body.pack(fill="both", expand=True, padx=36, pady=(6, 28))

        # Schritt 1: Datei
        c1 = Card(body)
        c1.pack(fill="x", pady=(0, 16))
        tk.Label(c1.inner, text="1 · Datei auswählen", bg=CARD, fg=PRIMARY,
                 font=(FONT, 14, "bold")).pack(anchor="w")
        row = tk.Frame(c1.inner, bg=CARD)
        row.pack(fill="x", pady=(12, 0))
        self._file_var = tk.StringVar(value="Noch keine Datei gewählt")
        drop = tk.Frame(row, bg="#F0F4F8", highlightbackground=BORDER, highlightthickness=1)
        drop.pack(side="left", fill="x", expand=True, ipady=14)
        tk.Label(drop, textvariable=self._file_var, bg="#F0F4F8", fg=INK,
                 font=(FONT, 11)).pack(side="left", padx=16)
        self._file_meta = tk.Label(drop, text="", bg="#F0F4F8", fg=MUTED, font=(FONT, 10))
        self._file_meta.pack(side="right", padx=16)
        PillButton(row, "Durchsuchen …", self._choose_file, kind="ghost").pack(side="left", padx=(12, 0))
        tk.Label(c1.inner, text="Unterstützt: .xlsx · .xls · .xlsm · .csv · .txt   (alte .xls werden automatisch konvertiert)",
                 bg=CARD, fg=MUTED, font=(FONT, 10)).pack(anchor="w", pady=(10, 0))

        # Schritt 2: Name + Start
        c2 = Card(body)
        c2.pack(fill="x", pady=(0, 16))
        tk.Label(c2.inner, text="2 · Bezeichnung & Start", bg=CARD, fg=PRIMARY,
                 font=(FONT, 14, "bold")).pack(anchor="w")
        row2 = tk.Frame(c2.inner, bg=CARD)
        row2.pack(fill="x", pady=(12, 0))
        tk.Label(row2, text="Name der Auswertung", bg=CARD, fg=MUTED, font=(FONT, 10)).pack(anchor="w")
        self._name_var = tk.StringVar(value="Testauswertung")
        ent = tk.Entry(row2, textvariable=self._name_var, font=(FONT, 12),
                       relief="flat", highlightbackground=BORDER, highlightthickness=1)
        ent.pack(side="left", fill="x", expand=True, ipady=7, pady=(4, 0))
        self._start_btn = PillButton(row2, "▶  Auswertung starten", self._start_auswertung, kind="success")
        self._start_btn.pack(side="left", padx=(12, 0), pady=(4, 0))

        # Schritt 3: Fortschritt (anfangs versteckt)
        self._progress_card = Card(body)
        self._build_progress_widgets(self._progress_card.inner)

    def _build_progress_widgets(self, parent):
        self._prog_title = tk.Label(parent, text="Auswertung läuft …", bg=CARD, fg=INK,
                                    font=(FONT, 15, "bold"))
        self._prog_title.pack(anchor="w")
        toprow = tk.Frame(parent, bg=CARD)
        toprow.pack(fill="x", pady=(2, 12))
        self._prog_sub = tk.Label(toprow, text="", bg=CARD, fg=MUTED, font=(FONT, 11))
        self._prog_sub.pack(side="left")
        self._prog_timer = tk.Label(toprow, text="00:00", bg=CARD, fg=ACCENT, font=(FONT, 13, "bold"))
        self._prog_timer.pack(side="right")

        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("NMG.Horizontal.TProgressbar", troughcolor="#E8EDF2",
                        background=ACCENT, thickness=20, borderwidth=0)
        self._pbar = ttk.Progressbar(parent, style="NMG.Horizontal.TProgressbar",
                                     mode="indeterminate", length=100)
        self._pbar.pack(fill="x")

        self._stage_frame = tk.Frame(parent, bg=CARD)
        self._stage_frame.pack(fill="x", pady=(14, 0))
        self._stage_labels: list[tk.Label] = []
        stages = ["Datei eingelesen", "Stammdaten geladen", "Positionen abgleichen", "Excel schreiben", "Fertig"]
        for i, s in enumerate(stages):
            lab = tk.Label(self._stage_frame, text=f"○  {s}", bg=CARD, fg=MUTED, font=(FONT, 11))
            lab.grid(row=0, column=i, padx=(0, 26), sticky="w")
            self._stage_labels.append(lab)

        self._result_frame = tk.Frame(parent, bg=CARD)
        self._result_frame.pack(fill="x", pady=(16, 0))

    def _set_stage(self, idx: int, state: str):
        marks = {"done": ("✓", SUCCESS), "active": ("●", ACCENT), "todo": ("○", MUTED)}
        labels = ["Datei eingelesen", "Stammdaten geladen", "Positionen abgleichen", "Excel schreiben", "Fertig"]
        for i, lab in enumerate(self._stage_labels):
            if i < idx:
                m, c = marks["done"]
            elif i == idx:
                m, c = marks[state] if state in marks else marks["active"]
            else:
                m, c = marks["todo"]
            lab.config(text=f"{m}  {labels[i]}", fg=c,
                       font=(FONT, 11, "bold" if i == idx else "normal"))

    # ----- Datei-Auswahl ----------------------------------------------------
    def _choose_file(self):
        if self._busy:
            return
        path = filedialog.askopenfilename(
            title="Rohdatei wählen",
            filetypes=[("Alle unterstützten", "*.xlsx *.xlsm *.xls *.csv *.txt"),
                       ("Excel neu", "*.xlsx *.xlsm"), ("Excel alt", "*.xls"),
                       ("CSV/Text", "*.csv *.txt"), ("Alle Dateien", "*.*")])
        if not path:
            return
        p = Path(path)
        if p.suffix.lower() not in SUPPORTED:
            messagebox.showwarning("Format", _T("'{p0}' wird nicht unterstützt.\nErlaubt: xlsx, xls, xlsm, csv, txt.", p0=p.suffix))
            return
        self._selected_file = p
        self._file_var.set(p.name)
        n = zeilen_schaetzen(p)
        self._erwartete_zeilen = n
        tag = "alte Excel – wird konvertiert" if p.suffix.lower() == ".xls" else p.suffix.lower().lstrip(".").upper()
        self._file_meta.config(text=_T('{p0} Zeilen · {p1}', p0=n, p1=tag))
        if not self._name_var.get().strip() or self._name_var.get() == "Testauswertung":
            self._name_var.set(p.stem[:40])

    # ----- Start / Hintergrund ---------------------------------------------
    def _start_auswertung(self):
        if self._busy:
            return
        if not self._selected_file:
            messagebox.showinfo("Datei fehlt", "Bitte zuerst eine Datei auswählen.")
            return
        name = self._name_var.get().strip() or "Testauswertung"

        self._busy = True
        self._start_btn.set_enabled(False)
        self._progress_card.pack(fill="x", pady=(0, 8))
        for f in self._result_frame.winfo_children():
            f.destroy()
        self._prog_title.config(text="Auswertung läuft …", fg=INK)
        self._prog_sub.config(text=_T('Verarbeite {p0}', p0=self._selected_file.name))
        self._pbar.config(mode="indeterminate")
        self._pbar.start(12)
        self._start_ts = time.time()
        self._set_stage(0, "active")
        self._tick_timer()
        self._poll_positions()

        threading.Thread(target=self._worker, args=(self._selected_file, name), daemon=True).start()

    def _worker(self, file: Path, name: str):
        """Laeuft im Hintergrund: ggf. .xls konvertieren, dann echte Engine."""
        from .exporter import create_vorlage_export
        try:
            src = file
            if file.suffix.lower() == ".xls":
                self.after(0, lambda: self._prog_sub.config(text="Alte .xls-Datei wird konvertiert …"))
                src = konvertiere_xls_zu_xlsx(file)
            self.after(0, lambda: self._set_stage(1, "active"))
            out = create_vorlage_export(str(src), name)
            self.after(0, lambda: self._finish_ok(Path(out)))
        except Exception as exc:
            self.after(0, lambda e=exc: self._finish_err(e))

    def _tick_timer(self):
        if not self._busy:
            return
        el = int(time.time() - self._start_ts)
        self._prog_timer.config(text=f"{el // 60:02d}:{el % 60:02d}")
        self.after(500, self._tick_timer)

    def _latest_position_count(self) -> int:
        try:
            with sqlite3.connect(DB_PATH) as con:
                row = con.execute(
                    "SELECT COUNT(*) FROM tbl_auswertungspositionen "
                    "WHERE auswertung_id = (SELECT MAX(id) FROM tbl_auswertungen)"
                ).fetchone()
                return int(row[0]) if row else 0
        except Exception:
            return 0

    def _poll_positions(self):
        """Echte Rueckmeldung: zeigt, wie viele Positionen schon in der DB sind."""
        if not self._busy:
            return
        cnt = self._latest_position_count()
        if cnt > 0:
            self._set_stage(2, "active")
            if self._erwartete_zeilen:
                pct = min(99, int(cnt / max(1, self._erwartete_zeilen) * 100))
                self._prog_sub.config(text=_T('{p0} von ~{p1} Positionen abgeglichen ({p2}%)', p0=cnt, p1=self._erwartete_zeilen, p2=pct))
            else:
                self._prog_sub.config(text=_T('{p0} Positionen abgeglichen …', p0=cnt))
        self.after(400, self._poll_positions)

    def _finish_ok(self, out: Path):
        self._busy = False
        self._pbar.stop()
        self._pbar.config(mode="determinate", value=100)
        self._set_stage(4, "done")
        self._prog_title.config(text="✓  Auswertung fertig", fg=SUCCESS)
        self._prog_sub.config(text="Die Excel-Datei wurde erstellt.")
        self._start_btn.set_enabled(True)
        for f in self._result_frame.winfo_children():
            f.destroy()
        box = tk.Frame(self._result_frame, bg="#EAF6EF", highlightbackground="#BfE3CD", highlightthickness=1)
        box.pack(fill="x")
        tk.Label(box, text=f"📄  {out.name}", bg="#EAF6EF", fg=INK, font=(FONT, 11, "bold")).pack(side="left", padx=16, pady=12)
        PillButton(box, "Ordner öffnen", lambda: self._open_folder(out.parent), kind="success").pack(side="right", padx=12, pady=8)
        PillButton(box, "Excel öffnen", lambda: self._open_file(out), kind="ghost").pack(side="right", pady=8)

    def _finish_err(self, exc: Exception):
        self._busy = False
        self._pbar.stop()
        self._set_stage(2, "todo")
        self._prog_title.config(text="✗  Auswertung fehlgeschlagen", fg=DANGER)
        self._prog_sub.config(text=str(exc)[:120])
        self._start_btn.set_enabled(True)
        for f in self._result_frame.winfo_children():
            f.destroy()
        msg = str(exc)
        hint = ""
        if "nicht erkannt" in msg.lower() or "format" in msg.lower():
            hint = "\n\nTipp: Im Original gibt es den Rohdaten-Formatassistenten für unbekannte Spalten."
        tk.Label(self._result_frame, text=msg[:300] + hint, bg=CARD, fg=DANGER,
                 font=(FONT, 10), justify="left", wraplength=820).pack(anchor="w")

    # ----- gemeinsame Helfer ------------------------------------------------
    def _open_file(self, path: Path):
        try:
            import os
            os.startfile(str(path))  # type: ignore[attr-defined]
        except Exception as exc:
            messagebox.showwarning("Öffnen", _T('Datei konnte nicht geöffnet werden:\n{p0}', p0=exc))

    def _open_folder(self, path: Path):
        try:
            import os
            os.startfile(str(path))  # type: ignore[attr-defined]
        except Exception as exc:
            messagebox.showwarning("Öffnen", _T('Ordner konnte nicht geöffnet werden:\n{p0}', p0=exc))

    def _scroll_page(self, title, subtitle):
        """Baut eine scrollbare Seite und liefert das innere Frame zurueck."""
        page = tk.Frame(self.content, bg=BG)
        page.pack(fill="both", expand=True)
        self._page_header(page, title, subtitle)
        canvas = tk.Canvas(page, bg=BG, highlightthickness=0)
        scroll = ttk.Scrollbar(page, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=BG)
        win = canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win, width=e.width))
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side="left", fill="both", expand=True, padx=(36, 0), pady=(0, 20))
        scroll.pack(side="right", fill="y")
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-e.delta / 120), "units"))
        return inner

    # ----- Seite: Kunden & ABC ----------------------------------------------
    ABC_COLORS = {"A": "#11823B", "B": "#C88200", "C": "#B3261E", None: "#8A97A5"}

    def _page_kunden(self):
        page = tk.Frame(self.content, bg=BG)
        page.pack(fill="both", expand=True)
        self._page_header(page, "Kunden & ABC-Analyse",
                          "Apotheken-Kunden verwalten, nach Wert einstufen und auf der Karte finden.")
        tabbar = tk.Frame(page, bg=BG)
        tabbar.pack(fill="x", padx=36, pady=(0, 4))
        self._kunden_tabbtns = {}
        for key, label in [("liste", "📋  Kundenliste"), ("abc", "🏆  ABC-Analyse"),
                           ("karte", "🗺️  Landkarte")]:
            b = tk.Label(tabbar, text=label, bg=BG, fg=PRIMARY, font=(FONT, 12, "bold"),
                         padx=16, pady=8, cursor="hand2")
            b.pack(side="left", padx=(0, 6))
            b.bind("<Button-1>", lambda e, k=key: self._kunden_tab(k))
            self._kunden_tabbtns[key] = b
        self._kunden_body = tk.Frame(page, bg=BG)
        self._kunden_body.pack(fill="both", expand=True, padx=36, pady=(6, 20))
        self._kunden_tab(getattr(self, "_kunden_active", "liste"))

    def _kunden_tab(self, key):
        self._kunden_active = key
        for k, b in self._kunden_tabbtns.items():
            active = (k == key)
            b.config(fg="#FFFFFF" if active else PRIMARY,
                     bg=PRIMARY if active else BG)
        for c in self._kunden_body.winfo_children():
            c.destroy()
        {"liste": self._kunden_liste, "abc": self._kunden_abc,
         "karte": self._kunden_karte}.get(key, self._kunden_liste)()

    def _load_kunden(self):
        try:
            return kunden_master()
        except Exception as exc:
            tk.Label(self._kunden_body, text=_T('Kunden konnten nicht geladen werden:\n{p0}', p0=exc),
                     bg=BG, fg=DANGER, font=(FONT, 12)).pack(anchor="w", pady=20)
            return None

    # --- Tab: Kundenliste ---
    def _kunden_liste(self):
        kunden = self._load_kunden()
        if kunden is None:
            return
        if not kunden:
            tk.Label(self._kunden_body, text="Noch keine Kunden im Kunden-Center hinterlegt.",
                     bg=BG, fg=MUTED, font=(FONT, 12)).pack(anchor="w", pady=20)
            return
        cont = tk.Frame(self._kunden_body, bg=BG)
        cont.pack(fill="both", expand=True)
        listcol = tk.Frame(cont, bg=BG, width=520)
        listcol.pack(side="left", fill="both", expand=True)
        self._kunden_detailcol = tk.Frame(cont, bg=BG, width=380)
        self._kunden_detailcol.pack(side="left", fill="both", padx=(16, 0))
        self._kunden_detailcol.pack_propagate(False)
        tk.Label(self._kunden_detailcol, text="← Kunde wählen, um Kontaktdaten zu sehen",
                 bg=BG, fg=MUTED, font=(FONT, 11)).pack(anchor="w", pady=20)

        head = tk.Frame(listcol, bg=BG)
        head.pack(fill="x", pady=(0, 4))
        for txt, w in [("ABC", 5), ("Kunde", 26), ("Ort", 18), ("Status", 10)]:
            tk.Label(head, text=txt, bg=BG, fg=MUTED, font=(FONT, 10, "bold"),
                     width=w, anchor="w").pack(side="left")
        for k in kunden:
            bg = CARD
            row = tk.Frame(listcol, bg=bg, highlightbackground=BORDER, highlightthickness=1, cursor="hand2")
            row.pack(fill="x", pady=2)
            klasse = k.get("abc")
            badge = tk.Label(row, text=f" {klasse or '–'} ", bg=self.ABC_COLORS.get(klasse),
                             fg="#FFFFFF", font=(FONT, 10, "bold"), width=3)
            badge.pack(side="left", padx=(8, 8), pady=8)
            tk.Label(row, text=(k.get("kundenname") or "—")[:26], bg=bg, fg=INK,
                     font=(FONT, 11, "bold"), width=24, anchor="w").pack(side="left")
            tk.Label(row, text=(k.get("ort") or "—")[:18], bg=bg, fg=MUTED,
                     font=(FONT, 10), width=18, anchor="w").pack(side="left")
            st = k.get("status") or "—"
            tk.Label(row, text=st, bg=bg, fg=(SUCCESS if st == "aktiv" else MUTED),
                     font=(FONT, 10), width=10, anchor="w").pack(side="left")
            for w in (row,) + tuple(row.winfo_children()):
                w.bind("<Button-1>", lambda e, kk=k: self._kunden_show_detail(kk))

    def _kunden_show_detail(self, kunde):
        col = self._kunden_detailcol
        for c in col.winfo_children():
            c.destroy()
        self._render_kunden_detail(col, kunde)

    def _render_kunden_detail(self, parent, k):
        """Kontakt-Steckbrief: alles, um sich beim Kunden zu melden."""
        card = Card(parent, padding=18)
        card.pack(fill="x")
        klasse = k.get("abc")
        top = tk.Frame(card.inner, bg=CARD)
        top.pack(fill="x")
        tk.Label(top, text=f" {klasse or '–'} ", bg=self.ABC_COLORS.get(klasse), fg="#FFFFFF",
                 font=(FONT, 12, "bold")).pack(side="left", padx=(0, 10))
        tk.Label(top, text=k.get("kundenname") or "—", bg=CARD, fg=INK,
                 font=(FONT, 15, "bold"), wraplength=300, justify="left").pack(side="left")
        if k.get("kundennummer"):
            tk.Label(card.inner, text=f"Kundennr. {k['kundennummer']}"
                     + (f"  ·  {k.get('kundentyp')}" if k.get("kundentyp") else ""),
                     bg=CARD, fg=MUTED, font=(FONT, 10)).pack(anchor="w", pady=(4, 0))

        def field(label, value, action=None, action_label=None):
            if not value:
                return
            r = tk.Frame(card.inner, bg=CARD)
            r.pack(fill="x", pady=3)
            tk.Label(r, text=label, bg=CARD, fg=MUTED, font=(FONT, 9), anchor="w").pack(anchor="w")
            line = tk.Frame(r, bg=CARD)
            line.pack(fill="x")
            tk.Label(line, text=str(value), bg=CARD, fg=INK, font=(FONT, 11),
                     anchor="w", justify="left", wraplength=240).pack(side="left")
            if action and action_label:
                PillButton(line, action_label, action, kind="ghost").pack(side="right")

        adresse = " ".join(str(x) for x in [k.get("strasse"), k.get("hausnummer")] if x).strip()
        ort_zeile = " ".join(str(x) for x in [k.get("plz"), k.get("ort")] if x).strip()
        voll_adr = (adresse + ", " + ort_zeile).strip(", ")
        inhaber = (k.get("inhaber") or " ".join(
            str(x) for x in [k.get("inhaber_vorname"), k.get("inhaber_zuname")] if x).strip())

        tk.Frame(card.inner, bg=BORDER, height=1).pack(fill="x", pady=(10, 8))
        field("Inhaber / Ansprechpartner", inhaber or k.get("ansprechpartner"))
        field("Telefon", k.get("telefon"),
              action=(lambda v=k.get("telefon"): self._copy_text(v, "Telefonnummer")) if k.get("telefon") else None,
              action_label="Kopieren" if k.get("telefon") else None)
        field("E-Mail", k.get("email"),
              action=(lambda v=k.get("email"): self._mailto(v)) if k.get("email") else None,
              action_label="Schreiben" if k.get("email") else None)
        field("Besteller", k.get("besteller_name"))
        field("Besteller-E-Mail", k.get("besteller_email"))
        field("Rechnungs-E-Mail", k.get("rechnungsemail"))
        field("Adresse", voll_adr or None,
              action=(lambda v=voll_adr: self._open_maps(v)) if voll_adr else None,
              action_label="Karte" if voll_adr else None)
        if k.get("abc_pot"):
            field("Rabatt-Potenzial (aus Auswertung)", self._eur(k["abc_pot"]))
        if k.get("notizen"):
            field("Notizen", k.get("notizen"))

        # Umsatz-Kurzinfo + Absprung in die Detailanalyse
        knr = k.get("kundennummer")
        if knr:
            try:
                u = kunde_umsatz(knr)
            except Exception:
                u = None
            if u and u["bestellungen"]:
                tk.Frame(card.inner, bg=BORDER, height=1).pack(fill="x", pady=(10, 8))
                ums = tk.Frame(card.inner, bg="#EFF5FB")
                ums.pack(fill="x")
                tk.Label(ums, text=_T('Umsatz gesamt: {p0}', p0=self._eur(u['umsatz'])), bg="#EFF5FB",
                         fg=PRIMARY, font=(FONT, 12, "bold")).pack(anchor="w", padx=12, pady=(8, 0))
                tk.Label(ums, text=_T('{p0} Bestellungen · {p1} bis {p2}', p0=u['bestellungen'], p1=u['von'], p2=u['bis']),
                         bg="#EFF5FB", fg=MUTED, font=(FONT, 10)).pack(anchor="w", padx=12, pady=(0, 8))
                PillButton(card.inner, "📊  Umsatz & Artikel ansehen",
                           lambda kk=k: self._open_umsatz_dialog(kk), kind="accent").pack(anchor="w", pady=(8, 0))
            else:
                tk.Frame(card.inner, bg=BORDER, height=1).pack(fill="x", pady=(10, 8))
                tk.Label(card.inner, text="Noch keine Bestellungen erfasst.", bg=CARD,
                         fg=MUTED, font=(FONT, 10)).pack(anchor="w")

    def _open_umsatz_dialog(self, kunde):
        """Eigenes Fenster: Umsatz mit Zeiträumen, Top-Artikel (Mehr→alle), Vorbestellungen."""
        knr = kunde.get("kundennummer")
        win = tk.Toplevel(self)
        win.title(f"Umsatz – {kunde.get('kundenname') or knr}")
        win.geometry("720x720")
        win.configure(bg=BG)
        win.transient(self)

        head = tk.Frame(win, bg=PRIMARY)
        head.pack(fill="x")
        tk.Label(head, text=kunde.get("kundenname") or "—", bg=PRIMARY, fg="#FFFFFF",
                 font=(FONT, 17, "bold")).pack(anchor="w", padx=20, pady=(14, 2))
        tk.Label(head, text=_T('Kundennr. {p0} · {p1}', p0=knr or '—', p1=kunde.get('ort') or ''), bg=PRIMARY,
                 fg="#CFE2F3", font=(FONT, 10)).pack(anchor="w", padx=20, pady=(0, 14))

        # Zeitraum-Umschalter
        self._um_zeitraum = tk.IntVar(value=0)  # 0 = alles
        bar = tk.Frame(win, bg=BG)
        bar.pack(fill="x", padx=20, pady=(12, 4))
        tk.Label(bar, text="Zeitraum:", bg=BG, fg=MUTED, font=(FONT, 10)).pack(side="left", padx=(0, 8))
        self._um_zr_buttons = {}
        for label, monate in [("Alles", 0), ("12 Monate", 12), ("6 Monate", 6), ("3 Monate", 3)]:
            b = tk.Label(bar, text=label, bg=CARD, fg=PRIMARY, font=(FONT, 10, "bold"),
                         padx=12, pady=6, cursor="hand2", highlightbackground=BORDER, highlightthickness=1)
            b.pack(side="left", padx=(0, 6))
            b.bind("<Button-1>", lambda e, m=monate: self._um_set_zeitraum(win, knr, m))
            self._um_zr_buttons[monate] = b

        self._um_body = tk.Frame(win, bg=BG)
        self._um_body.pack(fill="both", expand=True, padx=20, pady=(4, 16))
        self._um_show_alle = False
        self._um_set_zeitraum(win, knr, 0)

    def _um_set_zeitraum(self, win, knr, monate):
        self._um_zeitraum.set(monate)
        for m, b in self._um_zr_buttons.items():
            active = (m == monate)
            b.config(bg=PRIMARY if active else CARD, fg="#FFFFFF" if active else PRIMARY)
        self._um_render(win, knr, monate)

    def _um_render(self, win, knr, monate):
        for c in self._um_body.winfo_children():
            c.destroy()
        mz = monate or None
        u = kunde_umsatz(knr, mz)
        # KPI-Kacheln
        grid = tk.Frame(self._um_body, bg=BG)
        grid.pack(fill="x", pady=(0, 12))
        kacheln = [("Umsatz", self._eur(u["umsatz"]), PRIMARY),
                   ("Bestellungen", str(u["bestellungen"]), ACCENT),
                   ("Artikelpositionen", str(u["positionen"]), SUCCESS)]
        for i, (lab, val, col) in enumerate(kacheln):
            t = tk.Frame(grid, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
            t.grid(row=0, column=i, sticky="nsew", padx=(0 if i == 0 else 10, 0))
            grid.columnconfigure(i, weight=1)
            tk.Label(t, text=lab, bg=CARD, fg=MUTED, font=(FONT, 10)).pack(anchor="w", padx=14, pady=(12, 0))
            tk.Label(t, text=val, bg=CARD, fg=col, font=(FONT, 19, "bold")).pack(anchor="w", padx=14, pady=(2, 12))

        # Top-Artikel
        alle = kunde_top_artikel(knr, mz, limit=None)
        anzeige = alle if self._um_show_alle else alle[:10]
        titel = tk.Frame(self._um_body, bg=BG)
        titel.pack(fill="x")
        tk.Label(titel, text=("Alle Artikel" if self._um_show_alle else "Top 10 Artikel")
                 + f"  ({len(alle)} insgesamt)", bg=BG, fg=INK, font=(FONT, 13, "bold")).pack(side="left")
        if len(alle) > 10:
            def _toggle():
                self._um_show_alle = not self._um_show_alle
                self._um_render(win, knr, monate)
            PillButton(titel, "Top 10 anzeigen" if self._um_show_alle else f"Alle {len(alle)} anzeigen",
                       _toggle, kind="ghost").pack(side="right")

        if not alle:
            tk.Label(self._um_body, text="Keine Artikel im Zeitraum.", bg=BG, fg=MUTED,
                     font=(FONT, 11)).pack(anchor="w", pady=12)
        else:
            hdr = tk.Frame(self._um_body, bg=BG)
            hdr.pack(fill="x", pady=(8, 2))
            for txt, w in [("PZN", 10), ("Artikel", 34), ("Menge", 7), ("Best.", 6), ("Umsatz", 13)]:
                tk.Label(hdr, text=txt, bg=BG, fg=MUTED, font=(FONT, 9, "bold"),
                         width=w, anchor="w").pack(side="left")
            cv = tk.Canvas(self._um_body, bg=BG, highlightthickness=0, height=240)
            sc = ttk.Scrollbar(self._um_body, orient="vertical", command=cv.yview)
            inner = tk.Frame(cv, bg=BG)
            w_ = cv.create_window((0, 0), window=inner, anchor="nw")
            inner.bind("<Configure>", lambda e: cv.configure(scrollregion=cv.bbox("all")))
            cv.bind("<Configure>", lambda e: cv.itemconfig(w_, width=e.width))
            cv.configure(yscrollcommand=sc.set)
            cv.pack(side="top", fill="both", expand=True)
            sc.pack(side="right", fill="y")
            for i, a in enumerate(anzeige):
                bg = CARD if i % 2 == 0 else "#F7FAFC"
                row = tk.Frame(inner, bg=bg)
                row.pack(fill="x", pady=1)
                tk.Label(row, text=a["pzn"], bg=bg, fg=INK, font=(FONT, 9), width=10, anchor="w").pack(side="left")
                tk.Label(row, text=str(a["name"] or "")[:36], bg=bg, fg=INK, font=(FONT, 10),
                         width=34, anchor="w").pack(side="left")
                tk.Label(row, text=f"{a['menge']:.0f}", bg=bg, fg=INK, font=(FONT, 9), width=7, anchor="w").pack(side="left")
                tk.Label(row, text=str(a["bestellungen"]), bg=bg, fg=MUTED, font=(FONT, 9), width=6, anchor="w").pack(side="left")
                tk.Label(row, text=self._eur(a["umsatz"]), bg=bg, fg=PRIMARY, font=(FONT, 9, "bold"),
                         width=13, anchor="w").pack(side="left")

        # Vorbestellungen (zeitraumunabhaengig: offene)
        vb = kunde_vorbestellungen(knr)
        vcard = tk.Frame(self._um_body, bg=BG)
        vcard.pack(fill="x", pady=(14, 0))
        tk.Label(vcard, text=_T('🕓  Offene Vorbestellungen ({p0})', p0=len(vb)), bg=BG, fg=WARNING,
                 font=(FONT, 13, "bold")).pack(anchor="w", pady=(0, 4))
        if not vb:
            tk.Label(vcard, text="Keine offenen Vorbestellungen.", bg=BG, fg=MUTED,
                     font=(FONT, 10)).pack(anchor="w")
        else:
            for v in vb[:8]:
                txt = f"   {v.get('datum') or ''}  ·  {str(v.get('artikelname') or '')[:40]}  ·  {v.get('menge') or 0} St"
                if v.get("liefertermin"):
                    txt += f"  ·  Liefertermin {v['liefertermin']}"
                tk.Label(vcard, text=txt, bg=BG, fg=INK, font=(FONT, 10)).pack(anchor="w")

    # --- Tab: ABC-Analyse ---
    def _kunden_abc(self):
        try:
            ranking = abc_ranking()
        except Exception as exc:
            tk.Label(self._kunden_body, text=_T('ABC-Analyse fehlgeschlagen:\n{p0}', p0=exc),
                     bg=BG, fg=DANGER, font=(FONT, 12)).pack(anchor="w", pady=20)
            return
        if not ranking:
            tk.Label(self._kunden_body, text="Keine Auswertungen für eine ABC-Analyse vorhanden.",
                     bg=BG, fg=MUTED, font=(FONT, 12)).pack(anchor="w", pady=20)
            return
        # Zusammenfassung
        summe = sum(a["pot"] for a in ranking)
        zaehler = {"A": 0, "B": 0, "C": 0}
        for a in ranking:
            zaehler[a["klasse"]] += 1
        tk.Label(self._kunden_body,
                 text="Apotheken nach Rabatt-Potenzial. A = die wertvollen 80 %, B = nächste 15 %, "
                      "C = restliche 5 % des Gesamtpotenzials (Pareto).",
                 bg=BG, fg=MUTED, font=(FONT, 11), justify="left", wraplength=860).pack(anchor="w", pady=(0, 10))
        sumrow = tk.Frame(self._kunden_body, bg=BG)
        sumrow.pack(fill="x", pady=(0, 12))
        for kl in ("A", "B", "C"):
            t = tk.Frame(sumrow, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
            t.pack(side="left", padx=(0 if kl == "A" else 10, 0))
            tk.Label(t, text=_T('  Klasse {p0}  ', p0=kl), bg=self.ABC_COLORS[kl], fg="#FFFFFF",
                     font=(FONT, 11, "bold")).pack(anchor="w")
            tk.Label(t, text=_T('{p0} Apotheken', p0=zaehler[kl]), bg=CARD, fg=INK,
                     font=(FONT, 13, "bold")).pack(anchor="w", padx=14, pady=(6, 10))
        tk.Label(self._kunden_body, text=_T('Gesamtpotenzial: {p0}', p0=self._eur(summe)),
                 bg=BG, fg=PRIMARY, font=(FONT, 12, "bold")).pack(anchor="w", pady=(0, 10))

        head = tk.Frame(self._kunden_body, bg=BG)
        head.pack(fill="x", padx=(0, 8))
        for txt, w in [("ABC", 5), ("Apotheke", 32), ("Absatz", 9), ("NMG", 6),
                       ("Rabatt-Potenzial", 16), ("kum. %", 8)]:
            tk.Label(head, text=txt, bg=BG, fg=MUTED, font=(FONT, 10, "bold"),
                     width=w, anchor="w").pack(side="left")
        canvas = tk.Canvas(self._kunden_body, bg=BG, highlightthickness=0)
        scroll = ttk.Scrollbar(self._kunden_body, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=BG)
        win = canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win, width=e.width))
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-e.delta / 120), "units"))
        for i, a in enumerate(ranking):
            bg = CARD if i % 2 == 0 else "#F7FAFC"
            row = tk.Frame(inner, bg=bg, highlightbackground=BORDER, highlightthickness=1)
            row.pack(fill="x", pady=1, padx=(0, 8))
            tk.Label(row, text=f" {a['klasse']} ", bg=self.ABC_COLORS[a["klasse"]], fg="#FFFFFF",
                     font=(FONT, 10, "bold"), width=3).pack(side="left", padx=(6, 6), pady=6)
            tk.Label(row, text=str(a["apotheke"])[:34], bg=bg, fg=INK, font=(FONT, 11, "bold"),
                     width=30, anchor="w").pack(side="left")
            tk.Label(row, text=f"{a['absatz']:.0f}", bg=bg, fg=INK, font=(FONT, 10),
                     width=9, anchor="w").pack(side="left")
            tk.Label(row, text=str(a["treffer"]), bg=bg, fg=SUCCESS, font=(FONT, 10),
                     width=6, anchor="w").pack(side="left")
            tk.Label(row, text=self._eur(a["pot"]), bg=bg, fg=PRIMARY, font=(FONT, 10, "bold"),
                     width=16, anchor="w").pack(side="left")
            tk.Label(row, text=f"{a['anteil_kum']*100:.0f}%", bg=bg, fg=MUTED, font=(FONT, 10),
                     width=8, anchor="w").pack(side="left")

    # --- Tab: Landkarte ---
    def _kunden_karte(self):
        kunden = self._load_kunden()
        if kunden is None:
            return
        platzierbar = [k for k in kunden if k.get("latlon")]
        cont = tk.Frame(self._kunden_body, bg=BG)
        cont.pack(fill="both", expand=True)
        mapcol = tk.Frame(cont, bg=BG)
        mapcol.pack(side="left", fill="both", expand=True)
        self._karte_detailcol = tk.Frame(cont, bg=BG, width=360)
        self._karte_detailcol.pack(side="left", fill="both", padx=(16, 0))
        self._karte_detailcol.pack_propagate(False)
        info = "Punkt anklicken für Kontaktdaten." if platzierbar else "Keine Kunden mit erkennbarer PLZ."
        tk.Label(self._karte_detailcol, text=info, bg=BG, fg=MUTED,
                 font=(FONT, 11), wraplength=320, justify="left").pack(anchor="w", pady=18)

        W, H = 560, 720
        cv = tk.Canvas(mapcol, bg="#EAF1F8", width=W, height=H, highlightthickness=1,
                       highlightbackground=BORDER)
        cv.pack(anchor="nw")
        self._karte_canvas = cv
        from app import geo_de
        # Umriss zeichnen
        pts = []
        for lat, lon in geo_de.GERMANY_OUTLINE:
            x, y = geo_de.project(lat, lon, W, H)
            pts += [x, y]
        cv.create_polygon(pts, fill="#FBFCFE", outline="#9DB6CE", width=2)
        # Legende
        for i, kl in enumerate(("A", "B", "C")):
            cv.create_oval(16, 16 + i * 22, 28, 28 + i * 22, fill=self.ABC_COLORS[kl], outline="")
            cv.create_text(36, 22 + i * 22, text=_T('Klasse {p0}', p0=kl), anchor="w",
                           fill=INK, font=(FONT, 9))
        # Kollisionsschutz fuer Labels: leichte Verschiebung bei Ueberlappung
        belegt = []
        for k in platzierbar:
            lat, lon = k["latlon"]
            x, y = geo_de.project(lat, lon, W, H)
            klasse = k.get("abc")
            col = self.ABC_COLORS.get(klasse)
            r = 7
            dot = cv.create_oval(x - r, y - r, x + r, y + r, fill=col, outline="#FFFFFF", width=2)
            ly = y - 12
            for (bx, by) in belegt:
                if abs(bx - x) < 70 and abs(by - ly) < 14:
                    ly += 16
            belegt.append((x, ly))
            # rechte Punkte: Label nach links, sonst ragt es ueber den Kartenrand
            if x > W * 0.66:
                lx, anchor = x - 10, "e"
            else:
                lx, anchor = x + 10, "w"
            label = cv.create_text(lx, ly, text=k.get("kundenname") or "", anchor=anchor,
                                   fill=INK, font=(FONT, 9, "bold"))
            for item in (dot, label):
                cv.tag_bind(item, "<Button-1>", lambda e, kk=k: self._karte_show(kk))
                cv.tag_bind(item, "<Enter>", lambda e: cv.config(cursor="hand2"))
                cv.tag_bind(item, "<Leave>", lambda e: cv.config(cursor=""))

    def _karte_show(self, kunde):
        col = self._karte_detailcol
        for c in col.winfo_children():
            c.destroy()
        self._render_kunden_detail(col, kunde)

    # --- Kontakt-Aktionen ---
    def _copy_text(self, value, was="Text"):
        try:
            self.clipboard_clear()
            self.clipboard_append(str(value))
            messagebox.showinfo("Kopiert", _T('{p0} in die Zwischenablage kopiert.', p0=was))
        except Exception:
            pass

    def _mailto(self, email):
        import webbrowser
        try:
            webbrowser.open(f"mailto:{email}")
        except Exception as exc:
            messagebox.showwarning("E-Mail", _T('Konnte E-Mail-Programm nicht öffnen:\n{p0}', p0=exc))

    def _open_maps(self, adresse):
        import webbrowser
        from urllib.parse import quote_plus
        try:
            webbrowser.open(f"https://www.google.com/maps/search/?api=1&query={quote_plus(adresse)}")
        except Exception as exc:
            messagebox.showwarning("Karte", _T('Konnte Karte nicht öffnen:\n{p0}', p0=exc))

    # ----- Seite: PZN-Schnellauskunft ---------------------------------------
    def _page_auskunft(self):
        page = tk.Frame(self.content, bg=BG)
        page.pack(fill="both", expand=True)
        self._page_header(page, "PZN-Schnellauskunft",
                          "Eine PZN oder einen Artikelnamen eingeben – alle Daten auf einen Blick.")
        top = tk.Frame(page, bg=BG)
        top.pack(fill="x", padx=36, pady=(4, 8))
        self._auskunft_var = tk.StringVar()
        ent = tk.Entry(top, textvariable=self._auskunft_var, font=(FONT, 14),
                       relief="flat", highlightbackground=BORDER, highlightthickness=1)
        ent.pack(side="left", fill="x", expand=True, ipady=9)
        ent.bind("<Return>", lambda e: self._auskunft_suchen())
        ent.focus_set()
        PillButton(top, "🔎  Suchen", self._auskunft_suchen, kind="primary").pack(side="left", padx=(12, 0))

        self._auskunft_body = tk.Frame(page, bg=BG)
        self._auskunft_body.pack(fill="both", expand=True, padx=36, pady=(6, 24))
        tk.Label(self._auskunft_body,
                 text="Beispiel: PZN 12457880  ·  oder Name wie „Aclasta“, „Avastin“, „Humira“",
                 bg=BG, fg=MUTED, font=(FONT, 11)).pack(anchor="w", pady=20)

    def _auskunft_suchen(self):
        begriff = self._auskunft_var.get().strip()
        for c in self._auskunft_body.winfo_children():
            c.destroy()
        if not begriff:
            return
        # reine PZN -> direkt 360°, sonst Trefferliste
        if begriff.replace(".", "").isdigit() and len(_norm_pzn(begriff)) == 8:
            data = artikel_360(begriff)
            if data["gefunden"]:
                self._render_artikel(data)
                return
        treffer = suche_artikel(begriff)
        if not treffer:
            tk.Label(self._auskunft_body, text=_T('Keine Treffer für „{p0}“.', p0=begriff),
                     bg=BG, fg=DANGER, font=(FONT, 12)).pack(anchor="w", pady=14)
            return
        if len(treffer) == 1:
            self._render_artikel(artikel_360(treffer[0]["pzn"]))
            return
        tk.Label(self._auskunft_body, text=_T('{p0} Treffer – bitte auswählen:', p0=len(treffer)),
                 bg=BG, fg=MUTED, font=(FONT, 11)).pack(anchor="w", pady=(0, 8))
        lst = tk.Frame(self._auskunft_body, bg=BG)
        lst.pack(fill="both", expand=True)
        for t in treffer:
            row = tk.Frame(lst, bg=CARD, highlightbackground=BORDER, highlightthickness=1, cursor="hand2")
            row.pack(fill="x", pady=3)
            txt = f"  {t['pzn']}   {t['name'] or ''}"
            if t.get("herst"):
                txt += f"   · {t['herst']}"
            lab = tk.Label(row, text=txt, bg=CARD, fg=INK, font=(FONT, 11), anchor="w", pady=9)
            lab.pack(side="left", fill="x", expand=True)
            for w in (row, lab):
                w.bind("<Button-1>", lambda e, p=t["pzn"]: self._auskunft_open(p))
                w.bind("<Enter>", lambda e, r=row, l=lab: (r.config(bg="#EEF3F8"), l.config(bg="#EEF3F8")))
                w.bind("<Leave>", lambda e, r=row, l=lab: (r.config(bg=CARD), l.config(bg=CARD)))

    def _auskunft_open(self, pzn: str):
        for c in self._auskunft_body.winfo_children():
            c.destroy()
        self._render_artikel(artikel_360(pzn))

    def _render_artikel(self, data: dict):
        body = self._auskunft_body
        nmg = data.get("nmg") or {}
        stamm = data.get("stamm") or {}
        name = nmg.get("artikelname") or stamm.get("artikel") or "(kein Name hinterlegt)"

        head = Card(body, padding=20)
        head.pack(fill="x", pady=(0, 14))
        tk.Label(head.inner, text=name, bg=CARD, fg=INK, font=(FONT, 18, "bold"),
                 justify="left", wraplength=820).pack(anchor="w")
        sub = f"PZN {data['pzn']}"
        herst = nmg.get("herstellerkuerzel") or stamm.get("herst")
        if herst:
            sub += f"   ·   {herst}"
        if nmg.get("menge") or nmg.get("einheit"):
            sub += f"   ·   {nmg.get('menge') or ''} {nmg.get('einheit') or ''}".rstrip()
        tk.Label(head.inner, text=sub, bg=CARD, fg=MUTED, font=(FONT, 12)).pack(anchor="w", pady=(3, 0))
        if nmg.get("wirkstoffe"):
            tk.Label(head.inner, text=f"💊  {nmg['wirkstoffe']}", bg=CARD, fg=PRIMARY,
                     font=(FONT, 11), justify="left", wraplength=820).pack(anchor="w", pady=(8, 0))

        # Preis-/Konditions-Kacheln
        grid = tk.Frame(body, bg=BG)
        grid.pack(fill="x", pady=(0, 14))
        apu = nmg.get("apu")
        rabatt = (data.get("rabatt") or {}).get("rabatt")
        ek = data.get("ek_effektiv")
        eff_apu = apu * (1 - rabatt) if (apu is not None and rabatt is not None) else None
        kacheln = [
            ("APU (Listenpreis)", self._eur(apu), PRIMARY),
            ("NMG-Rabatt", (f"{rabatt*100:.1f} %" if rabatt is not None else "–"), SUCCESS if rabatt else MUTED),
            ("APU nach Rabatt", self._eur(eff_apu), SUCCESS if eff_apu is not None else MUTED),
            ("Taxe-EK / Taxe-VK", f"{self._eur(nmg.get('taxe_ek'))} / {self._eur(nmg.get('taxe_vk'))}", ACCENT),
        ]
        for i, (lab, val, col) in enumerate(kacheln):
            k = tk.Frame(grid, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
            k.grid(row=0, column=i, sticky="nsew", padx=(0 if i == 0 else 10, 0))
            grid.columnconfigure(i, weight=1)
            tk.Label(k, text=lab, bg=CARD, fg=MUTED, font=(FONT, 10)).pack(anchor="w", padx=14, pady=(12, 0))
            tk.Label(k, text=val, bg=CARD, fg=col, font=(FONT, 17, "bold")).pack(anchor="w", padx=14, pady=(2, 12))

        # Ersparnis-Hinweis (Marge gegenueber Taxe-EK)
        taxe_ek = nmg.get("taxe_ek")
        if eff_apu is not None and taxe_ek:
            diff = taxe_ek - eff_apu
            pct = diff / taxe_ek * 100 if taxe_ek else 0
            box = tk.Frame(body, bg="#EAF6EF", highlightbackground="#BfE3CD", highlightthickness=1)
            box.pack(fill="x", pady=(0, 14))
            tk.Label(box, text=_T('💡  Einkaufsvorteil ggü. Taxe-EK: {p0} pro Packung ({p1:.1f} %)', p0=self._eur(diff), p1=pct),
                     bg="#EAF6EF", fg=SUCCESS if diff > 0 else DANGER, font=(FONT, 12, "bold")).pack(anchor="w", padx=16, pady=11)

        # Mengen-/Konditionsrechner + Kopier-Funktion
        calc = Card(body, padding=18)
        calc.pack(fill="x", pady=(0, 14))
        toprow = tk.Frame(calc.inner, bg=CARD)
        toprow.pack(fill="x")
        tk.Label(toprow, text="🧮  Mengenrechner", bg=CARD, fg=PRIMARY,
                 font=(FONT, 13, "bold")).pack(side="left")
        PillButton(toprow, "📋  Als Text kopieren",
                   lambda: self._kopiere_auskunft(data, name, apu, rabatt, eff_apu, taxe_ek),
                   kind="ghost").pack(side="right")
        inrow = tk.Frame(calc.inner, bg=CARD)
        inrow.pack(fill="x", pady=(10, 6))
        tk.Label(inrow, text="Menge (Packungen):", bg=CARD, fg=MUTED, font=(FONT, 11)).pack(side="left")
        menge_var = tk.StringVar(value="10")
        sp = tk.Spinbox(inrow, from_=1, to=100000, textvariable=menge_var, width=8,
                        font=(FONT, 12), relief="flat", highlightbackground=BORDER, highlightthickness=1)
        sp.pack(side="left", padx=(8, 0), ipady=3)
        res = tk.Label(calc.inner, text="", bg=CARD, fg=INK, font=(FONT, 12), justify="left")
        res.pack(anchor="w", pady=(8, 0))

        def _rechne(*_a):
            try:
                m = max(0, int(float(menge_var.get().replace(",", "."))))
            except Exception:
                res.config(text="Bitte eine gültige Menge eingeben.", fg=DANGER)
                return
            teile = []
            if eff_apu is not None:
                teile.append(f"Einkauf gesamt (APU nach Rabatt): {self._eur(eff_apu * m)}")
            elif apu is not None:
                teile.append(f"Einkauf gesamt (APU): {self._eur(apu * m)}")
            if eff_apu is not None and taxe_ek:
                teile.append(f"Einkaufsvorteil ggü. Taxe-EK gesamt: {self._eur((taxe_ek - eff_apu) * m)}")
            res.config(text=("\n".join(teile) if teile else "Keine Preisdaten für eine Berechnung hinterlegt."),
                       fg=INK if teile else MUTED)

        menge_var.trace_add("write", _rechne)
        _rechne()

        # Detail-Spalten
        cols = tk.Frame(body, bg=BG)
        cols.pack(fill="both", expand=True)
        left = tk.Frame(cols, bg=BG)
        left.pack(side="left", fill="both", expand=True)
        right = tk.Frame(cols, bg=BG)
        right.pack(side="left", fill="both", expand=True, padx=(14, 0))

        # Stammdaten
        rows = []
        if stamm:
            rows += [("Artikel (Stamm)", stamm.get("artikel")), ("Darreichungsform", stamm.get("df")),
                     ("Packung", stamm.get("pck")), ("EK (Stamm)", self._eur(stamm.get("ek")))]
        if ek is not None:
            rows.append(("EK effektiv", self._eur(ek)))
        rows.append(("Quelle Rabatt", (data.get("rabatt") or {}).get("quelle")))
        self._detail_card(left, "Stamm- & Preisdaten", rows)

        # Lieferfaehigkeit
        liefer = data.get("liefer")
        lrows = []
        if liefer:
            lieferbar = "ja" if liefer.get("lieferbar") == "X" else "—"
            bev = "empfohlen" if liefer.get("bevorratung_angeraten") == "X" else "—"
            lrows = [("Lieferbar", lieferbar), ("Bevorratung", bev),
                     ("Liefervorschlag", liefer.get("liefervorschlag") or "—")]
        else:
            lrows = [("Lieferfähigkeit", "keine Angabe hinterlegt")]
        self._detail_card(right, "Verfügbarkeit", lrows)

        # Biosimilar / Alternativen
        bio = data.get("biosimilar")
        if bio:
            card = Card(body, padding=18)
            card.pack(fill="x", pady=(14, 0))
            tk.Label(card.inner, text=f"🧬  Wirkstoff-Gruppe: {bio.get('wirkstoff') or '?'}"
                     + (f"  ·  Referenz: {bio['referenz']}" if bio.get("referenz") else ""),
                     bg=CARD, fg=PRIMARY, font=(FONT, 13, "bold")).pack(anchor="w")
            tk.Label(card.inner, text=_T('Dieses Produkt: {p0} ({p1})', p0=bio.get('name'), p1=bio.get('rolle') or 'Produkt'),
                     bg=CARD, fg=INK, font=(FONT, 11)).pack(anchor="w", pady=(4, 8))
            if bio.get("alternativen"):
                tk.Label(card.inner, text="Austauschbare Produkte derselben Gruppe:",
                         bg=CARD, fg=MUTED, font=(FONT, 10)).pack(anchor="w")
                for a in bio["alternativen"]:
                    tk.Label(card.inner, text=f"   •  {a['name']}  ({a.get('rolle') or 'Produkt'})",
                             bg=CARD, fg=INK, font=(FONT, 11)).pack(anchor="w")

        # Austausch -> NMG
        aus = data.get("austausch_zu_nmg") or []
        if aus:
            card = Card(body, padding=18)
            card.pack(fill="x", pady=(14, 0))
            tk.Label(card.inner, text=_T('🔁  In der Austauschdatenbank ({p0} NMG-Vorschläge)', p0=len(aus)),
                     bg=CARD, fg=PRIMARY, font=(FONT, 13, "bold")).pack(anchor="w", pady=(0, 6))
            for a in aus:
                txt = f"   →  NMG-PZN {a['pzn_nmg']}"
                if a.get("artikel_nmg"):
                    txt += f"   {a['artikel_nmg']}"
                tk.Label(card.inner, text=txt, bg=CARD, fg=INK, font=(FONT, 11)).pack(anchor="w")

    def _kopiere_auskunft(self, data, name, apu, rabatt, eff_apu, taxe_ek):
        """Kopiert eine kompakte Text-Zusammenfassung in die Zwischenablage
        (praktisch fuer Angebote per E-Mail an Apotheken)."""
        zeilen = [name, f"PZN {data['pzn']}"]
        if apu is not None:
            zeilen.append(f"APU: {self._eur(apu)}")
        if rabatt is not None:
            zeilen.append(f"NMG-Rabatt: {rabatt*100:.1f} %")
        if eff_apu is not None:
            zeilen.append(f"APU nach Rabatt: {self._eur(eff_apu)}")
        if taxe_ek:
            zeilen.append(f"Taxe-EK: {self._eur(taxe_ek)}")
            if eff_apu is not None:
                zeilen.append(f"Einkaufsvorteil/Pck: {self._eur(taxe_ek - eff_apu)}")
        text = "\n".join(zeilen)
        try:
            self.clipboard_clear()
            self.clipboard_append(text)
            messagebox.showinfo("Kopiert", "Die Zusammenfassung wurde in die Zwischenablage kopiert.")
        except Exception as exc:
            messagebox.showwarning("Kopieren", _T('Konnte nicht kopieren:\n{p0}', p0=exc))

    def _detail_card(self, parent, title, rows):
        card = Card(parent, padding=18)
        card.pack(fill="x")
        tk.Label(card.inner, text=title, bg=CARD, fg=PRIMARY, font=(FONT, 13, "bold")).pack(anchor="w", pady=(0, 8))
        for lab, val in rows:
            r = tk.Frame(card.inner, bg=CARD)
            r.pack(fill="x", pady=2)
            tk.Label(r, text=lab, bg=CARD, fg=MUTED, font=(FONT, 10), width=18, anchor="w").pack(side="left")
            tk.Label(r, text=("—" if val in (None, "") else str(val)), bg=CARD, fg=INK,
                     font=(FONT, 11), anchor="w", justify="left", wraplength=360).pack(side="left", fill="x", expand=True)

    @staticmethod
    def _eur(v) -> str:
        if v in (None, ""):
            return "–"
        try:
            return f"{float(v):,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")
        except Exception:
            return str(v)

    # ----- Seite: Markt-Insights --------------------------------------------
    def _page_markt(self):
        page = tk.Frame(self.content, bg=BG)
        page.pack(fill="both", expand=True)
        self._page_header(page, "Markt-Insights",
                          "Was deine Apotheken-Kunden über alle Auswertungen hinweg wirklich nachfragen.")
        toolbar = tk.Frame(page, bg=BG)
        toolbar.pack(fill="x", padx=36, pady=(0, 4))
        self._markt_modus = getattr(self, "_markt_modus", "nmg")
        self._markt_btn_nmg = PillButton(
            toolbar, "💰  NMG nach Rabatt-Potenzial", lambda: self._markt_set("nmg"),
            kind="primary" if self._markt_modus == "nmg" else "ghost")
        self._markt_btn_nmg.pack(side="left")
        self._markt_btn_bedarf = PillButton(
            toolbar, "📊  Gesamt-Bedarf (Absatz)", lambda: self._markt_set("bedarf"),
            kind="primary" if self._markt_modus == "bedarf" else "ghost")
        self._markt_btn_bedarf.pack(side="left", padx=(10, 0))
        PillButton(toolbar, "⬇  CSV", lambda: self._markt_csv(), kind="ghost").pack(side="right")

        self._markt_body = tk.Frame(page, bg=BG)
        self._markt_body.pack(fill="both", expand=True, padx=36, pady=(8, 24))
        self._markt_render()

    def _markt_set(self, modus):
        self._markt_modus = modus
        self._markt_btn_nmg.config(bg=(PRIMARY if modus == "nmg" else CARD),
                                   fg=("#FFFFFF" if modus == "nmg" else PRIMARY))
        self._markt_btn_nmg._base = PRIMARY if modus == "nmg" else CARD
        self._markt_btn_bedarf.config(bg=(PRIMARY if modus == "bedarf" else CARD),
                                      fg=("#FFFFFF" if modus == "bedarf" else PRIMARY))
        self._markt_btn_bedarf._base = PRIMARY if modus == "bedarf" else CARD
        self._markt_render()

    def _markt_render(self):
        for c in self._markt_body.winfo_children():
            c.destroy()
        modus = self._markt_modus
        try:
            rows = markt_top(modus, limit=30)
        except Exception as exc:
            tk.Label(self._markt_body, text=_T('Konnte Markt-Daten nicht laden:\n{p0}', p0=exc),
                     bg=BG, fg=DANGER, font=(FONT, 12)).pack(anchor="w", pady=20)
            return
        self._markt_rows = rows
        if not rows:
            tk.Label(self._markt_body, text="Noch keine Auswertungs-Positionen vorhanden.",
                     bg=BG, fg=MUTED, font=(FONT, 12)).pack(anchor="w", pady=20)
            return

        if modus == "nmg":
            erkl = ("Nur NMG-Treffer, sortiert nach gesamtem Rabatt-Potenzial über alle Apotheken. "
                    "Oben stehen die Produkte, die sich für deinen Einkauf am meisten lohnen.")
            cols = [("PZN", 10), ("Produkt", 34), ("Apo.", 5), ("Absatz", 9), ("Ø Rabatt", 9), ("Rabatt-Potenzial", 16)]
        else:
            erkl = ("Alle Produkte, sortiert nach Gesamt-Absatz (6 Monate) über alle Apotheken. "
                    "Zeigt die gesamte Marktnachfrage; NMG-Treffer sind grün markiert.")
            cols = [("PZN", 10), ("Produkt", 34), ("Apo.", 5), ("Absatz", 9), ("NMG", 5), ("Rabatt-Potenzial", 16)]
        tk.Label(self._markt_body, text=erkl, bg=BG, fg=MUTED, font=(FONT, 11),
                 justify="left", wraplength=860).pack(anchor="w", pady=(0, 10))

        header = tk.Frame(self._markt_body, bg=BG)
        header.pack(fill="x", padx=(0, 8))
        for txt, w in cols:
            tk.Label(header, text=txt, bg=BG, fg=MUTED, font=(FONT, 10, "bold"),
                     width=w, anchor="w").pack(side="left")

        canvas = tk.Canvas(self._markt_body, bg=BG, highlightthickness=0)
        scroll = ttk.Scrollbar(self._markt_body, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=BG)
        win = canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win, width=e.width))
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-e.delta / 120), "units"))

        for i, r in enumerate(rows):
            bg = CARD if i % 2 == 0 else "#F7FAFC"
            row = tk.Frame(inner, bg=bg, highlightbackground=BORDER, highlightthickness=1, cursor="hand2")
            row.pack(fill="x", pady=1, padx=(0, 8))
            def cell(text, w, fg=INK, bold=False):
                tk.Label(row, text=text, bg=bg, fg=fg, font=(FONT, 10, "bold" if bold else "normal"),
                         width=w, anchor="w").pack(side="left", pady=7, padx=(6 if w == cols[0][1] else 0, 0))
            cell(r["pzn"], cols[0][1])
            cell(str(r["name"] or "")[:40], cols[1][1], bold=True)
            cell(str(r["apotheken"]), cols[2][1])
            cell(f"{r['absatz']:.0f}", cols[3][1])
            if modus == "nmg":
                rs = r.get("rabatt_satz")
                cell(f"{rs*100:.0f} %" if rs else "–", cols[4][1])
            else:
                cell("✓" if r.get("ist_nmg") else "–", cols[4][1],
                     fg=SUCCESS if r.get("ist_nmg") else MUTED, bold=bool(r.get("ist_nmg")))
            cell(self._eur(r.get("rabatt_pot")), cols[5][1], fg=PRIMARY, bold=True)
            pzn = r["pzn"]
            for w in (row,) + tuple(row.winfo_children()):
                w.bind("<Button-1>", lambda e, p=pzn: self._markt_open_pzn(p))

    def _markt_open_pzn(self, pzn):
        self.show_page("auskunft")
        self._auskunft_var.set(pzn)
        self._auskunft_suchen()

    def _markt_csv(self):
        rows = getattr(self, "_markt_rows", None)
        if not rows:
            return
        import csv
        from datetime import datetime
        modus = self._markt_modus
        path = filedialog.asksaveasfilename(
            title="Markt-Insights als CSV speichern", defaultextension=".csv",
            initialfile=f"Markt_Insights_{modus}_{datetime.now():%Y%m%d}.csv",
            filetypes=[("CSV-Datei", "*.csv")])
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as fh:
                w = csv.writer(fh, delimiter=";")
                w.writerow(["PZN", "Produkt", "Apotheken", "Absatz_6M", "Ø_Rabatt", "Rabatt_Potenzial", "NMG"])
                for r in rows:
                    rs = r.get("rabatt_satz")
                    w.writerow([r["pzn"], r.get("name") or "", r["apotheken"],
                                f"{r['absatz']:.0f}", (f"{rs*100:.0f}%" if rs else ""),
                                f"{(r.get('rabatt_pot') or 0):.2f}",
                                "ja" if r.get("ist_nmg") else ""])
            messagebox.showinfo("Export", _T('Gespeichert:\n{p0}', p0=path))
        except Exception as exc:
            messagebox.showwarning("Export", _T('Konnte nicht speichern:\n{p0}', p0=exc))

    # ----- Seite: Datenbank-Dashboard ---------------------------------------
    def _page_dashboard(self):
        inner = self._scroll_page("Datenbank-Dashboard",
                                  "Überblick über den aktuellen Datenbestand der NMGone-Datenbank.")
        try:
            k = db_kennzahlen()
        except Exception as exc:
            tk.Label(inner, text=_T('Datenbank konnte nicht gelesen werden:\n{p0}', p0=exc),
                     bg=BG, fg=DANGER, font=(FONT, 12)).pack(anchor="w", pady=20)
            return

        # Kennzahl-Kacheln
        tiles = [
            ("Artikelstamm", f"{k['artikel']:,}".replace(",", "."), "Artikel gesamt", PRIMARY),
            ("NMG-Stamm", f"{k['nmg_stamm']:,}".replace(",", "."), "Hochpreiser mit APU", ACCENT),
            ("NMG-Rabatte", f"{k['rabatte']:,}".replace(",", "."), "Partnerkonditionen", SUCCESS),
            ("Austausch-DB", f"{k['austausch']:,}".replace(",", "."), "Austausch-Einträge", WARNING),
            ("Biosimilar-Gruppen", f"{k['biosimilar_gruppen']:,}".replace(",", "."), f"{k['biosimilar_pzn']:,} PZN".replace(",", "."), PRIMARY),
            ("Auswertungen", f"{k['auswertungen']:,}".replace(",", "."), f"{k['positionen']:,} Positionen".replace(",", "."), ACCENT),
        ]
        grid = tk.Frame(inner, bg=BG)
        grid.pack(fill="x", pady=(0, 18), padx=(0, 8))
        cols = 3
        for i, (title, big, small, col) in enumerate(tiles):
            r, c = divmod(i, cols)
            tile = tk.Frame(grid, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
            tile.grid(row=r, column=c, sticky="nsew", padx=(0 if c == 0 else 12, 0), pady=(0, 12))
            grid.columnconfigure(c, weight=1)
            tk.Label(tile, text=title, bg=CARD, fg=MUTED, font=(FONT, 11)).pack(anchor="w", padx=16, pady=(14, 0))
            tk.Label(tile, text=big, bg=CARD, fg=col, font=(FONT, 26, "bold")).pack(anchor="w", padx=16)
            tk.Label(tile, text=small, bg=CARD, fg=MUTED, font=(FONT, 10)).pack(anchor="w", padx=16, pady=(0, 14))

        # Rabatt-Übersicht
        if k.get("rabatt_avg") is not None:
            card = Card(inner, padding=18)
            card.pack(fill="x", pady=(0, 14), padx=(0, 8))
            tk.Label(card.inner, text="NMG-Rabatte – Verteilung", bg=CARD, fg=PRIMARY,
                     font=(FONT, 14, "bold")).pack(anchor="w", pady=(0, 8))
            line = (f"Ø {k['rabatt_avg']*100:.1f} %     "
                    f"Spanne {k['rabatt_min']*100:.0f} – {k['rabatt_max']*100:.0f} %")
            tk.Label(card.inner, text=line, bg=CARD, fg=INK, font=(FONT, 13)).pack(anchor="w")

        # Top-Hersteller (Balken)
        if k.get("top_herst"):
            card = Card(inner, padding=18)
            card.pack(fill="x", pady=(0, 14), padx=(0, 8))
            tk.Label(card.inner, text="Top-Anbieter im Artikelstamm", bg=CARD, fg=PRIMARY,
                     font=(FONT, 14, "bold")).pack(anchor="w", pady=(0, 10))
            maxn = max(h["n"] for h in k["top_herst"]) or 1
            for h in k["top_herst"]:
                row = tk.Frame(card.inner, bg=CARD)
                row.pack(fill="x", pady=3)
                tk.Label(row, text=(h["herst"] or "?")[:24], bg=CARD, fg=INK, font=(FONT, 11),
                         width=24, anchor="w").pack(side="left")
                barwrap = tk.Frame(row, bg="#EEF3F8", height=18)
                barwrap.pack(side="left", fill="x", expand=True, padx=(6, 8))
                barwrap.pack_propagate(False)
                w = max(0.02, h["n"] / maxn)
                bar = tk.Frame(barwrap, bg=ACCENT)
                bar.place(relwidth=w, relheight=1)
                tk.Label(row, text=str(h["n"]), bg=CARD, fg=MUTED, font=(FONT, 10), width=6, anchor="e").pack(side="left")

        # Verfügbarkeit + letzte Auswertung
        info = Card(inner, padding=18)
        info.pack(fill="x", pady=(0, 14), padx=(0, 8))
        tk.Label(info.inner, text="Weitere Kennzahlen", bg=CARD, fg=PRIMARY,
                 font=(FONT, 14, "bold")).pack(anchor="w", pady=(0, 8))
        if "lieferbar" in k:
            tk.Label(info.inner, text=_T('Als lieferbar markiert: {p0}   ·   Bevorratung empfohlen: {p1}', p0=k['lieferbar'], p1=k['bevorratung']),
                     bg=CARD, fg=INK, font=(FONT, 12)).pack(anchor="w", pady=2)
        if k.get("letzte"):
            le = k["letzte"]
            kunde = le.get("kundenname") or le.get("apotheke") or "?"
            tk.Label(info.inner, text=_T('Letzte Auswertung: {p0}  ·  {p1}', p0=le.get('datum') or '?', p1=kunde),
                     bg=CARD, fg=INK, font=(FONT, 12)).pack(anchor="w", pady=2)
        db_name = Path(DB_PATH).name
        tk.Label(info.inner, text=_T('Datenbank-Datei: {p0}', p0=db_name), bg=CARD, fg=MUTED,
                 font=(FONT, 10)).pack(anchor="w", pady=(8, 0))

    # ----- Seite: Auswertungs-Verlauf ---------------------------------------
    def _page_verlauf(self):
        inner = self._scroll_page("Auswertungs-Verlauf",
                                  "Alle bisher erstellten Auswertungen – jüngste zuerst.")
        try:
            rows = auswertungen_liste()
        except Exception as exc:
            tk.Label(inner, text=_T('Konnte Verlauf nicht laden:\n{p0}', p0=exc), bg=BG, fg=DANGER,
                     font=(FONT, 12)).pack(anchor="w", pady=20)
            return
        if not rows:
            tk.Label(inner, text="Noch keine Auswertungen vorhanden.", bg=BG, fg=MUTED,
                     font=(FONT, 12)).pack(anchor="w", pady=20)
            return

        bar = tk.Frame(inner, bg=BG)
        bar.pack(fill="x", pady=(0, 8), padx=(0, 8))
        tk.Label(bar, text=_T('{p0} Auswertungen', p0=len(rows)), bg=BG, fg=MUTED,
                 font=(FONT, 11)).pack(side="left")
        PillButton(bar, "⬇  Als CSV exportieren", lambda: self._export_verlauf_csv(rows),
                   kind="ghost").pack(side="right")

        # Kopfzeile
        header = tk.Frame(inner, bg=BG)
        header.pack(fill="x", pady=(0, 6), padx=(0, 8))
        spalten = [("Datum", 16), ("Kunde", 30), ("Pos.", 7), ("NMG", 7), ("Absatz", 14), ("", 10)]
        for txt, w in spalten:
            tk.Label(header, text=txt, bg=BG, fg=MUTED, font=(FONT, 10, "bold"),
                     width=w, anchor="w").pack(side="left")

        for a in rows:
            card = tk.Frame(inner, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
            card.pack(fill="x", pady=2, padx=(0, 8))
            datum = (a.get("datum") or "")[:16]
            tk.Label(card, text=datum, bg=CARD, fg=INK, font=(FONT, 10), width=16, anchor="w").pack(side="left", pady=8, padx=(8, 0))
            tk.Label(card, text=(a.get("kunde") or "—")[:30], bg=CARD, fg=INK, font=(FONT, 11, "bold"),
                     width=30, anchor="w").pack(side="left")
            tk.Label(card, text=str(a.get("anzahl_positionen") or 0), bg=CARD, fg=INK,
                     font=(FONT, 10), width=7, anchor="w").pack(side="left")
            tk.Label(card, text=str(a.get("nmg_treffer") or 0), bg=CARD, fg=SUCCESS,
                     font=(FONT, 10), width=7, anchor="w").pack(side="left")
            tk.Label(card, text=self._eur(a.get("gesamt_absatz")), bg=CARD, fg=PRIMARY,
                     font=(FONT, 10), width=14, anchor="w").pack(side="left")
            out = a.get("ausgabedatei")
            if out and Path(out).exists():
                PillButton(card, "Öffnen", lambda p=out: self._open_file(Path(p)), kind="ghost").pack(side="right", padx=8, pady=4)
            else:
                tk.Label(card, text="(Datei fehlt)", bg=CARD, fg=MUTED, font=(FONT, 9),
                         width=12, anchor="e").pack(side="right", padx=8)

    def _export_verlauf_csv(self, rows):
        import csv
        from datetime import datetime
        path = filedialog.asksaveasfilename(
            title="Verlauf als CSV speichern", defaultextension=".csv",
            initialfile=f"Auswertungs_Verlauf_{datetime.now():%Y%m%d}.csv",
            filetypes=[("CSV-Datei", "*.csv")])
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as fh:
                w = csv.writer(fh, delimiter=";")
                w.writerow(["Datum", "Kunde", "Positionen", "NMG-Treffer", "Nicht-NMG",
                            "Gesamt-Absatz", "Datenquelle"])
                for a in rows:
                    w.writerow([a.get("datum") or "", a.get("kunde") or "",
                                a.get("anzahl_positionen") or 0, a.get("nmg_treffer") or 0,
                                a.get("nicht_nmg") or 0, a.get("gesamt_absatz") or 0,
                                a.get("datenquelle") or ""])
            messagebox.showinfo("Export", _T('Verlauf gespeichert:\n{p0}', p0=path))
        except Exception as exc:
            messagebox.showwarning("Export", _T('Konnte nicht speichern:\n{p0}', p0=exc))

    # ----- Seite: Hilfe -----------------------------------------------------
    def _page_hilfe(self):
        page = tk.Frame(self.content, bg=BG)
        page.pack(fill="both", expand=True)
        self._page_header(page, "Hilfe & Anleitung",
                          "In drei Schritten zur fertigen Auswertung – mit Bildern erklärt.")

        canvas = tk.Canvas(page, bg=BG, highlightthickness=0)
        scroll = ttk.Scrollbar(page, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=BG)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side="left", fill="both", expand=True, padx=(36, 0), pady=(0, 20))
        scroll.pack(side="right", fill="y")
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-e.delta / 120), "units"))

        # Werkzeug-Übersicht (neu)
        tools = Card(inner, padding=22)
        tools.pack(fill="x", pady=(0, 18), padx=(0, 8))
        tk.Label(tools.inner, text="Werkzeuge in dieser Vorschau", bg=CARD, fg=PRIMARY,
                 font=(FONT, 17, "bold")).pack(anchor="w", pady=(0, 10))
        werkzeuge = [
            ("📊  Neue Auswertung", "Rohdatei einlesen und auswerten – mit sichtbarem Fortschritt."),
            ("👥  Kunden & ABC", "Komplette Kundenverwaltung: Kundenliste mit Kontakt-Steckbrief, "
             "ABC-Analyse nach Rabatt-Potenzial und eine Deutschlandkarte – Punkt anklicken zeigt "
             "alle Daten, um sich beim Kunden zu melden. Pro Kunde: Umsatz nach Zeitraum, Top-Artikel "
             "(bis zu allen) und offene Vorbestellungen."),
            ("🔎  PZN-Schnellauskunft", "PZN oder Name eingeben und sofort APU, Rabatt, Einkaufsvorteil, "
             "Lieferfähigkeit, Wirkstoff-Gruppe und Austausch-Vorschläge sehen."),
            ("📦  Markt-Insights", "Über alle Auswertungen aggregiert: welche hochpreisigen Produkte "
             "viele Apotheken nachfragen und das größte Rabatt-Potenzial haben – Klick öffnet die Schnellauskunft."),
            ("📈  Datenbank-Dashboard", "Überblick: wie viele Artikel, Hochpreiser, Rabatte und Auswertungen "
             "stecken in der Datenbank, plus Top-Anbieter."),
            ("🗂️  Auswertungs-Verlauf", "Alle bisher erstellten Auswertungen mit Kennzahlen – Ausgabedatei direkt öffnen."),
        ]
        for titel, txt in werkzeuge:
            r = tk.Frame(tools.inner, bg=CARD)
            r.pack(fill="x", pady=5)
            tk.Label(r, text=titel, bg=CARD, fg=INK, font=(FONT, 12, "bold"),
                     width=24, anchor="w").pack(side="left", anchor="n")
            tk.Label(r, text=txt, bg=CARD, fg=MUTED, font=(FONT, 11), justify="left",
                     wraplength=620).pack(side="left", fill="x", expand=True)

        sections = [
            ("prozess_flow.png",
             "Wie eine Auswertung entsteht",
             "Du wählst eine Datei und startest. Den Rest macht das Programm automatisch: "
             "Es erkennt die Spalten, gleicht jede PZN gegen Stammdaten und Biosimilars ab, "
             "berechnet Absatz und Umsatz und schreibt am Ende eine fertige Excel-Datei."),
            ("formate.png",
             "Welche Dateien funktionieren",
             "Moderne Excel-Dateien (.xlsx), alte Excel-Dateien (.xls), sowie CSV- und Text-"
             "tabellen. Alte .xls-Dateien werden automatisch in das neue Format umgewandelt – "
             "du musst nichts vorbereiten."),
            ("fortschritt.png",
             "Du siehst, dass gearbeitet wird",
             "Während der Auswertung laufen ein Fortschrittsbalken und eine Stoppuhr. Die "
             "Stufen-Checkliste zeigt, wo das Programm gerade steht, und die Zahl der bereits "
             "abgeglichenen Positionen steigt sichtbar an."),
        ]
        for img_name, title, text in sections:
            card = Card(inner, padding=22)
            card.pack(fill="x", pady=(0, 18), padx=(0, 8))
            tk.Label(card.inner, text=title, bg=CARD, fg=PRIMARY, font=(FONT, 17, "bold")).pack(anchor="w")
            tk.Label(card.inner, text=text, bg=CARD, fg=INK, font=(FONT, 11),
                     justify="left", wraplength=820).pack(anchor="w", pady=(6, 14))
            img = self._img(img_name, max_w=860)
            if img:
                tk.Label(card.inner, image=img, bg=CARD).pack(anchor="w")
            else:
                tk.Label(card.inner, text="(Bild konnte nicht geladen werden)", bg=CARD, fg=MUTED).pack(anchor="w")

    # ----- Seite: Info ------------------------------------------------------
    def _page_info(self):
        page = tk.Frame(self.content, bg=BG)
        page.pack(fill="both", expand=True)
        self._page_header(page, "Über diese Vorschau",
                          "Was diese Testoberfläche ist – und was nicht.")
        card = Card(page, padding=24)
        card.pack(fill="x", padx=36, pady=(6, 0))
        punkte = [
            ("Isoliert vom Original", "Diese Oberfläche ist ein eigenes Fenster und verändert das bestehende NMGone-Programm nicht."),
            ("Echte Engine", "Die Auswertung nutzt dieselbe Berechnung wie das Original (create_vorlage_export) – die Ergebnisse sind echt."),
            ("Sichtbarer Fortschritt", "Balken, Stoppuhr, Stufen und die live steigende Positionszahl zeigen, dass gearbeitet wird."),
            (".xls wird gelesen", "Alte Excel-Dateien werden automatisch konvertiert und ausgewertet."),
            ("Neue Werkzeuge", "Kunden & ABC, PZN-Schnellauskunft, Markt-Insights, Datenbank-Dashboard und "
             "Auswertungs-Verlauf – alle lesen nur (read-only) aus der Datenbank und verändern nichts."),
            ("Offline-Karte", "Die Deutschlandkarte braucht kein Internet: PLZ werden über ihre Leitregion "
             "platziert (Stadt-/Regionsebene), gezeichnet auf einer Canvas-Fläche."),
            ("Zweck", "Entwurf für ein moderneres, professionelleres und leichter verständliches Design."),
        ]
        for t, b in punkte:
            r = tk.Frame(card.inner, bg=CARD)
            r.pack(fill="x", pady=7)
            tk.Label(r, text="●", bg=CARD, fg=ACCENT, font=(FONT, 12)).pack(side="left", anchor="n", padx=(0, 10))
            col = tk.Frame(r, bg=CARD)
            col.pack(side="left", fill="x", expand=True)
            tk.Label(col, text=t, bg=CARD, fg=INK, font=(FONT, 12, "bold")).pack(anchor="w")
            tk.Label(col, text=b, bg=CARD, fg=MUTED, font=(FONT, 11), justify="left", wraplength=820).pack(anchor="w")


def main():
    app = TestOberflaeche()
    app.mainloop()


if __name__ == "__main__":
    main()
