from pathlib import Path
from collections import defaultdict
import re
import csv
from openpyxl import load_workbook
from .db import connect, _pzn

PZN_ALIASES = {
    "pzn", "artikel pzn", "artikel-pzn", "pharmazentralnummer", "pharmazentral-nr", "pzn nr", "pzn-nr"
}
HERSTELLER_ALIASES = {
    "hersteller", "herst", "herst.", "herstellerkuerzel", "hersteller kuerzel",
    "herstellerkürzel", "hersteller kürzel", "herst kuerzel", "herst kürzel",
    "firma", "anbieter", "lieferant"
}
EK_ALIASES = {
    "ek", "apo ek", "apo-ek", "apo_ek", "lauer ek", "lauer-ek", "lauer_ek",
    "tax ek", "tax-ek", "tax_ek", "taxe ek", "taxe-ek", "taxe_ek",
    "einkaufspreis", "einkaufs preis", "ek netto", "ek-netto", "netto ek", "netto-ek"
}
ABSATZ_ALIASES = {
    "verkaufsmenge", "verkaufsmenge der letzten 6 monate", "abverkaeufe 6 monate",
    "abverkäufe 6 monate", "absatz", "abs", "abs pack", "abs packungen",
    "abs abg pack", "menge", "verkaufte menge", "abverkauf",
    "abverkauf letzte 6 monate", "abverkauf 6 monate", "abverkaeufe letzte 6 monate",
    "abverkäufe letzte 6 monate", "verbrauch 6 monate", "verbrauch letzte 6 monate"
}
ARTIKEL_ALIASES = {
    "artikelname", "artikelbezeichnung", "artikel", "bezeichnung", "name",
    "artikelbez.", "artikelbez", "artbez", "artbez.", "artikel bez", "artikel-bez"
}
DF_ALIASES = {"df", "dar", "darreichungsform"}
PACK_ALIASES = {"pck", "pack gr", "pack.gr", "packung", "packungsgröße", "pack groesse"}


def norm_header(value):
    text = str(value or "").strip().lower()
    text = text.replace("_", " ").replace("-", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def pzn_value(value):
    return _pzn(value)


def parse_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("€", "").replace(" ", "")
    if not text:
        return None
    try:
        if "," in text and "." in text:
            # Deutsch: 1.234,56
            text = text.replace(".", "").replace(",", ".")
        elif "," in text:
            text = text.replace(",", ".")
        return float(text)
    except Exception:
        return None


def clean_hersteller(value):
    """
    Bereinigung für Herstellerkürzel/-texte.
    In den NMG-Daten steht meist nur das Herstellerkürzel, z. B. OTSUK, PFIZ, 1APH.
    - Reine Zahlenblöcke am Anfang werden nur entfernt, wenn danach eine Leerzeile folgt.
    - Alphanumerische Hersteller/Kürzel wie "1A Pharma" oder "1APH" bleiben erhalten.
    """
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return ""
    lines = [line.strip() for line in text.split("\n")]
    i = 0
    saw_numeric_prefix = False
    while i < len(lines) and re.fullmatch(r"\d+", lines[i] or ""):
        saw_numeric_prefix = True
        i += 1
    blank_count = 0
    while i < len(lines) and lines[i] == "":
        blank_count += 1
        i += 1
    if saw_numeric_prefix and blank_count >= 1 and i < len(lines):
        remaining = [line for line in lines[i:] if line]
        return " ".join(remaining).strip()
    return " ".join(line for line in lines if line).strip()


def find_columns(ws, scan_rows=15):
    """Sucht Header-Zeile und Spalten. Erkennt auch Monatsverbrauchslisten."""
    monthly = detect_monthly_consumption_layout(ws, scan_rows=scan_rows)
    if monthly:
        return monthly
    best = None
    for row_idx in range(1, min(scan_rows, ws.max_row) + 1):
        raw_headers = [c.value for c in ws[row_idx]]
        headers = [norm_header(h) for h in raw_headers]
        mapping = {"header_row": row_idx, "headers": raw_headers}
        for idx, header in enumerate(headers, start=1):
            if not header:
                continue
            if "pzn" not in mapping and (header in PZN_ALIASES or header == "pzn" or "pharmazentral" in header):
                mapping["pzn"] = idx
                mapping["pzn_header"] = raw_headers[idx-1]
            if "artikel" not in mapping:
                h_compact_artikel = header.replace(" ", "").replace(".", "")
                if (
                    header in ARTIKEL_ALIASES
                    or "artikelbezeichnung" in header
                    or "artikelbez" in h_compact_artikel
                    or "artbez" in h_compact_artikel
                ):
                    mapping["artikel"] = idx
                    mapping["artikel_header"] = raw_headers[idx-1]
            if "df" not in mapping and (header in DF_ALIASES):
                mapping["df"] = idx
                mapping["df_header"] = raw_headers[idx-1]
            if "packung" not in mapping and (header in PACK_ALIASES or header.startswith("pack")):
                mapping["packung"] = idx
                mapping["packung_header"] = raw_headers[idx-1]
            if "hersteller" not in mapping and (header in HERSTELLER_ALIASES or header.startswith("herst")):
                mapping["hersteller"] = idx
                mapping["hersteller_header"] = raw_headers[idx-1]
            if "absatz" not in mapping:
                h_compact_abs = header.replace(" ", "")
                if (
                    header in ABSATZ_ALIASES
                    or h_compact_abs in {a.replace(" ", "") for a in ABSATZ_ALIASES}
                    or header.startswith("abs.")
                    or "abg.pack" in header
                    or "abg pack" in header
                    or ("abverkauf" in header and ("6" in header or "monat" in header))
                    or ("verbrauch" in header and ("6" in header or "monat" in header))
                ):
                    mapping["absatz"] = idx
                    mapping["absatz_header"] = raw_headers[idx-1]
            if "ek" not in mapping:
                h_compact = header.replace(" ", "")
                if header in EK_ALIASES or h_compact in {a.replace(" ", "") for a in EK_ALIASES}:
                    mapping["ek"] = idx
                    mapping["ek_header"] = raw_headers[idx-1]
                elif ("ek" in header and "pzn" not in header) or "einkauf" in header or "taxe" in header or "lauer" in header:
                    mapping["ek"] = idx
                    mapping["ek_header"] = raw_headers[idx-1]
        score = int("pzn" in mapping) * 4 + int("artikel" in mapping) + int("df" in mapping) + int("packung" in mapping) + int("hersteller" in mapping) + int("ek" in mapping) + int("absatz" in mapping)
        if score and (best is None or score > best[0]):
            best = (score, mapping)
        if score >= 5:
            return mapping
    return best[1] if best else None



def _looks_like_month_header(value):
    """Erkennt Monatsüberschriften im Format YYYYMM, z.B. 202606."""
    if value is None:
        return False
    text = str(value).strip()
    if text.endswith('.0'):
        text = text[:-2]
    return bool(re.fullmatch(r"20\d{2}(0[1-9]|1[0-2])", text))


def detect_monthly_consumption_layout(ws, scan_rows=10):
    """Erkennt Monatsverbrauchslisten wie EKM.

    Typisches Layout:
    A=PZN, danach mehrere Monats-Spalten YYYYMM,
    später Artikelname/Artbez., Hersteller, 6 Monate, 12 Monate.
    Manuelle Auswertungsspalten nach 12 Monate werden NICHT übernommen.
    """
    best = None
    for row_idx in range(1, min(scan_rows, ws.max_row) + 1):
        raw = [c.value for c in ws[row_idx]]
        norm = [norm_header(v) for v in raw]
        month_cols = [i + 1 for i, v in enumerate(raw) if _looks_like_month_header(v)]
        if len(month_cols) < 3:
            continue
        mapping = {"header_row": row_idx, "headers": raw, "format": "monatsverbrauch", "month_cols": month_cols}
        for idx, h in enumerate(norm, start=1):
            hc = h.replace(" ", "")
            if "pzn" not in mapping and (h == "pzn" or "pharmazentral" in h):
                mapping["pzn"] = idx; mapping["pzn_header"] = raw[idx-1]
            if "artikel" not in mapping and ("artbez" in hc or "artikel" in h or "bezeichnung" in h or "langname" in h):
                mapping["artikel"] = idx; mapping["artikel_header"] = raw[idx-1]
            if "hersteller" not in mapping and (h in HERSTELLER_ALIASES or h.startswith("herst") or "anbieter" in h):
                mapping["hersteller"] = idx; mapping["hersteller_header"] = raw[idx-1]
            # Absatzspalten in Hr.-Marx/EKM-Dateien sind nicht immer gleich benannt.
            # Beispiele: "Abverkauf 6 Monate", "Abverkäufe 6 Monate",
            # gelegentlich auch mit Tippfehlern. Wichtig ist: 6 Monate = Rohdaten-Absatz.
            if "absatz" not in mapping:
                if ("6" in hc and "monat" in hc) and (
                    "abverk" in h or "verkauf" in h or "absatz" in h or hc in {"6monate", "6monat", "sechsmonate"}
                ):
                    mapping["absatz"] = idx; mapping["absatz_header"] = raw[idx-1]
            if "absatz_12" not in mapping:
                if ("12" in hc and "monat" in hc) and (
                    "abverk" in h or "verkauf" in h or "absatz" in h or hc in {"12monate", "12monat", "zwoelfmonate"}
                ):
                    mapping["absatz_12"] = idx; mapping["absatz_12_header"] = raw[idx-1]
        # Fallback für genau bekannte EKM/Hr.-Marx-Positionen, falls Header leicht anders ist.
        if "pzn" not in mapping and raw and norm_header(raw[0]) == "pzn":
            mapping["pzn"] = 1; mapping["pzn_header"] = raw[0]
        # Sehr häufiges Layout: A=PZN, B-M=Monate, N=Artikel, O=Hersteller, P=6M, Q=12M.
        # Dieses Fallback greift nur, wenn die Monatsstruktur eindeutig erkannt wurde.
        if len(month_cols) >= 6:
            if "artikel" not in mapping and len(raw) >= 14:
                mapping["artikel"] = 14; mapping["artikel_header"] = raw[13]
            if "hersteller" not in mapping and len(raw) >= 15:
                mapping["hersteller"] = 15; mapping["hersteller_header"] = raw[14]
            if "absatz" not in mapping and len(raw) >= 16:
                mapping["absatz"] = 16; mapping["absatz_header"] = raw[15]
            if "absatz_12" not in mapping and len(raw) >= 17:
                mapping["absatz_12"] = 17; mapping["absatz_12_header"] = raw[16]
        score = int("pzn" in mapping) * 5 + int("artikel" in mapping) * 2 + int("hersteller" in mapping) * 2 + int("absatz" in mapping) * 3 + len(month_cols)
        if score >= 14:
            return mapping
        if best is None or score > best[0]:
            best = (score, mapping)
    return best[1] if best and best[0] >= 10 else None


def _clean_basis_value(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return re.sub(r"\s+", " ", text)


def register_basisdaten(con, pzn, artikelname=None, hersteller=None, df=None, pck=None, quelle=""):
    """Lernt feste PZN-Basisdaten: Artikelname, Herstellerkürzel, DF und PCK.
    Für Hersteller/DF/PCK gewinnt jeweils der häufigste Wert. Artikelname: längster nicht-leerer Wert.
    """
    pzn = pzn_value(pzn)
    if not pzn or not pzn.isdigit():
        return False
    changed = False
    values = {
        "herstellerkuerzel": clean_hersteller(hersteller),
        "df": _clean_basis_value(df),
        "pck": _clean_basis_value(pck),
    }
    for feld, wert in values.items():
        if not wert:
            continue
        con.execute(
            """
            INSERT INTO tbl_pzn_basis_stimmen(pzn, feld, wert, anzahl, quelle, letzte_aktualisierung)
            VALUES (?, ?, ?, 1, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(pzn, feld, wert) DO UPDATE SET
                anzahl=anzahl+1,
                quelle=excluded.quelle,
                letzte_aktualisierung=CURRENT_TIMESTAMP
            """,
            (pzn, feld, wert, quelle),
        )
        changed = True
    def winner(feld):
        row = con.execute(
            """
            SELECT wert, anzahl FROM tbl_pzn_basis_stimmen
            WHERE pzn=? AND feld=?
            ORDER BY anzahl DESC, letzte_aktualisierung DESC
            LIMIT 1
            """,
            (pzn, feld),
        ).fetchone()
        return row[0] if row else None
    artikel = _clean_basis_value(artikelname)
    existing = con.execute("SELECT artikelname FROM tbl_pzn_basisdaten WHERE pzn=?", (pzn,)).fetchone()
    if existing and existing[0] and len(str(existing[0])) >= len(artikel):
        artikel_final = existing[0]
    else:
        artikel_final = artikel or (existing[0] if existing else None)
    herst_final = winner("herstellerkuerzel")
    df_final = winner("df")
    pck_final = winner("pck")
    if any([artikel_final, herst_final, df_final, pck_final]):
        con.execute(
            """
            INSERT INTO tbl_pzn_basisdaten(pzn, artikelname, herstellerkuerzel, df, pck, treffer, quelle, letzte_aktualisierung)
            VALUES (?, ?, ?, ?, ?, 1, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(pzn) DO UPDATE SET
                artikelname=COALESCE(excluded.artikelname, tbl_pzn_basisdaten.artikelname),
                herstellerkuerzel=COALESCE(excluded.herstellerkuerzel, tbl_pzn_basisdaten.herstellerkuerzel),
                df=COALESCE(excluded.df, tbl_pzn_basisdaten.df),
                pck=COALESCE(excluded.pck, tbl_pzn_basisdaten.pck),
                treffer=tbl_pzn_basisdaten.treffer+1,
                quelle=excluded.quelle,
                letzte_aktualisierung=CURRENT_TIMESTAMP
            """,
            (pzn, artikel_final, herst_final, df_final, pck_final, quelle),
        )
        changed = True
    if hersteller:
        register_hersteller(con, pzn, hersteller, quelle)
    return changed


def lookup_basisdaten(con, pzn):
    pzn = pzn_value(pzn)
    if not pzn:
        return {}
    row = con.execute("SELECT artikelname, herstellerkuerzel, df, pck FROM tbl_pzn_basisdaten WHERE pzn=?", (pzn,)).fetchone()
    if not row:
        return {}
    return {"artikelname": row[0], "herstellerkuerzel": row[1], "df": row[2], "pck": row[3]}

def register_hersteller(con, pzn, hersteller, quelle):
    pzn = pzn_value(pzn)
    hersteller = clean_hersteller(hersteller)
    if not pzn or not pzn.isdigit() or not hersteller:
        return False
    con.execute(
        """
        INSERT INTO tbl_hersteller_stimmen(pzn, herstellerkuerzel, anzahl, quelle, letzte_aktualisierung)
        VALUES (?, ?, 1, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(pzn, herstellerkuerzel) DO UPDATE SET
            anzahl=anzahl+1,
            quelle=excluded.quelle,
            letzte_aktualisierung=CURRENT_TIMESTAMP
        """,
        (pzn, hersteller, quelle),
    )
    winner = con.execute(
        """
        SELECT herstellerkuerzel, anzahl FROM tbl_hersteller_stimmen
        WHERE pzn=?
        ORDER BY anzahl DESC, letzte_aktualisierung DESC
        LIMIT 1
        """,
        (pzn,),
    ).fetchone()
    if winner:
        con.execute(
            """
            INSERT INTO tbl_hersteller_lern(pzn, herstellerkuerzel, herstellername, treffer, quelle, letzte_aktualisierung)
            VALUES (?, ?, '', ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(pzn) DO UPDATE SET
                herstellerkuerzel=excluded.herstellerkuerzel,
                treffer=excluded.treffer,
                quelle=excluded.quelle,
                letzte_aktualisierung=CURRENT_TIMESTAMP
            """,
            (pzn, winner[0], winner[1], quelle),
        )
    return True


def register_ek(con, pzn, ek, quelle, spalte=""):
    """Speichert ausschließlich EKs, die in echten Rohdaten erkannt wurden.
    Taxe-EK/APU/HAP aus Stammdatenlisten dürfen hier nicht importiert werden.
    """
    pzn = pzn_value(pzn)
    ek_num = parse_number(ek)
    if not pzn or not pzn.isdigit() or ek_num is None:
        return False
    con.execute(
        """
        INSERT INTO tbl_pzn_ek_rohdaten(pzn, ek, quelle_datei, ek_spalte, importdatum)
        VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
        """,
        (pzn, ek_num, quelle, str(spalte or "")),
    )
    return True


def lookup_hersteller(con, pzn):
    pzn = pzn_value(pzn)
    if not pzn:
        return ""
    row = con.execute("SELECT herstellerkuerzel FROM tbl_hersteller_lern WHERE pzn=?", (pzn,)).fetchone()
    return row[0] if row else ""


def lookup_latest_ek(con, pzn):
    pzn = pzn_value(pzn)
    if not pzn:
        return None
    row = con.execute(
        """
        SELECT ek FROM tbl_pzn_ek_rohdaten
        WHERE pzn=?
        ORDER BY importdatum DESC, id DESC
        LIMIT 1
        """,
        (pzn,),
    ).fetchone()
    return row[0] if row else None


def _norm_csv_header_row(row):
    return [norm_header(v) for v in row]


def _find_columns_from_values(rows, scan_rows=15):
    best = None
    for row_idx, raw_headers in enumerate(rows[:scan_rows], start=1):
        headers = _norm_csv_header_row(raw_headers)
        mapping = {"header_row": row_idx, "headers": raw_headers}
        for idx, header in enumerate(headers, start=1):
            if not header:
                continue
            if "pzn" not in mapping and (header in PZN_ALIASES or header == "pzn" or "pharmazentral" in header):
                mapping["pzn"] = idx; mapping["pzn_header"] = raw_headers[idx-1]
            if "artikel" not in mapping and (header in ARTIKEL_ALIASES or "artikelbezeichnung" in header):
                mapping["artikel"] = idx; mapping["artikel_header"] = raw_headers[idx-1]
            if "hersteller" not in mapping and (header in HERSTELLER_ALIASES or header.startswith("herst")):
                mapping["hersteller"] = idx; mapping["hersteller_header"] = raw_headers[idx-1]
            if "df" not in mapping and header in DF_ALIASES:
                mapping["df"] = idx; mapping["df_header"] = raw_headers[idx-1]
            if "packung" not in mapping and (header in PACK_ALIASES or header.startswith("pack")):
                mapping["packung"] = idx; mapping["packung_header"] = raw_headers[idx-1]
            if "ek" not in mapping:
                h_compact = header.replace(" ", "")
                if header in EK_ALIASES or h_compact in {a.replace(" ", "") for a in EK_ALIASES}:
                    mapping["ek"] = idx; mapping["ek_header"] = raw_headers[idx-1]
                elif ("ek" in header and "pzn" not in header) or "einkauf" in header or "taxe" in header or "lauer" in header:
                    mapping["ek"] = idx; mapping["ek_header"] = raw_headers[idx-1]
        score = int("pzn" in mapping) * 4 + int("hersteller" in mapping) + int("ek" in mapping)
        if score and (best is None or score > best[0]):
            best = (score, mapping)
        if score >= 5:
            return mapping
    return best[1] if best else None


def import_learning_list(path: str | Path) -> dict:
    """Importiert aus Excel- oder CSV-Liste PZN-Herstellerkürzel und/oder PZN-EK."""
    path = Path(path)
    stats = defaultdict(int)
    with connect() as con:
        if path.suffix.lower() == ".csv":
            raw = path.read_text(encoding="utf-8-sig", errors="replace")
            try:
                dialect = csv.Sniffer().sniff(raw[:4096], delimiters=";,\t,")
            except Exception:
                dialect = csv.excel
                dialect.delimiter = ";"
            rows = list(csv.reader(raw.splitlines(), dialect))
            mapping = _find_columns_from_values(rows)
            if not mapping or "pzn" not in mapping:
                return {"skipped": len(rows)}
            con.execute(
                """
                INSERT INTO tbl_rohdaten_mapping(dateiname, pzn_spalte, hersteller_spalte, ek_spalte, header_zeile, letzte_aktualisierung)
                VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(dateiname) DO UPDATE SET
                    pzn_spalte=excluded.pzn_spalte,
                    hersteller_spalte=excluded.hersteller_spalte,
                    ek_spalte=excluded.ek_spalte,
                    header_zeile=excluded.header_zeile,
                    letzte_aktualisierung=CURRENT_TIMESTAMP
                """,
                (path.name, str(mapping.get("pzn_header") or ""), str(mapping.get("hersteller_header") or ""), str(mapping.get("ek_header") or ""), mapping["header_row"]),
            )
            pzn_col = mapping["pzn"]; herst_col = mapping.get("hersteller"); ek_col = mapping.get("ek"); artikel_col = mapping.get("artikel"); df_col = mapping.get("df"); pck_col = mapping.get("packung")
            for row in rows[mapping["header_row"]:]:
                pzn = pzn_value(row[pzn_col - 1] if len(row) >= pzn_col else None)
                if not pzn or not pzn.isdigit():
                    stats["skipped"] += 1; continue
                changed = False
                if register_basisdaten(
                    con, pzn,
                    row[artikel_col - 1] if artikel_col and len(row) >= artikel_col else None,
                    row[herst_col - 1] if herst_col and len(row) >= herst_col else None,
                    row[df_col - 1] if df_col and len(row) >= df_col else None,
                    row[pck_col - 1] if pck_col and len(row) >= pck_col else None,
                    path.name,
                ):
                    stats["basisdaten"] += 1; changed = True
                if herst_col:
                    stats["hersteller"] += 1
                if ek_col and register_ek(con, pzn, row[ek_col - 1] if len(row) >= ek_col else None, path.name, mapping.get("ek_header")):
                    stats["ek"] += 1; changed = True
                if changed: stats["rows"] += 1
                else: stats["skipped"] += 1
        else:
            wb = load_workbook(path, data_only=True, read_only=True)
            for ws in wb.worksheets:
                mapping = find_columns(ws)
                if not mapping or "pzn" not in mapping:
                    continue
                pzn_col = mapping["pzn"]
                herst_col = mapping.get("hersteller")
                ek_col = mapping.get("ek")
                artikel_col = mapping.get("artikel")
                df_col = mapping.get("df")
                pck_col = mapping.get("packung")
                start = mapping["header_row"] + 1
                con.execute(
                    """
                    INSERT INTO tbl_rohdaten_mapping(dateiname, pzn_spalte, hersteller_spalte, ek_spalte, header_zeile, letzte_aktualisierung)
                    VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                    ON CONFLICT(dateiname) DO UPDATE SET
                        pzn_spalte=excluded.pzn_spalte,
                        hersteller_spalte=excluded.hersteller_spalte,
                        ek_spalte=excluded.ek_spalte,
                        header_zeile=excluded.header_zeile,
                        letzte_aktualisierung=CURRENT_TIMESTAMP
                    """,
                    (path.name, str(mapping.get("pzn_header") or ""), str(mapping.get("hersteller_header") or ""), str(mapping.get("ek_header") or ""), mapping["header_row"]),
                )
                for row in ws.iter_rows(min_row=start, values_only=True):
                    pzn = pzn_value(row[pzn_col - 1] if len(row) >= pzn_col else None)
                    if not pzn or not pzn.isdigit():
                        stats["skipped"] += 1
                        continue
                    changed = False
                    if register_basisdaten(
                        con, pzn,
                        row[artikel_col - 1] if artikel_col and len(row) >= artikel_col else None,
                        row[herst_col - 1] if herst_col and len(row) >= herst_col else None,
                        row[df_col - 1] if df_col and len(row) >= df_col else None,
                        row[pck_col - 1] if pck_col and len(row) >= pck_col else None,
                        path.name,
                    ):
                        stats["basisdaten"] += 1; changed = True
                    if herst_col:
                        stats["hersteller"] += 1
                    if ek_col:
                        ek = row[ek_col - 1] if len(row) >= ek_col else None
                        if register_ek(con, pzn, ek, path.name, mapping.get("ek_header")):
                            stats["ek"] += 1; changed = True
                    if changed: stats["rows"] += 1
                    else: stats["skipped"] += 1
        con.commit()
    return dict(stats)


# --- Version 1.4: Lernimport aus fertigen / händisch geprüften Auswertungen ---
EVAL_ALIASES = {
    "im_sortiment": {"imsortiment"},
    "nmg_pzn": {"pznnmg"},
    "apu_nmg": {"apunmg", "apu"},
    "rabatt": {"nmgrabatt", "rabattnmg", "rabatt"},
    "lieferbar": {"lieferbar"},
    "bevorratung_angeraten": {"bevorratungangeraten"},
    "liefervorschlag": {"liefervorschlag"},
    "austauschbar_gegen": {"austauschbargegen"},
}

NEGATIVE_AUSTAUSCH_TEXTE = (
    "manuell", "prüfen", "pruefen", "keine zuordnung", "austauschdatenbank ergänzen", "austauschdatenbank ergaenzen"
)

def _compact_header(value):
    text = str(value or "").strip().lower()
    text = text.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    return re.sub(r"[^a-z0-9]+", "", text)


def _find_checked_columns(ws, scan_rows=15):
    best = None
    for row_idx in range(1, min(scan_rows, ws.max_row) + 1):
        rows_here = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        if not rows_here:
            continue
        raw = list(rows_here[0])
        comp = [_compact_header(v) for v in raw]
        mapping = {"header_row": row_idx, "headers": raw}
        # Basis-Spalten
        for idx, h in enumerate(comp, start=1):
            if not h:
                continue
            if "pzn" not in mapping and (h == "pzn" or h.endswith("pzn") or "pharmazentral" in h):
                mapping["pzn"] = idx; mapping["pzn_header"] = raw[idx-1]
            if "artikel" not in mapping and ("artikel" in h or "bezeichnung" in h or h == "name"):
                mapping["artikel"] = idx; mapping["artikel_header"] = raw[idx-1]
            if "hersteller" not in mapping and (h in {"herst", "hersteller", "anbieter", "lieferant", "firma"} or h.startswith("herst")):
                mapping["hersteller"] = idx; mapping["hersteller_header"] = raw[idx-1]
            if "df" not in mapping and h in {"df", "dar", "darreichungsform"}:
                mapping["df"] = idx; mapping["df_header"] = raw[idx-1]
            if "packung" not in mapping and h in {"pck", "packgr", "packung", "packungsgroesse", "packungsgröße"}:
                mapping["packung"] = idx; mapping["packung_header"] = raw[idx-1]
        # Auswertungs-Spalten mit engeren Regeln, damit PZN NMG nicht PZN überschreibt
        for idx, h in enumerate(comp, start=1):
            for key, aliases in EVAL_ALIASES.items():
                if key not in mapping and h in aliases:
                    mapping[key] = idx; mapping[f"{key}_header"] = raw[idx-1]
        score = int("pzn" in mapping) * 5 + sum(1 for k in EVAL_ALIASES if k in mapping) + int("hersteller" in mapping)
        if score and (best is None or score > best[0]):
            best = (score, mapping)
        if "pzn" in mapping and ("nmg_pzn" in mapping or "austauschbar_gegen" in mapping or "im_sortiment" in mapping):
            return mapping
    return best[1] if best else None


def _cell(row, mapping, key):
    col = mapping.get(key)
    if not col or len(row) < col:
        return None
    return row[col - 1]


def _valid_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text == "0":
        return ""
    return text


def _is_negative_austausch(text):
    low = str(text or "").lower()
    return any(tok in low for tok in NEGATIVE_AUSTAUSCH_TEXTE)


def import_checked_auswertung(path: str | Path) -> dict:
    """Lernt aus einer fertigen/kontrollierten Auswertung Hersteller, H-O, Austausch und Lieferfähigkeit.

    Die Funktion ist absichtlich konservativ: Negative Platzhalter wie "Manuell prüfen" werden nicht als
    Austauschartikel gelernt. Ein Austausch wird nur gespeichert, wenn eine NMG-PZN vorhanden ist oder ein
    sinnvoller Austauschtext ohne Negativmarker vorliegt.
    """
    path = Path(path)
    stats = defaultdict(int)
    wb = load_workbook(path, data_only=True, read_only=True)
    with connect() as con:
        for ws in wb.worksheets:
            mapping = _find_checked_columns(ws)
            if not mapping or "pzn" not in mapping:
                continue
            stats["sheets"] += 1
            start = mapping["header_row"] + 1
            for row in ws.iter_rows(min_row=start, values_only=True):
                original_pzn = pzn_value(_cell(row, mapping, "pzn"))
                if not original_pzn:
                    stats["skipped"] += 1
                    continue
                stats["rows"] += 1
                artikel = _valid_text(_cell(row, mapping, "artikel"))
                herst = _cell(row, mapping, "hersteller")
                if register_basisdaten(con, original_pzn, artikel, herst, _cell(row, mapping, "df"), _cell(row, mapping, "packung"), f"Handprüfung:{path.name}"):
                    stats["basisdaten"] += 1
                if herst:
                    stats["hersteller"] += 1

                im_sortiment = _valid_text(_cell(row, mapping, "im_sortiment")) or None
                nmg_pzn = pzn_value(_cell(row, mapping, "nmg_pzn"))
                apu_nmg = parse_number(_cell(row, mapping, "apu_nmg"))
                rabatt = parse_number(_cell(row, mapping, "rabatt"))
                # Falls Rabatt als Prozent-Zahl 10 statt 0,10 geliefert wird, konservativ normalisieren.
                if rabatt is not None and rabatt > 1:
                    rabatt = rabatt / 100.0
                lieferbar = _valid_text(_cell(row, mapping, "lieferbar")) or None
                bev = _valid_text(_cell(row, mapping, "bevorratung_angeraten")) or None
                liefervorschlag = _valid_text(_cell(row, mapping, "liefervorschlag")) or None
                austext = _valid_text(_cell(row, mapping, "austauschbar_gegen")) or None

                meaningful_austext = bool(austext and not _is_negative_austausch(austext))
                has_eval = any(v not in (None, "") for v in [nmg_pzn, apu_nmg, rabatt, lieferbar, bev]) or meaningful_austext
                # Negative Platzhalter ohne echte NMG-Zuordnung nicht in H-O übernehmen.
                if has_eval:
                    # H-O Referenz lernen. Diese Tabelle hat im Export die höchste Priorität.
                    con.execute(
                        """
                        INSERT INTO tbl_referenz_h_o(original_pzn, im_sortiment, nmg_pzn, apu_nmg, rabatt, lieferbar, bevorratung_angeraten, liefervorschlag, austauschbar_gegen)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(original_pzn) DO UPDATE SET
                            im_sortiment=COALESCE(excluded.im_sortiment, tbl_referenz_h_o.im_sortiment),
                            nmg_pzn=COALESCE(excluded.nmg_pzn, tbl_referenz_h_o.nmg_pzn),
                            apu_nmg=COALESCE(excluded.apu_nmg, tbl_referenz_h_o.apu_nmg),
                            rabatt=COALESCE(excluded.rabatt, tbl_referenz_h_o.rabatt),
                            lieferbar=COALESCE(excluded.lieferbar, tbl_referenz_h_o.lieferbar),
                            bevorratung_angeraten=COALESCE(excluded.bevorratung_angeraten, tbl_referenz_h_o.bevorratung_angeraten),
                            liefervorschlag=COALESCE(excluded.liefervorschlag, tbl_referenz_h_o.liefervorschlag),
                            austauschbar_gegen=COALESCE(excluded.austauschbar_gegen, tbl_referenz_h_o.austauschbar_gegen)
                        """,
                        (original_pzn, im_sortiment, nmg_pzn or None, apu_nmg, rabatt, lieferbar, bev, liefervorschlag, austext),
                    )
                    stats["referenz_h_o"] += 1

                if nmg_pzn or meaningful_austext:
                    bemerkung = "gelernt aus Handprüfung"
                    con.execute(
                        """
                        INSERT INTO tbl_austauschartikel(original_pzn, original_artikel, nmg_pzn, austauschbar_gegen, quelle, letzte_aktualisierung, treffer_anzahl, bemerkung)
                        VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP, 1, ?)
                        ON CONFLICT(original_pzn) DO UPDATE SET
                            original_artikel=COALESCE(excluded.original_artikel, tbl_austauschartikel.original_artikel),
                            nmg_pzn=COALESCE(excluded.nmg_pzn, tbl_austauschartikel.nmg_pzn),
                            austauschbar_gegen=COALESCE(excluded.austauschbar_gegen, tbl_austauschartikel.austauschbar_gegen),
                            quelle=excluded.quelle,
                            letzte_aktualisierung=CURRENT_TIMESTAMP,
                            treffer_anzahl=COALESCE(tbl_austauschartikel.treffer_anzahl,0)+1,
                            bemerkung=excluded.bemerkung
                        """,
                        (original_pzn, artikel, nmg_pzn or None, austext, f"Handprüfung:{path.name}", bemerkung),
                    )
                    stats["austausch"] += 1

                if nmg_pzn and any(v for v in [lieferbar, bev, liefervorschlag]):
                    con.execute(
                        """
                        INSERT INTO tbl_lieferfaehigkeit(nmg_pzn, lieferbar, bevorratung_angeraten, liefervorschlag, quelle, letzte_aktualisierung)
                        VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                        ON CONFLICT(nmg_pzn) DO UPDATE SET
                            lieferbar=COALESCE(excluded.lieferbar, tbl_lieferfaehigkeit.lieferbar),
                            bevorratung_angeraten=COALESCE(excluded.bevorratung_angeraten, tbl_lieferfaehigkeit.bevorratung_angeraten),
                            liefervorschlag=COALESCE(excluded.liefervorschlag, tbl_lieferfaehigkeit.liefervorschlag),
                            quelle=excluded.quelle,
                            letzte_aktualisierung=CURRENT_TIMESTAMP
                        """,
                        (nmg_pzn, lieferbar, bev, liefervorschlag, f"Handprüfung:{path.name}"),
                    )
                    stats["lieferfaehigkeit"] += 1

                if has_eval:
                    con.execute(
                        """
                        INSERT INTO tbl_lernhistorie(zeitpunkt, quelle, original_pzn, original_artikel, nmg_pzn, austauschbar_gegen, aktion, status)
                        VALUES (CURRENT_TIMESTAMP, ?, ?, ?, ?, ?, 'Handauswertung gelernt', 'ok')
                        """,
                        (path.name, original_pzn, artikel, nmg_pzn or '', austext or ''),
                    )
        con.execute(
            "INSERT INTO tbl_import_log(datei, typ, datensaetze, meldung) VALUES (?, 'handauswertung', ?, ?)",
            (path.name, stats["rows"], f"{stats['referenz_h_o']} H-O, {stats['austausch']} Austausch, {stats['lieferfaehigkeit']} Lieferfähigkeit gelernt"),
        )
        con.commit()
    return dict(stats)
