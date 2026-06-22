from pathlib import Path
from datetime import datetime
import re
import sqlite3
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .config import DB_PATH, OUTPUT_DIR
from .austausch_db import ensure_austauschdatenbank_table


def _num(value):
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def _safe(text):
    return str(text or "").strip()


def _norm_text(value):
    text = str(value or "").strip().upper()
    text = text.replace("Ä", "AE").replace("Ö", "OE").replace("Ü", "UE").replace("ß", "SS")
    return re.sub(r"[^A-Z0-9]+", " ", text).strip()


def _product_key(value):
    """Grobe Produktkennung fuer den Abgleich gegen NMG-Artikelname.

    Zweck: Wirkstoff-/Markennamen wie DUPIXENT oder STELARA sollen nicht als
    'keine NMG ohne Austausch' erscheinen, wenn NMG produktmaessig denselben
    Artikel im Sortiment hat, auch wenn die PZN abweicht.
    """
    norm = _norm_text(value)
    if not norm:
        return ""
    stop = {"N", "PZN", "MG", "ML", "ST", "FTA", "TAB", "KAP", "LOESUNG", "INJ", "FS", "FERTIGSPRITZE"}
    for token in norm.split():
        if len(token) >= 4 and token not in stop and not token.isdigit():
            return token
    return ""


def _datenquelle_clause(datenquelle, alias="p"):
    dq = (datenquelle or "ALLE").upper()
    if dq in ("NMG", "PK"):
        return f"COALESCE({alias}.datenquelle,'NMG') = 'NMG'", []
    if dq == "ZF":
        return f"COALESCE({alias}.datenquelle,'NMG') = 'ZF'", []
    return "1=1", []


def _analysis_clause(auswertung_id, alias="p"):
    if auswertung_id:
        return f"{alias}.auswertung_id = ?", [int(auswertung_id)]
    return "1=1", []


def _fetch_nmg_product_keys():
    with sqlite3.connect(DB_PATH) as con:
        rows = con.execute("SELECT artikelname FROM tbl_nmg_stamm WHERE COALESCE(artikelname,'') <> ''").fetchall()
    return {k for (name,) in rows if (k := _product_key(name))}


def _fetch_positions(datenquelle="ALLE", auswertung_id=None):
    dq_sql, dq_params = _datenquelle_clause(datenquelle, "p")
    id_sql, id_params = _analysis_clause(auswertung_id, "p")
    sql = f"""
        SELECT
            p.pzn,
            COALESCE(
                MAX(NULLIF(ast.artikel, '')),
                MAX(NULLIF(b.artikelname, '')),
                MAX(NULLIF(n.artikelname, '')),
                MAX(NULLIF(p.artikelname, '')),
                ''
            ) AS artikel,
            COALESCE(
                MAX(NULLIF(ast.df, '')),
                MAX(NULLIF(b.df, '')),
                MAX(CASE
                    WHEN NULLIF(p.df, '') IS NOT NULL
                     AND LENGTH(TRIM(p.df)) <= 12
                     AND LOWER(p.df) NOT LIKE '%gmbh%'
                     AND LOWER(p.df) NOT LIKE '%pharma%'
                     AND LOWER(p.df) NOT LIKE '% kg%'
                    THEN p.df ELSE NULL END),
                ''
            ) AS df,
            COALESCE(
                MAX(NULLIF(ast.pck, '')),
                MAX(NULLIF(b.pck, '')),
                MAX(CASE
                    WHEN NULLIF(p.pck, '') IS NOT NULL
                     AND LOWER(p.pck) NOT LIKE '%gmbh%'
                     AND LOWER(p.pck) NOT LIKE '%pharma%'
                    THEN p.pck ELSE NULL END),
                ''
            ) AS pck,
            COALESCE(
                MAX(NULLIF(h.herstellerkuerzel, '')),
                MAX(NULLIF(ast.herst, '')),
                MAX(NULLIF(b.herstellerkuerzel, '')),
                MAX(NULLIF(n.herstellerkuerzel, '')),
                MAX(NULLIF(p.herstellerkuerzel, '')),
                ''
            ) AS herst,
            SUM(COALESCE(p.absatz_6m, 0)) AS menge,
            MAX(COALESCE(p.ist_nmg_treffer, 0)) AS ist_nmg_treffer,
            MAX(CASE WHEN n.pzn IS NOT NULL THEN 1 ELSE 0 END) AS ist_nmg_stamm,
            MAX(CASE
                WHEN adb.id IS NOT NULL THEN 1
                WHEN alt.original_pzn IS NOT NULL THEN 1
                WHEN ref.original_pzn IS NOT NULL AND COALESCE(ref.nmg_pzn,'') <> '' THEN 1
                WHEN COALESCE(p.pzn_nmg,'') <> '' THEN 1
                WHEN COALESCE(p.austauschbar_gegen,'') <> '' THEN 1
                ELSE 0
            END) AS hat_austausch,
            MAX(COALESCE(adb.pzn_nmg, alt.nmg_pzn, ref.nmg_pzn, p.pzn_nmg, '')) AS pzn_nmg,
            MAX(COALESCE(adb.freitext_austausch, alt.austauschbar_gegen, ref.austauschbar_gegen, p.austauschbar_gegen, '')) AS austauschbar_gegen
        FROM tbl_auswertungspositionen p
        LEFT JOIN tbl_artikelstamm ast ON ast.pzn = p.pzn
        LEFT JOIN tbl_pzn_basisdaten b ON b.pzn = p.pzn
        LEFT JOIN tbl_hersteller_lern h ON h.pzn = p.pzn
        LEFT JOIN tbl_nmg_stamm n ON n.pzn = p.pzn
        LEFT JOIN tbl_austauschdatenbank adb
            ON adb.pzn_alt = p.pzn AND COALESCE(adb.status,'aktiv') = 'aktiv'
        LEFT JOIN tbl_austauschartikel alt ON alt.original_pzn = p.pzn
        LEFT JOIN tbl_referenz_h_o ref ON ref.original_pzn = p.pzn
        WHERE {dq_sql}
          AND {id_sql}
          AND COALESCE(p.absatz_6m, 0) > 0
          AND p.pzn IS NOT NULL AND p.pzn <> ''
        GROUP BY p.pzn
    """
    with sqlite3.connect(DB_PATH) as con:
        con.row_factory = sqlite3.Row
        return con.execute(sql, dq_params + id_params).fetchall()


def _append_header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _autosize(ws):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 10
        for cell in ws[letter]:
            width = max(width, min(45, len(str(cell.value or "")) + 2))
        ws.column_dimensions[letter].width = width


def _write_rows(ws, rows, include_exchange=False):
    headers = ["PZN", "Artikel", "DF", "PCK", "Herst", "Menge"]
    if include_exchange:
        headers += ["PZN NMG", "austauschbar gegen"]
    _append_header(ws, headers)
    for r in rows:
        line = [
            _safe(r["pzn"]), _safe(r["artikel"]), _safe(r["df"]), _safe(r["pck"]),
            _safe(r["herst"]), _num(r["menge"]),
        ]
        if include_exchange:
            line += [_safe(r["pzn_nmg"]), _safe(r["austauschbar_gegen"])]
        ws.append(line)
    _autosize(ws)


def export_produktanalyse_neu(limit=500, min_apotheken=1, datenquelle="ALLE", auswertung_id=None):
    """Erstellt die neue Produktanalyse mit 3 Reitern.

    Reiter 1: Nicht-NMG ohne vorhandenen Austausch und ohne NMG-Artikelgleichheit.
    Reiter 2: Nicht-NMG mit Abgabemenge und NMG-Austausch.
    Reiter 3: Herstellerabsatz komplett, nach PZN summiert und Datenquelle gefiltert.
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ensure_austauschdatenbank_table()
    rows = list(_fetch_positions(datenquelle=datenquelle, auswertung_id=auswertung_id))
    nmg_product_keys = _fetch_nmg_product_keys()

    nicht_nmg = [r for r in rows if int(r["ist_nmg_treffer"] or 0) == 0 and int(r["ist_nmg_stamm"] or 0) == 0]
    ohne_austausch = [
        r for r in nicht_nmg
        if int(r["hat_austausch"] or 0) == 0
        and (_product_key(r["artikel"]) not in nmg_product_keys)
    ]
    mit_austausch = [r for r in nicht_nmg if int(r["hat_austausch"] or 0) == 1]

    hersteller_absatz = sorted(rows, key=lambda r: (_safe(r["herst"]).lower(), _safe(r["artikel"]).lower(), _safe(r["pzn"])))
    ohne_austausch = sorted(ohne_austausch, key=lambda r: _num(r["menge"]), reverse=True)
    mit_austausch = sorted(mit_austausch, key=lambda r: _num(r["menge"]), reverse=True)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "1 Keine NMG ohne Austausch"
    _write_rows(ws1, ohne_austausch[:limit] if limit else ohne_austausch, include_exchange=False)

    ws2 = wb.create_sheet("2 Austausch zu NMG")
    _write_rows(ws2, mit_austausch[:limit] if limit else mit_austausch, include_exchange=True)

    ws3 = wb.create_sheet("3 Hersteller Absatz")
    _write_rows(ws3, hersteller_absatz, include_exchange=False)

    label = {"NMG": "PK", "PK": "PK", "ZF": "ZW", "ALLE": "PK_ZW"}.get((datenquelle or "ALLE").upper(), str(datenquelle or "ALLE"))
    suffix = f"_{int(auswertung_id)}" if auswertung_id else ""
    out = OUTPUT_DIR / f"Produktanalyse_{label}{suffix}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    wb.save(out)
    return out
