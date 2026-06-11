from pathlib import Path
from datetime import datetime
import getpass
import sqlite3

from openpyxl import load_workbook
from .file_loader import load_worksheet

from .config import DB_PATH


def _clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return " ".join(text.split())


def _pzn(value):
    text = _clean(value)
    if not text:
        return ""
    digits = "".join(ch for ch in text if ch.isdigit())
    if digits:
        return digits.zfill(8)
    return text


def _norm_header(value):
    return _clean(value).lower().replace(" ", "_").replace("-", "_")


def ensure_austauschdatenbank_table():
    with sqlite3.connect(DB_PATH) as con:
        con.execute("""
            CREATE TABLE IF NOT EXISTS tbl_austauschdatenbank (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pzn_alt TEXT,
                artikel_alt TEXT,
                pzn_nmg TEXT,
                artikel_nmg TEXT,
                freitext_austausch TEXT NOT NULL,
                quelle TEXT NOT NULL DEFAULT 'Manuell',
                status TEXT NOT NULL DEFAULT 'aktiv',
                erstellt_am TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                aktualisiert_am TEXT,
                bearbeiter TEXT,
                gueltig_ab TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                gueltig_bis TEXT,
                ersetzt_durch_id INTEGER,
                bemerkung TEXT
            )
        """)
        cols = {row[1] for row in con.execute("PRAGMA table_info(tbl_austauschdatenbank)").fetchall()}
        for col, definition in {
            "bearbeiter": "TEXT",
            "gueltig_ab": "TEXT",
            "gueltig_bis": "TEXT",
            "ersetzt_durch_id": "INTEGER",
            "bemerkung": "TEXT",
        }.items():
            if col not in cols:
                con.execute(f"ALTER TABLE tbl_austauschdatenbank ADD COLUMN {col} {definition}")
        con.execute("""
            UPDATE tbl_austauschdatenbank
            SET gueltig_ab = COALESCE(gueltig_ab, erstellt_am, CURRENT_TIMESTAMP)
            WHERE gueltig_ab IS NULL OR gueltig_ab = ''
        """)
        con.execute("""
            CREATE INDEX IF NOT EXISTS idx_austauschdatenbank_pzn_alt
            ON tbl_austauschdatenbank (pzn_alt)
        """)
        con.execute("""
            CREATE INDEX IF NOT EXISTS idx_austauschdatenbank_pzn_nmg
            ON tbl_austauschdatenbank (pzn_nmg)
        """)
        con.commit()


def _find_column(header_map, possible_names):
    for name in possible_names:
        key = _norm_header(name)
        if key in header_map:
            return header_map[key]
    return None


def import_austausch_excel(file_path, quelle="Excel-Import"):
    """
    Importiert eine Austauschdatenbank aus einer Excel-Datei.

    Erwartete Minimalspalten:
    - PZN
    - PZN NMG
    - austauschbar gegen

    Es werden keine vorhandenen Einträge doppelt angelegt.
    Wenn ein vorhandener Eintrag gefunden wird, wird er aktualisiert.
    """
    ensure_austauschdatenbank_table()

    path = Path(file_path)
    ws, input_typ = load_worksheet(path)

    header_map = {}
    for col in range(1, ws.max_column + 1):
        header = _norm_header(ws.cell(1, col).value)
        if header:
            header_map[header] = col

    col_pzn_alt = _find_column(header_map, ["PZN", "PZN alt", "PZN verkauft", "verkaufte PZN"])
    col_pzn_nmg = _find_column(header_map, ["PZN NMG", "PZN Austausch", "PZN austauschbar", "PZN neu"])
    col_freitext = _find_column(header_map, ["austauschbar gegen", "Austausch", "Austauschartikel", "Artikel"])

    missing = []
    if col_pzn_alt is None:
        missing.append("PZN")
    if col_freitext is None:
        missing.append("austauschbar gegen")
    if missing:
        raise ValueError(
            "Austauschdatenbank konnte nicht importiert werden. Fehlende Spalte(n): "
            + ", ".join(missing)
        )

    inserted = 0
    updated = 0
    skipped = 0

    with sqlite3.connect(DB_PATH) as con:
        for row in range(2, ws.max_row + 1):
            pzn_alt = _pzn(ws.cell(row, col_pzn_alt).value)
            pzn_nmg = _pzn(ws.cell(row, col_pzn_nmg).value) if col_pzn_nmg else ""
            freitext = _clean(ws.cell(row, col_freitext).value)

            if not pzn_alt and not freitext:
                skipped += 1
                continue
            if not freitext:
                skipped += 1
                continue

            existing = con.execute("""
                SELECT id
                FROM tbl_austauschdatenbank
                WHERE COALESCE(pzn_alt, '') = ?
                  AND COALESCE(pzn_nmg, '') = ?
                  AND freitext_austausch = ?
                LIMIT 1
            """, (pzn_alt, pzn_nmg, freitext)).fetchone()

            if existing:
                con.execute("""
                    UPDATE tbl_austauschdatenbank
                    SET quelle = ?,
                        status = 'aktiv',
                        aktualisiert_am = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (quelle, existing[0]))
                updated += 1
            else:
                con.execute("""
                    INSERT INTO tbl_austauschdatenbank (
                        pzn_alt,
                        artikel_alt,
                        pzn_nmg,
                        artikel_nmg,
                        freitext_austausch,
                        quelle,
                        status,
                        erstellt_am
                    )
                    VALUES (?, '', ?, '', ?, ?, 'aktiv', CURRENT_TIMESTAMP)
                """, (pzn_alt, pzn_nmg, freitext, quelle))
                inserted += 1

        con.commit()

    return {
        "file": str(path),
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "total_processed": inserted + updated + skipped,
    }


def count_austauschdatenbank():
    ensure_austauschdatenbank_table()
    with sqlite3.connect(DB_PATH) as con:
        return con.execute("""
            SELECT COUNT(*)
            FROM tbl_austauschdatenbank
            WHERE status = 'aktiv'
        """).fetchone()[0]


def find_austausch_by_pzn(pzn_alt):
    ensure_austauschdatenbank_table()
    pzn_alt = _pzn(pzn_alt)
    if not pzn_alt:
        return []

    with sqlite3.connect(DB_PATH) as con:
        con.row_factory = sqlite3.Row
        return con.execute("""
            SELECT id, pzn_alt, artikel_alt, pzn_nmg, artikel_nmg, freitext_austausch, quelle
            FROM tbl_austauschdatenbank
            WHERE status = 'aktiv'
              AND pzn_alt = ?
            ORDER BY id DESC
        """, (pzn_alt,)).fetchall()


def add_austausch_entry(
    pzn_alt="",
    pzn_nmg="",
    freitext_austausch="",
    quelle="Schulbank",
    artikel_alt="",
    artikel_nmg="",
    bearbeiter=None,
    bemerkung="",
):
    """
    Legt einen aktiven Austauschdatensatz an.

    Grundsatz:
    - alte Einträge werden nicht gelöscht
    - wenn für dieselbe pzn_alt ein anderer aktiver Austausch existiert,
      wird dieser auf inaktiv gesetzt und der neue Datensatz aktiv angelegt
    - doppelte identische aktive Einträge werden nur aktualisiert
    """
    ensure_austauschdatenbank_table()
    pzn_alt = _pzn(pzn_alt)
    pzn_nmg = _pzn(pzn_nmg)
    freitext_austausch = _clean(freitext_austausch)
    artikel_alt = _clean(artikel_alt)
    artikel_nmg = _clean(artikel_nmg)
    bearbeiter = _clean(bearbeiter) or getpass.getuser()
    bemerkung = _clean(bemerkung)

    if not freitext_austausch:
        if artikel_nmg:
            freitext_austausch = artikel_nmg
        elif pzn_nmg:
            freitext_austausch = f"PZN NMG: {pzn_nmg}"
        else:
            raise ValueError("Kein Austauschartikel angegeben.")

    with sqlite3.connect(DB_PATH) as con:
        existing = con.execute("""
            SELECT id
            FROM tbl_austauschdatenbank
            WHERE COALESCE(pzn_alt, '') = ?
              AND COALESCE(pzn_nmg, '') = ?
              AND freitext_austausch = ?
            LIMIT 1
        """, (pzn_alt, pzn_nmg, freitext_austausch)).fetchone()

        if existing:
            con.execute("""
                UPDATE tbl_austauschdatenbank
                SET status = 'aktiv',
                    quelle = ?,
                    aktualisiert_am = CURRENT_TIMESTAMP,
                    bearbeiter = ?,
                    gueltig_bis = NULL,
                    bemerkung = COALESCE(NULLIF(?, ''), bemerkung)
                WHERE id = ?
            """, (quelle, bearbeiter, bemerkung, existing[0]))
            con.commit()
            return {"created": False, "id": existing[0], "reactivated": True}

        cur = con.execute("""
            INSERT INTO tbl_austauschdatenbank (
                pzn_alt,
                artikel_alt,
                pzn_nmg,
                artikel_nmg,
                freitext_austausch,
                quelle,
                status,
                erstellt_am,
                bearbeiter,
                gueltig_ab,
                bemerkung
            )
            VALUES (?, ?, ?, ?, ?, ?, 'aktiv', CURRENT_TIMESTAMP, ?, CURRENT_TIMESTAMP, ?)
        """, (pzn_alt, artikel_alt, pzn_nmg, artikel_nmg, freitext_austausch, quelle, bearbeiter, bemerkung))
        new_id = cur.lastrowid

        # Andere aktive Einträge für dieselbe alte PZN außer Kraft setzen, aber nicht löschen.
        if pzn_alt:
            con.execute("""
                UPDATE tbl_austauschdatenbank
                SET status = 'inaktiv',
                    gueltig_bis = CURRENT_TIMESTAMP,
                    ersetzt_durch_id = ?,
                    aktualisiert_am = CURRENT_TIMESTAMP,
                    bearbeiter = ?,
                    bemerkung = CASE
                        WHEN bemerkung IS NULL OR bemerkung = '' THEN 'Durch neueren Schulbank-Eintrag ersetzt'
                        ELSE bemerkung || ' | Durch neueren Schulbank-Eintrag ersetzt'
                    END
                WHERE pzn_alt = ?
                  AND id <> ?
                  AND status = 'aktiv'
            """, (new_id, bearbeiter, pzn_alt, new_id))

        con.commit()
        return {"created": True, "id": new_id, "reactivated": False}
