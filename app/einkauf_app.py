"""NMG Einkauf-App (eigenstaendig) - Beschaffung EU-Ausland & Margen.

Das "Sorglos-Paket" fuer den Einkaeufer eines Pharma-Parallelimporteurs.
Eigenes Fenster / Taskleisten-Icon (run_standalone), teilt sich die Datenbank
mit NMGone (app/config.py). Kerngedanken:

  * Aufgaben/Wiedervorlagen mit Faelligkeit -> faellige & ueberfaellige Punkte
    erscheinen als Meldungen auf dem Dashboard ("melde dich bei ...").
  * Importquoten- & Margenkalkulation nach §129 SGB V: Import-EK (Fremdwaehrung)
    -> deutscher AVP/Festbetrag -> garantierter Preisabstand + eigene Marge.
  * Lieferanten-/Beschaffungsmodul EU-Ausland: Quellen je Land, EK in
    Fremdwaehrung, Lieferzeit, Mindestabnahme -> Basis fuer Beschaffung & Marge.

Datenmodell siehe migrations.py (tbl_einkauf_*). Optik aus app/theme.py
(gemeinsamer Look mit NMGone/Kasse/Faktura).
"""
from __future__ import annotations

import getpass
from .i18n import T as _T
import json
import math
import os
import re
import sqlite3
import webbrowser
from datetime import datetime, date, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog

from .config import DB_PATH, ASSETS_DIR, OUTPUT_DIR
from . import theme

# Palette aus dem zentralen Theme.
BG = theme.CARD
SHELL_BG = theme.BG
ACCENT = theme.PRIMARY
ACCENT_LIGHT = theme.SELECT_BG
BORDER = theme.BORDER
TEXT = theme.INK
MUTED = theme.MUTED
OK_GREEN = theme.SUCCESS

# §129-Preisabstand (Importarzneimittel) - Standardwerte, in den Einstellungen
# anpassbar. Auslegung der gaengigen Staffel des Rahmenvertrags:
#   AVP bis 100 €              -> Abstand mind. 15 %
#   AVP ueber 100 € bis 300 €  -> Abstand mind. 15 €
#   AVP ueber 300 €            -> Abstand mind. 5 %
P129_DEFAULTS = {
    "p129_t1_grenze": "100",
    "p129_t1_proz": "15",
    "p129_t2_grenze": "300",
    "p129_t2_eur": "15",
    "p129_t3_proz": "5",
}

SETTING_DEFAULTS = {"standard_waehrung": "EUR"}
SETTING_DEFAULTS.update(P129_DEFAULTS)

# Annahmen für die Importkandidaten-Bewertung (in den Einstellungen pflegbar):
#   kandidat_eu_ek_proz   = angenommener EU-Einkaufspreis als % des deutschen Preisniveaus
#   kandidat_min_marge_proz = Schwelle für die §129-Machbarkeits-Ampel (grün ab dieser Marge)
KANDIDAT_SETTINGS = {"kandidat_eu_ek_proz": "65", "kandidat_min_marge_proz": "5",
                     "kandidat_fixkosten": "0"}
SETTING_DEFAULTS.update(KANDIDAT_SETTINGS)

# Pipeline-Status eines Importkandidaten.
KANDIDAT_STATUS = ["Neu", "Angefragt", "Angebot/Muster", "Eingelistet", "Verworfen"]

# Maximal gerenderte Tabellenzeilen (Performance). Filter/Export nutzen alle Treffer.
IK_MAX_ROWS = 400

# Optionale Spalten der Importkandidaten-Tabelle (key, Anzeigename) – in den
# Einstellungen ein-/ausblendbar. PZN/Artikel/Hersteller sind immer sichtbar.
IK_OPT_COLUMNS = [
    ("typ", "Typ"), ("wirkstoff", "Wirkstoff"), ("absatz", "Absatz 6M"),
    ("apo", "Apotheken"), ("konkurrenz", "Konkurrenz"), ("trend", "Trend"),
    ("preis", "Preis"), ("ampel", "§129-Ampel"), ("potenzial", "Potenzial/Jahr"),
    ("status", "Status"),
]


def kandidat_spalten() -> dict[str, bool]:
    """Sichtbarkeit der optionalen Spalten (key->bool). Fehlende = sichtbar."""
    vis = {k: True for k, _ in IK_OPT_COLUMNS}
    raw = get_setting("kandidat_spalten", "")
    if raw:
        try:
            data = json.loads(raw)
            if isinstance(data, dict):
                for k in vis:
                    if k in data:
                        vis[k] = bool(data[k])
        except (ValueError, TypeError):
            pass
    return vis


def set_kandidat_spalten(vis: dict[str, bool]) -> None:
    set_setting("kandidat_spalten", json.dumps({k: bool(v) for k, v in vis.items()},
                                               ensure_ascii=False))

# Start-Wechselkurse (EUR je 1 Einheit der Fremdwaehrung). Werden beim ersten
# Start angelegt und sind danach in der App pflegbar (Tageskurs eintragen).
WECHSELKURS_SEED = {
    "EUR": 1.0, "GBP": 1.17, "CHF": 1.05, "PLN": 0.23, "DKK": 0.134,
    "SEK": 0.088, "NOK": 0.086, "CZK": 0.040, "HUF": 0.0025, "USD": 0.92,
}

AUFGABE_KATEGORIEN = ["Wiedervorlage", "Anfrage", "Bestellung", "Rueckruf",
                      "Reklamation", "Liefertermin", "Sonstiges"]
PRIO_LABEL = {1: "Hoch", 2: "Mittel", 3: "Niedrig"}
PRIO_FARBE = {1: theme.DANGER, 2: theme.WARNING, 3: MUTED}


# ── Helfer ───────────────────────────────────────────────────────────────────
def _con() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH)
    con.execute("PRAGMA foreign_keys = ON")
    return con


def _table_exists(con, name) -> bool:
    return con.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
                       (name,)).fetchone() is not None


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _heute() -> str:
    return date.today().isoformat()


def _parse_datum(s, default_iso: str | None = None) -> str:
    """'2026-06-30' / '30.06.2026' / '' -> ISO-Datum. Leer/ungueltig -> default."""
    s = str(s or "").strip()
    fallback = default_iso if default_iso is not None else ""
    if not s:
        return fallback
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return fallback


def _datum_de(iso: str) -> str:
    """ISO -> '30.06.2026' fuer die Anzeige. Leere/ungueltige Werte -> '—'."""
    iso = str(iso or "").strip()
    if not iso:
        return "—"
    try:
        return datetime.strptime(iso[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
    except ValueError:
        return iso


def _eur(v) -> str:
    if v is None:
        return "—"
    return f"{v:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")


def _parse_num(s) -> float:
    """'1.234,56 €' / '1234.56' / '' -> float. Robust gegen DE/EN-Format."""
    s = str(s or "").replace("€", "").replace("%", "").strip()
    s = s.replace(" ", "")
    if not s:
        return 0.0
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def get_setting(schluessel: str, default: str = "") -> str:
    try:
        with _con() as con:
            row = con.execute("SELECT wert FROM tbl_einkauf_einstellungen WHERE schluessel=?",
                              (schluessel,)).fetchone()
        if row and row[0] is not None:
            return row[0]
    except sqlite3.Error:
        pass
    return SETTING_DEFAULTS.get(schluessel, default)


def set_setting(schluessel: str, wert: str) -> None:
    with _con() as con:
        con.execute("INSERT INTO tbl_einkauf_einstellungen(schluessel, wert) VALUES(?,?) "
                    "ON CONFLICT(schluessel) DO UPDATE SET wert=excluded.wert",
                    (schluessel, str(wert)))
        con.commit()


def _log(aktion: str, details: str = "") -> None:
    try:
        with _con() as con:
            con.execute("INSERT INTO tbl_einkauf_log(bearbeiter, aktion, details) "
                        "VALUES(?,?,?)", (getpass.getuser(), aktion, details))
            con.commit()
    except sqlite3.Error:
        pass


def _ensure_kandidat_status_tabelle(con) -> None:
    """Pipeline-Status + Merkliste je Importkandidat (PZN). Idempotent."""
    con.execute(
        "CREATE TABLE IF NOT EXISTS tbl_einkauf_kandidat_status("
        "pzn TEXT PRIMARY KEY, status TEXT DEFAULT 'Neu', notiz TEXT DEFAULT '', "
        "wiedervorlage TEXT DEFAULT '', beobachtet INTEGER DEFAULT 0, "
        "geaendert_am TEXT, bearbeiter TEXT)")
    # ALTER-Guard für bestehende DBs ohne 'beobachtet'.
    cols = {r[1] for r in con.execute("PRAGMA table_info(tbl_einkauf_kandidat_status)")}
    if "beobachtet" not in cols:
        con.execute("ALTER TABLE tbl_einkauf_kandidat_status ADD COLUMN beobachtet INTEGER DEFAULT 0")


def _ensure_lieferant_hersteller_tabelle(con) -> None:
    """n:m-Verknüpfung Lieferant ↔ Hersteller (welcher Lieferant beschafft wessen
    Produkte). Lieferant und Hersteller bleiben getrennte Stammdaten. Idempotent."""
    con.execute(
        "CREATE TABLE IF NOT EXISTS tbl_einkauf_lieferant_hersteller("
        "lieferant_id INTEGER NOT NULL, hersteller TEXT NOT NULL, "
        "PRIMARY KEY(lieferant_id, hersteller), "
        "FOREIGN KEY(lieferant_id) REFERENCES tbl_einkauf_lieferanten(id) ON DELETE CASCADE)")


def _ensure_seed() -> None:
    """Legt beim ersten Start die Standard-Wechselkurse an (idempotent)."""
    try:
        with _con() as con:
            _ensure_kandidat_status_tabelle(con)
            _ensure_lieferant_hersteller_tabelle(con)
            # Lieferanten: Flag 'bevorzugt' nachrüsten (ALTER-Guard).
            if _table_exists(con, "tbl_einkauf_lieferanten"):
                lcols = {r[1] for r in con.execute("PRAGMA table_info(tbl_einkauf_lieferanten)")}
                if "bevorzugt" not in lcols:
                    con.execute("ALTER TABLE tbl_einkauf_lieferanten ADD COLUMN bevorzugt INTEGER DEFAULT 0")
            if not _table_exists(con, "tbl_einkauf_wechselkurse"):
                con.commit()
                return
            for waehrung, kurs in WECHSELKURS_SEED.items():
                con.execute(
                    "INSERT OR IGNORE INTO tbl_einkauf_wechselkurse(waehrung, kurs_eur, aktualisiert_am) "
                    "VALUES(?,?,?)", (waehrung, kurs, _now()))
            con.commit()
    except sqlite3.Error:
        pass


def kurs_eur(waehrung: str) -> float:
    """EUR je 1 Einheit der Fremdwaehrung. Unbekannt/leer -> 1.0 (EUR)."""
    waehrung = (waehrung or "EUR").strip().upper()
    if waehrung == "EUR":
        return 1.0
    try:
        with _con() as con:
            row = con.execute("SELECT kurs_eur FROM tbl_einkauf_wechselkurse WHERE waehrung=?",
                              (waehrung,)).fetchone()
        if row and row[0]:
            return float(row[0])
    except sqlite3.Error:
        pass
    return WECHSELKURS_SEED.get(waehrung, 1.0)


def waehrungen() -> list[str]:
    try:
        with _con() as con:
            rows = con.execute("SELECT waehrung FROM tbl_einkauf_wechselkurse ORDER BY "
                               "CASE WHEN waehrung='EUR' THEN 0 ELSE 1 END, waehrung").fetchall()
        if rows:
            return [r[0] for r in rows]
    except sqlite3.Error:
        pass
    return list(WECHSELKURS_SEED.keys())


# ── §129-Logik ───────────────────────────────────────────────────────────────
def _p129_staffel() -> tuple:
    """Die fünf §129-Staffelparameter einmalig aus den Einstellungen lesen.
    Für Schleifen (z. B. Importkandidaten) einmal laden und an
    erforderlicher_abstand() durchreichen, statt 5 get_setting() je Artikel."""
    return (_parse_num(get_setting("p129_t1_grenze")), _parse_num(get_setting("p129_t1_proz")),
            _parse_num(get_setting("p129_t2_grenze")), _parse_num(get_setting("p129_t2_eur")),
            _parse_num(get_setting("p129_t3_proz")))


def erforderlicher_abstand(avp: float, staffel: tuple | None = None) -> tuple[float, str]:
    """Liefert (erforderlicher Preisabstand in €, Regel-Text) fuer einen
    Referenz-AVP nach der konfigurierten §129-Staffel. `staffel` optional
    vorgeladen (siehe _p129_staffel), sonst werden die Einstellungen gelesen."""
    t1_grenze, t1_proz, t2_grenze, t2_eur, t3_proz = staffel if staffel else _p129_staffel()
    avp = max(0.0, avp)
    if avp <= t1_grenze:
        return avp * t1_proz / 100.0, f"≥ {t1_proz:.0f} % (AVP bis {t1_grenze:.0f} €)"
    if avp <= t2_grenze:
        return t2_eur, f"≥ {t2_eur:.0f} € (AVP {t1_grenze:.0f}–{t2_grenze:.0f} €)"
    return avp * t3_proz / 100.0, f"≥ {t3_proz:.0f} % (AVP über {t2_grenze:.0f} €)"


def margen_rechnung(referenz_avp: float, import_avp: float,
                    import_ek_eur: float, unser_vk: float) -> dict:
    """Kernrechnung der Importmarge. Liefert ein Dict mit allen Kennzahlen."""
    abstand_soll, regel = erforderlicher_abstand(referenz_avp)
    max_import_avp = max(0.0, referenz_avp - abstand_soll)
    kassen_ersparnis = referenz_avp - import_avp           # Ersparnis ggü. Original
    abstand_ist_proz = (kassen_ersparnis / referenz_avp * 100.0) if referenz_avp else 0.0
    p129_ok = import_avp <= max_import_avp + 1e-6 and import_avp > 0
    marge = unser_vk - import_ek_eur
    marge_proz = (marge / unser_vk * 100.0) if unser_vk else 0.0
    return {
        "abstand_soll": abstand_soll, "regel": regel, "max_import_avp": max_import_avp,
        "kassen_ersparnis": kassen_ersparnis, "abstand_ist_proz": abstand_ist_proz,
        "p129_ok": p129_ok, "marge": marge, "marge_proz": marge_proz,
    }


# ── Daten-Zugriffe ───────────────────────────────────────────────────────────
def artikel_by_pzn(pzn: str) -> dict | None:
    """NMG-Stammdaten zu einer PZN (Name, APU/HAP, Taxe-EK, Taxe-VK=AVP)."""
    pzn = str(pzn or "").strip()
    if not pzn:
        return None
    with _con() as con:
        if not _table_exists(con, "tbl_nmg_stamm"):
            return None
        row = con.execute(
            "SELECT pzn, artikelname, apu, taxe_ek, taxe_vk FROM tbl_nmg_stamm "
            "WHERE pzn=? LIMIT 1", (pzn,)).fetchone()
    if not row:
        return None
    return {"pzn": row[0], "artikelname": row[1], "apu": row[2],
            "taxe_ek": row[3], "taxe_vk": row[4]}


def artikel_suche(text: str, limit: int = 25) -> list[tuple]:
    like = f"%{text}%"
    with _con() as con:
        if not _table_exists(con, "tbl_nmg_stamm"):
            return []
        return con.execute(
            "SELECT pzn, artikelname, taxe_vk FROM tbl_nmg_stamm "
            "WHERE pzn LIKE ? OR artikelname LIKE ? ORDER BY artikelname LIMIT ?",
            (like, like, limit)).fetchall()


def lieferanten_liste(nur_aktiv: bool = False) -> list[sqlite3.Row]:
    with _con() as con:
        con.row_factory = sqlite3.Row
        sql = ("SELECT * FROM tbl_einkauf_lieferanten "
               + ("WHERE aktiv=1 " if nur_aktiv else "")
               + "ORDER BY name")
        return con.execute(sql).fetchall()


def lieferant_namen() -> dict[int, str]:
    with _con() as con:
        return {r[0]: r[1] for r in con.execute(
            "SELECT id, name FROM tbl_einkauf_lieferanten ORDER BY name")}


def bearbeiter() -> str:
    return getpass.getuser()


def absatz_je_pzn() -> dict[str, int]:
    """Summiert die verkaufte Menge je PZN aus den Kasse-Verkäufen
    (tbl_bestellpositionen). Grundlage für Bestellvorschläge."""
    out: dict[str, int] = {}
    try:
        with _con() as con:
            if not _table_exists(con, "tbl_bestellpositionen"):
                return out
            for pzn, menge in con.execute(
                    "SELECT pzn, SUM(COALESCE(menge,0)) FROM tbl_bestellpositionen "
                    "WHERE pzn IS NOT NULL AND pzn!='' GROUP BY pzn"):
                out[str(pzn)] = int(menge or 0)
    except sqlite3.Error:
        pass
    return out


def quelle_kennzahlen(row) -> dict:
    """Reichert eine Quellen-Zeile (sqlite3.Row mit ek_fremd/waehrung/pzn) um
    EK(EUR), AVP, §129-Status und Marge an. Einheitliche Logik für Liste,
    Vorschlag und Export."""
    ek_eur = (row["ek_fremd"] or 0) * kurs_eur(row["waehrung"])
    art = artikel_by_pzn(row["pzn"]) or {}
    avp = art.get("taxe_vk")
    erg = {"ek_eur": ek_eur, "avp": avp, "artikelname": row["artikelname"] or art.get("artikelname") or "",
           "marge": None, "marge_proz": 0.0, "p129_ok": None, "max_import_avp": None}
    if avp:
        import_avp = max(0.0, avp - erforderlicher_abstand(avp)[0])
        m = margen_rechnung(avp, import_avp, ek_eur, art.get("taxe_ek") or avp)
        erg.update({"marge": m["marge"], "marge_proz": m["marge_proz"],
                    "p129_ok": m["p129_ok"], "max_import_avp": m["max_import_avp"]})
    return erg


# ── Importkandidaten (Marktchancen aus Bedarfsanalysen) ──────────────────────
def wichtige_hersteller() -> set[str]:
    """Manuell als 'wichtig im Markt' markierte Hersteller (Setting-Key, JSON-Liste)."""
    raw = get_setting("wichtige_hersteller", "")
    if not raw:
        return set()
    try:
        data = json.loads(raw)
        if isinstance(data, list):
            return {str(x) for x in data}
    except (ValueError, TypeError):
        pass
    return set()


def set_hersteller_wichtig(name: str, wichtig: bool) -> None:
    name = (name or "").strip()
    if not name:
        return
    s = wichtige_hersteller()
    if wichtig:
        s.add(name)
    else:
        s.discard(name)
    set_setting("wichtige_hersteller", json.dumps(sorted(s), ensure_ascii=False))
    _log("Hersteller-Wichtigkeit gesetzt", f"{name}={'wichtig' if wichtig else 'normal'}")


# Bekannte Reimporteure / Parallelimporteure: Lauer-Herstellerkürzel + Namensbestandteile.
# Artikel dieser Anbieter sind bereits Parallelimporte und daher KEINE neuen
# Importkandidaten. In der App erweiterbar (Setting 'reimporteur_hersteller').
REIMPORTEUR_CODES = {
    "KOHL", "EURIM", "EMRA", "ORIF", "CCPHA", "ABACU", "ACAMU", "AXICO", "AXIC",
    "BERAG", "HAEMA", "PARAN", "MILIN", "WESTE", "GERKE", "DOCPH", "FDPHA", "2CARE",
    "EMRAM", "ORIFA",
}
REIMPORTEUR_NAMEN = {
    "KOHLPHARMA", "EURIM", "EMRA-MED", "EMRAMED", "EMRA MED", "ORIFARM", "CC-PHARMA",
    "CC PHARMA", "ABACUS", "ACA MULLER", "ACA MÜLLER", "AXICORP", "BERAGENA", "HAEMATO",
    "PARANOVA", "MILINDA", "WESTEN PHARMA", "GERKE", "DOCPHARM", "FD PHARMA", "2CARE4",
    "MTK PHARMA", "PHARMORE",
}


def benutzer_reimporteure() -> set[str]:
    """Vom Nutzer zusätzlich als Reimporteur markierte Hersteller (Setting, JSON-Liste)."""
    raw = get_setting("reimporteur_hersteller", "")
    if not raw:
        return set()
    try:
        data = json.loads(raw)
        if isinstance(data, list):
            return {str(x) for x in data}
    except (ValueError, TypeError):
        pass
    return set()


def set_hersteller_reimporteur(name: str, reimport: bool) -> None:
    name = (name or "").strip()
    if not name:
        return
    s = benutzer_reimporteure()
    if reimport:
        s.add(name)
    else:
        s.discard(name)
    set_setting("reimporteur_hersteller", json.dumps(sorted(s), ensure_ascii=False))
    _log("Reimporteur-Markierung gesetzt", f"{name}={'reimport' if reimport else 'original'}")


def _ist_reimporteur(hersteller: str, benutzer: set[str]) -> bool:
    """True, wenn der Anbieter ein (Re-)Parallelimporteur ist. `benutzer` = einmal
    vorab geladene Nutzer-Markierungen (kein DB-Zugriff je Zeile)."""
    name = (hersteller or "").strip()
    if not name:
        return False
    if name in benutzer:
        return True
    up = name.upper()
    if up in REIMPORTEUR_CODES:
        return True
    return any(sub in up for sub in REIMPORTEUR_NAMEN)


def hersteller_typ(hersteller: str) -> str:
    """'Reimport' oder 'Original' für einen einzelnen Anbieter (mit DB-Lookup)."""
    return "Reimport" if _ist_reimporteur(hersteller, benutzer_reimporteure()) else "Original"


def kandidat_status_map() -> dict[str, dict]:
    """PZN -> {status, notiz, wiedervorlage} aller gepflegten Pipeline-Einträge."""
    out: dict[str, dict] = {}
    try:
        with _con() as con:
            con.row_factory = sqlite3.Row
            _ensure_kandidat_status_tabelle(con)
            for r in con.execute("SELECT * FROM tbl_einkauf_kandidat_status"):
                out[str(r["pzn"])] = {"status": r["status"] or "Neu",
                                      "notiz": r["notiz"] or "",
                                      "wiedervorlage": r["wiedervorlage"] or "",
                                      "beobachtet": bool(r["beobachtet"])}
    except sqlite3.Error:
        pass
    return out


def lieferant_hersteller(lieferant_id) -> list[str]:
    """Hersteller, die dieser Lieferant beschaffen kann."""
    try:
        with _con() as con:
            _ensure_lieferant_hersteller_tabelle(con)
            return [r[0] for r in con.execute(
                "SELECT hersteller FROM tbl_einkauf_lieferant_hersteller "
                "WHERE lieferant_id=? ORDER BY hersteller", (lieferant_id,))]
    except sqlite3.Error:
        return []


def set_lieferant_hersteller(lieferant_id, herstellers: list[str]) -> None:
    """Setzt die Hersteller-Zuordnung eines Lieferanten (ersetzt die bisherige)."""
    with _con() as con:
        _ensure_lieferant_hersteller_tabelle(con)
        con.execute("DELETE FROM tbl_einkauf_lieferant_hersteller WHERE lieferant_id=?",
                    (lieferant_id,))
        for h in sorted({(x or "").strip() for x in herstellers if (x or "").strip()}):
            con.execute("INSERT OR IGNORE INTO tbl_einkauf_lieferant_hersteller"
                        "(lieferant_id, hersteller) VALUES(?,?)", (lieferant_id, h))
        con.commit()


def lieferanten_fuer_hersteller(hersteller: str) -> list[tuple]:
    """Aktive Lieferanten, die den genannten Hersteller beschaffen können
    (bevorzugte zuerst). Liefert (id, name, bevorzugt)."""
    hersteller = (hersteller or "").strip()
    if not hersteller:
        return []
    try:
        with _con() as con:
            con.row_factory = sqlite3.Row
            _ensure_lieferant_hersteller_tabelle(con)
            rows = con.execute(
                "SELECT l.id, l.name, COALESCE(l.bevorzugt,0) AS bevorzugt "
                "FROM tbl_einkauf_lieferant_hersteller lh "
                "JOIN tbl_einkauf_lieferanten l ON l.id=lh.lieferant_id "
                "WHERE lh.hersteller=? AND l.aktiv=1 "
                "ORDER BY l.bevorzugt DESC, l.name", (hersteller,)).fetchall()
        return [(r["id"], r["name"], bool(r["bevorzugt"])) for r in rows]
    except sqlite3.Error:
        return []


def lieferanten_erfolg() -> list[dict]:
    """Erfolgs-/Aktivitätskennzahlen je Lieferant für die Statistik: Anzahl
    Beschaffungsquellen, Ø-Marge, Aufgaben/Kontakte, Bestellungen."""
    rows: list[dict] = []
    with _con() as con:
        con.row_factory = sqlite3.Row
        if not _table_exists(con, "tbl_einkauf_lieferanten"):
            return rows
        liefs = con.execute("SELECT * FROM tbl_einkauf_lieferanten").fetchall()
        for l in liefs:
            lid = l["id"]
            quellen = con.execute("SELECT * FROM tbl_einkauf_quellen WHERE lieferant_id=? AND aktiv=1",
                                  (lid,)).fetchall()
            margen = [k["marge_proz"] for k in (quelle_kennzahlen(q) for q in quellen)
                      if k["marge"] is not None]
            aufg = con.execute("SELECT COUNT(*) FROM tbl_einkauf_aufgaben WHERE lieferant_id=?",
                               (lid,)).fetchone()[0] if _table_exists(con, "tbl_einkauf_aufgaben") else 0
            erl = con.execute("SELECT COUNT(*) FROM tbl_einkauf_aufgaben WHERE lieferant_id=? "
                              "AND status='erledigt'", (lid,)).fetchone()[0] if _table_exists(con, "tbl_einkauf_aufgaben") else 0
            best = con.execute("SELECT COUNT(*) FROM tbl_einkauf_aufgaben WHERE lieferant_id=? "
                               "AND kategorie='Bestellung'", (lid,)).fetchone()[0] if _table_exists(con, "tbl_einkauf_aufgaben") else 0
            rows.append({
                "id": lid, "name": l["name"], "land": l["land"] or "",
                "bevorzugt": bool(l["bevorzugt"]) if "bevorzugt" in l.keys() else False,
                "aktiv": bool(l["aktiv"]), "quellen": len(quellen),
                "marge_proz": (sum(margen) / len(margen)) if margen else None,
                "aufgaben": aufg, "erledigt": erl, "bestellungen": best,
            })
    rows.sort(key=lambda r: (r["bevorzugt"], r["bestellungen"], r["quellen"]), reverse=True)
    return rows


def hersteller_erfolg() -> list[dict]:
    """Pipeline-Erfolg je Hersteller aus den Kandidaten-Status: wie viele Artikel
    angefragt/eingelistet/verworfen wurden + Erfolgsquote (eingelistet / bearbeitet)."""
    out: dict[str, dict] = {}
    with _con() as con:
        con.row_factory = sqlite3.Row
        if not _table_exists(con, "tbl_einkauf_kandidat_status"):
            return []
        _ensure_kandidat_status_tabelle(con)
        rows = con.execute(
            "SELECT ks.pzn, ks.status, ks.beobachtet, "
            "(SELECT herstellerkuerzel FROM tbl_auswertungspositionen p "
            " WHERE p.pzn=ks.pzn AND COALESCE(herstellerkuerzel,'')<>'' LIMIT 1) AS herst "
            "FROM tbl_einkauf_kandidat_status ks").fetchall()
    for r in rows:
        herst = (r["herst"] or "").strip() or "(unbekannt)"
        d = out.setdefault(herst, {"hersteller": herst, "gesamt": 0, "beobachtet": 0,
                                    "angefragt": 0, "angebot": 0, "eingelistet": 0, "verworfen": 0})
        d["gesamt"] += 1
        if r["beobachtet"]:
            d["beobachtet"] += 1
        s = r["status"] or "Neu"
        if s == "Angefragt":
            d["angefragt"] += 1
        elif s == "Angebot/Muster":
            d["angebot"] += 1
        elif s == "Eingelistet":
            d["eingelistet"] += 1
        elif s == "Verworfen":
            d["verworfen"] += 1
    res = list(out.values())
    for d in res:
        bearbeitet = d["angefragt"] + d["angebot"] + d["eingelistet"] + d["verworfen"]
        d["erfolgsquote"] = (d["eingelistet"] / bearbeitet * 100.0) if bearbeitet else None
    res.sort(key=lambda d: (d["eingelistet"], d["gesamt"]), reverse=True)
    return res


def set_kandidat_beobachtet(pzn: str, beobachtet: bool) -> None:
    """Setzt/entfernt die Merk-Markierung eines Artikels (legt Zeile bei Bedarf an)."""
    pzn = str(pzn or "").strip()
    if not pzn:
        return
    with _con() as con:
        _ensure_kandidat_status_tabelle(con)
        con.execute(
            "INSERT INTO tbl_einkauf_kandidat_status(pzn, beobachtet, geaendert_am, bearbeiter) "
            "VALUES(?,?,?,?) ON CONFLICT(pzn) DO UPDATE SET beobachtet=excluded.beobachtet, "
            "geaendert_am=excluded.geaendert_am, bearbeiter=excluded.bearbeiter",
            (pzn, 1 if beobachtet else 0, _now(), bearbeiter()))
        con.commit()
    _log("Artikel gemerkt" if beobachtet else "Merkung entfernt", f"PZN {pzn}")


def set_kandidat_status(pzn: str, status: str, notiz: str = "", wiedervorlage: str = "") -> None:
    pzn = str(pzn or "").strip()
    if not pzn:
        return
    with _con() as con:
        _ensure_kandidat_status_tabelle(con)
        con.execute(
            "INSERT INTO tbl_einkauf_kandidat_status(pzn, status, notiz, wiedervorlage, geaendert_am, bearbeiter) "
            "VALUES(?,?,?,?,?,?) ON CONFLICT(pzn) DO UPDATE SET status=excluded.status, "
            "notiz=excluded.notiz, wiedervorlage=excluded.wiedervorlage, "
            "geaendert_am=excluded.geaendert_am, bearbeiter=excluded.bearbeiter",
            (pzn, status, notiz, wiedervorlage, _now(), bearbeiter()))
        con.commit()
    _log("Kandidaten-Status gesetzt", f"PZN {pzn} -> {status}")


def _trend_map(con, dq_sql: str, dq_p: list, id_sql: str, id_p: list) -> dict[str, tuple]:
    """PZN -> (symbol, label, prozent) als Nachfrage-Trend über die Bedarfsanalysen.
    Vergleicht den Ø-Absatz der neueren gegen die älteren Auswertungen (nach Datum).
    Weniger als 2 Auswertungen -> ('—','n/a',0)."""
    rows = con.execute(
        f"SELECT p.pzn AS pzn, a.datum AS datum, SUM(COALESCE(p.absatz_6m,0)) AS ab "
        f"FROM tbl_auswertungspositionen p JOIN tbl_auswertungen a ON a.id=p.auswertung_id "
        f"WHERE COALESCE(p.absatz_6m,0)>0 AND p.pzn IS NOT NULL AND p.pzn<>'' "
        f"AND ({dq_sql}) AND ({id_sql}) GROUP BY p.pzn, p.auswertung_id",
        dq_p + id_p).fetchall()
    series: dict[str, list] = {}
    for r in rows:
        series.setdefault(str(r["pzn"]), []).append((r["datum"] or "", float(r["ab"] or 0)))
    out: dict[str, tuple] = {}
    for pzn, vals in series.items():
        if len(vals) < 2:
            out[pzn] = ("—", "n/a", 0.0)
            continue
        vals.sort(key=lambda t: t[0])
        mid = len(vals) // 2
        alt = [v for _, v in vals[:mid]] or [vals[0][1]]
        neu = [v for _, v in vals[mid:]]
        a_avg = sum(alt) / len(alt)
        n_avg = sum(neu) / len(neu)
        if a_avg <= 0:
            out[pzn] = ("—", "n/a", 0.0)
            continue
        ch = (n_avg - a_avg) / a_avg * 100.0
        if ch >= 15:
            out[pzn] = ("↑", "steigend", ch)
        elif ch <= -15:
            out[pzn] = ("↓", "fallend", ch)
        else:
            out[pzn] = ("→", "stabil", ch)
    return out


def _reimporteur_je_wirkstoff(con, wmap: dict[str, str], benutzer: set[str]) -> dict[str, set]:
    """Wirkstoff -> Menge der Reimporteur-Anbieter, die ihn bereits bedienen.
    Basis für die Wettbewerbsdichte (wie viele Parallelimporteure sind schon aktiv)."""
    out: dict[str, set] = {}
    for pzn, herst in con.execute(
            "SELECT DISTINCT pzn, herstellerkuerzel FROM tbl_auswertungspositionen "
            "WHERE COALESCE(absatz_6m,0)>0 AND COALESCE(herstellerkuerzel,'')<>''"):
        if not _ist_reimporteur(herst, benutzer):
            continue
        w = wmap.get(str(pzn))
        if w:
            out.setdefault(w, set()).add(herst)
    return out


def _bedarf_dq_clause(datenquelle: str) -> tuple[str, list]:
    """Datenquelle-Filter auf tbl_auswertungen.datenquelle (analog produktanalyse_neu)."""
    dq = (datenquelle or "ALLE").upper()
    if dq in ("NMG", "PK"):
        return "COALESCE(a.datenquelle,'NMG')='NMG'", []
    if dq in ("ZF", "ZW"):
        return "COALESCE(a.datenquelle,'NMG') IN ('ZW','ZF')", []
    return "1=1", []


def _wirkstoff_map(con) -> dict[str, str]:
    """PZN -> Wirkstoff(e) aus tbl_wirkstoff_staerke (breite Marktquelle, 1:N je PZN).
    Separat geladen, damit der 1:N-JOIN die Absatz-Summen nicht aufbläht."""
    out: dict[str, str] = {}
    if not _table_exists(con, "tbl_wirkstoff_staerke"):
        return out
    # Nur Wirkstoffe zu PZN, die überhaupt in Bedarfsanalysen vorkommen (statt der
    # kompletten ~160k-Zeilen-Tabelle) – deutlich schneller beim Laden.
    sql = ("SELECT w.pzn, GROUP_CONCAT(DISTINCT w.wirkstoff) FROM tbl_wirkstoff_staerke w "
           "WHERE COALESCE(w.wirkstoff,'')<>'' ")
    if _table_exists(con, "tbl_auswertungspositionen"):
        sql += ("AND w.pzn IN (SELECT DISTINCT pzn FROM tbl_auswertungspositionen "
                "WHERE COALESCE(absatz_6m,0)>0 AND pzn IS NOT NULL AND pzn<>'') ")
    sql += "GROUP BY w.pzn"
    for pzn, wirk in con.execute(sql):
        out[str(pzn)] = wirk or ""
    return out


def _apu_hap_map(con) -> dict[str, tuple]:
    """PZN -> (APU, HAP) aus tbl_apu_hap_import (gepflegte Preisreferenz), separat geladen."""
    out: dict[str, tuple] = {}
    if not _table_exists(con, "tbl_apu_hap_import"):
        return out
    for pzn, apu, hap in con.execute(
            "SELECT pzn, MAX(apu), MAX(hap) FROM tbl_apu_hap_import GROUP BY pzn"):
        out[str(pzn)] = (apu, hap)
    return out


def _austausch_pzn_set(con) -> set[str]:
    """PZN mit aktivem Eintrag in der Austauschdatenbank (zusätzliche Ersatz-Quelle)."""
    s: set[str] = set()
    if _table_exists(con, "tbl_austauschdatenbank"):
        for (pzn,) in con.execute(
                "SELECT DISTINCT pzn_alt FROM tbl_austauschdatenbank "
                "WHERE COALESCE(status,'aktiv')='aktiv' AND COALESCE(pzn_alt,'')<>''"):
            s.add(str(pzn))
    return s


def lade_importkandidaten(datenquelle: str = "ALLE", auswertung_id=None) -> list[dict]:
    """Aggregiert alle Bedarfsanalyse-Positionen zu Importkandidaten je PZN.

    Kandidat = Artikel mit Nachfrage (absatz_6m>0), der NICHT im NMG-Stamm liegt und
    für den KEIN NMG-Austausch bekannt ist ('Neue Produktchance'). Liefert je Kandidat
    Nachfrage, Apothekenbreite, abgeleiteten Preis (APU/HAP wenn gepflegt, sonst
    Ø-Stückpreis aus Umsatz/Absatz), Wirkstoff sowie Hersteller-Kennzahlen.
    """
    with _con() as con:
        con.row_factory = sqlite3.Row
        if not (_table_exists(con, "tbl_auswertungspositionen")
                and _table_exists(con, "tbl_auswertungen")):
            return []
        dq_sql, dq_p = _bedarf_dq_clause(datenquelle)
        id_sql, id_p = ("a.id=?", [int(auswertung_id)]) if auswertung_id else ("1=1", [])
        has_aa = _table_exists(con, "tbl_austauschartikel")
        aa_join = "LEFT JOIN tbl_austauschartikel aa ON aa.original_pzn=p.pzn" if has_aa else ""
        aa_case = "WHEN aa.original_pzn IS NOT NULL THEN 1" if has_aa else ""
        sql = f"""
            SELECT p.pzn AS pzn,
                COALESCE(MAX(NULLIF(p.artikelname,'')), MAX(NULLIF(n.artikelname,'')), '') AS artikel,
                COALESCE(MAX(NULLIF(p.df,'')), '') AS df,
                COALESCE(MAX(NULLIF(p.pck,'')), '') AS pck,
                COALESCE(MAX(NULLIF(p.herstellerkuerzel,'')), MAX(NULLIF(n.herstellerkuerzel,'')), '') AS hersteller,
                COUNT(DISTINCT a.apotheke) AS apotheken,
                SUM(COALESCE(p.absatz_6m,0)) AS absatz,
                SUM(COALESCE(p.umsatz,0)) AS umsatz,
                SUM(COALESCE(p.ek,0)) AS ek_summe,
                MAX(n.pzn) AS nmg_pzn,
                MAX(n.apu) AS nmg_apu,
                MAX(n.taxe_ek) AS nmg_taxe_ek,
                MAX(NULLIF(n.wirkstoffe,'')) AS nmg_wirkstoff,
                MAX(CASE {aa_case}
                         WHEN COALESCE(p.pzn_nmg,'')<>'' THEN 1
                         WHEN COALESCE(p.austauschbar_gegen,'')<>'' THEN 1
                         ELSE 0 END) AS hat_austausch
            FROM tbl_auswertungspositionen p
            JOIN tbl_auswertungen a ON a.id=p.auswertung_id
            LEFT JOIN tbl_nmg_stamm n ON n.pzn=p.pzn
            {aa_join}
            WHERE COALESCE(p.absatz_6m,0)>0 AND p.pzn IS NOT NULL AND p.pzn<>''
              AND ({dq_sql}) AND ({id_sql})
            GROUP BY p.pzn
            HAVING nmg_pzn IS NULL AND hat_austausch=0
        """
        rows = con.execute(sql, dq_p + id_p).fetchall()
        wmap = _wirkstoff_map(con)
        ahmap = _apu_hap_map(con)
        adb = _austausch_pzn_set(con)
        benutzer_reimp = benutzer_reimporteure()
        reimp_je_wirk = _reimporteur_je_wirkstoff(con, wmap, benutzer_reimp)
        trend = _trend_map(con, dq_sql, dq_p, id_sql, id_p)

    wichtig = wichtige_hersteller()
    status_map = kandidat_status_map()
    eu_ek_rate = max(0.0, _parse_num(get_setting("kandidat_eu_ek_proz")) / 100.0)
    min_marge = _parse_num(get_setting("kandidat_min_marge_proz"))
    staffel = _p129_staffel()  # einmal laden statt je Kandidat
    cand: list[dict] = []
    herst_absatz: dict[str, float] = {}
    herst_count: dict[str, int] = {}
    for r in rows:
        pzn = str(r["pzn"])
        if pzn in adb:  # zusätzliche Ersatz-Quelle (Austauschdatenbank) -> kein Kandidat
            continue
        absatz = float(r["absatz"] or 0)
        umsatz = float(r["umsatz"] or 0)
        ek_summe = float(r["ek_summe"] or 0)
        ah = ahmap.get(pzn)
        apu = float(ah[0]) if ah and ah[0] else (float(r["nmg_apu"]) if r["nmg_apu"] else None)
        markt_preis = (umsatz / absatz) if absatz > 0 and umsatz > 0 else None
        if apu is not None:
            preis, quelle = apu, "APU"
        elif markt_preis is not None:
            preis, quelle = markt_preis, "Ø Markt"
        else:
            preis, quelle = None, "—"
        ek_stk = (ek_summe / absatz) if absatz > 0 and ek_summe > 0 else None
        taxe_ek = (float(r["nmg_taxe_ek"]) if r["nmg_taxe_ek"]
                   else (float(ah[1]) if ah and ah[1] else ek_stk))
        herst = (r["hersteller"] or "").strip() or "(unbekannt)"
        wirk = wmap.get(pzn) or (r["nmg_wirkstoff"] or "")

        # §129-Machbarkeit & Potenzial (Schätzung): angenommener EU-EK = eu_ek_rate
        # des deutschen Preisniveaus; max. zulässiger Import-AVP nach §129-Staffel.
        marge_stueck = marge_proz = max_import_avp = potenzial_jahr = umsatz_jahr = None
        ampel = "—"
        if preis and preis > 0:
            umsatz_jahr = preis * absatz * 2.0  # 6 Monate -> Jahr
            max_import_avp = max(0.0, preis - erforderlicher_abstand(preis, staffel)[0])
            import_ek = preis * eu_ek_rate
            marge_stueck = max_import_avp - import_ek
            marge_proz = (marge_stueck / max_import_avp * 100.0) if max_import_avp > 0 else 0.0
            potenzial_jahr = marge_stueck * absatz * 2.0
            ampel = ("gruen" if marge_proz >= min_marge else
                     ("gelb" if marge_stueck > 0 else "rot"))

        konkurrenz = len(reimp_je_wirk.get(wirk, ())) if wirk else 0
        tr_sym, tr_lbl, tr_proz = trend.get(pzn, ("—", "n/a", 0.0))
        cand.append({
            "pzn": pzn, "artikel": r["artikel"] or "", "hersteller": herst,
            "df": r["df"] or "", "pck": r["pck"] or "", "wirkstoff": wirk,
            "absatz": absatz, "apotheken": int(r["apotheken"] or 0), "umsatz": umsatz,
            "preis": preis, "preis_quelle": quelle, "apu": apu, "taxe_ek": taxe_ek,
            "markt_preis": markt_preis, "wichtig": herst in wichtig,
            "typ": "Reimport" if _ist_reimporteur(herst, benutzer_reimp) else "Original",
            "marge_stueck": marge_stueck, "marge_proz": marge_proz,
            "max_import_avp": max_import_avp, "ampel": ampel,
            "potenzial_jahr": potenzial_jahr, "umsatz_jahr": umsatz_jahr,
            "konkurrenz": konkurrenz,
            "konkurrenz_namen": sorted(reimp_je_wirk.get(wirk, ())) if wirk else [],
            "trend_sym": tr_sym, "trend_label": tr_lbl, "trend_proz": tr_proz,
            "status": status_map.get(pzn, {}).get("status", "Neu"),
            "beobachtet": status_map.get(pzn, {}).get("beobachtet", False),
            "score": 100.0 + absatz + int(r["apotheken"] or 0) * 10,
        })
        herst_absatz[herst] = herst_absatz.get(herst, 0.0) + absatz
        herst_count[herst] = herst_count.get(herst, 0) + 1
    for d in cand:
        d["herst_absatz"] = herst_absatz.get(d["hersteller"], 0.0)
        d["herst_artikel"] = herst_count.get(d["hersteller"], 0)
    # Standard: gemerkte Artikel ganz oben, dann höchstes geschätztes Jahres-Potenzial
    # (Artikel ohne Preis ans Ende).
    cand.sort(key=lambda d: (d["beobachtet"], d["potenzial_jahr"] is not None,
                             d["potenzial_jahr"] or 0.0, d["absatz"]), reverse=True)
    return cand


def _filter_kandidaten(cands: list[dict], filters: dict) -> list[dict]:
    """Filtert eine geladene Kandidatenliste nach Preis, Hersteller, Wirkstoff,
    Apothekenbreite und 'nur wichtige Hersteller'. Reine In-Memory-Operation."""
    pmin = filters.get("preis_min")
    pmax = filters.get("preis_max")
    hf = (filters.get("hersteller") or "").strip()
    wf = (filters.get("wirkstoff") or "").strip().upper()
    min_apo = int(filters.get("min_apotheken") or 0)
    nur_w = bool(filters.get("nur_wichtige_hersteller"))
    typ = (filters.get("typ") or "Alle")
    nur_machbar = bool(filters.get("nur_machbar"))
    nur_ohne_konk = bool(filters.get("nur_ohne_konkurrenz"))
    nur_beob = bool(filters.get("nur_beobachtet"))
    status_f = (filters.get("status") or "Alle")
    out = []
    for d in cands:
        if nur_beob and not d["beobachtet"]:
            continue
        if typ == "Original" and d["typ"] != "Original":
            continue
        if typ == "Reimport" and d["typ"] != "Reimport":
            continue
        if nur_machbar and d["ampel"] not in ("gruen", "gelb"):
            continue
        if nur_ohne_konk and d["konkurrenz"] > 0:
            continue
        if status_f and status_f != "Alle" and d["status"] != status_f:
            continue
        if pmin is not None and (d["preis"] is None or d["preis"] < pmin):
            continue
        if pmax is not None and (d["preis"] is None or d["preis"] > pmax):
            continue
        if hf and hf not in ("", "Alle") and d["hersteller"] != hf:
            continue
        if wf and wf not in (d["wirkstoff"] or "").upper():
            continue
        if min_apo and d["apotheken"] < min_apo:
            continue
        if nur_w and not d["wichtig"]:
            continue
        out.append(d)
    return out


def kandidat_hersteller_namen() -> list[str]:
    """Leichte Hersteller-Liste für das Filter-Dropdown, ohne die teure
    Vollberechnung der Kandidaten (nur DISTINCT über die Positionen)."""
    try:
        with _con() as con:
            if not _table_exists(con, "tbl_auswertungspositionen"):
                return []
            rows = con.execute(
                "SELECT DISTINCT COALESCE(NULLIF(herstellerkuerzel,''),'(unbekannt)') AS h "
                "FROM tbl_auswertungspositionen WHERE COALESCE(absatz_6m,0)>0 ORDER BY h").fetchall()
        return [r[0] for r in rows]
    except sqlite3.Error:
        return []


def importkandidaten(filters: dict | None = None) -> list[dict]:
    """Geladene + gefilterte Importkandidaten (Komfort-Wrapper für Export/Tests)."""
    filters = filters or {}
    cands = lade_importkandidaten(filters.get("datenquelle", "ALLE"),
                                  filters.get("auswertung_id"))
    return _filter_kandidaten(cands, filters)


# ── Export & Tabellen-Komfort ────────────────────────────────────────────────
def _export_xlsx(headers, rows, blatt_titel, dateibasis):
    """Schreibt eine einfache, sauber formatierte Excel-Datei in OUTPUT_DIR und
    gibt den Pfad zurück. headers=Liste, rows=Liste von Listen."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = blatt_titel[:31]
    ws.append(list(headers))
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0B4A86")
        c.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append(list(r))
    for i, h in enumerate(headers, 1):
        breite = max(len(str(h)) + 2, *(len(str(r[i - 1])) + 2 for r in rows)) if rows else len(str(h)) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(48, max(12, breite))
    ws.freeze_panes = "A2"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ziel = OUTPUT_DIR / f"{dateibasis}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(ziel)
    return ziel


def _datei_oeffnen(pfad: Path) -> None:
    try:
        if os.name == "nt":
            os.startfile(str(pfad))  # type: ignore[attr-defined]
        else:
            webbrowser.open(pfad.as_uri())
    except Exception:
        pass


_NUM_RE = re.compile(r"-?\d[\d.\s]*,?\d*")


def _sortable(tree: ttk.Treeview, numerisch=()):
    """Macht die Spalten einer Treeview durch Klick auf die Überschrift sortierbar.
    numerisch = Spaltennamen, die als Zahl interpretiert werden (€, %, Mengen)."""
    def key(val, col):
        if col in numerisch:
            m = _NUM_RE.search(str(val))
            return _parse_num(m.group(0)) if m else -1e18
        return str(val).lower()

    def sortiere(col, ab):
        daten = [(tree.set(k, col), k) for k in tree.get_children("")]
        daten.sort(key=lambda t: key(t[0], col), reverse=ab)
        for idx, (_, k) in enumerate(daten):
            tree.move(k, "", idx)
        tree.heading(col, command=lambda: sortiere(col, not ab))

    for col in tree["columns"]:
        tree.heading(col, command=lambda c=col: sortiere(c, False))


# ── Beispiel-/Demodaten ──────────────────────────────────────────────────────
def demo_daten_anlegen() -> dict:
    """Legt realistische Beispieldaten an (Lieferanten, Quellen zu echten PZN,
    Aufgaben), damit die App sofort testbar ist. Nicht-destruktiv: vorhandene
    Daten bleiben erhalten, es wird nur ergänzt."""
    lieferanten = [
        ("EuroPharma UK Ltd", "Großbritannien", "GBP", "James Carter", "sales@europharma.co.uk",
         "+44 20 1234 5678", 5, 1000.0, 30, 1),
        ("Polska Hurt Sp. z o.o.", "Polen", "PLN", "Anna Kowalska", "zakup@polskahurt.pl",
         "+48 22 123 4567", 7, 500.0, 21, 1),
        ("Hellas Pharma A.E.", "Griechenland", "EUR", "Nikos Papadopoulos", "info@hellaspharma.gr",
         "+30 21 0123 456", 10, 750.0, 30, 1),
        ("Iberia Distribución S.L.", "Spanien", "EUR", "María García", "compras@iberiadist.es",
         "+34 91 123 45 67", 8, 600.0, 45, 0),
        ("Bohemia Pharma s.r.o.", "Tschechien", "CZK", "Petr Novák", "nakup@bohemiapharma.cz",
         "+420 2 1234 5678", 6, 400.0, 30, 1),
    ]
    with _con() as con:
        con.row_factory = sqlite3.Row
        pzns = [r["pzn"] for r in con.execute(
            "SELECT pzn, apu FROM tbl_nmg_stamm WHERE apu IS NOT NULL AND apu>0 "
            "ORDER BY apu DESC LIMIT 8")]
        apu_map = {r["pzn"]: r["apu"] for r in con.execute(
            "SELECT pzn, apu FROM tbl_nmg_stamm WHERE apu IS NOT NULL")}

        lief_ids = []
        for L in lieferanten:
            cur = con.execute(
                "INSERT INTO tbl_einkauf_lieferanten(name, land, waehrung, ansprechpartner, email, "
                "telefon, lieferzeit_tage, mindestbestellwert, zahlungsziel_tage, gdp_zertifiziert, aktiv) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,1)", L)
            lief_ids.append(cur.lastrowid)

        # Quellen: jede PZN bei 2 Lieferanten mit leicht unterschiedlichem Preis,
        # damit der Beschaffungsvorschlag echte Alternativen vergleichen kann.
        n_quellen = 0
        for i, pzn in enumerate(pzns):
            apu = apu_map.get(pzn, 100.0)
            for j, faktor in enumerate((0.80, 0.74)):  # zwei Angebote: EK ~ unter Herstellerabgabepreis
                lid = lief_ids[(i + j) % len(lief_ids)]
                waehrung = next(L[2] for k, L in enumerate(lieferanten) if lief_ids[k] == lid)
                ek_eur = apu * faktor
                ek_fremd = round(ek_eur / kurs_eur(waehrung), 2)
                con.execute(
                    "INSERT INTO tbl_einkauf_quellen(lieferant_id, pzn, artikelname, ek_fremd, waehrung, "
                    "mindestabnahme, lieferzeit_tage, aktiv, gueltig_ab, notiz) "
                    "VALUES(?,?,?,?,?,?,?,1,?,?)",
                    (lid, pzn, "", ek_fremd, waehrung, 10 * (j + 1), 5 + j,
                     _heute(), "Beispieldaten"))
                n_quellen += 1

        aufgaben = [
            ("Preisliste Q3 anfordern", "Aktuelle Preisliste mit Mengenstaffel erbitten.",
             "Anfrage", 2, (date.today() + timedelta(days=5)).isoformat(), lief_ids[0]),
            ("Rückruf wegen Liefertermin", "Verzögerung klären – Kunde wartet.",
             "Rueckruf", 1, (date.today() - timedelta(days=1)).isoformat(), lief_ids[1]),
            ("Reklamation: Charge prüfen", "Beanstandete Charge mit Lieferant abstimmen.",
             "Reklamation", 1, date.today().isoformat(), lief_ids[2]),
            ("Zahlungsziel verhandeln", "Verlängerung auf 60 Tage anstreben.",
             "Wiedervorlage", 3, (date.today() + timedelta(days=14)).isoformat(), lief_ids[3]),
            ("Bestellung bestätigen lassen", "Auftragsbestätigung nachfassen.",
             "Bestellung", 2, date.today().isoformat(), lief_ids[4]),
        ]
        for titel, beschr, kat, prio, faellig, lid in aufgaben:
            con.execute(
                "INSERT INTO tbl_einkauf_aufgaben(titel, beschreibung, kategorie, prioritaet, "
                "faellig_am, lieferant_id, status, bearbeiter) VALUES(?,?,?,?,?,?,'offen',?)",
                (titel, beschr, kat, prio, faellig, lid, bearbeiter()))
        con.commit()
    _log("Beispieldaten angelegt", f"{len(lieferanten)} Lieferanten, {n_quellen} Quellen")
    return {"lieferanten": len(lieferanten), "quellen": n_quellen, "aufgaben": len(aufgaben)}


# ── Panel ────────────────────────────────────────────────────────────────────
class EinkaufPanel(tk.Frame):
    """Hauptpanel der Einkauf-App: dunkle Sidebar + wechselnde Seiten."""

    def __init__(self, master, on_close=None):
        super().__init__(master, bg=SHELL_BG)
        self.on_close = on_close
        _ensure_seed()

        self.sidebar = theme.Sidebar(self, title="Einkauf", subtitle="Beschaffung & Margen")
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.add_section("Übersicht")
        self.sidebar.add_item("start", "🏠", "Dashboard", lambda: self.show("start"), active=True)
        self.sidebar.add_item("aufgaben", "✅", "Aufgaben & Meldungen", lambda: self.show("aufgaben"))
        self.sidebar.add_section("Beschaffung")
        self.sidebar.add_item("lieferanten", "🏭", "Lieferanten", lambda: self.show("lieferanten"))
        self.sidebar.add_item("quellen", "📦", "Beschaffungsquellen", lambda: self.show("quellen"))
        self.sidebar.add_item("vorschlag", "🧭", "Beschaffungsvorschlag", lambda: self.show("vorschlag"))
        self.sidebar.add_section("Marktchancen")
        self.sidebar.add_item("importkandidaten", "🧪", "Importkandidaten", lambda: self.show("importkandidaten"))
        self.sidebar.add_item("statistik", "📊", "Statistik & Erfolg", lambda: self.show("statistik"))
        self.sidebar.add_section("Kalkulation")
        self.sidebar.add_item("marge", "📐", "Margenrechner §129", lambda: self.show("marge"))
        self.sidebar.add_item("kurse", "💱", "Wechselkurse", lambda: self.show("kurse"))
        self.sidebar.add_section("Konfiguration")
        self.sidebar.add_item("einstellungen", "⚙", "Einstellungen", lambda: self.show("einstellungen"))
        self.sidebar.add_footer_note("NMG Einkauf · Beschaffung EU-Ausland")

        self.content = tk.Frame(self, bg=SHELL_BG)
        self.content.pack(side="left", fill="both", expand=True)
        self._marge_prefill: dict | None = None
        # Importkandidaten: Filter über Seitenwechsel hinweg merken (bis zur nächsten Änderung).
        self._ik_state: dict = {"herst": "Alle", "wirk": "", "pmin": "", "pmax": "",
                                "minapo": "", "nurwichtig": 0, "typ": "Nur Original",
                                "machbar": 0, "ohnekonk": 0, "status": "Alle", "nurbeob": 0}
        # Cache der berechneten Kandidatenliste (teuer) – über Seitenwechsel hinweg,
        # damit Wiederöffnen sofort ist. Erst-Berechnung erst bei „Anzeigen".
        self._ik_cache: list | None = None
        # Gemerkte Filter der übrigen Seiten (bleiben bis zur nächsten Änderung erhalten).
        self._filter_state: dict = {"aufg": "offen", "quellen_lief": "Alle",
                                    "vor_nur129": 0, "vor_nur_absatz": 0}
        self.show("start")

    # ── Infrastruktur ────────────────────────────────────────────────────────
    def show(self, key, **kwargs):
        for w in self.content.winfo_children():
            w.destroy()
        self.sidebar.set_active(key)
        if key == "marge" and "prefill" in kwargs:
            self._marge_prefill = kwargs["prefill"]
        {"start": self._page_start, "aufgaben": self._page_aufgaben,
         "lieferanten": self._page_lieferanten, "quellen": self._page_quellen,
         "vorschlag": self._page_vorschlag, "importkandidaten": self._page_importkandidaten,
         "statistik": self._page_statistik,
         "marge": self._page_marge, "kurse": self._page_kurse,
         "einstellungen": self._page_einstellungen}[key]()

    def _scrollseite(self, breite=940):
        canvas = tk.Canvas(self.content, bg=SHELL_BG, highlightthickness=0)
        scroll = ttk.Scrollbar(self.content, orient="vertical", command=canvas.yview)
        wrap = tk.Frame(canvas, bg=SHELL_BG)
        wrap.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=wrap, anchor="nw", width=breite)
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side="left", fill="both", expand=True, padx=(24, 0))
        scroll.pack(side="right", fill="y")
        def _wheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _wheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))
        return wrap

    def _kpi_cards(self, parent, items):
        """items = Liste aus (label, wert, farbe)."""
        grid = tk.Frame(parent, bg=SHELL_BG)
        grid.pack(fill="x", padx=24)
        for i, (label, wert, farbe) in enumerate(items):
            grid.columnconfigure(i, weight=1)
            card = tk.Frame(grid, bg=BG, highlightbackground=BORDER, highlightthickness=1)
            card.grid(row=0, column=i, sticky="ew", padx=(0 if i == 0 else 8, 0))
            tk.Label(card, text=label, bg=BG, fg=MUTED, font=theme.SMALL).pack(anchor="w", padx=14, pady=(12, 0))
            tk.Label(card, text=wert, bg=BG, fg=farbe, font=(theme.FONT, 22, "bold")).pack(anchor="w", padx=14, pady=(0, 12))

    def _make_tree(self, parent, columns, widths, height=12, numerisch=()):
        wrap = tk.Frame(parent, bg=SHELL_BG)
        tree = ttk.Treeview(wrap, columns=columns, show="headings", height=height, style="NMG.Treeview")
        for col, w in zip(columns, widths):
            tree.heading(col, text=col)
            tree.column(col, width=w, anchor="w")
        vs = ttk.Scrollbar(wrap, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vs.set)
        tree.pack(side="left", fill="both", expand=True)
        vs.pack(side="right", fill="y")
        _sortable(tree, numerisch=numerisch)
        return wrap, tree

    # ── Seite: Dashboard ─────────────────────────────────────────────────────
    def _page_start(self):
        theme.page_header(self.content, "Einkauf",
                          _T('Angemeldet: {p0} · {p1}', p0=bearbeiter(), p1=_datum_de(_heute())),
                          bg=SHELL_BG).pack(fill="x", padx=24, pady=(20, 12))

        with _con() as con:
            offen = con.execute("SELECT COUNT(*) FROM tbl_einkauf_aufgaben WHERE status='offen'").fetchone()[0]
            ueberfaellig = con.execute(
                "SELECT COUNT(*) FROM tbl_einkauf_aufgaben WHERE status='offen' "
                "AND faellig_am!='' AND faellig_am < ?", (_heute(),)).fetchone()[0]
            lief = con.execute("SELECT COUNT(*) FROM tbl_einkauf_lieferanten WHERE aktiv=1").fetchone()[0]
            quellen = con.execute("SELECT COUNT(*) FROM tbl_einkauf_quellen WHERE aktiv=1").fetchone()[0]
        self._kpi_cards(self.content, [
            ("Offene Aufgaben", str(offen), ACCENT),
            ("Überfällig", str(ueberfaellig), theme.DANGER if ueberfaellig else OK_GREEN),
            ("Lieferanten aktiv", str(lief), theme.PURPLE),
            ("Beschaffungsquellen", str(quellen), OK_GREEN),
        ])

        # Leerer Start: freundlicher Einstieg mit Beispieldaten
        if lief == 0 and quellen == 0 and offen == 0:
            hint = theme.Card(self.content)
            hint.pack(fill="x", padx=24, pady=(16, 0))
            tk.Label(hint.inner, text="👋  Willkommen im Einkauf", bg=BG, fg=TEXT,
                     font=theme.SECTION).pack(anchor="w")
            tk.Label(hint.inner,
                     text=("Noch keine Daten erfasst. Lege Beispieldaten an, um Lieferanten, "
                           "Beschaffungsquellen, den §129-Margenrechner und den "
                           "Beschaffungsvorschlag sofort auszuprobieren."),
                     bg=BG, fg=MUTED, font=theme.BODY, wraplength=820, justify="left").pack(anchor="w", pady=(6, 10))
            theme.PillButton(hint.inner, "🎁  Beispieldaten anlegen", self._demo_laden,
                             kind="success", padx=16, pady=9).pack(anchor="w")

        # Meldungen: faellige & ueberfaellige Aufgaben ("melde dich bei ...")
        grenze = (date.today() + timedelta(days=2)).isoformat()
        with _con() as con:
            con.row_factory = sqlite3.Row
            meldungen = con.execute(
                "SELECT a.*, l.name AS lieferant FROM tbl_einkauf_aufgaben a "
                "LEFT JOIN tbl_einkauf_lieferanten l ON l.id=a.lieferant_id "
                "WHERE a.status='offen' AND a.faellig_am!='' AND a.faellig_am <= ? "
                "ORDER BY a.faellig_am, a.prioritaet", (grenze,)).fetchall()

        box = theme.Card(self.content)
        box.pack(fill="x", padx=24, pady=(16, 8))
        kopf = tk.Frame(box.inner, bg=BG)
        kopf.pack(fill="x")
        tk.Label(kopf, text="🔔  Meldungen – jetzt erledigen", bg=BG, fg=TEXT,
                 font=theme.SECTION).pack(side="left")
        theme.PillButton(kopf, "Neue Aufgabe", lambda: self._aufgabe_dialog(),
                         kind="primary", font_size=10, padx=12, pady=6).pack(side="right")

        if not meldungen:
            tk.Label(box.inner, text="Keine fälligen Meldungen. Alles im grünen Bereich. ✓",
                     bg=BG, fg=MUTED, font=theme.BODY).pack(anchor="w", pady=(10, 2))
        else:
            for m in meldungen:
                self._meldung_zeile(box.inner, m)

        # Schnellzugriff
        bar = tk.Frame(self.content, bg=SHELL_BG)
        bar.pack(fill="x", padx=24, pady=(8, 18))
        theme.PillButton(bar, "🧪  Importkandidaten", lambda: self.show("importkandidaten"),
                         kind="accent", padx=16, pady=10).pack(side="left", padx=(0, 8))
        theme.PillButton(bar, "📐  Margenrechner §129", lambda: self.show("marge"),
                         kind="neutral", padx=16, pady=10).pack(side="left", padx=8)
        theme.PillButton(bar, "📦  Beschaffungsquellen", lambda: self.show("quellen"),
                         kind="neutral", padx=16, pady=10).pack(side="left", padx=8)
        theme.PillButton(bar, "🏭  Lieferanten", lambda: self.show("lieferanten"),
                         kind="neutral", padx=16, pady=10).pack(side="left", padx=8)

    def _meldung_zeile(self, parent, m):
        ueberfaellig = m["faellig_am"] and m["faellig_am"] < _heute()
        row = tk.Frame(parent, bg=BG)
        row.pack(fill="x", pady=4)
        dot = "🔴" if ueberfaellig else "🟡"
        prio = PRIO_LABEL.get(m["prioritaet"], "Mittel")
        wer = f"  →  {m['lieferant']}" if m["lieferant"] else ""
        txt = f"{dot}  {m['titel']}{wer}"
        sub = (f"{'Überfällig seit' if ueberfaellig else 'Fällig'} {_datum_de(m['faellig_am'])}"
               f"  ·  {m['kategorie']}  ·  Prio {prio}")
        left = tk.Frame(row, bg=BG)
        left.pack(side="left", fill="x", expand=True)
        tk.Label(left, text=txt, bg=BG, fg=TEXT, font=theme.BODY_BOLD, anchor="w").pack(anchor="w")
        tk.Label(left, text=sub, bg=BG, fg=(theme.DANGER if ueberfaellig else MUTED),
                 font=theme.SMALL, anchor="w").pack(anchor="w")
        theme.PillButton(row, "Erledigt ✓", lambda i=m["id"]: self._aufgabe_erledigen(i),
                         kind="success", font_size=10, padx=10, pady=6).pack(side="right")
        theme.PillButton(row, "Öffnen", lambda i=m["id"]: self._aufgabe_dialog(i),
                         kind="ghost", font_size=10, padx=10, pady=6).pack(side="right", padx=6)

    # ── Seite: Aufgaben ──────────────────────────────────────────────────────
    def _page_aufgaben(self):
        theme.page_header(self.content, "Aufgaben & Meldungen",
                          "Wiedervorlagen, Anfragen, Rückrufe – nichts geht verloren.",
                          bg=SHELL_BG).pack(fill="x", padx=24, pady=(20, 8))
        bar = tk.Frame(self.content, bg=SHELL_BG)
        bar.pack(fill="x", padx=24, pady=(0, 8))
        theme.PillButton(bar, "➕  Neue Aufgabe", lambda: self._aufgabe_dialog(),
                         kind="primary", padx=14, pady=8).pack(side="left")
        self._aufg_filter = tk.StringVar(value=self._filter_state["aufg"])
        for label, val in (("Offen", "offen"), ("Erledigt", "erledigt"), ("Alle", "alle")):
            ttk.Radiobutton(bar, text=label, value=val, variable=self._aufg_filter,
                            command=self._aufg_reload).pack(side="left", padx=(12, 0))

        wrap, self._aufg_tree = self._make_tree(
            self.content,
            ("Fällig", "Titel", "Kategorie", "Lieferant", "Prio", "Status"),
            (90, 320, 120, 180, 70, 90), height=16)
        wrap.pack(fill="both", expand=True, padx=24, pady=(4, 4))
        self._aufg_tree.bind("<Double-1>", lambda e: self._aufg_edit_selected())

        akt = tk.Frame(self.content, bg=SHELL_BG)
        akt.pack(fill="x", padx=24, pady=(0, 16))
        theme.PillButton(akt, "Bearbeiten", self._aufg_edit_selected, kind="neutral",
                         font_size=10, padx=12, pady=6).pack(side="left")
        theme.PillButton(akt, "Erledigt ✓", self._aufg_erledigen_selected, kind="success",
                         font_size=10, padx=12, pady=6).pack(side="left", padx=6)
        theme.PillButton(akt, "Löschen", self._aufg_loeschen_selected, kind="danger",
                         font_size=10, padx=12, pady=6).pack(side="left")
        self._aufg_reload()

    def _aufg_reload(self):
        for i in self._aufg_tree.get_children():
            self._aufg_tree.delete(i)
        f = self._aufg_filter.get()
        self._filter_state["aufg"] = f
        where = ""
        if f == "offen":
            where = "WHERE a.status='offen'"
        elif f == "erledigt":
            where = "WHERE a.status='erledigt'"
        with _con() as con:
            con.row_factory = sqlite3.Row
            rows = con.execute(
                f"SELECT a.*, l.name AS lieferant FROM tbl_einkauf_aufgaben a "
                f"LEFT JOIN tbl_einkauf_lieferanten l ON l.id=a.lieferant_id {where} "
                f"ORDER BY CASE WHEN a.status='offen' THEN 0 ELSE 1 END, "
                f"a.faellig_am='' , a.faellig_am, a.prioritaet").fetchall()
        for r in rows:
            ueberf = r["status"] == "offen" and r["faellig_am"] and r["faellig_am"] < _heute()
            tag = "ueberf" if ueberf else ("erl" if r["status"] == "erledigt" else "")
            self._aufg_tree.insert(
                "", "end", iid=str(r["id"]),
                values=(_datum_de(r["faellig_am"]), r["titel"], r["kategorie"],
                        r["lieferant"] or "—", PRIO_LABEL.get(r["prioritaet"], "Mittel"),
                        "✓ erledigt" if r["status"] == "erledigt" else "offen"),
                tags=(tag,))
        self._aufg_tree.tag_configure("ueberf", foreground=theme.DANGER)
        self._aufg_tree.tag_configure("erl", foreground=MUTED)

    def _aufg_selected_id(self):
        sel = self._aufg_tree.selection()
        return int(sel[0]) if sel else None

    def _aufg_edit_selected(self):
        i = self._aufg_selected_id()
        if i:
            self._aufgabe_dialog(i)
        else:
            messagebox.showinfo("Aufgaben", "Bitte zuerst eine Aufgabe auswählen.", parent=self)

    def _aufg_erledigen_selected(self):
        i = self._aufg_selected_id()
        if i:
            self._aufgabe_erledigen(i)

    def _aufg_loeschen_selected(self):
        i = self._aufg_selected_id()
        if not i:
            return
        if messagebox.askyesno("Löschen", "Diese Aufgabe wirklich löschen?", parent=self):
            with _con() as con:
                con.execute("DELETE FROM tbl_einkauf_aufgaben WHERE id=?", (i,))
                con.commit()
            _log("Aufgabe gelöscht", f"id={i}")
            self._aufg_reload()

    def _aufgabe_erledigen(self, aufgabe_id):
        with _con() as con:
            con.execute("UPDATE tbl_einkauf_aufgaben SET status='erledigt', erledigt_am=? WHERE id=?",
                        (_now(), aufgabe_id))
            con.commit()
        _log("Aufgabe erledigt", f"id={aufgabe_id}")
        # Aktive Seite passend aktualisieren
        if hasattr(self, "_aufg_tree") and self._aufg_tree.winfo_exists():
            self._aufg_reload()
        else:
            self.show("start")

    def _aufgabe_dialog(self, aufgabe_id=None):
        AufgabeDialog(self, aufgabe_id, on_save=lambda: self._after_aufgabe_save())

    def _after_aufgabe_save(self):
        if hasattr(self, "_aufg_tree") and self._aufg_tree.winfo_exists():
            self._aufg_reload()
        else:
            self.show("start")

    # ── Seite: Lieferanten ───────────────────────────────────────────────────
    def _page_lieferanten(self):
        theme.page_header(self.content, "Lieferanten",
                          "Beschaffungsquellen je Land – Währung, Lieferzeit, Mindestbestellwert.",
                          bg=SHELL_BG).pack(fill="x", padx=24, pady=(20, 8))
        bar = tk.Frame(self.content, bg=SHELL_BG)
        bar.pack(fill="x", padx=24, pady=(0, 8))
        theme.PillButton(bar, "➕  Neuer Lieferant", lambda: self._lieferant_dialog(),
                         kind="primary", padx=14, pady=8).pack(side="left")

        wrap, self._lief_tree = self._make_tree(
            self.content,
            ("★", "Name", "Land", "Währung", "Lieferzeit", "Mind.-Wert", "GDP", "Status"),
            (28, 230, 100, 78, 88, 108, 56, 78), height=15,
            numerisch=("Lieferzeit", "Mind.-Wert"))
        wrap.pack(fill="both", expand=True, padx=24, pady=(4, 4))
        self._lief_tree.bind("<Double-1>", lambda e: self._lief_edit_selected())

        akt = tk.Frame(self.content, bg=SHELL_BG)
        akt.pack(fill="x", padx=24, pady=(0, 16))
        theme.PillButton(akt, "★ Bevorzugt", self._lief_toggle_bevorzugt, kind="accent",
                         font_size=10, padx=12, pady=6).pack(side="left", padx=(0, 6))
        theme.PillButton(akt, "Bearbeiten", self._lief_edit_selected, kind="neutral",
                         font_size=10, padx=12, pady=6).pack(side="left")
        theme.PillButton(akt, "Quellen anzeigen", lambda: self._lief_zeige_quellen(),
                         kind="ghost", font_size=10, padx=12, pady=6).pack(side="left", padx=6)
        theme.PillButton(akt, "✉ E-Mail", self._lief_email, kind="ghost",
                         font_size=10, padx=12, pady=6).pack(side="left", padx=6)
        theme.PillButton(akt, "Wiedervorlage", self._lief_wiedervorlage, kind="ghost",
                         font_size=10, padx=12, pady=6).pack(side="left", padx=6)
        theme.PillButton(akt, "⤓ Excel", self._lief_export, kind="neutral",
                         font_size=10, padx=12, pady=6).pack(side="left", padx=6)
        theme.PillButton(akt, "Löschen", self._lief_loeschen_selected, kind="danger",
                         font_size=10, padx=12, pady=6).pack(side="left")
        self._lief_reload()

    def _lief_reload(self):
        for i in self._lief_tree.get_children():
            self._lief_tree.delete(i)
        def bev(r):
            return bool(r["bevorzugt"]) if "bevorzugt" in r.keys() else False
        # Bevorzugte zuerst, dann nach Name.
        for r in sorted(lieferanten_liste(), key=lambda r: (not bev(r), (r["name"] or "").lower())):
            tag = "bevorzugt" if bev(r) else ("" if r["aktiv"] else "inaktiv")
            self._lief_tree.insert(
                "", "end", iid=str(r["id"]),
                values=("★" if bev(r) else "", r["name"], r["land"] or "—", r["waehrung"] or "EUR",
                        f"{r['lieferzeit_tage']} Tage" if r["lieferzeit_tage"] else "—",
                        _eur(r["mindestbestellwert"]) if r["mindestbestellwert"] else "—",
                        "✓" if r["gdp_zertifiziert"] else "—",
                        "aktiv" if r["aktiv"] else "inaktiv"),
                tags=(tag,))
        self._lief_tree.tag_configure("inaktiv", foreground=MUTED)
        self._lief_tree.tag_configure("bevorzugt", foreground=ACCENT)

    def _lief_toggle_bevorzugt(self):
        i = self._lief_selected_id()
        if not i:
            messagebox.showinfo("Lieferanten", "Bitte zuerst einen Lieferanten auswählen.", parent=self)
            return
        r = self._lief_row(i)
        aktuell = bool(r["bevorzugt"]) if r and "bevorzugt" in r.keys() else False
        with _con() as con:
            con.execute("UPDATE tbl_einkauf_lieferanten SET bevorzugt=? WHERE id=?",
                        (0 if aktuell else 1, i))
            con.commit()
        _log("Lieferant bevorzugt" if not aktuell else "Lieferant nicht mehr bevorzugt", f"id={i}")
        self._lief_reload()
        if str(i) in self._lief_tree.get_children():
            self._lief_tree.selection_set(str(i))
            self._lief_tree.see(str(i))

    def _lief_selected_id(self):
        sel = self._lief_tree.selection()
        return int(sel[0]) if sel else None

    def _lief_edit_selected(self):
        i = self._lief_selected_id()
        if i:
            self._lieferant_dialog(i)
        else:
            messagebox.showinfo("Lieferanten", "Bitte zuerst einen Lieferanten auswählen.", parent=self)

    def _lief_zeige_quellen(self):
        i = self._lief_selected_id()
        if i:
            self.show("quellen")
            self._quellen_filter_lief.set(lieferant_namen().get(i, "Alle"))
            self._quellen_reload()

    def _lief_row(self, lieferant_id):
        with _con() as con:
            con.row_factory = sqlite3.Row
            return con.execute("SELECT * FROM tbl_einkauf_lieferanten WHERE id=?",
                               (lieferant_id,)).fetchone()

    def _lief_email(self):
        i = self._lief_selected_id()
        if not i:
            messagebox.showinfo("Lieferanten", "Bitte zuerst einen Lieferanten auswählen.", parent=self)
            return
        r = self._lief_row(i)
        if not r or not (r["email"] or "").strip():
            messagebox.showinfo("E-Mail", "Für diesen Lieferanten ist keine E-Mail hinterlegt.", parent=self)
            return
        try:
            webbrowser.open(f"mailto:{r['email'].strip()}")
        except Exception:
            messagebox.showinfo("E-Mail", _T('E-Mail-Adresse: {p0}', p0=r['email']), parent=self)

    def _lief_wiedervorlage(self):
        i = self._lief_selected_id()
        if not i:
            messagebox.showinfo("Lieferanten", "Bitte zuerst einen Lieferanten auswählen.", parent=self)
            return
        r = self._lief_row(i)
        name = r["name"] if r else ""
        AufgabeDialog(self, None, on_save=self._after_aufgabe_save,
                      vorgabe={"titel": f"Bei {name} melden", "kategorie": "Wiedervorlage",
                               "lieferant_id": i})

    def _lief_export(self):
        rows = lieferanten_liste()
        if not rows:
            messagebox.showinfo("Export", "Keine Lieferanten zum Exportieren.", parent=self)
            return
        headers = ["Name", "Land", "Währung", "Ansprechpartner", "E-Mail", "Telefon",
                   "Lieferzeit (Tage)", "Mind.-Bestellwert (€)", "Zahlungsziel (Tage)",
                   "GDP", "Status"]
        data = [[r["name"], r["land"], r["waehrung"], r["ansprechpartner"], r["email"],
                 r["telefon"], r["lieferzeit_tage"], r["mindestbestellwert"],
                 r["zahlungsziel_tage"], "ja" if r["gdp_zertifiziert"] else "nein",
                 "aktiv" if r["aktiv"] else "inaktiv"] for r in rows]
        self._export_und_oeffnen(headers, data, "Lieferanten", "Einkauf_Lieferanten")

    def _export_und_oeffnen(self, headers, data, blatt, basis):
        try:
            pfad = _export_xlsx(headers, data, blatt, basis)
        except Exception as exc:
            messagebox.showerror("Export", _T('Export fehlgeschlagen:\n{p0}', p0=exc), parent=self)
            return
        if messagebox.askyesno("Export", _T('Excel erstellt:\n{p0}\n\nJetzt öffnen?', p0=pfad), parent=self):
            _datei_oeffnen(pfad)

    def _lief_loeschen_selected(self):
        i = self._lief_selected_id()
        if not i:
            return
        if messagebox.askyesno("Löschen", "Lieferant inkl. aller Beschaffungsquellen löschen?",
                               parent=self):
            with _con() as con:
                con.execute("DELETE FROM tbl_einkauf_lieferanten WHERE id=?", (i,))
                con.commit()
            _log("Lieferant gelöscht", f"id={i}")
            self._lief_reload()

    def _lieferant_dialog(self, lieferant_id=None):
        LieferantDialog(self, lieferant_id, on_save=self._lief_reload)

    # ── Seite: Beschaffungsquellen ───────────────────────────────────────────
    def _page_quellen(self):
        theme.page_header(self.content, "Beschaffungsquellen",
                          "Einkaufspreise je Artikel & Lieferant – Basis für Marge und Bestellvorschlag.",
                          bg=SHELL_BG).pack(fill="x", padx=24, pady=(20, 8))
        bar = tk.Frame(self.content, bg=SHELL_BG)
        bar.pack(fill="x", padx=24, pady=(0, 8))
        theme.PillButton(bar, "➕  Neue Quelle", lambda: self._quelle_dialog(),
                         kind="primary", padx=14, pady=8).pack(side="left")
        tk.Label(bar, text="Lieferant:", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(side="left", padx=(14, 4))
        namen = ["Alle"] + list(lieferant_namen().values())
        gespeichert = self._filter_state["quellen_lief"]
        self._quellen_filter_lief = tk.StringVar(value=gespeichert if gespeichert in namen else "Alle")
        ttk.Combobox(bar, textvariable=self._quellen_filter_lief, values=namen,
                     state="readonly", width=24, style="NMG.TCombobox").pack(side="left")
        self._quellen_filter_lief.trace_add("write", lambda *_: self._quellen_reload())

        wrap, self._quellen_tree = self._make_tree(
            self.content,
            ("PZN", "Artikel", "Lieferant", "EK (FW)", "EK (EUR)", "Min.", "AVP DE", "Marge"),
            (90, 250, 150, 110, 100, 60, 100, 130), height=15,
            numerisch=("PZN", "EK (FW)", "EK (EUR)", "Min.", "AVP DE", "Marge"))
        wrap.pack(fill="both", expand=True, padx=24, pady=(4, 4))
        self._quellen_tree.bind("<Double-1>", lambda e: self._quelle_rechnen_selected())

        akt = tk.Frame(self.content, bg=SHELL_BG)
        akt.pack(fill="x", padx=24, pady=(0, 16))
        theme.PillButton(akt, "📐  Marge rechnen", self._quelle_rechnen_selected, kind="accent",
                         font_size=10, padx=12, pady=6).pack(side="left")
        theme.PillButton(akt, "Bearbeiten", self._quelle_edit_selected, kind="neutral",
                         font_size=10, padx=12, pady=6).pack(side="left", padx=6)
        theme.PillButton(akt, "⤓ Excel", self._quellen_export, kind="neutral",
                         font_size=10, padx=12, pady=6).pack(side="left", padx=6)
        theme.PillButton(akt, "Löschen", self._quelle_loeschen_selected, kind="danger",
                         font_size=10, padx=12, pady=6).pack(side="left")
        self._quellen_reload()

    def _quellen_reload(self):
        if not hasattr(self, "_quellen_tree") or not self._quellen_tree.winfo_exists():
            return
        for i in self._quellen_tree.get_children():
            self._quellen_tree.delete(i)
        flt = self._quellen_filter_lief.get()
        self._filter_state["quellen_lief"] = flt
        with _con() as con:
            con.row_factory = sqlite3.Row
            sql = ("SELECT q.*, l.name AS lieferant FROM tbl_einkauf_quellen q "
                   "LEFT JOIN tbl_einkauf_lieferanten l ON l.id=q.lieferant_id ")
            params = ()
            if flt and flt != "Alle":
                sql += "WHERE l.name=? "
                params = (flt,)
            sql += "ORDER BY q.artikelname, q.ek_fremd"
            rows = con.execute(sql, params).fetchall()
        for r in rows:
            k = quelle_kennzahlen(r)
            marge_txt = "—"
            if k["marge"] is not None:
                marge_txt = f"{_eur(k['marge'])} ({k['marge_proz']:.0f}%)"
            self._quellen_tree.insert(
                "", "end", iid=str(r["id"]),
                values=(r["pzn"], k["artikelname"] or "—", r["lieferant"] or "—",
                        f"{(r['ek_fremd'] or 0):.2f} {r['waehrung']}", _eur(k["ek_eur"]),
                        r["mindestabnahme"] or "—", _eur(k["avp"]) if k["avp"] else "—", marge_txt),
                tags=("" if r["aktiv"] else "inaktiv",))
        self._quellen_tree.tag_configure("inaktiv", foreground=MUTED)

    def _quellen_export(self):
        flt = self._quellen_filter_lief.get()
        with _con() as con:
            con.row_factory = sqlite3.Row
            sql = ("SELECT q.*, l.name AS lieferant FROM tbl_einkauf_quellen q "
                   "LEFT JOIN tbl_einkauf_lieferanten l ON l.id=q.lieferant_id ")
            params = ()
            if flt and flt != "Alle":
                sql += "WHERE l.name=? "
                params = (flt,)
            sql += "ORDER BY q.artikelname, q.ek_fremd"
            rows = con.execute(sql, params).fetchall()
        if not rows:
            messagebox.showinfo("Export", "Keine Quellen zum Exportieren.", parent=self)
            return
        headers = ["PZN", "Artikel", "Lieferant", "EK (Fremdwährung)", "Währung", "EK (EUR)",
                   "Mindestabnahme", "AVP DE", "Marge (€)", "Marge (%)", "§129", "Status"]
        data = []
        for r in rows:
            k = quelle_kennzahlen(r)
            data.append([r["pzn"], k["artikelname"], r["lieferant"] or "", r["ek_fremd"],
                         r["waehrung"], round(k["ek_eur"], 2), r["mindestabnahme"],
                         round(k["avp"], 2) if k["avp"] else "", round(k["marge"], 2) if k["marge"] is not None else "",
                         round(k["marge_proz"], 1) if k["marge"] is not None else "",
                         ("erfüllt" if k["p129_ok"] else "nicht erfüllt") if k["p129_ok"] is not None else "",
                         "aktiv" if r["aktiv"] else "inaktiv"])
        self._export_und_oeffnen(headers, data, "Beschaffungsquellen", "Einkauf_Quellen")

    def _quelle_selected_id(self):
        sel = self._quellen_tree.selection()
        return int(sel[0]) if sel else None

    def _quelle_edit_selected(self):
        i = self._quelle_selected_id()
        if i:
            self._quelle_dialog(i)
        else:
            messagebox.showinfo("Quellen", "Bitte zuerst eine Quelle auswählen.", parent=self)

    def _quelle_loeschen_selected(self):
        i = self._quelle_selected_id()
        if not i:
            return
        if messagebox.askyesno("Löschen", "Diese Beschaffungsquelle löschen?", parent=self):
            with _con() as con:
                con.execute("DELETE FROM tbl_einkauf_quellen WHERE id=?", (i,))
                con.commit()
            _log("Quelle gelöscht", f"id={i}")
            self._quellen_reload()

    def _quelle_rechnen_selected(self):
        i = self._quelle_selected_id()
        if not i:
            messagebox.showinfo("Quellen", "Bitte zuerst eine Quelle auswählen.", parent=self)
            return
        with _con() as con:
            con.row_factory = sqlite3.Row
            r = con.execute("SELECT * FROM tbl_einkauf_quellen WHERE id=?", (i,)).fetchone()
        if r:
            self.show("marge", prefill={"pzn": r["pzn"], "ek_fremd": r["ek_fremd"],
                                        "waehrung": r["waehrung"]})

    def _quelle_dialog(self, quelle_id=None):
        QuelleDialog(self, quelle_id, on_save=self._quellen_reload)

    # ── Seite: Beschaffungsvorschlag ─────────────────────────────────────────
    def _page_vorschlag(self):
        theme.page_header(self.content, "Beschaffungsvorschlag",
                          "Beste Quelle je Artikel – günstigster EK, §129-Konformität und Marge.",
                          bg=SHELL_BG).pack(fill="x", padx=24, pady=(20, 8))
        bar = tk.Frame(self.content, bg=SHELL_BG)
        bar.pack(fill="x", padx=24, pady=(0, 8))
        self._vor_nur129 = tk.IntVar(value=self._filter_state["vor_nur129"])
        tk.Checkbutton(bar, text="Nur §129-konforme anzeigen", variable=self._vor_nur129,
                       bg=SHELL_BG, fg=TEXT, activebackground=SHELL_BG, font=theme.SMALL,
                       command=self._vorschlag_reload).pack(side="left")
        self._vor_nur_absatz = tk.IntVar(value=self._filter_state["vor_nur_absatz"])
        tk.Checkbutton(bar, text="Nur mit Absatz", variable=self._vor_nur_absatz,
                       bg=SHELL_BG, fg=TEXT, activebackground=SHELL_BG, font=theme.SMALL,
                       command=self._vorschlag_reload).pack(side="left", padx=(12, 0))
        theme.PillButton(bar, "⤓ Excel", self._vorschlag_export, kind="neutral",
                         font_size=10, padx=12, pady=6).pack(side="right")

        wrap, self._vor_tree = self._make_tree(
            self.content,
            ("PZN", "Artikel", "Beste Quelle", "EK (EUR)", "AVP DE", "§129", "Marge", "Marge %", "Min.", "Absatz"),
            (85, 230, 160, 95, 95, 70, 95, 75, 55, 70), height=15,
            numerisch=("PZN", "EK (EUR)", "AVP DE", "Marge", "Marge %", "Min.", "Absatz"))
        wrap.pack(fill="both", expand=True, padx=24, pady=(4, 4))
        self._vor_tree.bind("<Double-1>", lambda e: self._vorschlag_rechnen())

        akt = tk.Frame(self.content, bg=SHELL_BG)
        akt.pack(fill="x", padx=24, pady=(0, 16))
        theme.PillButton(akt, "📐  Marge rechnen", self._vorschlag_rechnen, kind="accent",
                         font_size=10, padx=12, pady=6).pack(side="left")
        theme.PillButton(akt, "📋  Nachbestellen (Aufgabe)", self._vorschlag_aufgabe, kind="primary",
                         font_size=10, padx=12, pady=6).pack(side="left", padx=6)
        self._vor_hinweis = tk.Label(akt, text="", bg=SHELL_BG, fg=MUTED, font=theme.SMALL)
        self._vor_hinweis.pack(side="right")
        self._vorschlag_reload()

    def _vorschlag_best(self):
        """Liefert je PZN die Quelle mit der höchsten Marge (sqlite3.Row + Kennzahlen)."""
        with _con() as con:
            con.row_factory = sqlite3.Row
            rows = con.execute(
                "SELECT q.*, l.name AS lieferant, l.id AS lid FROM tbl_einkauf_quellen q "
                "JOIN tbl_einkauf_lieferanten l ON l.id=q.lieferant_id "
                "WHERE q.aktiv=1 AND l.aktiv=1").fetchall()
        absatz = absatz_je_pzn()
        best: dict[str, tuple] = {}
        for r in rows:
            k = quelle_kennzahlen(r)
            score = k["marge"] if k["marge"] is not None else -1e18
            cur = best.get(r["pzn"])
            cur_score = cur[1]["marge"] if cur and cur[1]["marge"] is not None else -1e18
            if cur is None or score > cur_score:
                best[r["pzn"]] = (r, k, absatz.get(r["pzn"], 0))
        return best

    def _vorschlag_reload(self):
        if not hasattr(self, "_vor_tree") or not self._vor_tree.winfo_exists():
            return
        for i in self._vor_tree.get_children():
            self._vor_tree.delete(i)
        self._vor_map = {}
        best = self._vorschlag_best()
        nur129 = self._vor_nur129.get()
        nur_absatz = self._vor_nur_absatz.get()
        self._filter_state["vor_nur129"] = nur129
        self._filter_state["vor_nur_absatz"] = nur_absatz
        eintraege = sorted(best.values(),
                           key=lambda t: (t[1]["marge"] if t[1]["marge"] is not None else -1e18),
                           reverse=True)
        gezeigt = 0
        for r, k, absatz in eintraege:
            if nur129 and not k["p129_ok"]:
                continue
            if nur_absatz and absatz <= 0:
                continue
            p129 = "—" if k["p129_ok"] is None else ("✓" if k["p129_ok"] else "✗")
            tag = "" if k["p129_ok"] in (None, True) else "warn"
            iid = str(r["id"])
            self._vor_map[iid] = {"pzn": r["pzn"], "lieferant_id": r["lid"],
                                  "lieferant": r["lieferant"], "artikel": k["artikelname"],
                                  "ek_fremd": r["ek_fremd"], "waehrung": r["waehrung"]}
            self._vor_tree.insert(
                "", "end", iid=iid,
                values=(r["pzn"], k["artikelname"] or "—", r["lieferant"] or "—",
                        _eur(k["ek_eur"]), _eur(k["avp"]) if k["avp"] else "—", p129,
                        _eur(k["marge"]) if k["marge"] is not None else "—",
                        f"{k['marge_proz']:.0f}%" if k["marge"] is not None else "—",
                        r["mindestabnahme"] or "—", absatz or "—"),
                tags=(tag,))
            gezeigt += 1
        self._vor_tree.tag_configure("warn", foreground=theme.DANGER)
        self._vor_hinweis.config(
            text=_T('{p0} Artikel · sortiert nach Marge. Doppelklick = im Rechner prüfen.', p0=gezeigt))

    def _vorschlag_sel(self):
        sel = self._vor_tree.selection()
        return self._vor_map.get(sel[0]) if sel else None

    def _vorschlag_rechnen(self):
        d = self._vorschlag_sel()
        if not d:
            messagebox.showinfo("Beschaffungsvorschlag", "Bitte zuerst eine Zeile auswählen.", parent=self)
            return
        self.show("marge", prefill={"pzn": d["pzn"], "ek_fremd": d["ek_fremd"],
                                    "waehrung": d["waehrung"]})

    def _vorschlag_aufgabe(self):
        d = self._vorschlag_sel()
        if not d:
            messagebox.showinfo("Beschaffungsvorschlag", "Bitte zuerst eine Zeile auswählen.", parent=self)
            return
        AufgabeDialog(self, None, on_save=self._after_aufgabe_save,
                      vorgabe={"titel": f"Nachbestellen: {d['artikel']}".strip()[:120],
                               "kategorie": "Bestellung", "lieferant_id": d["lieferant_id"],
                               "beschreibung": f"PZN {d['pzn']} bei {d['lieferant']} "
                                               f"(EK {d['ek_fremd']} {d['waehrung']})"})

    def _vorschlag_export(self):
        best = self._vorschlag_best()
        if not best:
            messagebox.showinfo("Export", "Keine Beschaffungsquellen vorhanden.", parent=self)
            return
        headers = ["PZN", "Artikel", "Beste Quelle", "EK (EUR)", "AVP DE", "§129",
                   "Marge (€)", "Marge (%)", "Mindestabnahme", "Absatz"]
        data = []
        for r, k, absatz in sorted(best.values(),
                                   key=lambda t: (t[1]["marge"] if t[1]["marge"] is not None else -1e18),
                                   reverse=True):
            data.append([r["pzn"], k["artikelname"], r["lieferant"], round(k["ek_eur"], 2),
                         round(k["avp"], 2) if k["avp"] else "",
                         ("erfüllt" if k["p129_ok"] else "nicht erfüllt") if k["p129_ok"] is not None else "",
                         round(k["marge"], 2) if k["marge"] is not None else "",
                         round(k["marge_proz"], 1) if k["marge"] is not None else "",
                         r["mindestabnahme"], absatz])
        self._export_und_oeffnen(headers, data, "Beschaffungsvorschlag", "Einkauf_Vorschlag")

    # ── Seite: Importkandidaten ──────────────────────────────────────────────
    def _page_importkandidaten(self):
        theme.page_header(self.content, "Importkandidaten",
                          "Lohnende Parallelimport-Chancen aus den Bedarfsanalysen – "
                          "nachgefragte Artikel, die NMG noch nicht führt.",
                          bg=SHELL_BG).pack(fill="x", padx=24, pady=(20, 8))

        # Beim Öffnen des Reiters immer frisch berechnen (Cache verwerfen), damit
        # neue Bedarfsanalysen/Markierungen sofort einfließen. Das Hersteller-Dropdown
        # kommt aus einer leichten DISTINCT-Abfrage ohne die teure Vollberechnung.
        self._ik_cache = None
        herst_namen = ["Alle"] + kandidat_hersteller_namen()
        if len(herst_namen) == 1:  # gar keine Bedarfsanalysen vorhanden
            card = theme.Card(self.content)
            card.pack(fill="x", padx=24, pady=(8, 0))
            tk.Label(card.inner, text="Noch keine Importkandidaten", bg=BG, fg=TEXT,
                     font=theme.SECTION).pack(anchor="w")
            tk.Label(card.inner,
                     text=("Grundlage sind gespeicherte Bedarfsanalysen in NMGone. Sobald "
                           "Apotheken-Auswertungen vorliegen, erscheinen hier die Artikel mit "
                           "Marktnachfrage, die NMG noch nicht als Parallelimport führt."),
                     bg=BG, fg=MUTED, font=theme.BODY, wraplength=820, justify="left").pack(
                anchor="w", pady=(6, 4))
            return

        st = self._ik_state  # gemerkte Filter (bis zur nächsten Änderung)

        # Filterleiste
        flt = theme.Card(self.content)
        flt.pack(fill="x", padx=24, pady=(0, 8))
        g = flt.inner
        for c in range(6):
            g.columnconfigure(c, weight=1 if c in (1, 3) else 0)
        tk.Label(g, text="Hersteller", bg=BG, fg=MUTED, font=theme.SMALL).grid(
            row=0, column=0, sticky="w", padx=(0, 6))
        herst_vorgabe = st.get("herst", "Alle") if st.get("herst", "Alle") in herst_namen else "Alle"
        self._ik_herst = tk.StringVar(value=herst_vorgabe)
        ttk.Combobox(g, textvariable=self._ik_herst, values=herst_namen, state="readonly",
                     width=22, style="NMG.TCombobox").grid(row=0, column=1, sticky="ew", padx=(0, 16))
        tk.Label(g, text="Wirkstoff enthält", bg=BG, fg=MUTED, font=theme.SMALL).grid(
            row=0, column=2, sticky="w", padx=(0, 6))
        self._ik_wirk = tk.StringVar(value=st.get("wirk", ""))
        ttk.Entry(g, textvariable=self._ik_wirk, width=20).grid(row=0, column=3, sticky="ew", padx=(0, 16))
        tk.Label(g, text="Preis €  von / bis", bg=BG, fg=MUTED, font=theme.SMALL).grid(
            row=0, column=4, sticky="w", padx=(0, 6))
        pbox = tk.Frame(g, bg=BG)
        pbox.grid(row=0, column=5, sticky="w")
        self._ik_pmin = tk.StringVar(value=st.get("pmin", ""))
        self._ik_pmax = tk.StringVar(value=st.get("pmax", ""))
        ttk.Entry(pbox, textvariable=self._ik_pmin, width=8).pack(side="left")
        tk.Label(pbox, text="–", bg=BG, fg=MUTED).pack(side="left", padx=2)
        ttk.Entry(pbox, textvariable=self._ik_pmax, width=8).pack(side="left")

        z2 = tk.Frame(g, bg=BG)
        z2.grid(row=1, column=0, columnspan=6, sticky="w", pady=(8, 0))
        tk.Label(z2, text="Artikel-Typ", bg=BG, fg=MUTED, font=theme.SMALL).pack(side="left")
        self._ik_typ = tk.StringVar(value=st.get("typ", "Nur Original"))
        ttk.Combobox(z2, textvariable=self._ik_typ, values=["Alle", "Nur Original", "Nur Reimport"],
                     state="readonly", width=14, style="NMG.TCombobox").pack(side="left", padx=(6, 16))
        tk.Label(z2, text="Min. Apotheken", bg=BG, fg=MUTED, font=theme.SMALL).pack(side="left")
        self._ik_minapo = tk.StringVar(value=st.get("minapo", ""))
        ttk.Entry(z2, textvariable=self._ik_minapo, width=6).pack(side="left", padx=(6, 16))
        tk.Label(z2, text="Status", bg=BG, fg=MUTED, font=theme.SMALL).pack(side="left")
        self._ik_status = tk.StringVar(value=st.get("status", "Alle"))
        ttk.Combobox(z2, textvariable=self._ik_status, values=["Alle"] + KANDIDAT_STATUS,
                     state="readonly", width=13, style="NMG.TCombobox").pack(side="left", padx=(6, 16))

        z3 = tk.Frame(g, bg=BG)
        z3.grid(row=2, column=0, columnspan=6, sticky="w", pady=(8, 0))
        self._ik_nurwichtig = tk.IntVar(value=st.get("nurwichtig", 0))
        tk.Checkbutton(z3, text="Nur wichtige Hersteller (★)", variable=self._ik_nurwichtig,
                       bg=BG, fg=TEXT, activebackground=BG, font=theme.SMALL,
                       command=self._importkand_reload).pack(side="left")
        self._ik_machbar = tk.IntVar(value=st.get("machbar", 0))
        tk.Checkbutton(z3, text="Nur §129-machbar", variable=self._ik_machbar,
                       bg=BG, fg=TEXT, activebackground=BG, font=theme.SMALL,
                       command=self._importkand_reload).pack(side="left", padx=(12, 0))
        self._ik_ohnekonk = tk.IntVar(value=st.get("ohnekonk", 0))
        tk.Checkbutton(z3, text="Nur ohne Reimporteur (unbesetzt)", variable=self._ik_ohnekonk,
                       bg=BG, fg=TEXT, activebackground=BG, font=theme.SMALL,
                       command=self._importkand_reload).pack(side="left", padx=(12, 0))
        self._ik_nurbeob = tk.IntVar(value=st.get("nurbeob", 0))
        tk.Checkbutton(z3, text="Nur gemerkte (📌)", variable=self._ik_nurbeob,
                       bg=BG, fg=TEXT, activebackground=BG, font=theme.SMALL,
                       command=self._importkand_reload).pack(side="left", padx=(12, 0))
        theme.PillButton(z3, "Anzeigen", self._importkand_reload, kind="primary",
                         font_size=10, padx=14, pady=6).pack(side="left", padx=(16, 0))
        theme.PillButton(z3, "Zurücksetzen", self._importkand_reset, kind="ghost",
                         font_size=10, padx=12, pady=6).pack(side="left", padx=6)
        theme.PillButton(z3, "↻ Neu berechnen", self._ik_refresh, kind="ghost",
                         font_size=10, padx=12, pady=6).pack(side="left")

        # KPIs
        self._ik_kpi = tk.Frame(self.content, bg=SHELL_BG)
        self._ik_kpi.pack(fill="x", padx=0, pady=(0, 8))

        # Spalten dynamisch nach Sichtbarkeits-Einstellung (PZN/Artikel/Hersteller fix).
        self._ik_colspec = self._ik_sichtbare_spalten()
        header = tuple(c["header"] for c in self._ik_colspec)
        widths = tuple(c["width"] for c in self._ik_colspec)
        numerisch = tuple(c["header"] for c in self._ik_colspec if c["numeric"])
        wrap, self._ik_tree = self._make_tree(self.content, header, widths, height=14,
                                              numerisch=numerisch)
        wrap.pack(fill="both", expand=True, padx=24, pady=(0, 4))
        self._ik_tree.bind("<Double-1>", lambda e: self._ik_detail())

        akt = tk.Frame(self.content, bg=SHELL_BG)
        akt.pack(fill="x", padx=24, pady=(0, 16))
        theme.PillButton(akt, "📌  Merken / Entfernen", self._ik_toggle_beobachtet,
                         kind="accent", font_size=10, padx=12, pady=6).pack(side="left")
        theme.PillButton(akt, "📋  Kontaktieren", self._ik_kontakt,
                         kind="primary", font_size=10, padx=12, pady=6).pack(side="left", padx=6)
        theme.PillButton(akt, "★  Hersteller bevorzugt", self._ik_toggle_wichtig, kind="neutral",
                         font_size=10, padx=12, pady=6).pack(side="left", padx=6)
        theme.PillButton(akt, "📐  Marge rechnen", self._ik_marge, kind="ghost",
                         font_size=10, padx=12, pady=6).pack(side="left", padx=6)
        theme.PillButton(akt, "⤓ Excel", self._ik_export, kind="neutral",
                         font_size=10, padx=12, pady=6).pack(side="left", padx=6)
        self._ik_hinweis = tk.Label(akt, text="", bg=SHELL_BG, fg=MUTED, font=theme.SMALL)
        self._ik_hinweis.pack(side="right")
        self._importkand_reload()  # beim Öffnen immer frisch laden & anzeigen

    def _ik_sichtbare_spalten(self) -> list[dict]:
        """Spalten-Spezifikation der Tabelle, gefiltert nach Sichtbarkeits-Einstellung.
        PZN/Artikel/Hersteller sind immer dabei; der Rest folgt kandidat_spalten()."""
        ampel_sym = {"gruen": "✓", "gelb": "~", "rot": "✗", "—": "—"}

        def f_pot(d):
            return (f"{d['potenzial_jahr']:,.0f} €".replace(",", ".")
                    if d["potenzial_jahr"] is not None else "—")

        alle = [
            ("pzn", "PZN", 76, True, lambda d: d["pzn"], False),
            ("artikel", "Artikel", 200, False,
             lambda d: (("📌 " if d["beobachtet"] else "") + ("★ " if d["wichtig"] else "")
                        + (d["artikel"] or "—"))[:48], False),
            ("hersteller", "Hersteller", 110, False, lambda d: d["hersteller"], False),
            ("typ", "Typ", 70, False, lambda d: d["typ"], True),
            ("wirkstoff", "Wirkstoff", 130, False, lambda d: (d["wirkstoff"] or "—")[:26], True),
            ("absatz", "Absatz 6M", 78, True, lambda d: f"{d['absatz']:,.0f}".replace(",", "."), True),
            ("apo", "Apo.", 44, True, lambda d: d["apotheken"], True),
            ("konkurrenz", "Konk.", 48, True,
             lambda d: (d["konkurrenz"] if d["wirkstoff"] else "?"), True),
            ("trend", "Trend", 60, False, lambda d: d["trend_sym"], True),
            ("preis", "Preis", 80, True,
             lambda d: _eur(d["preis"]) if d["preis"] is not None else "—", True),
            ("ampel", "§129", 52, False, lambda d: ampel_sym.get(d["ampel"], "—"), True),
            ("potenzial", "Potenzial/J", 96, True, f_pot, True),
            ("status", "Status", 96, False, lambda d: d["status"], True),
        ]
        vis = kandidat_spalten()
        out = []
        for key, header, width, numeric, value, optional in alle:
            if optional and not vis.get(key, True):
                continue
            out.append({"key": key, "header": header, "width": width,
                        "numeric": numeric, "value": value})
        return out

    def _ik_filters(self) -> dict:
        typ_map = {"Alle": "Alle", "Nur Original": "Original", "Nur Reimport": "Reimport"}
        return {
            "hersteller": self._ik_herst.get(),
            "wirkstoff": self._ik_wirk.get(),
            "typ": typ_map.get(self._ik_typ.get(), "Alle"),
            "preis_min": _parse_num(self._ik_pmin.get()) if self._ik_pmin.get().strip() else None,
            "preis_max": _parse_num(self._ik_pmax.get()) if self._ik_pmax.get().strip() else None,
            "min_apotheken": _parse_num(self._ik_minapo.get()) if self._ik_minapo.get().strip() else 0,
            "nur_wichtige_hersteller": bool(self._ik_nurwichtig.get()),
            "nur_machbar": bool(self._ik_machbar.get()),
            "nur_ohne_konkurrenz": bool(self._ik_ohnekonk.get()),
            "nur_beobachtet": bool(self._ik_nurbeob.get()),
            "status": self._ik_status.get(),
        }

    def _importkand_reset(self):
        self._ik_herst.set("Alle")
        self._ik_wirk.set("")
        self._ik_pmin.set("")
        self._ik_pmax.set("")
        self._ik_minapo.set("")
        self._ik_nurwichtig.set(0)
        self._ik_typ.set("Nur Original")
        self._ik_machbar.set(0)
        self._ik_ohnekonk.set(0)
        self._ik_nurbeob.set(0)
        self._ik_status.set("Alle")
        self._importkand_reload()

    def _ik_cache_laden(self) -> bool:
        """Sorgt dafür, dass die (teure) Kandidatenliste im Cache liegt. Zeigt
        währenddessen einen kurzen Hinweis + Wartecursor. Liefert True bei Daten."""
        if self._ik_cache is None:
            self._ik_hinweis.config(text="Kandidaten werden berechnet … (einmalig)")
            try:
                self.config(cursor="watch")
                self.update_idletasks()
            except tk.TclError:
                pass
            self._ik_cache = lade_importkandidaten()
            try:
                self.config(cursor="")
            except tk.TclError:
                pass
        self._ik_all = self._ik_cache
        return bool(self._ik_cache)

    def _ik_refresh(self):
        """Verwirft den Cache und berechnet neu (z. B. nach neuen Bedarfsanalysen)."""
        self._ik_cache = None
        self._importkand_reload()

    def _importkand_reload(self):
        if not hasattr(self, "_ik_tree") or not self._ik_tree.winfo_exists():
            return
        # Aktuelle Filter merken, damit sie einen Seitenwechsel überleben.
        self._ik_state = {"herst": self._ik_herst.get(), "wirk": self._ik_wirk.get(),
                          "pmin": self._ik_pmin.get(), "pmax": self._ik_pmax.get(),
                          "minapo": self._ik_minapo.get(), "nurwichtig": self._ik_nurwichtig.get(),
                          "typ": self._ik_typ.get(), "machbar": self._ik_machbar.get(),
                          "ohnekonk": self._ik_ohnekonk.get(), "status": self._ik_status.get(),
                          "nurbeob": self._ik_nurbeob.get()}
        self._ik_cache_laden()
        daten = _filter_kandidaten(self._ik_all, self._ik_filters())
        self._ik_rows = daten
        for i in self._ik_tree.get_children():
            self._ik_tree.delete(i)
        # Anzeige deckeln – das Rendern tausender Zeilen ist langsam; Export/Filter
        # arbeiten weiter mit der vollständigen Treffermenge.
        zeige = daten[:IK_MAX_ROWS]
        for idx, d in enumerate(zeige):
            tag = ("beobachtet" if d["beobachtet"] else
                   ("wichtig" if d["wichtig"] else
                    ("reimport" if d["typ"] == "Reimport" else "")))
            self._ik_tree.insert(
                "", "end", iid=str(idx),
                values=tuple(c["value"](d) for c in self._ik_colspec), tags=(tag,))
        self._ik_tree.tag_configure("wichtig", foreground=ACCENT)
        self._ik_tree.tag_configure("reimport", foreground=MUTED)
        self._ik_tree.tag_configure("beobachtet", background=theme.SELECT_BG)

        # KPIs
        for w in self._ik_kpi.winfo_children():
            w.destroy()
        sum_pot = sum(d["potenzial_jahr"] or 0 for d in daten)
        ohne_konk = sum(1 for d in daten if d["konkurrenz"] == 0 and d["wirkstoff"])
        self._kpi_cards(self._ik_kpi, [
            ("Kandidaten", f"{len(daten):,}".replace(",", "."), ACCENT),
            ("Σ Potenzial/Jahr (geschätzt)", _eur(sum_pot), theme.PURPLE),
            ("Unbesetzt (kein Reimporteur)", f"{ohne_konk:,}".replace(",", "."), OK_GREEN),
        ])
        mehr = (f" · Anzeige auf {IK_MAX_ROWS} begrenzt – Filter verfeinern"
                if len(daten) > IK_MAX_ROWS else "")
        self._ik_hinweis.config(
            text=_T('{p0} Treffer · sortiert nach Jahres-Potenzial{p1} · Doppelklick = Details.', p0=len(daten), p1=mehr))

    def _ik_selected(self) -> dict | None:
        sel = self._ik_tree.selection()
        if not sel:
            return None
        try:
            return self._ik_rows[int(sel[0])]
        except (ValueError, IndexError):
            return None

    def _ik_detail(self, d=None):
        if d is None:
            d = self._ik_selected()
        if not d:
            messagebox.showinfo("Importkandidaten", "Bitte zuerst einen Artikel auswählen.", parent=self)
            return
        KandidatDialog(self, d)

    def _ik_kontakt(self, d=None):
        if d is None:
            d = self._ik_selected()
        if not d:
            messagebox.showinfo("Importkandidaten", "Bitte zuerst einen Artikel auswählen.", parent=self)
            return
        preis = _eur(d["preis"]) if d["preis"] is not None else "—"
        # Falls dem Hersteller ein Lieferant zugeordnet ist, gleich vorbelegen
        # (bevorzugter zuerst).
        liefs = lieferanten_fuer_hersteller(d["hersteller"])
        vorgabe = {"titel": f"Parallelimport prüfen: {d['artikel']}".strip()[:120],
                   "kategorie": "Anfrage",
                   "beschreibung": (f"PZN {d['pzn']} · Hersteller {d['hersteller']} · "
                                    f"Absatz 6M {d['absatz']:.0f} in {d['apotheken']} Apotheken · "
                                    f"Preis {preis} ({d['preis_quelle']}). "
                                    f"Bei Firma anfragen, ob als Parallelimport beschaffbar.")}
        if liefs:
            vorgabe["lieferant_id"] = liefs[0][0]
        AufgabeDialog(self, None, on_save=self._after_aufgabe_save, vorgabe=vorgabe)

    def _ik_toggle_wichtig(self, d=None):
        if d is None:
            d = self._ik_selected()
        if not d:
            messagebox.showinfo("Importkandidaten", "Bitte zuerst einen Artikel auswählen.", parent=self)
            return
        neu = not d["wichtig"]
        merk_pzn = d["pzn"]
        set_hersteller_wichtig(d["hersteller"], neu)
        # In-Memory-Liste nachziehen, damit Filter/★ ohne Neuladen stimmen.
        wichtig = wichtige_hersteller()
        for c in self._ik_all:
            c["wichtig"] = c["hersteller"] in wichtig
        self._importkand_reload()
        # Auswahl auf derselben Zeile halten (sofern weiterhin sichtbar).
        for idx, row in enumerate(getattr(self, "_ik_rows", [])):
            if row["pzn"] == merk_pzn:
                self._ik_tree.selection_set(str(idx))
                self._ik_tree.see(str(idx))
                break
        return neu

    def _ik_toggle_beobachtet(self, d=None):
        if d is None:
            d = self._ik_selected()
        if not d:
            messagebox.showinfo("Importkandidaten", "Bitte zuerst einen Artikel auswählen.", parent=self)
            return
        neu = not d["beobachtet"]
        set_kandidat_beobachtet(d["pzn"], neu)
        d["beobachtet"] = neu
        merk_pzn = d["pzn"]
        # Cache neu sortieren (gemerkte nach oben) und Auswahl halten.
        if self._ik_cache is not None:
            self._ik_cache.sort(key=lambda c: (c["beobachtet"], c["potenzial_jahr"] is not None,
                                               c["potenzial_jahr"] or 0.0, c["absatz"]), reverse=True)
        self._importkand_reload()
        for idx, row in enumerate(getattr(self, "_ik_rows", [])[:IK_MAX_ROWS]):
            if row["pzn"] == merk_pzn:
                self._ik_tree.selection_set(str(idx))
                self._ik_tree.see(str(idx))
                break
        return neu

    def _ik_toggle_reimporteur(self, d=None):
        if d is None:
            d = self._ik_selected()
        if not d:
            messagebox.showinfo("Importkandidaten", "Bitte zuerst einen Artikel auswählen.", parent=self)
            return
        neu = d["typ"] != "Reimport"
        set_hersteller_reimporteur(d["hersteller"], neu)
        # Typ aller Zeilen desselben Anbieters neu bewerten.
        benutzer = benutzer_reimporteure()
        for c in self._ik_all:
            c["typ"] = "Reimport" if _ist_reimporteur(c["hersteller"], benutzer) else "Original"
        self._importkand_reload()
        return "Reimport" if d["typ"] == "Reimport" else "Original"

    def _ik_marge(self, d=None):
        if d is None:
            d = self._ik_selected()
        if not d:
            messagebox.showinfo("Importkandidaten", "Bitte zuerst einen Artikel auswählen.", parent=self)
            return
        prefill = {"pzn": d["pzn"]}
        if d["preis"] is not None:
            prefill["referenz_avp"] = f"{d['preis']:.2f}".replace(".", ",")
        self.show("marge", prefill=prefill)

    def _ik_export(self):
        daten = getattr(self, "_ik_rows", None)
        if not daten:
            messagebox.showinfo("Export", "Keine Importkandidaten zum Exportieren.", parent=self)
            return
        ampel_txt = {"gruen": "machbar", "gelb": "knapp", "rot": "keine Marge", "—": ""}
        headers = ["PZN", "Artikel", "Hersteller", "Typ", "Wichtig", "Wirkstoff", "Absatz 6M",
                   "Trend", "Apotheken", "Reimporteure aktiv", "Umsatz 6M (€)", "APU/Preis (€)",
                   "Preisquelle", "§129-machbar", "Marge/Stück (€)", "Marge (%)",
                   "Potenzial/Jahr (€)", "Status", "Chancen-Score"]
        rows = [[d["pzn"], d["artikel"], d["hersteller"], d["typ"], "ja" if d["wichtig"] else "",
                 d["wirkstoff"], round(d["absatz"], 0), d.get("trend_label", ""),
                 d["apotheken"], d["konkurrenz"], round(d["umsatz"], 2),
                 round(d["preis"], 2) if d["preis"] is not None else "", d["preis_quelle"],
                 ampel_txt.get(d["ampel"], ""),
                 round(d["marge_stueck"], 2) if d["marge_stueck"] is not None else "",
                 round(d["marge_proz"], 1) if d["marge_proz"] is not None else "",
                 round(d["potenzial_jahr"], 0) if d["potenzial_jahr"] is not None else "",
                 d["status"], round(d["score"], 0)]
                for d in daten]
        self._export_und_oeffnen(headers, rows, "Importkandidaten", "Einkauf_Importkandidaten")

    # ── Seite: Statistik & Erfolg ────────────────────────────────────────────
    def _page_statistik(self):
        wrap = self._scrollseite()
        theme.page_header(wrap, "Statistik & Erfolg",
                          "Welche Lieferanten/Hersteller am besten kooperieren – Quellen, "
                          "Marge, Kontakte und Pipeline-Erfolg.",
                          bg=SHELL_BG).pack(fill="x", pady=(20, 12), anchor="w")

        lief = lieferanten_erfolg()
        herst = hersteller_erfolg()
        eingelistet = sum(h["eingelistet"] for h in herst)
        bevorzugt = sum(1 for l in lief if l["bevorzugt"])
        quellen = sum(l["quellen"] for l in lief)
        kpi = tk.Frame(wrap, bg=SHELL_BG)
        kpi.pack(fill="x", pady=(0, 12))
        self._kpi_cards(kpi, [
            ("Lieferanten aktiv", str(sum(1 for l in lief if l["aktiv"])), ACCENT),
            ("davon bevorzugt ★", str(bevorzugt), theme.PURPLE),
            ("Beschaffungsquellen", str(quellen), OK_GREEN),
            ("Artikel eingelistet", str(eingelistet), theme.WARNING),
        ])

        # Lieferanten-Erfolg
        c1 = tk.Frame(wrap, bg=SHELL_BG)
        c1.pack(fill="x", pady=(0, 4))
        tk.Label(c1, text="🏭  Lieferanten-Erfolg", bg=SHELL_BG, fg=TEXT, font=theme.SECTION).pack(side="left")
        theme.PillButton(c1, "⤓ Excel", self._stat_export_lief, kind="neutral",
                         font_size=10, padx=12, pady=6).pack(side="right")
        lwrap, ltree = self._make_tree(
            wrap, ("★", "Lieferant", "Land", "Quellen", "Ø Marge", "Aufgaben", "erledigt", "Bestellungen"),
            (28, 200, 100, 70, 80, 80, 72, 90), height=8,
            numerisch=("Quellen", "Ø Marge", "Aufgaben", "erledigt", "Bestellungen"))
        lwrap.pack(fill="x", pady=(2, 12))
        for l in lief:
            ltree.insert("", "end", values=(
                "★" if l["bevorzugt"] else "", l["name"], l["land"] or "—", l["quellen"],
                f"{l['marge_proz']:.0f}%" if l["marge_proz"] is not None else "—",
                l["aufgaben"], l["erledigt"], l["bestellungen"]),
                tags=("bev" if l["bevorzugt"] else ("" if l["aktiv"] else "inaktiv"),))
        ltree.tag_configure("bev", foreground=ACCENT)
        ltree.tag_configure("inaktiv", foreground=MUTED)
        if not lief:
            tk.Label(wrap, text="Noch keine Lieferanten erfasst.", bg=SHELL_BG, fg=MUTED,
                     font=theme.BODY).pack(anchor="w", pady=(0, 12))

        # Hersteller-Pipeline-Erfolg
        c2 = tk.Frame(wrap, bg=SHELL_BG)
        c2.pack(fill="x", pady=(0, 4))
        tk.Label(c2, text="🧪  Hersteller-Pipeline (Importkandidaten)", bg=SHELL_BG, fg=TEXT,
                 font=theme.SECTION).pack(side="left")
        theme.PillButton(c2, "⤓ Excel", self._stat_export_herst, kind="neutral",
                         font_size=10, padx=12, pady=6).pack(side="right")
        hwrap, htree = self._make_tree(
            wrap, ("Hersteller", "Kandidaten", "gemerkt", "angefragt", "Angebot", "eingelistet",
                   "verworfen", "Erfolgsquote"),
            (140, 90, 70, 80, 80, 90, 80, 90), height=8,
            numerisch=("Kandidaten", "gemerkt", "angefragt", "Angebot", "eingelistet",
                       "verworfen", "Erfolgsquote"))
        hwrap.pack(fill="x", pady=(2, 12))
        for h in herst:
            htree.insert("", "end", values=(
                h["hersteller"], h["gesamt"], h["beobachtet"], h["angefragt"], h["angebot"],
                h["eingelistet"], h["verworfen"],
                f"{h['erfolgsquote']:.0f}%" if h["erfolgsquote"] is not None else "—"))
        if not herst:
            tk.Label(wrap,
                     text="Sobald du in den Importkandidaten Artikel auf „Angefragt“, „Eingelistet“ "
                          "usw. setzt, erscheint hier die Erfolgsauswertung je Hersteller.",
                     bg=SHELL_BG, fg=MUTED, font=theme.BODY, wraplength=820, justify="left").pack(
                anchor="w", pady=(0, 16))

    def _stat_export_lief(self):
        lief = lieferanten_erfolg()
        if not lief:
            messagebox.showinfo("Export", "Keine Lieferanten zum Exportieren.", parent=self)
            return
        headers = ["Lieferant", "Land", "Bevorzugt", "Aktiv", "Beschaffungsquellen", "Ø Marge (%)",
                   "Aufgaben", "erledigt", "Bestellungen"]
        rows = [[l["name"], l["land"], "ja" if l["bevorzugt"] else "", "ja" if l["aktiv"] else "nein",
                 l["quellen"], round(l["marge_proz"], 1) if l["marge_proz"] is not None else "",
                 l["aufgaben"], l["erledigt"], l["bestellungen"]] for l in lief]
        self._export_und_oeffnen(headers, rows, "Lieferanten-Erfolg", "Einkauf_Statistik_Lieferanten")

    def _stat_export_herst(self):
        herst = hersteller_erfolg()
        if not herst:
            messagebox.showinfo("Export", "Noch keine Pipeline-Daten – setze zuerst Kandidaten-Status.",
                                parent=self)
            return
        headers = ["Hersteller", "Kandidaten", "gemerkt", "angefragt", "Angebot/Muster",
                   "eingelistet", "verworfen", "Erfolgsquote (%)"]
        rows = [[h["hersteller"], h["gesamt"], h["beobachtet"], h["angefragt"], h["angebot"],
                 h["eingelistet"], h["verworfen"],
                 round(h["erfolgsquote"], 1) if h["erfolgsquote"] is not None else ""] for h in herst]
        self._export_und_oeffnen(headers, rows, "Hersteller-Pipeline", "Einkauf_Statistik_Hersteller")

    # ── Seite: Margenrechner §129 ────────────────────────────────────────────
    def _page_marge(self):
        wrap = self._scrollseite()
        theme.page_header(wrap, "Margenrechner §129 SGB V",
                          "Import-EK → deutscher AVP/Festbetrag → garantierter Preisabstand & Marge.",
                          bg=SHELL_BG).pack(fill="x", pady=(20, 12), anchor="w")

        pre = self._marge_prefill or {}
        self._marge_prefill = None

        # Eingabe-Karte
        eingabe = theme.Card(wrap)
        eingabe.pack(fill="x", pady=(0, 12))
        grid = eingabe.inner
        for c in range(4):
            grid.columnconfigure(c, weight=1 if c in (1, 3) else 0)

        self._m_vars = {}

        def feld(row, col, label, key, default="", width=16):
            tk.Label(grid, text=label, bg=BG, fg=MUTED, font=theme.SMALL).grid(
                row=row, column=col, sticky="w", padx=(0, 8), pady=(8, 0))
            v = tk.StringVar(value=str(default))
            e = ttk.Entry(grid, textvariable=v, width=width)
            e.grid(row=row + 1, column=col, sticky="ew", padx=(0, 16), pady=(0, 6))
            self._m_vars[key] = v
            return e

        # PZN mit Suche/Übernahme
        tk.Label(grid, text="PZN", bg=BG, fg=MUTED, font=theme.SMALL).grid(
            row=0, column=0, sticky="w", padx=(0, 8), pady=(8, 0))
        pzn_frame = tk.Frame(grid, bg=BG)
        pzn_frame.grid(row=1, column=0, sticky="ew", padx=(0, 16), pady=(0, 6))
        self._m_vars["pzn"] = tk.StringVar(value=str(pre.get("pzn", "")))
        ttk.Entry(pzn_frame, textvariable=self._m_vars["pzn"], width=12).pack(side="left")
        theme.PillButton(pzn_frame, "Laden", self._marge_pzn_laden, kind="neutral",
                         font_size=9, padx=8, pady=4).pack(side="left", padx=(6, 0))

        self._m_artikel_lbl = tk.Label(grid, text="—", bg=BG, fg=TEXT, font=theme.BODY_BOLD,
                                       anchor="w")
        self._m_artikel_lbl.grid(row=1, column=1, columnspan=3, sticky="w", pady=(0, 6))

        feld(2, 0, "Referenz-AVP DE (Original) €", "referenz_avp", pre.get("referenz_avp", ""))
        feld(2, 1, "Import-AVP (unser Listenpreis) €", "import_avp", pre.get("import_avp", ""))
        feld(2, 2, "Import-EK (Fremdwährung)", "ek_fremd", pre.get("ek_fremd", ""))

        tk.Label(grid, text="Währung", bg=BG, fg=MUTED, font=theme.SMALL).grid(
            row=2, column=3, sticky="w", padx=(0, 8), pady=(8, 0))
        self._m_vars["waehrung"] = tk.StringVar(value=str(pre.get("waehrung", get_setting("standard_waehrung", "EUR"))))
        ttk.Combobox(grid, textvariable=self._m_vars["waehrung"], values=waehrungen(),
                     state="readonly", width=8, style="NMG.TCombobox").grid(
            row=3, column=3, sticky="w", padx=(0, 16), pady=(0, 6))

        feld(4, 0, "Unser VK netto (an Apotheke/GH) €", "unser_vk", pre.get("unser_vk", ""))

        btnrow = tk.Frame(grid, bg=BG)
        btnrow.grid(row=5, column=0, columnspan=4, sticky="w", pady=(8, 0))
        theme.PillButton(btnrow, "Berechnen", self._marge_berechnen, kind="primary",
                         padx=18, pady=9).pack(side="left")
        theme.PillButton(btnrow, "Als Aufgabe merken", self._marge_als_aufgabe, kind="ghost",
                         padx=14, pady=9).pack(side="left", padx=8)

        # Ergebnis-Bereich
        self._marge_ergebnis = tk.Frame(wrap, bg=SHELL_BG)
        self._marge_ergebnis.pack(fill="x", pady=(4, 24))

        if pre.get("pzn"):
            self._marge_pzn_laden()
        if any(pre.get(k) for k in ("referenz_avp", "ek_fremd")):
            self._marge_berechnen()

    def _marge_pzn_laden(self):
        pzn = self._m_vars["pzn"].get().strip()
        art = artikel_by_pzn(pzn)
        if not art:
            self._m_artikel_lbl.config(text="Keine NMG-Stammdaten zu dieser PZN gefunden.",
                                       fg=theme.DANGER)
            return
        self._m_artikel_lbl.config(text=art["artikelname"] or "—", fg=TEXT)
        if art.get("taxe_vk"):
            self._m_vars["referenz_avp"].set(f"{art['taxe_vk']:.2f}".replace(".", ","))
            abstand = erforderlicher_abstand(art["taxe_vk"])[0]
            self._m_vars["import_avp"].set(f"{max(0.0, art['taxe_vk'] - abstand):.2f}".replace(".", ","))
        if art.get("taxe_ek"):
            self._m_vars["unser_vk"].set(f"{art['taxe_ek']:.2f}".replace(".", ","))

    def _marge_berechnen(self):
        referenz = _parse_num(self._m_vars["referenz_avp"].get())
        import_avp = _parse_num(self._m_vars["import_avp"].get())
        ek_fremd = _parse_num(self._m_vars["ek_fremd"].get())
        waehrung = self._m_vars["waehrung"].get()
        unser_vk = _parse_num(self._m_vars["unser_vk"].get())
        ek_eur = ek_fremd * kurs_eur(waehrung)
        if referenz <= 0:
            messagebox.showinfo("Margenrechner", "Bitte einen Referenz-AVP eingeben (oder PZN laden).",
                                parent=self)
            return
        if import_avp <= 0:
            import_avp = referenz
        r = margen_rechnung(referenz, import_avp, ek_eur, unser_vk)

        for w in self._marge_ergebnis.winfo_children():
            w.destroy()

        # §129-Status-Banner
        ok = r["p129_ok"]
        banner = tk.Frame(self._marge_ergebnis, bg=(OK_GREEN if ok else theme.DANGER))
        banner.pack(fill="x", pady=(8, 12))
        status = "✓ §129 erfüllt" if ok else "✗ §129 NICHT erfüllt"
        tk.Label(banner, text=_T('{p0}   ·   erf. Preisabstand {p1}', p0=status, p1=r['regel']),
                 bg=(OK_GREEN if ok else theme.DANGER), fg="#FFFFFF",
                 font=(theme.FONT, 13, "bold")).pack(anchor="w", padx=16, pady=10)

        # Kennzahlen
        karten = [
            ("Import-EK (EUR)", _eur(ek_eur), ACCENT),
            ("Erf. Preisabstand", _eur(r["abstand_soll"]), theme.WARNING),
            ("Max. zul. Import-AVP", _eur(r["max_import_avp"]), theme.PURPLE),
            ("Ersparnis Kasse", f"{_eur(r['kassen_ersparnis'])} ({r['abstand_ist_proz']:.1f}%)",
             OK_GREEN if r["kassen_ersparnis"] > 0 else theme.DANGER),
        ]
        grid = tk.Frame(self._marge_ergebnis, bg=SHELL_BG)
        grid.pack(fill="x")
        for i, (label, wert, farbe) in enumerate(karten):
            grid.columnconfigure(i, weight=1)
            c = tk.Frame(grid, bg=BG, highlightbackground=BORDER, highlightthickness=1)
            c.grid(row=0, column=i, sticky="ew", padx=(0 if i == 0 else 8, 0))
            tk.Label(c, text=label, bg=BG, fg=MUTED, font=theme.SMALL).pack(anchor="w", padx=12, pady=(10, 0))
            tk.Label(c, text=wert, bg=BG, fg=farbe, font=(theme.FONT, 16, "bold")).pack(anchor="w", padx=12, pady=(0, 10))

        # Marge-Karte (Verkaufsargument)
        marge_card = theme.Card(self._marge_ergebnis)
        marge_card.pack(fill="x", pady=(12, 0))
        tk.Label(marge_card.inner, text="💶  Unsere Marge", bg=BG, fg=TEXT,
                 font=theme.SECTION).pack(anchor="w")
        zeile = tk.Frame(marge_card.inner, bg=BG)
        zeile.pack(fill="x", pady=(8, 0))
        marge_farbe = OK_GREEN if r["marge"] > 0 else theme.DANGER
        tk.Label(zeile, text=_eur(r["marge"]), bg=BG, fg=marge_farbe,
                 font=(theme.FONT, 26, "bold")).pack(side="left")
        tk.Label(zeile, text=_T('  ({p0:.1f} % vom VK)', p0=r['marge_proz']), bg=BG, fg=MUTED,
                 font=theme.BODY).pack(side="left", anchor="s", pady=(0, 6))
        tk.Label(marge_card.inner,
                 text=(_T('Garantierter Preisabstand zum Original: {p0} – das ist Ihr Verkaufsargument nach §129 SGB V.', p0=_eur(r['kassen_ersparnis']))),
                 bg=BG, fg=MUTED, font=theme.SMALL, wraplength=820, justify="left"
                 ).pack(anchor="w", pady=(8, 0))

    def _marge_als_aufgabe(self):
        pzn = self._m_vars["pzn"].get().strip()
        name = self._m_artikel_lbl.cget("text")
        titel = f"Beschaffung prüfen: {name}" if name not in ("—", "") else "Beschaffung prüfen"
        AufgabeDialog(self, None, on_save=self._after_aufgabe_save,
                      vorgabe={"titel": titel, "kategorie": "Anfrage",
                               "beschreibung": f"PZN {pzn}, Import-EK {self._m_vars['ek_fremd'].get()} "
                                               f"{self._m_vars['waehrung'].get()}"})

    # ── Seite: Wechselkurse ──────────────────────────────────────────────────
    def _page_kurse(self):
        theme.page_header(self.content, "Wechselkurse",
                          "EUR je 1 Einheit Fremdwährung. Tageskurs eintragen für korrekte Margen.",
                          bg=SHELL_BG).pack(fill="x", padx=24, pady=(20, 8))
        bar = tk.Frame(self.content, bg=SHELL_BG)
        bar.pack(fill="x", padx=24, pady=(0, 8))
        theme.PillButton(bar, "➕  Währung/Kurs", lambda: self._kurs_dialog(),
                         kind="primary", padx=14, pady=8).pack(side="left")

        wrap, self._kurs_tree = self._make_tree(
            self.content, ("Währung", "Kurs (EUR je 1)", "Beispiel 100 →", "Aktualisiert"),
            (120, 160, 160, 200), height=12)
        wrap.pack(fill="both", expand=True, padx=24, pady=(4, 4))
        self._kurs_tree.bind("<Double-1>", lambda e: self._kurs_edit_selected())

        akt = tk.Frame(self.content, bg=SHELL_BG)
        akt.pack(fill="x", padx=24, pady=(0, 16))
        theme.PillButton(akt, "Bearbeiten", self._kurs_edit_selected, kind="neutral",
                         font_size=10, padx=12, pady=6).pack(side="left")
        self._kurs_reload()

    def _kurs_reload(self):
        for i in self._kurs_tree.get_children():
            self._kurs_tree.delete(i)
        with _con() as con:
            rows = con.execute("SELECT waehrung, kurs_eur, aktualisiert_am FROM tbl_einkauf_wechselkurse "
                               "ORDER BY CASE WHEN waehrung='EUR' THEN 0 ELSE 1 END, waehrung").fetchall()
        for w, k, ts in rows:
            beispiel = _eur((k or 0) * 100)
            self._kurs_tree.insert("", "end", iid=w,
                                   values=(w, f"{k:.4f}".replace(".", ","), beispiel,
                                           _datum_de(ts) if ts else "—"))

    def _kurs_edit_selected(self):
        sel = self._kurs_tree.selection()
        if not sel:
            messagebox.showinfo("Wechselkurse", "Bitte zuerst eine Währung auswählen.", parent=self)
            return
        self._kurs_dialog(sel[0])

    def _kurs_dialog(self, waehrung=None):
        if waehrung is None:
            w = simpledialog.askstring("Währung", "Währungskürzel (z.B. GBP, CHF, PLN):", parent=self)
            if not w:
                return
            waehrung = w.strip().upper()
        aktuell = kurs_eur(waehrung) if waehrung != "EUR" else 1.0
        wert = simpledialog.askstring(
            "Wechselkurs", _T('Wie viel EUR ist 1 {p0} wert?', p0=waehrung),
            initialvalue=f"{aktuell:.4f}".replace(".", ","), parent=self)
        if wert is None:
            return
        kurs = _parse_num(wert)
        if kurs <= 0:
            messagebox.showerror("Wechselkurs", "Bitte einen Kurs größer 0 eingeben.", parent=self)
            return
        with _con() as con:
            con.execute("INSERT INTO tbl_einkauf_wechselkurse(waehrung, kurs_eur, aktualisiert_am) "
                        "VALUES(?,?,?) ON CONFLICT(waehrung) DO UPDATE SET "
                        "kurs_eur=excluded.kurs_eur, aktualisiert_am=excluded.aktualisiert_am",
                        (waehrung, kurs, _now()))
            con.commit()
        _log("Wechselkurs gesetzt", f"{waehrung}={kurs}")
        self._kurs_reload()

    # ── Seite: Einstellungen ─────────────────────────────────────────────────
    def _page_einstellungen(self):
        wrap = self._scrollseite()
        theme.page_header(wrap, "Einstellungen",
                          "§129-Preisabstand und Standardwerte des Einkaufs.",
                          bg=SHELL_BG).pack(fill="x", pady=(20, 12), anchor="w")

        card = theme.Card(wrap)
        card.pack(fill="x", pady=(0, 12))
        tk.Label(card.inner, text="§129 SGB V – Preisabstand für Importarzneimittel",
                 bg=BG, fg=TEXT, font=theme.SECTION).pack(anchor="w")
        tk.Label(card.inner,
                 text=("Staffel des erforderlichen Preisabstands. Standard: AVP bis 100 € → 15 %, "
                       "AVP 100–300 € → 15 €, AVP über 300 € → 5 %."),
                 bg=BG, fg=MUTED, font=theme.SMALL, wraplength=820, justify="left").pack(anchor="w", pady=(4, 10))

        grid = tk.Frame(card.inner, bg=BG)
        grid.pack(fill="x")
        self._set_vars = {}
        felder = [
            ("p129_t1_grenze", "Stufe 1: AVP-Grenze (€)"),
            ("p129_t1_proz", "Stufe 1: Abstand (%)"),
            ("p129_t2_grenze", "Stufe 2: AVP-Grenze (€)"),
            ("p129_t2_eur", "Stufe 2: Abstand (€)"),
            ("p129_t3_proz", "Stufe 3: Abstand (%)"),
        ]
        for i, (key, label) in enumerate(felder):
            r, c = divmod(i, 2)
            tk.Label(grid, text=label, bg=BG, fg=MUTED, font=theme.SMALL).grid(
                row=r * 2, column=c, sticky="w", padx=(0, 16), pady=(6, 0))
            v = tk.StringVar(value=get_setting(key))
            ttk.Entry(grid, textvariable=v, width=14).grid(
                row=r * 2 + 1, column=c, sticky="w", padx=(0, 16), pady=(0, 4))
            self._set_vars[key] = v

        cardk = theme.Card(wrap)
        cardk.pack(fill="x", pady=(0, 12))
        tk.Label(cardk.inner, text="Importkandidaten – Bewertungs-Annahmen", bg=BG, fg=TEXT,
                 font=theme.SECTION).pack(anchor="w")
        tk.Label(cardk.inner,
                 text=("Grundlage für Potenzial in € und die §129-Machbarkeits-Ampel. Der EU-EK-Anteil "
                       "ist der angenommene Einkaufspreis im EU-Ausland in Prozent des deutschen "
                       "Preisniveaus; die Mindestmarge legt fest, ab welcher geschätzten §129-Marge die "
                       "Ampel grün wird. Werte sind Schätzungen – keine echten EU-Preise."),
                 bg=BG, fg=MUTED, font=theme.SMALL, wraplength=820, justify="left").pack(anchor="w", pady=(4, 10))
        gridk = tk.Frame(cardk.inner, bg=BG)
        gridk.pack(fill="x")
        for i, (key, label) in enumerate([("kandidat_eu_ek_proz", "Angenommener EU-EK (% vom DE-Preis)"),
                                          ("kandidat_min_marge_proz", "Mindestmarge für Ampel grün (%)"),
                                          ("kandidat_fixkosten", "Einführungskosten je Artikel (€)")]):
            tk.Label(gridk, text=label, bg=BG, fg=MUTED, font=theme.SMALL).grid(
                row=0, column=i, sticky="w", padx=(0, 16), pady=(0, 0))
            v = tk.StringVar(value=get_setting(key))
            ttk.Entry(gridk, textvariable=v, width=14).grid(row=1, column=i, sticky="w", padx=(0, 16), pady=(0, 4))
            self._set_vars[key] = v

        cardsp = theme.Card(wrap)
        cardsp.pack(fill="x", pady=(0, 12))
        tk.Label(cardsp.inner, text="Importkandidaten – Spalten anzeigen", bg=BG, fg=TEXT,
                 font=theme.SECTION).pack(anchor="w")
        tk.Label(cardsp.inner,
                 text="Welche Spalten in der Importkandidaten-Tabelle sichtbar sind. PZN, Artikel "
                      "und Hersteller bleiben immer eingeblendet.",
                 bg=BG, fg=MUTED, font=theme.SMALL, wraplength=820, justify="left").pack(anchor="w", pady=(4, 8))
        spgrid = tk.Frame(cardsp.inner, bg=BG)
        spgrid.pack(fill="x")
        self._spalten_vars = {}
        vis = kandidat_spalten()
        for i, (key, label) in enumerate(IK_OPT_COLUMNS):
            r, c = divmod(i, 3)
            var = tk.IntVar(value=1 if vis.get(key, True) else 0)
            tk.Checkbutton(spgrid, text=label, variable=var, bg=BG, fg=TEXT,
                           activebackground=BG, font=theme.SMALL).grid(
                row=r, column=c, sticky="w", padx=(0, 18), pady=2)
            self._spalten_vars[key] = var

        card2 = theme.Card(wrap)
        card2.pack(fill="x", pady=(0, 12))
        tk.Label(card2.inner, text="Standard-Währung für neue Quellen", bg=BG, fg=TEXT,
                 font=theme.SECTION).pack(anchor="w")
        self._set_vars["standard_waehrung"] = tk.StringVar(value=get_setting("standard_waehrung", "EUR"))
        ttk.Combobox(card2.inner, textvariable=self._set_vars["standard_waehrung"],
                     values=waehrungen(), state="readonly", width=10,
                     style="NMG.TCombobox").pack(anchor="w", pady=(8, 0))

        theme.PillButton(wrap, "Speichern", self._einstellungen_speichern, kind="primary",
                         padx=20, pady=10).pack(anchor="w", pady=(4, 12))

        card3 = theme.Card(wrap)
        card3.pack(fill="x", pady=(0, 24))
        tk.Label(card3.inner, text="Beispieldaten", bg=BG, fg=TEXT, font=theme.SECTION).pack(anchor="w")
        tk.Label(card3.inner,
                 text=("Legt Beispiel-Lieferanten (EU-Ausland), Beschaffungsquellen zu echten "
                       "Artikeln und einige Aufgaben an, damit du die App sofort ausprobieren "
                       "kannst. Vorhandene Daten bleiben erhalten."),
                 bg=BG, fg=MUTED, font=theme.SMALL, wraplength=820, justify="left").pack(anchor="w", pady=(4, 10))
        theme.PillButton(card3.inner, "🎁  Beispieldaten anlegen", self._demo_laden,
                         kind="success", padx=16, pady=9).pack(anchor="w")

    def _demo_laden(self):
        if not messagebox.askyesno(
                "Beispieldaten",
                "Beispiel-Lieferanten, -Quellen und -Aufgaben anlegen?\n"
                "(Vorhandene Daten bleiben erhalten.)", parent=self):
            return
        try:
            res = demo_daten_anlegen()
        except Exception as exc:
            messagebox.showerror("Beispieldaten", _T('Konnte nicht angelegt werden:\n{p0}', p0=exc), parent=self)
            return
        messagebox.showinfo(
            "Beispieldaten",
            _T('Angelegt: {p0} Lieferanten, {p1} Quellen, {p2} Aufgaben.', p0=res['lieferanten'], p1=res['quellen'], p2=res['aufgaben']), parent=self)
        self.show("start")

    def _einstellungen_speichern(self):
        for key, var in self._set_vars.items():
            set_setting(key, var.get().strip())
        if hasattr(self, "_spalten_vars"):
            set_kandidat_spalten({k: bool(v.get()) for k, v in self._spalten_vars.items()})
        _log("Einstellungen gespeichert")
        messagebox.showinfo("Einstellungen", "Gespeichert.", parent=self)


# ── Dialoge ──────────────────────────────────────────────────────────────────
class _BaseDialog(tk.Toplevel):
    def __init__(self, master, title, breite=560, hoehe=520):
        super().__init__(master)
        self.title(title)
        self.configure(bg=SHELL_BG)
        self.transient(master.winfo_toplevel())
        self.resizable(False, False)
        self.geometry(f"{breite}x{hoehe}")
        try:
            self.grab_set()
        except tk.TclError:
            pass

    def _feld(self, parent, label, default="", width=44, multiline=False):
        tk.Label(parent, text=label, bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w", pady=(8, 0))
        if multiline:
            txt = tk.Text(parent, height=3, width=width, font=theme.BODY,
                          highlightbackground=BORDER, highlightthickness=1, relief="flat")
            txt.insert("1.0", default or "")
            txt.pack(fill="x")
            return txt
        v = tk.StringVar(value=str(default or ""))
        ttk.Entry(parent, textvariable=v, width=width).pack(fill="x")
        return v


class AufgabeDialog(_BaseDialog):
    def __init__(self, master, aufgabe_id=None, on_save=None, vorgabe=None):
        super().__init__(master, "Aufgabe", breite=560, hoehe=600)
        self.aufgabe_id = aufgabe_id
        self.on_save = on_save
        vorgabe = vorgabe or {}

        daten = {}
        if aufgabe_id:
            with _con() as con:
                con.row_factory = sqlite3.Row
                row = con.execute("SELECT * FROM tbl_einkauf_aufgaben WHERE id=?", (aufgabe_id,)).fetchone()
            daten = dict(row) if row else {}

        body = tk.Frame(self, bg=SHELL_BG)
        body.pack(fill="both", expand=True, padx=20, pady=16)
        tk.Label(body, text="Aufgabe bearbeiten" if aufgabe_id else "Neue Aufgabe",
                 bg=SHELL_BG, fg=TEXT, font=theme.H2).pack(anchor="w", pady=(0, 6))

        self.v_titel = self._feld(body, "Titel *", daten.get("titel", vorgabe.get("titel", "")))
        self.v_beschr = self._feld(body, "Beschreibung", daten.get("beschreibung", vorgabe.get("beschreibung", "")),
                                   multiline=True)

        zeile = tk.Frame(body, bg=SHELL_BG)
        zeile.pack(fill="x", pady=(8, 0))
        links = tk.Frame(zeile, bg=SHELL_BG)
        links.pack(side="left", fill="x", expand=True, padx=(0, 8))
        tk.Label(links, text="Kategorie", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self.v_kat = tk.StringVar(value=daten.get("kategorie", vorgabe.get("kategorie", "Wiedervorlage")))
        ttk.Combobox(links, textvariable=self.v_kat, values=AUFGABE_KATEGORIEN,
                     state="readonly", style="NMG.TCombobox").pack(fill="x")
        rechts = tk.Frame(zeile, bg=SHELL_BG)
        rechts.pack(side="left", fill="x", expand=True)
        tk.Label(rechts, text="Priorität", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self.v_prio = tk.StringVar(value=PRIO_LABEL.get(daten.get("prioritaet", 2), "Mittel"))
        ttk.Combobox(rechts, textvariable=self.v_prio, values=list(PRIO_LABEL.values()),
                     state="readonly", style="NMG.TCombobox").pack(fill="x")

        zeile2 = tk.Frame(body, bg=SHELL_BG)
        zeile2.pack(fill="x", pady=(8, 0))
        l2 = tk.Frame(zeile2, bg=SHELL_BG)
        l2.pack(side="left", fill="x", expand=True, padx=(0, 8))
        tk.Label(l2, text="Fällig am (TT.MM.JJJJ)", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self.v_faellig = tk.StringVar(value=_datum_de(daten.get("faellig_am", "")) if daten.get("faellig_am") else "")
        ttk.Entry(l2, textvariable=self.v_faellig).pack(fill="x")
        r2 = tk.Frame(zeile2, bg=SHELL_BG)
        r2.pack(side="left", fill="x", expand=True)
        tk.Label(r2, text="Lieferant", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self._lief_map = {"— keiner —": None}
        self._lief_map.update({v: k for k, v in lieferant_namen().items()})
        akt_lief = "— keiner —"
        vorgabe_lief = daten.get("lieferant_id") or vorgabe.get("lieferant_id")
        if vorgabe_lief:
            akt_lief = lieferant_namen().get(vorgabe_lief, "— keiner —")
        self.v_lief = tk.StringVar(value=akt_lief)
        ttk.Combobox(r2, textvariable=self.v_lief, values=list(self._lief_map.keys()),
                     state="readonly", style="NMG.TCombobox").pack(fill="x")

        # Schnell-Buttons fuer Faelligkeit
        quick = tk.Frame(body, bg=SHELL_BG)
        quick.pack(fill="x", pady=(8, 0))
        tk.Label(quick, text="Schnell:", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(side="left")
        for label, tage in (("Heute", 0), ("Morgen", 1), ("+3 Tage", 3), ("+1 Woche", 7)):
            theme.PillButton(quick, label, lambda t=tage: self.v_faellig.set(
                _datum_de((date.today() + timedelta(days=t)).isoformat())),
                kind="ghost", font_size=9, padx=8, pady=4).pack(side="left", padx=(6, 0))

        btn = tk.Frame(body, bg=SHELL_BG)
        btn.pack(fill="x", pady=(16, 0))
        theme.PillButton(btn, "Speichern", self._speichern, kind="primary",
                         padx=18, pady=9).pack(side="left")
        theme.PillButton(btn, "Abbrechen", self.destroy, kind="neutral",
                         padx=14, pady=9).pack(side="left", padx=8)

    def _speichern(self):
        titel = self.v_titel.get().strip()
        if not titel:
            messagebox.showerror("Aufgabe", "Bitte einen Titel eingeben.", parent=self)
            return
        prio = {v: k for k, v in PRIO_LABEL.items()}.get(self.v_prio.get(), 2)
        faellig = _parse_datum(self.v_faellig.get())
        lief_id = self._lief_map.get(self.v_lief.get())
        beschr = self.v_beschr.get("1.0", "end").strip()
        with _con() as con:
            if self.aufgabe_id:
                con.execute(
                    "UPDATE tbl_einkauf_aufgaben SET titel=?, beschreibung=?, kategorie=?, "
                    "prioritaet=?, faellig_am=?, lieferant_id=? WHERE id=?",
                    (titel, beschr, self.v_kat.get(), prio, faellig, lief_id, self.aufgabe_id))
            else:
                con.execute(
                    "INSERT INTO tbl_einkauf_aufgaben(titel, beschreibung, kategorie, prioritaet, "
                    "faellig_am, lieferant_id, status, bearbeiter) VALUES(?,?,?,?,?,?,'offen',?)",
                    (titel, beschr, self.v_kat.get(), prio, faellig, lief_id, bearbeiter()))
            con.commit()
        _log("Aufgabe gespeichert", titel)
        if self.on_save:
            self.on_save()
        self.destroy()


class LieferantDialog(_BaseDialog):
    def __init__(self, master, lieferant_id=None, on_save=None):
        super().__init__(master, "Lieferant", breite=560, hoehe=780)
        self.lieferant_id = lieferant_id
        self.on_save = on_save
        daten = {}
        if lieferant_id:
            with _con() as con:
                con.row_factory = sqlite3.Row
                row = con.execute("SELECT * FROM tbl_einkauf_lieferanten WHERE id=?", (lieferant_id,)).fetchone()
            daten = dict(row) if row else {}

        body = tk.Frame(self, bg=SHELL_BG)
        body.pack(fill="both", expand=True, padx=20, pady=16)
        tk.Label(body, text="Lieferant bearbeiten" if lieferant_id else "Neuer Lieferant",
                 bg=SHELL_BG, fg=TEXT, font=theme.H2).pack(anchor="w", pady=(0, 6))

        self.v_name = self._feld(body, "Name *", daten.get("name", ""))
        zeile = tk.Frame(body, bg=SHELL_BG)
        zeile.pack(fill="x", pady=(8, 0))
        l = tk.Frame(zeile, bg=SHELL_BG); l.pack(side="left", fill="x", expand=True, padx=(0, 8))
        tk.Label(l, text="Land", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self.v_land = tk.StringVar(value=daten.get("land", ""))
        ttk.Entry(l, textvariable=self.v_land).pack(fill="x")
        r = tk.Frame(zeile, bg=SHELL_BG); r.pack(side="left", fill="x", expand=True)
        tk.Label(r, text="Währung", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self.v_waehrung = tk.StringVar(value=daten.get("waehrung", get_setting("standard_waehrung", "EUR")))
        ttk.Combobox(r, textvariable=self.v_waehrung, values=waehrungen(),
                     state="readonly", style="NMG.TCombobox").pack(fill="x")

        self.v_ap = self._feld(body, "Ansprechpartner", daten.get("ansprechpartner", ""))
        zeile2 = tk.Frame(body, bg=SHELL_BG)
        zeile2.pack(fill="x", pady=(8, 0))
        l2 = tk.Frame(zeile2, bg=SHELL_BG); l2.pack(side="left", fill="x", expand=True, padx=(0, 8))
        tk.Label(l2, text="E-Mail", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self.v_email = tk.StringVar(value=daten.get("email", ""))
        ttk.Entry(l2, textvariable=self.v_email).pack(fill="x")
        r2 = tk.Frame(zeile2, bg=SHELL_BG); r2.pack(side="left", fill="x", expand=True)
        tk.Label(r2, text="Telefon", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self.v_tel = tk.StringVar(value=daten.get("telefon", ""))
        ttk.Entry(r2, textvariable=self.v_tel).pack(fill="x")

        zeile3 = tk.Frame(body, bg=SHELL_BG)
        zeile3.pack(fill="x", pady=(8, 0))
        a = tk.Frame(zeile3, bg=SHELL_BG); a.pack(side="left", fill="x", expand=True, padx=(0, 8))
        tk.Label(a, text="Lieferzeit (Tage)", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self.v_lz = tk.StringVar(value=daten.get("lieferzeit_tage") or "")
        ttk.Entry(a, textvariable=self.v_lz).pack(fill="x")
        b = tk.Frame(zeile3, bg=SHELL_BG); b.pack(side="left", fill="x", expand=True, padx=(0, 8))
        tk.Label(b, text="Mind.-Bestellwert (€)", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self.v_mbw = tk.StringVar(value=daten.get("mindestbestellwert") or "")
        ttk.Entry(b, textvariable=self.v_mbw).pack(fill="x")
        c = tk.Frame(zeile3, bg=SHELL_BG); c.pack(side="left", fill="x", expand=True)
        tk.Label(c, text="Zahlungsziel (Tage)", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self.v_zz = tk.StringVar(value=daten.get("zahlungsziel_tage") or "")
        ttk.Entry(c, textvariable=self.v_zz).pack(fill="x")

        self.v_notiz = self._feld(body, "Notizen", daten.get("notizen", ""), multiline=True)

        chk = tk.Frame(body, bg=SHELL_BG)
        chk.pack(fill="x", pady=(8, 0))
        self.v_gdp = tk.IntVar(value=daten.get("gdp_zertifiziert", 0) or 0)
        tk.Checkbutton(chk, text="GDP-zertifiziert", variable=self.v_gdp, bg=SHELL_BG,
                       fg=TEXT, activebackground=SHELL_BG, font=theme.BODY).pack(side="left")
        self.v_aktiv = tk.IntVar(value=daten.get("aktiv", 1) if lieferant_id else 1)
        tk.Checkbutton(chk, text="Aktiv", variable=self.v_aktiv, bg=SHELL_BG,
                       fg=TEXT, activebackground=SHELL_BG, font=theme.BODY).pack(side="left", padx=16)

        # Hersteller-Zuordnung: welche Marken dieser Lieferant beschaffen kann
        # (Lieferant ≠ Hersteller, aber verknüpfbar).
        tk.Label(body, text="Beschaffbare Hersteller (welche Marken dieser Lieferant liefern kann)",
                 bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w", pady=(10, 0))
        hz = tk.Frame(body, bg=SHELL_BG)
        hz.pack(fill="x")
        self._link_herst = lieferant_hersteller(lieferant_id) if lieferant_id else []
        self._link_lb = tk.Listbox(hz, height=4, font=theme.SMALL, highlightthickness=1,
                                   highlightbackground=BORDER, relief="flat", activestyle="none")
        self._link_lb.pack(side="left", fill="x", expand=True)
        for h in self._link_herst:
            self._link_lb.insert("end", h)
        hb = tk.Frame(hz, bg=SHELL_BG)
        hb.pack(side="left", padx=(6, 0), fill="y")
        self._link_combo = ttk.Combobox(hb, values=kandidat_hersteller_namen(), state="normal",
                                        width=16, style="NMG.TCombobox")
        self._link_combo.pack(anchor="w")
        hbtn = tk.Frame(hb, bg=SHELL_BG)
        hbtn.pack(anchor="w", pady=(4, 0))
        theme.PillButton(hbtn, "➕ Hinzufügen", self._link_add, kind="neutral",
                         font_size=9, padx=8, pady=4).pack(side="left")
        theme.PillButton(hbtn, "➖ Entfernen", self._link_remove, kind="ghost",
                         font_size=9, padx=8, pady=4).pack(side="left", padx=(4, 0))

        btn = tk.Frame(body, bg=SHELL_BG)
        btn.pack(fill="x", pady=(16, 0))
        theme.PillButton(btn, "Speichern", self._speichern, kind="primary",
                         padx=18, pady=9).pack(side="left")
        theme.PillButton(btn, "Abbrechen", self.destroy, kind="neutral",
                         padx=14, pady=9).pack(side="left", padx=8)

    def _link_add(self):
        h = (self._link_combo.get() or "").strip()
        if h and h not in self._link_herst:
            self._link_herst.append(h)
            self._link_lb.insert("end", h)
        self._link_combo.set("")

    def _link_remove(self):
        sel = self._link_lb.curselection()
        if not sel:
            return
        idx = sel[0]
        h = self._link_lb.get(idx)
        self._link_lb.delete(idx)
        if h in self._link_herst:
            self._link_herst.remove(h)

    def _speichern(self):
        name = self.v_name.get().strip()
        if not name:
            messagebox.showerror("Lieferant", "Bitte einen Namen eingeben.", parent=self)
            return
        werte = (name, self.v_land.get().strip(), self.v_waehrung.get(),
                 self.v_ap.get().strip(), self.v_email.get().strip(), self.v_tel.get().strip(),
                 int(_parse_num(self.v_lz.get())) or None,
                 _parse_num(self.v_mbw.get()) or None,
                 int(_parse_num(self.v_zz.get())) or None,
                 self.v_gdp.get(), self.v_aktiv.get(),
                 self.v_notiz.get("1.0", "end").strip())
        with _con() as con:
            if self.lieferant_id:
                con.execute(
                    "UPDATE tbl_einkauf_lieferanten SET name=?, land=?, waehrung=?, ansprechpartner=?, "
                    "email=?, telefon=?, lieferzeit_tage=?, mindestbestellwert=?, zahlungsziel_tage=?, "
                    "gdp_zertifiziert=?, aktiv=?, notizen=?, geaendert_am=? WHERE id=?",
                    werte + (_now(), self.lieferant_id))
                lid = self.lieferant_id
            else:
                cur = con.execute(
                    "INSERT INTO tbl_einkauf_lieferanten(name, land, waehrung, ansprechpartner, email, "
                    "telefon, lieferzeit_tage, mindestbestellwert, zahlungsziel_tage, gdp_zertifiziert, "
                    "aktiv, notizen) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)", werte)
                lid = cur.lastrowid
            con.commit()
        set_lieferant_hersteller(lid, getattr(self, "_link_herst", []))
        _log("Lieferant gespeichert", name)
        if self.on_save:
            self.on_save()
        self.destroy()


class QuelleDialog(_BaseDialog):
    def __init__(self, master, quelle_id=None, on_save=None):
        super().__init__(master, "Beschaffungsquelle", breite=560, hoehe=560)
        self.quelle_id = quelle_id
        self.on_save = on_save
        daten = {}
        if quelle_id:
            with _con() as con:
                con.row_factory = sqlite3.Row
                row = con.execute("SELECT * FROM tbl_einkauf_quellen WHERE id=?", (quelle_id,)).fetchone()
            daten = dict(row) if row else {}

        body = tk.Frame(self, bg=SHELL_BG)
        body.pack(fill="both", expand=True, padx=20, pady=16)
        tk.Label(body, text="Quelle bearbeiten" if quelle_id else "Neue Beschaffungsquelle",
                 bg=SHELL_BG, fg=TEXT, font=theme.H2).pack(anchor="w", pady=(0, 6))

        # PZN + Laden
        tk.Label(body, text="PZN *", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w", pady=(8, 0))
        pzn_row = tk.Frame(body, bg=SHELL_BG)
        pzn_row.pack(fill="x")
        self.v_pzn = tk.StringVar(value=daten.get("pzn", ""))
        ttk.Entry(pzn_row, textvariable=self.v_pzn, width=16).pack(side="left")
        theme.PillButton(pzn_row, "Artikel laden", self._pzn_laden, kind="neutral",
                         font_size=9, padx=8, pady=4).pack(side="left", padx=(6, 0))
        self.v_artikel = self._feld(body, "Artikelname", daten.get("artikelname", ""))

        tk.Label(body, text="Lieferant *", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w", pady=(8, 0))
        self._lief_map = {v: k for k, v in lieferant_namen().items()}
        akt = ""
        if daten.get("lieferant_id"):
            akt = lieferant_namen().get(daten["lieferant_id"], "")
        self.v_lief = tk.StringVar(value=akt)
        ttk.Combobox(body, textvariable=self.v_lief, values=list(self._lief_map.keys()),
                     state="readonly", style="NMG.TCombobox").pack(fill="x")

        zeile = tk.Frame(body, bg=SHELL_BG)
        zeile.pack(fill="x", pady=(8, 0))
        a = tk.Frame(zeile, bg=SHELL_BG); a.pack(side="left", fill="x", expand=True, padx=(0, 8))
        tk.Label(a, text="Einkaufspreis (Fremdwährung)", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self.v_ek = tk.StringVar(value=daten.get("ek_fremd") or "")
        ttk.Entry(a, textvariable=self.v_ek).pack(fill="x")
        b = tk.Frame(zeile, bg=SHELL_BG); b.pack(side="left", fill="x", expand=True)
        tk.Label(b, text="Währung", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self.v_waehrung = tk.StringVar(value=daten.get("waehrung", get_setting("standard_waehrung", "EUR")))
        ttk.Combobox(b, textvariable=self.v_waehrung, values=waehrungen(),
                     state="readonly", style="NMG.TCombobox").pack(fill="x")
        self.v_waehrung.trace_add("write", lambda *_: self._lief_waehrung_sync())

        zeile2 = tk.Frame(body, bg=SHELL_BG)
        zeile2.pack(fill="x", pady=(8, 0))
        c = tk.Frame(zeile2, bg=SHELL_BG); c.pack(side="left", fill="x", expand=True, padx=(0, 8))
        tk.Label(c, text="Mindestabnahme (Stück)", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self.v_min = tk.StringVar(value=daten.get("mindestabnahme") or "")
        ttk.Entry(c, textvariable=self.v_min).pack(fill="x")
        d = tk.Frame(zeile2, bg=SHELL_BG); d.pack(side="left", fill="x", expand=True)
        tk.Label(d, text="Lieferzeit (Tage)", bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self.v_lz = tk.StringVar(value=daten.get("lieferzeit_tage") or "")
        ttk.Entry(d, textvariable=self.v_lz).pack(fill="x")

        self.v_notiz = self._feld(body, "Notiz", daten.get("notiz", ""))
        self.v_aktiv = tk.IntVar(value=daten.get("aktiv", 1) if quelle_id else 1)
        tk.Checkbutton(body, text="Aktiv", variable=self.v_aktiv, bg=SHELL_BG,
                       fg=TEXT, activebackground=SHELL_BG, font=theme.BODY).pack(anchor="w", pady=(8, 0))

        # Live-EK-EUR-Hinweis
        self._ek_hinweis = tk.Label(body, text="", bg=SHELL_BG, fg=ACCENT, font=theme.SMALL)
        self._ek_hinweis.pack(anchor="w", pady=(6, 0))
        self.v_ek.trace_add("write", lambda *_: self._update_ek_hinweis())
        self._update_ek_hinweis()

        btn = tk.Frame(body, bg=SHELL_BG)
        btn.pack(fill="x", pady=(14, 0))
        theme.PillButton(btn, "Speichern", self._speichern, kind="primary",
                         padx=18, pady=9).pack(side="left")
        theme.PillButton(btn, "Abbrechen", self.destroy, kind="neutral",
                         padx=14, pady=9).pack(side="left", padx=8)

    def _lief_waehrung_sync(self):
        self._update_ek_hinweis()

    def _update_ek_hinweis(self):
        ek = _parse_num(self.v_ek.get())
        w = self.v_waehrung.get()
        eur = ek * kurs_eur(w)
        self._ek_hinweis.config(text=_T('≈ {p0}  (Kurs 1 {p1} = {p2:.4f} EUR)', p0=_eur(eur), p1=w, p2=kurs_eur(w)))

    def _pzn_laden(self):
        art = artikel_by_pzn(self.v_pzn.get().strip())
        if art and isinstance(self.v_artikel, tk.StringVar):
            self.v_artikel.set(art.get("artikelname") or "")
        elif not art:
            messagebox.showinfo("Artikel", "Keine NMG-Stammdaten zu dieser PZN.", parent=self)

    def _speichern(self):
        pzn = self.v_pzn.get().strip()
        lief_id = self._lief_map.get(self.v_lief.get())
        if not pzn:
            messagebox.showerror("Quelle", "Bitte eine PZN eingeben.", parent=self)
            return
        if not lief_id:
            messagebox.showerror("Quelle", "Bitte einen Lieferanten wählen (ggf. zuerst anlegen).", parent=self)
            return
        werte = (lief_id, pzn, self.v_artikel.get().strip(), _parse_num(self.v_ek.get()),
                 self.v_waehrung.get(), int(_parse_num(self.v_min.get())) or None,
                 int(_parse_num(self.v_lz.get())) or None, self.v_notiz.get().strip(),
                 self.v_aktiv.get())
        with _con() as con:
            if self.quelle_id:
                con.execute(
                    "UPDATE tbl_einkauf_quellen SET lieferant_id=?, pzn=?, artikelname=?, ek_fremd=?, "
                    "waehrung=?, mindestabnahme=?, lieferzeit_tage=?, notiz=?, aktiv=? WHERE id=?",
                    werte + (self.quelle_id,))
            else:
                con.execute(
                    "INSERT INTO tbl_einkauf_quellen(lieferant_id, pzn, artikelname, ek_fremd, waehrung, "
                    "mindestabnahme, lieferzeit_tage, notiz, aktiv, gueltig_ab) VALUES(?,?,?,?,?,?,?,?,?,?)",
                    werte + (_heute(),))
            con.commit()
        _log("Quelle gespeichert", f"PZN {pzn}")
        if self.on_save:
            self.on_save()
        self.destroy()


class KandidatDialog(_BaseDialog):
    """Detailfenster zu einem Importkandidaten: alle Artikel- und Herstellerdaten
    plus Direktaktionen (Marge rechnen, Hersteller kontaktieren, wichtig markieren)."""

    def __init__(self, panel, d):
        super().__init__(panel, f"Artikel {d['pzn']}", breite=660, hoehe=720)
        self.resizable(False, True)
        self.panel = panel
        self.d = d
        self._status_daten = kandidat_status_map().get(d["pzn"], {})

        # Feste Aktionsleiste unten, scrollbarer Inhalt darüber.
        btn = tk.Frame(self, bg=SHELL_BG)
        btn.pack(side="bottom", fill="x", padx=20, pady=10)
        outer = tk.Frame(self, bg=SHELL_BG)
        outer.pack(side="top", fill="both", expand=True)
        canvas = tk.Canvas(outer, bg=SHELL_BG, highlightthickness=0)
        sb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        body = tk.Frame(canvas, bg=SHELL_BG)
        body.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=body, anchor="nw", width=612)
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left", fill="both", expand=True, padx=(20, 0), pady=(14, 0))
        sb.pack(side="right", fill="y")
        canvas.bind("<Enter>", lambda e: canvas.bind_all(
            "<MouseWheel>", lambda ev: canvas.yview_scroll(int(-ev.delta / 120), "units")))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        tk.Label(body, text=d["artikel"] or "—", bg=SHELL_BG, fg=TEXT, font=theme.H2,
                 wraplength=580, justify="left").pack(anchor="w")
        tk.Label(body, text=_T('PZN {p0}', p0=d['pzn']), bg=SHELL_BG, fg=MUTED, font=theme.SMALL).pack(anchor="w", pady=(0, 8))

        # Bewertung (Potenzial & §129-Machbarkeit)
        bew = theme.Card(body)
        bew.pack(fill="x", pady=(0, 10))
        tk.Label(bew.inner, text="💶  Bewertung (geschätzt)", bg=BG, fg=TEXT,
                 font=theme.SECTION).pack(anchor="w", pady=(0, 4))
        ampel_txt = {"gruen": "✓ machbar", "gelb": "~ knapp", "rot": "✗ keine Marge",
                     "—": "kein Preis hinterlegt"}.get(d["ampel"], "—")
        ampel_farbe = {"gruen": OK_GREEN, "gelb": theme.WARNING, "rot": theme.DANGER}.get(d["ampel"], MUTED)
        ar = tk.Frame(bew.inner, bg=BG)
        ar.pack(fill="x", pady=2)
        tk.Label(ar, text="§129-Machbarkeit", bg=BG, fg=MUTED, font=theme.SMALL, width=24, anchor="w").pack(side="left")
        tk.Label(ar, text=ampel_txt, bg=BG, fg=ampel_farbe, font=theme.BODY_BOLD, anchor="w").pack(side="left")
        if d["potenzial_jahr"] is not None:
            self._row(bew.inner, "Potenzial / Jahr (Deckungsbeitrag)", _eur(d["potenzial_jahr"]))
            self._row(bew.inner, "geschätzte Marge / Stück",
                      f"{_eur(d['marge_stueck'])}  ({d['marge_proz']:.0f} %)")
            self._row(bew.inner, "max. zul. Import-AVP (§129)", _eur(d["max_import_avp"]))
            self._row(bew.inner, "Umsatz-Potenzial / Jahr", _eur(d["umsatz_jahr"]))
            tk.Label(bew.inner,
                     text=_T('Annahme EU-EK {p0} % vom DE-Preis · in den Einstellungen anpassbar.', p0=get_setting('kandidat_eu_ek_proz')),
                     bg=BG, fg=MUTED, font=theme.SMALL, wraplength=560, justify="left").pack(anchor="w", pady=(4, 0))
        else:
            tk.Label(bew.inner, text="Ohne hinterlegten Preis ist keine Margen-/Potenzialschätzung möglich.",
                     bg=BG, fg=MUTED, font=theme.SMALL, wraplength=560, justify="left").pack(anchor="w", pady=(2, 0))

        # Artikel-Daten
        art = theme.Card(body)
        art.pack(fill="x", pady=(0, 10))
        tk.Label(art.inner, text="📦  Artikel", bg=BG, fg=TEXT, font=theme.SECTION).pack(anchor="w", pady=(0, 4))
        preis_txt = (f"{_eur(d['preis'])}  ({d['preis_quelle']})" if d["preis"] is not None else "—")
        self._row(art.inner, "Darreichung / Packung",
                  " · ".join(x for x in (d.get("df"), d.get("pck")) if x) or "—")
        self._row(art.inner, "Wirkstoff", d["wirkstoff"] or "—")
        self._row(art.inner, "Absatz (6 Monate)", f"{d['absatz']:,.0f}".replace(",", "."))
        self._row(art.inner, "Apothekenbreite", f"{d['apotheken']} Apotheke(n)")
        trend_txt = d.get("trend_label", "n/a")
        if d.get("trend_label") not in (None, "n/a") and d.get("trend_proz"):
            trend_txt = f"{d['trend_sym']} {d['trend_label']} ({d['trend_proz']:+.0f} %)"
        self._row(art.inner, "Nachfrage-Trend", trend_txt)
        self._row(art.inner, "Umsatz (6 Monate)", _eur(d["umsatz"]) if d["umsatz"] else "—")
        self._row(art.inner, "APU / Preis", preis_txt)
        self._row(art.inner, "APU (gepflegt)", _eur(d["apu"]) if d["apu"] is not None else "—")
        self._row(art.inner, "EK (geschätzt)", _eur(d["taxe_ek"]) if d["taxe_ek"] is not None else "—")

        # Hersteller-Daten + Wettbewerb
        her = theme.Card(body)
        her.pack(fill="x", pady=(0, 10))
        kopf = tk.Frame(her.inner, bg=BG)
        kopf.pack(fill="x")
        tk.Label(kopf, text="🏭  Hersteller & Wettbewerb", bg=BG, fg=TEXT, font=theme.SECTION).pack(side="left")
        self._stern_lbl = tk.Label(kopf, text="", bg=BG, fg=ACCENT, font=theme.BODY_BOLD)
        self._stern_lbl.pack(side="left", padx=(8, 0))
        self._row(her.inner, "Hersteller", d["hersteller"])
        tr = tk.Frame(her.inner, bg=BG)
        tr.pack(fill="x", pady=2)
        tk.Label(tr, text="Typ", bg=BG, fg=MUTED, font=theme.SMALL, width=24, anchor="w").pack(side="left")
        self._typ_lbl = tk.Label(tr, text="", bg=BG, fg=TEXT, font=theme.BODY_BOLD, anchor="w")
        self._typ_lbl.pack(side="left")
        konk = d.get("konkurrenz_namen") or []
        konk_txt = (f"{len(konk)} aktiv: " + ", ".join(konk)) if konk else (
            "noch kein Reimporteur – unbesetzt ✓" if d["wirkstoff"] else "kein Wirkstoff hinterlegt")
        self._row(her.inner, "Reimporteure (gleicher Wirkstoff)", konk_txt)
        liefs = lieferanten_fuer_hersteller(d["hersteller"])
        lief_txt = ", ".join((("★ " if bev else "") + nm) for _, nm, bev in liefs) if liefs else \
            "— noch kein Lieferant zugeordnet (im Lieferanten-Dialog hinterlegbar)"
        self._row(her.inner, "Eure Lieferanten dafür", lief_txt)
        self._row(her.inner, "Markt-Gesamtabsatz (Kandidaten)",
                  f"{d['herst_absatz']:,.0f}".replace(",", "."))
        self._row(her.inner, "Kandidaten-Artikel", f"{d['herst_artikel']}")
        andere = sorted((c for c in getattr(panel, "_ik_all", [])
                         if c["hersteller"] == d["hersteller"] and c["pzn"] != d["pzn"]),
                        key=lambda c: c["absatz"], reverse=True)[:5]
        if andere:
            tk.Label(her.inner, text="Weitere Kandidaten dieses Herstellers:", bg=BG, fg=MUTED,
                     font=theme.SMALL).pack(anchor="w", pady=(6, 0))
            for c in andere:
                tk.Label(her.inner, text=f"• {c['artikel'][:46]}  ({c['absatz']:,.0f})".replace(",", "."),
                         bg=BG, fg=TEXT, font=theme.SMALL, anchor="w").pack(anchor="w")
        self._stern_aktualisieren()

        # Mengen-/Break-even-Rechner
        rech = theme.Card(body)
        rech.pack(fill="x", pady=(0, 10))
        tk.Label(rech.inner, text="🧮  Mengen- & Break-even-Rechner", bg=BG, fg=TEXT,
                 font=theme.SECTION).pack(anchor="w", pady=(0, 2))
        tk.Label(rech.inner, text="Echten EU-Einkaufspreis eingeben → Marge/Stück, Break-even-Menge "
                                  "und Deckungsbeitrag bei der Mindestabnahme.",
                 bg=BG, fg=MUTED, font=theme.SMALL, wraplength=560, justify="left").pack(anchor="w", pady=(0, 6))
        rr = tk.Frame(rech.inner, bg=BG)
        rr.pack(fill="x")
        ek_f = tk.Frame(rr, bg=BG); ek_f.pack(side="left", padx=(0, 8))
        tk.Label(ek_f, text="EU-EK (Fremdwährung)", bg=BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self._be_ek = tk.StringVar()
        ttk.Entry(ek_f, textvariable=self._be_ek, width=12).pack()
        w_f = tk.Frame(rr, bg=BG); w_f.pack(side="left", padx=(0, 8))
        tk.Label(w_f, text="Währung", bg=BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self._be_waehrung = tk.StringVar(value=get_setting("standard_waehrung", "EUR"))
        ttk.Combobox(w_f, textvariable=self._be_waehrung, values=waehrungen(), state="readonly",
                     width=7, style="NMG.TCombobox").pack()
        m_f = tk.Frame(rr, bg=BG); m_f.pack(side="left", padx=(0, 8))
        tk.Label(m_f, text="Mindestabnahme", bg=BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self._be_moq = tk.StringVar()
        ttk.Entry(m_f, textvariable=self._be_moq, width=10).pack()
        fx_f = tk.Frame(rr, bg=BG); fx_f.pack(side="left")
        tk.Label(fx_f, text="Einführungskosten €", bg=BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self._be_fix = tk.StringVar(value=get_setting("kandidat_fixkosten"))
        ttk.Entry(fx_f, textvariable=self._be_fix, width=10).pack()
        theme.PillButton(rech.inner, "Rechnen", self._be_rechnen, kind="primary",
                         font_size=10, padx=12, pady=6).pack(anchor="w", pady=(8, 4))
        self._be_result = tk.Label(rech.inner, text="", bg=BG, fg=TEXT, font=theme.BODY,
                                   wraplength=560, justify="left")
        self._be_result.pack(anchor="w")

        # Pipeline / Status
        pipe = theme.Card(body)
        pipe.pack(fill="x", pady=(0, 10))
        tk.Label(pipe.inner, text="📋  Pipeline", bg=BG, fg=TEXT, font=theme.SECTION).pack(anchor="w", pady=(0, 4))
        prow = tk.Frame(pipe.inner, bg=BG)
        prow.pack(fill="x", pady=(2, 0))
        sp = tk.Frame(prow, bg=BG); sp.pack(side="left", fill="x", expand=True, padx=(0, 8))
        tk.Label(sp, text="Status", bg=BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self._st_status = tk.StringVar(value=self._status_daten.get("status", "Neu"))
        ttk.Combobox(sp, textvariable=self._st_status, values=KANDIDAT_STATUS,
                     state="readonly", style="NMG.TCombobox").pack(fill="x")
        wv = tk.Frame(prow, bg=BG); wv.pack(side="left", fill="x", expand=True)
        tk.Label(wv, text="Wiedervorlage (TT.MM.JJJJ)", bg=BG, fg=MUTED, font=theme.SMALL).pack(anchor="w")
        self._st_wv = tk.StringVar(value=_datum_de(self._status_daten.get("wiedervorlage", ""))
                                   if self._status_daten.get("wiedervorlage") else "")
        ttk.Entry(wv, textvariable=self._st_wv).pack(fill="x")
        tk.Label(pipe.inner, text="Notiz", bg=BG, fg=MUTED, font=theme.SMALL).pack(anchor="w", pady=(8, 0))
        self._st_notiz = tk.StringVar(value=self._status_daten.get("notiz", ""))
        ttk.Entry(pipe.inner, textvariable=self._st_notiz).pack(fill="x")
        theme.PillButton(pipe.inner, "Status speichern", self._status_speichern, kind="primary",
                         font_size=10, padx=12, pady=6).pack(anchor="w", pady=(8, 0))

        # Aktionen
        self._merk_btn = theme.PillButton(btn, "📌 Merken", self._toggle_beobachtet, kind="accent",
                                          padx=14, pady=8)
        self._merk_btn.pack(side="left")
        theme.PillButton(btn, "📋 Kontaktieren", self._kontakt, kind="primary",
                         padx=14, pady=8).pack(side="left", padx=6)
        theme.PillButton(btn, "📐 Marge", self._marge, kind="neutral",
                         padx=14, pady=8).pack(side="left")
        self._stern_btn = theme.PillButton(btn, "★ bevorzugt", self._toggle_wichtig, kind="neutral",
                                           padx=14, pady=8)
        self._stern_btn.pack(side="left", padx=6)
        theme.PillButton(btn, "↔ Typ", self._toggle_reimport, kind="neutral",
                         padx=14, pady=8).pack(side="left")
        theme.PillButton(btn, "Schließen", self.destroy, kind="ghost",
                         padx=14, pady=8).pack(side="right")

    def _be_rechnen(self):
        d = self.d
        if not d["preis"] or d["preis"] <= 0:
            self._be_result.config(text="Für diesen Artikel ist kein Preis hinterlegt – "
                                        "keine Margenschätzung möglich.", fg=MUTED)
            return
        ek_eur = _parse_num(self._be_ek.get()) * kurs_eur(self._be_waehrung.get())
        if ek_eur <= 0:
            self._be_result.config(text="Bitte einen EU-Einkaufspreis eingeben.", fg=theme.DANGER)
            return
        max_avp = max(0.0, d["preis"] - erforderlicher_abstand(d["preis"])[0])
        marge = max_avp - ek_eur
        moq = int(_parse_num(self._be_moq.get()))
        fix = _parse_num(self._be_fix.get())
        if marge <= 0:
            self._be_result.config(
                text=_T('EK {p0}/Stk · max. Import-AVP {p1} → keine positive Marge bei diesem Einkaufspreis.', p0=_eur(ek_eur), p1=_eur(max_avp)), fg=theme.DANGER)
            return
        be_menge = math.ceil(fix / marge) if fix > 0 else 0
        db_moq = marge * moq - fix
        teile = [f"EK {_eur(ek_eur)}/Stk", f"Marge {_eur(marge)}/Stk ({marge / max_avp * 100:.0f} %)"]
        if fix > 0:
            teile.append(f"Break-even ab {be_menge:,} Stk".replace(",", "."))
        if moq > 0:
            moq_str = f"{moq:,}".replace(",", ".")
            teile.append(f"DB bei {moq_str} Stk: {_eur(db_moq)}")
        self._be_result.config(text="   ·   ".join(teile),
                               fg=(OK_GREEN if (moq <= 0 or db_moq > 0) else theme.WARNING))

    def _status_speichern(self):
        wv_iso = _parse_datum(self._st_wv.get())
        set_kandidat_status(self.d["pzn"], self._st_status.get(),
                            self._st_notiz.get().strip(), wv_iso)
        self.d["status"] = self._st_status.get()
        # In-Memory-Liste + Tabelle nachziehen.
        for c in self.panel._ik_all:
            if c["pzn"] == self.d["pzn"]:
                c["status"] = self._st_status.get()
        self.panel._importkand_reload()
        messagebox.showinfo("Pipeline", "Status gespeichert.", parent=self)

    def _row(self, parent, label, value):
        r = tk.Frame(parent, bg=BG)
        r.pack(fill="x", pady=2)
        tk.Label(r, text=label, bg=BG, fg=MUTED, font=theme.SMALL, width=24, anchor="w").pack(side="left")
        tk.Label(r, text=value, bg=BG, fg=TEXT, font=theme.BODY, anchor="w",
                 wraplength=360, justify="left").pack(side="left", fill="x", expand=True)

    def _stern_aktualisieren(self):
        self._stern_lbl.config(text="★ als wichtig markiert" if self.d["wichtig"] else "")
        if hasattr(self, "_stern_btn"):
            self._stern_btn.config(text="☆ als normal" if self.d["wichtig"] else "★ wichtig")
        reimport = self.d["typ"] == "Reimport"
        self._typ_lbl.config(text=("Reimport (bereits Parallelimport)" if reimport else "Original-Hersteller"),
                             fg=(MUTED if reimport else OK_GREEN))
        if hasattr(self, "_merk_btn"):
            self._merk_btn.config(text="📌 Gemerkt ✓" if self.d["beobachtet"] else "📌 Merken")

    def _toggle_wichtig(self):
        self.panel._ik_toggle_wichtig(self.d)  # aktualisiert _ik_all (inkl. self.d) + Tabelle
        self._stern_aktualisieren()

    def _toggle_beobachtet(self):
        self.panel._ik_toggle_beobachtet(self.d)  # setzt Flag + sortiert + lädt Tabelle neu
        self._stern_aktualisieren()

    def _toggle_reimport(self):
        self.panel._ik_toggle_reimporteur(self.d)  # aktualisiert _ik_all (inkl. self.d) + Tabelle
        self._stern_aktualisieren()

    def _marge(self):
        d = self.d
        self.destroy()
        self.panel._ik_marge(d)

    def _kontakt(self):
        d = self.d
        self.destroy()
        self.panel._ik_kontakt(d)


# ── Standalone ───────────────────────────────────────────────────────────────
def run_standalone():
    """Startet die Einkauf-App als eigenständiges Fenster (eigenes Taskleisten-Icon).
    Wird von start_einkauf.py / start.py --einkauf genutzt."""
    if os.name == "nt":
        try:
            import ctypes
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("NMG.Einkauf")
        except Exception:
            pass
    try:
        from .migrations import run_migrations
        run_migrations()
    except Exception:
        pass
    root = tk.Tk()
    root.title("NMG Einkauf")
    root.geometry("1160x760")
    root.minsize(1000, 640)
    # Im Vollbild (maximiert) starten. 'zoomed' = Windows; sonst -zoomed/Bildschirmgroesse.
    try:
        root.state("zoomed")
    except tk.TclError:
        try:
            root.attributes("-zoomed", True)
        except tk.TclError:
            root.geometry(f"{root.winfo_screenwidth()}x{root.winfo_screenheight()}+0+0")
    root.configure(bg=SHELL_BG)
    theme.apply_theme(root)
    theme.apply_widget_defaults(root)
    for ico in ("Einkauf.ico", "GDP.ico", "NMGone.ico"):
        try:
            root.iconbitmap(str(ASSETS_DIR / ico))
            break
        except Exception:
            continue
    EinkaufPanel(root, on_close=root.destroy).pack(fill="both", expand=True)
    root.mainloop()


if __name__ == "__main__":
    run_standalone()
