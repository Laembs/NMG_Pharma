from __future__ import annotations

import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .config import OUTPUT_DIR, DB_PATH

try:
    from .protocol_manager import PROTOCOL_ROOT, ensure_protocol_dirs, log_event
except Exception:  # pragma: no cover - Fallback fuer portable/alte Stände
    PROTOCOL_ROOT = OUTPUT_DIR.parent / "protokolle"

    def ensure_protocol_dirs():
        Path(PROTOCOL_ROOT).mkdir(parents=True, exist_ok=True)

    def log_event(*args, **kwargs):
        return None


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return " ".join(text.split())


def _pzn(value: Any) -> str:
    text = _clean(value)
    if not text:
        return ""
    digits = "".join(ch for ch in text if ch.isdigit())
    if digits:
        return digits.zfill(8)
    return text


def _norm_text(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _rowdict(row) -> dict:
    return dict(row) if row else {}


def _table_exists(con: sqlite3.Connection, table: str) -> bool:
    try:
        return con.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (table,),
        ).fetchone() is not None
    except Exception:
        return False


def _article_name_by_pzn(con: sqlite3.Connection, pzn: str) -> str:
    pzn = _pzn(pzn)
    if not pzn:
        return ""
    lookups = (
        ("tbl_nmg_stamm", "artikelname"),
        ("tbl_artikelstamm", "artikel"),
        ("tbl_artikelstamm", "artikelname"),
        ("tbl_pzn_basisdaten", "artikelname"),
    )
    for table, col in lookups:
        try:
            if not _table_exists(con, table):
                continue
            row = con.execute(f"SELECT {col} FROM {table} WHERE pzn=? LIMIT 1", (pzn,)).fetchone()
            if row and _clean(row[0]):
                return _clean(row[0])
        except Exception:
            continue
    return ""


def _variant_key(item: dict) -> tuple[str, str]:
    return (_pzn(item.get("pzn_nmg")), _norm_text(item.get("freitext_austausch")))


def build_trace_index(con: sqlite3.Connection) -> dict:
    """Lädt einmalig alle Quell-Tabellen in In-Memory-Dicts, damit
    trace_lookup_source pro Zeile nur noch Dict-Lookups macht statt
    18k Austausch-Zeilen erneut zu scannen. Ohne diesen Index dauert
    eine Auswertung über 400 Zeilen ca. 3 Minuten (>90 % in trace_lookup_source).
    """
    index = {
        "austausch_by_pzn": {},
        "lernvorschlaege_by_pzn": {},
        "referenz_h_o_by_pzn": {},
        "nmg_pzns": set(),
        "has_austausch": False,
        "has_lernvorschlaege": False,
        "has_referenz_h_o": False,
    }

    if _table_exists(con, "tbl_austauschdatenbank"):
        index["has_austausch"] = True
        try:
            rows = con.execute(
                """
                SELECT *
                FROM tbl_austauschdatenbank
                WHERE COALESCE(status, 'aktiv') = 'aktiv'
                  AND COALESCE(pzn_alt, '') <> ''
                ORDER BY datetime(COALESCE(aktualisiert_am, erstellt_am, gueltig_ab, '1970-01-01')) DESC, id DESC
                """
            ).fetchall()
        except Exception:
            rows = []
        bucket = index["austausch_by_pzn"]
        for row in rows:
            item = _rowdict(row)
            key = _pzn(item.get("pzn_alt"))
            if not key:
                continue
            bucket.setdefault(key, []).append(item)

    if _table_exists(con, "tbl_lernvorschlaege"):
        index["has_lernvorschlaege"] = True
        try:
            rows = con.execute(
                """
                SELECT *
                FROM tbl_lernvorschlaege
                WHERE status = 'uebernommen'
                  AND COALESCE(pzn_alt, '') <> ''
                ORDER BY datetime(COALESCE(bearbeitet_am, erstellt_am, '1970-01-01')) DESC, id DESC
                """
            ).fetchall()
        except Exception:
            rows = []
        bucket = index["lernvorschlaege_by_pzn"]
        for row in rows:
            item = _rowdict(row)
            key = _pzn(item.get("pzn_alt"))
            if not key:
                continue
            bucket.setdefault(key, []).append(item)

    if _table_exists(con, "tbl_referenz_h_o"):
        index["has_referenz_h_o"] = True
        try:
            rows = con.execute("SELECT * FROM tbl_referenz_h_o WHERE COALESCE(original_pzn,'') <> ''").fetchall()
        except Exception:
            rows = []
        bucket = index["referenz_h_o_by_pzn"]
        for row in rows:
            item = _rowdict(row)
            key = _pzn(item.get("original_pzn"))
            if not key:
                continue
            bucket.setdefault(key, item)

    nmg_set = index["nmg_pzns"]
    for table, col in (("tbl_nmg_stamm", "pzn"), ("nmg_rabatte", "nmg_pzn"), ("tbl_lieferfaehigkeit", "nmg_pzn")):
        if not _table_exists(con, table):
            continue
        try:
            for (val,) in con.execute(f"SELECT DISTINCT {col} FROM {table} WHERE COALESCE({col},'') <> ''"):
                key = _pzn(val)
                if key:
                    nmg_set.add(key)
        except Exception:
            continue

    return index


def trace_lookup_source(con: sqlite3.Connection, original_pzn: str, hit: dict | None = None, index: dict | None = None) -> dict:
    """Ermittelt fuer das Protokoll, aus welcher Datenbank die Neue Auswertung liest.

    Die Funktion veraendert keine Daten. Sie bildet die Prioritaet aus exporter._lookup()
    nach und markiert mehrdeutige aktive Austauschdaten als Konflikt fuer die manuelle Pruefung.
    """
    original_pzn = _pzn(original_pzn)
    hit = hit or {}
    base = {
        "pzn_alt": original_pzn,
        "quelle": "",
        "trefferstatus": "nicht gefunden",
        "pzn_nmg": _pzn(hit.get("nmg_pzn")),
        "austauschbar_gegen": _clean(hit.get("austauschbar_gegen")),
        "im_sortiment": _clean(hit.get("im_sortiment")),
        "grund": "Keine passende Zuordnung in den geprüften Datenbanken gefunden.",
        "hinweis_manuelle_pruefung": "",
        "anzahl_varianten": 0,
    }
    if not original_pzn:
        base["grund"] = "Keine gueltige PZN in der Rohdatei."
        return base

    # 1) Hauptquelle: aktive Austauschdatenbank.
    try:
        has_austausch = index["has_austausch"] if index is not None else _table_exists(con, "tbl_austauschdatenbank")
        if has_austausch:
            if index is not None:
                matches = list(index["austausch_by_pzn"].get(original_pzn, ()))
            else:
                rows = con.execute("""
                    SELECT *
                    FROM tbl_austauschdatenbank
                    WHERE COALESCE(status, 'aktiv') = 'aktiv'
                      AND COALESCE(pzn_alt, '') <> ''
                    ORDER BY datetime(COALESCE(aktualisiert_am, erstellt_am, gueltig_ab, '1970-01-01')) DESC, id DESC
                """).fetchall()
                matches = []
                for candidate in rows:
                    item = _rowdict(candidate)
                    if _pzn(item.get("pzn_alt")) == original_pzn:
                        matches.append(item)
            variants = {_variant_key(item) for item in matches if _variant_key(item) != ("", "")}
            if len(variants) > 1:
                pzns = []
                texts = []
                for pzn_nmg, text in sorted(variants):
                    if pzn_nmg and pzn_nmg not in pzns:
                        pzns.append(pzn_nmg)
                    if text and text not in texts:
                        texts.append(text)
                base.update({
                    "quelle": "tbl_austauschdatenbank",
                    "trefferstatus": "Konflikt - manuelle Prüfung",
                    "pzn_nmg": "; ".join(pzns),
                    "austauschbar_gegen": "; ".join(texts),
                    "grund": "Mehrere aktive Varianten zur selben Alt-PZN gefunden. Die Neue Auswertung übernimmt bewusst keinen automatischen Treffer.",
                    "hinweis_manuelle_pruefung": "Ja - in Schulbank > Manuelle Prüfung anzeigen/entscheiden.",
                    "anzahl_varianten": len(variants),
                })
                return base
            if matches:
                item = matches[0]
                pzn_nmg = _pzn(item.get("pzn_nmg"))
                freitext = _clean(item.get("freitext_austausch"))
                artikel_nmg = _clean(item.get("artikel_nmg"))
                if pzn_nmg and (not freitext or freitext.lower().startswith("pzn nmg:")):
                    freitext = artikel_nmg or _article_name_by_pzn(con, pzn_nmg) or freitext
                base.update({
                    "quelle": "tbl_austauschdatenbank",
                    "trefferstatus": "gefunden",
                    "pzn_nmg": pzn_nmg,
                    "austauschbar_gegen": freitext,
                    "im_sortiment": "X" if pzn_nmg else "X Austausch mögl",
                    "grund": "Aktiver Eintrag aus der Austauschdatenbank verwendet.",
                    "anzahl_varianten": len(variants) or 1,
                })
                return base
    except Exception as exc:
        base.update({
            "quelle": "tbl_austauschdatenbank",
            "trefferstatus": "Prüfung fehlgeschlagen",
            "grund": f"Austauschdatenbank konnte nicht geprüft werden: {exc}",
        })
        return base

    # 2) Sicherheitsnetz: übernommene Lernvorschläge.
    try:
        has_lern = index["has_lernvorschlaege"] if index is not None else _table_exists(con, "tbl_lernvorschlaege")
        if has_lern:
            if index is not None:
                candidates = list(index["lernvorschlaege_by_pzn"].get(original_pzn, ()))
            else:
                rows = con.execute("""
                    SELECT *
                    FROM tbl_lernvorschlaege
                    WHERE status = 'uebernommen'
                      AND COALESCE(pzn_alt, '') <> ''
                    ORDER BY datetime(COALESCE(bearbeitet_am, erstellt_am, '1970-01-01')) DESC, id DESC
                """).fetchall()
                candidates = [_rowdict(r) for r in rows]
            for item in candidates:
                if not isinstance(item, dict):
                    item = _rowdict(item)
                if _pzn(item.get("pzn_alt")) != original_pzn:
                    continue
                pzn_nmg = _pzn(item.get("pzn_nmg"))
                freitext = _clean(item.get("freitext_austausch") or item.get("produkt_neu"))
                if pzn_nmg and (not freitext or freitext.lower().startswith("pzn nmg:")):
                    freitext = _article_name_by_pzn(con, pzn_nmg) or freitext
                base.update({
                    "quelle": "tbl_lernvorschlaege",
                    "trefferstatus": "gefunden",
                    "pzn_nmg": pzn_nmg,
                    "austauschbar_gegen": freitext,
                    "im_sortiment": "X" if pzn_nmg else "X Austausch mögl",
                    "grund": "Übernommener Lernvorschlag als Sicherheitsnetz verwendet.",
                    "anzahl_varianten": 1,
                })
                return base
    except Exception:
        pass

    # 3) Alte geprüfte Referenz.
    try:
        has_ref = index["has_referenz_h_o"] if index is not None else _table_exists(con, "tbl_referenz_h_o")
        if has_ref:
            if index is not None:
                ref = index["referenz_h_o_by_pzn"].get(original_pzn, {})
            else:
                row = con.execute("SELECT * FROM tbl_referenz_h_o WHERE original_pzn=? LIMIT 1", (original_pzn,)).fetchone()
                ref = _rowdict(row)
            if ref:
                base.update({
                    "quelle": "tbl_referenz_h_o",
                    "trefferstatus": "gefunden",
                    "pzn_nmg": _pzn(ref.get("nmg_pzn")),
                    "austauschbar_gegen": _clean(ref.get("austauschbar_gegen")),
                    "im_sortiment": _clean(ref.get("im_sortiment")),
                    "grund": "Alte geprüfte Referenz verwendet, weil keine neuere Austauschdatenbank-Zuordnung gefunden wurde.",
                    "anzahl_varianten": 1,
                })
                return base
    except Exception:
        pass

    # 4) PZN ist selbst NMG / Rabatt- oder Lieferdaten vorhanden.
    try:
        if index is not None:
            has_nmg = original_pzn in index["nmg_pzns"]
        else:
            has_nmg = False
            for table, col in (("tbl_nmg_stamm", "pzn"), ("nmg_rabatte", "nmg_pzn"), ("tbl_lieferfaehigkeit", "nmg_pzn")):
                if not _table_exists(con, table):
                    continue
                row = con.execute(f"SELECT 1 FROM {table} WHERE {col}=? LIMIT 1", (original_pzn,)).fetchone()
                if row:
                    has_nmg = True
                    break
        if has_nmg:
            base.update({
                "quelle": "tbl_nmg_stamm / nmg_rabatte / tbl_lieferfaehigkeit",
                "trefferstatus": "gefunden",
                "pzn_nmg": original_pzn,
                "im_sortiment": "X",
                "grund": "Die abgegebene PZN ist selbst als NMG-/PK-Artikel oder mit NMG-Daten vorhanden.",
                "anzahl_varianten": 1,
            })
            return base
    except Exception:
        pass

    return base


def _safe_filename(value: str) -> str:
    value = re.sub(r"[\\/:*?\"<>|]+", "_", str(value or "").strip())
    value = re.sub(r"\s+", "_", value).strip("_")
    return value[:80] or "Analyse"


def write_abgleichartikel_protocol(
    rows: list[dict],
    analyse_name: str,
    input_file: str | Path = "",
    output_file: str | Path = "",
    auswertung_id: int | None = None,
) -> Path:
    """Schreibt das automatische Abgleichartikel-Protokoll fuer Neue Auswertung."""
    ensure_protocol_dirs()
    out_dir = Path(PROTOCOL_ROOT) / "Neue_Auswertung"
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = out_dir / f"Abgleichartikel_{_safe_filename(analyse_name)}_{stamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Abgleichartikel"

    ws["A1"] = "NMGone - Abgleichartikel bei Neuer Auswertung"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:M1")
    ws["A2"] = f"Analyse: {analyse_name or ''}"
    ws["A3"] = f"Rohdatei: {Path(input_file).name if input_file else ''}"
    ws["A4"] = f"Auswertung-ID: {auswertung_id or ''}"
    ws["A5"] = f"Ausgabedatei: {output_file}"
    ws["A6"] = f"Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"

    headers = [
        "Zeile",
        "PZN alt",
        "Artikel alt",
        "DF",
        "PCK",
        "Hersteller",
        "Absatz 6M",
        "Datenbank / Quelle",
        "Trefferstatus",
        "PZN NMG",
        "Austauschbar gegen",
        "Grund / Entscheidung",
        "Manuelle Prüfung",
    ]
    header_row = 8
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, item in enumerate(rows, start=header_row + 1):
        values = [
            item.get("zeile", ""),
            item.get("pzn_alt", ""),
            item.get("artikel_alt", ""),
            item.get("df", ""),
            item.get("pck", ""),
            item.get("hersteller", ""),
            item.get("absatz_6m", ""),
            item.get("quelle", ""),
            item.get("trefferstatus", ""),
            item.get("pzn_nmg", ""),
            item.get("austauschbar_gegen", ""),
            item.get("grund", ""),
            item.get("hinweis_manuelle_pruefung", ""),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(idx, col, value)

    last_row = max(header_row, header_row + len(rows))
    thin = Side(style="thin", color="B7B7B7")
    for row in ws.iter_rows(min_row=header_row, max_row=last_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [8, 12, 34, 8, 10, 14, 12, 30, 24, 14, 40, 56, 28]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A9"
    ws.auto_filter.ref = f"A{header_row}:M{last_row}"

    ws2 = wb.create_sheet("Zusammenfassung")
    summary = {}
    for item in rows:
        key = item.get("trefferstatus", "") or "unbekannt"
        summary[key] = summary.get(key, 0) + 1
    ws2.append(["Trefferstatus", "Anzahl"])
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    for key, count in sorted(summary.items()):
        ws2.append([key, count])
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 12

    wb.save(out)
    try:
        log_event("neue_auswertung", "Abgleichartikel-Protokoll erstellt", str(out))
    except Exception:
        pass
    return out
