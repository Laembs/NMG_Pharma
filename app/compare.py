from pathlib import Path
from datetime import datetime
import getpass
import sqlite3

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .config import OUTPUT_DIR, DB_PATH


COMPARE_FIELDS = [
    ("im_sortiment", "im Sortiment"),
    ("pzn_nmg", "PZN NMG"),
    ("apu_nmg", "APU NMG"),
    ("nmg_rabatt", "NMG Rabatt"),
    ("lieferbar", "lieferbar"),
    ("bevorratung", "Bevorratung angeraten"),
    ("liefervorschlag", "Liefervor- schlag"),
    ("austausch", "austauschbar gegen"),
    ("rabatt_euro", "NMG Rabatt in Euro"),
    ("rabatt_gesamt", "NMG Rabatt Gesamt nach Absatz"),
    ("umsatz", "Umsatz"),
]

RAW_FIELDS = [
    ("artikel", "Artikelname"),
    ("df", "DF"),
    ("pck", "Pck"),
    ("herst", "Herst"),
    ("ek", "EK"),
    ("absatz", "Abverkäufe 6 Monate"),
]


def _norm_header(v):
    if v is None:
        return ""
    return str(v).strip().lower().replace("€", "euro").replace("  ", " ")


def _pzn(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    if s.isdigit():
        return s.zfill(8)
    return s


def _clean(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        return s[:-2]
    return " ".join(s.split())


def _num(v):
    if v in (None, ""):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("€", "")
    try:
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        return None


def _ensure_lernvorschlaege_table():
    with sqlite3.connect(DB_PATH) as con:
        con.execute("""
            CREATE TABLE IF NOT EXISTS tbl_lernvorschlaege (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                produkt_alt TEXT NOT NULL DEFAULT '',
                produkt_neu TEXT NOT NULL DEFAULT '',
                pzn_alt TEXT,
                artikel_alt TEXT,
                pzn_nmg TEXT,
                freitext_austausch TEXT,
                quelle_datei TEXT,
                status TEXT NOT NULL DEFAULT 'neu',
                erstellt_am TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                bearbeitet_am TEXT,
                bearbeiter TEXT,
                historie_ab TEXT,
                historie_bis TEXT,
                austausch_id INTEGER
            )
        """)
        cols = {row[1] for row in con.execute("PRAGMA table_info(tbl_lernvorschlaege)").fetchall()}
        for col, definition in {
            "pzn_alt": "TEXT",
            "artikel_alt": "TEXT",
            "pzn_nmg": "TEXT",
            "freitext_austausch": "TEXT",
            "quelle_datei": "TEXT",
            "bearbeiter": "TEXT",
            "historie_ab": "TEXT",
            "historie_bis": "TEXT",
            "austausch_id": "INTEGER",
        }.items():
            if col not in cols:
                con.execute(f"ALTER TABLE tbl_lernvorschlaege ADD COLUMN {col} {definition}")
        con.execute("UPDATE tbl_lernvorschlaege SET artikel_alt = COALESCE(NULLIF(artikel_alt, ''), produkt_alt) WHERE artikel_alt IS NULL OR artikel_alt = ''")
        con.execute("UPDATE tbl_lernvorschlaege SET freitext_austausch = COALESCE(NULLIF(freitext_austausch, ''), produkt_neu) WHERE freitext_austausch IS NULL OR freitext_austausch = ''")
        con.commit()


def _lookup_artikelname(con, pzn):
    """Sucht den Artikelnamen zu einer PZN in den bekannten Stammtabellen."""
    pzn = _pzn(pzn)
    if not pzn:
        return ""
    lookups = [
        ("tbl_artikelstamm", "artikel"),
        ("tbl_artikelstamm", "artikelname"),
        ("tbl_pzn_basisdaten", "artikelname"),
        ("tbl_nmg_stamm", "artikelname"),
    ]
    for table, col in lookups:
        try:
            row = con.execute(f"SELECT {col} FROM {table} WHERE pzn = ? LIMIT 1", (pzn,)).fetchone()
            if row and _clean(row[0]):
                return _clean(row[0])
        except Exception:
            continue
    return ""


def _austausch_ist_bereits_gelernt(con, pzn_alt):
    """Ein bereits aktiver Eintrag für die abgegebene PZN gilt als gelernt."""
    pzn_alt = _pzn(pzn_alt)
    if not pzn_alt:
        return False
    try:
        row = con.execute("""
            SELECT id
            FROM tbl_austauschdatenbank
            WHERE status = 'aktiv'
              AND pzn_alt = ?
            LIMIT 1
        """, (pzn_alt,)).fetchone()
        return row is not None
    except Exception:
        return False


def _lernvorschlag_existiert(con, pzn_alt, pzn_nmg, freitext):
    """Verhindert doppelte Schulbank-Vorschläge auch bei wiederholter gleicher Abweichungsanalyse."""
    pzn_alt = _pzn(pzn_alt)
    pzn_nmg = _pzn(pzn_nmg)
    freitext = _clean(freitext)
    if not pzn_alt:
        return True

    # Primär: gleiche abgegebene PZN + gleiche PZN NMG gilt als derselbe Fall.
    # Falls keine PZN NMG vorhanden ist, wird zusätzlich der Freitext verglichen.
    if pzn_nmg:
        row = con.execute("""
            SELECT id
            FROM tbl_lernvorschlaege
            WHERE COALESCE(pzn_alt, '') = ?
              AND COALESCE(pzn_nmg, '') = ?
              AND status IN ('neu', 'uebernommen', 'abgelehnt')
            LIMIT 1
        """, (pzn_alt, pzn_nmg)).fetchone()
    else:
        row = con.execute("""
            SELECT id
            FROM tbl_lernvorschlaege
            WHERE COALESCE(pzn_alt, '') = ?
              AND COALESCE(pzn_nmg, '') = ''
              AND COALESCE(freitext_austausch, produkt_neu, '') = ?
              AND status IN ('neu', 'uebernommen', 'abgelehnt')
            LIMIT 1
        """, (pzn_alt, freitext)).fetchone()
    return row is not None


def _save_lernvorschlaege_from_abweichung(manual, program, common):
    """
    Neue Schulbank-Regel aus der manuellen Anpassung:
    - Spalte H / Feld 'im_sortiment' gefüllt => potentieller Lernvorschlag.
    - Bereits aktiv gelernte pzn_alt in tbl_austauschdatenbank => ignorieren.
    - PZN NMG darf leer sein.
    - Wenn PZN NMG gefüllt und 'austauschbar gegen' leer ist, wird der Artikelname
      zur PZN NMG aus dem Artikelstamm/NMG-Stamm als Freitext vorbelegt.
    - Vorschläge werden gesammelt gespeichert, nicht einzeln während der Analyse abgefragt.
    """
    _ensure_lernvorschlaege_table()
    created = 0
    bearbeiter = getpass.getuser()

    with sqlite3.connect(DB_PATH) as con:
        for p in common:
            m = manual[p]
            g = program[p]

            # Entscheidend ist die händisch geprüfte Spalte H / 'im Sortiment'.
            if not _clean(m.get("im_sortiment")):
                continue

            pzn_alt = _pzn(p)
            if not pzn_alt:
                continue

            # Bereits gelernte alte PZN nicht erneut vorschlagen.
            if _austausch_ist_bereits_gelernt(con, pzn_alt):
                continue

            pzn_nmg = _pzn(m.get("pzn_nmg"))
            artikel_alt = _clean(m.get("artikel") or g.get("artikel")) or _lookup_artikelname(con, pzn_alt)
            freitext = _clean(m.get("austausch"))

            # Wenn eine konkrete PZN NMG da ist, aber kein Freitext, den Artikelnamen vorbelegen.
            if pzn_nmg and not freitext:
                freitext = _lookup_artikelname(con, pzn_nmg) or f"PZN NMG: {pzn_nmg}"

            # Wenn keine PZN NMG und kein Freitext vorhanden ist, trotzdem als offener Prüffall aufnehmen.
            produkt_alt = artikel_alt or pzn_alt
            produkt_neu = freitext or ""

            if _lernvorschlag_existiert(con, pzn_alt, pzn_nmg, produkt_neu):
                continue

            con.execute("""
                INSERT INTO tbl_lernvorschlaege (
                    produkt_alt,
                    produkt_neu,
                    pzn_alt,
                    artikel_alt,
                    pzn_nmg,
                    freitext_austausch,
                    quelle_datei,
                    status,
                    erstellt_am,
                    bearbeiter
                )
                VALUES (?, ?, ?, ?, ?, ?, 'Abweichungsanalyse', 'neu', CURRENT_TIMESTAMP, ?)
            """, (produkt_alt, produkt_neu, pzn_alt, artikel_alt, pzn_nmg, freitext, bearbeiter))

            created += 1

        con.commit()

    return created

def _same(a, b, field=None):
    # Leere Felder gleich behandeln.
    if _clean(a) == "" and _clean(b) == "":
        return True

    # Zahlen mit kleiner Toleranz vergleichen, damit Formel/Festwert/Rundung nicht stört.
    if field in {
        "apu_nmg",
        "nmg_rabatt",
        "rabatt_euro",
        "rabatt_gesamt",
        "umsatz",
        "ek",
        "absatz",
        "liefervorschlag",
    }:
        na, nb = _num(a), _num(b)
        if na is not None and nb is not None:
            return abs(na - nb) < 0.01

    return _clean(a).lower() == _clean(b).lower()


def _header_map(ws):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    return {_norm_header(h): i + 1 for i, h in enumerate(headers)}


def _detect_layout(ws):
    headers = [_norm_header(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

    # Marx/EKM manuelle Datei: A-Q Rohdaten, R-AB händische Auswertung.
    if ws.max_column >= 28 and "im sortiment" in headers[17] and "pzn nmg" in headers[18]:
        return "marx_manuell"

    # Linden/Test-Layout: A-R direkt.
    if headers and headers[0] == "pzn" and any("pzn nmg" == h for h in headers):
        return "linden"

    return "unbekannt"


def _read_positions(path):
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    layout = _detect_layout(ws)

    if layout == "unbekannt":
        raise ValueError(f"Dateiformat für Abweichungsanalyse nicht erkannt: {path.name}")

    positions = {}

    if layout == "marx_manuell":
        mapping = {
            "pzn": 1,
            "artikel": 14,
            "df": None,
            "pck": None,
            "herst": 15,
            "ek": None,
            "absatz": 16,
            "im_sortiment": 18,
            "pzn_nmg": 19,
            "apu_nmg": 20,
            "nmg_rabatt": 21,
            "lieferbar": 22,
            "bevorratung": 23,
            "liefervorschlag": 24,
            "austausch": 25,
            "rabatt_euro": 26,
            "rabatt_gesamt": 27,
            "umsatz": 28,
        }
    else:
        mapping = {
            "pzn": 1,
            "artikel": 2,
            "df": 3,
            "pck": 4,
            "herst": 5,
            "ek": 6,
            "absatz": 7,
            "im_sortiment": 8,
            "pzn_nmg": 9,
            "apu_nmg": 10,
            "nmg_rabatt": 11,
            "lieferbar": 12,
            "bevorratung": 13,
            "liefervorschlag": 14,
            "austausch": 15,
            "rabatt_euro": 16,
            "rabatt_gesamt": 17,
            "umsatz": 18,
        }

    for r in range(2, ws.max_row + 1):
        p = _pzn(ws.cell(r, mapping["pzn"]).value)
        if not p:
            continue

        row = {}
        for k, col in mapping.items():
            row[k] = ws.cell(r, col).value if col else None

        positions[p] = row

    return positions, layout


def _append_header(ws, headers):
    ws.append(headers)
    fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7B7B7")

    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)


def _autosize(ws, max_width=48):
    for col in range(1, ws.max_column + 1):
        width = 10
        for row in range(1, min(ws.max_row, 250) + 1):
            val = ws.cell(row, col).value
            if val is not None:
                width = max(width, min(max_width, len(str(val)) + 2))

        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def export_abweichungsanalyse(manuelle_datei, programm_datei):
    man_path = Path(manuelle_datei)
    prog_path = Path(programm_datei)

    manual, manual_layout = _read_positions(man_path)
    program, program_layout = _read_positions(prog_path)

    common = sorted(set(manual) & set(program))
    only_manual = sorted(set(manual) - set(program))
    only_program = sorted(set(program) - set(manual))

    wb = Workbook()

    ws_stat = wb.active
    ws_stat.title = "Statistik"
    ws_stat.append(["Manuelle Datei", man_path.name])
    ws_stat.append(["Programm-Datei", prog_path.name])
    ws_stat.append(["Manuelles Layout", manual_layout])
    ws_stat.append(["Programm-Layout", program_layout])
    ws_stat.append(["Verglichene PZN", len(common)])
    ws_stat.append(["Nur manuell", len(only_manual)])
    ws_stat.append(["Nur Programm", len(only_program)])
    ws_stat.append([])
    ws_stat.append(["Bereich", "Abweichungen"])

    # Alle Abweichungen sammeln.
    field_counts = {k: 0 for k, _ in RAW_FIELDS + COMPARE_FIELDS}
    all_diffs = []

    for p in common:
        m = manual[p]
        g = program[p]

        for k, label in RAW_FIELDS + COMPARE_FIELDS:
            if not _same(m.get(k), g.get(k), k):
                field_counts[k] += 1
                all_diffs.append([
                    p,
                    m.get("artikel") or g.get("artikel"),
                    label,
                    m.get(k),
                    g.get(k),
                ])

    groups = {
        "NMG_ZUORDNUNG": ["pzn_nmg", "apu_nmg", "nmg_rabatt"],
        "LIEFERVORSCHLAG": ["liefervorschlag"],
        "AUSTAUSCH": ["austausch"],
        "LIEFERBAR_BEVORRATUNG": ["im_sortiment", "lieferbar", "bevorratung"],
        "ROHDATEN": ["artikel", "df", "pck", "herst", "ek", "absatz"],
    }

    for name, keys in groups.items():
        affected = set()
        for p in common:
            if any(not _same(manual[p].get(k), program[p].get(k), k) for k in keys):
                affected.add(p)

        ws_stat.append([name, len(affected)])

    ws_stat.append(["Alle Feldabweichungen", len(all_diffs)])

    # Schulbank-Lernvorschläge aus Austausch-Abweichungen speichern.
    lernvorschlaege = _save_lernvorschlaege_from_abweichung(manual, program, common)
    ws_stat.append(["Neue Lernvorschläge Schulbank", lernvorschlaege])

    for c in range(1, 3):
        ws_stat.cell(9, c).font = Font(bold=True)
        ws_stat.cell(9, c).fill = PatternFill("solid", fgColor="D9EAF7")

    _autosize(ws_stat)

    def make_group_sheet(title, keys):
        ws = wb.create_sheet(title)
        headers = ["PZN", "Artikel"]
        labels = dict(RAW_FIELDS + COMPARE_FIELDS)

        for k in keys:
            headers += [
                f"Manuell {labels[k]}",
                f"Programm {labels[k]}",
            ]

        headers.append("Abweichende Felder")
        _append_header(ws, headers)

        for p in common:
            diff_keys = [
                k for k in keys
                if not _same(manual[p].get(k), program[p].get(k), k)
            ]

            if not diff_keys:
                continue

            row = [
                p,
                manual[p].get("artikel") or program[p].get("artikel"),
            ]

            for k in keys:
                row += [
                    manual[p].get(k),
                    program[p].get(k),
                ]

            row.append(", ".join(labels[k] for k in diff_keys))
            ws.append(row)

        _autosize(ws)

    make_group_sheet("NMG_ZUORDNUNG", groups["NMG_ZUORDNUNG"])
    make_group_sheet("LIEFERVORSCHLAG", groups["LIEFERVORSCHLAG"])
    make_group_sheet("AUSTAUSCH", groups["AUSTAUSCH"])
    make_group_sheet("LIEFERBAR_BEVORRATUNG", groups["LIEFERBAR_BEVORRATUNG"])
    make_group_sheet("ROHDATEN", groups["ROHDATEN"])

    ws_all = wb.create_sheet("ALLE_ABWEICHUNGEN")
    _append_header(ws_all, ["PZN", "Artikel", "Feld", "Manuell", "Programm"])

    for row in all_diffs:
        ws_all.append(row)

    _autosize(ws_all)

    ws_missing = wb.create_sheet("PZN_FEHLEN")
    _append_header(ws_missing, ["Status", "PZN", "Artikel"])

    for p in only_manual:
        ws_missing.append([
            "Nur manuell",
            p,
            manual[p].get("artikel"),
        ])

    for p in only_program:
        ws_missing.append([
            "Nur Programm",
            p,
            program[p].get("artikel"),
        ])

    _autosize(ws_missing)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    safe = ''.join(
        ch if ch.isalnum() or ch in '-_' else '_'
        for ch in man_path.stem
    )[:45]

    out = OUTPUT_DIR / f"Abweichungsanalyse_{safe}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    wb.save(out)

    return out