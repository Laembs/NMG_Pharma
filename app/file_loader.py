from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


SUPPORTED_DATA_EXTENSIONS = {".xlsx", ".xlsm", ".xls", ".csv", ".txt"}
SUPPORTED_DATA_FILETYPES = [
    ("Daten", "*.xlsx *.xlsm *.xls *.csv *.txt"),
    ("Excel", "*.xlsx *.xlsm *.xls"),
    ("Excel (alt)", "*.xls"),
    ("CSV", "*.csv"),
    ("Text", "*.txt"),
    ("Alle Dateien", "*.*"),
]


class UnsupportedFileFormatError(ValueError):
    pass


class TableData:
    """Einheitliche Tabellendaten fuer Excel, CSV und TXT.

    headers: erste erkannte Zeile
    rows: Datenzeilen danach
    source_type: excel/csv/txt
    delimiter: bei csv/txt erkanntes Trennzeichen
    """

    def __init__(self, headers, rows, source_type="", delimiter=None, sheet_name=""):
        self.headers = list(headers or [])
        self.rows = [list(r) for r in (rows or [])]
        self.source_type = source_type
        self.delimiter = delimiter
        self.sheet_name = sheet_name

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def column_count(self) -> int:
        return max([len(self.headers)] + [len(r) for r in self.rows] + [0])


class _Cell:
    def __init__(self, value):
        self.value = value


class TableWorksheet:
    """Kleine openpyxl-kompatible Huelle.

    Damit koennen bestehende Funktionen wie find_columns(ws), Diagnose und
    iter_rows(values_only=True) auch CSV/TXT verarbeiten.
    """

    def __init__(self, table: TableData):
        self._rows = [table.headers] + table.rows
        self.max_row = len(self._rows)
        self.max_column = max((len(r) for r in self._rows), default=0)
        self.title = table.sheet_name or table.source_type or "Tabelle"

    def __getitem__(self, row_idx):
        row = self._rows[row_idx - 1] if 1 <= row_idx <= self.max_row else []
        padded = list(row) + [None] * max(0, self.max_column - len(row))
        return [_Cell(v) for v in padded]

    def cell(self, row, column):
        try:
            return _Cell(self._rows[row - 1][column - 1])
        except Exception:
            return _Cell(None)

    def iter_rows(self, min_row=1, max_row=None, min_col=1, max_col=None, values_only=False, **kwargs):
        max_row = max_row or self.max_row
        max_col = max_col or self.max_column
        for r in range(min_row, max_row + 1):
            source = self._rows[r - 1] if 1 <= r <= self.max_row else []
            vals = []
            for c in range(min_col, max_col + 1):
                vals.append(source[c - 1] if c - 1 < len(source) else None)
            if values_only:
                yield tuple(vals)
            else:
                yield tuple(_Cell(v) for v in vals)


def _read_text_with_fallback(path: Path) -> str:
    last_error = None
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            return path.read_text(encoding=enc)
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error:
        return path.read_text(encoding="latin1", errors="replace")
    return path.read_text(errors="replace")


def _detect_delimiter(sample: str) -> str:
    """Erkennt uebliche Trenner fuer CSV/TXT.

    Reihenfolge/Logik:
    1. csv.Sniffer
    2. haeufigster Trenner aus ; , TAB |
    3. Semikolon als deutscher Standard
    """
    if not sample:
        return ";"

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";\t,|")
        if dialect.delimiter:
            return dialect.delimiter
    except Exception:
        pass

    candidates = [";", "\t", "|", ","]
    counts = {d: sample.count(d) for d in candidates}
    best = max(counts, key=counts.get)
    return best if counts[best] > 0 else ";"


def _strip_empty_tail(row):
    row = list(row)
    while row and row[-1] in (None, ""):
        row.pop()
    return row


def _rows_from_text(path: Path) -> tuple[list[list[str]], str]:
    text = _read_text_with_fallback(path)
    sample = text[:8192]
    delimiter = _detect_delimiter(sample)

    rows = []
    reader = csv.reader(text.splitlines(), delimiter=delimiter)
    for row in reader:
        cleaned = [cell.strip() if isinstance(cell, str) else cell for cell in row]
        if any(cell not in (None, "") for cell in cleaned):
            rows.append(_strip_empty_tail(cleaned))
    return rows, delimiter


def _rows_from_xls(path: Path, sheet_name: str | None = None) -> tuple[list, list, str]:
    """Liest .xls-Dateien.

    openpyxl kann nur xlsx/xlsm. Fuer echtes Excel 97-2003 (BIFF) wird xlrd
    benutzt. Viele Apotheken-/Grosshandels-Exporte sind aber als .xls deklariert,
    in Wirklichkeit aber XML- (SpreadsheetML 2003) oder HTML-Tabellen. Schlaegt
    xlrd fehl, wird auf diese Formate zurueckgefallen.

    Zahlen, die ganzzahlig sind, werden als int geliefert, damit PZNs nicht als
    12345678.0 ankommen.
    """
    import xlrd  # nur bei Bedarf laden

    try:
        book = xlrd.open_workbook(str(path))
    except Exception:
        # Keine echte Binaer-.xls -> XML/HTML-Fallback.
        return _rows_from_pseudo_xls(path)

    sheet = book.sheet_by_name(sheet_name) if sheet_name else book.sheet_by_index(0)

    def _cell(value, ctype):
        if ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR):
            return None
        if ctype == xlrd.XL_CELL_NUMBER and float(value).is_integer():
            return int(value)
        return value

    all_rows = []
    for r in range(sheet.nrows):
        values = [_cell(sheet.cell_value(r, c), sheet.cell_type(r, c)) for c in range(sheet.ncols)]
        values = _strip_empty_tail(values)
        if any(v not in (None, "") for v in values):
            all_rows.append(values)
    headers = all_rows[0] if all_rows else []
    rows = all_rows[1:] if len(all_rows) > 1 else []
    return headers, rows, sheet.name


def _coerce_value(text):
    """Wandelt reine Ganzzahl-Strings in int (z.B. PZN '12345678'); Zahlen
    (auch numpy-Typen) bleiben Zahl; sonst Text."""
    if text is None:
        return None
    if isinstance(text, (int, float)):  # numpy int64/float64 sind Subklassen
        return text
    s = str(text).strip()
    if s == "":
        return None
    if s.isdigit():
        try:
            return int(s)
        except ValueError:
            return s
    return s


def _rows_from_pseudo_xls(path: Path) -> tuple[list, list, str]:
    """Liest als .xls deklarierte XML- (SpreadsheetML 2003) oder HTML-Tabellen.

    Solche Dateien entstehen haeufig bei Apotheken-/Grosshandelssoftware, die
    'Excel-Export' anbietet, intern aber XML oder HTML schreibt.
    """
    text = _read_text_with_fallback(path)
    head = text.lstrip("﻿")[:4096].lower()

    if "urn:schemas-microsoft-com:office:spreadsheet" in text or head.startswith("<?xml"):
        all_rows, title = _rows_from_spreadsheetml(text)
    elif "<table" in head or "<html" in head:
        all_rows, title = _rows_from_html_table(path)
    else:
        raise UnsupportedFileFormatError(
            "Die .xls-Datei ist weder eine echte Excel-Datei noch eine erkennbare "
            "XML-/HTML-Tabelle und konnte nicht gelesen werden."
        )

    all_rows = [_strip_empty_tail(r) for r in all_rows]
    all_rows = [r for r in all_rows if any(v not in (None, "") for v in r)]
    headers = all_rows[0] if all_rows else []
    rows = all_rows[1:] if len(all_rows) > 1 else []
    return headers, rows, title


def _rows_from_spreadsheetml(text: str) -> tuple[list, str]:
    """Parst Excel-2003-XML (SpreadsheetML). Beruecksichtigt ss:Index (luecken-
    hafte Zellen) und nimmt das erste Worksheet/Table."""
    import xml.etree.ElementTree as ET

    SS = "urn:schemas-microsoft-com:office:spreadsheet"
    ns_row = f"{{{SS}}}Row"
    ns_cell = f"{{{SS}}}Cell"
    ns_data = f"{{{SS}}}Data"
    ns_index = f"{{{SS}}}Index"

    root = ET.fromstring(text.encode("utf-8") if isinstance(text, str) else text)
    title = "Tabelle"
    for ws in root.iter(f"{{{SS}}}Worksheet"):
        name = ws.get(f"{{{SS}}}Name")
        if name:
            title = name
        table = ws.find(f"{{{SS}}}Table")
        if table is None:
            continue
        rows = []
        for row in table.findall(ns_row):
            cells: list = []
            col = 0
            for cell in row.findall(ns_cell):
                idx = cell.get(ns_index)
                if idx:
                    col = int(idx) - 1
                while len(cells) < col:
                    cells.append(None)
                data = cell.find(ns_data)
                cells.append(_coerce_value(data.text if data is not None else None))
                col += 1
            rows.append(cells)
        return rows, title
    return [], title


def _rows_from_html_table(path: Path) -> tuple[list, str]:
    """Parst HTML-Tabellen, die als .xls exportiert wurden (via pandas.read_html).

    thousands='.'/decimal=',' interpretiert deutsche Zahlenformate korrekt
    (z.B. '1.234,56' -> 1234.56). Die Spaltenkoepfe werden als erste Zeile
    uebernommen, weil pandas die Kopfzeile sonst in die Spaltennamen zieht.
    """
    import pandas as pd

    tables = pd.read_html(str(path), thousands=".", decimal=",")
    if not tables:
        return [], "Tabelle"
    # Groesste Tabelle nehmen (Layout-Tabellen sind meist kleiner).
    df = max(tables, key=lambda t: t.shape[0] * t.shape[1])

    def _label(c):
        if isinstance(c, tuple):
            c = c[-1]
        return _coerce_value(c)

    rows = []
    cols = list(df.columns)
    # Echte Kopfzeile (keine 0,1,2,...-Default-Spalten) als erste Zeile retten.
    if not all(isinstance(c, int) and not isinstance(c, bool) for c in cols):
        rows.append([_label(c) for c in cols])
    for row in df.itertuples(index=False):
        rows.append([_coerce_value(v) if not pd.isna(v) else None for v in row])
    return rows, "Tabelle"


def load_table(path: str | Path, sheet_name: str | None = None) -> TableData:
    """Laedt xlsx/xlsm/xls/csv/txt als einheitliche Tabelle.

    Bei Excel wird das erste aktive Blatt verwendet, falls kein sheet_name angegeben ist.
    Bei CSV/TXT wird die erste nicht-leere Zeile als Header verwendet.
    """
    path = Path(path)
    suffix = path.suffix.lower()

    if suffix not in SUPPORTED_DATA_EXTENSIONS:
        raise UnsupportedFileFormatError(
            f"Nicht unterstütztes Dateiformat: {suffix}. Erlaubt sind .xlsx, .xlsm, .xls, .csv und .txt."
        )

    if suffix == ".xls":
        headers, rows, sheet_title = _rows_from_xls(path, sheet_name)
        return TableData(headers=headers, rows=rows, source_type="excel", sheet_name=sheet_title)

    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            values = _strip_empty_tail(row)
            if any(v not in (None, "") for v in values):
                all_rows.append(values)
        headers = all_rows[0] if all_rows else []
        rows = all_rows[1:] if len(all_rows) > 1 else []
        return TableData(headers=headers, rows=rows, source_type="excel", sheet_name=ws.title)

    rows, delimiter = _rows_from_text(path)
    headers = rows[0] if rows else []
    data_rows = rows[1:] if len(rows) > 1 else []
    return TableData(headers=headers, rows=data_rows, source_type="txt" if suffix == ".txt" else "csv", delimiter=delimiter)


def load_worksheet(path: str | Path, sheet_name: str | None = None):
    """Laedt Dateien fuer bestehende openpyxl-nahe Verarbeitung.

    Rueckgabe:
    - Excel: echtes Worksheet
    - CSV/TXT: TableWorksheet
    - source_type: excel/csv/txt
    """
    path = Path(path)
    suffix = path.suffix.lower()

    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        return ws, "excel"

    table = load_table(path, sheet_name=sheet_name)
    return TableWorksheet(table), table.source_type


def normalize_header(value) -> str:
    import re
    text = str(value or "").strip().lower()
    text = text.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    text = text.replace("_", " ").replace("-", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def compact_header(value) -> str:
    import re
    return re.sub(r"[^a-z0-9]+", "", normalize_header(value))


def find_column(headers: Iterable, aliases: Iterable[str], contains: Iterable[str] = ()):
    normalized = [normalize_header(h) for h in headers]
    compacted = [compact_header(h) for h in headers]
    alias_norm = {normalize_header(a) for a in aliases}
    alias_compact = {compact_header(a) for a in aliases}
    contains_compact = [compact_header(c) for c in contains]

    for idx, h in enumerate(normalized):
        if h in alias_norm:
            return idx

    for idx, h in enumerate(compacted):
        if h in alias_compact:
            return idx
        if any(part and part in h for part in contains_compact):
            return idx

    return None
