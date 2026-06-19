from __future__ import annotations
from pathlib import Path
from datetime import datetime
import re
import sqlite3
from openpyxl import load_workbook
from .config import DB_PATH, HISTORICAL_ANALYSIS_DIR
from .db import connect, init_db, _pzn
from .learning_db import clean_hersteller, parse_number, find_columns, lookup_basisdaten, register_basisdaten, lookup_latest_ek, register_ek
from .exporter import _lookup

STANDARD_HEADERS = [
    "PZN", "Artikelname", "DF", "Pck", "Herst", "EK", "Abverkäufe 6 Monate",
    "im Sortiment", "PZN NMG", "APU NMG", "NMG Rabatt", "lieferbar",
    "Bevorratung angeraten", "Liefervor- schlag", "austauschbar gegen",
    "NMG Rabatt in Euro", "NMG Rabatt Gesamt nach Absatz", "Umsatz"
]


def _norm(text):
    return re.sub(r"[^a-z0-9]+", "", str(text or "").strip().lower().replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss"))


def _apotheke_from_filename(path: Path) -> str:
    stem = path.stem
    stem = re.sub(r"^Hochpreiser[_\s-]*", "", stem, flags=re.I)
    stem = re.sub(r"_?\d{1,2}[.]\d{1,2}[.]\d{2,4}.*$", "", stem)
    stem = re.sub(r"_?\d{8,}.*$", "", stem)
    stem = stem.replace("_", " ").strip(" -")
    return stem or path.stem


def _find_standard_header_row(ws, scan_rows=20):
    required = {_norm("PZN"), _norm("Artikelname"), _norm("Abverkäufe 6 Monate")}
    for r in range(1, min(ws.max_row, scan_rows) + 1):
        vals = [_norm(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 30) + 1)]
        if required.issubset(set(vals)):
            return r
        # häufig exakt A-R ohne sauber normalisierte Absatzüberschrift
        if "pzn" in vals and len(vals) >= 5 and any(x in vals for x in ("artikelname", "artikel", "artikelbezeichnung", "vollstartikelname")):
            return r
        if "pzn" in vals and any("nmg" in x for x in vals):
            return r
    return None


def _header_map(ws, header_row):
    mapping = {}
    for c in range(1, ws.max_column + 1):
        h = _norm(ws.cell(header_row, c).value)
        if h:
            mapping[h] = c
    def find(*names, default=None, contains=()):
        keys = [_norm(n) for n in names]
        for key in keys:
            if key in mapping:
                return mapping[key]
        for h, c in mapping.items():
            if any(part in h for part in contains):
                return c
        return default
    return {
        "pzn": find("PZN", default=1, contains=("pzn",)),
        "artikel": find("Artikelname", "Artikel", "Artikelbezeichnung", "Vollst. Artikelname", default=2, contains=("artikel", "bezeichnung", "vollstartikelname")),
        "df": find("DF", "DAR", default=3, contains=("dar",)),
        "pck": find("Pck", "Packung", "Pack.Gr", "Einheit", default=4, contains=("pack", "einheit")),
        "hersteller": find("Herst", "Hersteller", "Anbieter", default=5, contains=("herst", "anbieter")),
        "ek": find("EK", "Apo-EK", "Lauer-EK", "Tax-EK", "Taxe-EK", "Durch.-EK", "Eigen-EK", default=6, contains=("ek",)),
        "absatz": find("Abverkäufe 6 Monate", "Verkaufsmenge der letzten 6 Monate", "Absatz", "Abs.", "Packungen (Abgegeben)", "Gesamtmenge", default=7, contains=("abverkaeufe", "verkaufsmenge", "abs", "packungenabgegeben", "gesamtmenge")),
        "im_sortiment": find("im Sortiment", default=8, contains=("imsortiment",)),
        "pzn_nmg": find("PZN NMG", default=9, contains=("pznnmg",)),
        "apu_nmg": find("APU NMG", "Apu NMG", default=10, contains=("apunmg",)),
        "nmg_rabatt": find("NMG Rabatt", "Rabatt NMG", default=11, contains=("nmgrabatt", "rabattnmg")),
        "lieferbar": find("lieferbar", default=12, contains=("lieferbar",)),
        "bevorratung": find("Bevorratung angeraten", default=13, contains=("bevorratung",)),
        "liefervorschlag": find("Liefervor- schlag", "Liefervorschlag", default=14, contains=("liefervor", "liefervorschlag")),
        "austausch": find("austauschbar gegen", default=15, contains=("austauschbar",)),
        "rabatt_euro": find("NMG Rabatt in Euro", default=16, contains=("rabattineuro",)),
        "rabatt_gesamt": find("NMG Rabatt Gesamt nach Absatz", default=17, contains=("rabattgesamtnachabsatz",)),
        "umsatz": find("Umsatz", default=18, contains=("umsatz",)),
    }


def _get(row, col):
    if not col or col < 1 or col > len(row):
        return None
    return row[col-1]


def _to_float(v):
    return parse_number(v)


def _clean_value(v):
    # SQLite kann datetime.time/date aus Excel nicht direkt binden.
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    return str(v).strip() if str(v).strip() else None


def _is_nmg_hit(im_sortiment, pzn_nmg, apu_nmg) -> int:
    if pzn_nmg not in (None, ""):
        return 1
    if apu_nmg not in (None, "") and _to_float(apu_nmg) is not None:
        return 1
    if im_sortiment not in (None, "") and str(im_sortiment).strip().upper().startswith("X"):
        return 1
    return 0


def _build_rohdaten_cache(con: sqlite3.Connection, pzns_norm: set[str]) -> dict:
    """V1.1 SP16: Pre-Cache fuer den rohdaten-Modus. Statt pro Zeile
    lookup_basisdaten + _lookup (= mehrere UDF-Queries ueber grosse
    Tabellen) wird hier EINMAL alles geladen, was die Schleife danach
    pro PZN braucht.

    Annahme: PZNs in tbl_pzn_basisdaten, tbl_austauschdatenbank,
    tbl_nmg_stamm, nmg_rabatte, tbl_lieferfaehigkeit sind beim Insert
    bereits normalisiert (8 Stellen, nur Ziffern). Damit funktioniert
    direkter =-Vergleich + IN ueber den jeweiligen Index.
    """
    caches = {
        "basisdaten": {},      # pzn -> {artikelname, herstellerkuerzel, df, pck}
        "austausch":  {},      # pzn_alt -> [list of austausch-records]
        "nmg_stamm":  {},      # pzn -> {apu, artikelname}
        "nmg_rabatte": {},     # pzn -> {rabatt, ...}
        "lieferfaehigkeit": {},  # pzn -> {lieferbar, ...}
    }
    if not pzns_norm:
        return caches
    pzn_list = list(pzns_norm)
    # SQLite: max 999 Parameter pro Query (Standard). Chunked.
    chunk = 900

    def _in_chunks(items):
        for i in range(0, len(items), chunk):
            yield items[i:i + chunk]

    # Basisdaten
    for ch in _in_chunks(pzn_list):
        placeholders = ",".join("?" * len(ch))
        for r in con.execute(
            f"SELECT pzn, artikelname, herstellerkuerzel, df, pck "
            f"FROM tbl_pzn_basisdaten WHERE pzn IN ({placeholders})", ch
        ).fetchall():
            caches["basisdaten"][r[0]] = {
                "artikelname": r[1], "herstellerkuerzel": r[2],
                "df": r[3], "pck": r[4],
            }

    # Austauschdatenbank: alle aktiven Eintraege fuer diese PZNs.
    # Sortier-Reihenfolge analog der bisherigen _lookup()-Logik:
    # 1. Treffer mit Rabatt vor solchen ohne, dann hoechster Rabatt
    # 2. Im NMG-Stamm vor sonstigen
    # 3. aktualisiert_am DESC / id DESC
    for ch in _in_chunks(pzn_list):
        placeholders = ",".join("?" * len(ch))
        try:
            rows = con.execute(f"""
                SELECT a.id, a.pzn_alt, a.pzn_nmg, a.artikel_nmg,
                       a.freitext_austausch, a.aktualisiert_am, a.erstellt_am,
                       r.rabatt AS _rabatt,
                       ns.pzn   AS _nmg_stamm
                FROM tbl_austauschdatenbank a
                LEFT JOIN nmg_rabatte r ON r.nmg_pzn = a.pzn_nmg
                LEFT JOIN tbl_nmg_stamm ns ON ns.pzn = a.pzn_nmg
                WHERE COALESCE(a.status, 'aktiv') = 'aktiv'
                  AND a.pzn_alt IN ({placeholders})
                ORDER BY
                    CASE WHEN r.rabatt IS NOT NULL THEN 0 ELSE 1 END,
                    COALESCE(r.rabatt, 0) DESC,
                    CASE WHEN ns.pzn IS NOT NULL THEN 0 ELSE 1 END,
                    datetime(COALESCE(a.aktualisiert_am, a.erstellt_am, '1970-01-01')) DESC,
                    a.id DESC
            """, ch).fetchall()
        except sqlite3.OperationalError:
            rows = []
        for r in rows:
            caches["austausch"].setdefault(r[1], []).append({
                "id": r[0], "pzn_alt": r[1], "pzn_nmg": r[2],
                "artikel_nmg": r[3], "freitext_austausch": r[4],
                "aktualisiert_am": r[5], "erstellt_am": r[6],
                "rabatt": r[7], "nmg_stamm": r[8],
            })

    # Welche NMG-PZNs werden ueber die Austausch-Cache referenziert?
    referenced_nmg_pzns = set()
    for lst in caches["austausch"].values():
        for entry in lst:
            if entry["pzn_nmg"]:
                referenced_nmg_pzns.add(_pzn(entry["pzn_nmg"]))
    nmg_pzn_list = list(referenced_nmg_pzns)

    # NMG-Stamm: apu pro nmg_pzn
    for ch in _in_chunks(nmg_pzn_list):
        placeholders = ",".join("?" * len(ch))
        try:
            for r in con.execute(
                f"SELECT pzn, apu, artikelname FROM tbl_nmg_stamm WHERE pzn IN ({placeholders})",
                ch,
            ).fetchall():
                caches["nmg_stamm"][r[0]] = {"apu": r[1], "artikelname": r[2]}
        except sqlite3.OperationalError:
            pass

    # NMG-Rabatte
    for ch in _in_chunks(nmg_pzn_list):
        placeholders = ",".join("?" * len(ch))
        try:
            for r in con.execute(
                f"SELECT nmg_pzn, rabatt FROM nmg_rabatte WHERE nmg_pzn IN ({placeholders})",
                ch,
            ).fetchall():
                caches["nmg_rabatte"][r[0]] = {"rabatt": r[1]}
        except sqlite3.OperationalError:
            pass

    # Lieferfaehigkeit
    for ch in _in_chunks(nmg_pzn_list):
        placeholders = ",".join("?" * len(ch))
        try:
            for r in con.execute(
                f"SELECT nmg_pzn, lieferbar, bevorratung_angeraten, liefervorschlag "
                f"FROM tbl_lieferfaehigkeit WHERE nmg_pzn IN ({placeholders})",
                ch,
            ).fetchall():
                caches["lieferfaehigkeit"][r[0]] = {
                    "lieferbar": r[1],
                    "bevorratung_angeraten": r[2],
                    "liefervorschlag": r[3],
                }
        except sqlite3.OperationalError:
            pass

    return caches


def _lookup_from_cache(caches: dict, pzn_norm: str) -> dict:
    """V1.1 SP16: Dict-Lookup-Variante von _lookup() fuer den rohdaten-Modus.
    Liefert die gleichen Felder wie _lookup, aber ohne DB-Queries.
    """
    austausch_list = caches["austausch"].get(pzn_norm) or []
    if not austausch_list:
        return {}
    main = austausch_list[0]
    nmg_pzn = _pzn(main.get("pzn_nmg")) or None
    freitext = (main.get("freitext_austausch") or "").strip()
    artikel_nmg = (main.get("artikel_nmg") or "").strip()
    if nmg_pzn and (not freitext or freitext.lower().startswith("pzn nmg:")):
        # Artikel-Name bevorzugt aus austausch-Eintrag, sonst aus NMG-Stamm.
        nmg_stamm = caches["nmg_stamm"].get(nmg_pzn, {})
        freitext = artikel_nmg or nmg_stamm.get("artikelname", "") or freitext

    # Weitere Austauschoptionen anhaengen (nur die mit Eintrag im NMG-Stamm).
    weitere_texte = []
    seen = {nmg_pzn} if nmg_pzn else set()
    for extra in austausch_list[1:]:
        ext_pzn = _pzn(extra.get("pzn_nmg"))
        if not ext_pzn or ext_pzn in seen:
            continue
        if ext_pzn not in caches["nmg_stamm"]:
            continue
        seen.add(ext_pzn)
        ext_name = (extra.get("artikel_nmg") or "").strip() or caches["nmg_stamm"].get(ext_pzn, {}).get("artikelname", "")
        weitere_texte.append(f"PZN {ext_pzn}" + (f" – {ext_name}" if ext_name else ""))
    if weitere_texte:
        zusatz = "weitere: " + " | ".join(weitere_texte)
        freitext = f"{freitext} | {zusatz}" if freitext else zusatz

    apu = None
    rabatt = None
    lieferbar = None
    bevorratung = None
    liefervorschlag = None
    if nmg_pzn:
        nmg_stamm = caches["nmg_stamm"].get(nmg_pzn, {})
        apu = nmg_stamm.get("apu")
        rab = caches["nmg_rabatte"].get(nmg_pzn, {})
        rabatt = rab.get("rabatt")
        lief = caches["lieferfaehigkeit"].get(nmg_pzn, {})
        lieferbar = lief.get("lieferbar")
        bevorratung = lief.get("bevorratung_angeraten")
        liefervorschlag = lief.get("liefervorschlag")

    return {
        "im_sortiment": "X" if nmg_pzn else "X Austausch mögl",
        "nmg_pzn": nmg_pzn,
        "apu_nmg": apu,
        "rabatt": rabatt,
        "lieferbar": lieferbar,
        "bevorratung": bevorratung,
        "liefervorschlag": liefervorschlag,
        "austauschbar_gegen": freitext or None,
    }


def import_historical_market_file(file_path: str | Path, con: sqlite3.Connection | None = None, datenquelle: str = "NMG", analyse_name: str | None = None) -> dict:
    """Importiert eine vorhandene händische Auswertung oder schwierige Rohdaten nur in die Marktanalyse.

    Wichtig: Diese Funktion lernt keine fachlichen H-O-Werte. Sie füllt nur tbl_auswertungen
    und tbl_auswertungspositionen, damit Marktanalysen über historische Dateien möglich sind.
    """
    path = Path(file_path)
    own_con = False
    if con is None:
        if not DB_PATH.exists():
            init_db(DB_PATH)
        con = connect(DB_PATH)
        own_con = True
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        apotheke = analyse_name or _apotheke_from_filename(path)
        datenquelle = (datenquelle or "NMG").upper()

        header_row = _find_standard_header_row(ws)
        mode = "auswertung_ar" if header_row else "rohdaten_erkannt"
        rows_to_insert = []

        if header_row:
            mp = _header_map(ws, header_row)
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                pzn = _pzn(_get(row, mp["pzn"]))
                if not pzn:
                    continue
                artikel = _get(row, mp["artikel"])
                df = _get(row, mp["df"])
                pck = _get(row, mp["pck"])
                herst = clean_hersteller(_get(row, mp["hersteller"]))
                ek = _to_float(_get(row, mp["ek"]))
                absatz = _to_float(_get(row, mp["absatz"])) or 0.0
                im_sortiment = _get(row, mp["im_sortiment"])
                pzn_nmg = _pzn(_get(row, mp["pzn_nmg"])) or None
                apu_nmg = _to_float(_get(row, mp["apu_nmg"]))
                rabatt = _to_float(_get(row, mp["nmg_rabatt"]))
                lieferbar = _get(row, mp["lieferbar"])
                bevorratung = _get(row, mp["bevorratung"])
                liefervorschlag = _get(row, mp["liefervorschlag"])
                austausch = _get(row, mp["austausch"])
                rabatt_euro = _to_float(_get(row, mp["rabatt_euro"]))
                rabatt_gesamt = _to_float(_get(row, mp["rabatt_gesamt"]))
                umsatz = _to_float(_get(row, mp["umsatz"]))
                rows_to_insert.append((pzn, _clean_value(artikel), _clean_value(df), _clean_value(pck), _clean_value(herst), ek, absatz, _clean_value(im_sortiment), pzn_nmg, apu_nmg, rabatt, _clean_value(lieferbar), _clean_value(bevorratung), _clean_value(liefervorschlag), _clean_value(austausch), rabatt_euro, rabatt_gesamt, umsatz, _is_nmg_hit(im_sortiment, pzn_nmg, apu_nmg)))
        else:
            mapping = find_columns(ws)
            if not mapping or not (mapping.get("pzn") and mapping.get("absatz")):
                raise ValueError("Format nicht erkannt: historische Datei wurde nicht importiert.")
            header_row = mapping.get("header_row", 1)

            # V1.1 SP16: Statt pro Zeile lookup_basisdaten + _lookup
            # (das macht UDF-WHERE-Klauseln ueber grosse Tabellen = Full-Scan)
            # erst ALLE PZNs sammeln, dann Pre-Cache in einem Rutsch.
            rohdaten_rows = []
            pzns_norm: set[str] = set()
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                pzn = _pzn(_get(row, mapping.get("pzn")))
                if not pzn:
                    continue
                rohdaten_rows.append((row, pzn))
                pzns_norm.add(pzn)

            caches = _build_rohdaten_cache(con, pzns_norm)

            for row, pzn in rohdaten_rows:
                basis = caches["basisdaten"].get(pzn, {})
                artikel = _get(row, mapping.get("artikel")) or basis.get("artikelname")
                df = _get(row, mapping.get("df")) or basis.get("df")
                pck = _get(row, mapping.get("packung")) or basis.get("pck")
                herst = clean_hersteller(_get(row, mapping.get("hersteller"))) or basis.get("herstellerkuerzel")
                ek = _to_float(_get(row, mapping.get("ek")))
                absatz = _to_float(_get(row, mapping.get("absatz"))) or 0.0
                # NMG-Status aus dem Cache (keine UDF-Joins mehr).
                hit = _lookup_from_cache(caches, pzn)
                pzn_nmg = hit.get("nmg_pzn")
                apu_nmg = hit.get("apu_nmg")
                rabatt = hit.get("rabatt")
                rabatt_euro = (_to_float(apu_nmg) or 0) * (_to_float(rabatt) or 0)
                rabatt_gesamt = rabatt_euro * absatz
                umsatz = (ek or 0) * absatz if ek is not None else None
                rows_to_insert.append((pzn, _clean_value(artikel), _clean_value(df), _clean_value(pck), _clean_value(herst), ek, absatz, _clean_value(hit.get("im_sortiment")), pzn_nmg, apu_nmg, rabatt, _clean_value(hit.get("lieferbar")), _clean_value(hit.get("bevorratung")), _clean_value(hit.get("liefervorschlag")), _clean_value(hit.get("austauschbar_gegen")), rabatt_euro, rabatt_gesamt, umsatz, _is_nmg_hit(hit.get("im_sortiment"), pzn_nmg, apu_nmg)))

        if not rows_to_insert:
            return {"file": path.name, "rows": 0, "imported": 0, "message": "keine PZN-Zeilen gefunden"}

        # idempotent pro Datei: alten Analyseimport zur Datei löschen und neu anlegen
        old_ids = [r[0] for r in con.execute("SELECT id FROM tbl_auswertungen WHERE quelldatei=? AND COALESCE(datenquelle,'NMG')=? AND bemerkung LIKE 'historischer Analyseimport%'", (path.name, datenquelle)).fetchall()]
        for aid in old_ids:
            con.execute("DELETE FROM tbl_auswertungen WHERE id=?", (aid,))
        cur = con.execute(
            """INSERT INTO tbl_auswertungen(apotheke, quelldatei, ausgabedatei, bemerkung, datenquelle)
               VALUES (?, ?, '', ?, ?)""",
            (apotheke, path.name, f"historischer Analyseimport ({mode}, {datenquelle})", datenquelle)
        )
        aid = cur.lastrowid
        # V1.1 SP13: Statistik + Batch-Tuples vorbereiten, dann ein einziges
        # executemany() statt N einzelne execute()-Calls. Bei 5000+ Zeilen
        # macht das den Unterschied zwischen 10s und 5+ Minuten.
        nmg = 0; nicht = 0; gesamt_absatz = 0.0
        batch_tuples = []
        for rec in rows_to_insert:
            ist = rec[-1]
            nmg += int(bool(ist))
            nicht += int(not bool(ist))
            gesamt_absatz += float(rec[6] or 0)
            batch_tuples.append((aid, *rec[:-1], ist, path.name, datenquelle))
        if batch_tuples:
            con.executemany(
                """INSERT INTO tbl_auswertungspositionen(
                    auswertung_id, pzn, artikelname, df, pck, herstellerkuerzel, ek, absatz_6m,
                    im_sortiment, pzn_nmg, apu_nmg, nmg_rabatt, lieferbar, bevorratung_angeraten,
                    liefervorschlag, austauschbar_gegen, nmg_rabatt_euro, nmg_rabatt_gesamt, umsatz,
                    ist_nmg_treffer, quelle, datenquelle
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                batch_tuples,
            )
        con.execute("UPDATE tbl_auswertungen SET anzahl_positionen=?, nmg_treffer=?, nicht_nmg=?, gesamt_absatz=? WHERE id=?", (len(rows_to_insert), nmg, nicht, gesamt_absatz, aid))
        con.execute("INSERT INTO tbl_import_log(datei, typ, datensaetze, meldung) VALUES (?, 'historische_marktanalyse', ?, ?)", (path.name, len(rows_to_insert), f"{apotheke} [{datenquelle}]: {nicht} Nicht-NMG / {nmg} NMG importiert"))
        con.commit()
        return {"file": path.name, "rows": len(rows_to_insert), "nmg": nmg, "nicht_nmg": nicht, "apotheke": apotheke, "mode": mode, "datenquelle": datenquelle}
    finally:
        if own_con:
            con.close()


def import_historical_market_folder(folder: str | Path = HISTORICAL_ANALYSIS_DIR, datenquelle: str = "NMG") -> dict:
    if not DB_PATH.exists():
        init_db(DB_PATH)
    folder = Path(folder)
    result = {"files": 0, "rows": 0, "nmg": 0, "nicht_nmg": 0, "errors": []}
    if not folder.exists():
        return result
    with connect(DB_PATH) as con:
        for path in sorted(folder.glob("*.xlsx")):
            try:
                r = import_historical_market_file(path, con=con, datenquelle=datenquelle)
                result["files"] += 1
                result["rows"] += int(r.get("rows", 0))
                result["nmg"] += int(r.get("nmg", 0))
                result["nicht_nmg"] += int(r.get("nicht_nmg", 0))
            except Exception as exc:
                result["errors"].append(f"{path.name}: {exc}")
                con.execute("INSERT INTO tbl_import_log(datei, typ, datensaetze, meldung) VALUES (?, 'historische_marktanalyse_fehler', 0, ?)", (path.name, str(exc)))
                con.commit()
    return result
