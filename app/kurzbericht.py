"""Kurzbericht / Management-Summary zur Bedarfsanalyse.

Erzeugt NEBEN der normalen Detail-Auswertung einen verdichteten Kurzbericht
fuer den Apotheken-Inhaber - einmal als Excel (native Diagramme, weiter
editierbar) und einmal als PDF (zum Vorzeigen). Die Detail-Auswertung selbst
wird nicht veraendert; dieses Modul liest sie nur aus.

Inhalte:
  - Einsparpotenzial gesamt (im Auswertungszeitraum) + Hochrechnung auf 12 Monate
  - Top-10-Hebel (Artikel mit der groessten Einsparung)
  - Kennzahlen / Umstellungsgrad (Positionen, umstellbar, Anteil im NMG-Sortiment)

Die Einsparung je Zeile entspricht der Spalte "NMG Rabatt Gesamt nach Absatz"
der Detail-Auswertung (= APU x Rabatt% x Absatz).
"""

from __future__ import annotations

import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Spaltenindex (1-basiert) in der Detail-Auswertung, siehe exporter.LINDEN_HEADERS.
COL_PZN = 1
COL_ARTIKEL = 2
COL_MENGE = 7          # Abverkaeufe im Zeitraum
COL_SORTIMENT = 8      # "X" / "X Austausch moegl" / leer
COL_PZN_NMG = 9        # NMG-PZN, auf die umgestellt werden kann
COL_EINSPARUNG = 17    # NMG Rabatt Gesamt nach Absatz
COL_UMSATZ = 18

ACCENT = "0B4A86"
ACCENT_LIGHT = "D9EAF7"
GREEN = "1E8E5A"


def _num(value) -> float:
    """Robuste Zahl-Konvertierung fuer Zellwerte (int/float/str, deutsch/englisch)."""
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        pass
    # Deutsches Format: 1.234,56
    try:
        return float(text.replace(".", "").replace(",", "."))
    except ValueError:
        return 0.0


def _eur(value: float, suffix: str = " €") -> str:
    """Deutsche Geldformatierung: 1.234,56 (Default mit Euro-Zeichen)."""
    s = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return s + suffix


# ── 1) Kennzahlen aus der Detail-Auswertung sammeln ──────────────────────────
def collect_summary(detail_xlsx: str | Path, zeitraum_monate: int | None = None,
                    apotheke: str = "") -> dict:
    """Liest die Detail-Auswertung und verdichtet sie zu Kennzahlen."""
    path = Path(detail_xlsx)
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    positionen = 0
    umstellbar = 0
    im_sortiment = 0
    einsparung_gesamt = 0.0
    umsatz_gesamt = 0.0
    hebel: list[tuple[str, str, float]] = []  # (pzn, artikel, einsparung)

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue
        pzn = str(row[COL_PZN - 1] or "").strip()
        if not pzn:
            continue
        positionen += 1
        einsp = _num(row[COL_EINSPARUNG - 1] if len(row) >= COL_EINSPARUNG else 0)
        umsatz_gesamt += _num(row[COL_UMSATZ - 1] if len(row) >= COL_UMSATZ else 0)
        nmg_pzn = str(row[COL_PZN_NMG - 1] or "").strip() if len(row) >= COL_PZN_NMG else ""
        sortiment = str(row[COL_SORTIMENT - 1] or "").strip() if len(row) >= COL_SORTIMENT else ""
        if nmg_pzn:
            umstellbar += 1
        if sortiment:
            im_sortiment += 1
        einsparung_gesamt += einsp
        if einsp > 0:
            artikel = str(row[COL_ARTIKEL - 1] or "").strip()
            hebel.append((pzn, artikel, einsp))

    wb.close()

    hebel.sort(key=lambda t: t[2], reverse=True)
    top10 = hebel[:10]

    z = int(zeitraum_monate) if zeitraum_monate else 6
    quote = (umstellbar / positionen * 100.0) if positionen else 0.0
    pro_monat = (einsparung_gesamt / z) if z else 0.0

    return {
        "apotheke": apotheke,
        "erstellt": datetime.now(),
        "zeitraum_monate": z,
        "positionen": positionen,
        "umstellbar": umstellbar,
        "im_sortiment": im_sortiment,
        "umstellungsgrad": quote,
        "einsparung_zeitraum": einsparung_gesamt,
        # Beide Sichten hochgerechnet aus dem tatsaechlichen Datenzeitraum, damit
        # der Vertrieb 6- und 12-Monats-Darstellung vergleichen kann.
        "einsparung_6m": pro_monat * 6,
        "einsparung_12m": pro_monat * 12,
        "umsatz_gesamt": umsatz_gesamt,
        "top10": top10,
    }


# ── 2) Excel-Kurzbericht (native Diagramme) ──────────────────────────────────
def _xl_kpi(ws, row, label, value, value_fill=ACCENT):
    ws.cell(row, 1, label).font = Font(bold=False, color="555555", size=10)
    c = ws.cell(row, 2, value)
    c.font = Font(bold=True, color=value_fill, size=14)


def create_kurzbericht_excel(summary: dict, out_path: str | Path) -> Path:
    out_path = Path(out_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Zusammenfassung"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16

    title = ws.cell(1, 1, "Bedarfsanalyse – Kurzbericht")
    title.font = Font(bold=True, size=16, color=ACCENT)
    sub = ws.cell(2, 1, f"{summary.get('apotheke') or ''}  ·  Stand {summary['erstellt']:%d.%m.%Y}"
                        f"  ·  Zeitraum {summary['zeitraum_monate']} Monate")
    sub.font = Font(size=10, color="777777")

    _xl_kpi(ws, 4, "Einsparpotenzial 6 Monate (hochgerechnet)",
            _eur(summary["einsparung_6m"]), GREEN)
    _xl_kpi(ws, 5, "Einsparpotenzial 12 Monate (hochgerechnet)",
            _eur(summary["einsparung_12m"]), GREEN)
    _xl_kpi(ws, 6, "Positionen gesamt", summary["positionen"])
    _xl_kpi(ws, 7, "davon umstellbar auf NMG", summary["umstellbar"])
    _xl_kpi(ws, 8, "Umstellungsgrad", f"{summary['umstellungsgrad']:.0f} %")
    _xl_kpi(ws, 9, "Umsatz gesamt (EK x Absatz)", _eur(summary["umsatz_gesamt"]))
    _xl_kpi(ws, 10, "Datengrundlage", f"{summary['zeitraum_monate']} Monate Absatz")

    # Top-10-Tabelle
    head_row = 11
    headers = ["#", "Artikel", "PZN", "Einsparung"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(head_row, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=ACCENT)
        cell.alignment = Alignment(horizontal="center")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(bottom=thin)
    for i, (pzn, artikel, einsp) in enumerate(summary["top10"], start=1):
        r = head_row + i
        ws.cell(r, 1, i).alignment = Alignment(horizontal="center")
        ws.cell(r, 2, artikel or "(ohne Namen)")
        ws.cell(r, 3, pzn)
        ec = ws.cell(r, 4, round(einsp, 2))
        ec.number_format = '#,##0.00 €'
        for c in range(1, 5):
            ws.cell(r, c).border = border

    # Balkendiagramm Top-10
    if summary["top10"]:
        n = len(summary["top10"])
        data_ref = Reference(ws, min_col=4, min_row=head_row, max_row=head_row + n)
        cats_ref = Reference(ws, min_col=2, min_row=head_row + 1, max_row=head_row + n)
        bar = BarChart()
        bar.type = "bar"
        bar.title = "Top-10 Einsparung je Artikel"
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.legend = None
        bar.height = 8
        bar.width = 18
        ws.add_chart(bar, "F4")

    # Kreisdiagramm Umstellungsgrad
    pie_anchor_row = head_row + len(summary["top10"]) + 3
    ws.cell(pie_anchor_row, 1, "Umstellbarkeit").font = Font(bold=True, color=ACCENT)
    ws.cell(pie_anchor_row + 1, 1, "umstellbar auf NMG")
    ws.cell(pie_anchor_row + 1, 2, summary["umstellbar"])
    rest = max(0, summary["positionen"] - summary["umstellbar"])
    ws.cell(pie_anchor_row + 2, 1, "keine NMG-Alternative")
    ws.cell(pie_anchor_row + 2, 2, rest)
    if summary["positionen"]:
        pie = PieChart()
        pie.title = "Umstellungsgrad"
        labels = Reference(ws, min_col=1, min_row=pie_anchor_row + 1, max_row=pie_anchor_row + 2)
        data = Reference(ws, min_col=2, min_row=pie_anchor_row + 1, max_row=pie_anchor_row + 2)
        pie.add_data(data, titles_from_data=False)
        pie.set_categories(labels)
        pie.height = 7
        pie.width = 9
        ws.add_chart(pie, f"F{pie_anchor_row}")

    wb.save(out_path)
    return out_path


# ── 3) Diagramme als PNG (PIL) fuer das PDF ──────────────────────────────────
def _font(size: int):
    from PIL import ImageFont
    for name in ("arial.ttf", "C:/Windows/Fonts/arial.ttf", "DejaVuSans.ttf"):
        try:
            return ImageFont.truetype(name, size)
        except Exception:
            continue
    return ImageFont.load_default()


def _bar_png(top10, path):
    from PIL import Image, ImageDraw
    W, H = 1000, 460
    pad_l, pad_r, pad_t, pad_b = 340, 175, 20, 20
    img = Image.new("RGB", (W, H), "white")
    d = ImageDraw.Draw(img)
    if not top10:
        img.save(path)
        return path
    maxv = max(t[2] for t in top10) or 1.0
    n = len(top10)
    avail_h = H - pad_t - pad_b
    bar_h = avail_h / n * 0.62
    gap = avail_h / n
    f = _font(18)
    for i, (pzn, artikel, einsp) in enumerate(top10):
        y = pad_t + i * gap + (gap - bar_h) / 2
        w = (W - pad_l - pad_r) * (einsp / maxv)
        d.rectangle([pad_l, y, pad_l + w, y + bar_h], fill=(11, 74, 134))
        name = (artikel or pzn)[:38]
        d.text((10, y + bar_h / 2 - 9), name, fill=(40, 40, 40), font=f)
        d.text((pad_l + w + 8, y + bar_h / 2 - 9), _eur(einsp, " EUR"), fill=(30, 142, 90), font=f)
    img.save(path)
    return path


def _donut_png(umstellbar, rest, path):
    from PIL import Image, ImageDraw
    W = H = 360
    img = Image.new("RGB", (W, H), "white")
    d = ImageDraw.Draw(img)
    total = umstellbar + rest
    box = [30, 30, W - 30, H - 30]
    if total <= 0:
        d.ellipse(box, fill=(220, 220, 220))
    else:
        ang = 360.0 * umstellbar / total
        d.pieslice(box, -90, -90 + ang, fill=(11, 74, 134))
        d.pieslice(box, -90 + ang, 270, fill=(210, 222, 235))
    # Loch (Donut)
    hole = [W / 2 - 70, H / 2 - 70, W / 2 + 70, H / 2 + 70]
    d.ellipse(hole, fill="white")
    f = _font(40)
    pct = (umstellbar / total * 100.0) if total else 0.0
    txt = f"{pct:.0f}%"
    try:
        bb = d.textbbox((0, 0), txt, font=f)
        tw, th = bb[2] - bb[0], bb[3] - bb[1]
    except Exception:
        tw, th = f.getsize(txt)
    d.text((W / 2 - tw / 2, H / 2 - th / 2 - 6), txt, fill=(11, 74, 134), font=f)
    img.save(path)
    return path


# ── 4) PDF-Kurzbericht (fpdf2) ───────────────────────────────────────────────
def create_kurzbericht_pdf(summary: dict, out_path: str | Path) -> Path:
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos

    out_path = Path(out_path)
    tmpdir = Path(tempfile.mkdtemp(prefix="nmg_kb_"))
    bar_png = _bar_png(summary["top10"], tmpdir / "bar.png")
    rest = max(0, summary["positionen"] - summary["umstellbar"])
    donut_png = _donut_png(summary["umstellbar"], rest, tmpdir / "donut.png")

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    def line(text, size=11, style="", color=(40, 40, 40), h=7):
        pdf.set_font("Helvetica", style, size)
        pdf.set_text_color(*color)
        pdf.cell(0, h, text, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    # Kopf
    line("Bedarfsanalyse - Kurzbericht", size=20, style="B", color=(11, 74, 134), h=11)
    line(f"{summary.get('apotheke') or ''}   |   Stand {summary['erstellt']:%d.%m.%Y}"
         f"   |   Zeitraum {summary['zeitraum_monate']} Monate", size=10, color=(120, 120, 120))
    pdf.ln(3)

    # Zwei grosse Einspar-Zahlen nebeneinander: 6 und 12 Monate.
    y0 = pdf.get_y()
    box_h = 26
    gap = 6
    box_w = (pdf.epw - gap) / 2
    boxes = [
        ("Einsparpotenzial 6 Monate", summary["einsparung_6m"]),
        ("Einsparpotenzial 12 Monate", summary["einsparung_12m"]),
    ]
    for i, (label, value) in enumerate(boxes):
        x = pdf.l_margin + i * (box_w + gap)
        pdf.set_fill_color(232, 245, 238)
        pdf.set_draw_color(232, 245, 238)
        pdf.rect(x, y0, box_w, box_h, style="F")
        pdf.set_xy(x + 4, y0 + 3)
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(90, 90, 90)
        pdf.cell(box_w - 8, 6, label, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_xy(x + 4, y0 + 11)
        pdf.set_font("Helvetica", "B", 20)
        pdf.set_text_color(30, 142, 90)
        pdf.cell(box_w - 8, 11, _eur(value, " EUR"))
    pdf.set_y(y0 + box_h + 3)
    pdf.set_font("Helvetica", "I", 9)
    pdf.set_text_color(140, 140, 140)
    pdf.cell(0, 5, f"hochgerechnet aus {summary['zeitraum_monate']} Monaten Absatz",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(1)

    # KPI-Zeilen
    kpis = [
        ("Positionen gesamt", str(summary["positionen"])),
        ("davon umstellbar auf NMG", str(summary["umstellbar"])),
        ("Umstellungsgrad", f"{summary['umstellungsgrad']:.0f} %"),
        ("Umsatz gesamt (EK x Absatz)", _eur(summary["umsatz_gesamt"], " EUR")),
    ]
    for label, val in kpis:
        pdf.set_font("Helvetica", "", 11)
        pdf.set_text_color(90, 90, 90)
        pdf.cell(110, 7, label, border=0)
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(40, 40, 40)
        pdf.cell(0, 7, val, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(2)

    # Donut rechts neben den KPIs (Bildplatzierung verschiebt den Cursor nicht).
    donut_y = y0 + box_h + 6
    pdf.image(str(donut_png), x=pdf.w - pdf.r_margin - 45, y=donut_y, w=42)

    # Top-10: Cursor unter Donut UND KPIs setzen, damit nichts ueberlappt.
    pdf.set_y(max(pdf.get_y(), donut_y + 44))
    pdf.ln(2)
    line("Top-10 Hebel-Artikel", size=13, style="B", color=(11, 74, 134), h=8)
    pdf.image(str(bar_png), x=pdf.l_margin, w=pdf.epw)

    pdf.output(str(out_path))
    return out_path


# ── 5) Orchestrator ──────────────────────────────────────────────────────────
def create_kurzbericht(detail_xlsx: str | Path, analyse_name: str, out_dir: str | Path,
                       zeitraum_monate: int | None = None, apotheke: str = "") -> dict:
    """Erzeugt Excel- und PDF-Kurzbericht aus der Detail-Auswertung.

    Liefert {"excel": Path|None, "pdf": Path|None, "summary": dict}. Einzelne
    Fehler (z.B. fehlende PDF-Lib) brechen den Gesamtlauf NICHT ab.
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    safe = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in (analyse_name or "Bedarfsanalyse")).strip("_") or "Bedarfsanalyse"

    summary = collect_summary(detail_xlsx, zeitraum_monate=zeitraum_monate,
                              apotheke=apotheke or analyse_name)
    result = {"excel": None, "pdf": None, "summary": summary}

    try:
        result["excel"] = create_kurzbericht_excel(summary, out_dir / f"Kurzbericht_{safe}.xlsx")
    except Exception:
        pass
    try:
        result["pdf"] = create_kurzbericht_pdf(summary, out_dir / f"Kurzbericht_{safe}.pdf")
    except Exception:
        pass
    return result
