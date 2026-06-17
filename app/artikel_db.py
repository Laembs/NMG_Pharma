from pathlib import Path
from datetime import datetime
import re
import sqlite3

from openpyxl import load_workbook
from .file_loader import load_worksheet

from .config import DB_PATH


def _clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return " ".join(text.split())


def _pzn(value):
    text = _clean(value)
    if not text:
        return ""
    digits = "".join(ch for ch in text if ch.isdigit())
    if not digits:
        return ""
    return digits.zfill(8)


def _norm_header(value):
    text = _clean(value).lower()
    text = text.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    return re.sub(r"[^a-z0-9]+", "", text)


def _find_column(header_map, possible_names):
    wanted = [_norm_header(name) for name in possible_names]
    for key in wanted:
        if key in header_map:
            return header_map[key]
    for header, col in header_map.items():
        if any(key and key in header for key in wanted):
            return col
    return None


def _decimal(value):
    """SP15: Konvertiert Excel-Werte ('5,99' / '5.99' / 5.99 / '') zu float oder None.
    Wird fuer EK / Listen-EK gebraucht.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", ".").replace(" ", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def ensure_artikelstamm_table():
    """Legt die zentrale PZN-Artikelbasis an, ohne bestehende Daten zu löschen.
    SP15: ek-Spalte (Listen-EK / Taxe-EK / Import-EK) hinzugefuegt - wird in
    der Produktanalyse fuer die Umsatzberechnung verwendet.
    """
    with sqlite3.connect(DB_PATH) as con:
        con.execute("""
            CREATE TABLE IF NOT EXISTS tbl_artikelstamm (
                pzn TEXT PRIMARY KEY,
                artikel TEXT,
                df TEXT,
                pck TEXT,
                herst TEXT,
                ek REAL,
                quelle TEXT NOT NULL DEFAULT 'Import',
                treffer INTEGER NOT NULL DEFAULT 1,
                erstellt_am TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                aktualisiert_am TEXT
            )
        """)
        # SP15: Migration fuer Bestands-DBs - falls die ek-Spalte fehlt, nachpflegen.
        cols = {r[1] for r in con.execute("PRAGMA table_info(tbl_artikelstamm)").fetchall()}
        if "ek" not in cols:
            con.execute("ALTER TABLE tbl_artikelstamm ADD COLUMN ek REAL")
        con.execute("""
            CREATE INDEX IF NOT EXISTS idx_artikelstamm_artikel
            ON tbl_artikelstamm (artikel)
        """)
        con.execute("""
            CREATE INDEX IF NOT EXISTS idx_artikelstamm_herst
            ON tbl_artikelstamm (herst)
        """)
        con.commit()


def _detect_columns(ws):
    header_row = 1
    best_score = -1
    best_map = {}

    for row in range(1, min(ws.max_row, 15) + 1):
        header_map = {}
        for col in range(1, ws.max_column + 1):
            key = _norm_header(ws.cell(row, col).value)
            if key:
                header_map[key] = col

        col_pzn = _find_column(header_map, ["PZN", "Pharmazentralnummer", "Pharmazentral-Nr", "PZN Nr"])
        col_artikel = _find_column(header_map, ["Artikel", "Artikelname", "Artikelbez.", "Artikelbez", "Artikelbezeichnung", "Bezeichnung", "Name"])
        col_df = _find_column(header_map, ["DF", "DAR", "Darreichungsform"])
        col_pck = _find_column(header_map, ["PCK", "Pck", "Packung", "Pack.Gr", "Packungsgröße", "Packungsgroesse"])
        col_herst = _find_column(header_map, ["Herst", "Herst.", "Hersteller", "Herstellerkürzel", "Anbieter", "Firma", "Lieferant"])
        # SP15: EK-Spalte (Listen-EK, Taxe-EK, Import-EK etc.)
        col_ek = _find_column(header_map, [
            "EK", "Listen-EK", "Listen EK", "ListenEK", "Import EK", "Import-EK",
            "ImportEK", "TAX-EK", "TAXEK", "Taxe EK", "Taxe-EK", "Einkaufspreis",
            "EK Preis", "Einkauf",
        ])

        score = int(bool(col_pzn)) * 10 + sum(int(bool(x)) for x in [col_artikel, col_df, col_pck, col_herst, col_ek])
        if score > best_score:
            best_score = score
            best_map = {
                "header_row": row,
                "pzn": col_pzn,
                "artikel": col_artikel,
                "df": col_df,
                "pck": col_pck,
                "herst": col_herst,
                "ek": col_ek,
            }

    if not best_map.get("pzn"):
        raise ValueError("Artikelstamm konnte nicht importiert werden. Fehlende Pflichtspalte: PZN")
    return best_map


def import_artikelstamm_excel(file_path, quelle="Artikelstamm-Import", progress_callback=None, batch_size=5000):
    """
    Schneller Import für große Artikelstamm-Dateien.

    Optimierungen:
    - read_only=True für große Exceldateien
    - eine Transaktion statt Commit pro Zeile
    - executemany in Blöcken
    - kein SELECT pro PZN
    - optionaler Fortschrittscallback für die GUI
    """
    ensure_artikelstamm_table()

    path = Path(file_path)
    ws, input_typ = load_worksheet(path)
    mapping = _detect_columns(ws)
    header_row = int(mapping["header_row"])

    total_rows = max(0, ws.max_row - header_row)
    inserted_or_updated = 0
    skipped = 0
    batch = []

    def col_index(key):
        col = mapping.get(key)
        return int(col) - 1 if col else None

    idx_pzn = col_index("pzn")
    idx_artikel = col_index("artikel")
    idx_df = col_index("df")
    idx_pck = col_index("pck")
    idx_herst = col_index("herst")
    idx_ek = col_index("ek")  # SP15

    def get(row, idx):
        if idx is None or idx >= len(row):
            return ""
        return _clean(row[idx])

    def get_decimal(row, idx):
        # SP15: numerische Spalten (EK) sicher parsen.
        if idx is None or idx >= len(row):
            return None
        return _decimal(row[idx])

    sql = """
        INSERT INTO tbl_artikelstamm (
            pzn, artikel, df, pck, herst, ek, quelle, treffer, erstellt_am, aktualisiert_am
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, 1, CURRENT_TIMESTAMP, NULL)
        ON CONFLICT(pzn) DO UPDATE SET
            artikel = COALESCE(NULLIF(excluded.artikel, ''), tbl_artikelstamm.artikel),
            df = COALESCE(NULLIF(excluded.df, ''), tbl_artikelstamm.df),
            pck = COALESCE(NULLIF(excluded.pck, ''), tbl_artikelstamm.pck),
            herst = COALESCE(NULLIF(excluded.herst, ''), tbl_artikelstamm.herst),
            ek = COALESCE(excluded.ek, tbl_artikelstamm.ek),
            quelle = excluded.quelle,
            treffer = tbl_artikelstamm.treffer + 1,
            aktualisiert_am = CURRENT_TIMESTAMP
    """

    with sqlite3.connect(DB_PATH) as con:
        # Schneller, aber weiterhin sicher genug für einen Import. Backup gibt es im Programm ohnehin.
        con.execute("PRAGMA journal_mode=WAL")
        con.execute("PRAGMA synchronous=NORMAL")
        con.execute("BEGIN")

        for current, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=1):
            pzn = _pzn(row[idx_pzn] if idx_pzn is not None and idx_pzn < len(row) else None)
            if not pzn:
                skipped += 1
            else:
                batch.append((
                    pzn,
                    get(row, idx_artikel),
                    get(row, idx_df),
                    get(row, idx_pck),
                    get(row, idx_herst),
                    get_decimal(row, idx_ek),  # SP15
                    quelle,
                ))

            if len(batch) >= batch_size:
                con.executemany(sql, batch)
                inserted_or_updated += len(batch)
                batch.clear()
                if progress_callback:
                    progress_callback(current, total_rows)

        if batch:
            con.executemany(sql, batch)
            inserted_or_updated += len(batch)
            batch.clear()

        con.commit()

    if progress_callback:
        progress_callback(total_rows, total_rows)

    return {
        "file": str(path),
        "imported_or_updated": inserted_or_updated,
        "inserted": inserted_or_updated,  # kompatibel zur alten GUI-Anzeige
        "updated": 0,                     # exakte Trennung wird zugunsten Geschwindigkeit nicht ermittelt
        "skipped": skipped,
        "total_processed": inserted_or_updated + skipped,
        "total_rows": total_rows,
    }


def count_artikelstamm():
    ensure_artikelstamm_table()
    with sqlite3.connect(DB_PATH) as con:
        return con.execute("SELECT COUNT(*) FROM tbl_artikelstamm").fetchone()[0]


def find_artikel_by_pzn(pzn):
    ensure_artikelstamm_table()
    pzn = _pzn(pzn)
    if not pzn:
        return None
    with sqlite3.connect(DB_PATH) as con:
        con.row_factory = sqlite3.Row
        return con.execute("""
            SELECT pzn, artikel, df, pck, herst, quelle, treffer, erstellt_am, aktualisiert_am
            FROM tbl_artikelstamm
            WHERE pzn = ?
        """, (pzn,)).fetchone()


def list_artikelstamm(limit=200):
    ensure_artikelstamm_table()
    with sqlite3.connect(DB_PATH) as con:
        con.row_factory = sqlite3.Row
        return con.execute("""
            SELECT pzn, artikel, df, pck, herst, quelle, treffer, erstellt_am, aktualisiert_am
            FROM tbl_artikelstamm
            ORDER BY COALESCE(aktualisiert_am, erstellt_am) DESC, pzn
            LIMIT ?
        """, (int(limit),)).fetchall()
