from __future__ import annotations

import csv
import hashlib
import shutil
import sqlite3
from pathlib import Path
from datetime import datetime
from tempfile import TemporaryDirectory
from typing import Iterable

from openpyxl import Workbook, load_workbook

from .config import DB_PATH, IMPORT_DIR, jahr_quartal_pfad
from .db import connect, init_db, _pzn
from .historical_import import import_historical_market_file


def _kopiere_in_archivordner(path: Path, analyse_typ: str) -> Path:
    """V1.1 SP12: Kopiert die Original-Datei nach
    IMPORT_DIR/<analyse_typ>/<Jahr>/Q<n>/ und liefert den Ziel-Pfad zurueck.
    analyse_typ: 'PK' oder 'ZF'. Dateiname bleibt erhalten; bei Konflikt
    wird ein _2, _3, ... angehaengt.
    """
    target_dir = jahr_quartal_pfad(IMPORT_DIR / analyse_typ)
    target = target_dir / path.name
    if target.exists():
        # Schon eine Datei mit dem Namen da -> Suffix anhaengen.
        stem, suffix = target.stem, target.suffix
        counter = 2
        while True:
            candidate = target_dir / f"{stem}_{counter}{suffix}"
            if not candidate.exists():
                target = candidate
                break
            counter += 1
    try:
        shutil.copy2(path, target)
    except Exception:
        # Kopie ist Best-Effort - wenn sie scheitert (Permissions o.ae.),
        # bleibt der Original-Pfad in tbl_importierte_analysen erhalten.
        return path
    return target


class ManualAnalysisFormatError(Exception):
    """Datei konnte nicht als manuelle Analyse erkannt/importiert werden."""


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _ensure_tables(con: sqlite3.Connection) -> None:
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS tbl_importierte_analysen (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dateiname TEXT,
            dateipfad TEXT,
            dateigroesse INTEGER,
            hashwert TEXT UNIQUE,
            analyse_typ TEXT,
            auswertung_id INTEGER,
            status TEXT DEFAULT 'importiert',
            importiert_am TEXT DEFAULT CURRENT_TIMESTAMP,
            bearbeiter TEXT,
            bemerkung TEXT
        )
        """
    )
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS tbl_lernvorschlaege (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            produkt_alt TEXT NOT NULL DEFAULT '',
            produkt_neu TEXT NOT NULL DEFAULT '',
            pzn_alt TEXT,
            artikel_alt TEXT,
            pzn_nmg TEXT,
            freitext_austausch TEXT,
            quelle_datei TEXT,
            status TEXT NOT NULL DEFAULT 'neu',
            erstellt_am TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            bearbeitet_am TEXT,
            bearbeiter TEXT,
            historie_ab TEXT,
            historie_bis TEXT,
            austausch_id INTEGER
        )
        """
    )
    cols = {row[1] for row in con.execute("PRAGMA table_info(tbl_lernvorschlaege)").fetchall()}
    additions = {
        "pzn_alt": "TEXT",
        "artikel_alt": "TEXT",
        "pzn_nmg": "TEXT",
        "freitext_austausch": "TEXT",
        "quelle_datei": "TEXT",
        "bearbeiter": "TEXT",
        "historie_ab": "TEXT",
        "historie_bis": "TEXT",
        "austausch_id": "INTEGER",
    }
    for col, definition in additions.items():
        if col not in cols:
            con.execute(f"ALTER TABLE tbl_lernvorschlaege ADD COLUMN {col} {definition}")
    con.commit()


def _norm_header(value) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    return "".join(ch for ch in text if ch.isalnum())


def _find_header_row_and_map(ws, max_scan_rows: int = 25):
    aliases = {
        "pzn": {"pzn", "pharmazentralnummer", "artikelpzn"},
        "artikel": {"artikel", "artikelname", "artikelbezeichnung", "bezeichnung"},
        "df": {"df", "dar", "darreichungsform"},
        "pck": {"pck", "packung", "packgr", "einheit"},
        "herst": {"herst", "hersteller", "anbieter"},
        "absatz": {"abverkaeufe6monate", "verkaufsmengederletzten6monate", "absatz", "menge", "gesamtmenge", "abverkauf"},
        "im_sortiment": {"imsortiment", "sortiment"},
        "pzn_nmg": {"pznnmg", "nmgpzn"},
        "austausch": {"austauschbargegen", "austausch", "austauschartikel", "ersatzartikel"},
    }
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        mapping = {}
        for c in range(1, ws.max_column + 1):
            h = _norm_header(ws.cell(r, c).value)
            if not h:
                continue
            for key, names in aliases.items():
                if key not in mapping and (h in names or any(name in h for name in names if len(name) > 5)):
                    mapping[key] = c
        if "pzn" in mapping and ("artikel" in mapping or "absatz" in mapping or "pzn_nmg" in mapping or "austausch" in mapping):
            return r, mapping
    return None, None


def _csv_or_txt_to_xlsx(path: Path, target_dir: Path) -> Path:
    # einfache robuste Konvertierung für CSV/TXT; Trennzeichen wird aus einer Probe erkannt.
    sample = path.read_text(encoding="utf-8-sig", errors="replace")[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,	|")
    except Exception:
        dialect = csv.excel
        dialect.delimiter = ";"
    out = target_dir / f"{path.stem}_konvertiert.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Auswertung"
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
        reader = csv.reader(f, dialect)
        for row in reader:
            ws.append(row)
    wb.save(out)
    return out


def _prepare_readable_file(path: Path, temp_dir: Path | None = None) -> Path:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return path
    if suffix in {".csv", ".txt"}:
        if temp_dir is None:
            raise ManualAnalysisFormatError("CSV/TXT benötigt ein temporäres Zielverzeichnis.")
        return _csv_or_txt_to_xlsx(path, temp_dir)
    raise ManualAnalysisFormatError(f"Dateityp wird nicht unterstützt: {path.suffix}")


def _cell(row, col):
    if not col:
        return None
    idx = col - 1
    return row[idx] if 0 <= idx < len(row) else None


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _insert_learning_suggestions(con: sqlite3.Connection, readable_file: Path, source_name: str, bearbeiter: str = "") -> int:
    """V1.1 SP13: Batch-Variante. Vorher pro Excel-Zeile 2 Queries
    (SELECT-Duplikat-Check + INSERT) = 5000 Zeilen -> 10.000 Queries +
    je eine implizite Transaction. Jetzt: alle vorhandenen Schluessel
    EINMAL laden, alle neuen Zeilen sammeln, dann ein executemany +
    eine Transaction.
    """
    wb = load_workbook(readable_file, data_only=True, read_only=True)

    # Bestehende Schluessel einmal als Set in den Speicher laden.
    existing = set()
    try:
        for r in con.execute(
            """SELECT COALESCE(pzn_alt,''), COALESCE(pzn_nmg,''),
                      COALESCE(freitext_austausch,'')
               FROM tbl_lernvorschlaege"""
        ).fetchall():
            existing.add((r[0], r[1], r[2]))
    except sqlite3.OperationalError:
        pass

    to_insert: list[tuple] = []
    seen_in_batch: set = set()

    for ws in wb.worksheets:
        header_row, mp = _find_header_row_and_map(ws)
        if not header_row or not mp:
            continue
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            pzn_alt = _pzn(_cell(row, mp.get("pzn")))
            if not pzn_alt:
                continue
            pzn_nmg = _pzn(_cell(row, mp.get("pzn_nmg")))
            freitext = _text(_cell(row, mp.get("austausch")))
            artikel_alt = _text(_cell(row, mp.get("artikel")))
            if not pzn_nmg and not freitext:
                continue
            key = (pzn_alt, pzn_nmg, freitext)
            if key in existing or key in seen_in_batch:
                continue
            seen_in_batch.add(key)
            to_insert.append(
                (artikel_alt, freitext, pzn_alt, artikel_alt, pzn_nmg,
                 freitext, source_name, bearbeiter)
            )

    if to_insert:
        con.execute("BEGIN")
        con.executemany(
            """INSERT INTO tbl_lernvorschlaege (
                produkt_alt, produkt_neu, pzn_alt, artikel_alt, pzn_nmg,
                freitext_austausch, quelle_datei, status, erstellt_am, bearbeiter
            ) VALUES (?, ?, ?, ?, ?, ?, ?, 'neu', CURRENT_TIMESTAMP, ?)""",
            to_insert,
        )
        con.commit()
    return len(to_insert)



def _insert_learning_suggestions_from_auswertung(con: sqlite3.Connection, auswertung_id: int | None, source_name: str, bearbeiter: str = "") -> int:
    """Fallback: Lernvorschläge aus den bereits gespeicherten Auswertungspositionen erzeugen.

    Grund: Manche manuelle Dateien werden zwar korrekt in tbl_auswertungen/tbl_auswertungspositionen
    importiert, aber die direkte Spaltenerkennung in der Originaldatei findet PZN NMG/Austausch
    nicht zuverlässig. Dann darf die Schulbank nicht leer bleiben.
    """
    if not auswertung_id:
        return 0

    # V1.1 SP13: Batch + Single Transaction. Vorher pro Position 2 Queries.
    rows = con.execute(
        """
        SELECT pzn, artikelname, pzn_nmg, austauschbar_gegen
        FROM tbl_auswertungspositionen
        WHERE auswertung_id = ?
          AND COALESCE(pzn, '') <> ''
          AND (COALESCE(pzn_nmg, '') <> '' OR COALESCE(austauschbar_gegen, '') <> '')
        """,
        (int(auswertung_id),),
    ).fetchall()
    if not rows:
        return 0

    existing = set()
    try:
        for r in con.execute(
            """SELECT COALESCE(pzn_alt,''), COALESCE(pzn_nmg,''),
                      COALESCE(freitext_austausch,'')
               FROM tbl_lernvorschlaege"""
        ).fetchall():
            existing.add((r[0], r[1], r[2]))
    except sqlite3.OperationalError:
        pass

    to_insert: list[tuple] = []
    seen_in_batch: set = set()
    for row in rows:
        pzn_alt = _pzn(row[0])
        artikel_alt = _text(row[1])
        pzn_nmg = _pzn(row[2])
        freitext = _text(row[3])
        if not pzn_alt or (not pzn_nmg and not freitext):
            continue
        key = (pzn_alt, pzn_nmg, freitext)
        if key in existing or key in seen_in_batch:
            continue
        seen_in_batch.add(key)
        to_insert.append(
            (artikel_alt, freitext, pzn_alt, artikel_alt, pzn_nmg,
             freitext, source_name, bearbeiter)
        )

    if to_insert:
        con.execute("BEGIN")
        con.executemany(
            """INSERT INTO tbl_lernvorschlaege (
                produkt_alt, produkt_neu, pzn_alt, artikel_alt, pzn_nmg,
                freitext_austausch, quelle_datei, status, erstellt_am, bearbeiter
            ) VALUES (?, ?, ?, ?, ?, ?, ?, 'neu', CURRENT_TIMESTAMP, ?)""",
            to_insert,
        )
        con.commit()
    return len(to_insert)

def _latest_auswertung_id(con: sqlite3.Connection) -> int | None:
    try:
        row = con.execute("SELECT id FROM tbl_auswertungen ORDER BY id DESC LIMIT 1").fetchone()
        return int(row[0]) if row else None
    except Exception:
        return None


def import_manual_analysis_files(paths: Iterable[str | Path], analyse_typ: str, bearbeiter: str = "") -> dict:
    """Importiert geprüfte manuelle Analysen als Historien-/Analysebasis.

    - keine neue fachliche Auswertung rechnen
    - keine Hersteller-Lernlogik erzwingen
    - Dubletten über SHA256 überspringen
    - Daten in tbl_auswertungen/tbl_auswertungspositionen übernehmen
    - vorhandene Austausch-/NMG-Entscheidungen als Schulbank-Vorschläge anlegen
    """
    analyse_typ = (analyse_typ or "").upper().strip()
    if analyse_typ == "PK":
        datenquelle = "NMG"
    elif analyse_typ == "ZF":
        datenquelle = "ZF"
    else:
        raise ValueError("Analyseart muss PK oder ZF sein.")

    if not DB_PATH.exists():
        init_db(DB_PATH)

    stats = {
        "selected": 0,
        "imported": 0,
        "duplicates": 0,
        "failed": 0,
        "positions": 0,
        "learning_suggestions": 0,
        "duplicate_files": [],
        "errors": [],
    }

    with connect(DB_PATH) as con:
        con.row_factory = sqlite3.Row
        _ensure_tables(con)

    with TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        for item in paths:
            stats["selected"] += 1
            path = Path(item)
            try:
                if not path.exists():
                    raise FileNotFoundError(path)
                digest = _sha256(path)
                size = path.stat().st_size
                with connect(DB_PATH) as con:
                    _ensure_tables(con)
                    # V1.1 SP10 Bug-Fix: Duplikat nur dann, wenn der bisherige
                    # Hash-Eintrag noch eine LEBENDE auswertung_id hat. Wenn die
                    # zugehoerige Auswertung in tbl_auswertungen geloescht wurde
                    # (Admin-Cleanup, Zeitraum-Loeschung, Archiv-Auslagerung),
                    # zaehlt das nicht mehr als Duplikat - die Datei darf neu
                    # eingespielt werden. Sonst landen Re-Imports nie wieder
                    # in der Produktanalyse.
                    existing = con.execute(
                        """SELECT ia.id, ia.auswertung_id
                           FROM tbl_importierte_analysen ia
                           WHERE ia.hashwert = ? LIMIT 1""",
                        (digest,),
                    ).fetchone()
                    existing_stale_id = None
                    if existing:
                        old_aw_id = existing["auswertung_id"] if isinstance(existing, sqlite3.Row) else existing[1]
                        ist_lebend = False
                        if old_aw_id is not None:
                            r = con.execute(
                                "SELECT 1 FROM tbl_auswertungen WHERE id = ? LIMIT 1",
                                (old_aw_id,),
                            ).fetchone()
                            ist_lebend = r is not None
                        if ist_lebend:
                            stats["duplicates"] += 1
                            stats["duplicate_files"].append(path.name)
                            continue
                        # Hash-Eintrag ist verwaist -> beim Neu-Import ueberschreiben.
                        existing_stale_id = existing["id"] if isinstance(existing, sqlite3.Row) else existing[0]

                readable = _prepare_readable_file(path, tmp_path)

                # Historienbasis für Produkt-/Marktanalyse füllen.
                result = import_historical_market_file(readable, datenquelle=datenquelle, analyse_name=path.stem)
                rows_imported = int(result.get("rows", 0) or result.get("imported", 0) or 0)

                # V1.1 SP12: Originaldatei in den passenden PK-/ZF-Ordner unter
                # gespeicherte_analysen kopieren. Auch in tbl_auswertungen den
                # ausgabedatei-Wert auf den Archiv-Pfad setzen, damit
                # "Auswertung oeffnen" in der GUI funktioniert.
                archiv_pfad = _kopiere_in_archivordner(path, analyse_typ)

                with connect(DB_PATH) as con:
                    con.row_factory = sqlite3.Row
                    _ensure_tables(con)
                    auswertung_id = _latest_auswertung_id(con)
                    suggestions = _insert_learning_suggestions(con, readable, path.name, bearbeiter)
                    if suggestions == 0 and auswertung_id:
                        suggestions = _insert_learning_suggestions_from_auswertung(con, auswertung_id, path.name, bearbeiter)
                    # ausgabedatei in tbl_auswertungen auf Archiv-Pfad setzen,
                    # damit der "Auswertung oeffnen"-Button in der GUI direkt
                    # die kopierte Datei oeffnet.
                    if auswertung_id is not None:
                        try:
                            con.execute(
                                "UPDATE tbl_auswertungen SET ausgabedatei=? WHERE id=?",
                                (str(archiv_pfad), auswertung_id),
                            )
                        except Exception:
                            pass
                    if existing_stale_id is not None:
                        # Verwaister Hash-Eintrag -> mit neuer auswertung_id ueberschreiben.
                        con.execute(
                            """UPDATE tbl_importierte_analysen
                               SET dateipfad=?, dateigroesse=?, analyse_typ=?, auswertung_id=?,
                                   status='importiert', importiert_am=CURRENT_TIMESTAMP,
                                   bearbeiter=?, bemerkung=?
                               WHERE id=?""",
                            (str(archiv_pfad), size, analyse_typ, auswertung_id, bearbeiter,
                             f"Positionen: {rows_imported}; Schulbank-Vorschläge: {suggestions} (Re-Import)",
                             existing_stale_id),
                        )
                    else:
                        con.execute(
                            """
                            INSERT INTO tbl_importierte_analysen (
                                dateiname, dateipfad, dateigroesse, hashwert, analyse_typ,
                                auswertung_id, status, importiert_am, bearbeiter, bemerkung
                            ) VALUES (?, ?, ?, ?, ?, ?, 'importiert', CURRENT_TIMESTAMP, ?, ?)
                            """,
                            (path.name, str(archiv_pfad), size, digest, analyse_typ, auswertung_id, bearbeiter, f"Positionen: {rows_imported}; Schulbank-Vorschläge: {suggestions}"),
                        )
                    con.commit()

                stats["imported"] += 1
                stats["positions"] += rows_imported
                stats["learning_suggestions"] += suggestions
            except Exception as exc:
                stats["failed"] += 1
                stats["errors"].append(f"{path.name}: {exc}")
    return stats
